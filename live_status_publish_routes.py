import logging
import re
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, make_response

from flask_login import login_required, current_user

from config import (
    ADMIN_USERS,
    TARGET_GROUP,
    JIRA_PDT_FILTER_ID,
    VIEWER_OVERRIDE_USERS,
    LIVE_STATUS_VIEWER_GROUP_ACCESS,
)
from dashboard_common import get_business_units, get_targets_for_bu, get_bu_for_target, get_display_name_for_target, load_metadata_config, get_auto_target_keys

logger = logging.getLogger(__name__)

from live_status_publish_service import (
    list_jobs,
    get_job,
    create_job,
    save_job_meta,
    save_job_rows,
    publish_job,
    revoke_job,


    update_viewer_heartbeat,
    delete_job,

    load_job_workspace_data,
    get_report_sidecar,
    _build_key,
    set_sidecar_jql,
    set_sidecar_report_cache,
    set_sidecar_exclusions,
    set_sidecar_swpdt_builds,
    get_swpdt_sidecar,
    _delete_sidecars,
    set_weekly_report_selection,
    _utc_now,
        _SWPDT_JSON,
)

from live_view_saved_jql_service import (
    list_tabs as list_saved_jql_tabs,
    save_tab as save_saved_jql_tab,
    delete_tab as delete_saved_jql_tab,
    get_tab as get_saved_jql_tab,
    get_cached_report as get_saved_jql_cached_report,
    get_cached_report_raw as get_saved_jql_cached_report_raw,
    set_cached_report as set_saved_jql_cached_report,
)

def _iframe_aware_redirect(url):
    """Return a tiny HTML page that navigates window.top (the landing page)
    to load the target URL inside the viewer iframe - never breaks out.
    """
    import json
    from flask import make_response
    safe_url_json = json.dumps(str(url or ''))
    html = '''<!doctype html><html><head><meta charset="utf-8"></head><body>
<script>
(function(){
  var u = __URL__;
  if(window.parent && window.parent !== window){
    // Tell the landing page to load this URL in the viewer iframe
    try{ window.parent.postMessage({type:'lsp_navigate',url:u}, '*'); }catch(e){}
  } else {
    window.location.href = u;
  }
})();
</script>
</body></html>'''.replace('__URL__', safe_url_json)
    resp = make_response(html, 200)
    resp.headers['Content-Type'] = 'text/html; charset=utf-8'
    return resp



# Axiom fetch enabled - controlled by ENABLE_SWPDT_AXIOM_POLLER env var.
AXIOM_FETCH_DISABLED = False

live_status_publish_bp = Blueprint('live_status_publish_bp', __name__)


def _target_group_access() -> bool:
    """Editor access remains controlled by qipl.target.pdt / admins only.
    Result is cached on Flask g for the duration of the request."""
    from flask import g
    cached = getattr(g, '_target_group_access_result', None)
    if cached is not None:
        return cached
    uid = getattr(current_user, 'id', '') or ''
    if uid in VIEWER_OVERRIDE_USERS:
        result = False
    elif uid in ADMIN_USERS:
        result = True
    else:
        try:
            import app as _app
            result = _app.is_user_in_group(uid, TARGET_GROUP)
        except Exception:
            result = False
    g._target_group_access_result = result
    return result


def _norm_access_list(values):
    if values is None:
        return set()
    if isinstance(values, str):
        values = [values]
    try:
        return {str(v or '').strip().upper() for v in values if str(v or '').strip()}
    except Exception:
        return set()


def _scope_target_patterns(scope):
    if not isinstance(scope, dict):
        return set()
    return _norm_access_list(scope.get('target_patterns') or scope.get('target_pattern') or scope.get('patterns'))


def _live_status_user_in_group(uid, group_name) -> bool:
    """Cached LDAP/test group membership check for this Flask request."""
    from flask import g
    uid = str(uid or '').strip().lower()
    group_name = str(group_name or '').strip()
    if not uid or not group_name:
        return False
    cache = getattr(g, '_live_status_group_membership_cache', None)
    if cache is None:
        cache = {}
        g._live_status_group_membership_cache = cache
    key = (uid, group_name)
    if key in cache:
        return cache[key]
    try:
        import app as _app
        result = bool(_app.is_user_in_group(uid, group_name))
    except Exception as exc:
        logger.warning('[LIVE STATUS ACCESS] group check failed for %s in %s: %s', uid, group_name, exc)
        result = False
    cache[key] = result
    return result


def _target_matches_access_scope(target_name, scope, bu_key=None) -> bool:
    if not scope:
        return False
    if scope.get('all'):
        return True
    target_upper = str(target_name or '').strip().upper()
    bu_upper = str(bu_key or get_bu_for_target(target_name) or '').strip().upper()
    if target_upper and target_upper in (scope.get('targets') or set()):
        return True
    if bu_upper and bu_upper in (scope.get('bus') or set()):
        return True
    for pattern in scope.get('target_patterns') or set():
        if pattern and target_upper and pattern in target_upper:
            return True
    return False


def _current_live_status_viewer_scope():
    """Return read-only Live Status scope for non-editor viewers.

    Config source: config.LIVE_STATUS_VIEWER_GROUP_ACCESS, keyed by LDAP group.
    Each value can include bus/bu, targets, and target_patterns. Matching groups
    are unioned. This does not grant edit/save/publish permissions.
    """
    from flask import g
    cached = getattr(g, '_live_status_viewer_scope_cache', None)
    if cached is not None:
        return cached
    if _target_group_access():
        result = {'all': True, 'bus': {'*'}, 'targets': {'*'}, 'target_patterns': {'*'}, 'matched_groups': []}
        g._live_status_viewer_scope_cache = result
        return result
    if not current_user.is_authenticated:
        result = {'all': False, 'bus': set(), 'targets': set(), 'target_patterns': set(), 'matched_groups': []}
        g._live_status_viewer_scope_cache = result
        return result

    uid = getattr(current_user, 'id', '') or ''
    cfg = LIVE_STATUS_VIEWER_GROUP_ACCESS or {}
    if not isinstance(cfg, dict) or not cfg:
        result = {'all': False, 'bus': set(), 'targets': set(), 'target_patterns': set(), 'matched_groups': []}
        g._live_status_viewer_scope_cache = result
        return result

    bus_scope = set()
    target_scope = set()
    pattern_scope = set()
    matched_groups = []
    try:
        for group_name, scope in cfg.items():
            group_name = str(group_name or '').strip()
            if not group_name:
                continue
            if not _live_status_user_in_group(uid, group_name):
                continue
            matched_groups.append(group_name)
            scope = scope or {}
            if isinstance(scope, (list, tuple, set, str)):
                # Shorthand: {"group": ["AUTO", "IOT"]} means BU scope.
                bus_scope |= _norm_access_list(scope)
                continue
            if not isinstance(scope, dict):
                continue
            if bool(scope.get('all')):
                result = {'all': True, 'bus': {'*'}, 'targets': {'*'}, 'target_patterns': {'*'}, 'matched_groups': matched_groups}
                g._live_status_viewer_scope_cache = result
                return result
            bus_scope |= _norm_access_list(scope.get('bus') or scope.get('bu') or scope.get('business_units'))
            target_scope |= _norm_access_list(scope.get('targets') or scope.get('target'))
            pattern_scope |= _scope_target_patterns(scope)
    except Exception as exc:
        logger.warning('[LIVE STATUS ACCESS] scope resolution failed for %s: %s', uid, exc)
        result = {'all': False, 'bus': set(), 'targets': set(), 'target_patterns': set(), 'matched_groups': []}
        g._live_status_viewer_scope_cache = result
        return result

    if '*' in bus_scope or 'ALL' in bus_scope or '*' in target_scope or 'ALL' in target_scope or '*' in pattern_scope or 'ALL' in pattern_scope:
        result = {'all': True, 'bus': {'*'}, 'targets': {'*'}, 'target_patterns': {'*'}, 'matched_groups': matched_groups}
    else:
        result = {'all': False, 'bus': bus_scope, 'targets': target_scope, 'target_patterns': pattern_scope, 'matched_groups': matched_groups}
    g._live_status_viewer_scope_cache = result
    return result




def _can_view_live_status_target(target_name, bu_key=None) -> bool:
    """True if current user may view this target in read-only Live Status."""
    if _target_group_access():
        return True
    return _target_matches_access_scope(target_name, _current_live_status_viewer_scope(), bu_key)



def _filter_live_status_jobs_for_current_user(jobs, *, can_edit=False):
    if can_edit:
        return list(jobs or [])
    out = []
    for job in jobs or []:
        first_target = (job.get('targets') or [''])[0]
        if first_target and _can_view_live_status_target(first_target, job.get('_bu_key')):
            out.append(job)
    return out


def _live_status_group_join_url(group_name, scope):
    if isinstance(scope, dict):
        explicit = scope.get('join_url') or scope.get('request_url') or scope.get('access_url')
        if explicit:
            return str(explicit)
    return 'https://groups.qualcomm.com/groups/' + str(group_name or '').strip()


def _live_status_access_groups_catalog(all_target_opts=None):
    """Public landing-page catalog of viewer groups and what they unlock."""
    cfg = LIVE_STATUS_VIEWER_GROUP_ACCESS or {}
    if not isinstance(cfg, dict) or not cfg:
        return []
    all_target_opts = all_target_opts or _all_targets_for_ui()
    by_target_upper = {str(r.get('target') or '').upper(): r for r in all_target_opts}
    business_units = get_business_units() or {}
    out = []
    for group_name, scope in sorted(cfg.items(), key=lambda item: str(item[0]).lower()):
        group_name = str(group_name or '').strip()
        if not group_name:
            continue
        if isinstance(scope, (list, tuple, set, str)):
            label = group_name
            bus_scope = _norm_access_list(scope)
            target_scope = set()
            pattern_scope = set()
            all_scope = False
        elif isinstance(scope, dict):
            label = str(scope.get('label') or scope.get('display_name') or group_name).strip()
            all_scope = bool(scope.get('all'))
            bus_scope = _norm_access_list(scope.get('bus') or scope.get('bu') or scope.get('business_units'))
            target_scope = _norm_access_list(scope.get('targets') or scope.get('target'))
            pattern_scope = _scope_target_patterns(scope)
        else:
            continue
        if all_scope or '*' in bus_scope or 'ALL' in bus_scope or '*' in target_scope or 'ALL' in target_scope or '*' in pattern_scope or 'ALL' in pattern_scope:
            bus_scope = {str(k).upper() for k in business_units.keys() if str(k).upper() != 'WEEKLY_QIPL_REPORTS'}
            target_scope = {str(r.get('target') or '').upper() for r in all_target_opts if r.get('target')}
            pattern_scope = set()

        bus_rows = []
        for bu_key in sorted(bus_scope):
            if bu_key == 'WEEKLY_QIPL_REPORTS':
                continue
            bu_info = business_units.get(bu_key) or business_units.get(bu_key.upper()) or {}
            bus_rows.append({
                'key': bu_key,
                'name': (bu_info or {}).get('display_name') or bu_key,
                'url': url_for('live_status_publish_bp.landing', bu_key=bu_key),
            })

        matched_target_uppers = set(target_scope)
        for row in all_target_opts:
            target = str(row.get('target') or '')
            target_upper = target.upper()
            if any(pattern and pattern in target_upper for pattern in pattern_scope):
                matched_target_uppers.add(target_upper)

        target_rows = []
        for target_upper in sorted(matched_target_uppers):
            row = by_target_upper.get(target_upper)
            target = (row or {}).get('target') or target_upper
            bu_key = str((row or {}).get('bu_key') or get_bu_for_target(target) or 'TARGET').upper()
            target_rows.append({
                'name': target,
                'bu_key': bu_key,
                'bu_name': (row or {}).get('bu_name') or bu_key,
                'url': url_for('live_status_publish_bp.live_status_target_by_bu', bu_key=bu_key, target_name=target),
            })
        out.append({
            'label': label,
            'group': group_name,
            'join_url': _live_status_group_join_url(group_name, scope),
            'bus': bus_rows,
            'targets': target_rows[:12],
            'target_count': len(target_rows),
            'patterns': sorted(pattern_scope),
        })

    return out



def _all_targets_for_ui():
    """Return active target options for UI dropdowns.

    Build Report should not use the process-wide BUSINESS_UNITS cache because
    that cache can include inactive/stale dashboard_status rows. A stale row can
    make one target appear under another display/product name (for example Maili
    showing Hawi). Build this list from active DB metadata each request.
    """
    try:
        metadata = load_metadata_config(active_only=True) or {}
        business_units = metadata.get('BUSINESS_UNITS', {}) or {}
        targets_config = metadata.get('TARGETS_CONFIG', {}) or {}
    except Exception:
        logger.warning('[TARGET OPTIONS] active metadata load failed; falling back to cached metadata', exc_info=True)
        business_units = get_business_units() or {}
        targets_config = {}

    rows = []
    seen = set()
    for bu_key, bu_info in sorted(business_units.items()):
        bu_key_upper = str(bu_key).upper()
        if bu_key_upper == 'WEEKLY_QIPL_REPORTS':
            continue
        if bu_key_upper == 'AUTO':
            targets = get_auto_target_keys({'BUSINESS_UNITS': business_units, 'TARGETS_CONFIG': targets_config})
        else:
            targets = list((bu_info or {}).get('targets') or [])
            if not targets:
                targets = get_targets_for_bu(bu_key_upper) or []
        for target in targets:
            target = str(target or '').strip()
            if not target:
                continue
            key = (bu_key_upper, target.lower())
            if key in seen:
                continue
            seen.add(key)
            info = targets_config.get(target) or next(
                (cfg for tk, cfg in targets_config.items() if str(tk).lower() == target.lower()),
                {},
            )
            display_name = str((info or {}).get('display_name') or '').strip()
            if not display_name:
                try:
                    display_name = get_display_name_for_target(target) or target
                except Exception:
                    display_name = target
            rows.append({
                'bu_key': bu_key_upper,
                'bu_name': (bu_info or {}).get('display_name', bu_key),
                'target': target,
                'display_name': display_name,
            })
    return rows



def _find_target_option(target_name):
    target_name = str(target_name or '').strip()
    if not target_name:
        return None
    for row in _all_targets_for_ui():
        if str(row.get('target')) == target_name:
            return row
    return None


def _is_active_job(job):
    return (job or {}).get('status') in ('draft', 'published')


def _job_type(job):
    return str((job or {}).get('job_type') or 'CRM').strip().upper()


def _is_core_deck_target(target_name: str) -> bool:
    """Core Deck is for Automotive/NORD PDT status targets.

    Some configs return AUTO, some AUTOMOTIVE, and some NORD targets are
    grouped under a product-specific BU. Keep this broader than exact AUTO so
    NORD_HQX/NORD_HGY pages show the Core Deck tab.
    auto_gen4.5 / 4.8.9.0 / 4.8.0.9 are not in TARGETS_CONFIG so
    get_bu_for_target returns None - match them explicitly by target name.
    """
    target = str(target_name or '').strip().upper()
    bu = str(get_bu_for_target(target_name) or '').strip().upper()
    return (
        bu in {'AUTO', 'AUTOMOTIVE', 'AUTO_TELEMATICS'}
        or target.startswith('NORD')
        or target.startswith('SECA')
        or 'NORD_' in target
        or 'NORD.' in target
        or 'SECA_' in target
        or 'SECA.' in target
        or target in {'AUTO_GEN4.5', 'AUTO_GEN45', '4.8.9.0', '4.8.0.9'}
        or target.startswith('AUTO_GEN4')
    )


def _find_existing_single_target_job(target_name, job_type='CRM'):
    target_name = str(target_name or '').strip()
    job_type = str(job_type or 'CRM').strip().upper()
    if not target_name:
        return None
    matches = []
    for job in list_jobs():
        targets = job.get('targets') or []
        if (_is_active_job(job) and _job_type(job) == job_type
                and len(targets) == 1 and str(targets[0]).lower() == target_name.lower()):
            matches.append(job)
    if not matches:
        return None
    # Prefer the latest updated row if casing or duplicate legacy rows exist.
    matches.sort(key=lambda row: str(row.get('updated_at') or row.get('published_at') or ''), reverse=True)
    return matches[0]


def _normal_live_status_tab(value, default='mtbf'):
    tab = str(value or '').strip().lower()
    return tab if tab in {'core', 'current', 'mtbf', 'weekly', 'opencrs', 'openjiras', 'buildreport'} else default


def _requested_live_status_tab(default='mtbf'):
    return _normal_live_status_tab(request.args.get('tab') or request.args.get('initial_tab'), default=default)


def _canonical_target_edit_url(target_name, tab=None):
    target_name = str(target_name or '').strip()
    bu_key = str(get_bu_for_target(target_name) or '').strip().upper() or 'TARGET'
    values = {'bu_key': bu_key, 'target_name': target_name}
    if tab:
        values['tab'] = _normal_live_status_tab(tab)
    return url_for('live_status_publish_bp.live_status_target_by_bu', **values)


def _canonical_target_editor_url(target_name, tab=None):
    return _canonical_target_edit_url(target_name, tab=tab)


def _render_current_report_editor(job, initial_tab=None):
    """Editors use live_status_publish_edit.html (the single canonical template)
    with can_edit=True, giving the full rich UI plus Save / Publish controls.
    """
    default_tab = 'mtbf'
    return _render_published_full_page(job, initial_tab=_normal_live_status_tab(initial_tab, default_tab), suppress_top_redirect=True)



def _count_active_eng_jobs(target_name):
    target_name = str(target_name or '').strip()
    return sum(
        1 for job in list_jobs()
        if _is_active_job(job)
        and _job_type(job) == 'ENG'
        and target_name in [str(t) for t in (job.get('targets') or [])]
    )


def _published_display_rows(rows, published_at):
    """Return display copies with final hours/MTBF calculated from publish time."""
    from datetime import datetime, timezone

    pub_ms = 0.0
    try:
        raw = str(published_at or '').strip()
        if raw:
            pub_dt = datetime.fromisoformat(raw.replace('Z', '+00:00'))
            if pub_dt.tzinfo is None:
                pub_dt = pub_dt.replace(tzinfo=timezone.utc)
            pub_ms = pub_dt.timestamp()
    except Exception:
        pub_ms = 0.0
    elapsed_hours = max(0.0, (datetime.now(timezone.utc).timestamp() - pub_ms) / 3600.0) if pub_ms else 0.0

    def _f(v):
        try:
            return float(v or 0)
        except Exception:
            return 0.0

    out = []
    for src in rows or []:
        r = dict(src or {})
        raw_hours = _f(r.get('hours'))
        reduction = max(0.0, min(100.0, _f(r.get('reduction_percent'))))
        crashes = _f(r.get('crashes'))
        has_calc_input = raw_hours > 0 or reduction > 0
        if has_calc_input:
            devices = _f(r.get('device_count')) or 1.0
            final_hours = raw_hours + (elapsed_hours * (1.0 - (reduction / 100.0)) * devices)
            r['display_hours'] = f'{round(final_hours, 1):.1f}'
            r['display_mtbf'] = f'{round(final_hours / crashes, 1):.1f}' if crashes > 0 and final_hours > 0 else 'NA'
        else:
            r['display_hours'] = 'NA'
            r['display_mtbf'] = 'NA'
        out.append(r)
    return out


def _read_mtbf_excel_rows(target_name, sheet_name_override=None):
    """Read configured MTBF Excel rows for public published Live Status widgets.

    For Compute MTBF workbooks, each project/view (for example Glymur,
    Mahua) is represented by a workbook sheet. This mirrors the MTBF Excel
    page selector instead of filtering mixed rows client-side.
    """
    import os
    import openpyxl
    from datetime import datetime as _dtt, date as _ddate
    from dashboard_routes import _get_target_excel_config, _normalize_excel_path

    cfg = (_get_target_excel_config(target_name) or {}).get('mtbf', {})
    excel_path = cfg.get('excel_path', '')
    sheet_name = str(sheet_name_override or cfg.get('sheet_name') or '').strip()
    if not excel_path:
        return {'success': False, 'message': 'Excel not configured.', 'headers': [], 'rows': []}
    path = _normalize_excel_path(excel_path)
    if not os.path.exists(path):
        return {'success': False, 'message': f'File not found: {path}', 'headers': [], 'rows': []}
    wb = openpyxl.load_workbook(path, data_only=True)
    actual_sheet = sheet_name if sheet_name in wb.sheetnames else ''
    if not actual_sheet and sheet_name:
        actual_sheet = next((s for s in wb.sheetnames if s.strip().lower() == sheet_name.lower()), '')
    if not actual_sheet and wb.sheetnames:
        actual_sheet = wb.sheetnames[0]
    if not actual_sheet:
        return {'success': False, 'message': 'No sheets found in workbook.', 'headers': [], 'rows': []}
    def _norm_header(value):
        import re as _re
        return _re.sub(r'\s+', ' ', _re.sub(r'[^a-z0-9 ]+', '', str(value or '').strip().lower().replace('_', ' ').replace('-', ' '))).strip()

    def _sheet_payload(sheet_title):
        ws = wb[sheet_title]

        # Match the dashboard MTBF table reader: flatten merged cells so the
        # published page sees the same effective values as the main MTBF page.
        merge_map = {}
        for mr in list(ws.merged_cells.ranges):
            val = ws.cell(mr.min_row, mr.min_col).value
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    merge_map[(row, col)] = val

        def _cv(row, col):
            value = merge_map.get((row, col), ws.cell(row, col).value)
            if isinstance(value, _dtt):
                return value.strftime('%Y-%m-%d')
            if isinstance(value, _ddate):
                return value.strftime('%Y-%m-%d')
            return '' if value is None else str(value).strip()

        header_tokens = {
            'hours', 'total hours', 'tested hours',
            'crashes', 'total crashes', 'crash count',
            'mtbf', 'product mtbf', 'qc mtbf',
            'meta id', 'meta', 'build', 'builds', 'builds full id',
        }
        best_header_row = 1
        best_score = -1
        scan_rows = min(ws.max_row or 1, 20)
        for rr in range(1, scan_rows + 1):
            vals = [_cv(rr, c) for c in range(1, (ws.max_column or 1) + 1)]
            norm_vals = {_norm_header(v) for v in vals if str(v).strip()}
            score = len(norm_vals & header_tokens)
            if score > best_score:
                best_score = score
                best_header_row = rr

        headers = [_cv(best_header_row, c) for c in range(1, (ws.max_column or 1) + 1)]
        rows = []
        for r in range(best_header_row + 1, (ws.max_row or 1) + 1):
            vals = [_cv(r, c) for c in range(1, (ws.max_column or 1) + 1)]
            if any(str(v).strip() for v in vals):
                rows.append({'excel_row': r, 'values': vals})
        return {
            'success': True,
            'headers': headers,
            'rows': rows,
            'sheet_name': sheet_title,
            'sheet_names': wb.sheetnames,
            'excel_path': excel_path,
            'header_row': best_header_row,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'detected_header_score': best_score,
        }

    data = _sheet_payload(actual_sheet)
    if not sheet_name_override and not data.get('rows'):
        # If the configured/default sheet has only headers, look for another
        # sheet in the same workbook that has MTBF headers plus real data rows.
        best = data
        for candidate in wb.sheetnames:
            if candidate == actual_sheet:
                continue
            cand = _sheet_payload(candidate)
            if len(cand.get('rows') or []) > len(best.get('rows') or []):
                best = cand
        data = best
    return data


def _get_published_job_for_api(job_id):
    job = get_job(job_id)
    if not job:
        return None, ({'ok': False, 'error': 'Job not found'}, 404)
    if job.get('status') != 'published' and not (current_user.is_authenticated and _target_group_access()):
        return None, ({'ok': False, 'error': 'Access denied'}, 403)
    requested_target = (request.args.get('target') or request.args.get('target_name') or '').strip()
    targets = [str(t or '').strip() for t in (job.get('targets') or []) if str(t or '').strip()]
    if requested_target:
        canonical_target = next((t for t in targets if t.lower() == requested_target.lower()), '')
        if canonical_target:
            job = dict(job)
            job['targets'] = [canonical_target] + [t for t in targets if t != canonical_target]
    return job, None


def _get_target_report_job_for_api(target_name):
    """Resolve a target-based public/report API to its backing job when needed."""
    target = str(target_name or '').strip()
    if not target:
        return None, ({'ok': False, 'error': 'Target is required'}, 400)
    job = _find_published_job_for_target(target)
    if not job and current_user.is_authenticated and _target_group_access():
        job = _find_existing_single_target_job(target, 'CRM')
    if job:
        targets = [str(t or '').strip() for t in (job.get('targets') or []) if str(t or '').strip()]
        canonical_target = next((t for t in targets if t.lower() == target.lower()), target)
        job = dict(job)
        job['targets'] = [canonical_target] + [t for t in targets if t != canonical_target]
        return job, None
    return {'id': '', 'status': 'published', 'targets': [target], 'published_rows': [], 'draft_rows': []}, None



def _lsp_safe_int(value, default=0):
    try:
        return int(value or default)
    except Exception:
        return default


def _lsp_parse_iso_dt(value):
    from datetime import datetime
    text = str(value or '').strip().replace('Z', '')
    if not text:
        return datetime.min
    try:
        return datetime.fromisoformat(text)
    except Exception:
        return datetime.min


def _lsp_extract_meta_from_jql(value):
    """Best-effort meta/build extraction for saved-filter JQL metadata badges."""
    text = str(value or '')
    patterns = [
        r'\b[A-Z][A-Z0-9_.]*\.LE\.[0-9.]+-[0-9]{3,6}-[A-Z0-9_.-]+(?:-[0-9]+)?\b',
        r'\b[A-Z][A-Z0-9_.-]+-[0-9]{3,6}-[A-Z0-9_.-]+(?:-[0-9]+)?\b',
        r'\b(?:META|BUILD)[-_ ]?0*([0-9]{3,6})\b',
    ]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.I)
        if m:
            return m.group(0)
    quoted = re.findall(r'"([^"\r\n]{6,160})"', text)
    return next((q for q in quoted if re.search(r'\d{3,6}', q) and re.search(r'[A-Za-z]', q)), '')


def _lsp_extract_meta_from_report(cached):
    """Extract a displayed meta/build from cached report rows when JQL is only filter=N.

    Saved-filter tabs often store/display only ``filter = 346152``. In that case
    there is no build text in the JQL itself, so derive the badge from the
    generated report rows exactly like WBC's cache metadata does.
    """
    if not isinstance(cached, dict):
        return ''
    candidates = []
    for key in ('build_id', 'meta_id', 'metabuild', 'build', 'build_name', 'build_full', 'display_build'):
        value = str(cached.get(key) or '').strip()
        if value:
            candidates.append(value)
    row_sets = [
        cached.get('rows'),
        cached.get('flat_rows'),
        cached.get('detail_rows'),
        cached.get('hierarchical_report'),
        cached.get('jiras'),
    ]
    for rows in row_sets:
        if isinstance(rows, dict):
            rows = list(rows.values())
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            for key in (
                'build_id', 'meta_id', 'metabuild', 'matched_build', 'build',
                'build_name', 'build_full', 'display_build', 'summary', 'title',
                'jira_title',
            ):
                value = str(row.get(key) or '').strip()
                if value:
                    candidates.append(value)
    for value in candidates:
        meta = _lsp_extract_meta_from_jql(value)
        if meta:
            return meta
    return next((v for v in candidates if v and not re.match(r'^\s*filter\s*=?\s*\d+\s*$', v, flags=re.I)), '')


def _lsp_saved_jql_cache_meta(cached, tab=None, resolved_jql=''):
    """Return WBC-style Last Run / Next Run / Meta-from-filter metadata."""
    from datetime import datetime, timedelta
    cached = cached if isinstance(cached, dict) else {}
    tab = tab if isinstance(tab, dict) else {}
    ttl_minutes = max(1, _lsp_safe_int(cached.get('cache_ttl_minutes') or cached.get('refresh_minutes') or tab.get('refresh_minutes') or 30, 30))
    ttl = timedelta(minutes=ttl_minutes)
    generated_at = _lsp_parse_iso_dt(cached.get('generated_at') or cached.get('last_run_at') or tab.get('last_run_at') or tab.get('last_run'))
    next_run_at = _lsp_parse_iso_dt(cached.get('next_run_at') or cached.get('next_auto_refresh_at') or tab.get('next_run_at'))
    next_run = next_run_at if next_run_at != datetime.min else (generated_at + ttl if generated_at != datetime.min else datetime.min)
    expired = bool(next_run != datetime.min and datetime.utcnow() >= next_run)
    rows = cached.get('rows') or cached.get('flat_rows') or cached.get('detail_rows') or []
    # Priority 1: extract build/meta name from the RESOLVED JQL (same as WBC approach).
    # The resolved JQL contains summary ~ "BuildName" clauses — extract the build name from those.
    # This is the correct approach: filter ID is just the lookup key, not the meta label.
    meta_from_filter = (
        _lsp_extract_meta_from_jql(resolved_jql)
        or _lsp_extract_meta_from_jql(cached.get('resolved_jql') or cached.get('jql') or cached.get('raw_jql'))
        or _lsp_extract_meta_from_jql(tab.get('name'))
    )
    # Priority 2: if no build name found in resolved JQL, fall back to filter ID as label
    if not meta_from_filter:
        _raw_jql_meta = str(tab.get('jql') or '').strip()
        _cached_jql_meta = str(cached.get('raw_jql') or cached.get('jql') or '').strip()
        _filter_ids_meta = (
            re.findall(r'filter\s*=\s*(\d+)', _raw_jql_meta, flags=re.I)
            or re.findall(r'filter\s*=\s*(\d+)', _cached_jql_meta, flags=re.I)
            or re.findall(r'filter\s*=\s*(\d+)', str(resolved_jql or ''), flags=re.I)
        )
        if _filter_ids_meta:
            meta_from_filter = 'Filter: ' + ', '.join(_filter_ids_meta)
    return {
        'has_cached_report': bool(cached),
        'cached_report_stale': expired,
        'last_run_at': cached.get('generated_at') or cached.get('last_run_at') or tab.get('last_run_at') or tab.get('last_run') or '',
        'next_run_at': next_run.isoformat() + 'Z' if next_run != datetime.min else '',
        'next_auto_refresh_at': next_run.isoformat() + 'Z' if next_run != datetime.min else '',
        'cache_ttl_minutes': ttl_minutes,
        'cached_row_count': _lsp_safe_int(cached.get('row_count') or cached.get('count') or len(rows)) if cached else 0,
        'cached_cr_count': _lsp_safe_int(cached.get('valid_cr_count') or cached.get('cr_count')) if cached else 0,
        'cached_mapped_jira_count': _lsp_safe_int(cached.get('valid_mapped_jira_count') or cached.get('mapped_jira_count')) if cached else 0,
        'cached_open_jira_count': _lsp_safe_int(cached.get('valid_open_jira_count') or cached.get('open_jira_count')) if cached else 0,
        'cached_jira_count': _lsp_safe_int(cached.get('valid_jira_count') or cached.get('jira_count') or cached.get('total_count')) if cached else 0,
        'cached_invalid_count': _lsp_safe_int(cached.get('invalid_count')) if cached else 0,
        'cache_status': 'stale' if expired else ('cached' if cached else 'not_run'),
        'meta_from_filter': meta_from_filter or '',
        'build_id': meta_from_filter or '',
    }


def _saved_jql_domain_or_400(value, target_name: str = ""):
    """Validate and normalise the domain parameter.

    Accepts ANY domain that is valid for the target — read from the same
    mtbf_domains.json that the MTBF Trend page uses.  This means CSP,
    SAFE-IVI, NONSAFE-IVI, HQX, HGY, ADAS, FLEX, IVI and any custom
    domain added by the user are all accepted.

    Only falls back to the target-name key if the domain is completely
    unknown (empty string).
    """
    domain = str(value or '').strip().upper()
    if not domain:
        if target_name:
            safe = str(target_name).strip().upper().replace(' ', '_')
            return safe, None
        return '', ({'ok': False, 'error': 'domain parameter is required'}, 400)

    # Accept any non-empty domain string — validation is done by the JS
    # which only shows domains returned by _get_target_domains().
    # Sanitise: allow letters, digits, underscore, hyphen only.
    import re as _re
    if _re.match(r'^[A-Z0-9_\-]+$', domain):
        return domain, None

    # Fallback for unexpected characters
    if target_name:
        safe = str(target_name).strip().upper().replace(' ', '_')
        return safe, None
    return '', ({'ok': False, 'error': f'Invalid domain: {domain}'}, 400)



@live_status_publish_bp.route('/api/live_status/targets/<target_name>/saved_jql_tabs', methods=['GET'])
@login_required
def api_live_status_saved_jql_tabs(target_name):
    if not _can_view_live_status_target(target_name):
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    domain, err = _saved_jql_domain_or_400(request.args.get('domain'), target_name)
    if err:
        return jsonify(err[0]), err[1]
    tabs = []
    for tab in list_saved_jql_tabs(target_name, domain):
        row = dict(tab)
        raw_jql = str(row.get('jql') or '').strip()
        resolved_jql = raw_jql
        filter_id = ''
        filter_resolved = False
        filter_error = ''
        try:
            from dashboard_routes import _jira_filter_id_from_jql, _resolve_jira_filter_jql
            filter_id = str(row.get('filter_id') or _jira_filter_id_from_jql(raw_jql) or '').strip()
            if filter_id:
                latest = str(_resolve_jira_filter_jql(filter_id) or '').strip()
                if latest:
                    resolved_jql = latest
                    filter_resolved = True
                else:
                    filter_error = 'Filter lookup returned empty JQL'
        except Exception as exc:
            filter_error = str(exc)
        cached = get_saved_jql_cached_report_raw(target_name, domain, row.get('id')) or {}
        row['raw_jql'] = raw_jql
        row['resolved_jql'] = resolved_jql
        row['filter_id'] = filter_id
        row['filter_resolved'] = filter_resolved
        row['filter_error'] = filter_error
        row.update(_lsp_saved_jql_cache_meta(cached, row, resolved_jql))
        tabs.append(row)
    return jsonify({'ok': True, 'tabs': tabs, 'target': target_name, 'domain': domain})


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/saved_jql_tabs', methods=['POST'])
@login_required
def api_live_status_saved_jql_tabs_save(target_name):
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    payload = request.get_json(silent=True) or {}
    domain, err = _saved_jql_domain_or_400(payload.get('domain'), target_name)
    if err:
        return jsonify(err[0]), err[1]
    try:
        raw_jql = str(payload.get('jql') or '').strip()
        explicit_filter_id = str(payload.get('filter_id') or '').strip()
        if not explicit_filter_id:
            try:
                from dashboard_routes import _jira_filter_id_from_jql
                explicit_filter_id = str(_jira_filter_id_from_jql(raw_jql) or '').strip()
            except Exception:
                explicit_filter_id = ''
        row = save_saved_jql_tab(
            target_name,
            domain,
            tab_id=payload.get('id'),
            name=payload.get('name'),
            jql=raw_jql,
            username=getattr(current_user, 'id', 'unknown'),
            filter_id=explicit_filter_id,
        )
        return jsonify({'ok': True, 'tab': row, 'tabs': list_saved_jql_tabs(target_name, domain)})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/saved_jql_tabs/<tab_id>', methods=['DELETE'])
@login_required
def api_live_status_saved_jql_tabs_delete(target_name, tab_id):
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    domain, err = _saved_jql_domain_or_400(request.args.get('domain'), target_name)
    if err:
        return jsonify(err[0]), err[1]
    deleted = delete_saved_jql_tab(target_name, domain, tab_id)
    return jsonify({'ok': True, 'deleted': bool(deleted), 'tabs': list_saved_jql_tabs(target_name, domain)})


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/saved_jql_tabs/<tab_id>/report', methods=['GET'])
@login_required
def api_live_status_saved_jql_tab_report(target_name, tab_id):
    if not _can_view_live_status_target(target_name):
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    domain, err = _saved_jql_domain_or_400(request.args.get('domain'), target_name)
    if err:
        return jsonify(err[0]), err[1]
    tab = get_saved_jql_tab(target_name, domain, tab_id)
    if not tab:
        return jsonify({'ok': False, 'error': 'Saved tab not found'}), 404

    force = str(request.args.get('force') or '').lower() in ('1', 'true', 'yes')

    # Saved-JQL tabs are scheduled centrally. Normal page opens should behave like
    # WBC/current-running-build: return the latest available cache immediately so
    # viewers never block on a long JIRA traversal. "Run Now" passes force=1 and
    # executes a foreground refresh.
    raw_jql = str(tab.get('jql') or '').strip()
    jql = raw_jql
    filter_id = ''
    filter_resolved = False
    filter_error = ''
    try:
        from dashboard_routes import _jira_filter_id_from_jql, _resolve_jira_filter_jql
        filter_id = str(tab.get('filter_id') or _jira_filter_id_from_jql(raw_jql) or '').strip()
        if filter_id:
            resolved_jql = str(_resolve_jira_filter_jql(filter_id) or '').strip()
            if resolved_jql:
                jql = resolved_jql
                filter_resolved = True
            else:
                filter_error = 'Filter lookup returned empty JQL'
    except Exception as exc:
        filter_error = str(exc)
        logger.warning('[SAVED JQL REPORT] filter resolve failed for %s/%s: %s', target_name, tab_id, exc)

    if not force:
        cached = get_saved_jql_cached_report(target_name, domain, tab_id) or get_saved_jql_cached_report_raw(target_name, domain, tab_id)
        if cached:
            cached = dict(cached)
            cached_filter_id = str(cached.get('filter_id') or '').strip()
            cached_resolved_jql = str(cached.get('resolved_jql') or '').strip()
            cached_effective_jql = str(cached.get('jql') or '').strip()
            cached_raw_jql = str(cached.get('raw_jql') or '').strip()
            # Match WBC behavior: for saved-filter tabs, the filter id itself is
            # not enough to trust an old cache because the Jira filter definition
            # can change while the id stays constant.  Reuse cache only when the
            # currently resolved JQL exactly matches cached resolved/effective JQL.
            if filter_id:
                cache_matches_filter = (
                    cached_resolved_jql == jql
                    or (
                        cached_effective_jql == jql
                        and cached_effective_jql not in (raw_jql, filter_id, f'filter = {filter_id}')
                    )
                )
            else:
                cache_matches_filter = (
                    cached_resolved_jql == jql
                    or cached_effective_jql == jql
                    or cached_raw_jql == raw_jql
                )
            if cache_matches_filter:
                cached['ok'] = True
                cached['from_cache'] = True
                cached['tab'] = tab
                cached['jql'] = jql
                cached['raw_jql'] = raw_jql
                cached['resolved_jql'] = jql
                cached['filter_id'] = filter_id
                cached['filter_resolved'] = filter_resolved
                cached['filter_error'] = filter_error
                cached.update(_lsp_saved_jql_cache_meta(cached, tab, jql))
                # If the scheduled next-run time is already overdue, do not keep
                # returning the stale raw cache.  Refresh now so the displayed
                # Last/Next Run behaves like the scheduler/WBC expectation.
                if cached.get('cached_report_stale'):
                    logger.info(
                        '[SAVED JQL REPORT] cache overdue for %s/%s; refreshing now (last=%s next=%s)',
                        target_name, tab_id, cached.get('last_run_at'), cached.get('next_run_at')
                    )
                    force = True
                else:
                    return jsonify(cached)
            logger.info(
                '[SAVED JQL REPORT] ignoring stale cache for %s/%s: cached_filter=%s current_filter=%s',
                target_name, tab_id, cached_filter_id, filter_id
            )
        if not force:
            # Do not run a foreground JIRA traversal from the initial page load
            # when no usable cache exists yet. The central scheduler will
            # populate this cache; editors can still use "Run Now" (force=1).
            return jsonify({
                'ok': True,
                'from_cache': False,
                'cache_status': 'pending',
                'run_error': 'Report cache is not ready yet. The background scheduler will generate it automatically; use Run Now to generate immediately.',
                'tab': tab,
                'target': target_name,
                'domain': domain,
                'jql': jql,
                'raw_jql': raw_jql,
                'resolved_jql': jql,
                'filter_id': filter_id,
                'filter_resolved': filter_resolved,
                'filter_error': filter_error,
                'rows': [],
                'flat_rows': [],
                'detail_rows': [],
                'count': 0,
                'total_count': 0,
                **_lsp_saved_jql_cache_meta({}, tab, jql),
            })

    # Forced/manual refresh — run the full consolidated report in foreground.
    # The central scheduler will keep this same cache fresh in the background.

    # Crash-type filter from query param (default: all types)
    _ct_raw = str(request.args.get('crash_types') or '').strip()
    crash_types = (
        {c.strip().lower() for c in _ct_raw.split(',') if c.strip()}
        if _ct_raw else {'system', 'ssr', 'process', 'open_jira'}
    )

    run_error = None
    payload_extra = {}

    if jql:
        progress = None
        progress_job_id = str(request.args.get('progress_job_id') or '').strip()
        try:
            import os as _os, sys as _sys
            _scripts_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'scripts')
            if _scripts_dir not in _sys.path:
                _sys.path.insert(0, _scripts_dir)
            from fetch_consolidated_report import run_consolidated_report, register_progress, unregister_progress
            from automotive_live_view_stats_routes import (
                _auto_flatten_consolidated_report,
            )

            if progress_job_id:
                progress = register_progress(progress_job_id)
                progress.update(stage='start', total=100, done=5, message='Resolving saved JQL and starting JIRA report...')

            try:
                from live_view_saved_jql_service import _extract_build_id as _sjql_extract_build_id
                _sjql_build_id = _sjql_extract_build_id(jql) or _sjql_extract_build_id(raw_jql) or _sjql_extract_build_id(tab.get('name'))
            except Exception:
                _sjql_build_id = ''

            if progress:
                progress.update(stage='jira', total=100, done=12, message='Running JIRA query for saved filter...')
            raw_report = run_consolidated_report(
                build_ids=[_sjql_build_id] if _sjql_build_id else [],
                filter_id=filter_id or JIRA_PDT_FILTER_ID,
                traverse=True,
                enrich_orbit=True,
                target_name=target_name,
                custom_jql=jql,
                progress=progress,
            ) or {}
            if progress:
                progress.update(stage='flatten', total=100, done=92, message='Preparing report rows...')

            all_rows = _auto_flatten_consolidated_report(raw_report)
            detail_rows = [
                row for row in all_rows
                if str(row.get('crash_type') or 'system').lower() in crash_types
            ]

            payload_extra = {
                'hierarchical_report':   raw_report.get('hierarchical_report') or [],
                'jiras':                 raw_report.get('jiras') or [],
                'cr_index':              raw_report.get('cr_index') or {},
                'summary':               raw_report.get('summary') or {},
                'detail_rows':           detail_rows,
                'count':                 len(detail_rows),
                'total_count':           len(all_rows),
                'crash_types_used':      sorted(crash_types),
                'crash_types_available': ['system', 'ssr', 'process', 'open_jira'],
            }
            if progress:
                progress.update(stage='done', total=100, done=100, message=f'Report ready: {len(detail_rows)} rows.')
        except Exception as exc:
            logger.warning('[SAVED JQL REPORT] consolidated report failed for %s/%s: %s', target_name, tab_id, exc)
            run_error = str(exc)
            # Fallback: plain JIRA search so the user still gets something
            try:
                from config import JIRA_PASSWORD, JIRA_SERVER_ENDPOINT, JIRA_USER
                from fetch_consolidated_report import connect_jira
                jira_obj = connect_jira(JIRA_USER, JIRA_PASSWORD, JIRA_SERVER_ENDPOINT)
                rows_out = []
                start = 0
                while True:
                    results = jira_obj.search_issues(
                        jql, startAt=start, maxResults=100,
                        fields='summary,status,assignee,priority,created,updated,issuetype'
                    )
                    issues = list(results or [])
                    if not issues:
                        break
                    for issue in issues:
                        f = issue.fields
                        rows_out.append({
                            'key':       str(issue.key or ''),
                            'summary':   str(getattr(f, 'summary', '') or ''),
                            'status':    str(getattr(getattr(f, 'status', None), 'name', '') or ''),
                            'assignee':  str(getattr(getattr(f, 'assignee', None), 'displayName', '') or ''),
                            'priority':  str(getattr(getattr(f, 'priority', None), 'name', '') or ''),
                            'issuetype': str(getattr(getattr(f, 'issuetype', None), 'name', '') or ''),
                            'created':   str(getattr(f, 'created', '') or '')[:10],
                            'updated':   str(getattr(f, 'updated', '') or '')[:10],
                        })
                    start += len(issues)
                    if start >= getattr(results, 'total', 0):
                        break
                payload_extra = {'rows': rows_out, 'count': len(rows_out), 'fallback': True}
            except Exception as exc2:
                logger.warning('[SAVED JQL REPORT] fallback search also failed: %s', exc2)
                payload_extra = {'rows': [], 'count': 0, 'fallback': True}
        finally:
            if progress_job_id:
                try:
                    unregister_progress(progress_job_id)
                except Exception:
                    pass

    payload = {
        'tab':             tab,
        'jql':             jql,
        'raw_jql':         raw_jql,
        'resolved_jql':    jql,
        'filter_id':       filter_id,
        'filter_resolved': filter_resolved,
        'filter_error':    filter_error,
        'target':          target_name,
        'domain':          domain,
        'from_cache':      False,
        'build_id':        _sjql_build_id if '_sjql_build_id' in locals() else _lsp_extract_meta_from_jql(jql),
    }
    payload.update(payload_extra)
    if run_error:
        payload['run_error'] = run_error
    stored = set_saved_jql_cached_report(target_name, domain, tab_id, payload)
    stored['ok'] = True
    stored['tab'] = tab
    stored.update(_lsp_saved_jql_cache_meta(stored, tab, jql))
    return jsonify(stored)


def _published_jobs_by_bu_context():
    """Build context for the external PDT MTBF published reports page."""
    all_target_opts = _all_targets_for_ui()
    target_to_bu = {
        row['target']: {'bu_key': str(row['bu_key']).upper(), 'bu_name': row['bu_name']}
        for row in all_target_opts
    }
    hidden_bus = {'WEEKLY_QIPL_REPORTS', 'HWPDT'}
    published_jobs = []
    for job in list_jobs():
        if job.get('status') != 'published' or not job.get('public_token'):
            continue
        first_target = (job.get('targets') or [''])[0]
        info = target_to_bu.get(first_target, {})
        bu_key = (info.get('bu_key') or 'OTHER').upper()
        if bu_key in hidden_bus:
            continue
        rows = job.get('published_rows') or []
        running_rows = [r for r in rows if str((r or {}).get('run_status', '')).lower() == 'running']
        build_rows = [
            r for r in rows
            if (r or {}).get('builds_tab') or str((r or {}).get('run_status', '')).lower() in ('builds', 'stopped', 'completed')
        ]
        enriched = dict(job)
        enriched['_bu_key'] = bu_key
        enriched['_bu_name'] = info.get('bu_name') or bu_key
        enriched['_target'] = first_target or (job.get('name') or '')
        enriched['_job_type'] = _job_type(job)
        enriched['_row_count'] = len(rows)
        enriched['_running_count'] = len(running_rows)
        enriched['_build_count'] = len(build_rows) if build_rows else len(rows)
        published_jobs.append(enriched)

    seen_bus = {}
    for job in published_jobs:
        seen_bus.setdefault(job['_bu_key'], job['_bu_name'])
    bu_list = [
        {'key': key, 'name': name, 'count': sum(1 for j in published_jobs if j['_bu_key'] == key)}
        for key, name in seen_bus.items()
    ]
    bu_list.sort(key=lambda row: (0 if row['key'] == 'MOBILE' else 1, row['name'].lower()))
    default_bu = 'MOBILE' if any(row['key'] == 'MOBILE' for row in bu_list) else (bu_list[0]['key'] if bu_list else '')
    selected_bu = (request.args.get('bu') or request.args.get('bu_key') or default_bu or '').strip().upper()
    return {
        'jobs': published_jobs,
        'bu_list': bu_list,
        'target_options': all_target_opts,
        'default_bu': default_bu,
        'selected_bu': selected_bu,
        'can_edit': current_user.is_authenticated and _target_group_access(),
    }


def _find_published_job_for_target(target_name):
    wanted = str(target_name or '').strip().lower()
    if not wanted:
        return None
    matches = []
    for job in list_jobs():
        if job.get('status') != 'published' or not job.get('public_token'):
            continue
        targets = [str(t or '').strip() for t in (job.get('targets') or [])]
        if any(t.lower() == wanted for t in targets):
            matches.append(job)
    if not matches:
        return None
    matches.sort(key=lambda row: str(row.get('published_at') or row.get('updated_at') or ''), reverse=True)
    return matches[0]




@live_status_publish_bp.route('/api/live_status/meta_jiras_json/<target_name>/<meta_id>')
@login_required
def meta_jiras_json(target_name, meta_id):
    """JSON list of JIRAs for a meta_id - used by the Exclude JIRAs modal in the MTBF table."""
    from dashboard_common import get_mysql_connection_db, fq_table_for_target
    conn = None
    cursor = None
    try:
        conn = get_mysql_connection_db()
        cursor = conn.cursor(dictionary=True)
        j_table = fq_table_for_target(target_name, 'jiras')
        o_table = fq_table_for_target(target_name, 'openjiras')
        like = '%' + meta_id + '%'
        def _tbl_ok(fq):
            n = fq.replace('`', '')
            try:
                s, t = n.split('.', 1)
            except ValueError:
                return True
            cursor.execute(
                'SELECT 1 FROM information_schema.tables '
                'WHERE table_schema=%s AND table_name=%s LIMIT 1', (s, t)
            )
            return cursor.fetchone() is not None
        if _tbl_ok(o_table):
            cursor.execute(
                f'SELECT j.stability_ticket AS jira_key, j.jira_title AS title,'
                f' j.serial_no, j.metabuild AS matched_build'
                f' FROM {j_table} j WHERE j.metabuild LIKE %s'
                f' UNION'
                f' SELECT o.stability_ticket, o.jira_title, o.serial_no, o.metabuild'
                f' FROM {o_table} o WHERE o.metabuild LIKE %s'
                f' ORDER BY jira_key',
                (like, like)
            )
        else:
            cursor.execute(
                f'SELECT stability_ticket AS jira_key, jira_title AS title,'
                f' serial_no, metabuild AS matched_build'
                f' FROM {j_table} WHERE metabuild LIKE %s ORDER BY jira_key',
                (like,)
            )
        rows = cursor.fetchall() or []
        jiras = [{
            'key': r.get('jira_key') or '',
            'title': r.get('title') or '',
            'serial_no': r.get('serial_no') or '',
            'matched_build': r.get('matched_build') or ''
        } for r in rows]
        return jsonify({'ok': True, 'meta_id': meta_id, 'jiras': jiras})
    except Exception as exc:
        logger.exception('[META JIRAS JSON] %s', exc)
        return jsonify({'ok': False, 'error': str(exc), 'jiras': []}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def _render_target_status_page(target_name, initial_tab=None):
    """Render the canonical per-target Live Status page after access is checked."""
    job = _find_published_job_for_target(target_name)

    if not job:
        if current_user.is_authenticated and _target_group_access():
            job = _find_existing_single_target_job(target_name, 'CRM')
        if not job:
            return render_template('coming_soon_template.html',
                                   title='Live Status',
                                   message='No report found for this target.'), 404

    if current_user.is_authenticated:
        try:
            update_viewer_heartbeat(job.get('id'), getattr(current_user, 'id', 'viewer'))
            job = get_job(job.get('id')) or job
        except Exception:
            pass

    targets = [str(t or '').strip() for t in (job.get('targets') or []) if str(t or '').strip()]
    canonical_target = next((t for t in targets if t.lower() == str(target_name or '').strip().lower()), target_name)
    if canonical_target and canonical_target in targets and (not targets or targets[0] != canonical_target):
        job = dict(job)
        job['targets'] = [canonical_target] + [t for t in targets if t != canonical_target]
    return _render_published_full_page(job, initial_tab or _requested_live_status_tab('mtbf'))



@live_status_publish_bp.route('/live_status_view/AUTO/<target_name>/sp')
@live_status_publish_bp.route('/live_status_view/auto/<target_name>/sp')
@login_required
def live_status_view_sp_page(target_name):
    """SP-aware MTBF page for Auto Gen5 targets (nord_hqx / nord_hgy)."""
    target_name = str(target_name or '').strip()
    if not target_name:
        return redirect(url_for('live_status_publish_bp.landing'))
    if not (current_user.is_authenticated and (
        _target_group_access() or _can_view_live_status_target(target_name)
    )):
        return render_template(
            'coming_soon_template.html',
            title='Live Status SP',
            message='You do not have access to this page.'
        ), 403
    from dashboard_common import get_target_info as _gti
    info = _gti(target_name) or {}
    display_name = str(info.get('display_name') or target_name).upper()
    can_edit = current_user.is_authenticated and _target_group_access()
    return render_template(
        'live_status_view_sp.html',
        target_name=target_name,
        display_name=display_name,
        can_edit=can_edit,
    )


@live_status_publish_bp.route('/pdt/<target_name>/ext_status')

@live_status_publish_bp.route('/pdt/<target_name>/ext-status')
def pdt_target_ext_status(target_name):
    """Legacy URL. Use /live_status_view/<BU>/<target> instead."""
    return redirect(_canonical_target_edit_url(target_name))


@live_status_publish_bp.route('/build-report')
@live_status_publish_bp.route('/build_report')
@login_required
def build_report_standalone():
    """Standalone Build Report page for generated-build JQL or direct JQL runs."""
    return render_template(
        'build_report_standalone.html',
        target_options=_all_targets_for_ui(),
        jira_pdt_filter_id=JIRA_PDT_FILTER_ID,
    )


@live_status_publish_bp.route('/api/build_report/running_builds', methods=['GET'])
@login_required
def api_build_report_running_builds():
    """List active builds from the local Axiom cache table.

    The Build Report page does not call Axiom directly. It reads
    pdt_stats_dashboard.axiom_job_summary, then filters by the selected target.
    """
    from datetime import date as _date, datetime as _datetime
    from dashboard_common import get_mysql_connection_db, get_target_info, update_global_targets_config, fq_table_for_target, get_schema_for_bu



    def _ser(value):
        if isinstance(value, (_datetime, _date)):
            return value.strftime('%Y-%m-%d %H:%M:%S') if isinstance(value, _datetime) else value.isoformat()
        return '' if value is None else str(value)

    def _chip_ids(value):
        import json as _json
        if value in (None, ''):
            return []
        if isinstance(value, (list, tuple, set)):
            raw = list(value)
        else:
            text = str(value or '').strip()
            if not text:
                return []
            try:
                parsed = _json.loads(text)
                raw = parsed if isinstance(parsed, list) else [parsed]
            except Exception:
                raw = re.split(r'[,;\s]+', text)
        out = []
        seen = set()
        for chip in raw:
            chip = str(chip or '').strip().strip('"\'')
            key = chip.upper()
            if chip and key not in seen:
                seen.add(key)
                out.append(chip)
        return out



    def _tail(raw):
        text = str(raw or '').strip()
        if not text:
            return ''
        parts = [p.strip() for p in text.replace('/', '\\').split('\\') if p.strip()]
        return parts[-1] if parts else text

    def _product_norm(value):
        """Normalize product/build text so Sinka_QC_1_0 and Sinka.QC.1.0 compare."""
        text = re.sub(r'[^A-Z0-9]+', '.', str(value or '').upper()).strip('.')
        text = re.sub(r'\.+', '.', text)
        return text

    def _family_prefix(value):
        """Return product-family prefix before numeric version.

        Example: Sinka.QC.1.0, Sinka.QC.1.0.r1, Sinka_QC_2_0 -> SINKA.QC.
        This lets a selected sp/display name with 1.0 match active builds on
        1.0.r1, 2.0, etc. under the same product family.
        """
        norm = _product_norm(value)
        match = re.search(r'(?:^|\.)(\d+)\.(\d+)(?:\.|$)', norm)
        if not match:
            return ''
        prefix = norm[:match.start()].strip('.')
        if not prefix:
            return ''
        return prefix + '.'

    def _terms_from_aliases(aliases):
        seen, alias_out = set(), []
        for alias in aliases or []:
            alias = str(alias or '').strip()
            key = alias.upper()
            if key and key not in seen:
                seen.add(key)
                alias_out.append(alias)

        stop = {'PDT', 'QIPL', 'CRM', 'ENG', 'LIVE', 'STATUS', 'TARGET', 'GENERIC', 'INT', 'STD', 'PERF'}
        tokens = []
        families = []
        for alias in alias_out:
            fam = _family_prefix(alias)
            if fam and fam not in families:
                families.append(fam)
            for tok in re.split(r'[^A-Z0-9]+', alias.upper()):
                if len(tok) >= 3 and tok not in stop and tok not in tokens:
                    tokens.append(tok)
        return alias_out, tokens, families

    def _target_match_terms(target_name):
        target_name = str(target_name or '').strip()
        if not target_name:
            return [], [], []
        try:
            update_global_targets_config()
        except Exception:
            pass
        info = get_target_info(target_name) or {}
        aliases = [target_name]
        for key in ('target_name', 'display_name', 'db_name', 'db_prefix', 'sp_name', 'program'):
            val = str(info.get(key) or '').strip()
            if val:
                aliases.append(val)
        for alias in (info.get('aliases') or []):
            alias = str(alias or '').strip()
            if alias:
                aliases.append(alias)
        return _terms_from_aliases(aliases)


    def _target_pl_terms(cursor, target_name):
        """Read distinct Product Line / PL values from the target's Jira tables."""
        target_name = str(target_name or '').strip()
        if not target_name:
            return []

        def _table_exists(fq_table):
            raw = str(fq_table or '').replace('`', '')
            try:
                schema, table = raw.split('.', 1)
            except ValueError:
                return True
            cursor.execute(
                'SELECT 1 FROM information_schema.tables '
                'WHERE table_schema=%s AND table_name=%s LIMIT 1',
                (schema, table),
            )
            return cursor.fetchone() is not None

        def _table_columns(fq_table):
            try:
                cursor.execute(f'SHOW COLUMNS FROM {fq_table}')
                return [str(r.get('Field') or '').strip() for r in (cursor.fetchall() or []) if r.get('Field')]
            except Exception:
                return []

        def _norm_col(value):
            return re.sub(r'[^a-z0-9]+', '', str(value or '').lower())

        pl_values = []
        seen = set()
        for suffix in ('jiras', 'openjiras'):
            try:
                table = fq_table_for_target(target_name, suffix)
                if not _table_exists(table):
                    continue
                columns = _table_columns(table)
                by_norm = {_norm_col(c): c for c in columns}
                pl_col = next((by_norm.get(_norm_col(c)) for c in (
                    'PL-ID', 'pl-id', 'PL_ID', 'pl_id', 'PL ID', 'PL', 'Product Line', 'Product_Line', 'product_line',
                    'productline', 'Program Line', 'program_line', 'chipset',
                ) if by_norm.get(_norm_col(c))), '')
                if not pl_col:
                    continue
                cursor.execute(
                    f'SELECT DISTINCT `{pl_col}` AS pl FROM {table} '
                    f'WHERE `{pl_col}` IS NOT NULL AND TRIM(`{pl_col}`) <> %s '
                    f'ORDER BY `{pl_col}` LIMIT 100',
                    ('',),
                )
                for row in cursor.fetchall() or []:
                    val = str(row.get('pl') or '').strip()
                    key = val.upper()
                    if val and key not in seen:
                        seen.add(key)
                        pl_values.append(val)
            except Exception as exc:
                logger.warning('[BUILD REPORT RUNNING BUILDS] PL lookup failed for %s %s: %s', target_name, suffix, exc)
        return pl_values

    def _norm_col(value):
        return re.sub(r'[^a-z0-9]+', '', str(value or '').lower())

    def _table_exists(cursor, fq_table):
        raw = str(fq_table or '').replace('`', '')
        try:
            schema, table = raw.split('.', 1)
        except ValueError:
            return True
        cursor.execute(
            'SELECT 1 FROM information_schema.tables '
            'WHERE table_schema=%s AND table_name=%s LIMIT 1',
            (schema, table),
        )
        return cursor.fetchone() is not None

    def _table_columns(cursor, fq_table):
        try:
            cursor.execute(f'SHOW COLUMNS FROM {fq_table}')
            return [str(r.get('Field') or '').strip() for r in (cursor.fetchall() or []) if r.get('Field')]
        except Exception:
            return []

    def _first_col(columns, candidates):
        by_norm = {_norm_col(c): c for c in (columns or [])}
        for cand in candidates:
            hit = by_norm.get(_norm_col(cand))
            if hit:
                return hit
        return ''

    def _dashboard_status_target(cursor, selected_target, selected_bu=''):
        """Resolve selected dropdown target/display to dashboard_status row."""
        selected_target = str(selected_target or '').strip()
        selected_bu = str(selected_bu or '').strip().upper()
        if not selected_target:
            return None
        params = [selected_target, selected_target, selected_target]
        bu_sql = ''
        if selected_bu:
            bu_sql = ' AND UPPER(bu) = %s'
            params.append(selected_bu)
        cursor.execute(f"""
            SELECT *
            FROM pdt_stats_dashboard.dashboard_status
            WHERE is_active = 1
              AND (
                    LOWER(target_name) = LOWER(%s)
                 OR LOWER(target_display) = LOWER(%s)
                 OR LOWER(db_name) = LOWER(%s)
              )
              {bu_sql}
            ORDER BY id DESC
            LIMIT 1
        """, tuple(params))
        return cursor.fetchone() or None

    def _fq_from_dashboard(row, suffix):
        if not row:
            return ''
        schema = get_schema_for_bu(str(row.get('bu') or '').strip().upper())
        prefix = str(row.get('db_name') or row.get('target_name') or '').strip('`.').lower()
        if not schema or not prefix:
            return ''
        return f'`{schema}`.`{prefix}_{suffix}`'

    def _target_pl_terms_from_dashboard(cursor, dashboard_row):
        """Read PL/Product Line values using dashboard_status -> actual target tables.
        Filters by product/target name column so shared tables (e.g. COMPUTE)
        only return PLs for the selected target, not all products in the table.
        """
        pl_values = []
        seen = set()

        # Candidate column names that identify the product/target in a shared table
        _PRODUCT_COLS = (
            'product', 'product_name', 'project', 'target', 'target_name',
            'sp_name', 'program', 'cpl', 'program_line', 'Program Line',
        )

        # Build a set of name tokens from the dashboard row to match against
        dr = dashboard_row or {}
        name_candidates = [
            str(dr.get('target_name') or '').strip(),
            str(dr.get('target_display') or '').strip(),
            str(dr.get('sp_name') or '').strip(),
            str(dr.get('db_name') or '').strip(),
            str(dr.get('program') or '').strip(),
            str(dr.get('cpl') or '').strip(),
        ]
        name_candidates = [n for n in name_candidates if n]

        for suffix in ('jiras', 'openjiras'):
            table = _fq_from_dashboard(dashboard_row, suffix)
            if not table:
                continue
            try:
                if not _table_exists(cursor, table):
                    continue
                columns = _table_columns(cursor, table)
                pl_col = _first_col(columns, (
                    'PL-ID', 'pl-id', 'PL_ID', 'pl_id', 'PL ID', 'PL', 'Product Line', 'Product_Line',
                    'product_line', 'productline', 'Program Line', 'program_line',
                    'chipset', 'software_product', 'software product',
                ))
                if not pl_col:
                    continue

                # Try to find a product/target name column to scope the query
                prod_col = _first_col(columns, _PRODUCT_COLS)

                if prod_col and name_candidates:
                    # Shared table: filter by product name so Kanapali PLs
                    # don't appear when Glymur is selected
                    like_parts = ' OR '.join(
                        [f'LOWER(`{prod_col}`) LIKE %s'] * len(name_candidates)
                    )
                    params = [f'%{n.lower()}%' for n in name_candidates] + ['']
                    cursor.execute(
                        f'SELECT DISTINCT `{pl_col}` AS pl FROM {table} '
                        f'WHERE ({like_parts}) '
                        f'  AND `{pl_col}` IS NOT NULL AND TRIM(`{pl_col}`) <> %s '
                        f'ORDER BY `{pl_col}` LIMIT 200',
                        tuple(params),
                    )
                    rows_found = cursor.fetchall() or []
                    # If the product col filter returns nothing, the column
                    # probably doesn't hold target names - fall back to unfiltered
                    if not rows_found:
                        cursor.execute(
                            f'SELECT DISTINCT `{pl_col}` AS pl FROM {table} '
                            f'WHERE `{pl_col}` IS NOT NULL AND TRIM(`{pl_col}`) <> %s '
                            f'ORDER BY `{pl_col}` LIMIT 200',
                            ('',),
                        )
                        rows_found = cursor.fetchall() or []
                else:
                    # No product col found - read all PLs (single-product table)
                    cursor.execute(
                        f'SELECT DISTINCT `{pl_col}` AS pl FROM {table} '
                        f'WHERE `{pl_col}` IS NOT NULL AND TRIM(`{pl_col}`) <> %s '
                        f'ORDER BY `{pl_col}` LIMIT 200',
                        ('',),
                    )
                    rows_found = cursor.fetchall() or []

                for row in rows_found:
                    val = str(row.get('pl') or '').strip()
                    key = val.upper()
                    if val and key not in seen:
                        seen.add(key)
                        pl_values.append(val)
            except Exception as exc:
                logger.warning('[BUILD REPORT RUNNING BUILDS] dashboard PL lookup failed for %s: %s', table, exc)
        return pl_values

    def _cr_lookup_keys(value):
        raw = re.sub(r'[^A-Z0-9]+', '', str(value or '').upper())
        if not raw:
            return set()
        keys = {raw}
        digits = ''.join(re.findall(r'\d+', raw))
        if 5 <= len(digits) <= 10:
            keys.add(digits)
            keys.add(f'CR{digits}')
        if raw.startswith('CR') and raw[2:]:
            keys.add(raw[2:])
        return {k for k in keys if k}

    def _build_cr_details_by_build(cursor, dashboard_row, build_names, pl_terms=None):
        """For each running build, find mapped CRs in jiras/openjiras and enrich from unique_crs."""
        if not dashboard_row or not build_names:
            return {}
        build_names = [str(b or '').strip() for b in build_names if str(b or '').strip()]
        if not build_names:
            return {}

        build_to_crs = {b: [] for b in build_names}
        all_cr_ids = []
        for suffix in ('jiras', 'openjiras'):
            table = _fq_from_dashboard(dashboard_row, suffix)
            if not table:
                continue
            try:
                if not _table_exists(cursor, table):
                    continue
                columns = _table_columns(cursor, table)
                mb_col = _first_col(columns, ('metabuild', 'MetaBuild', 'meta_build', 'build', 'build_id', 'builds'))
                cr_col = _first_col(columns, ('mapped_cr', 'Mapped CR', 'mapped cr', 'cr', 'cr_number', 'CR', 'Change Request'))
                ticket_col = _first_col(columns, ('stability_ticket', 'jira_key', 'JIRA', 'key'))
                pl_col = _first_col(columns, ('PL-ID', 'pl-id', 'PL_ID', 'pl_id', 'PL ID', 'PL', 'Product Line', 'Product_Line', 'product_line', 'productline', 'software_product'))
                if not mb_col or not cr_col:
                    continue
                select_parts = [f'`{mb_col}` AS metabuild', f'`{cr_col}` AS cr']
                select_parts.append(f'`{ticket_col}` AS jira_key' if ticket_col else 'NULL AS jira_key')
                select_parts.append(f'`{pl_col}` AS pl' if pl_col else 'NULL AS pl')
                pl_norms = {_product_norm(pl) for pl in (pl_terms or []) if str(pl or '').strip()}
                for build in build_names:
                    cursor.execute(
                        f"SELECT {', '.join(select_parts)} FROM {table} "
                        f"WHERE `{mb_col}` LIKE %s AND `{cr_col}` IS NOT NULL AND TRIM(`{cr_col}`) <> %s "
                        f"ORDER BY `{mb_col}` DESC LIMIT 100",
                        (f'%{build}%', ''),
                    )
                    for jr in cursor.fetchall() or []:
                        if pl_norms and pl_col:
                            row_pl_norm = _product_norm(jr.get('pl'))
                            if row_pl_norm and row_pl_norm not in pl_norms:
                                continue
                        cr = str(jr.get('cr') or '').strip()
                        if not cr:
                            continue
                        build_to_crs.setdefault(build, []).append({'cr': cr, 'jira_key': _ser(jr.get('jira_key')), 'pl': _ser(jr.get('pl'))})
                        all_cr_ids.append(cr)
            except Exception as exc:
                logger.warning('[BUILD REPORT RUNNING BUILDS] CR lookup failed for %s: %s', table, exc)

        if not all_cr_ids:
            return {}

        unique_table = _fq_from_dashboard(dashboard_row, 'unique_crs')
        unique_by_key = {}
        try:
            if unique_table and _table_exists(cursor, unique_table):
                columns = _table_columns(cursor, unique_table)
                key_cols = [c for c in (
                    _first_col(columns, ('mapped_cr', 'Mapped CR', 'mapped cr')),
                    _first_col(columns, ('cr', 'CR')),
                    _first_col(columns, ('cr_number', 'CR Number', 'stability_ticket')),
                ) if c]
                if key_cols:
                    status_col = _first_col(columns, ('cr_status', 'CR Status', 'status', 'final_status'))
                    image_col = _first_col(columns, ('si_image', 'SI Image', 'simage', 'image', 'build_image', 'cr_si_image'))
                    age_col = _first_col(columns, ('cr_age', 'CR Age', 'overall_age', 'age'))
                    title_col = _first_col(columns, ('cr_title', 'CR Title', 'title', 'jira_title', 'summary'))
                    wanted = sorted(set().union(*[_cr_lookup_keys(x) for x in all_cr_ids]))
                    placeholders = ','.join(['%s'] * len(wanted))

                    def _sql_norm(col):
                        expr = f"UPPER(TRIM(`{col}`))"
                        for old in (' ', '-', '_', ',', '.0', '.'):
                            expr = f"REPLACE({expr}, '{old}', '')"
                        return expr

                    where = ' OR '.join([f'{_sql_norm(c)} IN ({placeholders})' for c in key_cols])
                    params = tuple(x for _ in key_cols for x in wanted)
                    key_expr = 'COALESCE(' + ', '.join([f"NULLIF(TRIM(`{c}`), '')" for c in key_cols]) + ') AS cr'
                    select_parts = [key_expr]
                    select_parts.append(f'`{status_col}` AS cr_status' if status_col else 'NULL AS cr_status')
                    select_parts.append(f'`{image_col}` AS si_image' if image_col else 'NULL AS si_image')
                    select_parts.append(f'`{age_col}` AS cr_age' if age_col else 'NULL AS cr_age')
                    select_parts.append(f'`{title_col}` AS cr_title' if title_col else 'NULL AS cr_title')
                    cursor.execute(f"SELECT {', '.join(select_parts)} FROM {unique_table} WHERE {where} LIMIT 1000", params)
                    for row in cursor.fetchall() or []:
                        detail = {k: _ser(v) for k, v in row.items()}
                        for key in _cr_lookup_keys(detail.get('cr')):
                            unique_by_key.setdefault(key, detail)
        except Exception as exc:
            logger.warning('[BUILD REPORT RUNNING BUILDS] unique_cr enrichment failed for %s: %s', unique_table, exc)

        enriched = {}
        for build, refs in build_to_crs.items():
            rows = []
            seen_crs = set()
            for ref in refs:
                cr = ref.get('cr') or ''
                lookup = None
                for key in _cr_lookup_keys(cr):
                    lookup = unique_by_key.get(key)
                    if lookup:
                        break
                item = dict(ref)
                if lookup:
                    item.update({k: lookup.get(k) for k in ('cr_status', 'si_image', 'cr_age', 'cr_title')})
                dedup_key = re.sub(r'[^A-Z0-9]+', '', cr.upper())
                if dedup_key and dedup_key not in seen_crs:
                    seen_crs.add(dedup_key)
                    rows.append(item)
            enriched[build] = rows
        return enriched


    q = str(request.args.get('q') or '').strip().lower()


    bu_raw = str(request.args.get('bu') or request.args.get('bu_key') or '').strip().upper()
    target_raw = str(request.args.get('target') or '').strip()
    target_config_bu = str(get_bu_for_target(target_raw) or '').strip().upper() if target_raw else ''
    limit_raw = request.args.get('limit') or '250'

    try:
        limit = max(1, min(int(limit_raw), 500))

    except Exception:
        limit = 250
    target_aliases, target_tokens, target_families = _target_match_terms(target_raw)
    target_pl_terms = []

    conn = None


    cur = None
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if not conn:
            return jsonify({'ok': False, 'error': 'DB connection failed', 'builds': []}), 500      
        cur = conn.cursor(dictionary=True)

        dashboard_target = _dashboard_status_target(cur, target_raw, bu_raw)
        if dashboard_target:
            target_config_bu = str(dashboard_target.get('bu') or target_config_bu or '').strip().upper()
            dashboard_aliases = [
                target_raw,
                dashboard_target.get('target_name'), dashboard_target.get('target_display'),
                dashboard_target.get('db_name'), dashboard_target.get('sp_name'),
                dashboard_target.get('program'), dashboard_target.get('cpl'),
            ]
            # Prefer the active dashboard_status row over the in-memory metadata
            # cache. The cache can contain inactive/stale rows, which previously
            # allowed Maili to inherit Hawi aliases and therefore Hawi builds.
            target_aliases, target_tokens, target_families = _terms_from_aliases(dashboard_aliases)

        target_pl_terms = _target_pl_terms_from_dashboard(cur, dashboard_target) if dashboard_target else _target_pl_terms(cur, target_raw)
        for pl in target_pl_terms:

            if pl not in target_aliases:
                target_aliases.append(pl)
            fam = _family_prefix(pl)
            if fam and fam not in target_families:
                target_families.append(fam)
            for tok in re.split(r'[^A-Z0-9]+', str(pl or '').upper()):
                if len(tok) >= 3 and tok not in target_tokens:
                    target_tokens.append(tok)

        cur.execute("""

            SELECT MAX(updated_at) AS axiom_last_updated_at,
                   MAX(fetched_at) AS axiom_last_fetched_at,
                   MAX(submitted_at) AS latest_submitted_at,
                   MAX(started_at) AS latest_started_at,
                   COUNT(*) AS active_total
                        FROM `pdt_stats_dashboard`.`axiom_job_summary`
            WHERE state = 'Running'
        """)

        meta = cur.fetchone() or {}

                # Build SQL WHERE from PL terms so we never pull 10 000 rows into Python.
        # Underscore is a LIKE wildcard - neutralise via REPLACE for ADAS/FLEX PLs.
        def _pl_sql_where(pl_values):
            parts, params = [], []
            seen = set()
            for pl in (pl_values or []):
                base = re.sub(r'\.[rc]\d+$', '', str(pl or '').strip(), flags=re.IGNORECASE).strip()
                if not base or base.upper() in seen:
                    continue
                seen.add(base.upper())
                if '_' in base:
                    safe = base.replace('_', '|')
                    parts.append("(software_product = %s OR REPLACE(software_product,'_','|') LIKE %s)")
                    params.extend([base, f'%{safe}%'])
                else:
                    parts.append("(software_product = %s OR software_product LIKE %s)")
                    params.extend([base, f'%{base}%'])
            return (' OR '.join(parts), params) if parts else (None, [])

        pl_where, pl_params = _pl_sql_where(target_pl_terms)
        if pl_where:
            # PL-scoped query - only rows matching this target's software_product
            cur.execute(f"""
                SELECT job_id, build_id, build_name, software_product,
                       taxonomy_path, team, state, device_count, chip_ids,
                       submitted_at, started_at, ended_at,
                       product_flavor, submitter, site, updated_at, fetched_at
                FROM `pdt_stats_dashboard`.`axiom_job_summary`
                WHERE state = 'Running'
                  AND ({pl_where})
                ORDER BY COALESCE(started_at, submitted_at) DESC
                LIMIT %s
            """, tuple(pl_params) + (limit,))
        else:
            # No PL terms - broad fetch, Python alias filter below still applies
            candidate_limit = limit if not target_raw else 10000
            cur.execute("""
                SELECT job_id, build_id, build_name, software_product,
                       taxonomy_path, team, state, device_count, chip_ids,
                       submitted_at, started_at, ended_at,
                       product_flavor, submitter, site, updated_at, fetched_at
                FROM `pdt_stats_dashboard`.`axiom_job_summary`
                WHERE state = 'Running'
                ORDER BY COALESCE(started_at, submitted_at) DESC
                LIMIT %s
            """, (candidate_limit,))

        rows = cur.fetchall() or []
        out = []
        by_build = {}
        matched_before_limit = 0

        for r in rows:
            build_full = str(r.get('build_name') or r.get('build_id') or '').strip()
            build = _tail(build_full)
            if not build:
                continue
            hay = ' '.join(str(r.get(k) or '') for k in (
                'build_id', 'build_name', 'software_product', 'taxonomy_path',
                'team', 'state', 'product_flavor', 'submitter', 'site'
                        ))
            hay_lower = hay.lower()

            if q and q not in hay_lower and q not in build.lower():
                continue

            # When PL terms exist, SQL already filtered - no Python re-filter needed.
            # When no PL terms, apply broad alias/token/family match in Python.
            if target_raw and not pl_where:
                hay_upper = (hay + ' ' + build).upper()
                hay_norm = _product_norm(hay + ' ' + build)
                alias_hit = any(str(alias or '').upper() in hay_upper for alias in target_aliases if str(alias or '').strip())
                family_hit = any(fam and fam in hay_norm for fam in target_families)
                token_hit = any(tok in hay_upper for tok in target_tokens)
                if not alias_hit and not family_hit and not token_hit:
                    continue




            matched_before_limit += 1

            key = build.upper()
            chip_ids = _chip_ids(r.get('chip_ids'))
            device_count = len(chip_ids) if chip_ids else int(r.get('device_count') or 0)
            existing = by_build.get(key)
            if existing:
                existing['job_count'] = int(existing.get('job_count') or 1) + 1
                if _ser(r.get('job_id')):
                    existing.setdefault('job_ids', []).append(_ser(r.get('job_id')))
                existing_chips = existing.setdefault('chip_ids', [])
                existing_chip_keys = {str(c or '').upper() for c in existing_chips}
                for chip in chip_ids:
                    if chip.upper() not in existing_chip_keys:
                        existing_chips.append(chip)
                        existing_chip_keys.add(chip.upper())
                if existing_chips:
                    existing['device_count'] = len(existing_chips)
                else:
                    existing['device_count'] = int(existing.get('device_count') or 0) + device_count
                continue

            if len(out) >= limit:
                continue
            item = {
                'job_id': _ser(r.get('job_id')),
                'job_ids': [_ser(r.get('job_id'))] if _ser(r.get('job_id')) else [],
                                'job_count': 1,
                'chip_ids': chip_ids,
                'build': build,

                'build_full': build_full,
                'software_product': _ser(r.get('software_product')),
                'taxonomy_path': _ser(r.get('taxonomy_path')),
                'team': _ser(r.get('team')),
                'state': _ser(r.get('state')),
                'device_count': device_count,
                'submitted_at': _ser(r.get('submitted_at')),
                'started_at': _ser(r.get('started_at')),
                'product_flavor': _ser(r.get('product_flavor')),
                'submitter': _ser(r.get('submitter')),
                'site': _ser(r.get('site')),
                'updated_at': _ser(r.get('updated_at')),
                'fetched_at': _ser(r.get('fetched_at')),
            }
            by_build[key] = item
            out.append(item)

        cr_by_build = _build_cr_details_by_build(cur, dashboard_target, [row.get('build') for row in out], target_pl_terms)
        for row in out:
            cr_rows = cr_by_build.get(row.get('build')) or []
            row['crs'] = cr_rows
            row['cr_count'] = len(cr_rows)
            row['cr_statuses'] = ', '.join(sorted({str(c.get('cr_status') or '').strip() for c in cr_rows if str(c.get('cr_status') or '').strip()}))
            row['si_images'] = ', '.join(sorted({str(c.get('si_image') or '').strip() for c in cr_rows if str(c.get('si_image') or '').strip()}))
            row['cr_ages'] = ', '.join(sorted({str(c.get('cr_age') or '').strip() for c in cr_rows if str(c.get('cr_age') or '').strip()}))
        return jsonify({

            'ok': True,
            'builds': out,
            'count': len(out),
            'matched_before_limit': matched_before_limit,
            'candidate_count': len(rows),
                        'limit': limit,
            'bu': bu_raw,
                        'target': target_raw,
            'target_bu': target_config_bu,
            'dashboard_target': {
                'bu': (dashboard_target or {}).get('bu') or '',
                'target_name': (dashboard_target or {}).get('target_name') or '',
                'db_name': (dashboard_target or {}).get('db_name') or '',
                'target_display': (dashboard_target or {}).get('target_display') or '',
                'sp_name': (dashboard_target or {}).get('sp_name') or '',
            },
            'target_tables': {
                'jiras': _fq_from_dashboard(dashboard_target, 'jiras') if dashboard_target else '',
                'openjiras': _fq_from_dashboard(dashboard_target, 'openjiras') if dashboard_target else '',
                'unique_crs': _fq_from_dashboard(dashboard_target, 'unique_crs') if dashboard_target else '',
            },
            'target_aliases': target_aliases,


            'target_tokens': target_tokens,

            'target_families': target_families,
            'target_pl_terms': target_pl_terms,
            'source': 'pdt_stats_dashboard.axiom_job_summary',

            'axiom_last_updated_at': _ser(meta.get('axiom_last_updated_at')),

            'axiom_last_fetched_at': _ser(meta.get('axiom_last_fetched_at')),
            'latest_submitted_at': _ser(meta.get('latest_submitted_at')),
            'latest_started_at': _ser(meta.get('latest_started_at')),
            'active_total': int(meta.get('active_total') or 0),
        })
    except Exception as exc:
        logger.exception('[BUILD REPORT RUNNING BUILDS] %s', exc)
        return jsonify({'ok': False, 'error': str(exc), 'builds': []}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()






@live_status_publish_bp.route('/live_status_view')
@login_required
def landing():
    """Live Status landing.

    TARGET_GROUP users get editor controls (Add Job / Manage Jobs). Other
    authenticated users see the same BU -> target published-report navigation,
    without editor controls.
    """

    requested_target = (request.args.get('target') or request.args.get('target_name') or '').strip()
    if requested_target:

        requested_tab = _requested_live_status_tab('mtbf')
        if _target_group_access():
            return redirect(_canonical_target_edit_url(requested_target, tab=requested_tab))
        if not _can_view_live_status_target(requested_target):
            return redirect(url_for('live_status_publish_bp.landing'))
        return redirect(_canonical_target_edit_url(requested_target, tab=requested_tab))

    can_edit = _target_group_access()

    # All published jobs - both editor and viewer see the same BU---target navigation.
    # Editor additionally sees Add Job / Manage Jobs controls (can_edit flag).
    all_jobs = list_jobs()
    all_target_opts = _all_targets_for_ui()
    target_to_bu = {row['target']: {'bu_key': row['bu_key'], 'bu_name': row['bu_name']} for row in all_target_opts}

    hidden_bus = {'WEEKLY_QIPL_REPORTS', 'HWPDT'}

        # Annotate every job with BU info (needed for admin table and bu_list)

    for job in all_jobs:
        first_target = (job.get('targets') or [''])[0]
        info = target_to_bu.get(first_target, {})
        job['_bu_key'] = (info.get('bu_key') or 'OTHER').upper()
        job['_bu_name'] = info.get('bu_name') or job['_bu_key']
        job['_job_type'] = _job_type(job)

    visible_jobs = _filter_live_status_jobs_for_current_user(all_jobs, can_edit=can_edit)

    # BU list: editors see published + draft, scoped viewers see only their published jobs
    seen_bus = {}
    for job in visible_jobs:
        if job.get('status') not in ('published', 'draft') if can_edit else job.get('status') != 'published':
            continue
        bk = job['_bu_key']
        if bk not in hidden_bus and bk not in seen_bus:
            seen_bus[bk] = job['_bu_name']
    bu_list = sorted(seen_bus.items(), key=lambda x: x[0])

    # Build per-BU target data for JS (published + draft for editors)
    visible_statuses = ('published', 'draft') if can_edit else ('published',)
    bu_targets_js = {}
    for job in visible_jobs:
        if job.get('status') not in visible_statuses:
            continue
        bk = job['_bu_key']
        if bk in hidden_bus:
            continue
        first_target = (job.get('targets') or [''])[0]
        if not first_target:
            continue
        entry = {
            'name': first_target,
            'bu_key': bk,
            'bu_name': job['_bu_name'],
            'target_url': url_for('live_status_publish_bp.live_status_target_by_bu', bu_key=bk, target_name=first_target),
            'token': job.get('public_token') or '',
            'published_at': job.get('published_at') or '',
            'updated_at': job.get('updated_at') or '',
            'published_by': job.get('published_by') or '',
            'job_type': _job_type(job),
            'meta_count': len(job.get('published_rows') or job.get('draft_rows') or []),
            'job_name': job.get('name') or '',
            'status': job.get('status') or 'draft',
                        'job_id': job.get('id') or '',

        }
        bu_targets_js.setdefault(bk, []).append(entry)

    viewer_bu_sections = []
    for bu_key, bu_name in bu_list:
        targets = [t for t in (bu_targets_js.get(bu_key) or []) if t.get('status') == 'published']
        if targets:
            viewer_bu_sections.append({
                'bu_key': bu_key,
                'bu_name': bu_name,
                                'targets': targets,

            })

    viewer_scope = _current_live_status_viewer_scope() if not can_edit else {'matched_groups': []}
    matched_groups_upper = {str(g or '').strip().upper() for g in (viewer_scope.get('matched_groups') or [])}

        # - Special "Live View Stats" BU sections (Automotive 4.5 + WBC) -

    # These are always shown as top-level cards that open their own dedicated
    # stats page directly - they are NOT backed by a Live Status publish job.
    # Editors (full target-group access) always see BOTH. Restricted viewers
    # only see the card whose BU their access scope explicitly covers - each
    # card has its OWN visibility flag so a user scoped only to WBC does not
    # also see the Automotive 4.5 card (and vice versa).
    _viewer_bus_scope = viewer_scope.get('bus') or set()
    _show_auto_gen45_section = (
        can_edit or bool(viewer_scope.get('all'))
        or 'PDTBUDDY.IVIGEN4.5' in matched_groups_upper
        or 'AUTO' in _viewer_bus_scope or 'AUTOMOTIVE' in _viewer_bus_scope
        or 'AUTO_GEN45' in _viewer_bus_scope or 'AUTOMOTIVE4.5' in _viewer_bus_scope
    )
    _show_wbc_section = (
        can_edit or bool(viewer_scope.get('all'))
        or 'WBC' in _viewer_bus_scope
    )

    # Strip regular WBC/synthetic sections only so WBC does not duplicate with the
    # special card below. Keep AUTO/AUTOMOTIVE because those are already-published
    # Automotive Gen5 reports such as HQX/HGY.
    viewer_bu_sections = [
        section for section in viewer_bu_sections
        if str(section.get('bu_key') or '').upper() not in {'WBC', 'AUTO_GEN45', 'AUTOMOTIVE4.5'}
    ]

    if _show_auto_gen45_section:
        # Automotive Gen 4.5 card
        viewer_bu_sections.insert(0, {
            'bu_key': 'AUTO_GEN45',
            'bu_name': 'Automotive Gen 4.5',
            'special_page': True,
            'targets': [{
                'name': 'Automotive Gen 4.5',
                'bu_key': 'AUTO_GEN45',
                'bu_name': 'Automotive Gen 4.5',
                'target_url': url_for('automotive_live_view_stats_bp.automotive_live_view_stats_page', target_name='4.8.9.0'),
                'token': '',
                'published_at': '',
                'updated_at': '',
                'published_by': '',
                'job_type': 'CRM',
                'meta_count': 0,
                'job_name': 'Automotive Gen 4.5 Live View Stats',
                'status': 'published',
                'job_id': '',
                'special_page': True,
                'special_label': 'Live View Stats',
                'special_icon': 'fa-car',
                'special_desc': 'Gen 4.5 - ADAS / FLEX / IVI',
            }],
        })

    if _show_wbc_section:
        # WBC card
        viewer_bu_sections.insert(1 if _show_auto_gen45_section else 0, {
            'bu_key': 'WBC',
            'bu_name': 'WBC',
            'special_page': True,
            'targets': [{
                'name': 'WBC',
                'bu_key': 'WBC',
                'bu_name': 'WBC',
                'target_url': url_for('wbc_live_view_stats_bp.wbc_live_view_status_page'),
                'token': '',
                'published_at': '',
                'updated_at': '',
                'published_by': '',
                'job_type': 'CRM',
                'meta_count': 0,
                'job_name': 'WBC Live View Stats',
                'status': 'published',
                'job_id': '',
                'special_page': True,
                'special_label': 'Live View Stats',
                'special_icon': 'fa-network-wired',
                'special_desc': 'Wireless Broadband Connectivity',
            }],
        })

    # BUs whose targets should NOT appear in the Add Job modal. Keep AUTO/AUTOMOTIVE
    # visible so already-published Automotive Gen5 targets can still be managed.
    _hidden_add_job_bus = {'WBC', 'AUTO_GEN45', 'AUTOMOTIVE4.5'}

    total_viewer_targets = sum(len(section.get('targets') or []) for section in viewer_bu_sections)
    # Never auto-redirect when special sections are present (multiple cards always shown)
    _has_special = any(section.get('special_page') for section in viewer_bu_sections)
    if not can_edit and total_viewer_targets == 1 and not _has_special:
        only_target = next((section['targets'][0] for section in viewer_bu_sections if section.get('targets')), None)
        if only_target:
            return redirect(only_target['target_url'])

    requested_bu = (request.args.get('bu_key') or '').strip().upper()
    visible_bu_keys = {str(row[0]).upper() for row in bu_list}
    auto_open_bu = requested_bu if requested_bu in visible_bu_keys else ''
    # NOTE: 'AUTO_GEN45' here is legacy - it used to auto-expand the old
    # accordion-style BU section for viewers whose only LDAP group is
    # PdtBuddy.IVIGen4.5. Automotive 4.5 is now rendered as a "special_page"
    # card that always navigates directly on click and must never be
    # force-expanded (see the `not is_special` guard in
    # live_status_publish_landing.html - without it, this assignment caused
    # the special card's hidden body to render a second time, showing what
    # looked like two duplicate "Automotive 4.5" cards stacked on the page).
    if not auto_open_bu and not can_edit and 'PDTBUDDY.IVIGEN4.5' in matched_groups_upper:
        auto_open_bu = 'AUTO_GEN45'
    elif not auto_open_bu and not can_edit and len(viewer_scope.get('matched_groups') or []) == 1 and len(bu_list) == 1:
        auto_open_bu = bu_list[0][0]



        
        
    # Filter target_options passed to Add Job modal - exclude AUTO/WBC BUs

    add_job_target_opts = [
        row for row in all_target_opts
        if str(row.get('bu_key') or '').upper() not in _hidden_add_job_bus
    ]

    response = make_response(render_template(
        'live_status_publish_landing.html',
        jobs=visible_jobs,
        bu_list=bu_list,
        bu_targets_js=bu_targets_js,
        target_options=add_job_target_opts,
        preselected_target=requested_target,
        preselected_bu=(request.args.get('bu_key') or '').strip(),
        can_edit=can_edit,
        access_groups=_live_status_access_groups_catalog(all_target_opts),
        auto_open_bu=auto_open_bu,
        viewer_groups=[],
        viewer_bu_sections=viewer_bu_sections,

    ))

    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@live_status_publish_bp.route('/live_status_view/new', methods=['POST'])
@login_required
def create_job_view():
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    name = (request.form.get('name') or '').strip()
    targets = request.form.getlist('targets')
    if not targets:
        one_target = (request.form.get('target') or '').strip()
        if one_target:
            targets = [one_target]
    job_type = (request.form.get('job_type') or 'CRM').strip().upper()
    primary_target = targets[0] if targets else ''
    if job_type == 'ENG' and primary_target and _count_active_eng_jobs(primary_target) >= 10:
        return jsonify({'ok': False, 'error': 'Maximum 10 active ENG jobs allowed per target.'}), 400
    if job_type != 'ENG':
        job_type = 'CRM'
        existing = _find_existing_single_target_job(primary_target, 'CRM') if primary_target else None
        if existing:
            edit_url = _canonical_target_editor_url(primary_target)
            return jsonify({'ok': True, 'edit_url': edit_url, 'job_id': existing['id']})
    job = create_job(name=name, targets=targets, username=getattr(current_user, 'id', 'unknown'), job_type=job_type)
    edit_url = _canonical_target_editor_url(primary_target) if primary_target else url_for('live_status_publish_bp.edit_job', job_id=job['id'])
    return jsonify({'ok': True, 'edit_url': edit_url, 'job_id': job['id']})


@live_status_publish_bp.route('/live_status_view/target/<target_name>')
@login_required
def open_target_workspace(target_name):
    """Backward-compatible target shortcut to the canonical BU/target page."""
    if _target_group_access():
        return redirect(_canonical_target_editor_url(target_name))
    return redirect(_canonical_target_edit_url(target_name))


@live_status_publish_bp.route('/live_status_view/<job_id>/edit')
@login_required
def edit_job(job_id):
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    job = get_job(job_id)
    if not job:
        return render_template('coming_soon_template.html', title='Live Status Publish', message='Job not found.'), 404
    primary_target = (job.get('targets') or [''])[0]
    if primary_target:
        return redirect(_canonical_target_editor_url(primary_target))
    return _render_current_report_editor(job)


@live_status_publish_bp.route('/live_status/<bu_key>/')
@live_status_publish_bp.route('/live_status_view/<bu_key>/')
@login_required
def live_status_bu_incomplete_url(bu_key):
    """Handle incomplete BU-only Live Status URLs gracefully.

    Example: /live_status_view/IOT/ should show the Live Status landing with
    IOT targets instead of Flask's default 404 page.
    """
    bu_key = str(bu_key or '').strip().upper()
    business_units = get_business_units() or {}
    known_bus = {str(k).upper() for k in business_units.keys()}
    if bu_key in known_bus:
        return redirect(url_for('live_status_publish_bp.landing', bu_key=bu_key))
    return redirect(url_for('live_status_publish_bp.landing'))


@live_status_publish_bp.route('/live_status/<bu_key>/<target_name>/<initial_tab_path>')
@live_status_publish_bp.route('/live_status_view/<bu_key>/<target_name>/<initial_tab_path>')
@live_status_publish_bp.route('/live_status/<bu_key>/<target_name>')
@live_status_publish_bp.route('/live_status_view/<bu_key>/<target_name>')
@login_required
def live_status_target_by_bu(bu_key, target_name, initial_tab_path=None):

    """Canonical per-target Live Status URL.

    Editors use this URL as the single Current Report edit/save/publish workspace.
        Viewers use the same URL for the published read-only report.
    """

    initial_tab = _normal_live_status_tab(initial_tab_path, _requested_live_status_tab('mtbf'))

    # ── BU-based routing ──────────────────────────────────────────────────────
    # Determine the effective BU for this target (prefer resolved BU over URL param).
    _bu = str(get_bu_for_target(target_name) or bu_key or '').strip().upper()
    _target_upper = str(target_name or '').upper()
    _is_auto_bu = _bu in {'AUTO', 'AUTOMOTIVE', 'AUTO_TELEMATICS'} or \
                  _target_upper.startswith('NORD') or \
                  _target_upper.startswith('SECA') or \
                  'NORD_' in _target_upper or \
                  'SECA_' in _target_upper

    # Non-AUTO / non-WBC targets → redirect ALL users (editors and viewers)
    # to the correct live-view-stats page for that BU.
    if not _is_auto_bu:
        # Auto Gen4.5 (e.g. 4.8.x targets)
        try:
            from automotive_live_view_stats_routes import _is_auto_gen45_target
            if _is_auto_gen45_target(target_name):
                return redirect(url_for(
                    'automotive_live_view_stats_bp.automotive_live_view_stats_page',
                    target_name=target_name,
                ))
        except Exception:
            pass
        # WBC
        if _bu == 'WBC':
            return redirect(url_for('wbc_live_view_stats_bp.wbc_live_view_status_page'))
        # Others (XR, Mobile, IoT, MBB, Compute, …)
        return redirect(url_for(
            'others_live_view_stats_bp.others_live_view_stats_page',
            target_name=target_name,
        ))
    # ─────────────────────────────────────────────────────────────────────────

    if current_user.is_authenticated and _target_group_access():

        # Editors always get the edit workspace - draft or published.
        job = _find_existing_single_target_job(target_name, 'CRM') or _find_published_job_for_target(target_name)
        if job:
            return _render_current_report_editor(job, initial_tab=initial_tab)
        # No job exists yet - send editor back to landing to create one.
        return redirect(url_for('live_status_publish_bp.landing'))


    # Viewers: check access first.
    if not _can_view_live_status_target(target_name, bu_key):
        return render_template(
            'coming_soon_template.html',
            title='Live Status',
            message='You do not have access to this target. Request the listed Live Status viewer group from the landing page.'
        ), 403

    # Viewers only see a published job - never a draft.
    if not _find_published_job_for_target(target_name):
        return render_template(
            'coming_soon_template.html',
            title='Live Status',
            message='No published report is available for this target yet.'
        ), 404
    return _render_target_status_page(target_name, initial_tab=initial_tab)


@live_status_publish_bp.route('/live_status_view/<job_id>')

@login_required
def view_job(job_id):
    """Legacy one-segment route.

    Prefer target/BU navigation. Only treat the segment as a legacy job id if a
    matching job exists; otherwise avoid exposing any job-id wording in the UI.
    """
    segment = str(job_id or '').strip()
    bu_key = segment.upper()
    business_units = get_business_units() or {}
    if bu_key in {str(k).upper() for k in business_units.keys()}:
        return redirect(url_for('live_status_publish_bp.landing', bu_key=bu_key))

    # If the segment is actually a target name, send users to the canonical
    # /live_status_view/<BU>/<target> URL.
    target_opt = _find_target_option(segment)
    if target_opt:
        return redirect(url_for(
            'live_status_publish_bp.live_status_target_by_bu',
            bu_key=str(target_opt.get('bu_key') or get_bu_for_target(segment) or 'TARGET').upper(),
            target_name=target_opt.get('target') or segment,
        ))

    job = get_job(segment)
    if not job:
        return redirect(url_for('live_status_publish_bp.landing'))
    target_name = (job.get('targets') or [''])[0]

    if target_name and _target_group_access():
        return redirect(_canonical_target_editor_url(target_name))
    if target_name and not _can_view_live_status_target(target_name):
        return render_template(
            'coming_soon_template.html',
            title='Live Status',
            message='You do not have access to this target.'
        ), 403
    if job.get('status') == 'published' and job.get('public_token'):
        if target_name:
            return redirect(_canonical_target_edit_url(target_name))
        return redirect(url_for('live_status_publish_bp.published_view', public_token=job['public_token']))
    return render_template('coming_soon_template.html', title='Live Status', message='This live status view is not yet published.'), 404




def _available_sjql_domains(target_name: str, is_auto_bu: bool) -> list:
    """Return the actual domain list for this target by reading the same
    mtbf_domains.json that the MTBF Trend page uses. Falls back to a
    sensible default only when the file does not exist yet.
    Only available for AUTO BU — non-AUTO BUs do not use domain-wise JQL tabs.
    """
    if not is_auto_bu:
        return []
    try:
        from live_status_view_api import _get_target_domains
        domains = _get_target_domains(target_name)
        if domains:
            return domains
    except Exception:
        pass
    # Fallback: ADAS/FLEX/IVI until domains file exists
    return ['ADAS', 'FLEX', 'IVI']

def _get_sp_siblings(primary_target: str) -> list:
    """Return SP switcher entries for the current NORD/AUTO target family.

    The URL must point to the actual Live Status job target. If we link to a
    dashboard_status target that has no job, editors are redirected to the
    landing page; therefore a sibling is clickable only when an active CRM job
    exists for one of that SP's candidate targets.
    """
    try:
        import re as _re
        from dashboard_common import get_mysql_connection_db

        primary_target = str(primary_target or '').strip()
        if not primary_target:
            return []
        bu = (get_bu_for_target(primary_target) or 'AUTO').upper()
        prefix = _re.sub(r'_([a-z]+)_[0-9_]+$', '', primary_target.lower())

        conn = get_mysql_connection_db(bu_key=bu)
        if not conn:
            return []
        cur = conn.cursor(dictionary=True)
        cur.execute(
            "SELECT target_name, cpl FROM pdt_stats_dashboard.dashboard_status "
            "WHERE target_name LIKE %s AND cpl IS NOT NULL AND is_active=1 "
            "ORDER BY cpl ASC, target_name ASC",
            (prefix + '%',)
        )
        rows = cur.fetchall() or []

        cur.execute(
            "SELECT cpl FROM pdt_stats_dashboard.dashboard_status "
            "WHERE target_name=%s LIMIT 1",
            (primary_target,)
        )
        own_row = cur.fetchone() or {}
        conn.close()
        own_cpl = str(own_row.get('cpl') or '').strip()

        targets_by_cpl = {}
        preferred_by_cpl = {}
        for r in rows:
            cpl = str(r.get('cpl') or '').strip()
            tgt = str(r.get('target_name') or '').strip()
            if not cpl or not tgt:
                continue
            targets_by_cpl.setdefault(cpl, [])
            if tgt not in targets_by_cpl[cpl]:
                targets_by_cpl[cpl].append(tgt)
            preferred_by_cpl.setdefault(cpl, tgt)
            if tgt.lower() == prefix + '_' + cpl.replace('.', '_'):
                preferred_by_cpl[cpl] = tgt

        if own_cpl:
            targets_by_cpl.setdefault(own_cpl, [])
            if primary_target not in targets_by_cpl[own_cpl]:
                targets_by_cpl[own_cpl].insert(0, primary_target)
            preferred_by_cpl.setdefault(own_cpl, primary_target)

        # HQX fix: DB has domain targets (nord_hqx_adas_5_7_7_0 etc) but
        # nord_hqx itself has no cpl row -> own_cpl is empty.
        # Derive own_cpl from DB rows and add primary_target to candidates.
        if not own_cpl and targets_by_cpl:
            own_cpl = sorted(targets_by_cpl.keys())[0]
        if own_cpl and primary_target not in targets_by_cpl.get(own_cpl, []):
            targets_by_cpl.setdefault(own_cpl, []).insert(0, primary_target)
            preferred_by_cpl.setdefault(own_cpl, primary_target)

        # Filesystem fallback for HQX (no dashboard_status CPL rows)
        # Scan mtbf_*_<key>.json to discover SPs e.g. mtbf_adas_5770.json -> 5.7.7.0
        fs_cpls = {}  # populated below if DB had no CPL rows
        if not targets_by_cpl:
            try:
                import os as _os
                from live_status_view_api import _adas_mtbf_folder as _mtbf_folder
                folder = _mtbf_folder(primary_target)
                for fname in _os.listdir(folder):
                    m = _re.match(r'^mtbf_[a-z\-]+_(\d{4,8})\.json$', fname)
                    if m:
                        sp_k = m.group(1)
                        d = sp_k.ljust(4, '0')[:4]
                        cpl = d[0]+'.'+d[1]+'.'+d[2]+'.'+d[3]
                        fs_cpls[cpl] = primary_target
                for cpl in sorted(fs_cpls):
                    targets_by_cpl.setdefault(cpl, [primary_target])
                    preferred_by_cpl.setdefault(cpl, primary_target)
                if fs_cpls and not own_cpl:
                    own_cpl = sorted(fs_cpls.keys())[0]
            except Exception:
                pass

        # Show SP bar when only 1 unique CPL exists (HQX: only 5.7.7.0)
        # Require 2+ only when multiple CPLs from DB (HGY: 5.1.7.0 + 5.1.9.0)
        _min_sps = 1 if (fs_cpls or len(targets_by_cpl) == 1) else 2
        if len(targets_by_cpl) < _min_sps:
            return []

        # job_targets: any CRM job (active OR revoked/published) so we can
        # build a URL for every SP that ever had a job.
        # active_job_targets: only active jobs (used to decide is_active).
        job_targets = {}
        active_job_targets = {}
        for job in list_jobs():
            if _job_type(job) != 'CRM':
                continue
            for t in (job.get('targets') or []):
                t = str(t or '').strip()
                if not t:
                    continue
                job_targets.setdefault(t.lower(), t)
                if _is_active_job(job):
                    active_job_targets.setdefault(t.lower(), t)
        # Always include primary_target itself as a known job target
        job_targets.setdefault(primary_target.lower(), primary_target)
        active_job_targets.setdefault(primary_target.lower(), primary_target)

        out = []
        for idx, cpl in enumerate(targets_by_cpl.keys()):
            candidates = targets_by_cpl.get(cpl) or []
            candidate_lowers = {t.lower() for t in candidates}

            # Use any job (incl. revoked) for URL; active-only for is_active
            job_target = next((job_targets.get(t.lower()) for t in candidates if job_targets.get(t.lower())), '')
            if not job_target:
                slug = cpl.replace('.', '_')
                job_target = next((real for low, real in job_targets.items() if low.startswith(prefix) and slug in low), '')

            active_target = next((active_job_targets.get(t.lower()) for t in candidates if active_job_targets.get(t.lower())), '')
            if not active_target:
                slug = cpl.replace('.', '_')
                active_target = next((real for low, real in active_job_targets.items() if low.startswith(prefix) and slug in low), '')

            is_active = (
                own_cpl == cpl
                or primary_target.lower() in candidate_lowers
                or bool(active_target and primary_target.lower() == active_target.lower())
            )
            if is_active and not job_target:
                job_target = primary_target

            route_bu = (get_bu_for_target(job_target) or bu or 'AUTO').upper() if job_target else bu
            url = '' if is_active else ('/live_status_view/{}/{}'.format(route_bu, job_target) if job_target else '')

            out.append({
                'cpl': cpl,
                'target': job_target or preferred_by_cpl.get(cpl, ''),
                'url': url,
                'active': is_active,
                'has_job': bool(job_target),
                'color_idx': idx % 6,
            })
        return out
    except Exception:
        logger.exception('[LIVE STATUS SP] failed to build SP siblings for %s', primary_target)
        return []

def _render_published_full_page(job, initial_tab='current', suppress_top_redirect=False):
    """
        Render the canonical Live Status page.
    Works for both published and draft jobs.
    """

    primary_target = (job.get('targets') or [''])[0]
    # Default to 'core' for AUTO BU, 'current' for ENG, 'mtbf' otherwise
    _default_tab = 'core' if _is_core_deck_target(primary_target) else 'mtbf'
    initial_tab = _normal_live_status_tab(initial_tab, _default_tab)

    embedded_core_deck = str(request.args.get('embed') or '').lower() in ('1', 'true', 'yes') and initial_tab == 'core'

    can_edit = current_user.is_authenticated and _target_group_access()

    # Editors work from draft_rows. Viewers see only the last explicitly
    # published snapshot; never fall back to draft rows on a published page.
    published_rows = (job.get('draft_rows') if can_edit else job.get('published_rows')) or []
    running_rows = _published_display_rows(
        [r for r in published_rows if str(r.get('run_status','')).lower() == 'running'],
        job.get('published_at')
    )
    all_rows = _published_display_rows(published_rows, job.get('published_at'))
    is_compute_mtbf = (get_bu_for_target(primary_target) or '').upper() == 'COMPUTE'
    is_auto_bu = _is_core_deck_target(primary_target)

            # Core Slides tab: only shown for AUTO BU targets or targets with

    # core_deck_enabled=True in their per-target config. Default: disabled.
    try:
        _meta = load_metadata_config() or {}
        _targets_cfg = _meta.get('TARGETS_CONFIG', {}) or {}
        _target_cfg  = _targets_cfg.get(primary_target) or next(
            (v for k, v in _targets_cfg.items() if str(k).lower() == str(primary_target).lower()),
            {}
        ) or {}
        _core_enabled = is_auto_bu or bool(_target_cfg.get('core_deck_enabled', False))
    except Exception:
        _core_enabled = is_auto_bu
    visible_tabs = ['core'] if _core_enabled else []
    if _job_type(job) == 'ENG':
        initial_tab = 'current'
    else:
        visible_tabs += ['weekly', 'opencrs', 'openjiras', 'buildreport']
    visible_tabs += ['current', 'mtbf']

    return render_template(
        'live_status_publish_edit.html',
        job=job,
        workspace_data=None,
        primary_target=primary_target,
        running_rows=running_rows,
        all_rows=all_rows,
        target_options=_all_targets_for_ui() if can_edit else [],
        jira_pdt_filter_id=JIRA_PDT_FILTER_ID,
        is_compute_mtbf=is_compute_mtbf,
        is_auto_bu=is_auto_bu,
        is_eng_job=_job_type(job) == 'ENG',
        can_edit=can_edit,
        initial_tab=initial_tab,
        mtbf_only=False,  # tab visibility controlled client-side via Customize Tabs
        embedded_core_deck=embedded_core_deck,
        suppress_top_redirect=suppress_top_redirect,
        # Domains available for this target
        available_domains=_available_sjql_domains(primary_target, is_auto_bu),
        visible_tabs=visible_tabs,
        sp_siblings=_get_sp_siblings(primary_target) if is_auto_bu else [],
        sp_configs=job.get('sp_configs') or {},
    )


@live_status_publish_bp.route('/published/live-status/<public_token>')
def published_view(public_token):
    job = next((j for j in list_jobs() if str(j.get('public_token')) == str(public_token)), None)
    if not job:
        return render_template('coming_soon_template.html', title='Published Live Status', message='Published view not found.'), 404
    if job.get('status') == 'revoked':
        return render_template('coming_soon_template.html', title='Published Live Status', message='This report has been revoked and is no longer active.'), 410
    if job.get('status') != 'published':
        return render_template('coming_soon_template.html', title='Published Live Status', message='This report is not yet published.'), 404
    target_name = (job.get('targets') or [''])[0]
    if target_name:
        return redirect(_canonical_target_edit_url(target_name))
    return _render_published_full_page(job, request.args.get('tab') or 'current')


@live_status_publish_bp.route('/published/live-status/<public_token>/current')
def published_view_current_tab(public_token):
    job = next((j for j in list_jobs() if str(j.get('public_token')) == str(public_token)), None)
    if not job or job.get('status') != 'published':
        return render_template('coming_soon_template.html', title='Published Live Status', message='Published view not found.'), 404
    primary_target = (job.get('targets') or [''])[0]
    if primary_target:
        return redirect(_canonical_target_edit_url(primary_target))
    return _render_published_full_page(job, 'current')


@live_status_publish_bp.route('/published/live-status/<public_token>/mtbf')
def published_view_mtbf_tab(public_token):
    job = next((j for j in list_jobs() if str(j.get('public_token')) == str(public_token)), None)
    if not job or job.get('status') != 'published':
        return render_template('coming_soon_template.html', title='Published Live Status', message='Published view not found.'), 404
    target_name = (job.get('targets') or [''])[0]
    if target_name:
        return redirect(_canonical_target_edit_url(target_name))
    return _render_published_full_page(job, 'current' if _job_type(job) == 'ENG' else 'mtbf')


@live_status_publish_bp.route('/published/live-status/<public_token>/weekly')
def published_view_weekly_tab(public_token):
    job = next((j for j in list_jobs() if str(j.get('public_token')) == str(public_token)), None)
    if not job or job.get('status') != 'published':
        return render_template('coming_soon_template.html', title='Published Live Status', message='Published view not found.'), 404
    target_name = (job.get('targets') or [''])[0]
    if target_name:
        return redirect(_canonical_target_edit_url(target_name))
    return _render_published_full_page(job, 'current' if _job_type(job) == 'ENG' else 'weekly')



@live_status_publish_bp.route('/api/live_status/targets/<target_name>/mtbf_dashboard', methods=['GET'])
def api_published_mtbf_dashboard(job_id=None, target_name=None):
    """Public API: return DB-backed MTBF trend/table for a target."""
    if target_name is not None:
        job, err = _get_target_report_job_for_api(target_name)
    else:
        job, err = _get_published_job_for_api(job_id)
    if err:
        return jsonify(err[0]), err[1]
    target = (job.get('targets') or [''])[0]

    def _num(value, default=0.0):

        try:
            if value in (None, ''):
                return default
            text = str(value).strip()
            if text.upper() in ('NA', 'N/A', '-', '---', 'NONE'):
                return default
            return float(text.replace(',', ''))
        except Exception:
            return default

    def _saved_mtbf_rows_from_job(saved_rows):
        """Prefer rows saved from the edit page (Builds tab) for published MTBF."""
        rows = [r for r in (saved_rows or []) if isinstance(r, dict)]
        builds_rows = [r for r in rows if r.get('builds_tab') or str(r.get('run_status', '')).lower() == 'builds']
        source_rows = builds_rows or rows
        if not source_rows:
            return [], []
        series = []
        details = []
        seen = set()
        for r in source_rows:
            meta = str(r.get('meta_id') or r.get('display_build') or '').strip()
            build = r.get('build_full') or r.get('display_build') or meta
            if r.get('isMerged') and r.get('merged_builds'):
                try:
                    build = '\n'.join([str(b) for b in (r.get('merged_builds') or []) if str(b).strip()]) or build
                except Exception:
                    pass
            key = (meta.upper(), str(build).strip().upper(), bool(r.get('builds_tab') or str(r.get('run_status', '')).lower() == 'builds'))
            if key in seen:
                continue
            seen.add(key)
            hours = _num(r.get('display_hours', r.get('hours')), 0.0)
            crashes = int(_num(r.get('crashes'), 0.0))
            mtbf_raw = r.get('display_mtbf', r.get('mtbf'))
            mtbf = _num(mtbf_raw, 0.0)
            if not mtbf and hours and crashes:
                mtbf = round(hours / crashes, 2)
            label = meta or str(build or '').split('\n')[0]
            if not label:
                continue
            series.append({
                'meta_id': label,
                'week': r.get('week') or r.get('first_submitted') or '',
                'total_hours': round(hours, 2),
                'crashes': crashes,
                'mtbf': round(mtbf, 2) if mtbf else 0,
                'source': r.get('source') or ('BUILDS' if r.get('builds_tab') else 'LIVE_STATUS'),
            })
            details.append({
                'meta_id': meta or label,
                'build_id': build or label,
                'hours': round(hours, 2),
                'crashes': crashes,
                'mtbf': round(mtbf, 2) if mtbf else '',
                'product_mtbf': r.get('product_mtbf') or '',
                'qc_mtbf': r.get('qc_mtbf') or '',
                'source': r.get('source') or ('BUILDS' if r.get('builds_tab') else 'LIVE_STATUS'),
                'mode': r.get('mode') or 'CRM',
                'mtbf_details': r.get('test_eng_comment') or r.get('comments') or '',
                'week': r.get('week') or r.get('first_submitted') or '',
            })
            return series, details

    saved_rows = (job.get('draft_rows') if current_user.is_authenticated and _target_group_access() else job.get('published_rows')) or []
    saved_series, saved_details = _saved_mtbf_rows_from_job(saved_rows)

    if saved_series or saved_details:
        return jsonify({
            'success': True,
            'target': target,
            'schema': 'live_status_job_json',
            'source': 'edit_page_saved_rows',
            'mtbf_series': saved_series,
            'mtbf_build_table': saved_details,
        })

    # ── Try JSON-backed MTBF data (same path as internal dashboard) ──
    try:
        from dashboard_routes import _load_mtbf_json_payload
        _json_payload = _load_mtbf_json_payload(target, 'MTBF')
        _json_rows = _json_payload.get('rows') or []
        if _json_rows:
            _series = []
            _details = []
            for _r in _json_rows:
                _meta_id = str(_r.get('meta_id') or _r.get('build') or '').strip()
                if not _meta_id:
                    continue
                _hours   = _num(_r.get('hours') or _r.get('total_hours'), 0.0)
                _crashes = int(_num(_r.get('total_crashes') or _r.get('crashes'), 0.0))
                _mtbf    = _num(_r.get('mtbf'), 0.0)
                if not _mtbf and _hours and _crashes:
                    _mtbf = round(_hours / _crashes, 2)
                _series.append({
                    'meta_id':     _meta_id,
                    'week':        _r.get('date') or _r.get('week') or '',
                    'total_hours': round(_hours, 2),
                    'crashes':     _crashes,
                    'mtbf':        round(_mtbf, 2) if _mtbf else 0,
                    'source':      'JSON',
                })
                _details.append({
                    'meta_id':     _meta_id,
                    'build_id':    str(_r.get('build_full') or _meta_id),
                    'hours':       round(_hours, 2),
                    'crashes':     _crashes,
                    'mtbf':        round(_mtbf, 2) if _mtbf else '',
                    'source':      'JSON',
                    'mode':        'CRM',
                    'mtbf_details': _r.get('comments') or '',
                    'week':        _r.get('date') or _r.get('week') or '',
                })
            if _series:
                return jsonify({
                    'success': True,
                    'target':  target,
                    'schema':  'json_backed',
                    'source':  'mtbf_json',
                    'mtbf_series':      _series,
                    'mtbf_build_table': _details,
                })
    except Exception:
        pass  # fall through to DB-backed path

    conn = None
    cur = None
    try:

        import json as _json
        from datetime import date as _date, datetime as _datetime
        from dashboard_common import get_schema_for_target, get_mysql_connection_db
        from dashboard_service import get_build_report_for_target, build_mtbf_dashboard_payload

        def _ser(value):
            if isinstance(value, (_datetime, _date)):
                return value.isoformat()
            return value

        def _round2(value):
            try:
                if value in (None, ''):
                    return value
                return round(float(value), 2)
            except Exception:
                return value

        def _parse_build_date(value):
            text = str(value or '')
            if not text:
                return None
            m = re.search(r'(20\d{2})[-_/]?(0[1-9]|1[0-2])[-_/]?([0-3]\d)', text)
            if m:
                try:
                    return _date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                except Exception:
                    return None
            for m in re.finditer(r'(?:^|[^0-9])((0[1-9]|1[0-2])([0-3]\d))(?:[^0-9]|$)', text):
                try:
                    return _date(_date.today().year, int(m.group(2)), int(m.group(3)))
                except Exception:
                    continue
            return None

        def _build_date(build):
            for key in ('submitted', 'first_submitted', 'week', 'date', 'created_at', 'updated_at', 'build_id', 'build'):
                parsed = _parse_build_date((build or {}).get(key))
                if parsed:
                    return parsed
            return None

        def _filter_rows(rows):
            cutoff = _date(_date.today().year, 5, 1)
            out = []
            for row in rows or []:
                meta_dt = _parse_build_date(row.get('first_jira_date') or row.get('jira_date'))
                if meta_dt and meta_dt < cutoff:
                    continue
                kept = []
                for build in row.get('builds') or []:
                    if not isinstance(build, dict):
                        continue
                    parsed = _build_date(build)
                    if meta_dt or parsed is None or parsed >= cutoff:
                        for key in ('mtbf', 'product_mtbf', 'qc_mtbf', 'hours'):
                            if key in build:
                                build[key] = _round2(build[key])
                        kept.append(build)
                if kept:
                    nr = dict(row)
                    nr['builds'] = kept
                    for key in ('mtbf', 'product_mtbf', 'qc_mtbf', 'total_hours', 'hours'):
                        if key in nr:
                            nr[key] = _round2(nr[key])
                    out.append(nr)
            return out

        schema_name = get_schema_for_target(target) or 'pdt_stats_mobile'
        conn = get_mysql_connection_db()
        if not conn:
            return jsonify({'success': False, 'message': 'Database connection failed'}), 500
        cur = conn.cursor(dictionary=True)
        report_payload = get_build_report_for_target(
            cur,
            target,
            schema_name=schema_name,
            pdt_type='SWPDT',
            toggle_mode='CRM',
            use_static_cache=False,
        ) or {}
        report_rows = report_payload.get('rows', []) if isinstance(report_payload, dict) else (report_payload or [])
        report_rows = _filter_rows([row for row in report_rows if isinstance(row, dict)])
        dashboard = build_mtbf_dashboard_payload(report_rows, pdt_type='SWPDT') or {}

        detail_rows = []
        for row in report_rows or []:
            meta_id = row.get('meta_id') or ''
            for build in row.get('builds') or []:
                if not isinstance(build, dict):
                    continue
                build_id = str(build.get('build_id') or '').strip()
                if not build_id or build_id == '__META__':
                    continue
                hours = float(build.get('hours') or 0)
                crashes = int(float(build.get('swpdt_crashes') or build.get('crashes') or 0))
                mtbf = build.get('mtbf')
                if mtbf in (None, ''):
                    mtbf = round(hours / crashes, 2) if hours and crashes else (round(hours, 2) if hours else '')
                mtbf = _round2(mtbf)
                detail_rows.append({
                    'meta_id': meta_id,
                    'build_id': build_id,
                    'hours': _round2(hours),
                    'crashes': crashes,
                    'mtbf': mtbf,
                    'product_mtbf': _round2(build.get('product_mtbf')),
                    'qc_mtbf': _round2(build.get('qc_mtbf')),
                    'source': build.get('build_source') or build.get('source') or 'MANUAL',
                    'mode': build.get('mode') or 'CRM',
                    'mtbf_details': build.get('mtbf_details') or build.get('comments') or build.get('notes') or '',
                    'week': build.get('week') or build.get('date') or build.get('submitted') or build.get('first_submitted') or row.get('first_jira_date') or '',
                })

        payload = {
            'success': True,
            'target': target,
            'schema': schema_name,
            'mtbf_series': dashboard.get('mtbf_series') or [],
            'mtbf_build_table': detail_rows or dashboard.get('mtbf_build_table') or [],
        }
        return jsonify(_json.loads(_json.dumps(payload, default=_ser)))
    except Exception as exc:
        logger.exception('[PUBLISHED MTBF DASHBOARD] %s', exc)
        return jsonify({'success': False, 'message': str(exc), 'mtbf_series': [], 'mtbf_build_table': []}), 500
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()




@live_status_publish_bp.route('/api/live_status/targets/<target_name>/auto_mtbf', methods=['GET'])
def api_target_auto_mtbf(target_name):
    """Public API: return JSON-backed AUTO MTBF rows for ADAS/FLEX/IVI by target."""
    target = str(target_name or '').strip()
    view = str(request.args.get('view') or 'ADAS').strip().upper()
    if view not in {'ADAS', 'FLEX', 'IVI'}:
        view = 'ADAS'
    requested_sp = str(request.args.get('sp') or '').strip()
    if not requested_sp:
        siblings = _get_sp_siblings(target) or []
        default_sp = next((row for row in siblings if row.get('active')), None)
        default_sp = default_sp or (siblings[0] if siblings else {})
        requested_sp = str(default_sp.get('cpl') or '').strip()
    try:
        from live_status_view_api import _load_adas_mtbf, _adas_rows_to_chart_data
        data = _load_adas_mtbf(target, view, requested_sp) or {}

        rows = data.get('rows') if isinstance(data.get('rows'), list) else []
        crash_types_raw = str(request.args.get('crash_types') or 'system,ssr,process').strip()
        crash_types = [c.strip().lower() for c in crash_types_raw.split(',') if c.strip()]
        if not crash_types:
            crash_types = ['system', 'ssr', 'process']
        return jsonify({
            'ok': True,
            'target': target,
                        'view': view,
            'sp': requested_sp,
            'views': ['ADAS', 'FLEX', 'IVI'],

            'rows': rows,
            'chart_data': _adas_rows_to_chart_data(rows, crash_types),
            'updated_at': data.get('updated_at') or '',
        })
    except Exception as exc:
        logger.exception('[TARGET AUTO MTBF] %s', exc)
        return jsonify({'ok': False, 'error': str(exc), 'rows': []}), 500


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/excel_rows', methods=['GET'])
def api_published_excel_rows(job_id=None, target_name=None):
    """Public API: return MTBF Excel rows for a target."""
    if target_name is not None:
        job, err = _get_target_report_job_for_api(target_name)
    else:
        job, err = _get_published_job_for_api(job_id)
    if err:
        return jsonify(err[0]), err[1]
    target = (job.get('targets') or [''])[0]
    sheet = (request.args.get('sheet') or '').strip()
    try:
        data = _read_mtbf_excel_rows(target, sheet_name_override=sheet or None)
        return jsonify(data)
    except Exception as exc:
        logger.exception('[EXCEL_ROWS] %s', exc)
        return jsonify({'success': False, 'message': str(exc), 'headers': [], 'rows': []}), 500


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/weekly_full', methods=['GET'])
def api_published_weekly_full(job_id=None, target_name=None):
    """Public API: full weekly raw data for a target.
    Returns:
      - cr_rows        : from unique_crs table (mapped_cr, cr_area, cr_subsystem,
                         cr_functionality, cr_status, cr_category, cr_age,
                         cr_occurrence, cr_title, jira_date, last_instance)
      - jira_rows      : from jiras + openjiras tables (stability_ticket, jira_date,
                         jira_title, metabuild)
      - open_jira_rows : from openjiras table only (same columns)
      - build_ids      : distinct metabuild values seen in jira_rows for the range
      - pie_status     : [{name, y}] aggregated from cr_rows
      - pie_area       : [{name, y}] aggregated from cr_rows
            - counts         : summary numbers
    """
    if target_name is not None:
        job, err = _get_target_report_job_for_api(target_name)
    else:
        job, err = _get_published_job_for_api(job_id)
    if err:
        return jsonify(err[0]), err[1]
    target = (job.get('targets') or [''])[0]
    from_arg = (request.args.get('from') or '').strip()
    to_arg   = (request.args.get('to')   or '').strip()
    force = str(request.args.get('force') or '').lower() in ('1', 'true', 'yes')
    cache_requested = str(request.args.get('cache') or '').lower() in ('1', 'true', 'yes')
    builds_only = str(request.args.get('builds_only') or '').lower() in ('1', 'true', 'yes')
    can_edit = current_user.is_authenticated and _target_group_access()
    weekly_selection = dict(job.get('weekly_report_selection') or {})
    requested_builds = [b.strip() for b in (request.args.get('builds') or '').split(',') if b.strip()]
    selected_builds = requested_builds or [str(b or '').strip() for b in (weekly_selection.get('selected_builds') or []) if str(b or '').strip()]
    try:
        import json as _json
        from collections import Counter
        from datetime import date as _date, datetime as _datetime, timedelta as _td
        from dashboard_common import (
            get_schema_for_target, get_mysql_connection_db,
            fetch_weekly_crs, norm_ymd,
        )

        if from_arg and to_arg:
            from_dt = _datetime.strptime(from_arg[:10], '%Y-%m-%d').date()
            to_dt = _datetime.strptime(to_arg[:10], '%Y-%m-%d').date()
        else:
            today = _date.today()
            # Default published weekly report is the last completed Monday-Sunday window.
            offset = 7 if today.weekday() == 6 else today.weekday() + 1
            to_dt = today - _td(days=offset)
            from_dt = to_dt - _td(days=6)

        from_s = norm_ymd(from_dt)
        to_s   = norm_ymd(to_dt)

        cached_payload = weekly_selection.get('cache') if isinstance(weekly_selection.get('cache'), dict) else None
        if selected_builds and cached_payload and not force:
            cached_builds = [str(b or '').strip() for b in (weekly_selection.get('selected_builds') or []) if str(b or '').strip()]
            if (cached_builds == selected_builds
                    and str(cached_payload.get('from_date') or '') == from_s
                    and str(cached_payload.get('to_date') or '') == to_s
                    and int(cached_payload.get('area_source_version') or 0) >= 6):
                return jsonify(cached_payload)

        schema = get_schema_for_target(target)
        if not schema:
            return jsonify({'success': False, 'message': 'Target schema not found'}), 404
        conn = get_mysql_connection_db(bu_key=schema)
        if not conn:
            return jsonify({'success': False, 'message': 'DB connection error'}), 500

        sc  = schema.strip('`')
        tgt = target.strip('`.')
        jiras_tbl  = f'`{sc}`.`{tgt}_jiras`'
        open_tbl   = f'`{sc}`.`{tgt}_openjiras`'
        unique_tbl = f'`{sc}`.`{tgt}_unique_crs`'

        cur = conn.cursor(dictionary=True)

        def _tbl_exists(fq):
            n = fq.replace('`', '')
            try:
                s, t = n.split('.', 1)
            except ValueError:
                return True
            cur.execute('SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1', (s, t))
            return cur.fetchone() is not None

        def _table_cols(fq):
            try:
                cur.execute(f'SHOW COLUMNS FROM {fq}')
                return {r.get('Field') for r in (cur.fetchall() or []) if r.get('Field')}
            except Exception:
                return set()

        def _norm_id(v):
            raw = str(v or '').strip().upper()
            if raw.endswith('.0'):
                raw = raw[:-2]
            return re.sub(r'[\s\-_,.]', '', raw)

        def _cr_digits(v):
            m = re.search(r'(\d{5,9})', str(v or ''))
            return m.group(1) if m else ''

        def _cr_lookup_keys(v):
            raw = _norm_id(v)
            if not raw:
                return set()
            keys = {raw}
            digits = _cr_digits(v) or _cr_digits(raw)
            if digits:
                keys.add(digits)
                keys.add(f'CR{digits}')
            if raw.startswith('CR') and raw[2:]:
                keys.add(raw[2:])
            return {k for k in keys if k}

        def _fetch_unique_cr_rows_by_ids(cr_ids):
            """Fetch and enrich authoritative CR details from *_unique_crs.

            Build-selected reports can reference CRs outside the current week,
            so this lookup ignores the date window. If the matching row is a
            duplicate (for example cr_occurrence = DUP), refill its CR details
            from the related mapped_cr row in unique_crs.
            """
            keys = set()
            for cr_id in cr_ids or []:
                keys.update(_cr_lookup_keys(cr_id))
            if not keys:
                return []
            ucols = _table_cols(unique_tbl)
            key_cols = [c for c in ('mapped_cr', 'cr', 'cr_number') if c in ucols]
            if not key_cols:
                return []

            def _col(name, alias=None):
                alias = alias or name
                return f'`{name}` AS `{alias}`' if name in ucols else f'NULL AS `{alias}`'

            def _sql_norm_expr(col):
                expr = f"UPPER(TRIM(`{col}`))"
                for old in (' ', '-', '_', ',', '.0', '.'):
                    expr = f"REPLACE({expr}, '{old}', '')"
                return expr

            last_inst_col = 'jira_date__last_instance' if 'jira_date__last_instance' in ucols else ('last_instance' if 'last_instance' in ucols else ('jira_date' if 'jira_date' in ucols else ''))
            current_month_col = next((c for c in ('cr_____current_month', 'current_month_occurrence', 'current_month_occurrence#', 'current_month_count', 'current_occurrence', 'current_month') if c in ucols), '')
            previous_month_col = next((c for c in ('cr_____previous_month', 'previous_month_occurrence', 'previous_month_occurrence#', 'previous_month_count', 'previous_occurrence', 'previous_month') if c in ucols), '')
            total_builds_col = next((c for c in ('cr_reported_build_count', 'total_builds_cr_reported', 'total_no_of_builds_cr_reported', 'CR_Reported_Build_count', 'total_build_count') if c in ucols), '')
            cr_candidates = [c for c in ('mapped_cr', 'cr', 'cr_number') if c in ucols]
            cr_expr = "COALESCE(" + ', '.join([f"NULLIF(TRIM(`{c}`), '')" for c in cr_candidates]) + ") AS `cr`" if cr_candidates else "NULL AS `cr`"
            select_parts = [
                cr_expr,
                _col('mapped_cr'),
                _col('cr_number'),
                _col('cr_occurrence', 'overall_cr_occurrence'),
                _col('cr_age', 'overall_age'),
                _col('cr_title'),
                _col('cr_area'),
                _col('cr_subsystem'),
                _col('cr_functionality'),
                _col('built_date', 'cr_date'),
                _col('cr_status'),
                _col('cr_category'),
                _col('jira_date'),
                f'`{last_inst_col}` AS `last_instance`' if last_inst_col else 'NULL AS `last_instance`',
                f'`{current_month_col}` AS `current_month_occurrence`' if current_month_col else 'NULL AS `current_month_occurrence`',
                f'`{previous_month_col}` AS `previous_month_occurrence`' if previous_month_col else 'NULL AS `previous_month_occurrence`',
                f'`{total_builds_col}` AS `total_builds_cr_reported`' if total_builds_col else 'NULL AS `total_builds_cr_reported`',
            ]
            order_col = last_inst_col or key_cols[0]

            def _query_unique_rows(wanted_keys):
                wanted_keys = sorted({k for k in wanted_keys if k})
                if not wanted_keys:
                    return []
                placeholders = ','.join(['%s'] * len(wanted_keys))
                where_parts = [f"{_sql_norm_expr(col)} IN ({placeholders})" for col in key_cols]
                flat_params = tuple(x for _ in where_parts for x in wanted_keys)
                cur.execute(
                    f"SELECT {', '.join(select_parts)} FROM {unique_tbl} "
                    f"WHERE {' OR '.join(where_parts)} "
                    f"ORDER BY `{order_col}` DESC",
                    flat_params,
                )
                return _ser_rows(cur.fetchall() or [])

            def _row_keys(row):
                row_keys = set()
                for ck in ('cr', 'mapped_cr', 'cr_number'):
                    row_keys.update(_cr_lookup_keys((row or {}).get(ck)))
                return row_keys

            def _is_dup_row(row):
                hay = ' '.join(str((row or {}).get(k) or '') for k in (
                    'overall_cr_occurrence', 'cr_occurrence', 'cr_category', 'cr_status'
                )).strip().upper()
                return bool(re.search(r'\bDUP(?:LICATE)?\b', hay))

            rows = _query_unique_rows(keys)

            # If the selected CR row is a duplicate, unique_crs normally stores
            # the authoritative details on the related mapped_cr. Fetch that row
            # too, then copy its area/subsystem/functionality/status/title/etc.
            related_keys = set()
            for row in rows:
                if not _is_dup_row(row):
                    continue
                self_keys = _cr_lookup_keys(row.get('cr') or row.get('cr_number'))
                mapped_keys = _cr_lookup_keys(row.get('mapped_cr'))
                if mapped_keys and mapped_keys != self_keys:
                    related_keys.update(mapped_keys)
            if related_keys:
                existing_keys = set()
                for row in rows:
                    existing_keys.update(_row_keys(row))
                rows.extend(_query_unique_rows(related_keys - existing_keys))

            by_key = {}
            for row in rows:
                for key in _row_keys(row):
                    by_key.setdefault(key, row)

            fill_fields = (
                'cr_area', 'cr_subsystem', 'cr_functionality', 'cr_status',
                'cr_category', 'cr_title', 'overall_age', 'cr_date', 'jira_date',
                'last_instance', 'current_month_occurrence',
                'previous_month_occurrence', 'total_builds_cr_reported',
            )
            for row in rows:
                if not _is_dup_row(row):
                    continue
                related = None
                for key in _cr_lookup_keys(row.get('mapped_cr')):
                    candidate = by_key.get(key)
                    if candidate is not row:
                        related = candidate
                        break
                if not related:
                    continue
                for field in fill_fields:
                    if str(related.get(field) or '').strip():
                        row[field] = related.get(field)
                row['duplicate_source_cr'] = row.get('cr') or row.get('cr_number') or ''
                row['duplicate_mapped_cr'] = related.get('cr') or row.get('mapped_cr') or ''
                row['duplicate_details_refilled'] = True

            seen = {}
            for row in rows:
                matched_keys = _row_keys(row) & keys
                matched = next(iter(sorted(matched_keys)), '') or _norm_id(row.get('cr') or row.get('mapped_cr') or row.get('cr_number'))
                if matched and matched not in seen:
                    seen[matched] = row
            return list(seen.values())

        def _row_build(row):
            return str((row or {}).get('metabuild') or '').strip()

        def _filter_by_selected(rows):
            if not selected_builds:
                return list(rows or [])
            sel = set(selected_builds)
            return [r for r in (rows or []) if _row_build(r) in sel]

        def _ser(v):
            if isinstance(v, (_datetime, _date)):
                return v.strftime('%Y-%m-%d %H:%M:%S') if hasattr(v, 'hour') else str(v)
            return v

        def _ser_rows(rows):
            return [{k: _ser(v) for k, v in r.items()} for r in (rows or [])]

        def _first_text(row, keys):
            for key in keys:
                val = (row or {}).get(key)
                if val is not None and str(val).strip():
                    return str(val).strip()
            return ''

        def _area_from_open_jira_title(value):
            """Bucket open/unmapped JIRAs from title text only."""
            text = str(value or '').strip().lower()
            if not text:
                return ''
            if any(token in text for token in ('wconnect', 'wcnss', 'cnss', 'wlan', 'wi-fi', 'wifi', 'btfm', 'bluetooth', 'wireless')):
                return 'WConnect'
            if ' bt ' in f' {text} ' or text.startswith('bt ') or text.endswith(' bt'):
                return 'WConnect'
            if any(token in text for token in ('modem', 'mpss', 'ril', 'data call', 'lte', '5g', 'nr', 'ims', 'qmi')):
                return 'Modem'
            if any(token in text for token in ('adsp', 'audio', 'qdsp')):
                return 'ADSP'
            if any(token in text for token in ('cdsp', 'compute dsp')):
                return 'CDSP'
            if any(token in text for token in ('trustzone', 'trust zone', 'qsee')) or text == 'tz' or ' tz ' in f' {text} ':
                return 'TZ'
            if any(token in text for token in ('apps', 'apss', 'android', 'kernel', 'framework', 'userspace')):
                return 'APPS'
            return ''

        # - 1. JIRA rows (jiras + openjiras) for the date range -
        jira_rows = []
        open_jira_rows = []
        build_ids = []

        base_jira_cols = ['stability_ticket', 'jira_date', 'jira_title', 'serial_no', 'metabuild']
        extra_cr_cols = [
            'mapped_cr', 'cr', 'cr_number',
            'cr_area', 'area', 'ChangeRequestParticipant.Area',
            'cr_subsystem', 'subsystem', 'ChangeRequestParticipant.Subsystem',
            'cr_function', 'cr_functionality', 'functionality', 'ChangeRequestParticipant.Functionality',
        ]

        def _build_where_for_selected(mb_col):
            vals = []
            for b in selected_builds or []:
                b = str(b or '').strip()
                if not b:
                    continue
                tail = b.replace('/', '\\').split('\\')[-1]
                vals.append(tail)
                m = re.search(r'-(\d{3,6})(?:\.\d+)?-(?:STD|PERF|SAFE|USER|ENG)', tail, re.IGNORECASE)
                if m:
                    n = str(int(m.group(1)))
                    vals.extend([f'-{n.zfill(3)}-', f'-{n}-'])
            vals = list(dict.fromkeys([v for v in vals if v]))
            if not vals:
                return '', ()
            return ' OR '.join([f'`{mb_col}` LIKE %s' for _ in vals]), tuple(f'%{v}%' for v in vals)

        # jiras table. If builds were explicitly requested by Core Deck, fetch by selected build/meta
        # across history instead of only the last weekly date window; otherwise old builds show 0 crashes.
        try:
            jcols = _table_cols(jiras_tbl)
            select_cols = [c for c in base_jira_cols + extra_cr_cols if c in jcols] or base_jira_cols
            jira_cols = ', '.join(f'`{c}`' for c in select_cols)
            mb_col = 'metabuild' if 'metabuild' in jcols else ('MetaBuild' if 'MetaBuild' in jcols else '')
            where_like, like_params = _build_where_for_selected(mb_col) if (requested_builds and mb_col) else ('', ())
            if where_like:
                cur.execute(f'SELECT {jira_cols} FROM {jiras_tbl} WHERE ({where_like}) ORDER BY jira_date DESC LIMIT 5000', like_params)
            else:
                cur.execute(
                    f'SELECT {jira_cols} FROM {jiras_tbl} '
                    f'WHERE jira_date BETWEEN %s AND %s ORDER BY jira_date DESC',
                    (from_s, to_s)
                )
            jira_rows = _ser_rows(cur.fetchall() or [])
        except Exception as e:
            logger.warning('[WEEKLY_FULL] jiras table error: %s', e)

        # openjiras table
        if _tbl_exists(open_tbl):
            try:
                ocols = _table_cols(open_tbl)
                select_cols = [c for c in base_jira_cols + extra_cr_cols if c in ocols] or base_jira_cols
                jira_cols = ', '.join(f'`{c}`' for c in select_cols)
                mb_col = 'metabuild' if 'metabuild' in ocols else ('MetaBuild' if 'MetaBuild' in ocols else '')
                where_like, like_params = _build_where_for_selected(mb_col) if (requested_builds and mb_col) else ('', ())
                if where_like:
                    cur.execute(f'SELECT {jira_cols} FROM {open_tbl} WHERE ({where_like}) ORDER BY jira_date DESC LIMIT 5000', like_params)
                else:
                    cur.execute(
                        f'SELECT {jira_cols} FROM {open_tbl} '
                        f'WHERE jira_date BETWEEN %s AND %s ORDER BY jira_date DESC',
                        (from_s, to_s)
                    )
                open_jira_rows = _ser_rows(cur.fetchall() or [])
            except Exception as e:
                logger.warning('[WEEKLY_FULL] openjiras table error: %s', e)

        # distinct build IDs seen in jira rows before applying the saved selection
        all_jira_rows_all = jira_rows + open_jira_rows
        seen_builds = {}
        for r in all_jira_rows_all:
            mb = str(r.get('metabuild') or '').strip()
            if mb and mb not in seen_builds:
                seen_builds[mb] = True
        available_build_ids = list(seen_builds.keys())
        if not requested_builds:
            selected_builds = [b for b in selected_builds if b in seen_builds]

        if builds_only:
            cur.close()
            return jsonify({
                'success': True,
                'from_date': from_s,
                'to_date': to_s,
                'available_build_ids': available_build_ids,
                'selected_build_ids': selected_builds or available_build_ids,
                'selection_enabled': bool(selected_builds),
                'selection_updated_at': weekly_selection.get('updated_at'),
            })

        jira_rows = _filter_by_selected(jira_rows)
        open_jira_rows = _filter_by_selected(open_jira_rows)
        all_jira_rows = jira_rows + open_jira_rows
        build_ids = selected_builds or available_build_ids

        selected_cr_ids = set()
        if selected_builds:
            for jr in all_jira_rows:
                for ck in ('mapped_cr', 'cr', 'cr_number'):
                    cv = _norm_id(jr.get(ck))
                    if cv:
                        selected_cr_ids.add(cv)

        # - 2. CR rows from unique_crs -
        # Start with date-window rows, then for selected builds enrich every
        # mapped CR directly from the full unique_crs table by CR id. The second
        # lookup is required because build-selected JIRAs can map to older CRs
        # outside this week's unique_crs date window; the weekly CR table must
        # still show the authoritative CR Area/Subsystem/Functionality.
        cr_rows = fetch_weekly_crs(conn, schema, target, from_dt, to_dt)
        if selected_cr_ids:
            wanted_keys = set()
            for cr_id in selected_cr_ids:
                wanted_keys.update(_cr_lookup_keys(cr_id))
            detail_rows = _fetch_unique_cr_rows_by_ids(selected_cr_ids)
            by_cr = {}
            for r in detail_rows + cr_rows:
                row_keys = set()
                for ck in ('cr', 'mapped_cr', 'cr_number'):
                    row_keys.update(_cr_lookup_keys(r.get(ck)))
                matching = row_keys & wanted_keys
                if matching:
                    key = sorted(matching)[0]
                    by_cr.setdefault(key, r)
            cr_rows = list(by_cr.values())
        elif selected_builds:
            # Some JIRA tables do not carry a CR id. In that case, keep the
            # selected build's JIRA/open-JIRA rows exact and leave CR rows
            # date-filtered rather than fabricating CR details.
            logger.info('[WEEKLY_FULL] selected builds have no CR mapping columns; CR rows left date-filtered')

        # - 3. Pie aggregations -
        status_ctr = Counter(str(r.get('cr_status') or '').strip() for r in cr_rows if str(r.get('cr_status') or '').strip())
        area_ctr   = Counter(str(r.get('cr_area')   or '').strip() for r in cr_rows if str(r.get('cr_area') or '').strip())
        pie_status = [{'name': k, 'y': v} for k, v in sorted(status_ctr.items(), key=lambda x: x[0].lower())]
        pie_area   = [{'name': k, 'y': v} for k, v in sorted(area_ctr.items(),   key=lambda x: x[0].lower())]

        # - 4. Per-build CR/JIRA area matrix -
        # Area source rule:
        #   - Open/unmapped JIRAs: title-based bucket when available.
        #   - CRs / mapped JIRAs: Orbit CR area from unique_crs.cr_area.
        #   - No signal: blank (do not fabricate Unknown).
        cr_lookup = {}
        for r in cr_rows:
            for ck in ('cr', 'mapped_cr', 'cr_number'):
                for key in _cr_lookup_keys(r.get(ck)):
                    cr_lookup[key] = r

        def _area_for_jira(row):
            # If this JIRA is mapped to a CR, use Orbit's CR Area from unique_crs.
            for ck in ('mapped_cr', 'cr', 'cr_number'):
                for lookup_key in _cr_lookup_keys(row.get(ck)):
                    cr_row = cr_lookup.get(lookup_key)
                    if cr_row:
                        return str(cr_row.get('cr_area') or '').strip()

            # Open/unmapped JIRAs are bucketed from title only when a bucket is clear.
            return _area_from_open_jira_title(row.get('jira_title'))

        build_area_matrix = {}
        area_totals = Counter()
        for r in all_jira_rows:
            mb = str(r.get('metabuild') or '').strip()
            if not mb:
                continue
            area = _area_for_jira(r)
            if not area:
                continue
            build_area_matrix.setdefault(mb, {})[area] = build_area_matrix.setdefault(mb, {}).get(area, 0) + 1
            area_totals[area] += 1

        if not area_totals:
            for r in cr_rows:
                area = str(r.get('cr_area') or '').strip()
                if area:
                    area_totals[area] += 1

        areas = [a for a, _ in area_totals.most_common()]
        for mb in build_ids:
            build_area_matrix.setdefault(mb, {})
            for area in areas:
                build_area_matrix[mb].setdefault(area, 0)

        # - 5. Counts -
        total_jiras  = len({r.get('stability_ticket') for r in all_jira_rows if r.get('stability_ticket')})
        open_jiras   = len({r.get('stability_ticket') for r in open_jira_rows if r.get('stability_ticket')})
        total_crs    = len(cr_rows)
        valid_cats   = {'built', 'undisposed'}
        valid_crs    = sum(1 for r in cr_rows if (r.get('cr_category') or '').strip().lower() in valid_cats)

        # overall CRs (all in unique_crs table, not date-filtered)
        try:
            cur.execute(f'SELECT COUNT(*) AS cnt FROM {unique_tbl}')
            overall_crs = (cur.fetchone() or {}).get('cnt', 0) or 0
        except Exception:
            overall_crs = total_crs

        cur.close()

        payload = {
            'success':        True,
            'from_date':      from_s,
            'to_date':        to_s,
            'cr_rows':        cr_rows,
            'jira_rows':      jira_rows,
            'open_jira_rows': open_jira_rows,
            'build_ids':      build_ids,
            'available_build_ids': available_build_ids,
            'selected_build_ids': selected_builds or available_build_ids,
            'selection_enabled': bool(selected_builds),
            'selection_updated_at': weekly_selection.get('updated_at'),
            'build_area_matrix': build_area_matrix,
            'areas':          areas,
            'area_source_version': 6,
            'pie_status':     pie_status,
            'pie_area':       pie_area,
            'counts': {
                'total_jiras':  total_jiras,
                'open_jiras':   open_jiras,
                'total_crs':    total_crs,
                'overall_crs':  overall_crs,
                'valid_crs':    valid_crs,
                'build_count':  len(build_ids),
            },
        }
        payload = _json.loads(_json.dumps(payload, default=str))
        if selected_builds and can_edit and cache_requested:
            updated_at = _utc_now()
            set_weekly_report_selection(job.get('id'), {
                'selected_builds': selected_builds,
                'from_date': from_s,
                'to_date': to_s,
                'updated_at': updated_at,
                'updated_by': getattr(current_user, 'id', 'unknown'),
                'cache': payload,
            })
            payload['selection_updated_at'] = updated_at
        return jsonify(payload)
    except Exception as exc:
        logger.exception('[WEEKLY_FULL] %s', exc)
        return jsonify({'success': False, 'message': str(exc)}), 500


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/open_crs_full', methods=['GET'])
def api_published_open_crs_full(job_id=None, target_name=None):
    """Public API: unique open/analysis CR table for the published Live Status page."""
    if target_name is not None:
        job, err = _get_target_report_job_for_api(target_name)
    else:
        job, err = _get_published_job_for_api(job_id)
    if err:
        return jsonify(err[0]), err[1]
    target = (job.get('targets') or [''])[0]
    domain_filter = str(request.args.get('domain') or '').strip().upper()
    if domain_filter not in {'ADAS', 'FLEX', 'IVI'}:
        domain_filter = ''
    try:
        from datetime import date as _date, datetime as _datetime
        from dashboard_common import get_mysql_connection_db, get_schema_for_target, get_target_info

        schema = (get_schema_for_target(target) or '').strip('`')
        info = get_target_info(target) or {}
        prefix = str(info.get('db_prefix') or target or '').strip('`').lower()
        if not schema or not prefix:
            return jsonify({'success': False, 'message': 'Target schema/prefix not found', 'rows': []}), 404
        conn = get_mysql_connection_db(bu_key=schema)
        if not conn:
            return jsonify({'success': False, 'message': 'DB connection error', 'rows': []}), 500
        cur = conn.cursor(dictionary=True)

        def _tbl_exists(table_name):
            cur.execute('SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1', (schema, table_name))
            return cur.fetchone() is not None

        # SP-aware table selection: use sp_configs if sp param provided
        sp_param = str(request.args.get('sp') or '').strip()
        domain_low = (domain_filter or 'ADAS').lower()
        sp_cfg = {}
        if sp_param:
            sp_cfg = ((job.get('sp_configs') or {}).get(sp_param) or {})
        # Build candidate table list - when SP is given, NEVER fall back to default table
        table_candidates = []
        if sp_cfg.get(domain_low + '_uniq_table'):
            # sp_configs stores fully-qualified: pdt_stats_auto.nord_hgy_flex_5_1_9_0_unique_crs
            fq = sp_cfg[domain_low + '_uniq_table']
            table_candidates.append(fq.split('.')[-1])
        if not sp_param:
            # No SP - use domain-specific then default fallback
            if domain_filter: table_candidates.append(f'{prefix}_{domain_filter.lower()}_unique_crs')
            table_candidates.append(f'{prefix}_unique_crs')
        elif not table_candidates:
            # SP given but no table in sp_configs - try pattern-based name only
            table_candidates.append(f'{prefix}_{domain_low}_{sp_param.replace(".","_")}_unique_crs')
        table_name = next((t for t in table_candidates if _tbl_exists(t)), None)
        if not table_name:
            return jsonify({'success': True, 'target': target, 'rows': [], 'message': 'No unique_crs table found'})
        tbl = f'`{schema}`.`{table_name}`'
        cur.execute(f'SHOW COLUMNS FROM {tbl}')
        cols = {r.get('Field') for r in (cur.fetchall() or []) if r.get('Field')}

        def _norm_col(value):
            return re.sub(r'[^a-z0-9]+', '', str(value or '').lower())

        by_norm = {_norm_col(c): c for c in cols}

        def _first(candidates):
            for cand in candidates:
                hit = by_norm.get(_norm_col(cand))
                if hit:
                    return hit
            return ''

        def _sel(candidates, alias):
            col = _first(candidates)
            return (f'`{col}` AS `{alias}`' if col else f'NULL AS `{alias}`'), col

        cr_expr, cr_col = _sel(['mapped_cr', 'mapped_crs', 'Mapped CRs', 'Mapped CR', 'cr', 'cr_number', 'stability_ticket'], 'cr')
        raw_expr, raw_col = _sel(['cr', 'cr_number', 'stability_ticket'], 'raw_cr')
        title_expr, title_col = _sel(['cr_title', 'jira_title', 'title', 'summary'], 'cr_title')
        area_expr, area_col = _sel(['cr_area', 'area', 'ChangeRequestParticipant.Area'], 'cr_area')
        sub_expr, sub_col = _sel(['cr_subsystem', 'subsystem', 'ChangeRequestParticipant.Subsystem'], 'cr_subsystem')
        func_expr, func_col = _sel(['cr_functionality', 'cr_function', 'functionality', 'ChangeRequestParticipant.Functionality'], 'cr_functionality')
        status_expr, status_col = _sel(['cr_status', 'status', 'final_status'], 'cr_status')
        cat_expr_sel, cat_col = _sel(['cr_category', 'category', 'CR Category'], 'cr_category')
        age_expr, age_col = _sel(['cr_age', 'overall_age', 'age'], 'cr_age')
        created_expr, created_col = _sel(['cr_date', 'created_date', 'date_added__created', 'created', 'created_on'], 'cr_created_date')
        first_expr, first_col = _sel(['first_seen_date', 'first_seen', 'jira_date__first_instance', 'jira_date', 'created_date', 'cr_date', 'built_date'], 'first_instance')
        last_expr, last_col = _sel(['last_seen_date', 'last_seen', 'jira_date__last_instance', 'last_instance', 'updated_date', 'jira_date'], 'last_instance')
        notes_expr, notes_col = _sel(['latest_cr_notes', 'latest_notes', 'latest_comment', 'latest_comments', 'analysis', 'debug_notes', 'cr_notes', 'notes', 'comment'], 'latest_cr_notes')
        occ_expr, occ_col = _sel(['cr_occurrence', 'overall_cr_occurrence', 'jira_count', 'cr_____current_month', 'current_month_occurrence'], 'occurrence')
        priority_expr, priority_col = _sel(['cr_priority', 'priority', 'severity', 'Severity'], 'priority')
        assignee_expr, assignee_col = _sel(['cr_assignee', 'assignee', 'owner', 'assigned_to', 'ChangeRequest.Assignee', 'cr_created_by', 'created_by'], 'cr_assignee')
        built_expr, built_col = _sel(['built_date', 'builtdate', 'cr_built_date', 'closed_date', 'resolved_date', 'fix_built_date'], 'built_date')
        si_expr, si_col = _sel(['si_image', 'SI Image', 'simage', 'si_last_seen', 'last_seen_image', 'build_image', 'cr_si_image', 'image', 'cr_image', 'si'], 'si_image')
        domain_expr, domain_col = _sel(['domain', 'Domain', 'cr_domain', 'crash_domain', 'software_domain', 'platform_domain', 'sub_domain'], 'domain')
        if not cr_col:
            return jsonify({'success': True, 'target': target, 'rows': [], 'message': 'No CR column found'})

        def _sql_norm(col):
            return f"LOWER(REPLACE(REPLACE(REPLACE(TRIM(`{col}`), ' ', ''), '_', ''), '-', ''))"

        where = [f"NULLIF(TRIM(`{cr_col}`), '') IS NOT NULL"]
        # Open CRs tab must be driven strictly by CR Status. Do not include
        # category-based rows (undisposed/no-SIR/image) or other statuses like
        # new/in-progress; the user requested only Open / Analysis CR statuses.
        if not status_col:
            return jsonify({'success': True, 'target': target, 'rows': [], 'message': 'No CR status column found'})
        where.append(f"{_sql_norm(status_col)} IN ('open','analysis','inanalysis')")
        select_sql = ', '.join([cr_expr, raw_expr, title_expr, area_expr, sub_expr, func_expr, status_expr, cat_expr_sel, age_expr, created_expr, first_expr, last_expr, notes_expr, occ_expr, priority_expr, assignee_expr, built_expr, si_expr, domain_expr])
        order_col = last_col or first_col or cr_col
        cur.execute(f"SELECT {select_sql} FROM {tbl} WHERE {' AND '.join(where)} ORDER BY `{order_col}` DESC LIMIT 5000")
        raw_rows = cur.fetchall() or []

        def _ser(v):
            if isinstance(v, (_datetime, _date)):
                return v.strftime('%Y-%m-%d %H:%M:%S') if isinstance(v, _datetime) else v.isoformat()
            return '' if v is None else str(v)

        def _cr_key(v):
            m = re.search(r'(\d{5,9})', str(v or ''))
            return m.group(1) if m else re.sub(r'[^A-Z0-9]+', '', str(v or '').upper())

        def _domain_for(row):
            explicit = str(row.get('domain') or '').strip().upper()
            if explicit in {'ADAS', 'FLEX', 'IVI'}:
                return explicit
            text = ' '.join(str(row.get(k) or '') for k in ('domain', 'cr_area', 'cr_subsystem', 'cr_functionality', 'cr_title', 'latest_cr_notes', 'si_image')).upper()
            if any(x in text for x in ('ADAS', 'ADP', 'RIDE', 'VISION', 'CAMERA')):
                return 'ADAS'
            if 'FLEX' in text or re.search(r'\bFLE\b', text):
                return 'FLEX'
            # Automotive Gen5 rows are primarily ADAS/FLEX/IVI. When the source
            # table does not carry an explicit domain marker, keep the row under
            # IVI rather than returning an empty/null domain so the Open CRs tab
            # is never blank solely because domain inference failed.
            return 'IVI' if _is_core_deck_target(target) else ''

        seen = {}
        allowed_statuses = {'open', 'analysis', 'inanalysis'}
        for r in raw_rows:
            row = {k: _ser(v) for k, v in dict(r).items()}
            if str(row.get('cr_age') or '').strip().upper() in {'', '-', 'NA', 'N/A', 'NONE', 'NULL'}:
                try:
                    # CR Age is age from CR created date to current date for open/analysis CRs.
                    start_raw = str(row.get('cr_created_date') or row.get('cr_date') or '').strip()[:10]
                    start_dt = _datetime.fromisoformat(start_raw).date() if start_raw else None
                    if start_dt:
                        row['cr_age'] = str(max(0, (_datetime.now().date() - start_dt).days))
                    else:
                        row['cr_age'] = ''
                except Exception:
                    row['cr_age'] = ''
            if re.sub(r'[^a-z0-9]+', '', str(row.get('cr_status') or '').lower()) not in allowed_statuses:
                continue
            key = _cr_key(row.get('cr') or row.get('raw_cr'))
            if not key:
                continue
            row['domain'] = _domain_for(row)
            if domain_filter and row['domain'] != domain_filter:
                continue
            row['cr_display'] = str(row.get('cr') or row.get('raw_cr') or '').strip()
            old = seen.get(key)
            if not old or str(row.get('last_instance') or row.get('first_instance') or '') >= str(old.get('last_instance') or old.get('first_instance') or ''):
                seen[key] = row
        rows = list(seen.values())
        rows.sort(key=lambda r: (str(r.get('last_instance') or r.get('first_instance') or ''), str(r.get('cr_display') or '')), reverse=True)
        status_counts = {}
        area_counts = {}
        domain_counts = {}
        for r in rows:
            status_counts[r.get('cr_status') or 'Unknown'] = status_counts.get(r.get('cr_status') or 'Unknown', 0) + 1
            area_counts[r.get('cr_area') or 'Unknown'] = area_counts.get(r.get('cr_area') or 'Unknown', 0) + 1
            if r.get('domain'):
                domain_counts[r.get('domain')] = domain_counts.get(r.get('domain'), 0) + 1
        return jsonify({
            'success': True,
            'target': target,
            'is_auto_bu': _is_core_deck_target(target),
            'domain': domain_filter,
            'rows': rows,
            'count': len(rows),
            'status_counts': status_counts,
            'area_counts': area_counts,
            'domain_counts': domain_counts,
            'source_table': f'{schema}.{table_name}',
        })
    except Exception as exc:
        logger.exception('[OPEN_CRS_FULL] %s', exc)
        return jsonify({'success': False, 'message': str(exc), 'rows': []}), 500
    finally:
        try:
            cur.close(); conn.close()
        except Exception:
            pass
@live_status_publish_bp.route('/api/live_status/targets/<target_name>/weekly_summary', methods=['GET'])
def api_published_weekly_summary(job_id=None, target_name=None):
    """Public API: return weekly CR summary for a target."""
    if target_name is not None:
        job, err = _get_target_report_job_for_api(target_name)
    else:
        job, err = _get_published_job_for_api(job_id)
    if err:
        return jsonify(err[0]), err[1]
    target = (job.get('targets') or [''])[0]
    from_arg = (request.args.get('from') or '').strip()
    to_arg   = (request.args.get('to')   or '').strip()
    try:
        from datetime import datetime as _datetime
        from weekly_summary_service import normalize_to_monday_sunday, write_target_weekly_summary
        import json as _json
        if from_arg and to_arg:
            req_start = _datetime.strptime(from_arg[:10], '%Y-%m-%d').date()
            req_end   = _datetime.strptime(to_arg[:10],   '%Y-%m-%d').date()
            week_start, week_end = normalize_to_monday_sunday(req_start, req_end)
        else:
            week_start, week_end = normalize_to_monday_sunday()
        path = write_target_weekly_summary(target, week_start, week_end)
        with open(path, 'r', encoding='utf-8') as fh:
            payload = _json.load(fh) or {}
        table_name = payload.get('table_name') or f'weekly_summary_{target.lower()}'
        rows = payload.get(table_name) or []
        return jsonify({'success': True, 'rows': rows, 'table_name': table_name, 'payload': payload})
    except Exception as exc:
        logger.exception('[WEEKLY_SUMMARY] %s', exc)
        return jsonify({'success': False, 'message': str(exc), 'rows': []}), 500


def _append_live_status_mtbf_rows(job, rows, sheet_name_override=None):
    """Append revoke-time Live Status rows into the configured MTBF workbook."""
    import os
    import re
    import openpyxl
    from openpyxl.styles import Alignment
    from dashboard_routes import _get_target_excel_config, _normalize_excel_path

    target = (job.get('targets') or [''])[0]
    cfg = (_get_target_excel_config(target) or {}).get('mtbf', {})
    excel_path = cfg.get('excel_path', '')
    sheet_name = str(sheet_name_override or cfg.get('sheet_name') or '').strip()
    if not excel_path:
        raise RuntimeError('MTBF Excel is not configured for this target.')
    path = _normalize_excel_path(excel_path)
    if not os.path.exists(path):
        raise RuntimeError(f'MTBF Excel file not found: {path}')

    wb = openpyxl.load_workbook(path)
    actual_sheet = sheet_name if sheet_name in wb.sheetnames else ''
    if not actual_sheet and sheet_name:
        actual_sheet = next((s for s in wb.sheetnames if s.strip().lower() == sheet_name.lower()), '')
    if not actual_sheet and wb.sheetnames:
        actual_sheet = wb.sheetnames[0]
    if not actual_sheet:
        raise RuntimeError('No sheets found in MTBF workbook.')
    ws = wb[actual_sheet]

    def _norm(value):
        return re.sub(r'\s+', ' ', re.sub(r'[^a-z0-9 ]+', '', str(value or '').strip().lower().replace('_', ' ').replace('-', ' '))).strip()

    header_tokens = {
        'hours', 'total hours', 'tested hours', 'hours tested', 'total tested hours',
        'crashes', 'total crashes', 'crash count', 'total crash count',
        'mtbf', 'meta id', 'meta', 'build', 'builds', 'full build', 'builds full id',
        'week', 'date', 'reduction', 'reduction percent', 'reduction ', 'source', 'status', 'comments', 'notes',
    }
    best_header_row = 1
    best_score = -1
    for rr in range(1, min(ws.max_row or 1, 20) + 1):
        vals = [ws.cell(rr, c).value for c in range(1, (ws.max_column or 1) + 1)]
        score = len({_norm(v) for v in vals if str(v or '').strip()} & header_tokens)
        if score > best_score:
            best_score = score
            best_header_row = rr

    def _headers():
        return [str(ws.cell(best_header_row, c).value or '').strip() for c in range(1, (ws.max_column or 1) + 1)]

    def _find(headers, candidates):
        norms = [_norm(h) for h in headers]
        cand_norms = [_norm(c) for c in candidates]
        for c in cand_norms:
            for i, h in enumerate(norms):
                if h == c:
                    return i + 1
        for c in cand_norms:
            for i, h in enumerate(norms):
                if c and c in h:
                    return i + 1
        return 0

    def _ensure(candidates, label):
        headers = _headers()
        col = _find(headers, candidates)
        if col:
            return col
        col = (ws.max_column or 0) + 1
        ws.cell(best_header_row, col).value = label
        return col

    col_week = _ensure(['Week', 'Date'], 'Week')
    col_meta = _ensure(['Meta-ID', 'META-ID', 'Meta ID', 'META'], 'Meta-ID')
    col_build = _ensure(['Build(s) Full ID', 'Full Build', 'Build(s)', 'Builds', 'Build'], 'Build(s) Full ID')
    col_hours = _ensure(['Total Hours', 'Tested Hours', 'Hours', 'Hours Tested', 'Total Tested Hours'], 'Total Hours')
    col_crashes = _ensure(['Total Crashes', 'Crashes', 'Crash Count', 'Total Crash Count'], 'Total Crashes')
    col_mtbf = _ensure(['MTBF', 'MTBF Hrs', 'MTBF Hours', 'MTBF (hrs)', 'MTBF (Hours)'], 'MTBF')
    col_reduction = _find(_headers(), ['Reduction %', 'Reduction Percent', 'Reduction'])
    col_source = _find(_headers(), ['Source'])
    col_status = _find(_headers(), ['Build Status', 'Run Status', 'Status'])
    col_comments = _find(_headers(), ['MTBF Details', 'Notes', 'Comments'])

    def _num_or_text(value):
        text = str(value or '').strip()
        if not text:
            return None
        try:
            return float(text.replace(',', ''))
        except Exception:
            return text

    appended = 0
    for item in rows or []:
        meta = str((item or {}).get('meta_id') or '').strip()
        builds = str((item or {}).get('builds') or '').strip()
        if not meta and not builds:
            continue
        rr = (ws.max_row or best_header_row) + 1
        ws.cell(rr, col_week).value = str((item or {}).get('week') or '')[:10]
        ws.cell(rr, col_meta).value = meta
        build_cell = ws.cell(rr, col_build)
        build_cell.value = builds
        build_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(rr, col_hours).value = _num_or_text((item or {}).get('hours'))
        ws.cell(rr, col_crashes).value = _num_or_text((item or {}).get('crashes'))
        ws.cell(rr, col_mtbf).value = _num_or_text((item or {}).get('mtbf'))
        if col_reduction:
            ws.cell(rr, col_reduction).value = _num_or_text((item or {}).get('reduction_percent'))
        if col_source:
            ws.cell(rr, col_source).value = 'Live Status Revoke'
        if col_status:
            ws.cell(rr, col_status).value = 'Revoked'
        if col_comments:
            ws.cell(rr, col_comments).value = 'Saved from Live Status revoke flow'
        appended += 1

    if not appended:
        raise RuntimeError('No MTBF rows to save.')
    wb.save(path)
    return {'success': True, 'saved_count': appended, 'excel_path': excel_path, 'sheet_name': actual_sheet, 'header_row': best_header_row}



@live_status_publish_bp.route('/api/live_status/swpdt_status', methods=['GET'])

@login_required
def api_swpdt_status():
    """Return SWPDT JSON file age, last generated_at, total jobs, and thread health."""
    import os, json, threading
    from datetime import datetime, timezone
    from live_status_publish_service import _get_swpdt_json_path, _SWPDT_JSON_NETWORK, _SWPDT_JSON_LOCAL

    active_path = _get_swpdt_json_path()
    result = {
        'file_exists':    False,
        'file_age_min':   None,
        'generated_at':   None,
        'total_jobs':     None,
        'state_counts':   None,
        'poller_thread':  None,
        'poller_alive':   False,
        'active_path':    active_path,
        'using_network':  active_path == _SWPDT_JSON_NETWORK,
        'network_exists': os.path.exists(_SWPDT_JSON_NETWORK),
        'local_exists':   os.path.exists(_SWPDT_JSON_LOCAL),
    }
    if os.path.exists(active_path):
        result['file_exists'] = True
        mtime = os.path.getmtime(active_path)
        age_min = (datetime.now(timezone.utc).timestamp() - mtime) / 60
        result['file_age_min'] = round(age_min, 1)
        try:
            with open(active_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            result['generated_at'] = data.get('generated_at')
            result['total_jobs']   = data.get('total_jobs')
            result['state_counts'] = data.get('state_counts')
        except Exception as e:
            result['read_error'] = str(e)
    # Thread health
    for t in threading.enumerate():
        if 'swpdt' in t.name.lower():
            result['poller_thread'] = t.name
            result['poller_alive']  = t.is_alive()
            break
    return jsonify({'ok': True, 'status': result})


def _swpdt_build_tail(raw):
    text = str(raw or '').strip()
    if not text:
        return ''
    parts = [p.strip() for p in text.replace('/', '\\').split('\\') if p.strip()]
    return parts[-1] if parts else text


def _swpdt_meta_from_build(raw):
    text = _swpdt_build_tail(raw)
    match = re.search(r'-0*(\d{3,6})(?:\.\d+)?[-_]', text)
    return f"META-{match.group(1).zfill(5)}" if match else ''


def _swpdt_domain_for_build(build):
    software_product = str(build.get('software_product') or '').upper()
    build_text = str(build.get('build_id') or build.get('build') or build.get('build_name') or '').upper()
    flavor = str(build.get('product_flavor') or build.get('flavor') or '').upper()
    hay = ' '.join([str(build.get('domain') or '').upper(), software_product, build_text, flavor, str(build.get('taxonomy_path') or '').upper()])
    if 'ADAS' in hay:
        return 'ADAS'
    if 'IVI' in flavor or 'SAFEIVI' in hay or 'SAFETYIVI' in hay or 'NONSAFE_IVI' in hay:
        return 'IVI'
    if 'FLEX' in software_product or '_FLEX' in build_text or '.FLEX' in build_text or 'FLEX_' in build_text or flavor.startswith('FLEX'):
        return 'FLEX'
    if 'IVI' in hay:
        return 'IVI'
    return 'OTHER'



def _swpdt_target_tokens(target_name):
    """Return program tokens used to restrict Automotive SWPDT rows to the selected target.

    Current Report has separate ADAS/FLEX/IVI domain filters, but all three
    domains still need to stay inside the selected automotive program, e.g.
    NORD_HQX must not show NORD_HGY builds.
    """
    raw = str(target_name or '').strip().upper()
    tokens = [t for t in re.split(r'[^A-Z0-9]+', raw) if t]
    stop = {'NORD', 'AUTO', 'AUTOMOTIVE', 'PDT', 'QIPL', 'LIVE', 'STATUS', 'CRM', 'ENG'}
    out = []
    for token in tokens:
        if token in stop or len(token) < 3:
            continue
        if token not in out:
            out.append(token)
    return out


def _swpdt_build_haystack(build):
    return ' '.join(str(build.get(k) or '') for k in (
        'build_id', 'build', 'build_name', 'software_product',
        'product_flavor', 'flavor', 'taxonomy_path'
    )).upper()


def _swpdt_matches_target(build, target_name):
    tokens = _swpdt_target_tokens(target_name)
    if not tokens:
        return True
    hay = _swpdt_build_haystack(build)
    return any(token in hay for token in tokens)


def _all_swpdt_build_rows():
    import json as _json
    import os as _os
    from live_status_publish_service import _get_swpdt_json_path

    path = _get_swpdt_json_path()
    if not _os.path.exists(path):
        return []
    with open(path, 'r', encoding='utf-8') as fh:
        payload = _json.load(fh) or {}
    raw = payload.get('builds') if isinstance(payload, dict) else None
    if isinstance(raw, dict):
        return [b for b in raw.values() if isinstance(b, dict)]
    jobs = payload.get('jobs') if isinstance(payload, dict) else None
    if isinstance(jobs, list):
        return [j for j in jobs if isinstance(j, dict)]
    return []


def _swpdt_rows_for_job(job, q='', domain='ALL', limit=300):
    targets = job.get('targets') or []
    primary = targets[0] if targets else ''
    q_lower = str(q or '').strip().lower()
    domain = str(domain or 'ALL').strip().upper()
    is_auto = _is_core_deck_target(primary)

    rows = []
    if is_auto:
        for item in _all_swpdt_build_rows():
            build_id = item.get('build_id') or item.get('build') or item.get('build_name') or ''
            build_name = _swpdt_build_tail(build_id)

            if not build_name:
                continue
            if not _swpdt_matches_target(item, primary):
                continue
            row_domain = _swpdt_domain_for_build(item)
            if row_domain not in {'ADAS', 'FLEX', 'IVI'}:
                continue
            if domain in {'ADAS', 'FLEX', 'IVI'} and row_domain != domain:
                continue
            meta_id = _swpdt_meta_from_build(build_name)
            software_product = str(item.get('software_product') or '').strip()
            state = str(item.get('state') or item.get('status') or item.get('run_status') or '').strip().lower()
            run_status = 'running' if state in ('running', 'submitted', 'dispatched') else 'completed'
            hay = ' '.join([build_name, meta_id, software_product, row_domain]).lower()

            if q_lower and not (q_lower in hay or q_lower in meta_id.lower().replace('meta-', '').lstrip('0')):
                continue
            rows.append({
                'meta_id': meta_id,
                'build_name': build_name,
                'build_full': build_name,
                'software_product': software_product,
                'domain': row_domain,
                'run_status': run_status,
                'job_count': int(item.get('job_count') or 1),
                'device_count': int(item.get('device_count') or 0),
                'first_submitted': str(item.get('submitted') or item.get('first_submitted') or '')[:10],
                '_submitted_sort': str(item.get('submitted') or item.get('first_submitted') or ''),
            })
    else:
        from live_status_publish_service import load_swpdt_running_builds
        for item in load_swpdt_running_builds(primary.capitalize()):
            build_name = _swpdt_build_tail(item.get('build_name') or item.get('build_full') or '')
            meta_id = item.get('meta_id') or _swpdt_meta_from_build(build_name)
            hay = ' '.join([build_name, meta_id, str(item.get('software_product') or '')]).lower()
            if q_lower and not (q_lower in hay or q_lower in meta_id.lower().replace('meta-', '').lstrip('0')):
                continue
            rows.append({**item, 'meta_id': meta_id, 'build_name': build_name, 'build_full': build_name, 'domain': _swpdt_domain_for_build(item), '_submitted_sort': str(item.get('first_submitted') or '')})

    dedup = {}
    for row in rows:
        key = str(row.get('build_name') or row.get('build_full') or '').strip().upper()
        if not key:
            continue
        existing = dedup.get(key)
        if not existing:
            dedup[key] = row
            continue
        existing['job_count'] = int(existing.get('job_count') or 0) + int(row.get('job_count') or 1)
        existing['device_count'] = int(existing.get('device_count') or 0) + int(row.get('device_count') or 0)
        if row.get('run_status') == 'running':
            existing['run_status'] = 'running'
        if str(row.get('_submitted_sort') or '') > str(existing.get('_submitted_sort') or ''):
            existing['first_submitted'] = row.get('first_submitted') or existing.get('first_submitted')
            existing['_submitted_sort'] = row.get('_submitted_sort') or existing.get('_submitted_sort')

    out = list(dedup.values())
    out.sort(key=lambda r: (str(r.get('_submitted_sort') or r.get('first_submitted') or ''), str(r.get('build_name') or '')), reverse=True)
    for row in out:
        row.pop('_submitted_sort', None)
    return out[:max(1, min(int(limit or 300), 1000))]


def _published_report_builds(job):

    """Return only real build IDs (last path segment) for JIRA JQL search.
    build_full may be a UNC path like \\\\server\\share\\Skyros.LA.1.0.r1-00340-PERF.INT-1
    We only need the last segment: Skyros.LA.1.0.r1-00340-PERF.INT-1
    """
    def _extract_build_id(raw):
        """Strip UNC/share path prefix, return only the build ID segment."""
        b = str(raw or '').strip()
        if not b:
            return ''
        # Normalise slashes and split; take the last non-empty part
        parts = [p.strip() for p in b.replace('/', '\\').split('\\') if p.strip()]
        return parts[-1] if parts else b

    rows = [r for r in (job.get('draft_rows') or []) if str(r.get('run_status', '')).lower() == 'running']
    builds = []
    for row in rows:
        if row.get('isMerged') and row.get('merged_builds'):
            builds.extend([_extract_build_id(b) for b in (row.get('merged_builds') or [])])
        else:
            bf = _extract_build_id(row.get('build_full') or '')
            if bf:
                builds.append(bf)
    cleaned = []
    seen = set()
    for b in builds:
        if b and b.upper() not in seen and not b.upper().startswith('META-'):
            seen.add(b.upper())
            cleaned.append(b)
    return cleaned


def _build_published_current_jql(builds):
    def _q(value):
        return str(value or '').replace('"', '\\"')
    parts = [f'summary ~ "{_q(b)}"' for b in (builds or []) if str(b or '').strip()]
    if not parts:
        return ''
    return f"({' OR '.join(parts)}) AND filter = {JIRA_PDT_FILTER_ID} AND summary !~ \"tombstone\" ORDER BY created ASC"


def _count_jira_for_jql(jql):
    import os as _os, sys as _sys
    _scripts_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'scripts')
    if _scripts_dir not in _sys.path:
        _sys.path.insert(0, _scripts_dir)
    from config import JIRA_PASSWORD, JIRA_SERVER_ENDPOINT, JIRA_USER
    from fetch_consolidated_report import connect_jira
    jira_obj = connect_jira(JIRA_USER, JIRA_PASSWORD, JIRA_SERVER_ENDPOINT)
    result = jira_obj.search_issues(jql, startAt=0, maxResults=0, fields='summary')
    return int(getattr(result, 'total', 0) or 0)


def _run_published_current_report(job, force=False, custom_jql=''):
    import os as _os, sys as _sys
    _scripts_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'scripts')
    if _scripts_dir not in _sys.path:
        _sys.path.insert(0, _scripts_dir)
    from fetch_consolidated_report import run_consolidated_report


    target = (job.get('targets') or [''])[0]
    builds = _published_report_builds(job)
    domain    = (request.args.get('domain') or '').strip().upper() or None
    build_key = _build_key(builds) if builds else _build_key(request.args.get('build_key') or '')
    sc = get_report_sidecar(target, build_key, domain)

    import logging as _logging
    _rlog = _logging.getLogger(__name__)
    _rlog.info(f"[CURRENT_REPORT] job_id={job.get('id')}, target={target!r}, builds={builds}, force={force}")
    # Only use the persisted JQL - never auto-generate from builds.
    # The editor is responsible for saving the JQL before publish.
    persisted_jql = (sc.get('jql') or '').strip()
    jql = (custom_jql or '').strip() or persisted_jql
    if not jql:
        return {'ok': False, 'error': 'No JQL saved for this report. Go to the editor, run the JIRA query, then publish again.'}, 400

    checked_at = _utc_now()
    cache = dict(sc.get('report_cache') or {})
    cached_report = cache.get('report') if isinstance(cache.get('report'), dict) else None
    old_count = cache.get('jira_count')
    new_count = _count_jira_for_jql(jql)

    # Check if cached report has CRs with missing enrichment (area/subsystem empty)
    def _cache_has_missing_cr_details(report):
        if not report:
            return False
        rows = report.get('hierarchical_report') or []
        cr_rows = [r for r in rows if r.get('cr') and r.get('cr') != 'NO_CR']
        if not cr_rows:
            return False
        missing = sum(1 for r in cr_rows if not (r.get('cr_area') or r.get('cr_subsystem') or r.get('cr_title')))
        # If more than half the CR rows are missing details, force re-enrich
        return missing > len(cr_rows) / 2

    cache_stale = _cache_has_missing_cr_details(cached_report)

    if cached_report and not force and not cache_stale and old_count == new_count and (cache.get('jql') or '') == jql:
        cache['last_fresh_check_at'] = checked_at
        set_sidecar_report_cache(target, build_key, domain, dict(cache, last_fresh_check_at=checked_at))
        cached_report.setdefault('meta', {})['cache_status'] = 'fresh_count_unchanged'
        cached_report['meta']['last_fresh_check_at'] = checked_at
        cached_report['meta']['jira_count'] = new_count
        cached_report['meta']['active_jql'] = jql
        cached_report['excluded_jiras'] = sc.get('excluded_jiras') or []
        return {'ok': True, 'report': cached_report, 'from_cache': True, 'active_jql': jql}, 200

    report = run_consolidated_report(
        build_ids=builds,
        filter_id=JIRA_PDT_FILTER_ID,
        traverse=True,
        enrich_orbit=True,
        target_name=target,
        custom_jql=jql,
    )
    report.setdefault('meta', {})['jira_count'] = new_count
    report['meta']['last_fresh_check_at'] = checked_at
    report['meta']['cache_status'] = 'rerun_stale_cr_details' if cache_stale else ('rerun_count_changed' if cached_report else 'initial_full_run')
    report['meta']['active_jql'] = jql
    report['excluded_jiras'] = sc.get('excluded_jiras') or []
    set_sidecar_report_cache(target, build_key, domain, {
        'jql': jql,
        'jira_count': new_count,
        'last_full_run_at': checked_at,
        'last_fresh_check_at': checked_at,
        'report': report,
    })
    return {'ok': True, 'report': report, 'from_cache': False, 'active_jql': jql}, 200


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/current_report', methods=['GET', 'POST'])
@live_status_publish_bp.route('/api/live_status/jobs/<job_id>/current_report', methods=['GET', 'POST'])
def api_published_current_report(job_id=None, target_name=None):
    if target_name is not None:
        job, err = _get_target_report_job_for_api(target_name)
        if err:
            return jsonify(err[0]), err[1]
    else:
        job = get_job(job_id)
        if not job:
            return jsonify({'ok': False, 'error': 'Job not found'}), 404
        # Published reports are publicly readable; write ops still need TARGET_GROUP
        if job.get('status') != 'published' and not (current_user.is_authenticated and _target_group_access()):
            return jsonify({'ok': False, 'error': 'Access denied'}), 403
    data = request.get_json(force=True, silent=True) or {}
    force = str(request.args.get('force') or '').lower() in ('1', 'true', 'yes') or bool(data.get('force'))
    custom_jql = data.get('custom_jql') or request.args.get('jql') or ''
    try:
        payload, status = _run_published_current_report(job, force=force, custom_jql=custom_jql)
        return jsonify(payload), status
    except Exception as exc:
        logger.exception('[PUBLISHED CURRENT REPORT] Failed: %s', exc)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@live_status_publish_bp.route('/api/live_status/swpdt_force_refresh', methods=['POST'])
@login_required
def api_swpdt_force_refresh():
    """
    Trigger a one-shot Axiom fetch right now and update SWPDT_job_summary.json.
    Returns the new file stats so the UI can show the result.
    """
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    if AXIOM_FETCH_DISABLED:
        logger.info('[SWPDT FORCE REFRESH] Axiom fetch disabled; skipping one-shot fetch.')
        return jsonify({'ok': False, 'disabled': True, 'error': 'Axiom fetch is temporarily disabled'}), 503
    import os, json, threading
    from datetime import datetime, timezone

    client_id     = os.environ.get('AXIOM_CLIENT_ID', '').strip()
    client_secret = os.environ.get('AXIOM_CLIENT_SECRET', '').strip()

    if not client_id or not client_secret:
        return jsonify({'ok': False, 'error': 'AXIOM_CLIENT_ID / AXIOM_CLIENT_SECRET not configured in .env'}), 400

    try:
        from scripts.fetch_axiom_jobs import (
            _get_token, fetch_swpdt_jobs, merge_and_prune, _save,
            DEFAULT_API_HOST, DEFAULT_APP_NAME, DEFAULT_PAGE_SIZE,
            DEFAULT_OUTPUT_DIR, OUTPUT_FILENAME, RETENTION_DAYS,
        )
        output_path = _SWPDT_JSON  # use the same path the service reads from

        logger.info('[SWPDT FORCE REFRESH] Starting one-shot fetch...')
        token     = _get_token(DEFAULT_API_HOST, client_id, client_secret)
        new_jobs  = fetch_swpdt_jobs(DEFAULT_API_HOST, token, DEFAULT_PAGE_SIZE, DEFAULT_APP_NAME)

        if not new_jobs:
            return jsonify({'ok': False, 'error': 'No jobs returned from Axiom - API may be down or no Running jobs today'}), 502

        final_jobs = merge_and_prune(output_path, new_jobs, RETENTION_DAYS)
        from datetime import datetime as _dt, timezone as _tz
        payload = {
            'generated_at':   _dt.now(_tz.utc).strftime('%Y-%m-%dT%H:%M:%SZ'),
            'taxonomy':       '/PDT',
            'hwpdt_excluded': '/PDT/QIPL/HW',
            'retention_days': RETENTION_DAYS,
            'total_jobs':     len(final_jobs),
            'total_devices':  sum(j.get('device_count', 0) for j in final_jobs),
            'state_counts':   {},
            'jobs':           final_jobs,
        }
        for j in final_jobs:
            s = j.get('state', 'Unknown')
            payload['state_counts'][s] = payload['state_counts'].get(s, 0) + 1

        _save(output_path, payload)
        logger.info('[SWPDT FORCE REFRESH] Done - %d jobs saved to %s', len(final_jobs), output_path)

        return jsonify({
            'ok':           True,
            'total_jobs':   len(final_jobs),
            'new_fetched':  len(new_jobs),
            'state_counts': payload['state_counts'],
            'generated_at': payload['generated_at'],
            'saved_to':     output_path,
        })
    except Exception as exc:
        logger.exception('[SWPDT FORCE REFRESH] Failed: %s', exc)
        return jsonify({'ok': False, 'error': str(exc)}), 500


# ---------------------------------------------------------------------------
# Save / Publish APIs (used by the editor hero buttons)
# Target-scoped: reads/writes the job JSON directly from the target folder
# Same pattern as Core Slides /api/core_deck/save
# ---------------------------------------------------------------------------

def _get_job_file_for_target(target_name: str):
    """Return (job_dict, job_file_path) by scanning target's jobs/ folder directly."""
    from live_status_publish_service import target_live_status_dir
    import json, os
    jobs_dir = os.path.join(target_live_status_dir(target_name), 'jobs')
    if not os.path.isdir(jobs_dir):
        return None, None
    best_job, best_path, best_ts = None, None, ''
    for fname in os.listdir(jobs_dir):
        if not fname.endswith('.json'):
            continue
        fpath = os.path.join(jobs_dir, fname)
        try:
            with open(fpath, 'r', encoding='utf-8') as fh:
                job = json.load(fh)
            if not isinstance(job, dict) or not job.get('id'):
                continue
            ts = str(job.get('updated_at') or job.get('created_at') or '')
            if best_job is None or ts > best_ts:
                best_job, best_path, best_ts = job, fpath, ts
        except Exception:
            continue
    return best_job, best_path


# ---------------------------------------------------------------------------
# SP Config API  — get and save per-SP table config inside the job JSON
# ---------------------------------------------------------------------------
@live_status_publish_bp.route('/api/live_status/targets/<target_name>/sp_configs', methods=['GET'])
@login_required
def api_get_sp_configs(target_name):
    job, _ = _get_job_file_for_target(target_name)
    if not job:
        return jsonify({'ok': False, 'error': 'No job found'}), 404
    sp_configs = job.get('sp_configs') or {}
    sp_siblings = _get_sp_siblings(target_name)
    return jsonify({'ok': True, 'sp_configs': sp_configs, 'sp_siblings': sp_siblings})


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/sp_tables', methods=['GET'])
@live_status_publish_bp.route('/api/live_status/targets/<target_name>/sp_table_options', methods=['GET'])
@login_required
def api_get_sp_tables(target_name):
    """Return all DB tables for this target grouped by SP and domain.
    Pattern: <prefix>_<domain>_<sp_slug>_<suffix>
    e.g. nord_hgy_adas_5_1_7_0_crs, nord_hgy_flex_5_1_9_0_openjiras
    """
    import re as _re
    from dashboard_common import get_mysql_connection_db
    target = str(target_name or '').strip().lower()
    if not target:
        return jsonify({'ok': False, 'error': 'target required'}), 400
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if not conn:
            return jsonify({'ok': False, 'error': 'DB connection failed'}), 500
        cur = conn.cursor()
        # New external-link jobs can use a display/URL target whose physical
        # tables are keyed by dashboard_status.db_name or target_name. Query
        # all plausible prefixes instead of relying on one URL slug.
        prefixes = {target}
        # Common external-link naming variants:
        # SECA_LE_IVI_1_0 may be stored as seca_ivi_1_0_* or seca_ivi_1_0.
        for variant in (
            target.replace('_le_', '_'),
            target.replace('_LE_', '_').lower(),
            target.replace('_safe_ivi_', '_ivi_'),
            target.replace('_nonsafe_ivi_', '_ivi_'),
        ):
            variant = str(variant or '').strip('_').lower()
            if variant:
                prefixes.add(variant)
        schema_hint = ''
        try:
            from dashboard_common import get_target_info, get_schema_for_target, get_mysql_connection_db
            info = get_target_info(target_name) or {}
            for key in ('target_name', 'target_display', 'db_name', 'db_prefix', 'sp_name'):
                value = str(info.get(key) or '').strip().lower()
                if value:
                    prefixes.add(value)
            schema_hint = str(get_schema_for_target(target_name) or '').strip('`')
        except Exception:
            pass
        # Also resolve dashboard_status directly. Newly created jobs often use
        # a URL/display target absent from TARGETS_CONFIG, while this row still
        # contains the physical db_name and BU schema.
        try:
            meta_conn = get_mysql_connection_db(bu_key=None)
            meta_cur = meta_conn.cursor(dictionary=True)
            meta_cur.execute(
                """SELECT target_name, target_display, db_name, db_prefix, sp_name, bu
                   FROM pdt_stats_dashboard.dashboard_status
                   WHERE is_active=1
                     AND (LOWER(target_name)=LOWER(%s)
                          OR LOWER(target_display)=LOWER(%s)
                          OR LOWER(db_name)=LOWER(%s))
                   ORDER BY id DESC LIMIT 1""",
                (target_name, target_name, target_name),
            )
            meta = meta_cur.fetchone() or {}
            for key in ('target_name', 'target_display', 'db_name', 'db_prefix', 'sp_name'):
                value = str(meta.get(key) or '').strip().lower()
                if value:
                    prefixes.add(value)
                    prefixes.add(value.replace('_le_', '_'))
            if not schema_hint and meta.get('bu'):
                from dashboard_common import get_schema_for_bu
                schema_hint = str(get_schema_for_bu(str(meta['bu']).upper()) or '').strip('`')
            meta_cur.close()
            meta_conn.close()
        except Exception as exc:
            logger.info('[SP TABLES] dashboard metadata lookup skipped for %s: %s', target_name, exc)
        all_tables = []
        seen_tables = set()
        for prefix in prefixes:
            cur.execute(
                'SELECT TABLE_SCHEMA, TABLE_NAME FROM information_schema.TABLES '
                'WHERE TABLE_NAME LIKE %s ORDER BY TABLE_SCHEMA, TABLE_NAME',
                (prefix.replace('_', r'\_') + '%',)
            )
            for row in cur.fetchall() or []:
                key = (row[0], row[1])
                if key not in seen_tables:
                    seen_tables.add(key)
                    all_tables.append(key)
        if schema_hint:
            all_tables = [row for row in all_tables if row[0] == schema_hint] or all_tables
        cur.close()
        conn.close()
    except Exception as exc:
        logger.exception('[SP TABLES] %s', exc)
        return jsonify({'ok': False, 'error': str(exc)}), 500

    # Suffixes we care about per domain
    # IMPORTANT: longer suffixes first - 'unique_crs' ends with 'crs', 'closed_jiras' ends with 'jiras'
    SUFFIXES = ['unique_crs', 'closed_jiras', 'openjiras', 'crs', 'jiras']
    DOMAINS  = ['adas', 'flex', 'ivi']

    # Group tables: result[sp_cpl][domain][suffix] = 'schema.table'.
    # Use the prefix that actually matched the DB table; a new external job
    # may be opened with target_display/target_name while tables use db_name.
    result = {}
    for schema, tname in all_tables:
        tl = tname.lower()
        for dom in DOMAINS:
            matched_prefix = next(
                (prefix + '_' + dom + '_' for prefix in prefixes
                 if tl.startswith(prefix + '_' + dom + '_')),
                ''
            )
            if not matched_prefix:
                continue
            rest = tl[len(matched_prefix):]  # e.g. '5_1_7_0_crs'
            for suf in SUFFIXES:
                if rest.endswith('_' + suf):
                    sp_slug = rest[: -(len(suf) + 1)]  # e.g. '5_1_7_0'
                    sp_cpl  = sp_slug.replace('_', '.')  # e.g. '5.1.7.0'
                    fq = schema + '.' + tname
                    result.setdefault(sp_cpl, {}).setdefault(dom.upper(), {})[suf] = fq
                    break

    # Fallback for target tables whose domain is embedded in the target name,
    # e.g. seca_ivi_1_0_jiras rather than seca_ivi_1_0_ivi_<cpl>_jiras.
    if not result:
        for schema, tname in all_tables:
            tl = tname.lower()
            suffix = next((s for s in SUFFIXES if tl.endswith('_' + s)), '')
            if not suffix:
                continue
            stem = tl[:-(len(suffix) + 1)]
            matched = next((p for p in prefixes if stem == p or stem.startswith(p + '_')), '')
            if not matched:
                continue
            domain = next((d.upper() for d in DOMAINS if f'_{d}_' in f'_{stem}_'), 'IVI')
            cpl_match = re.search(r'(\d+(?:_\d+){1,3})$', stem)
            cpl = cpl_match.group(1).replace('_', '.') if cpl_match else 'DEFAULT'
            result.setdefault(cpl, {}).setdefault(domain, {})[suffix] = f'{schema}.{tname}'

    return jsonify({
        'ok': True,
        'tables': result,
        'target': target_name,
        'prefixes_checked': sorted(prefixes),
        'table_count': len(all_tables),
        'matched_table_count': sum(len(domains) for sp in result.values() for domains in sp.values()),
    })


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/add_sp', methods=['POST'])
@login_required
def api_add_sp(target_name):
    """Insert a new SP (cpl) row into dashboard_status for this target."""
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    payload = request.get_json(silent=True) or {}
    cpl = str(payload.get('cpl') or '').strip()
    import re as _re
    if not cpl or not _re.match(r'^\d+\.\d+\.\d+\.\d+$', cpl):
        return jsonify({'ok': False, 'error': 'Invalid CPL format. Use e.g. 5.1.9.0'}), 400
    target = str(target_name or '').strip()
    if not target:
        return jsonify({'ok': False, 'error': 'target required'}), 400
    try:
        from dashboard_common import get_mysql_connection_db
        bu = (get_bu_for_target(target) or 'AUTO').upper()
        conn = get_mysql_connection_db(bu_key=None)
        if not conn:
            return jsonify({'ok': False, 'error': 'DB connection failed'}), 500
        cur = conn.cursor()
        # Check if already exists
        cur.execute(
            'SELECT id FROM pdt_stats_dashboard.dashboard_status '
            'WHERE target_name=%s AND cpl=%s LIMIT 1',
            (target, cpl)
        )
        existing = cur.fetchone()
        if existing:
            # Re-activate if inactive
            cur.execute(
                'UPDATE pdt_stats_dashboard.dashboard_status '
                'SET is_active=1 WHERE target_name=%s AND cpl=%s',
                (target, cpl)
            )
            conn.commit()
            cur.close(); conn.close()
            return jsonify({'ok': True, 'cpl': cpl, 'action': 'reactivated'})
        # Insert new row
        cur.execute(
            'INSERT INTO pdt_stats_dashboard.dashboard_status '
            '(target_name, cpl, bu, is_active) VALUES (%s, %s, %s, 1)',
            (target, cpl, bu)
        )
        conn.commit()
        cur.close(); conn.close()
        logger.info('[ADD SP] %s added cpl=%s to target=%s', getattr(current_user,'id','?'), cpl, target)
        return jsonify({'ok': True, 'cpl': cpl, 'action': 'inserted'})
    except Exception as exc:
        logger.exception('[ADD SP] %s', exc)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/sp_configs', methods=['POST'])
@login_required
def api_save_sp_configs(target_name):
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    job, _ = _get_job_file_for_target(target_name)
    if not job:
        return jsonify({'ok': False, 'error': 'No job found'}), 404
    payload = request.get_json(silent=True) or {}
    sp_configs = payload.get('sp_configs') or {}
    if not isinstance(sp_configs, dict):
        return jsonify({'ok': False, 'error': 'sp_configs must be a dict'}), 400
    from live_status_publish_service import save_job_meta
    saved = save_job_meta(job['id'], {'sp_configs': sp_configs})
    if not saved:
        return jsonify({'ok': False, 'error': 'Save failed'}), 500
    return jsonify({'ok': True, 'sp_configs': saved.get('sp_configs') or {}})


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/save', methods=['POST'])
@login_required
def api_save_job(target_name):
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    job, _job_path = _get_job_file_for_target(target_name)
    if not job:
        return jsonify({'ok': False, 'error': 'No job found for target'}), 404

    payload = request.get_json(silent=True) or {}
    try:
        if isinstance(payload, dict) and 'rows' in payload:
            saved = save_job_rows(job['id'], payload.get('rows') or [], getattr(current_user, 'id', 'unknown'))
        else:
            # Metadata/no-op save: keep existing draft rows intact, but refresh the
            # job timestamp through the service writer so index/locks stay consistent.
            saved = save_job_meta(job['id'], {})
        if not saved:
            return jsonify({'ok': False, 'error': 'Save failed'}), 500
        return jsonify({'ok': True, 'updated_at': saved.get('updated_at'), 'row_count': len(saved.get('draft_rows') or [])})
    except Exception as exc:
        logger.exception('[LIVE STATUS TARGET SAVE] %s', exc)
        return jsonify({'ok': False, 'error': str(exc)}), 500


@live_status_publish_bp.route('/api/live_status/targets/<target_name>/publish', methods=['POST'])
@login_required
def api_publish_job(target_name):
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    job, _job_path = _get_job_file_for_target(target_name)
    if not job:
        return jsonify({'ok': False, 'error': 'No job found for target'}), 404

    payload = request.get_json(silent=True) or {}
    username = getattr(current_user, 'username', None) or getattr(current_user, 'id', 'unknown')
    try:
        # Persist browser rows first when supplied. publish_job snapshots
        # draft_rows into published_rows, so viewers don't depend on localStorage.
        if isinstance(payload, dict) and 'rows' in payload:
            job = save_job_rows(job['id'], payload.get('rows') or [], username) or job
        published = publish_job(job['id'], username)
        if not published:
            return jsonify({'ok': False, 'error': 'Publish failed'}), 500
        return jsonify({
            'ok': True,
            'published_at': published.get('published_at'),
            'published_by': published.get('published_by'),
            'status': published.get('status'),
            'row_count': len(published.get('published_rows') or []),
        })
    except Exception as exc:
        logger.exception('[LIVE STATUS TARGET PUBLISH] %s', exc)
        return jsonify({'ok': False, 'error': str(exc)}), 500


# - Delete job -
@live_status_publish_bp.route('/api/live_status/jobs/<job_id>/delete', methods=['POST'])
@login_required
def api_delete_job(job_id):
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    job = get_job(job_id)
    if not job:
        return jsonify({'ok': False, 'error': 'Job not found'}), 404
    if job.get('status') == 'published':
        return jsonify({'ok': False, 'error': 'Cannot delete a published job. Revoke it first.'}), 400
    ok = delete_job(job_id)
    if ok:
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'Delete failed'}), 500


# - Revoke job -
@live_status_publish_bp.route('/api/live_status/jobs/<job_id>/revoke', methods=['POST'])
@login_required
def api_revoke_job(job_id):
    if not _target_group_access():
        return jsonify({'ok': False, 'error': 'Access denied'}), 403
    job = get_job(job_id)
    if not job:
        return jsonify({'ok': False, 'error': 'Job not found'}), 404
    username = getattr(current_user, 'username', None) or getattr(current_user, 'id', 'unknown')
    data = request.get_json(silent=True) or {}
    reason = str(data.get('reason') or '').strip()
    updated = revoke_job(job_id, username=username, reason=reason)
    if updated:
        return jsonify({'ok': True, 'status': 'revoked', 'revoked_by': username})
    return jsonify({'ok': False, 'error': 'Revoke failed'}), 500


# - Build-wise consolidated report -
@live_status_publish_bp.route('/api/live_status/targets/<target_name>/build_wise_report', methods=['GET'])
@login_required
def api_build_wise_report(target_name):
    """All builds (latest---oldest) with per-build consolidated CR+JIRA data.

    Uses the same jiras/openjiras/unique_crs join logic as weekly_full but
    without a date window - scans ALL rows so every historical build appears.

    Query params:
      build       : single build ID - return full CR detail rows for that build
      domain      : ADAS | FLEX | IVI  (AUTO BU only)
      crash_types : comma list: system,ssr,process,open_jira  (AUTO BU only)
    """
    import re as _re
    from collections import defaultdict, Counter
    from datetime import date as _date, datetime as _datetime
    from dashboard_common import get_schema_for_target, get_mysql_connection_db

    job, err = _get_target_report_job_for_api(target_name)
    if err:
        return jsonify(err[0]), err[1]
    target   = (job.get('targets') or [''])[0]
    is_auto  = _is_core_deck_target(target)

    selected_build = (request.args.get('build') or '').strip()
    domain_filter  = (request.args.get('domain') or '').strip().upper()
    sp_filter      = (request.args.get('sp') or '').strip()          # e.g. 5.1.9.0
    ct_raw         = (request.args.get('crash_types') or 'system,ssr,process,open_jira').strip()
    crash_types    = {c.strip().lower() for c in ct_raw.split(',') if c.strip()} or {'system','ssr','process','open_jira'}

    schema = get_schema_for_target(target)
    if not schema:
        return jsonify({'success': False, 'message': 'Target schema not found'}), 404
    conn = get_mysql_connection_db(bu_key=schema)
    if not conn:
        return jsonify({'success': False, 'message': 'DB connection error'}), 500

    sc  = schema.strip('`')
    tgt = target.strip('`.')

        # SP-aware table selection: use sp_configs tables when sp param provided
    def _sp_tables(sp, domain, sc, tgt):
        """Return list of (jiras_tbl, open_tbl, unique_tbl) tuples for the given SP+domain.
        When domain is empty (All), returns tables for every configured domain.
        """
        if not sp:
            # No SP - use default tables (all domains mixed)
            return [(f'`{sc}`.`{tgt}_jiras`',
                     f'`{sc}`.`{tgt}_openjiras`',
                     f'`{sc}`.`{tgt}_unique_crs`')]
        sp_cfg = ((job.get('sp_configs') or {}).get(sp) or {})
        if not domain:
            # All domains - collect every domain that has tables configured
            seen, result = set(), []
            for dom_key in ['adas', 'flex', 'ivi', 'csp', 'safe-ivi', 'nonsafe-ivi']:
                j_fq = sp_cfg.get(dom_key + '_jiras_table', '')
                o_fq = sp_cfg.get(dom_key + '_openjiras_table', '')
                u_fq = sp_cfg.get(dom_key + '_uniq_table', '')
                if j_fq or o_fq:
                    j_tbl = f'`{sc}`.`{j_fq.split(".")[-1]}`' if j_fq else f'`{sc}`.`{tgt}_{dom_key}_{sp.replace(".","_")}_jiras`'
                    o_tbl = f'`{sc}`.`{o_fq.split(".")[-1]}`' if o_fq else f'`{sc}`.`{tgt}_{dom_key}_{sp.replace(".","_")}_openjiras`'
                    u_tbl = f'`{sc}`.`{u_fq.split(".")[-1]}`' if u_fq else f'`{sc}`.`{tgt}_{dom_key}_{sp.replace(".","_")}_unique_crs`'
                    key = j_tbl
                    if key not in seen:
                        seen.add(key)
                        result.append((j_tbl, o_tbl, u_tbl))
            if result:
                return result
            # No sp_configs entries - fall back to default tables
            return [(f'`{sc}`.`{tgt}_jiras`',
                     f'`{sc}`.`{tgt}_openjiras`',
                     f'`{sc}`.`{tgt}_unique_crs`')]
        # Specific domain requested
        dom_low = domain.lower()
        def _tbl(key, fallback):
            fq = sp_cfg.get(dom_low + key, '')
            if fq:
                return f'`{sc}`.`{fq.split(".")[-1]}`'
            sp_slug = sp.replace('.', '_')
            return f'`{sc}`.`{tgt}_{dom_low}_{sp_slug}{fallback}`'
        return [(_tbl('_jiras_table', '_jiras'),
                 _tbl('_openjiras_table', '_openjiras'),
                 _tbl('_uniq_table', '_unique_crs'))]

    table_sets = _sp_tables(sp_filter, domain_filter, sc, tgt)
    # Primary tables (first set) used for unique_crs lookup
    jiras_tbl, open_tbl, unique_tbl = table_sets[0]

    cur = conn.cursor(dictionary=True)

    # - helpers (same as weekly_full) -
    def _tbl_exists(fq):
        n = fq.replace('`', '')
        try:
            s, t = n.split('.', 1)
        except ValueError:
            return True
        cur.execute('SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1', (s, t))
        return cur.fetchone() is not None

    def _table_cols(fq):
        try:
            cur.execute(f'SHOW COLUMNS FROM {fq}')
            return {r.get('Field') for r in (cur.fetchall() or []) if r.get('Field')}
        except Exception:
            return set()

    def _ser(v):
        if isinstance(v, (_datetime, _date)):
            return v.strftime('%Y-%m-%d %H:%M:%S') if hasattr(v, 'hour') else str(v)
        return v

    def _ser_rows(rows):
        return [{k: _ser(v) for k, v in r.items()} for r in (rows or [])]

    def _norm_id(v):
        raw = str(v or '').strip().upper()
        if raw.endswith('.0'):
            raw = raw[:-2]
        return _re.sub(r'[\s\-_,.]', '', raw)

    def _cr_digits(v):
        m = _re.search(r'(\d{5,9})', str(v or ''))
        return m.group(1) if m else ''

    def _cr_lookup_keys(v):
        raw = _norm_id(v)
        if not raw:
            return set()
        keys = {raw}
        digits = _cr_digits(v) or _cr_digits(raw)
        if digits:
            keys.add(digits)
            keys.add(f'CR{digits}')
        if raw.startswith('CR') and raw[2:]:
            keys.add(raw[2:])
        return {k for k in keys if k}

    def _norm_build(raw):
        s = str(raw or '').strip()
        return s.replace('/', '\\').split('\\')[-1] if s else ''

    def _crash_type_from_title(title):
        """Classify jira_title into system / ssr / process using the same
        keyword rules as the JQL crash-type filters:

          Process : summary ~ ProcessDump OR ProcessCrash OR QNX OR Undetermined
          SSR     : (summary ~ sleep OR ssr) AND NOT (ProcessDump OR ProcessCrash OR QNX)
          System  : everything else (no ProcessDump/ProcessCrash/QNX/sleep/ssr/Undetermined)
        """
        t = str(title or '').lower()

        # Process crash keywords (highest priority)
        process_kw = ('processdump', 'processcrash', 'process dump', 'process crash',
                      'qnx', 'undetermined')
        is_process = any(kw in t for kw in process_kw)
        if is_process:
            return 'process'

        # SSR keywords - only when no process keywords present
        ssr_kw = ('ssr', 'sleep', 'subsystem restart')
        is_ssr = any(kw in t for kw in ssr_kw)
        if is_ssr:
            return 'ssr'

        # Default: System crash
        return 'system'

    def _area_from_open_jira_title(value):
        text = str(value or '').strip().lower()
        if not text:
            return ''
        if any(x in text for x in ('wconnect','wcnss','cnss','wlan','wi-fi','wifi','btfm','bluetooth','wireless')):
            return 'WConnect'
        if ' bt ' in f' {text} ' or text.startswith('bt ') or text.endswith(' bt'):
            return 'WConnect'
        if any(x in text for x in ('modem','mpss','ril','data call','lte','5g','nr','ims','qmi')):
            return 'Modem'
        if any(x in text for x in ('adsp','audio','qdsp')):
            return 'ADSP'
        if any(x in text for x in ('cdsp','compute dsp')):
            return 'CDSP'
        if any(x in text for x in ('trustzone','trust zone','qsee')) or text == 'tz' or ' tz ' in f' {text} ':
            return 'TZ'
        if any(x in text for x in ('apps','apss','android','kernel','framework','userspace')):
            return 'APPS'
        return ''

    def _domain_from_build_id(build_id):
        """Primary domain signal: read directly from the build/metabuild name.
        SA8797P_ADAS.HGX... - ADAS, CI_SA8797P_FLEX.HGX... - FLEX, rest - IVI.
        """
        b = str(build_id or '').upper()
        if '_ADAS' in b or '.ADAS' in b or 'ADAS_' in b or 'ADAS.' in b:
            return 'ADAS'
        if '_FLEX' in b or '.FLEX' in b or 'FLEX_' in b or 'FLEX.' in b:
            return 'FLEX'
        return 'IVI'   # everything else in AUTO BU is IVI

    def _domain_from_cr(area, sub, func, title):
        """Fallback only: derive domain from CR metadata when build name has no signal."""
        text = ' '.join([str(x or '') for x in (area, sub, func, title)]).upper()
        if any(x in text for x in ('ADAS','ADP','RIDE','VISION','CAMERA')):
            return 'ADAS'
        if 'FLEX' in text or _re.search(r'\bFLE\b', text):
            return 'FLEX'
        return 'IVI'

    try:
        # - 1. Read ALL rows from jiras + openjiras (no date filter) -
        base_cols = ['stability_ticket', 'jira_date', 'jira_title', 'serial_no', 'metabuild']
        extra_cr_cols = [
            'mapped_cr', 'cr', 'cr_number',
            'cr_area', 'area', 'ChangeRequestParticipant.Area',
            'cr_subsystem', 'subsystem', 'ChangeRequestParticipant.Subsystem',
            'cr_function', 'cr_functionality', 'functionality', 'ChangeRequestParticipant.Functionality',
            'application_domain',
        ]

        all_jira_rows  = []   # from jiras table
        all_open_rows  = []   # from openjiras table

        for jiras_tbl, open_tbl, unique_tbl in table_sets:
            for tbl, store, source in [(jiras_tbl, 'jira', 'jira'), (open_tbl, 'openjira', 'openjira')]:
                if not _tbl_exists(tbl):
                    continue
                cols   = _table_cols(tbl)
                sel    = [c for c in base_cols + extra_cr_cols if c in cols]
                if not sel:
                    continue
                mb_col = next((c for c in ('metabuild','MetaBuild','meta_build') if c in cols), None)
                if not mb_col:
                    continue
                try:
                    if selected_build:
                        tail = _norm_build(selected_build)
                        cur.execute(
                            f'SELECT {", ".join("`"+c+"`" for c in sel)} FROM {tbl} '
                            f'WHERE `{mb_col}` LIKE %s ORDER BY jira_date DESC LIMIT 5000',
                            (f'%{tail}%',)
                        )
                    else:
                        cur.execute(
                            f'SELECT {", ".join("`"+c+"`" for c in sel)} FROM {tbl} '
                            f'ORDER BY jira_date DESC LIMIT 30000'
                        )
                    for row in _ser_rows(cur.fetchall() or []):
                        row['_source']     = source
                        row['_build']      = _norm_build(row.get(mb_col) or '')
                        row['_crash_type'] = 'open_jira' if source == 'openjira' else _crash_type_from_title(row.get('jira_title') or '')
                        row['_domain_raw'] = str(row.get('application_domain') or '').strip().upper()
                        if source == 'jira':
                            all_jira_rows.append(row)
                        else:
                            all_open_rows.append(row)
                except Exception as e:
                    logger.warning('[BUILD_WISE] %s error: %s', tbl, e)

        all_rows = all_jira_rows + all_open_rows

        # - 2. Collect all CR IDs seen across all rows -
        all_cr_ids = set()
        for row in all_rows:
            for ck in ('mapped_cr', 'cr', 'cr_number'):
                cid = str(row.get(ck) or '').strip()
                if cid:
                    all_cr_ids.add(cid)
                    break

        # - 3. Fetch unique_crs for ALL cr_ids (same logic as weekly_full) -
        cr_detail = {}   # normalised_key - enriched row
        if _tbl_exists(unique_tbl) and all_cr_ids:
            try:
                ucols = _table_cols(unique_tbl)
                key_cols = [c for c in ('mapped_cr', 'cr', 'cr_number') if c in ucols]
                if key_cols:
                    def _col(name, alias=None):
                        alias = alias or name
                        return f'`{name}` AS `{alias}`' if name in ucols else f'NULL AS `{alias}`'

                    def _sql_norm_expr(col):
                        expr = f"UPPER(TRIM(`{col}`))"
                        for old in (' ', '-', '_', ',', '.0', '.'):
                            expr = f"REPLACE({expr}, '{old}', '')"
                        return expr

                    last_inst_col = next((c for c in ('jira_date__last_instance','last_instance','jira_date') if c in ucols), '')
                    cr_candidates = [c for c in ('mapped_cr','cr','cr_number') if c in ucols]
                    cr_expr = ("COALESCE(" + ', '.join([f"NULLIF(TRIM(`{c}`), '')" for c in cr_candidates]) + ") AS `cr`"
                               if cr_candidates else "NULL AS `cr`")
                    select_parts = [
                        cr_expr,
                        _col('mapped_cr'), _col('cr_number'),
                        _col('cr_title'), _col('cr_area'), _col('cr_subsystem'), _col('cr_functionality'),
                        _col('cr_status'), _col('cr_category'), _col('cr_age', 'overall_age'),
                        _col('jira_date'),
                        f'`{last_inst_col}` AS `last_instance`' if last_inst_col else 'NULL AS `last_instance`',
                    ]

                    # query in batches of 500
                    all_keys = sorted({k for cid in all_cr_ids for k in _cr_lookup_keys(cid)})
                    fetched_rows = []
                    batch = 500
                    for i in range(0, len(all_keys), batch):
                        chunk = all_keys[i:i+batch]
                        placeholders = ','.join(['%s'] * len(chunk))
                        where_parts  = [f"{_sql_norm_expr(col)} IN ({placeholders})" for col in key_cols]
                        flat_params  = tuple(x for _ in where_parts for x in chunk)
                        cur.execute(
                            f"SELECT {', '.join(select_parts)} FROM {unique_tbl} "
                            f"WHERE {' OR '.join(where_parts)} "
                            f"ORDER BY `{last_inst_col or key_cols[0]}` DESC",
                            flat_params,
                        )
                        fetched_rows.extend(_ser_rows(cur.fetchall() or []))

                    for row in fetched_rows:
                        area  = str(row.get('cr_area') or '').strip()
                        sub   = str(row.get('cr_subsystem') or '').strip()
                        func  = str(row.get('cr_functionality') or '').strip()
                        title = str(row.get('cr_title') or '').strip()
                        enriched = {
                            'cr_title':         title,
                            'cr_status':        str(row.get('cr_status') or '').strip(),
                            'cr_age':           str(row.get('overall_age') or '').strip(),
                            'si_last_seen':     str(row.get('last_instance') or row.get('jira_date') or '').strip(),
                            'cr_area':          area,
                            'cr_subsystem':     sub,
                            'cr_functionality': func,
                            'cr_category':      str(row.get('cr_category') or '').strip(),
                            'last_instance':    str(row.get('last_instance') or row.get('jira_date') or '').strip(),
                            'domain':           _domain_from_cr(area, sub, func, title) if is_auto else '',
                        }
                        for ck in ('cr', 'mapped_cr', 'cr_number'):
                            for key in _cr_lookup_keys(row.get(ck)):
                                cr_detail.setdefault(key, enriched)
            except Exception as e:
                logger.warning('[BUILD_WISE] unique_crs error: %s', e)

        # - 4. Build cr_lookup for area resolution (same as weekly_full) -
        cr_lookup = {}
        for key, detail in cr_detail.items():
            cr_lookup[key] = detail

        def _area_for_jira(row):
            for ck in ('mapped_cr', 'cr', 'cr_number'):
                for lk in _cr_lookup_keys(row.get(ck)):
                    d = cr_lookup.get(lk)
                    if d:
                        return d.get('cr_area', '')
            return _area_from_open_jira_title(row.get('jira_title'))

        # - 5. Enrich each row with domain -
        for row in all_rows:
            cr_id = ''
            for ck in ('mapped_cr', 'cr', 'cr_number'):
                cid = str(row.get(ck) or '').strip()
                if cid:
                    cr_id = cid
                    break
            detail = {}
            for lk in _cr_lookup_keys(cr_id):
                detail = cr_lookup.get(lk, {})
                if detail:
                    break
            if is_auto:
                # Build name is the ground truth - SA8797P_ADAS/FLEX/other.
                # CR metadata is only a fallback for rows with no build signal.
                row['_domain'] = (
                    _domain_from_build_id(row.get('_build') or '') or
                    row['_domain_raw'] or
                    detail.get('domain') or
                    _domain_from_cr(
                        row.get('cr_area') or row.get('area') or '',
                        row.get('cr_subsystem') or row.get('subsystem') or '',
                        row.get('cr_functionality') or row.get('functionality') or '',
                        row.get('jira_title') or ''
                    )
                )
            else:
                row['_domain'] = ''

        # - 6. Apply domain + crash_type filters -
        def _passes(row):
            if is_auto and domain_filter and row.get('_domain') != domain_filter:
                return False
            if row.get('_crash_type', 'system') not in crash_types:
                return False
            return True

        filtered = [r for r in all_rows if _passes(r)]

        # - 7. Group by build -
        build_map = defaultdict(list)
        for row in filtered:
            build_map[row.get('_build') or 'UNKNOWN'].append(row)

        def _build_sort_key(bid):
            m = _re.search(r'-(\d{3,6})(?:\.\d+)?-', bid)
            return -int(m.group(1)) if m else 0

        sorted_builds = sorted(build_map.keys(), key=_build_sort_key)

        # - 8. Build summary list -
        builds_summary = []
        for bid in sorted_builds:
            rows = build_map[bid]
            # Count unique CRs: use raw jira-table CR columns first;
            # if empty, resolve via stability_ticket -> cr_detail (unique_crs join).
            cr_ids = set()
            for r in rows:
                raw_cr = ''
                for ck in ('mapped_cr', 'cr', 'cr_number'):
                    cid = str(r.get(ck) or '').strip()
                    if cid:
                        raw_cr = cid
                        break
                if raw_cr:
                    norm = _re.sub(r'[^0-9]', '', raw_cr)
                    cr_ids.add(norm or raw_cr)
                    continue
                # Ticket -> unique_crs lookup
                ticket = str(r.get('stability_ticket') or '').strip()
                if ticket:
                    for lk in _cr_lookup_keys(ticket):
                        if cr_lookup.get(lk):
                            cr_ids.add(lk)
                            break
            builds_summary.append({
                'build_id':        bid,
                'total_crashes':   len(rows),
                'jira_count':      sum(1 for r in rows if r.get('_source') == 'jira'),
                'open_jira_count': sum(1 for r in rows if r.get('_source') == 'openjira'),
                'system_count':    sum(1 for r in rows if r.get('_crash_type') == 'system'),
                'ssr_count':       sum(1 for r in rows if r.get('_crash_type') == 'ssr'),
                'process_count':   sum(1 for r in rows if r.get('_crash_type') == 'process'),
                'cr_count':        len(cr_ids),
            })

        # - 9. Detail rows for selected build -
        detail_rows = []
        if selected_build:
            tail = _norm_build(selected_build)
            matched_bid = next((b for b in sorted_builds if tail.lower() in b.lower()), None)
            rows = build_map.get(matched_bid or tail, [])
            seen_cr = {}
            sno = 0
            for row in rows:
                cr_id = ''
                for ck in ('mapped_cr', 'cr', 'cr_number'):
                    cid = str(row.get(ck) or '').strip()
                    if cid:
                        cr_id = cid
                        break
                ticket   = str(row.get('stability_ticket') or '').strip()
                jira_ttl = str(row.get('jira_title') or '').strip()
                source   = row.get('_source', 'jira')
                crash_t  = row.get('_crash_type', 'system')
                domain   = row.get('_domain', '')

                # get enriched CR detail
                detail = {}
                for lk in _cr_lookup_keys(cr_id):
                    detail = cr_lookup.get(lk, {})
                    if detail:
                        break

                key = cr_id or ticket
                if key and key in seen_cr:
                    seen_cr[key]['cr_count'] = seen_cr[key].get('cr_count', 1) + 1
                    continue
                sno += 1
                entry = {
                    'sno':              sno,
                    'cr':               cr_id,
                    'jira':             ticket,
                    'title':            detail.get('cr_title') or jira_ttl,
                    'cr_count':         1,
                    'cr_status':        detail.get('cr_status', ''),
                    'cr_age':           detail.get('cr_age', ''),
                    'si_last_seen':     detail.get('si_last_seen', ''),
                    'cr_area':          detail.get('cr_area') or _area_for_jira(row),
                    'cr_subsystem':     detail.get('cr_subsystem') or str(row.get('cr_subsystem') or row.get('subsystem') or '').strip(),
                    'cr_functionality': detail.get('cr_functionality') or str(row.get('cr_functionality') or row.get('functionality') or '').strip(),
                    'cr_category':      detail.get('cr_category', ''),
                    'last_instance':    detail.get('last_instance', ''),
                    'source':           source,
                    'crash_type':       crash_t,
                    'domain':           domain,
                    'jira_date':        str(row.get('jira_date') or '').strip(),
                }
                if key:
                    seen_cr[key] = entry
                detail_rows.append(entry)

        cur.close()
        conn.close()
        return jsonify({
            'success':        True,
            'target':         target,
            'is_auto':        is_auto,
            'domain_filter':  domain_filter,
            'crash_types':    list(crash_types),
            'builds':         builds_summary,
            'detail_rows':    detail_rows,
            'selected_build': selected_build,
        })

    except Exception as exc:
        logger.exception('[BUILD_WISE] %s', exc)
        try:
            cur.close(); conn.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': str(exc)}), 500


@live_status_publish_bp.route('/api/live_status/target_group_members', methods=['GET'])
@login_required
def api_target_group_members():
    """Return list of TARGET_GROUP members for the Customize modal Members section."""
    try:
        members = sorted(list(TARGET_GROUP)) if TARGET_GROUP else []
        return jsonify({'ok': True, 'members': members})
    except Exception as exc:
        return jsonify({'ok': False, 'members': [], 'error': str(exc)})


@live_status_publish_bp.route('/build_report/api_docs')
@login_required
def build_report_api_docs():
    """Build Report API documentation page — shows how to share the API with external tools."""
    return render_template('public_build_report_api.html')
