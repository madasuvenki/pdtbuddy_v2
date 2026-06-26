import logging
logger = logging.getLogger(__name__)
import traceback
import uuid
import time
from datetime import datetime, date,timedelta
import json
from flask import request, url_for, flash
from flask_login import login_required, current_user
import os
from datetime import datetime as _dt


from collections import defaultdict  
  

from flask import (
    Blueprint,
    render_template,
    request,
    url_for,
    redirect,
    flash,
        session,jsonify,
    send_file,
)


from qdt_client import get_rework_info_from_qdt
import pandas as pd
from werkzeug.utils import secure_filename


from mysql.connector import Error
from mysql.connector import errorcode
from dashboard_service import save_meta_report_bulk, ensure_meta_builds_table


from dashboard_common import (
    VALID_TARGETS,
    get_mysql_connection_db,
    get_mysql_connection_db,
    get_schema_for_target,
    fq_table_for_target,
    get_target_info,
    get_bu_for_target,
    validate_target_availability,
    get_weekly_report_data,
    
)
from src.utils import (
    execute_and_fetch_all,
    execute_and_fetch_one_or_zero,
)
from dashboard_state import GLOBAL_REPORT_DATA_STORAGE

from dashboard_service import (
    get_build_report_for_target,
    build_mtbf_dashboard_payload,
    _round_if_number,
)
from dashboard_common import get_display_name_for_target

dashboard_bp = Blueprint("dashboard_bp", __name__)

# Persistent user/data storage. Do NOT keep generated Excel/config under static/,
# because static/ can be replaced when Buddy is rebuilt/recompiled/redeployed.
_PDTBUDDY_DATA_ROOT = os.environ.get(
    'PDTBUDDY_DATA_ROOT',
    r'\\sphere\pdtqipl_internal\PDTBuddy'
)
_EXCEL_PAGE_CONFIG_PATH = os.path.join(
    _PDTBUDDY_DATA_ROOT,
    'config',
    'target_excel_page_config.json',
)
_MANAGED_MTBF_EXCEL_DIR = os.path.join(
    _PDTBUDDY_DATA_ROOT,
    'excel_uploads',
    'mtbf',
)



def _load_target_excel_page_config():
    try:
        if os.path.exists(_EXCEL_PAGE_CONFIG_PATH):
            with open(_EXCEL_PAGE_CONFIG_PATH, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
                return data if isinstance(data, dict) else {}
    except Exception:
        logger.debug('[EXCEL CONFIG] load failed', exc_info=True)
    return {}

def _save_target_excel_page_config(config):
    os.makedirs(os.path.dirname(_EXCEL_PAGE_CONFIG_PATH), exist_ok=True)
    with open(_EXCEL_PAGE_CONFIG_PATH, 'w', encoding='utf-8') as fh:
        json.dump(config or {}, fh, indent=2)

def _get_target_excel_config(target_name):
    data = _load_target_excel_page_config()
    return (data.get(target_name) or {}) if isinstance(data, dict) else {}

def _update_target_excel_config(target_name, page_key, payload):
    data = _load_target_excel_page_config()
    target_cfg = data.get(target_name) or {}
    target_cfg[page_key] = {
        **(target_cfg.get(page_key) or {}),
        **(payload or {}),
        'updated_at': _dt.utcnow().isoformat() + 'Z',
    }
    data[target_name] = target_cfg
    _save_target_excel_page_config(data)
    return target_cfg[page_key]

def _normalize_excel_path(path_value):
    path = str(path_value or '').strip().strip('"').strip("'")
    return os.path.expanduser(path) if path else ''


def _safe_target_slug(target_name):
    import re
    return re.sub(r'[^A-Za-z0-9_.-]+', '_', str(target_name or 'target')).strip('._') or 'target'


_DEFAULT_COMPUTE_CR_TAG_ALIASES = [
    {
        'name': 'GMK',

        'aliases': [
            'GlymurWP_PDT', 'PDT_GLYMUR', 'glymur_pdt', 'mahua_pdt', 'PDT_Mahua',
            'kalambo_pdt', 'CNPDT_PL_Glymur', 'PDT_Glymur_EXTENDED',
            'PDT_Glymur_ExtendedStability', 'PCIE_GEN5_Supported',
        ],
    },
    {
        'name': 'Refresh MSM/License',
                'aliases': ['pdt_glymur_refresh', 'pdt_mahua_refresh', 'pdt_kalambo_refresh', 'likett hi'],

    },
    {
        'name': 'Karnali',
        'aliases': [
            'pdt_qipl_mahua_karnali', 'pdt_qipl_glymur_karnali',
            'MahuaWP_PDT_Karnali', 'PDT_Mahua_Karnali', 'MahuaWP_PDT_Karnali',
            'Desktop_PDT',
        ],
    },
    {
        'name': 'Kenai',
        'aliases': ['Kenai_PDT', 'qipl_PDT_Kenai', 'KenaiWP_PDT'],
    },
]


def _compute_cr_tag_alias_config_path():
    return os.path.join(_PDTBUDDY_DATA_ROOT, 'managed_excel', 'COMPUTE', 'cr_tag_aliases.json')


def _compute_cr_tag_cache_dir():
    path = os.path.join(_PDTBUDDY_DATA_ROOT, 'managed_excel', 'COMPUTE', 'GLYMUR')
    os.makedirs(path, exist_ok=True)
    return path


def _compute_cr_tag_cache_path(target_name):
    return os.path.join(_compute_cr_tag_cache_dir(), f"cr_tag_cache_{_safe_target_slug(target_name).lower()}.json")


def _load_compute_cr_tag_cache(target_name):
    path = _compute_cr_tag_cache_path(target_name)
    try:
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            rows = data.get('rows') if isinstance(data, dict) else {}
            if isinstance(rows, dict):
                return rows, data.get('updated_at')
    except Exception:
        logger.debug('[CR TAG CACHE] load failed', exc_info=True)
    return {}, None


def _save_compute_cr_tag_cache(target_name, rows):
    path = _compute_cr_tag_cache_path(target_name)
    payload = {
        'target': target_name,
        'updated_at': _dt.utcnow().isoformat() + 'Z',
        'rows': rows or {},
    }
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, indent=2)
    os.replace(tmp, path)
    return payload


def _load_compute_cr_tag_alias_config():
    path = _compute_cr_tag_alias_config_path()
    try:
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            groups = data.get('groups') if isinstance(data, dict) else data
            if isinstance(groups, list):
                cleaned = []
                for g in groups:
                    if not isinstance(g, dict):
                        continue
                    name = str(g.get('name') or '').strip()
                    aliases = [str(a).strip() for a in (g.get('aliases') or []) if str(a).strip()]
                    if name and aliases:
                        cleaned.append({'name': name, 'aliases': aliases})
                if cleaned:
                    return cleaned
    except Exception:
        logger.debug('[CR TAG ALIAS] load failed', exc_info=True)
    return [dict(name=g['name'], aliases=list(g['aliases'])) for g in _DEFAULT_COMPUTE_CR_TAG_ALIASES]


def _save_compute_cr_tag_alias_config(groups):
    cleaned = []
    for g in groups or []:
        if not isinstance(g, dict):
            continue
        name = str(g.get('name') or '').strip()
        aliases = []
        seen = set()
        for a in (g.get('aliases') or []):
            alias = str(a).strip()
            key = alias.lower()
            if alias and key not in seen:
                aliases.append(alias)
                seen.add(key)
        if name and aliases:
            cleaned.append({'name': name, 'aliases': aliases})
    if not cleaned:
        cleaned = [dict(name=g['name'], aliases=list(g['aliases'])) for g in _DEFAULT_COMPUTE_CR_TAG_ALIASES]
    path = _compute_cr_tag_alias_config_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as fh:
        json.dump({'groups': cleaned, 'updated_at': _dt.utcnow().isoformat() + 'Z'}, fh, indent=2)
    os.replace(tmp, path)
    return cleaned


_PDT_CR_TAG_JOBS = {}


def _match_compute_cr_tag_aliases(tags, groups=None):

    tags = [str(t).strip() for t in (tags or []) if str(t).strip()]
    tag_lut = {t.lower(): t for t in tags}
    out_groups = []
    out_aliases = []
    for g in (groups or _load_compute_cr_tag_alias_config()):
        name = str(g.get('name') or '').strip()
        matched = []
        for alias in (g.get('aliases') or []):
            actual = tag_lut.get(str(alias).strip().lower())
            if actual:
                matched.append(actual)
        if name and matched:
            out_groups.append(name)
            out_aliases.extend(matched)
    return ', '.join(dict.fromkeys(out_groups)), ', '.join(dict.fromkeys(out_aliases))


def _is_wcdma_target(target_name):
    """Return True for the Compute WCDMA target, where CR TAG UI/fetch is disabled."""
    import re as _re
    candidates = [target_name]
    try:
        info = get_target_info(target_name) or {}
        candidates.extend([
            info.get('target_name'),
            info.get('display_name'),
            info.get('db_prefix'),
            info.get('db_name'),
            info.get('chip_name'),
            info.get('product'),
        ])
    except Exception:
        pass
    for value in candidates:
        normalized = _re.sub(r'[^A-Z0-9]+', '', str(value or '').upper())
        if normalized == 'WCDMA':
            return True
    return False


def _is_compute_cr_tag_enabled_target(target_name):
    """CR TAG support is Compute-only, except for WCDMA."""
    from dashboard_common import get_bu_for_target as _get_bu_for_target
    return (_get_bu_for_target(target_name) or '').upper() == 'COMPUTE' and not _is_wcdma_target(target_name)


_MTBF_JSON_VIEW_NAMES = ["Glymur", "Mahua"]
_MTBF_JSON_HEADERS = ["Meta ID", "Build(s)", "Date", "Hours", "Total Crashes", "QC Crashes", "Product MTBF", "QC MTBF", "Comments"]
_MTBFSIMPLE_JSON_HEADERS = ["Meta ID", "Build(s)", "Date", "Hours", "Total Crashes", "MTBF", "Comments"]


def _mtbf_json_headers(is_compute=False):
    return list(_MTBF_JSON_HEADERS if is_compute else _MTBFSIMPLE_JSON_HEADERS)


def _mtbf_json_dir(target_name, view_name=None):
    # MTBF is JSON-backed. Compute keeps the historical shared GLYMUR folder for
    # Glymur/Mahua compatibility; all other targets get target-specific JSON.
    try:
        from dashboard_common import get_bu_for_target
        bu_key = (get_bu_for_target(target_name) or '').upper()
    except Exception:
        bu_key = ''
    if bu_key == 'COMPUTE':
        path = os.path.join(_PDTBUDDY_DATA_ROOT, 'managed_excel', 'COMPUTE', 'GLYMUR')
    else:
        path = os.path.join(_PDTBUDDY_DATA_ROOT, 'managed_excel', _safe_target_slug(bu_key or 'GENERAL'), _safe_target_slug(target_name))
    os.makedirs(path, exist_ok=True)
    return path




def _legacy_mtbf_json_path(target_name, view_name):
    return os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        'static',
        'mtbf_json',
        _safe_target_slug(target_name),
        f'{_mtbf_json_view_name(view_name).lower()}.json',
    )


def _mtbf_json_view_name(view_name):

    raw = str(view_name or '').strip()
    if raw.lower() == 'mtbf':
        return 'MTBF'
    for name in _MTBF_JSON_VIEW_NAMES:
        if raw.lower() == name.lower():
            return name
    return _MTBF_JSON_VIEW_NAMES[0]



def _mtbf_json_path(target_name, view_name):
    view = _mtbf_json_view_name(view_name)
    return os.path.join(_mtbf_json_dir(target_name, view), f'mtbf_{view.lower()}.json')



def _load_mtbf_json_payload(target_name, view_name):
    view = _mtbf_json_view_name(view_name)
    path = _mtbf_json_path(target_name, view)
    read_path = path
    if not os.path.exists(read_path):
        legacy_path = _legacy_mtbf_json_path(target_name, view)
        if os.path.exists(legacy_path):
            read_path = legacy_path
    if os.path.exists(read_path):
        try:
            with open(read_path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                data.setdefault('target', target_name)
                data.setdefault('view', view)
                data.setdefault('rows', [])
                # Auto-copy legacy static JSON into the managed network location.
                if read_path != path:
                    try:
                        _save_mtbf_json_payload(target_name, view, data)
                    except Exception:
                        logger.debug('[MTBF JSON] legacy copy failed: %s -> %s', read_path, path, exc_info=True)
                return data
        except Exception:
            logger.debug('[MTBF JSON] load failed: %s', read_path, exc_info=True)
    try:
        from dashboard_common import get_bu_for_target
        is_compute = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
    except Exception:
        is_compute = False
    return {'target': target_name, 'view': view, 'headers': _mtbf_json_headers(is_compute), 'rows': []}




def _save_mtbf_json_payload(target_name, view_name, payload):
    view = _mtbf_json_view_name(view_name)
    data = payload if isinstance(payload, dict) else {}
    data['target'] = target_name
    data['view'] = view
    try:
        from dashboard_common import get_bu_for_target
        is_compute = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
    except Exception:
        is_compute = False
    data['headers'] = _mtbf_json_headers(is_compute)

    data['updated_at'] = _dt.utcnow().isoformat() + 'Z'
    data['rows'] = data.get('rows') if isinstance(data.get('rows'), list) else []
    path = _mtbf_json_path(target_name, view)
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)
    os.replace(tmp, path)
    return data


def _num_or_blank(v, integer=False):
    if v in (None, ''):
        return ''
    try:
        n = float(str(v).replace(',', '').strip())
        return int(n) if integer else round(n, 2)
    except Exception:
        return ''


def _mtbf_json_row_from_payload(payload):
    build = str(payload.get('build') or payload.get('meta_id') or payload.get('build_full') or '').strip()
    build_full = str(payload.get('build_full') or payload.get('full_build') or build).strip()
    return {
        'id': str(payload.get('id') or '').strip() or (_dt.utcnow().strftime('%Y%m%d%H%M%S%f')),
        'build': build,
        'build_full': build_full,
        'date': str(payload.get('date') or payload.get('week') or '').strip(),
        'total_crashes': _num_or_blank(payload.get('total_crashes', payload.get('crashes')), integer=True),
        'qc_crashes': _num_or_blank(payload.get('qc_crashes'), integer=True),
        'product_mtbf': _num_or_blank(payload.get('product_mtbf')),
        'qc_mtbf': _num_or_blank(payload.get('qc_mtbf')),
        'mtbf': _num_or_blank(payload.get('mtbf', payload.get('qc_mtbf'))),
        'hours': _num_or_blank(payload.get('hours')),

        'comments': str(payload.get('comments') or payload.get('mtbf_details') or '').strip(),
    }


def _mtbf_json_to_preview_rows(rows, is_compute=False):
    out = []
    for i, r in enumerate(rows or [], start=1):
        common = [
            {'v': str(i), 'rs': 1, 'cs': 1, 'skip': False},
            {'v': str(r.get('build') or ''), 'rs': 1, 'cs': 1, 'skip': False},
            {'v': str(r.get('build_full') or r.get('full_build') or r.get('build') or ''), 'rs': 1, 'cs': 1, 'skip': False},
            {'v': str(r.get('date') or ''), 'rs': 1, 'cs': 1, 'skip': False},
            {'v': str(r.get('hours') or ''), 'rs': 1, 'cs': 1, 'skip': False},
            {'v': str(r.get('total_crashes') or ''), 'rs': 1, 'cs': 1, 'skip': False},
        ]
        if is_compute:
            common.extend([
                {'v': str(r.get('qc_crashes') or ''), 'rs': 1, 'cs': 1, 'skip': False},
                {'v': str(r.get('product_mtbf') or ''), 'rs': 1, 'cs': 1, 'skip': False},
                {'v': str(r.get('qc_mtbf') or ''), 'rs': 1, 'cs': 1, 'skip': False},
            ])
        else:
            common.append({'v': str(r.get('mtbf') or r.get('qc_mtbf') or ''), 'rs': 1, 'cs': 1, 'skip': False})
        common.append({'v': str(r.get('comments') or ''), 'rs': 1, 'cs': 1, 'skip': False})
        out.append(common)
    return out





def _mtbf_excel_sheet_to_json_rows(excel_path, sheet_name):
    """Convert one MTBF Excel sheet into the JSON row shape used by Compute MTBF.

    This is intentionally tolerant: it accepts the historical workbook headers and
    maps only the fields needed by the new JSON-backed chart/table.
    """
    import openpyxl
    from datetime import datetime as _dtt, date as _ddate
    path = _normalize_excel_path(excel_path)
    if not path or not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    actual_sheet = None
    for s in wb.sheetnames:
        if str(s).strip().lower() == str(sheet_name).strip().lower():
            actual_sheet = s
            break
    if not actual_sheet:
        return []
    ws = wb[actual_sheet]
    headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]

    def _ci(candidates):
        normalized = [str(c).strip().lower() for c in candidates]
        for cand in normalized:
            for i, h in enumerate(headers):
                if h == cand:
                    return i
        for cand in normalized:
            for i, h in enumerate(headers):
                if cand in h:
                    return i
        return None

    build_cols = [i for i, h in enumerate(headers) if 'build' in h]
    build_i = build_cols[-1] if build_cols else _ci(['meta-id', 'meta id', 'meta'])
    date_i = _ci(['date', 'week', 'run date'])
    total_crashes_i = _ci(['total crashes', 'product crashes', 'crash count', 'crashes', 'crash'])
    qc_crashes_i = _ci(['qc crashes', 'qc crash count', 'total qc crashes'])
    comments_i = _ci(['comments', 'mtbf details', 'notes'])
    hours_i = _ci(['hours', 'test hours', 'total hours', 'run hours'])
    sheet_key = str(sheet_name or '').strip().lower()
    product_mtbf_i = _ci([f'{sheet_key} product mtbf', 'product mtbf', 'prod mtbf', 'product mtbf hrs'])
    qc_mtbf_i = _ci([f'{sheet_key} qc mtbf', 'qc mtbf', 'qcmtbf', 'qc mtbf hrs'])

    def _cell(rn, idx):
        if idx is None:
            return ''
        v = ws.cell(rn, idx + 1).value
        if v is None:
            return ''
        if isinstance(v, (_dtt, _ddate)):
            return v.strftime('%Y-%m-%d')
        return str(v).strip()

    rows = []
    for rn in range(2, ws.max_row + 1):
        build = _cell(rn, build_i)
        if not build:
            continue
            row = {
            'id': f'excel-{sheet_key}-{rn}',
            'build': build,
            'date': _cell(rn, date_i),
            'total_crashes': _num_or_blank(_cell(rn, total_crashes_i), integer=True),
            'qc_crashes': _num_or_blank(_cell(rn, qc_crashes_i), integer=True),
            'product_mtbf': _num_or_blank(_cell(rn, product_mtbf_i)),
            'qc_mtbf': _num_or_blank(_cell(rn, qc_mtbf_i)),
            'hours': _num_or_blank(_cell(rn, hours_i)),
            'comments': _cell(rn, comments_i),
        }
        rows.append(row)
    return rows


def _migrate_compute_mtbf_excel_to_json_if_needed(target_name, excel_path):
    """One-time migration from the old Compute MTBF Excel sheets to JSON.

    If JSON already contains user-added rows but was never migrated, Excel rows
    are merged in without duplicating existing build/date pairs. After migration
    is marked complete, the Excel workbook is ignored and all future updates are
    JSON-only.
    """

    if not excel_path:
        return {}
    migrated = {}
    for view in _MTBF_JSON_VIEW_NAMES:
        current = _load_mtbf_json_payload(target_name, view)
        current_rows = current.get('rows') or []
        if current.get('migrated_from_excel'):
            migrated[view] = 'already_migrated'
            continue
        rows = _mtbf_excel_sheet_to_json_rows(excel_path, view)
        if rows:
            existing_keys = {
                (str(r.get('build') or '').strip().lower(), str(r.get('date') or '').strip())
                for r in current_rows
            }
            added = []
            for r in rows:
                key = (str(r.get('build') or '').strip().lower(), str(r.get('date') or '').strip())
                if key not in existing_keys:
                    added.append(r)
                    existing_keys.add(key)
            current['rows'] = added + current_rows
            current['migrated_from_excel'] = True
            current['source_excel_path'] = str(excel_path)
            _save_mtbf_json_payload(target_name, view, current)
            migrated[view] = f'migrated_{len(added)}'
        else:
            current['migrated_from_excel'] = True
            _save_mtbf_json_payload(target_name, view, current)
            migrated[view] = 'no_excel_rows'

    return migrated

def _mtbf_json_to_chart_data(rows, is_compute=False):

    data = []
    for r in rows or []:
        build = str(r.get('build') or '').strip()
        if not build:
            continue
        total_crashes = _num_or_blank(r.get('total_crashes'), integer=True) or 0
        qc_crashes = _num_or_blank(r.get('qc_crashes'), integer=True) or 0
        product_mtbf = _num_or_blank(r.get('product_mtbf')) or 0
        qc_mtbf = _num_or_blank(r.get('qc_mtbf')) or 0
        mtbf = _num_or_blank(r.get('mtbf')) or qc_mtbf or product_mtbf or 0

        hours_raw = r.get('hours')
        hours = float(hours_raw) if hours_raw not in (None, '', 0, '0') else 0
        full_build = str(r.get('build_full') or r.get('full_build') or build).strip()
        data.append({
            'build': build,
            'full_build': full_build,
            'label': build,
            'product_line': '',
            'week': str(r.get('date') or ''),
            'hours': hours,
            'crashes': total_crashes,
            'qc_crashes': qc_crashes,
            'product_mtbf': product_mtbf,
            'qc_mtbf': qc_mtbf,
                        'mtbf': mtbf,

            'comments': str(r.get('comments') or ''),
        })
    return data

def _mtbf_header_groups_for_validation(is_compute=False):
    # Uploaded files may contain extra columns, but they must include enough
    # headers for the chart parser to identify build + values.
    if is_compute:
        return [
            ('Build', ['build', 'meta']),
            ('Product MTBF or QC MTBF', ['product mtbf', 'qc mtbf', 'mtbf']),
        ]
    return [
        ('Build', ['build', 'meta']),
        ('MTBF', ['mtbf']),
    ]


def _validate_mtbf_workbook(path, is_compute=False):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    groups = _mtbf_header_groups_for_validation(is_compute)
    checked = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [str(ws.cell(1, c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]
        missing = []
        for label, needles in groups:
            if not any(any(n in h for n in needles) for h in headers):
                missing.append(label)
        checked.append({'sheet': sheet, 'missing': missing})
        if not missing:
            return sheet, []
    first_missing = checked[0]['missing'] if checked else [g[0] for g in groups]
    return '', first_missing


def _managed_mtbf_upload_dir(target_name):
    path = os.path.join(_MANAGED_MTBF_EXCEL_DIR, _safe_target_slug(target_name))
    os.makedirs(path, exist_ok=True)
    return path



def _read_excel_sheet_names(path_value):
    path = _normalize_excel_path(path_value)
    if not path:
        raise ValueError('Excel path is required.')
    if not os.path.exists(path):
        raise FileNotFoundError(f'Excel file not found: {path}')
    xl = pd.ExcelFile(path)
    return path, list(xl.sheet_names or [])

def _read_excel_sheet_preview(path_value, sheet_name, max_rows=200):
    """
    Read an Excel sheet with full merged-cell awareness.
    Returns (path, columns, rows, total_rows) where:
      - columns : list of header strings (row 1)
      - rows    : list of row-lists; each cell is a dict
                  {'v': display_value, 'rs': rowspan, 'skip': bool}
                  'skip'=True means this cell is covered by a merge above it
                  (the template must not render a <td> for it).
      - total_rows : total data rows in the sheet (excluding header)
    """
    import openpyxl
    path = _normalize_excel_path(path_value)
    if not path:
        raise ValueError('Excel path is required.')
    if not os.path.exists(path):
        raise FileNotFoundError(f'Excel file not found: {path}')
    if not sheet_name:
        raise ValueError('Sheet name is required.')

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found.')
    ws = wb[sheet_name]

    max_col = ws.max_column
    max_row = ws.max_row

    # Build merge map: (row, col) -> (top_left_row, top_left_col, row_span, col_span)
    # Cells that are NOT the top-left of a merge are marked as "skip".
    merge_info = {}   # (r,c) -> {'v': val, 'rs': rowspan, 'cs': colspan, 'skip': bool}
    for mr in list(ws.merged_cells.ranges):
        val = ws.cell(mr.min_row, mr.min_col).value
        rs  = mr.max_row  - mr.min_row + 1
        cs  = mr.max_col  - mr.min_col + 1
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                is_origin = (r == mr.min_row and c == mr.min_col)
                merge_info[(r, c)] = {
                    'v':    val,
                    'rs':   rs if is_origin else 1,
                    'cs':   cs if is_origin else 1,
                    'skip': not is_origin,
                }

    def _cell_val(r, c):
        info = merge_info.get((r, c))
        if info:
            return info
        raw = ws.cell(r, c).value
        from datetime import datetime as _dtt, date as _ddate
        if isinstance(raw, _dtt):
            raw = raw.strftime('%Y-%m-%d')   # strip time component
        elif isinstance(raw, _ddate):
            raw = raw.strftime('%Y-%m-%d')
        return {'v': '' if raw is None else str(raw).strip(), 'rs': 1, 'cs': 1, 'skip': False}

    # Row 1 = headers
    columns = []
    for c in range(1, max_col + 1):
        info = _cell_val(1, c)
        columns.append('' if info['skip'] else str(info['v'] or '').strip())

    total_rows = max(0, max_row - 1)
    limit = min(max_rows, total_rows)

    rows = []
    for r in range(2, 2 + limit):
        row = []
        for c in range(1, max_col + 1):
            info = _cell_val(r, c)
            row.append({
                'v':    str(info['v'] or '').strip() if not info['skip'] else '',
                'rs':   info['rs'],
                'cs':   info['cs'],
                'skip': info['skip'],
            })
        rows.append(row)

    return path, columns, rows, total_rows


def _read_excel_merged_aware(path_value, sheet_name):
    """Read an Excel sheet handling merged cells by forward-filling them."""
    import openpyxl
    path = _normalize_excel_path(path_value)
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f'Excel file not found: {path}')
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found.')
    ws = wb[sheet_name]
    # Unmerge and forward-fill merged cells
    merge_map = {}
    for merge_range in list(ws.merged_cells.ranges):
        min_row, min_col = merge_range.min_row, merge_range.min_col
        val = ws.cell(min_row, min_col).value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_map[(row, col)] = val
    rows_data = []
    for row in ws.iter_rows():
        row_vals = []
        for cell in row:
            v = merge_map.get((cell.row, cell.column), cell.value)
            row_vals.append('' if v is None else str(v))
        rows_data.append(row_vals)
    if not rows_data:
        return [], []
    headers = rows_data[0]
    data_rows = rows_data[1:]
    return headers, data_rows


def _build_device_summary_table(path_value, sheet_name):
    """
    Parse SW PDT deployment table from Excel.
    Handles merged cells (forward-fill), dynamic site detection,
    any number of sites (QIPL / CH / SD / etc.).
    Row 1: Form Factor | MCN | Storage | <SITE_A> | None | <SITE_B> | None | ... | Total | None
    Row 2: None        | None| None    | Delivered | Deployed | ...  | Delivered | Deployed
    Data starts row 3.
    """
    import openpyxl
    path = _normalize_excel_path(path_value)
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f'Excel file not found: {path}')
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found.')
    ws = wb[sheet_name]
    max_col = ws.max_column
    max_row = ws.max_row

    # -- build merge map: every cell -> its effective value --
    merge_map = {}
    for mr in list(ws.merged_cells.ranges):
        val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = val

    def cv(r, c):
        """Get cell value, respecting merge map."""
        v = merge_map.get((r, c), ws.cell(r, c).value)
        return '' if v is None else str(v).strip()

    def safe_int(v):
        try:
            s = str(v).strip()
            if s in ('', '-', '�', 'None', 'nan', 'none'): return 0
            return int(float(s))
        except Exception:
            return 0

    # -- Row 1: detect column roles --
    row1 = [cv(1, c) for c in range(1, max_col + 1)]   # 0-based list
    row2 = [cv(2, c) for c in range(1, max_col + 1)]

    SKIP_UPPER = {'FORM FACTOR', 'FORM', 'FACTOR', 'MCN', 'STORAGE',
                  'STORAGE TYPE', 'TOTAL', 'GRAND TOTAL', '', 'DELIVERED', 'DEPLOYED'}
    TOTAL_UPPER = {'TOTAL', 'GRAND TOTAL'}

    ff_col = mcn_col = sto_col = None
    for i, v in enumerate(row1):
        vu = v.upper()
        if ('FORM' in vu or vu == 'FF') and ff_col is None:   ff_col  = i
        elif 'MCN' in vu and mcn_col is None:                  mcn_col = i
        elif 'STOR' in vu and sto_col is None:                 sto_col = i

    # -- Detect site columns dynamically --
    # A site column in row1 is any non-empty value that is NOT a skip/total word.
    # Its DEL col = first col in row2 (same or next) with 'DELIV', DEP = 'DEPLO'.
    site_cols = {}   # ordered dict: site_name -> {'del': 0-based-idx, 'dep': 0-based-idx}
    i = 0
    while i < len(row1):
        v   = row1[i]
        vu  = v.upper()
        if v and vu not in SKIP_UPPER and vu not in TOTAL_UPPER:
            del_idx = dep_idx = None
            # scan row2 from same col up to +3
            for j in range(i, min(i + 4, len(row2))):
                r2u = row2[j].upper()
                if 'DELIV' in r2u and del_idx is None:  del_idx = j
                elif 'DEPLO' in r2u and dep_idx is None: dep_idx = j
            if del_idx is not None or dep_idx is not None:
                site_cols[v] = {'del': del_idx, 'dep': dep_idx}
                i = max(filter(lambda x: x is not None, [del_idx, dep_idx, i])) + 1
                continue
        i += 1

    if not site_cols:
        return None

    sites     = list(site_cols.keys())
    totals    = {s: {'del': 0, 'dep': 0} for s in sites}
    rows_out  = []
    grand_del = grand_dep = 0
    last_ff   = ''

    for row_num in range(3, max_row + 1):
        row = [cv(row_num, c) for c in range(1, max_col + 1)]
        if not any(row): continue

        ff  = row[ff_col]  if ff_col  is not None and ff_col  < len(row) else ''
        mcn = row[mcn_col] if mcn_col is not None and mcn_col < len(row) else ''
        sto = row[sto_col] if sto_col is not None and sto_col < len(row) else ''

        if not mcn: continue          # skip blank / grand-total rows
        if ff:  last_ff = ff
        else:   ff = last_ff          # carry forward merged Form Factor

        site_data = {}
        row_del = row_dep = 0
        for site, cols in site_cols.items():
            d = safe_int(row[cols['del']]) if cols['del'] is not None and cols['del'] < len(row) else 0
            p = safe_int(row[cols['dep']]) if cols['dep'] is not None and cols['dep'] < len(row) else 0
            site_data[site] = {'del': d, 'dep': p}
            totals[site]['del'] += d
            totals[site]['dep'] += p
            row_del += d
            row_dep += p
        grand_del += row_del
        grand_dep += row_dep
        rows_out.append({
            'form_factor': ff, 'mcn': mcn, 'storage': sto,
            'sites': site_data, 'total_del': row_del, 'total_dep': row_dep
        })

    # MCN chart: aggregate delivered per MCN
    mcn_agg = {}
    for r in rows_out:
        k = r['mcn'] or 'Unknown'
        mcn_agg[k] = mcn_agg.get(k, 0) + r['total_del']
    mcn_chart_data = [
        {'mcn': k, 'total': v}
        for k, v in sorted(mcn_agg.items(), key=lambda x: -x[1])
    ]

    return {
        'sites': sites, 'rows': rows_out, 'totals': totals,
        'grand_del': grand_del, 'grand_dep': grand_dep,
        'mcn_chart_data': mcn_chart_data
    }

# Use the single common BU icon map from config.py so dashboard/sidebar/shell
# pages render the same Font Awesome 5-compatible icons everywhere.
from config import BU_ICONS



def _build_bu_shell_context(active_bu_key=None):
    business_units = get_business_units() or {}
    bu_list = []
    for bu_key, bu_info in business_units.items():
        bu_key_upper = str(bu_key).upper()
        targets = get_targets_for_bu(bu_key_upper) if bu_key_upper != 'WEEKLY_QIPL_REPORTS' else []
        bu_list.append({
            'key': bu_key,
            'display_name': (bu_info or {}).get('display_name', bu_key),
            'targets_count': len(targets or []),
        })
    bu_list.sort(key=lambda x: str(x.get('display_name') or x.get('key') or '').upper())
    return {
        'active_bu_key': (active_bu_key or '').upper(),
        'bu_list': bu_list,
        'BU_ICONS': BU_ICONS,
        'shell_title': 'Business Unit Workspace',
    }



def _perf_now():
    return time.perf_counter()


def _perf_elapsed_ms(start_ts):
    return round((time.perf_counter() - start_ts) * 1000.0, 1)


def _perf_log_dashboard(target_name, section, phase_timings, extra=None):
    try:
        parts = [f"{name}={value}ms" for name, value in phase_timings]
        msg = f"[DASHBOARD PERF] target={target_name} section={section} " + " | ".join(parts)
        if extra:
            msg += " | " + " ".join(f"{k}={v}" for k, v in extra.items())
        logger.info(msg)
    except Exception:
        logger.debug("[DASHBOARD PERF] failed to log timings", exc_info=True)


def _jira_sort_key(row):
    jira_date = str((row or {}).get("jira_date") or (row or {}).get("date") or "").strip()
    metabuild = str((row or {}).get("metabuild") or (row or {}).get("build_id") or "").strip()
    stability_ticket = str((row or {}).get("stability_ticket") or "").strip()
    return (jira_date, metabuild, stability_ticket)



# -----------------------------------------------------------------------------


# -----------------------------------------------------------------------------
# CR OVERVIEW  �  direct read from unique_crs + jiras with 5-min in-memory cache
# Powered by src/cr_overview_service.py
# -----------------------------------------------------------------------------

from src.cr_overview_service import (
    fetch_cr_overview_data,
    fetch_cr_rows              as _svc_fetch_cr_rows,
    fetch_area_target_breakdown as _svc_fetch_area_targets,
    clear_cache                as _svc_clear_cr_cache,
    get_cache_info             as _svc_cr_cache_info,
    warmup_cache               as _svc_cr_warmup,
    SITE_KEYS                  as _CR_SITE_KEYS,
    SITE_LABELS                as _CR_SITE_LABELS,
)

# -- excluded-targets JSON path (kept for the excluded_targets UI) ------------
_EXCLUDED_TARGETS_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    'static', 'cr_overview_excluded_targets.json'
)

def _load_excluded_targets() -> set:
    try:
        if os.path.exists(_EXCLUDED_TARGETS_PATH):
            with open(_EXCLUDED_TARGETS_PATH, 'r', encoding='utf-8') as f:
                return set(json.load(f).get('excluded', []))
    except Exception:
        pass
    return set()

def _save_excluded_targets(excluded: list):
    os.makedirs(os.path.dirname(_EXCLUDED_TARGETS_PATH), exist_ok=True)
    with open(_EXCLUDED_TARGETS_PATH, 'w', encoding='utf-8') as f:
        json.dump({'excluded': sorted(set(excluded))}, f, indent=2)


# -----------------------------------------------------------------------------
# /api/cr_overview  � summary payload (hero KPIs, BU cards, charts, pivot)
# Source: direct read from unique_crs + jiras via cr_overview_service
# -----------------------------------------------------------------------------
@dashboard_bp.route('/api/cr_overview')
@login_required
def api_cr_overview():
    """
    Main summary API for the CR Overview landing page.

        Query params:
      bu            = ALL | <BU_KEY>
      target        = ALL | <target_name>
      dim           = cr_area | cr_status | cr_functionality | cr_subsystem
      status_filter = all (default, excludes invalid) | invalid
      site          = ALL (default) | PDT_QIPL | PDT_SD | PDT_CH |
                      PDT_QIPL_AND_CH | PDT_QIPL_AND_SD | PDT_ALL | PDT_SD_AND_CH
        """
    bu_filter     = (request.args.get('bu')            or 'ALL').strip().upper()
    tgt_filter    = (request.args.get('target')        or 'ALL').strip()
    # Multi-target support: ?targets=T1,T2,T3 overrides single ?target=
    _multi_tgts_raw = (request.args.get('targets') or '').strip()
    if _multi_tgts_raw:
        _multi_tgts = [t.strip() for t in _multi_tgts_raw.split(',') if t.strip()]
        if len(_multi_tgts) == 1:
            tgt_filter = _multi_tgts[0]
        elif len(_multi_tgts) > 1:
            tgt_filter = '__MULTI__:' + ','.join(_multi_tgts)
    dimension     = (request.args.get('dim')           or 'cr_area').strip().lower()
    status_filter = (request.args.get('status_filter') or 'all').strip().lower()

    # BU-card status checkbox filter (comma-separated cr_status values from frontend)
    _sfl_raw = (request.args.get('status_filter_list') or '').strip()
    status_filter_list = [s.strip() for s in _sfl_raw.split(',') if s.strip()] if _sfl_raw else []
    site_filter   = (request.args.get('site')          or 'ALL').strip().upper()
    date_from     = (request.args.get('date_from')     or '').strip()[:10]
    date_to       = (request.args.get('date_to')       or '').strip()[:10]

    # Column-level filters pushed up from CR Detail Table
    flt_cr       = (request.args.get('flt_cr')       or '').strip().lower()
    flt_area     = (request.args.get('flt_area')     or '').strip().lower()
    flt_sub      = (request.args.get('flt_sub')      or '').strip().lower()
    flt_func     = (request.args.get('flt_func')     or '').strip().lower()
    flt_proj     = (request.args.get('flt_proj')     or '').strip().lower()
    flt_age_min  = (request.args.get('flt_age_min')  or '').strip()
    flt_age_max  = (request.args.get('flt_age_max')  or '').strip()
    _flt_st_raw  = (request.args.get('flt_statuses') or '').strip()
    flt_statuses = [s.strip() for s in _flt_st_raw.split(',') if s.strip()] if _flt_st_raw else []
    _flt_si_raw  = (request.args.get('flt_sites') or '').strip()
    flt_sites    = [s.strip() for s in _flt_si_raw.split(',') if s.strip()] if _flt_si_raw else []
    flt_age_unit = (request.args.get('flt_age_unit') or 'days').strip().lower()

    if site_filter not in ('ALL',) + tuple(_CR_SITE_KEYS):
        site_filter = 'ALL'

    VALID_DIMS = {'bu_key', 'cr_area', 'cr_status', 'cr_functionality', 'cr_subsystem'}
    if dimension not in VALID_DIMS:
        dimension = 'cr_area'

    try:
        data, err = fetch_cr_overview_data(
            bu_filter=bu_filter,
            tgt_filter=tgt_filter,
            status_filter=status_filter,
            status_filter_list=status_filter_list,
            dimension=dimension,
            site_filter=site_filter,
            date_from=date_from,
            date_to=date_to,
            use_cache=True,
            flt_cr=flt_cr,
            flt_area=flt_area,
            flt_sub=flt_sub,
            flt_func=flt_func,
            flt_proj=flt_proj,
            flt_age_min=flt_age_min,
            flt_age_max=flt_age_max,
            flt_age_unit=flt_age_unit,
            flt_statuses=flt_statuses,
            flt_sites=flt_sites,
        )

        if err:
            return jsonify({'error': err}), 500

        # Always include the full site key list so the frontend can build the filter UI
        data['site_keys'] = _CR_SITE_KEYS
        data['site_filter'] = site_filter
        return jsonify(data)
    except Exception as e:
        logger.info(f'[CR OVERVIEW API] {e}')
        return jsonify({'error': str(e)}), 500



# -----------------------------------------------------------------------------
# /api/cr_overview/cr_rows  � paginated detail rows
# Source: direct read from unique_crs + jiras via cr_overview_service
# -----------------------------------------------------------------------------
@dashboard_bp.route('/api/cr_overview/cr_rows')
@login_required
def api_cr_overview_cr_rows():
    """
    Paginated detailed CR rows.

    Query params:
      bu        = ALL | <BU_KEY>
      target    = ALL | <target_name>
      dim       = cr_area | cr_status | cr_functionality | cr_subsystem
      dim_val   = filter by specific dimension value  (e.g. 'Camera')
      category  = all | undisposed | built | invalid | nosir  (default: undisposed)
      sort      = age_desc | age_asc | jira_desc       (default: age_desc)
      page      = 1-based  (default 1)
      per_page  = rows per page  (default 200, max 100000)
      site          = ALL | <SITE_KEY>
      status_filter  = all | invalid | nosir
      date_from / date_to (YYYY-MM-DD)
    """
    bu_filter     = (request.args.get('bu')       or 'ALL').strip().upper()
    tgt_filter    = (request.args.get('target')   or 'ALL').strip()
    # Multi-target support
    _multi_tgts_raw = (request.args.get('targets') or '').strip()
    if _multi_tgts_raw:
        _multi_tgts = [t.strip() for t in _multi_tgts_raw.split(',') if t.strip()]
        if len(_multi_tgts) == 1:
            tgt_filter = _multi_tgts[0]
        elif len(_multi_tgts) > 1:
            tgt_filter = '__MULTI__:' + ','.join(_multi_tgts)
    dimension     = (request.args.get('dim')      or 'cr_area').strip().lower()
    dim_val       = (request.args.get('dim_val')  or '').strip()

    # map legacy 'open_analysis' ? 'undisposed' so old frontend calls still work
    _cat_raw      = (request.args.get('category') or 'undisposed').strip().lower()
    category      = 'undisposed' if _cat_raw == 'open_analysis' else _cat_raw

    sort_by       = (request.args.get('sort')     or 'age_desc').strip().lower()
    site_filter   = (request.args.get('site')     or 'ALL').strip().upper()
    status_filter = (request.args.get('status_filter') or 'all').strip().lower()

    # BU-card status checkbox filter (comma-separated cr_status values from frontend)
    _sfl_raw = (request.args.get('status_filter_list') or '').strip()
    status_filter_list = [s.strip() for s in _sfl_raw.split(',') if s.strip()] if _sfl_raw else []

    # map legacy 'occ_desc' ? 'jira_desc'
    if sort_by == 'occ_desc':
        sort_by = 'jira_desc'

        # Validate site_filter
    if site_filter not in ('ALL',) + tuple(_CR_SITE_KEYS):
        site_filter = 'ALL'

    try:
        page = max(1, int(request.args.get('page', 1)))
        per_page = min(100000, max(10, int(request.args.get('per_page', 200))))
    except (ValueError, TypeError):
        page, per_page = 1, 200

    VALID_DIMS = {'bu_key', 'cr_area', 'cr_status', 'cr_functionality', 'cr_subsystem'}
    if dimension not in VALID_DIMS:
        dimension = 'cr_area'

    # If page is in Invalid or NoSIR mode, detail table must return that same category
    if status_filter == 'invalid':
        category = 'invalid'
    elif status_filter == 'nosir':
        category = 'nosir'

    try:
        result, err = _svc_fetch_cr_rows(

            bu_filter=bu_filter,
            tgt_filter=tgt_filter,
            category=category,
            dimension=dimension,
            dim_val=dim_val,
            sort_by=sort_by,
            site_filter=site_filter,
            date_from=(request.args.get('date_from') or '').strip()[:10],
            date_to=(request.args.get('date_to')   or '').strip()[:10],
            page=page,
            per_page=per_page,
            status_filter_list=status_filter_list,
            flt_age_min=(request.args.get('flt_age_min') or '').strip(),
            flt_age_max=(request.args.get('flt_age_max') or '').strip(),
            flt_age_unit=(request.args.get('flt_age_unit') or 'days').strip().lower(),
            flt_proj=(request.args.get('flt_proj') or '').strip(),
        )

        if err:
            return jsonify({'error': err}), 500
        return jsonify(result)
    except Exception as e:
        logger.info(f'[CR ROWS API] {e}')
        return jsonify({'error': str(e)}), 500

# -----------------------------------------------------------------------------
# /api/cr_overview/area_targets  � per-target breakdown for a dimension value
# e.g. ?area=Multimedia&dim=cr_area&bu=MOBILE
# Returns: targets[], all_areas[], site_keys[], site_labels{}
# -----------------------------------------------------------------------------
@dashboard_bp.route('/api/cr_overview/area_targets')
@login_required
def api_cr_overview_area_targets():
    """
    For a given dimension value (area/subsystem/functionality),
    return per-target CR count + avg age breakdown.
    Used by the Area?Target drill-down panel on the CR Overview page.
    """
    area_value    = (request.args.get('area')          or '').strip()
    dimension     = (request.args.get('dim')           or 'cr_area').strip().lower()
    bu_filter     = (request.args.get('bu')            or 'ALL').strip().upper()
    tgt_filter    = (request.args.get('target')        or 'ALL').strip()
    # Multi-target support
    _multi_tgts_raw = (request.args.get('targets') or '').strip()
    if _multi_tgts_raw:
        _multi_tgts = [t.strip() for t in _multi_tgts_raw.split(',') if t.strip()]
        if len(_multi_tgts) == 1:
            tgt_filter = _multi_tgts[0]
        elif len(_multi_tgts) > 1:
            tgt_filter = '__MULTI__:' + ','.join(_multi_tgts)

    status_filter = (request.args.get('status_filter') or 'all').strip().lower()
    site_filter   = (request.args.get('site')          or 'ALL').strip().upper()
    date_from     = (request.args.get('date_from')     or '').strip()[:10]
    date_to       = (request.args.get('date_to')       or '').strip()[:10]

    _sfl_raw = (request.args.get('status_filter_list') or '').strip()
    status_filter_list = [s.strip() for s in _sfl_raw.split(',') if s.strip()] if _sfl_raw else []


    valid_dims = {'bu_key', 'cr_area', 'cr_status', 'cr_functionality', 'cr_subsystem'}
    if dimension not in valid_dims:
        dimension = 'cr_area'
    if site_filter not in ('ALL',) + tuple(_CR_SITE_KEYS):
        site_filter = 'ALL'

    try:
        data, err = _svc_fetch_area_targets(
            area_value=area_value,
            dimension=dimension,
            bu_filter=bu_filter,
            tgt_filter=tgt_filter,
            status_filter=status_filter,
            status_filter_list=status_filter_list,
            site_filter=site_filter,
            date_from=date_from,
            date_to=date_to,
            flt_age_min=(request.args.get('flt_age_min') or '').strip(),
            flt_age_max=(request.args.get('flt_age_max') or '').strip(),
            flt_age_unit=(request.args.get('flt_age_unit') or 'days').strip().lower(),
        )

        if err:
            return jsonify({'error': err}), 500

        return jsonify(data)
    except Exception as e:
        logger.info(f'[CR AREA TARGETS API] {e}')
        return jsonify({'error': str(e)}), 500



# -----------------------------------------------------------------------------
# /api/cr_overview/targets  � active targets for a given BU
# -----------------------------------------------------------------------------
@dashboard_bp.route('/api/cr_overview/targets')
@login_required
def api_cr_overview_targets():
    """Return list of active (non-excluded) targets for a BU from dashboard_status."""
    from dashboard_common import get_business_units, get_targets_config
    bu_filter = (request.args.get('bu') or 'ALL').strip().upper()
    excluded  = _load_excluded_targets()

    try:
        targets_config = get_targets_config()
        business_units = get_business_units()

        if bu_filter == 'ALL':
            all_targets = sorted(targets_config.keys())
        else:
            bu_info = business_units.get(bu_filter) or {}
            all_targets = sorted(bu_info.get('targets') or [])

        targets = [t for t in all_targets if t not in excluded]
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    return jsonify({'targets': targets, 'bu': bu_filter})


# -----------------------------------------------------------------------------
# /api/cr_overview/excluded_targets  GET / POST
# -----------------------------------------------------------------------------


@dashboard_bp.route('/api/cr_overview/excluded_targets', methods=['GET'])
@login_required
def api_get_excluded_targets():
    """Return current excluded targets + full target list per BU from dashboard_status."""
    from dashboard_common import get_business_units, get_targets_config
    excluded = _load_excluded_targets()

    try:
        targets_config = get_targets_config()
        business_units = get_business_units()

        by_bu: dict = {}
        for bu_key, bu_info in sorted(business_units.items()):
            if bu_key in ('WEEKLY_QIPL_REPORTS',):
                continue
            # AUTO stores targets in admin_hierarchy, not a flat list
            if bu_key.upper() == 'AUTO':
                from dashboard_common import get_auto_target_keys, load_metadata_config
                tgt_list = sorted(get_auto_target_keys(load_metadata_config()))
            else:
                tgt_list = sorted(bu_info.get('targets') or [])
            if not tgt_list:
                continue
            by_bu[bu_key] = {
                'display_name': (bu_info.get('display_name') or bu_key).upper(),
                'targets': [
                    {'key': t, 'display': t, 'excluded': t in excluded}
                    for t in tgt_list
                ],
            }
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    return jsonify({'excluded': sorted(excluded), 'by_bu': by_bu})


@dashboard_bp.route('/api/cr_overview/excluded_targets', methods=['POST'])
@login_required
def api_save_excluded_targets():
    """Save excluded targets list {excluded: [...]} and bust the CR overview cache."""
    data     = request.get_json(force=True) or {}
    excluded = [str(t).strip() for t in (data.get('excluded') or []) if str(t).strip()]
    try:
        _save_excluded_targets(excluded)
        _svc_clear_cr_cache()   # force fresh data on next overview call
        return jsonify({'ok': True, 'excluded': sorted(excluded)})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# -----------------------------------------------------------------------------
# /admin/cr_overview/clear_cache  � manual cache bust (admin only)
# /admin/cr_overview/cache_stats  � inspect live cache entries (admin only)
# -----------------------------------------------------------------------------
@dashboard_bp.route('/api/hwpdt/excluded_targets', methods=['GET'])
@login_required
def api_get_hwpdt_excluded_targets():
    """Return current HWPDT excluded targets + full target list.

    Merges two sources:
      1. dashboard_status WHERE is_hwpdt=1  (managed targets, source='db')
      2. axiom_job_summary DISTINCT software_product WHERE team='HWPDT'
         (Axiom-discovered targets, source='axiom')

    Axiom-only targets (not in dashboard_status) are shown with source='axiom'
    so the UI can badge them and let admins exclude them.
    """
    import json as _json
    from dashboard_common import get_all_hwpdt_targets
    _path = r'\\sphere\pdtqipl_internal\PDTBuddy\HWPDT\hwpdt_excluded_targets.json'
    try:
        excluded = set(_json.load(open(_path, encoding='utf-8')).get('excluded', []))
    except Exception:
        excluded = set()

    # -- Source 1: dashboard_status is_hwpdt=1 targets ----------------------
    all_rows = get_all_hwpdt_targets()
    targets  = [
        {'key':     r['target_name'],
         'display': r.get('display_name') or r['target_name'],
         'bu_key':  r.get('bu_key', ''),
         'sp_name': r.get('sp_name', ''),
         'source':  'db',
         'excluded': r['target_name'] in excluded}
        for r in all_rows
    ]
    db_sp_names = {str(r.get('sp_name') or '').strip().upper() for r in all_rows if r.get('sp_name')}
    db_keys     = {r['target_name'].upper() for r in all_rows}

    # -- Source 2: Axiom axiom_job_summary HWPDT software_products ----------
    try:
        _conn = get_mysql_connection_db(bu_key=None)
        if _conn:
            _cur = _conn.cursor(dictionary=True)
            _cur.execute("""
                SELECT DISTINCT software_product,
                       MAX(submitted_at) AS last_seen,
                       COUNT(*)          AS job_count
                FROM pdt_stats_dashboard.axiom_job_summary
                WHERE team = 'HWPDT'
                  AND software_product IS NOT NULL
                  AND software_product != ''
                GROUP BY software_product
                ORDER BY software_product
            """)
            axiom_sps = _cur.fetchall() or []
            _cur.close()
            _conn.close()

            for row in axiom_sps:
                sp = str(row.get('software_product') or '').strip()
                if not sp:
                    continue
                # Skip if already covered by a db target (by sp_name or key match)
                if sp.upper() in db_sp_names or sp.upper() in db_keys:
                    continue
                targets.append({
                    'key':       sp,
                    'display':   sp,
                    'bu_key':    'HWPDT',
                    'sp_name':   sp,
                    'source':    'axiom',
                    'last_seen': str(row.get('last_seen') or '')[:10],
                    'job_count': int(row.get('job_count') or 0),
                    'excluded':  sp in excluded,
                })
    except Exception as _ax_err:
        logger.warning('[HWPDT EXCLUDED] Axiom SP fetch failed: %s', _ax_err)

    # Sort: db targets first (alpha), then axiom-only (alpha)
    targets.sort(key=lambda t: (0 if t['source'] == 'db' else 1, t['key'].lower()))
    return jsonify({'excluded': sorted(excluded), 'targets': targets})


@dashboard_bp.route('/api/hwpdt/excluded_targets', methods=['POST'])
@login_required
def api_save_hwpdt_excluded_targets():
    """Save updated HWPDT excluded targets list (admin only)."""
    import json as _json
    if getattr(current_user, 'role', None) != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json(force=True, silent=True) or {}
    excluded = sorted(set(data.get('excluded', [])))
    _path = r'\\sphere\pdtqipl_internal\PDTBuddy\HWPDT\hwpdt_excluded_targets.json'
    try:
        with open(_path, 'w', encoding='utf-8') as f:
            _json.dump({'excluded': excluded}, f, indent=2)
        return jsonify({'ok': True, 'excluded': excluded})
    except Exception as e:
        return jsonify({'error': str(e)}), 500



# -- HWPDT Playlist Aliases  GET / POST --------------------------------------
# Aliases are stored server-side in hwpdt_playlist_aliases.json on the network
# share so they are shared across all users and machines (not localStorage).
_HWPDT_ALIASES_NET   = r'\\sphere\pdtqipl_internal\PDTBuddy\HWPDT\hwpdt_playlist_aliases.json'
_HWPDT_ALIASES_LOCAL = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                    'hwpdt_playlist_aliases_local_backup.json')

def _load_hwpdt_aliases() -> list:
    """Load aliases from network path, fall back to local backup."""
    import json as _json
    for path in [_HWPDT_ALIASES_NET, _HWPDT_ALIASES_LOCAL]:
        if os.path.exists(path):
            try:
                with open(path, 'r', encoding='utf-8') as fh:
                    data = _json.load(fh)
                aliases = data.get('aliases') if isinstance(data, dict) else data
                if isinstance(aliases, list):
                    return aliases
            except Exception:
                pass
    return []

def _save_hwpdt_aliases(aliases: list) -> list:
    """Save aliases to network path and local backup. Returns saved list."""
    import json as _json
    from datetime import datetime as _dt
    payload = {
        'aliases':    aliases,
        'updated_at': _dt.utcnow().isoformat() + 'Z',
    }
    saved = []
    net_dir = os.path.dirname(_HWPDT_ALIASES_NET)
    if os.path.exists(net_dir):
        try:
            tmp = _HWPDT_ALIASES_NET + '.tmp'
            with open(tmp, 'w', encoding='utf-8') as fh:
                _json.dump(payload, fh, indent=2, ensure_ascii=False)
            os.replace(tmp, _HWPDT_ALIASES_NET)
            saved.append('network')
        except Exception as _e:
            logger.warning('[HWPDT ALIASES] network write failed: %s', _e)
    try:
        tmp = _HWPDT_ALIASES_LOCAL + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as fh:
            _json.dump(payload, fh, indent=2, ensure_ascii=False)
        os.replace(tmp, _HWPDT_ALIASES_LOCAL)
        saved.append('local')
    except Exception as _e:
        logger.warning('[HWPDT ALIASES] local write failed: %s', _e)
    return aliases


@dashboard_bp.route('/api/hwpdt/aliases', methods=['GET'])
@login_required
def api_get_hwpdt_aliases():
    """Return shared HWPDT playlist aliases from the network JSON file."""
    try:
        aliases = _load_hwpdt_aliases()
        return jsonify({'ok': True, 'aliases': aliases})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e), 'aliases': []}), 500


@dashboard_bp.route('/api/hwpdt/aliases', methods=['POST'])
@login_required
def api_save_hwpdt_aliases():
    """Save HWPDT playlist aliases to the shared network JSON file.
    Body: { aliases: [{id, name, playlists:[...]}] }
    Any logged-in user can save (aliases are a shared team resource).
    """
    try:
        data    = request.get_json(force=True, silent=True) or {}
        aliases = data.get('aliases')
        if not isinstance(aliases, list):
            return jsonify({'ok': False, 'error': 'aliases must be a list'}), 400
        # Sanitise each alias entry
        clean = []
        for a in aliases:
            if not isinstance(a, dict):
                continue
            name = str(a.get('name') or '').strip()
            if not name:
                continue
            pls = [str(p).strip() for p in (a.get('playlists') or []) if str(p).strip()]
            clean.append({
                'id':        a.get('id') or int(__import__('time').time() * 1000),
                'name':      name,
                'playlists': pls,
            })
        _save_hwpdt_aliases(clean)
        logger.info('[HWPDT ALIASES] saved %d aliases by %s',
                    len(clean), getattr(current_user, 'id', 'unknown'))
        return jsonify({'ok': True, 'aliases': clean})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@dashboard_bp.route('/admin/cr_overview/clear_cache', methods=['POST'])
@login_required
def admin_clear_cr_overview_cache():
    """Admin-only: evict all CR overview cache entries."""
    from config import ADMIN_USERS
    if getattr(current_user, 'id', '') not in ADMIN_USERS:
        return jsonify({'error': 'Admin only.'}), 403
    _svc_clear_cr_cache()
    return jsonify({'ok': True, 'message': 'CR overview cache cleared.'})


@dashboard_bp.route('/admin/cr_overview/cache_stats')
@login_required
def admin_cr_overview_cache_stats():
    """Admin-only: return live cache entry metadata."""
    from config import ADMIN_USERS
    if getattr(current_user, 'id', '') not in ADMIN_USERS:
        return jsonify({'error': 'Admin only.'}), 403
    return jsonify({'ok': True, 'entries': _svc_cr_cache_info()})


# ---------------------------------------------------------------------------------
# TOOL FEEDBACK / RATING  �  pdt_stats_dashboard.tool_feedback
# Access restricted to TARGET_GROUP = "qipl.target.pdt"
# ---------------------------------------------------------------------------------
_FEEDBACK_DB  = 'pdt_stats_dashboard'
_FEEDBACK_TBL = 'pdt_stats_dashboard.tool_feedback'


def _check_target_group_access():
    """Return True if current_user belongs to TARGET_GROUP (qipl.target.pdt).
    Admins always pass. Falls back to True on LDAP error so dev env works."""
    from config import TARGET_GROUP, ADMIN_USERS
    uid = getattr(current_user, 'id', '') or ''
    if uid in ADMIN_USERS:
        return True
    try:
        import app as _app
        return _app.is_user_in_group(uid, TARGET_GROUP)
    except Exception:
        return True


@dashboard_bp.route('/api/feedback/submit', methods=['POST'])
@login_required
def api_feedback_submit():
    """Submit a star rating + hours-saved feedback.
    Body: {rating: 1-5, hours_saved: float, feedback_text: str, page: str}
    Restricted to TARGET_GROUP.
    """
    if not _check_target_group_access():
        return jsonify({'ok': False, 'error': 'Access restricted to qipl.target.pdt group.'}), 403

    data         = request.get_json(force=True) or {}
    rating       = int(data.get('rating')       or 0)
    hours_saved  = float(data.get('hours_saved') or 0)
    feedback_txt = str(data.get('feedback_text') or '')[:500].strip()
    page         = str(data.get('page')          or 'cr_overview')[:64].strip()
    username     = str(getattr(current_user, 'id', '') or 'unknown')[:64]

    if not (1 <= rating <= 5):
        return jsonify({'ok': False, 'error': 'Rating must be 1-5.'}), 400
    if hours_saved < 0 or hours_saved > 100:
        return jsonify({'ok': False, 'error': 'hours_saved out of range.'}), 400

    conn = cursor = None
    try:
        conn   = get_mysql_connection_db(_FEEDBACK_DB)
        cursor = conn.cursor()
        cursor.execute(
            f'INSERT INTO {_FEEDBACK_TBL} '
            '(username, rating, hours_saved, feedback_text, page) '
            'VALUES (%s, %s, %s, %s, %s)',
            (username, rating, hours_saved, feedback_txt or None, page)
        )
        conn.commit()
        return jsonify({'ok': True, 'message': 'Thank you for your feedback!'})
    except Exception as e:
        logger.info(f'[FEEDBACK] {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn:   conn.close()


@dashboard_bp.route('/api/feedback/stats')
@login_required
def api_feedback_stats():
    """Admin-only: return all feedback records + aggregates."""
    from config import ADMIN_USERS
    uid = getattr(current_user, 'id', '') or ''
    if uid not in ADMIN_USERS:
        return jsonify({'error': 'Admin only.'}), 403

    conn = cursor = None
    try:
        conn   = get_mysql_connection_db(_FEEDBACK_DB)
        cursor = conn.cursor(dictionary=True)

        cursor.execute(f"""
            SELECT
                COUNT(*)                        AS total_responses,
                ROUND(AVG(rating), 2)           AS avg_rating,
                SUM(hours_saved)                AS total_hours_saved,
                ROUND(AVG(hours_saved), 1)      AS avg_hours_saved,
                SUM(CASE WHEN rating=5 THEN 1 ELSE 0 END) AS five_star,
                SUM(CASE WHEN rating=4 THEN 1 ELSE 0 END) AS four_star,
                SUM(CASE WHEN rating=3 THEN 1 ELSE 0 END) AS three_star,
                SUM(CASE WHEN rating=2 THEN 1 ELSE 0 END) AS two_star,
                SUM(CASE WHEN rating=1 THEN 1 ELSE 0 END) AS one_star
            FROM {_FEEDBACK_TBL}
        """)
        agg = cursor.fetchone() or {}

        cursor.execute(f"""
            SELECT username,
                   COUNT(*)               AS submissions,
                   ROUND(AVG(rating),1)   AS avg_rating,
                   SUM(hours_saved)       AS total_hours_saved,
                   MAX(submitted_at)      AS last_submitted
            FROM {_FEEDBACK_TBL}
            GROUP BY username
            ORDER BY total_hours_saved DESC
        """)
        by_user = cursor.fetchall() or []

        cursor.execute(f"""
            SELECT id, username, rating, hours_saved, feedback_text, page, submitted_at
            FROM {_FEEDBACK_TBL}
            ORDER BY submitted_at DESC
            LIMIT 50
        """)
        recent = cursor.fetchall() or []

        for r in recent + by_user:
            for k in ('submitted_at', 'last_submitted'):
                v = r.get(k)
                if v and not isinstance(v, str):
                    r[k] = str(v)
        for k, v in agg.items():
            if hasattr(v, 'isoformat'):
                agg[k] = str(v)

        return jsonify({'ok': True, 'aggregates': agg, 'by_user': by_user, 'recent': recent})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn:   conn.close()


@dashboard_bp.route('/api/feedback/check_access')
@login_required
def api_feedback_check_access():
    """Returns whether current user can submit feedback."""
    allowed = _check_target_group_access()
    return jsonify({'allowed': allowed, 'username': getattr(current_user, 'id', '')})


# ---------------------------------------------------------------------
# Helper: build minimal sidebar context for standalone pages
# ---------------------------------------------------------------------
def _get_pv_hidden_tabs(target_name):
    """Return list of hidden tab keys for a target from the page_visibility config."""
    import json as _json, os as _os
    pv_path = _os.path.join(
        _os.environ.get('PDTBUDDY_DATA_ROOT', r'\\sphere\pdtqipl_internal\PDTBuddy'),
        'config', 'page_visibility.json'
    )
    try:
        if _os.path.exists(pv_path):
            with open(pv_path, 'r', encoding='utf-8') as fh:
                all_data = _json.load(fh)
            settings = all_data.get(target_name) or {}
            return [k for k, v in settings.items() if v is False]
    except Exception:
        pass
    return []


def _build_sidebar_context(target_name, active_section="mtbf-table"):
    try:
        target_display_name = get_display_name_for_target(target_name).upper()
    except Exception:
        target_display_name = target_name.upper()
    try:
        schema_name = get_schema_for_target(target_name) or "pdt_stats_mobile"
        is_compute_bu = (schema_name == "pdt_stats_compute")
    except Exception:
        schema_name = "pdt_stats_mobile"
        is_compute_bu = False
    try:
        dash_meta = get_dashboard_meta_for_target(target_name) or {}
        raw_dt = dash_meta.get("dashboard_latest_update") or dash_meta.get("unique_cr_last_update")

        if isinstance(raw_dt, datetime.datetime):
            target_update = raw_dt.strftime("%Y-%m-%d %H:%M:%S")
        elif raw_dt:
            target_update = str(raw_dt)
        else:
            target_update = "N/A"
        dashboard_latest_update = dash_meta.get("dashboard_latest_update")
        unique_cr_last_update = dash_meta.get("unique_cr_last_update")
        milestones = {

            "ES": dash_meta.get("ES") or "TBD",
            "FC": dash_meta.get("FC") or "TBD",
            "CS": dash_meta.get("CS") or "TBD",
        }
    except Exception:
        target_update = "N/A"
        dashboard_latest_update = None
        unique_cr_last_update = None
        milestones = {"ES": "TBD", "FC": "TBD", "CS": "TBD"}

    glance = {
        "mapped_jiras": 0, "open_jiras": 0, "closed_jiras": 0,
        "total_jiras": 0, "total_crs": 0,
    }
    active_bu_key = (get_bu_for_target(target_name) or '').upper()
    milestone_phase = build_milestone_phase_context(target_name)
    return {
        "target_name": target_name,
        "target_display_name": target_display_name,
        "active_section": active_section,
        **_build_bu_shell_context(active_bu_key),
        "toggle_mode": "CRM",
        "pdt_type": "SWPDT",
        "schema_name": schema_name,
        "is_compute_bu": is_compute_bu,
        "compute_bu": False,
        "hwpdt_available": False,
        "glance": glance,
        "milestones": milestones,
                "milestone_phase": milestone_phase,
        "target_update": target_update,
        "dashboard_latest_update": dashboard_latest_update,
        "unique_cr_last_update": unique_cr_last_update,

        "cr_age_chart_built": {"categories": [], "cr_count": [], "avg_cr_age": []},
                "cr_age_chart_undisposed": {"categories": [], "cr_count": [], "avg_cr_age": []},
        "cr_age_buckets": {"5_20": 0, "20_40": 0, "over_40": 0},
        "cr_age_list_5_20": [], "cr_age_list_20_40": [], "cr_age_list_over_40": [],
        "undisp_status_counts": {"open": 0, "analysis": 0, "other": 0},
        "cr_rows": [],
        "mapped_jiras_url": "#", "open_jiras_url": "#",
        "closed_jiras_url": "#", "total_crs_url": "#",
        "pv_hidden_tabs": _get_pv_hidden_tabs(target_name),
    }


# ---------------------------------------------------------------------
# Helper: clean_data_for_session
# ---------------------------------------------------------------------
def clean_data_for_session(rows):
    from datetime import datetime as _dts, date as _ds
    cleaned = []
    for row in rows:
        new_row = {}
        for k, v in row.items():
            if isinstance(v, (_dts, _ds)):
                new_row[k] = v.isoformat()
            else:
                new_row[k] = v
        cleaned.append(new_row)
    return cleaned


# ---------------------------------------------------------------------
# Helper: redirect_to_full_table
# ---------------------------------------------------------------------
def redirect_to_full_table(query, target_name, table_name="Data Table"):
    import uuid as _uuid
    from mysql.connector import Error as _Error
    conn = get_mysql_connection_db()
    if not conn:
        flash("Database connection error.", "danger")
        return redirect(url_for("bu_selection"))
    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(query)
        res = cursor.fetchall() or []
        if not res:
            flash(f"No records found in this table for {target_name}.", "info")
            return redirect(url_for("dashboard_bp.dashboard", target_name=target_name, section="dashboard"))
        clean_res = clean_data_for_session(res)
        res_id = str(_uuid.uuid4())
        session[f"query_results_{res_id}"] = clean_res
        session[f"table_name_{res_id}"] = table_name
        from app import _sign_result_id
        token = _sign_result_id(res_id, current_user.get_id())
        return redirect(url_for("view_query_table", token=token))
    except _Error as e:
        import traceback; logger.debug(traceback.format_exc())
        flash(f"Error fetching table data: {e}", "danger")
        return redirect(url_for("dashboard_bp.dashboard", target_name=target_name, section="dashboard"))
    finally:
        cursor.close()
        conn.close()


# ---------------------------------------------------------------------
# Helper: group_hwpdt_by_meta
# ---------------------------------------------------------------------
def group_hwpdt_by_meta(flat_rows):
    from collections import defaultdict as _dd
    grouped = _dd(list)
    for r in flat_rows:
        mid = r.get("meta_id")
        if not mid:
            continue
        grouped[mid].append(r)
    result = []
    for idx, (meta_id, builds) in enumerate(grouped.items(), start=1):
        total_crashes = sum(int(b.get("hwpdt_crashes") or b.get("crashes") or 0) for b in builds)
        mode = builds[0].get("mode") or builds[0].get("build_mode")
        chipmd = builds[0].get("chipmd_ticket") or builds[0].get("chipmd") or ""
        result.append({"s_no": idx, "meta_id": meta_id, "builds": builds,
                        "hwpdt_crashes": total_crashes, "mode": mode, "chipmd_ticket": chipmd})
    result.sort(key=lambda r: r["meta_id"])
    return result


@dashboard_bp.route("/mtbf_table_save/<string:target_name>", methods=["POST"])
@login_required
def mtbf_table_save_view(target_name):

    conn = get_mysql_connection_db()
    cursor = conn.cursor()
    try:
        payload_json = request.form.get("payload_json") or ""
        data = json.loads(payload_json) if payload_json else {}

        rows = data.get("rows", [])
        pdt_type = data.get("pdt_type")
        toggle_mode = data.get("toggle_mode")

        schema_name = get_schema_for_target(target_name) or "pdt_stats_mobile"
        is_compute_bu = (schema_name == "pdt_stats_compute")

        table_name = ensure_meta_builds_table(cursor, schema_name, target_name)
        meta_builds_table = f"`{schema_name}`.`{table_name}`"

        AGG_BUILD_ID = "__META__"

        upsert_sql = f"""
            INSERT INTO {meta_builds_table}
                (meta_id,
                 build_id,
                 pdt_type,
                 mode,
                 hours,
                 swpdt_crashes,
                 hwpdt_crashes,
                 mtbf,
                 product_mtbf,
                 qc_mtbf,
                 is_selected,
                 build_source,
                 is_manual_entry,
                 is_active)
            VALUES
                (%s, %s, %s, %s, %s,
                 %s, %s, %s, %s, %s,
                 %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                hours           = VALUES(hours),
                swpdt_crashes   = VALUES(swpdt_crashes),
                hwpdt_crashes   = VALUES(hwpdt_crashes),
                mtbf            = VALUES(mtbf),
                product_mtbf    = VALUES(product_mtbf),
                qc_mtbf         = VALUES(qc_mtbf),
                is_selected     = VALUES(is_selected),
                is_manual_entry = VALUES(is_manual_entry),
                build_source    = VALUES(build_source),
                is_active       = VALUES(is_active),
                mode            = VALUES(mode),
                pdt_type        = VALUES(pdt_type)
        """

        for r in rows:
            meta_id = r.get("meta_id")
            mode = r.get("mode") or "CRM"
            pdt = r.get("pdt_type") or pdt_type
            crashes = int(r.get("crashes") or 0)
            hours_val = float(r.get("hours") or 0.0)
            reset_to_auto_meta = bool(r.get("reset_to_auto_meta"))
            excluded_tickets_raw = r.get("excluded_tickets") or ""
            excluded_tickets = [
                t.strip() for t in str(excluded_tickets_raw).split(",") if t.strip()
            ]
            meta_notes = None
            if excluded_tickets:
                meta_notes = json.dumps({"excluded_tickets": excluded_tickets})

            if not meta_id:
                continue

            if reset_to_auto_meta:
                # Clear META-level override so we fall back to auto raw_crashes/mtbf
                cursor.execute(
                    f"""
                    UPDATE {meta_builds_table}
                    SET
                        swpdt_crashes = NULL,
                        hwpdt_crashes = NULL,
                        mtbf          = NULL,
                        product_mtbf  = NULL,
                        qc_mtbf       = NULL,
                        meta_notes    = NULL
                    WHERE meta_id = %s
                      AND build_id = %s
                      AND pdt_type = %s
                    """,
                    (meta_id, AGG_BUILD_ID, pdt),
                )
            else:
                if is_compute_bu:
                    qc_mtbf = round(float(r.get("qc_mtbf") or 0.0), 2)
                    prod_mtbf = round(float(r.get("product_mtbf") or 0.0), 2)
                    mtbf_val = round(qc_mtbf + prod_mtbf, 2)
                else:
                    mtbf_val = round(float(r.get("mtbf") or 0.0), 2)
                    prod_mtbf = 0.0
                    qc_mtbf = 0.0

                # store META-level aggregate row
                cursor.execute(
                    upsert_sql,
                    (
                        meta_id,
                        AGG_BUILD_ID,
                        pdt,
                        mode,
                        hours_val,
                        crashes,
                        0,          # hwpdt_crashes
                        mtbf_val,
                        prod_mtbf,
                        qc_mtbf,
                        1,          # is_selected
                        "MANUAL",
                        1,          # is_manual_entry
                        1,          # is_active
                    ),
                )

                if meta_notes:
                    cursor.execute(
                        f"""
                        UPDATE {meta_builds_table}
                        SET meta_notes = %s
                        WHERE meta_id = %s
                          AND build_id = %s
                          AND pdt_type = %s
                        """,
                        (meta_notes, meta_id, AGG_BUILD_ID, pdt),
                    )

            # store each build row
            for b in r.get("builds", []):
                build_id = b.get("build_id")
                if not build_id:
                    continue

                b_hours = float(b.get("hours") or 0.0)
                b_swpdt = int(b.get("swpdt_crashes") or 0)
                b_hwpdt = int(b.get("hwpdt_crashes") or 0)
                b_source = b.get("build_source") or "MANUAL"
                b_is_selected = 1 if b.get("is_selected") else 0
                b_is_active = 1 if b_is_selected else 0

                if is_compute_bu:
                    b_qc_mtbf = float(b.get("qc_mtbf") or 0.0)
                    b_prod_mtbf = float(b.get("product_mtbf") or 0.0)
                    b_mtbf = round(b_qc_mtbf + b_prod_mtbf, 2)
                else:
                    b_mtbf = float(b.get("mtbf") or 0.0)
                    b_prod_mtbf = 0.0
                    b_qc_mtbf = 0.0

                cursor.execute(
                    upsert_sql,
                    (
                        meta_id,
                        build_id,
                        pdt,
                        mode,
                        b_hours,
                        b_swpdt,
                        b_hwpdt,
                        b_mtbf,
                        b_prod_mtbf,
                        b_qc_mtbf,
                        b_is_selected,
                        b_source,
                        1,
                        b_is_active,
                    ),
                )

        conn.commit()

        return redirect(
            url_for(
                "dashboard_bp.dashboard",
                target_name=target_name,
                section="mtbf-table",
                toggle_mode=toggle_mode,
                pdt_type=pdt_type,
                compute_bu="1" if is_compute_bu else "0",
            )
        )

    except Exception as e:
        conn.rollback()
        return f"Error saving MTBF table: {str(e)}", 500
    finally:
        cursor.close()
        conn.close()

@dashboard_bp.route("/api/mtbf_jiras/<string:target_name>/<string:meta_id>")
@login_required
def api_mtbf_jiras(target_name, meta_id):

    conn = None
    cursor = None
    try:
        conn = get_mysql_connection_db()
        cursor = conn.cursor(dictionary=True)

        schema_name = get_schema_for_target(target_name) or "pdt_stats_mobile"
        j_table = fq_table_for_target(target_name, "jiras")
        # Try fq_table_for_target first; fall back to direct DB search
        o_table = None
        try:
            o_table = fq_table_for_target(target_name, "openjiras")
        except Exception as _fq_err:
            pass

        if not o_table:
            tbl_pattern = target_name.lower().replace('-', '_').replace(' ', '_') + '_openjiras'
            cursor.execute(
                "SELECT TABLE_SCHEMA, TABLE_NAME FROM information_schema.TABLES "
                "WHERE TABLE_NAME = %s LIMIT 1",
                (tbl_pattern,)
            )
            _row = cursor.fetchone()
            if _row:
                _s = _row.get('TABLE_SCHEMA') or _row.get('table_schema', '')
                _n = _row.get('TABLE_NAME')   or _row.get('table_name', '')
                o_table = f"`{_s}`.`{_n}`"
            else:
                return jsonify({"success": True, "rows": [], "area_summary": [],
                                "notice": "Open JIRAs table not available for this target."})

        metabuild_like = "%" + meta_id + "%"
        AGG_BUILD_ID = "__META__"
        meta_builds_table_name = ensure_meta_builds_table(cursor, schema_name, target_name)
        meta_builds_table = f"`{schema_name}`.`{meta_builds_table_name}`"

        # ─ load excluded tickets from meta_notes ─
        cursor.execute(
            f"""
            SELECT meta_notes
            FROM {meta_builds_table}
            WHERE meta_id = %s AND build_id = %s AND pdt_type = %s AND is_active = 1
            """,
            (meta_id, AGG_BUILD_ID, "SWPDT"),
        )
        meta_row = cursor.fetchone() or {}
        excluded_tickets = set()
        notes = meta_row.get("meta_notes")
        if notes:
            try:
                notes_obj = json.loads(notes)
                for t in (notes_obj.get("excluded_tickets") or []):
                    if t:
                        excluded_tickets.add(str(t).strip())
            except Exception:
                pass

        # ─ builds: from ?builds= param (real-time DOM) OR saved DB selection ─
        builds_param = request.args.get('builds', '').strip()
        if builds_param:
            selected_builds = [b.strip() for b in builds_param.split(',') if b.strip()]
        else:
            cursor.execute(
                f"""
                SELECT build_id FROM {meta_builds_table}
                WHERE meta_id = %s AND build_id <> %s
                  AND pdt_type = %s AND is_selected = 1 AND is_active = 1
                """,
                (meta_id, AGG_BUILD_ID, "SWPDT"),
            )
            selected_builds = [r["build_id"] for r in (cursor.fetchall() or []) if r.get("build_id")]

        if selected_builds:
            ph = ",".join(["%s"] * len(selected_builds))
            jira_where = f"j.metabuild IN ({ph})"
            open_where = f"o.metabuild IN ({ph})"
            params = tuple(selected_builds) * 2
        else:
            jira_where = "j.metabuild LIKE %s"
            open_where = "o.metabuild LIKE %s"
            params = (metabuild_like, metabuild_like)

                # Guard: openjiras table may not exist for all targets
        def _o_tbl_ok(fq_name):
            n = fq_name.replace("`", "")
            try:
                s, t = n.split(".", 1)
            except ValueError:
                return True
            cursor.execute("SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1", (s, t))
            return cursor.fetchone() is not None

        if _o_tbl_ok(o_table):
            sql = (
                f"SELECT j.stability_ticket, j.jira_date, j.jira_title, j.serial_no, j.metabuild"
                f" FROM {j_table} j WHERE {jira_where}"
                f" UNION"
                f" SELECT o.stability_ticket, o.jira_date, o.jira_title, o.serial_no, o.metabuild"
                f" FROM {o_table} o WHERE {open_where}"
                f" ORDER BY stability_ticket"
            )
            cursor.execute(sql, params)
        else:
            logger.info(f"[MTBF JIRAs] openjiras table missing for target={target_name}, using jiras only")
            if selected_builds:
                ph2 = ",".join(["%s"] * len(selected_builds))
                cursor.execute(
                    f"SELECT j.stability_ticket, j.jira_date, j.jira_title, j.serial_no, j.metabuild"
                    f" FROM {j_table} j WHERE j.metabuild IN ({ph2}) ORDER BY stability_ticket",
                    tuple(selected_builds),
                )
            else:
                cursor.execute(
                    f"SELECT j.stability_ticket, j.jira_date, j.jira_title, j.serial_no, j.metabuild"
                    f" FROM {j_table} j WHERE j.metabuild LIKE %s ORDER BY stability_ticket",
                    (metabuild_like,),
                )
        all_rows = cursor.fetchall() or []

        # ─ exclude saved excluded tickets ─
        rows = []
        for r in all_rows:
            ticket = (r.get("stability_ticket") or "").strip()
            if ticket and ticket not in excluded_tickets:
                r["crash_count"] = 1
                r["build_id"] = r.get("metabuild")
                rows.append(r)

        logger.info(f"[MTBF JIRAs] total={len(all_rows)} excluded={len(excluded_tickets)} shown={len(rows)}")
        return jsonify({"meta_id": meta_id, "jiras": rows})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@dashboard_bp.route("/mtbf_meta_jiras/<string:target_name>/<string:meta_id>")
@login_required
def mtbf_meta_jiras_view(target_name, meta_id):

    conn = None
    cursor = None
    try:
        conn = get_mysql_connection_db()
        cursor = conn.cursor(dictionary=True)

        schema_name = get_schema_for_target(target_name) or "pdt_stats_mobile"
        meta_builds_table_name = ensure_meta_builds_table(cursor, schema_name, target_name)
        meta_builds_table = f"`{schema_name}`.`{meta_builds_table_name}`"
        AGG_BUILD_ID = "__META__"

        # Load excluded tickets from meta_notes for this META's aggregate row
        cursor.execute(
            f"""
            SELECT meta_notes
            FROM {meta_builds_table}
            WHERE meta_id = %s
              AND build_id = %s
              AND pdt_type = %s
              AND is_active = 1
            """,
            (meta_id, AGG_BUILD_ID, "SWPDT"),
        )
        row = cursor.fetchone() or {}
        excluded_tickets = set()
        notes = row.get("meta_notes")
        if notes:
            try:
                notes_obj = json.loads(notes)
                for t in notes_obj.get("excluded_tickets") or []:
                    if t:
                        excluded_tickets.add(str(t).strip())
            except Exception:
                pass

        j_table = fq_table_for_target(target_name, "jiras")
        # Try fq_table_for_target first; fall back to direct DB search
        o_table = None
        try:
            o_table = fq_table_for_target(target_name, "openjiras")
        except Exception as _fq_err:
            pass

        if not o_table:
            tbl_pattern = target_name.lower().replace('-', '_').replace(' ', '_') + '_openjiras'
            cursor.execute(
                "SELECT TABLE_SCHEMA, TABLE_NAME FROM information_schema.TABLES "
                "WHERE TABLE_NAME = %s LIMIT 1",
                (tbl_pattern,)
            )
            _row = cursor.fetchone()
            if _row:
                _s = _row.get('TABLE_SCHEMA') or _row.get('table_schema', '')
                _n = _row.get('TABLE_NAME')   or _row.get('table_name', '')
                o_table = f"`{_s}`.`{_n}`"
            else:
                return jsonify({"success": True, "rows": [], "area_summary": [],
                                "notice": "Open JIRAs table not available for this target."})
        metabuild_like = "%" + meta_id + "%"

        # ─ Load SELECTED build IDs for this meta from meta_builds_table ─
        cursor.execute(
            f"""
            SELECT build_id
            FROM {meta_builds_table}
            WHERE meta_id = %s
              AND build_id <> %s
              AND pdt_type = %s
              AND is_selected = 1
              AND is_active   = 1
            """,
            (meta_id, AGG_BUILD_ID, "SWPDT"),
        )
        selected_build_rows = cursor.fetchall() or []
        selected_builds = [r["build_id"] for r in selected_build_rows if r.get("build_id")]

        # If we have explicit selected builds, filter JIRAs to those builds only.
        # Otherwise fall back to the full meta LIKE match (no saved selection yet).
        if selected_builds:
            build_placeholders = ",".join(["%s"] * len(selected_builds))
            # Each build_id IS the metabuild value stored in the jiras table
            # (exact match per build, not LIKE on meta prefix)
            jira_where  = f"j.metabuild IN ({build_placeholders})"
            open_where  = f"o.metabuild IN ({build_placeholders})"
            jira_params = tuple(selected_builds)
            open_params = tuple(selected_builds)
        else:
            jira_where  = "j.metabuild LIKE %s"
            open_where  = "o.metabuild LIKE %s"
            jira_params = (metabuild_like,)
            open_params = (metabuild_like,)

                # Guard: openjiras table may not exist for all targets
        def _o_tbl_ok_meta(fq_name):
            n = fq_name.replace("`", "")
            try:
                s, t = n.split(".", 1)
            except ValueError:
                return True
            cursor.execute("SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1", (s, t))
            return cursor.fetchone() is not None

        if _o_tbl_ok_meta(o_table):
            sql = f"""
                SELECT
                    j.stability_ticket,
                    j.jira_date,
                    j.jira_title,
                    j.serial_no,
                    j.metabuild
                FROM {j_table} j
                WHERE {jira_where}

                UNION

                SELECT
                    o.stability_ticket,
                    o.jira_date,
                    o.jira_title,
                    o.serial_no,
                    o.metabuild
                FROM {o_table} o
                WHERE {open_where}

                ORDER BY stability_ticket
            """
            cursor.execute(sql, jira_params + open_params)
        else:
            logger.info(f"[MTBF META JIRAS] openjiras table missing for target={target_name}, using jiras only")
            sql = f"""
                SELECT
                    j.stability_ticket,
                    j.jira_date,
                    j.jira_title,
                    j.serial_no,
                    j.metabuild
                FROM {j_table} j
                WHERE {jira_where}
                ORDER BY stability_ticket
            """
            cursor.execute(sql, jira_params)
        rows = cursor.fetchall() or []

        # ─ Split into active (non-excluded) and excluded ─
        # The JIRA table shows only active rows.
        # CR occurrence count + JQL link use only active rows.
        # Excluded tickets are stored in meta_notes and should not
        # appear in the count or the JQL hyperlink.
        active_rows   = []
        excluded_rows = []
        for r in rows:
            ticket = (r.get("stability_ticket") or "").strip()
            if ticket in excluded_tickets:
                r["is_excluded"] = True
                excluded_rows.append(r)
            else:
                r["is_excluded"] = False
                active_rows.append(r)

        # jiras passed to template = active only (excluded ones not shown)
        rows = active_rows

        # ──────────────────────────────────────────────────────────────────
        # CR FETCH
        # Source of truth for ticket count = active_rows (JIRA table)
        # Step 1: build ticket -> CR map from jiras table (has cr column)
        # Step 2: group active_rows tickets by CR  -> cr_to_tickets
        # Step 3: fetch CR details from unique_crs
        # Step 4: merge jira_count + jira_display onto each CR row
        # ──────────────────────────────────────────────────────────────────
        cr_rows = []
        try:
            u_table = fq_table_for_target(target_name, "unique_crs")

            # check which CR column exists in jiras table
            cursor.execute(f"SHOW COLUMNS FROM {j_table}")
            j_cols = {c['Field'] for c in (cursor.fetchall() or [])}
            cr_col = 'cr' if 'cr' in j_cols else ('mapped_crs' if 'mapped_crs' in j_cols else None)

            # STEP 1: build ticket -> CR map from jiras table
            # Query only for the active tickets (from active_rows)
            # Use DISTINCT to avoid duplicate rows per ticket
            ticket_to_cr = {}   # { 'QSTABILITY-123': 'CR4012140', ... }
            if cr_col:
                active_tickets = list({
                    (r.get('stability_ticket') or '').strip()
                    for r in active_rows
                    if r.get('stability_ticket')
                })
                if active_tickets:
                    at_ph = ",".join(["%s"] * len(active_tickets))
                    cursor.execute(
                        f"SELECT DISTINCT `{cr_col}` AS cr_num, stability_ticket "
                        f"FROM {j_table} "
                        f"WHERE stability_ticket IN ({at_ph}) "
                        f"  AND `{cr_col}` IS NOT NULL AND `{cr_col}` <> ''",
                        tuple(active_tickets),
                    )
                    for jrow in (cursor.fetchall() or []):
                        t = (jrow.get('stability_ticket') or '').strip()
                        c = (jrow.get('cr_num') or '').strip()
                        # one ticket -> one CR (first wins)
                        if t and c and t not in ticket_to_cr:
                            ticket_to_cr[t] = c

            # STEP 2: group active_rows tickets by CR
            # source of truth = active_rows only, no extras
            cr_to_tickets = {}
            for r in active_rows:
                ticket = (r.get('stability_ticket') or '').strip()
                if not ticket:
                    continue
                cr_num = ticket_to_cr.get(ticket)
                if cr_num:
                    if ticket not in cr_to_tickets.get(cr_num, []):
                        cr_to_tickets.setdefault(cr_num, []).append(ticket)

            #logger.info(f"[MTBF META JIRAS] cr_to_tickets: { {k: len(v) for k, v in cr_to_tickets.items()} }")

                            # STEP 2b: resolve raw CRs to their canonical mapped_cr in unique_crs
            # e.g. raw CR4154709 (Dup) -> mapped_cr CR4171973
            if cr_to_tickets:
                raw_cr_list = list(cr_to_tickets.keys())
                ph2 = ','.join(['%s'] * len(raw_cr_list))
                cursor.execute(
                    f"SELECT cr, mapped_cr FROM {u_table} "
                    f"WHERE cr IN ({ph2}) OR mapped_cr IN ({ph2})",
                    tuple(raw_cr_list) * 2,
                )
                cr_to_mapped = {}
                for row2 in (cursor.fetchall() or []):
                    raw = (row2.get('cr') or '').strip()
                    mapped = (row2.get('mapped_cr') or '').strip()
                    if raw and mapped:
                        cr_to_mapped[raw] = mapped
                    if mapped:
                        cr_to_mapped[mapped] = mapped  # identity mapping

                # re-key cr_to_tickets using canonical mapped_cr
                resolved = {}
                for raw_cr, tickets in cr_to_tickets.items():
                    canonical = cr_to_mapped.get(raw_cr, raw_cr)
                    existing = resolved.get(canonical, [])
                    for t in tickets:
                        if t not in existing:
                            existing.append(t)
                    resolved[canonical] = existing
                cr_to_tickets = resolved

            #logger.info(f"[MTBF META JIRAS] cr_to_tickets (resolved): { {k: len(v) for k, v in cr_to_tickets.items()} }")

                # STEP 3: fetch CR details from unique_crs
            if cr_to_tickets:
                cursor.execute(f"SHOW COLUMNS FROM {u_table}")
                u_cols = {c['Field'] for c in (cursor.fetchall() or [])}

                def _col(name):
                    return f"`{name}`" if name in u_cols else f"NULL AS `{name}`"

                # Use image column as cr_si if cr_si doesn't exist
                cr_si_expr = ("`cr_si`" if "cr_si" in u_cols
                              else ("`image` AS `cr_si`" if "image" in u_cols
                                    else "NULL AS `cr_si`"))

                select_sql = ", ".join([
                    "`mapped_cr`",
                    _col("cr_occurrence"),
                    _col("cr_title"),
                    _col("cr_area"),
                    _col("cr_subsystem"),
                    _col("cr_functionality"),
                    _col("cr_status"),
                    _col("cr_age"),
                    cr_si_expr,
                    _col("built_date"),
                    _col("jira_date"),
                    _col("cr_category"),
                ])

                cr_list = list(cr_to_tickets.keys())
                ph = ",".join(["%s"] * len(cr_list))
                # Fetch ALL rows for these mapped_crs (including Dup rows)
                # We deduplicate ourselves below � do NOT filter cr_occurrence here
                cursor.execute(
                    f"SELECT {select_sql} "
                    f"FROM {u_table} "
                    f"WHERE mapped_cr IN ({ph})",
                    tuple(cr_list),
                )
                raw_cr_rows = cursor.fetchall() or []

                # deduplicate: one row per mapped_cr
                # Prefer non-Dup rows; among those keep highest numeric cr_occurrence
                seen_cr = {}
                for cr_row in raw_cr_rows:
                    mapped = (cr_row.get('mapped_cr') or '').strip()
                    if not mapped:
                        continue
                    occ_raw = str(cr_row.get('cr_occurrence') or '').strip()
                    is_dup = occ_raw.lower() == 'dup'
                    existing = seen_cr.get(mapped)
                    if existing is None:
                        seen_cr[mapped] = cr_row
                    else:
                        ex_occ = str(existing.get('cr_occurrence') or '').strip()
                        ex_is_dup = ex_occ.lower() == 'dup'
                        # Always prefer non-Dup over Dup
                        if ex_is_dup and not is_dup:
                            seen_cr[mapped] = cr_row
                        elif not ex_is_dup and not is_dup:
                            # Both non-Dup: keep higher numeric occurrence
                            try:
                                if int(occ_raw or 0) > int(ex_occ or 0):
                                    seen_cr[mapped] = cr_row
                            except (ValueError, TypeError):
                                pass
                raw_cr_rows = list(seen_cr.values())

                # STEP 4: merge jira_count + jira_display onto each CR row
                for cr_row in raw_cr_rows:
                    mapped = (cr_row.get('mapped_cr') or '').strip()
                    tickets = cr_to_tickets.get(mapped, [])
                    cr_row['jira_count']   = len(tickets)
                    cr_row['jira_display'] = tickets[:45]

                # Only keep CRs that actually have JIRA hits
                raw_cr_rows = [r for r in raw_cr_rows if r.get('jira_count', 0) > 0]

                # sort by jira_count descending
                cr_rows = sorted(raw_cr_rows, key=lambda x: x.get('jira_count', 0), reverse=True)

            # ─ remove CR-mapped tickets from the JIRA table ─
            # tickets already shown in CR table should not appear in Open JIRAs
            cr_mapped_tickets = {
                t
                for tickets in cr_to_tickets.values()
                for t in tickets
            }
            rows = [
                r for r in rows
                if (r.get('stability_ticket') or '').strip() not in cr_mapped_tickets
                        ]

            #logger.info(f"[MTBF META JIRAS] Final CR rows: {len(cr_rows)}, JIRA table (unmapped only): {len(rows)}")

        except Exception as cr_err:
            logger.info(f"[MTBF META JIRAS] CR fetch error: {cr_err}")
            import traceback as _tb; _tb.print_exc()
            cr_rows = []

        total_jiras_count = len(active_rows)  # all active (non-excluded) JIRAs for this META

        # Build sidebar context so target_layout.html renders the nav correctly
        sidebar_ctx = _build_sidebar_context(target_name, active_section="mtbf-table")
        # Add page-specific keys that are NOT in sidebar_ctx
        sidebar_ctx["meta_id"]           = meta_id
        sidebar_ctx["jiras"]             = rows
        sidebar_ctx["cr_rows"]           = cr_rows          # overwrite empty list from sidebar_ctx
        sidebar_ctx["excluded_tickets"]  = excluded_tickets
        sidebar_ctx["total_jiras_count"] = total_jiras_count
        sidebar_ctx["page_heading"]      = f"{sidebar_ctx['target_display_name']} \u2014 CRs & JIRAs \u2014 {meta_id}"

        return render_template(
            "mtbf_meta_jiras.html",
            **sidebar_ctx,
                )
    except Exception as e:
        return f"Error loading JIRAs for {meta_id}: {e}", 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
def _load_hwpdt_job_audit_data():
    """
    Load HWPDT job audit data.

    PRIMARY source: pdt_stats_dashboard.axiom_job_summary DB table
      (populated every 5 min by the Axiom combined poller).
    FALLBACK: HWPDT_job_audit.json on the network share / local backup
      (used when DB has no HWPDT rows yet or DB is unreachable).

    The normalised output always contains:
      generated_at  - ISO timestamp string
      chip_lookup   - { chip_id: [{job_id, software_product, start_time, playlist_name}] }
      jobs          - list of job dicts
      source        - 'db' | 'json_old' | 'json_new'
    """
    import json as _json

    # -- PRIMARY: read from axiom_job_summary DB table ----------------------
    try:
        _conn = get_mysql_connection_db(bu_key=None)
        if _conn:
            _cur = _conn.cursor(dictionary=True)
            try:
                # Best effort only: increase this session's sort/work memory, not global MySQL.
                # The query below avoids MySQL ORDER BY on large JSON rows; Python sorts after fetch.
                _cur.execute("SET SESSION sort_buffer_size = 268435456")
            except Exception:
                pass
            _cur.execute("""
                SELECT job_id, build_id, software_product, chip_ids,
                       playlist_name, certicom_playlist, state, device_count,
                       submitted_at, started_at, updated_at
                FROM pdt_stats_dashboard.axiom_job_summary
                WHERE team = 'HWPDT'
                  AND state IN ('Completed', 'Aborted', 'Running', 'JobSetup')
            """)
            db_rows = _cur.fetchall() or []
            db_rows.sort(key=lambda r: str(r.get("submitted_at") or ""), reverse=True)
            _cur.close()
            _conn.close()

            if db_rows:
                chip_lookup = {}
                jobs_list   = []
                latest_ts   = ""

                for row in db_rows:
                    job_id           = str(row.get("job_id") or "").strip()
                    software_product = str(row.get("software_product") or "").strip()
                    start_time       = str(row.get("started_at") or row.get("submitted_at") or "").strip()
                    playlist_name    = str(row.get("playlist_name") or "").strip()
                    build_id         = str(row.get("build_id") or "").strip()
                    status           = str(row.get("state") or "").strip()
                    updated_at       = str(row.get("updated_at") or "").strip()

                    try:
                        chip_ids = _json.loads(row.get("chip_ids") or "[]")
                        chip_ids = [str(c).strip().upper() for c in chip_ids if str(c).strip()]
                    except Exception:
                        chip_ids = []

                    try:
                        certicom_playlist = row.get("certicom_playlist") or []
                        if isinstance(certicom_playlist, str):
                            certicom_playlist = _json.loads(certicom_playlist or "[]")
                        if not isinstance(certicom_playlist, list):
                            certicom_playlist = []
                    except Exception:
                        certicom_playlist = []

                    if updated_at and updated_at > latest_ts:
                        latest_ts = updated_at

                    jobs_list.append({
                        "job_id":           job_id,
                        "software_product": software_product,
                        "chip_ids":         chip_ids,
                                                "start_time":         start_time,
                        "playlist_name":      playlist_name,
                        "playlist":           "",
                        "certicom_playlist":  certicom_playlist,
                        "build_id":           build_id,
                        "status":             status,
                    })

                    entry = {
                        "job_id":             job_id,
                        "software_product":   software_product,
                        "start_time":         start_time,
                        "playlist_name":      playlist_name,
                        "playlist":           "",
                        "certicom_playlist":  certicom_playlist,
                    }
                    for chip_id in chip_ids:
                        chip_lookup.setdefault(chip_id, []).append(entry)

                logger.info(
                    "[HWPDT CHIP DATA] Loaded %d jobs / %d chips from axiom_job_summary DB",
                    len(jobs_list), len(chip_lookup),
                )
                return {
                    "generated_at":   latest_ts or "",
                    "chip_lookup":    chip_lookup,
                    "jobs":           jobs_list,
                    "job_count":      len(jobs_list),
                    "taxonomy":       "/PDT/QIPL/HW",
                    "retention_days": 20,
                    "total_builds":   len(jobs_list),
                    "source":         "db",
                }
    except Exception as _db_err:
        logger.warning("[HWPDT CHIP DATA] DB load failed, falling back to JSON: %s", _db_err)

    # -- FALLBACK: read from flat JSON files --------------------------------
    logger.info("[HWPDT CHIP DATA] Falling back to HWPDT_job_audit.json")
    network_path = r"\\sphere\pdtqipl_internal\PDTBuddy\HWPDT\HWPDT_job_audit.json"
    local_backup = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HWPDT_job_audit_local_backup.json")

    raw = None
    best_count = -1
    for path in [network_path, local_backup]:
        if not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                _d = _json.load(f)
            _count = len(_d.get("builds") or _d.get("chip_lookup") or {})
            if _count > best_count:
                best_count = _count
                raw = _d
        except Exception:
            continue

    if not raw:
        return {}

    # Already in old format (has chip_lookup) - return as-is
    if raw.get("chip_lookup"):
        raw["source"] = "json_old"
        return raw

    # New format: builds is a dict keyed by job_id
    builds_raw = raw.get("builds")
    if not builds_raw or not isinstance(builds_raw, dict):
        return raw

    chip_lookup_fb = {}
    jobs_list_fb   = []

    for job_id_key, job in builds_raw.items():
        if not isinstance(job, dict):
            continue

        job_id           = str(job.get("job_id") or job_id_key).strip()
        software_product = str(job.get("software_product") or "").strip()
        start_time       = str(job.get("submitted") or job.get("start_time") or "").strip()
        playlist_name    = str(job.get("playlist_name") or "").strip()
        playlist         = str(job.get("playlist") or "").strip()
        certicom_playlist = job.get("certicom_playlist") or []
        if not isinstance(certicom_playlist, list):
            certicom_playlist = []
        build_id         = str(job.get("build_id") or "").strip()
        chip_ids         = [str(c).strip().upper() for c in (job.get("chip_ids") or []) if str(c).strip()]

        jobs_list_fb.append({
            "job_id":           job_id,
            "software_product": software_product,
            "chip_ids":         chip_ids,
                        "start_time":         start_time,
            "playlist_name":      playlist_name,
            "playlist":           playlist,
            "certicom_playlist":  certicom_playlist,
            "build_id":           build_id,
            "status":             str(job.get("status") or "").strip(),
        })

        entry = {
            "job_id":             job_id,
            "software_product":   software_product,
            "start_time":         start_time,
            "playlist_name":      playlist_name,
            "playlist":           playlist,
            "certicom_playlist":  certicom_playlist,
        }
        for chip_id in chip_ids:
            chip_lookup_fb.setdefault(chip_id, []).append(entry)

    return {
        "generated_at":   raw.get("generated_at", ""),
        "chip_lookup":    chip_lookup_fb,
        "jobs":           jobs_list_fb,
        "job_count":      len(jobs_list_fb),
        "taxonomy":       raw.get("taxonomy", ""),
        "retention_days": raw.get("retention_days"),
        "total_builds":   raw.get("total_builds"),
        "source":         "json_new",
    }


def _get_projected_parts(target_name: str):
    """Read projected_parts for a target from HWPDT_projected.json (network then local backup)."""
    import json as _json
    network_path = r"\\sphere\pdtqipl_internal\PDTBuddy\HWPDT\HWPDT_projected.json"
    local_backup = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HWPDT_projected_local_backup.json")
    for path in [network_path, local_backup]:
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = _json.load(f)
                entry = data.get(target_name) or {}
                val = entry.get("projected_parts")
                if val and int(val) > 0:
                    return int(val)
            except Exception:
                pass
    return None


@dashboard_bp.route("/api/hwpdt_projected/<string:target_name>", methods=["POST"])
@login_required
def api_save_hwpdt_projected(target_name):
    """Save projected_parts for a target to HWPDT_projected.json."""
    import json as _json
    try:
        body = request.get_json(force=True) or {}
        projected = int(body.get("projected_parts") or 0)
        if projected < 1:
            return jsonify({"success": False, "message": "projected_parts must be > 0"}), 400

        network_path = r"\\sphere\pdtqipl_internal\PDTBuddy\HWPDT\HWPDT_projected.json"
        local_backup = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HWPDT_projected_local_backup.json")

        # Read existing
        proj_data = {}
        for path in [network_path, local_backup]:
            if os.path.exists(path):
                try:
                    with open(path, "r", encoding="utf-8") as f:
                        proj_data = _json.load(f)
                    break
                except Exception:
                    pass

        from datetime import datetime as _dt
        proj_data[target_name] = {
            "projected_parts": projected,
            "updated_at": _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
            "updated_by": getattr(current_user, 'username', 'unknown'),
        }

        saved_to = []
        net_dir = os.path.dirname(network_path)
        if os.path.exists(net_dir):
            try:
                with open(network_path, "w", encoding="utf-8") as f:
                    _json.dump(proj_data, f, indent=2)
                saved_to.append("network")
            except Exception as ex:
                logger.warning(f"[HWPDT PROJECTED] Network write failed: {ex}")
        try:
            with open(local_backup, "w", encoding="utf-8") as f:
                _json.dump(proj_data, f, indent=2)
            saved_to.append("local")
        except Exception as ex:
            logger.warning(f"[HWPDT PROJECTED] Local backup write failed: {ex}")

        if not saved_to:
            return jsonify({"success": False, "message": "Could not write to any path"}), 500

        logger.info(f"[HWPDT PROJECTED] {target_name} = {projected} saved to {saved_to}")
        return jsonify({"success": True, "projected_parts": projected, "saved_to": saved_to})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


def _hwpdt_sp_family_prefix(sp_name: str, target_name: str = "") -> str:
    """
    Return the software-product family prefix used for HWPDT chip rollups.

    Examples:
      Balsam.LA.1.0    -> BALSAM.LA
      Balsam.LA.1.0.r1 -> BALSAM.LA
      SW6100.LAW.1.0   -> SW6100.LAW
    """
    import re as _re

    sp_upper = str(sp_name or "").strip().upper()
    match = _re.match(r"^([A-Z0-9_-]+\.LA[A-Z]*)\b", sp_upper)
    if match:
        return match.group(1)

    target_seg0 = str(target_name or "").strip().upper().split("_")[0].split(".")[0]
    if target_seg0:
        return f"{target_seg0}.LA"
    return ""


@dashboard_bp.route("/api/hwpdt_chip_parts/<string:target_name>", methods=["GET"])
@login_required
def api_hwpdt_chip_parts(target_name):
    """Return tested HWPDT parts for a target.

    Matching rolls up all softwareProduct variants under the same family.
    Example: target Balsam / SP Balsam.LA.1.0 includes Balsam.LA.1.0,
    Balsam.LA.1.0.r1, Balsam.LA.2.0, etc.
    """


    # 1. Read sp_name from dashboard_status
    sp_name = ""
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if conn:
            cur = conn.cursor(dictionary=True)
            cur.execute(
                """
                SELECT sp_name
                FROM pdt_stats_dashboard.dashboard_status
                WHERE target_name = %s AND is_active = 1
                ORDER BY id DESC LIMIT 1
                """,
                (target_name,),
            )
            row = cur.fetchone() or {}
            sp_name = (row.get("sp_name") or "").strip()
            cur.close()
            conn.close()
    except Exception as ex:
        logger.warning(f"[HWPDT CHIP PARTS] DB error: {ex}")

    if not sp_name:
        return jsonify({"success": False, "message": "sp_name not configured for this target."}), 200

    # 2. Load audit JSON � single source of truth (certicom file no longer needed)
    audit_data  = _load_hwpdt_job_audit_data() or {}
    chip_lookup = audit_data.get("chip_lookup") or {}
    chip_data   = audit_data   # for generated_at reference below

    if not chip_lookup:
        return jsonify({"success": False, "message": "HWPDT chip data not available yet."}), 200

    # Derive softwareProduct -> [chip_ids] map directly from audit chip_lookup
    _sp_chips: dict = {}
    for _cid, _entries in chip_lookup.items():
        for _je in (_entries or []):
            _sp = str(_je.get("software_product") or "").strip()
            if _sp:
                _sp_chips.setdefault(_sp, set()).add(_cid.strip().upper())
    chip_map = {sp: sorted(chips) for sp, chips in _sp_chips.items()}

    # 3. Match by software-product family first.

    # If Balsam.LA.* exists in the certicom JSON, an HWPDT Balsam target should
    # count all Balsam.LA variants, not only the exact configured SP revision.
    sp_upper      = sp_name.upper()
    target_upper  = target_name.upper()
    family_prefix = _hwpdt_sp_family_prefix(sp_name, target_name)
    target_seg0   = target_upper.split("_")[0].split(".")[0]

    matched_products = []
    if family_prefix:
        matched_products = [
            sw for sw in chip_map
            if sw.upper() == family_prefix or sw.upper().startswith(family_prefix + ".")
        ]

    # Fallbacks for non-LA naming or unusual SP values.
    if not matched_products:
        matched_products = [
            sw for sw in chip_map
            if sp_upper in sw.upper() or sw.upper().startswith(sp_upper)
        ]
    if not matched_products and target_seg0:
        matched_products = [
            sw for sw in chip_map
            if sw.upper().split(".")[0].split("_")[0] == target_seg0
        ]

    matched_products = sorted(set(matched_products))

    if not matched_products:
        return jsonify({
            "success": True,
            "sp_name": sp_name,
            "family_prefix": family_prefix,
            "matched_products": [],
            "tested_parts": 0,
            "projected_parts": _get_projected_parts(target_name),
            "chip_ids": [],
            "generated_at": chip_data.get("generated_at"),
            "message": f"No softwareProduct matched for sp_name='{sp_name}'",
        }), 200

        # -- Build chip list from audit chip_lookup (already loaded above) ----------
    # chip_lookup: { chip_id: [{job_id, software_product, ...}] }
    all_chips = set()
    for chip_id, job_entries in chip_lookup.items():
        for je in (job_entries or []):
            if str(je.get("software_product") or "").strip() in matched_products:
                all_chips.add(chip_id.strip().upper())
                break

    chip_ids = sorted(all_chips)
    tested_parts = len(chip_ids)

    # -- enrich with job audit data (audit_data + chip_lookup already loaded above) --
    chip_rows        = []
    seen_pl          = set()
    playlist_filters = []

    for idx, chip_id in enumerate(chip_ids, start=1):
        job_entries = chip_lookup.get(chip_id) or []
        filtered    = (
            [je for je in job_entries
             if str(je.get("software_product") or "").strip() in matched_products]
            or job_entries
        )
        # deduplicate by job_id � keep only unique jobs
        seen_jids      = set()
        job_ids        = []
        job_details    = []   # [{job_id, start_date, playlist_name}]
        playlist_names = []
        certicom_entries = []
        seen_certicom_entries = set()
        for je in filtered:
            jid = je.get("job_id")
            if jid in (None, ""):
                continue
            jid_str = str(jid)
            if jid_str in seen_jids:
                continue          # same job_id already added � skip duplicate
            seen_jids.add(jid_str)
            job_ids.append(jid_str)
            # extract date from start_time
            st = str(je.get("start_time") or "").strip()
            date_str = st[:10] if len(st) >= 10 else ""
            pl = str(je.get("playlist_name") or je.get("playlist") or "").strip()
            job_details.append({"job_id": jid_str, "date": date_str, "playlist": pl})
            if pl and pl not in playlist_names:
                playlist_names.append(pl)
            if pl and pl not in seen_pl:
                seen_pl.add(pl)
                playlist_filters.append(pl)
            for pe in (je.get("certicom_playlist") or []):
                if not isinstance(pe, dict):
                    continue
                pe_name = str(pe.get("playlist_name") or pe.get("name") or pl or "Unknown").strip() or "Unknown"
                pe_id = str(pe.get("playlist_id") or pe.get("id") or pe_name).strip()
                pe_ids_raw = pe.get("certicom_ids") or pe.get("chip_ids") or pe.get("deviceSerialNumbers") or []
                pe_ids = [str(c).strip().upper() for c in pe_ids_raw if str(c).strip()] if isinstance(pe_ids_raw, list) else []
                if not pe_ids:
                    pe_ids = [chip_id]
                pe_key = (pe_id, pe_name, ",".join(pe_ids))
                if pe_key in seen_certicom_entries:
                    continue
                seen_certicom_entries.add(pe_key)
                certicom_entries.append({
                    "playlist_id": pe_id,
                    "playlist_name": pe_name,
                    "certicom_ids": pe_ids,
                })
        # multi_job = chip tested in more than one unique job ? highlight
        multi_job = len(job_ids) > 1
        chip_rows.append({
            "s_no":              idx,
            "chip_id":           chip_id,
            "job_ids":           job_ids,
            "job_details":       job_details,
            "job_ids_display":   ", ".join(job_ids),
            "playlists":         playlist_names,
            "playlists_display": ", ".join(playlist_names),
            "certicom_playlist": certicom_entries,
            "multi_job":         multi_job,
        })

    return jsonify({
        "success":            True,
        "sp_name":            sp_name,
        "family_prefix":      family_prefix,
        "matched_products":   matched_products,
        "tested_parts":       tested_parts,
        "projected_parts":    _get_projected_parts(target_name),
        "chip_ids":           chip_ids,
        "chip_rows":          chip_rows,
        "playlist_filters":   playlist_filters,
        "generated_at":       chip_data.get("generated_at"),
        "audit_generated_at": audit_data.get("generated_at"),
        "data_source":        audit_data.get("source", "unknown"),
    })




@dashboard_bp.route("/api/hwpdt_cr_venn/<string:target_name>", methods=["GET"])
@login_required
def api_hwpdt_cr_venn(target_name):
    conn = None
    cursor = None
    try:
        conn = get_mysql_connection_db()
        cursor = conn.cursor(dictionary=True)

        # ── helper: resolve table via dashboard_status first, then
        #    fall back to information_schema search by name pattern.
        #    This handles Axiom-only targets not yet in dashboard_status.
        def _resolve_table(suffix):
            """Return fully-qualified `schema`.`table` or None."""
            # 1. Try the managed target lookup (dashboard_status)
            try:
                return fq_table_for_target(target_name, suffix)
            except Exception:
                pass
            # 2. Fallback: search information_schema by <target>_<suffix>
            #    Try both the raw name and a dot-stripped variant
            candidates = [
                target_name.lower().replace('-', '_').replace(' ', '_').replace('.', '_') + '_' + suffix,
                target_name.lower().replace('-', '_').replace(' ', '_') + '_' + suffix,
            ]
            for pat in dict.fromkeys(candidates):   # deduplicate, preserve order
                cursor.execute(
                    "SELECT TABLE_SCHEMA, TABLE_NAME "
                    "FROM information_schema.TABLES "
                    "WHERE TABLE_NAME = %s LIMIT 1",
                    (pat,)
                )
                row = cursor.fetchone()
                if row:
                    _s = row.get('TABLE_SCHEMA') or row.get('table_schema', '')
                    _n = row.get('TABLE_NAME')   or row.get('table_name', '')
                    return f"`{_s}`.`{_n}`"
            return None

        # ── resolve all three tables ──────────────────────────────────────
        j_table = _resolve_table("jiras")
        if not j_table:
            return jsonify({"success": True, "rows": [], "area_summary": [],
                            "hero_cards": {"total_hwpdt_jiras": 0},
                            "summary": {"hwpdt_only": 0, "swpdt_only": 0, "both": 0},
                            "notice": "JIRAs table not available for this target."})

        o_table = _resolve_table("openjiras")
        if not o_table:
            return jsonify({"success": True, "rows": [], "area_summary": [],
                            "hero_cards": {"total_hwpdt_jiras": 0},
                            "summary": {"hwpdt_only": 0, "swpdt_only": 0, "both": 0},
                            "notice": "Open JIRAs table not available for this target."})

        u_table = _resolve_table("unique_crs")

        cursor.execute(f"SHOW COLUMNS FROM {j_table}")
        j_cols = {c['Field'] for c in (cursor.fetchall() or [])}

        ticket_col   = 'stability_ticket' if 'stability_ticket' in j_cols else ('jira_id' if 'jira_id' in j_cols else None)
        team_col     = 'test_team' if 'test_team' in j_cols else None
        area_col     = 'area' if 'area' in j_cols else None
        title_col    = 'jira_title' if 'jira_title' in j_cols else ('title' if 'title' in j_cols else None)
        j_mapped_col = 'mapped_crs' if 'mapped_crs' in j_cols else ('mapped_cr' if 'mapped_cr' in j_cols else ('cr' if 'cr' in j_cols else None))

        if not ticket_col or not j_mapped_col:
            return jsonify({"success": False, "message": "Required JIRA columns not available."}), 500

        def _hwpdt_tbl_ok(fq_name):
            n = str(fq_name or '').replace('`', '')
            try:
                s, t = n.split('.', 1)
            except ValueError:
                return True
            cursor.execute("SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1", (s, t))
            return cursor.fetchone() is not None

        select_parts_j = [
            f"TRIM(COALESCE(`{j_mapped_col}`, '')) AS raw_cr",
            f"TRIM(COALESCE(`{ticket_col}`, '')) AS stability_ticket",
            f"TRIM(COALESCE(`{team_col}`, '')) AS test_team" if team_col else "'' AS test_team",
            f"TRIM(COALESCE(`{area_col}`, '')) AS jira_area" if area_col else "'' AS jira_area",
            f"TRIM(COALESCE(`{title_col}`, '')) AS jira_title" if title_col else "'' AS jira_title",
        ]
        queries = [f"SELECT {', '.join(select_parts_j)} FROM {j_table} WHERE TRIM(COALESCE(`{j_mapped_col}`, '')) <> ''"]

        if _hwpdt_tbl_ok(o_table):
            cursor.execute(f"SHOW COLUMNS FROM {o_table}")
            o_cols = {c['Field'] for c in (cursor.fetchall() or [])}
            o_ticket_col   = 'stability_ticket' if 'stability_ticket' in o_cols else ('jira_id' if 'jira_id' in o_cols else None)
            o_team_col     = 'test_team' if 'test_team' in o_cols else None
            o_area_col     = 'area' if 'area' in o_cols else None
            o_title_col    = 'jira_title' if 'jira_title' in o_cols else ('title' if 'title' in o_cols else None)
            o_mapped_col   = 'mapped_crs' if 'mapped_crs' in o_cols else ('mapped_cr' if 'mapped_cr' in o_cols else ('cr' if 'cr' in o_cols else None))
            if o_ticket_col and o_mapped_col:
                select_parts_o = [
                    f"TRIM(COALESCE(`{o_mapped_col}`, '')) AS raw_cr",
                    f"TRIM(COALESCE(`{o_ticket_col}`, '')) AS stability_ticket",
                    f"TRIM(COALESCE(`{o_team_col}`, '')) AS test_team" if o_team_col else "'' AS test_team",
                    f"TRIM(COALESCE(`{o_area_col}`, '')) AS jira_area" if o_area_col else "'' AS jira_area",
                    f"TRIM(COALESCE(`{o_title_col}`, '')) AS jira_title" if o_title_col else "'' AS jira_title",
                ]
                queries.append(f"SELECT {', '.join(select_parts_o)} FROM {o_table} WHERE TRIM(COALESCE(`{o_mapped_col}`, '')) <> ''")

        cursor.execute(" UNION DISTINCT ".join(queries))
        jira_rows = cursor.fetchall() or []

        cr_raw_values = sorted({str(r.get('raw_cr') or '').strip() for r in jira_rows if str(r.get('raw_cr') or '').strip()})
        canonical_map = {}
        if cr_raw_values and _hwpdt_tbl_ok(u_table):
            ph = ','.join(['%s'] * len(cr_raw_values))
            cursor.execute(
                f"SELECT TRIM(COALESCE(`cr`, '')) AS cr, TRIM(COALESCE(`mapped_cr`, '')) AS mapped_cr "
                f"FROM {u_table} WHERE TRIM(COALESCE(`cr`, '')) IN ({ph}) OR TRIM(COALESCE(`mapped_cr`, '')) IN ({ph})",
                tuple(cr_raw_values) * 2,
            )
            for row in (cursor.fetchall() or []):
                raw    = str(row.get('cr') or '').strip()
                mapped = str(row.get('mapped_cr') or '').strip()
                if raw and mapped:   canonical_map[raw]    = mapped
                if mapped:           canonical_map[mapped] = mapped

        cr_groups = {}
        for row in jira_rows:
            raw_cr = str(row.get('raw_cr') or '').strip()
            if not raw_cr: continue
            canonical_cr = canonical_map.get(raw_cr, raw_cr)
            bucket = 'HWPDT' if str(row.get('test_team') or '').strip().upper() == 'PDT_QIPL_HWPDT' else 'SWPDT'
            grp = cr_groups.setdefault(canonical_cr, {
                'cr_id': canonical_cr, 'buckets': set(),
                'jira_tickets': set(), 'test_teams': set(),
                'jira_area': str(row.get('jira_area') or '').strip(),
                'jira_title': str(row.get('jira_title') or '').strip(),
            })
            grp['buckets'].add(bucket)
            ticket = str(row.get('stability_ticket') or '').strip()
            if ticket: grp['jira_tickets'].add(ticket)
            team = str(row.get('test_team') or '').strip()
            if team: grp['test_teams'].add(team)

        unique_meta = {}
        if cr_groups and _hwpdt_tbl_ok(u_table):
            cr_keys = list(cr_groups.keys())
            ph = ','.join(['%s'] * len(cr_keys))
            cursor.execute(f"SHOW COLUMNS FROM {u_table}")
            u_cols = {c['Field'] for c in (cursor.fetchall() or [])}
            def _col(name): return f"`{name}`" if name in u_cols else f"NULL AS `{name}`"
            cursor.execute(
                f"SELECT `mapped_cr`, {_col('cr_title')}, {_col('cr_area')}, {_col('cr_subsystem')}, "
                f"{_col('cr_status')}, {_col('cr_age')}, {_col('cr_occurrence')} "
                f"FROM {u_table} WHERE TRIM(COALESCE(`mapped_cr`, '')) IN ({ph})",
                tuple(cr_keys),
            )
            for row in (cursor.fetchall() or []):
                mapped  = str(row.get('mapped_cr') or '').strip()
                if not mapped: continue
                occ_raw = str(row.get('cr_occurrence') or '').strip().lower()
                prev    = unique_meta.get(mapped)
                if prev is None:
                    unique_meta[mapped] = row
                elif str(prev.get('cr_occurrence') or '').strip().lower() == 'dup' and occ_raw != 'dup':
                    unique_meta[mapped] = row

        # total CRs from jira cr column
        total_crs_from_jira_cr = 0
        j_cr_col = 'cr' if 'cr' in j_cols else None
        if j_cr_col:
            cr_values = set()
            cursor.execute(f"SELECT TRIM(COALESCE(`{j_cr_col}`, '')) AS cr_val FROM {j_table} WHERE TRIM(COALESCE(`{j_cr_col}`, '')) <> ''")
            for row in (cursor.fetchall() or []):
                v = str(row.get('cr_val') or '').strip()
                if v: cr_values.add(v)
            if _hwpdt_tbl_ok(o_table):
                cursor.execute(f"SHOW COLUMNS FROM {o_table}")
                o_cols_cr = {c['Field'] for c in (cursor.fetchall() or [])}
                o_cr_col  = 'cr' if 'cr' in o_cols_cr else None
                if o_cr_col:
                    cursor.execute(f"SELECT TRIM(COALESCE(`{o_cr_col}`, '')) AS cr_val FROM {o_table} WHERE TRIM(COALESCE(`{o_cr_col}`, '')) <> ''")
                    for row in (cursor.fetchall() or []):
                        v = str(row.get('cr_val') or '').strip()
                        if v: cr_values.add(v)
            total_crs_from_jira_cr = len(cr_values)

        def _count_hwpdt_team_jiras(fq_name):
            """Count distinct JIRA tickets where test_team is HWPDT for a resolved table."""
            if not fq_name or not _hwpdt_tbl_ok(fq_name):
                return 0
            cursor.execute(f"SHOW COLUMNS FROM {fq_name}")
            cols = {c['Field'] for c in (cursor.fetchall() or [])}
            ticket = 'stability_ticket' if 'stability_ticket' in cols else ('jira_id' if 'jira_id' in cols else None)
            team = 'test_team' if 'test_team' in cols else None
            if not ticket or not team:
                return 0
            cursor.execute(
                f"""
                SELECT COUNT(DISTINCT TRIM(COALESCE(`{ticket}`, ''))) AS cnt
                FROM {fq_name}
                WHERE TRIM(COALESCE(`{ticket}`, '')) <> ''
                  AND UPPER(TRIM(COALESCE(`{team}`, ''))) = 'PDT_QIPL_HWPDT'
                """
            )
            row = cursor.fetchone() or {}
            return int(row.get('cnt') or 0)

        # HWPDT JIRA widgets: count the HWPDT test-team tickets from each table.
        jira_table_hwpdt_jiras = _count_hwpdt_team_jiras(j_table)
        open_hwpdt_jiras       = _count_hwpdt_team_jiras(o_table)
        c_table                = _resolve_table("closed_jiras")
        closed_hwpdt_jiras     = _count_hwpdt_team_jiras(c_table)
        total_hwpdt_jiras      = jira_table_hwpdt_jiras + open_hwpdt_jiras + closed_hwpdt_jiras

        # invalid jiras from closed_jiras
        invalid_jiras = 0
        if c_table and _hwpdt_tbl_ok(c_table):
            cursor.execute(f"SHOW COLUMNS FROM {c_table}")
            c_cols = {c['Field'] for c in (cursor.fetchall() or [])}
            c_ticket_col = 'stability_ticket' if 'stability_ticket' in c_cols else ('jira_id' if 'jira_id' in c_cols else None)
            c_res_col    = 'resolution' if 'resolution' in c_cols else None
            if c_ticket_col:
                sel = [f"TRIM(COALESCE(`{c_ticket_col}`, '')) AS stability_ticket"]
                sel.append(f"TRIM(COALESCE(`{c_res_col}`, '')) AS resolution" if c_res_col else "'' AS resolution")
                cursor.execute(f"SELECT {', '.join(sel)} FROM {c_table}")
                invalid_seen = set()
                for row in (cursor.fetchall() or []):
                    ticket     = str(row.get('stability_ticket') or '').strip()
                    resolution = str(row.get('resolution') or '').strip().lower()
                    if not ticket or not ticket.upper().startswith('CHIPMD-'): continue
                    if resolution in ('manual force crash', 'test log'):       continue
                    invalid_seen.add(ticket)
                invalid_jiras = len(invalid_seen)

        # build details
        details = []
        hw_only = sw_only = overlap = 0
        for cr_id, grp in cr_groups.items():
            buckets = grp['buckets']
            if buckets == {'HWPDT'}:   category = 'HWPDT_ONLY'; hw_only += 1
            elif buckets == {'SWPDT'}: category = 'SWPDT_ONLY'; sw_only += 1
            else:                      category = 'OVERLAP';    overlap  += 1
            meta = unique_meta.get(cr_id) or {}
            details.append({
                'cr_id':      cr_id,
                'category':   category,
                'area':       str(meta.get('cr_area')    or grp.get('jira_area') or '').strip(),
                'subsystem':  str(meta.get('cr_subsystem') or '').strip(),
                'status':     str(meta.get('cr_status')  or '').strip(),
                'cr_age':     meta.get('cr_age'),
                'cr_title':   str(meta.get('cr_title')   or grp.get('jira_title') or '').strip(),
                'jira_count': len(grp.get('jira_tickets') or []),
                'test_teams': ', '.join(sorted(grp.get('test_teams') or [])),
            })

        details.sort(key=lambda r: (0 if r['category'] == 'OVERLAP' else 1, -(int(r['jira_count'] or 0)), str(r['cr_id'] or '')))

        return jsonify({
            'success': True,
            'hero_cards': {
                'total_crs':          total_crs_from_jira_cr or len(details),
                'swpdt_crs':          sw_only + overlap,
                                'hwpdt_crs':          hw_only + overlap,
                'total_hwpdt_jiras':       total_hwpdt_jiras,
                'jira_table_hwpdt_jiras':  jira_table_hwpdt_jiras,
                'open_hwpdt_jiras':        open_hwpdt_jiras,
                'closed_hwpdt_jiras':      closed_hwpdt_jiras,
                'invalid_jiras':           invalid_jiras,
                'total_tested_parts':      0,
            },
            'summary': {
                'all':        len(details),
                'hwpdt_only': hw_only,
                'swpdt_only': sw_only,
                'overlap':    overlap,
            },
            'rows': details,
        })

    except Exception as e:
        logger.error(f"[HWPDT CR VENN] Failed for {target_name}: {e}")
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        if cursor: cursor.close()
        if conn:   conn.close()



@dashboard_bp.route("/hwpdt_save/<string:target_name>", methods=["POST"])
@login_required
def hwpdt_save_view(target_name):

    try:
        payload_json = request.form.get("payload_json") or ""
        data = json.loads(payload_json) if payload_json else {}
        # data = { "pdt_type": ..., "toggle_mode": ..., "rows": [...] }

        # TODO: persist MODE / crashes per meta/build as needed.
        # For now you can just log it:
        logger.info("[HWPDT SAVE]")

        return redirect(
            url_for("dashboard_bp.dashboard",
                    target_name=target_name,
                    section="hwpdt",
                    toggle_mode=data.get("toggle_mode", "CRM"),
                    pdt_type=data.get("pdt_type", "HWPDT"))
        )
    except Exception as e:
        return f"Error saving HWPDT table: {str(e)}", 500
    

def build_cr_area_age_query(table_u, category):

    """
    Build an area summary query for a specific CR category.
    category: 'Built' or 'Undisposed'  (case-insensitive; DB stores lowercase)
    """
    cat = (category or "").strip().lower()
    if cat not in ("built", "undisposed"):
        raise ValueError(f"Unsupported CR category for age query: {category!r}")

    return f"""
        SELECT
            cr_area,
            COUNT(*) AS cr_count,
            AVG(CAST(NULLIF(cr_age, '') AS UNSIGNED)) AS avg_cr_age
        FROM {table_u}
        WHERE LOWER(TRIM(cr_category)) = '{cat}'
          AND CAST(NULLIF(cr_age, '') AS UNSIGNED) > 0
        GROUP BY cr_area
        ORDER BY cr_count DESC
    """


def get_dashboard_meta_for_target(target_name):

    """
    Fetch dashboard_latest_update / unique_cr_last_update and milestone dates
    from pdt_stats_dashboard.dashboard_status.

    Lookup order:
    1) exact target_name
    2) lowercase target_name
    3) base token before first underscore (e.g. GLYMUR_MAHUA_KALAMBO -> glymur)

    Returns a dict with keys: dashboard_latest_update, unique_cr_last_update, ES, FC, CS.
    """
    conn = get_mysql_connection_db("pdt_stats_dashboard")
    try:
        cursor = conn.cursor(dictionary=True)
        raw_target = str(target_name or "").strip()
        raw_upper = raw_target.upper()
        candidates = []
        if raw_upper in ("GLYMUR", "GLYMUR_MAHUA_KALAMBO"):
            for cand in (raw_target, raw_target.lower(), "glymur", "GLYMUR", "GLYMUR_MAHUA_KALAMBO", "glymur_mahua_kalambo"):
                if cand and cand not in candidates:
                    candidates.append(cand)
        else:
            for cand in (raw_target, raw_target.lower()):
                if cand and cand not in candidates:
                    candidates.append(cand)

        for cand in candidates:
            cursor.execute(
                """
                SELECT
                    dashboard_latest_update,
                    unique_cr_last_update,
                    es_date,
                    fc_date,
                    cs_date
                FROM pdt_stats_dashboard.dashboard_status
                WHERE LOWER(target_name) = LOWER(%s)
                  AND is_active = 1
                ORDER BY id DESC
                LIMIT 1
                """,
                (cand,),
            )
            row = cursor.fetchone()
            if row:
                return {
                    "dashboard_latest_update": row.get("dashboard_latest_update"),
                    "unique_cr_last_update": row.get("unique_cr_last_update"),
                    "ES": row.get("es_date"),
                    "FC": row.get("fc_date"),
                    "CS": row.get("cs_date"),
                }
        return {}
    finally:
        conn.close()


def build_milestone_phase_context(target_name):
    """
    Returns a milestone_phase dict for target_layout.html sidebar card.
    Compares milestone dates against today to compute phase/progress/status.
    has_milestone_data=True when at least one date is set (hides Refetch button).
    """
    try:
        from datetime import date as _date, datetime as _dt
        dash_meta = get_dashboard_meta_for_target(target_name) or {}
        def _to_date(val):
            if isinstance(val, _date): return val
            if isinstance(val, _dt):   return val.date()
            if not val or str(val).strip().upper() == "TBD": return None
            try: return _dt.fromisoformat(str(val)).date()
            except Exception: return None
        es_dt = _to_date(dash_meta.get("ES"))
        fc_dt = _to_date(dash_meta.get("FC"))
        cs_dt = _to_date(dash_meta.get("CS"))
        today = _date.today()

        # True when at least one milestone date is actually set
        has_milestone_data = any([es_dt, fc_dt, cs_dt])

        def _fmt(d): return d.strftime("%Y-%m-%d") if d else "TBD"
        def _status(d):
            if not d: return "pending"
            return "done" if today >= d else "upcoming"

        milestones = [
            {"label": "ES", "date": _fmt(es_dt), "status": _status(es_dt)},
            {"label": "FC", "date": _fmt(fc_dt), "status": _status(fc_dt)},
            {"label": "CS", "date": _fmt(cs_dt), "status": _status(cs_dt)},
        ]

        if not es_dt or today < es_dt:
            phase, pct = "Pre-ES", 0
        elif fc_dt and today < fc_dt:
            seg = max(0.0, min(1.0, (today - es_dt).days / max(1, (fc_dt - es_dt).days)))
            phase, pct = "ES ? FC", round(33 + seg * 33)
        elif cs_dt and today < cs_dt:
            seg = max(0.0, min(1.0, (today - fc_dt).days / max(1, (cs_dt - fc_dt).days)))
            phase, pct = "FC ? CS", round(66 + seg * 34)
        else:
            phase, pct = "CS Done ?", 100

        return {
            "phase_label":        "Release Milestones",
            "phase_name":         phase,
            "progress_pct":       pct,
            "milestones":         milestones,
            "has_milestone_data": has_milestone_data,
        }
    except Exception as _e:
        import traceback
        print(f"[build_milestone_phase_context] ERROR for {target_name}: {_e}\n{traceback.format_exc()}")
        return None


def build_weekly_report_context(target_name, request):
    """Helper to build weekly report data for a target."""
    today = date.today()
    # Default = last completed Mon-Sun week
    days_since_monday = today.weekday()  # 0=Mon, 6=Sun
    this_monday  = today - timedelta(days=days_since_monday)
    last_monday  = this_monday - timedelta(days=7)
    last_sunday  = last_monday + timedelta(days=6)
    default_from = last_monday
    default_to   = last_sunday

    schema_name = get_schema_for_target(target_name)
    if not schema_name:
        return None, None, None, "Could not determine BU/schema for this target."

    conn = get_mysql_connection_db(bu_key=schema_name)
    if not conn:
        return None, None, None, "Database connection error."

    # ---- 1) Determine date range (defaults, user input, or ALL) ----
    if request.method == "POST":
        range_type = (request.form.get("range_type") or "week").strip().lower()
        from_date_str = request.form.get("from_date") or default_from.strftime("%Y-%m-%d")
        to_date_str = request.form.get("to_date") or default_to.strftime("%Y-%m-%d")
    else:
        range_type = (request.args.get("range_type") or "week").strip().lower()
        from_date_str = request.args.get("from_date") or default_from.strftime("%Y-%m-%d")
        to_date_str = request.args.get("to_date") or default_to.strftime("%Y-%m-%d")

    if range_type == "all":
        from_dt = to_dt = "all"
        from_date_str = to_date_str = "ALL"
    else:
        try:
            from_dt = datetime.strptime(from_date_str, "%Y-%m-%d").date()
            to_dt = datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            return None, from_date_str, to_date_str, "Invalid date format. Please use YYYY-MM-DD."

        # ---- 2) Always fetch weekly data for the chosen range ----
    report_summary = get_weekly_report_data(
        conn,
        schema_name,
        target_name,
        from_date=from_dt,
        to_date=to_dt,
    )

    # ---- 3) Apply CR title exclude filter (Compute only) ----
    from dashboard_common import get_bu_for_target as _get_bu
    if (_get_bu(target_name) or '').upper() == 'COMPUTE' and report_summary:
        _excl = _get_cr_title_exclude(target_name)
        if _excl['enabled'] and _excl['keywords']:
            _kws = [k.lower() for k in _excl['keywords']]
            orig_rows = report_summary.get('cr_rows') or []
            filtered_rows = [
                r for r in orig_rows
                if not any(kw in str(r.get('cr_title') or '').lower() for kw in _kws)
            ]
            # Recompute counts and pie data from filtered rows
            from collections import Counter as _Counter
            report_summary['cr_rows']          = filtered_rows
            report_summary['num_crs_reported'] = len(filtered_rows)
            report_summary['num_crs_week']     = len(filtered_rows)
            _status_counts = _Counter((r.get('cr_status') or 'Unknown').strip() for r in filtered_rows)
            _area_counts   = _Counter((r.get('cr_area')   or 'Unknown').strip() for r in filtered_rows)
            report_summary['cr_status_counts'] = dict(_status_counts)
            report_summary['cr_status_pie']    = [{'name': k, 'y': v} for k, v in sorted(_status_counts.items(), key=lambda x: x[0].lower())]
            report_summary['cr_area_pie']      = [{'name': k, 'y': v} for k, v in sorted(_area_counts.items(),   key=lambda x: x[0].lower())]

    return report_summary, from_date_str, to_date_str, None


# ---------------------------------------------------------------------
# HELPERS FOR CR AGE (detail + status)
# ---------------------------------------------------------------------
def fetch_undisposed_crs_in_age_band(table_u, min_age, max_age=None):

    """
    Build a query for individual undisposed CR rows whose age is:
    - between min_age and max_age (inclusive lower, exclusive upper), or
        - >= min_age when max_age is None.
    """
    conditions = [
        "LOWER(TRIM(cr_category)) = 'undisposed'",
        "CAST(NULLIF(cr_age, '') AS UNSIGNED) > 0",
        f"CAST(NULLIF(cr_age, '') AS UNSIGNED) >= {min_age}",
    ]
    if max_age is not None:
        conditions.append(f"CAST(NULLIF(cr_age, '') AS UNSIGNED) < {max_age}")

    where_clause = " AND ".join(conditions)
    return f"""
        SELECT
            mapped_cr    AS cr_id,
            cr_title,
            cr_status,
            cr_area,
            CAST(NULLIF(cr_age, '') AS UNSIGNED) AS cr_age
        FROM {table_u}
        WHERE {where_clause}
        ORDER BY cr_age DESC
    """


def fetch_undisposed_status_counts(cursor, table_u):

    """
        Return counts of undisposed CRs by status (Open / Analysis / Other).
    """
    q = f"""
        SELECT
            cr_status,
            COUNT(*) AS cnt
        FROM {table_u}
        WHERE LOWER(TRIM(cr_category)) = 'undisposed'
          AND CAST(NULLIF(cr_age, '') AS UNSIGNED) > 0
        GROUP BY cr_status
    """
    cursor.execute(q)
    rows = cursor.fetchall() or []

    result = {"open": 0, "analysis": 0, "other": 0}
    for r in rows:
        status = (r["cr_status"] or "").strip().lower()
        count = int(r["cnt"] or 0)
        if "open" in status:
            result["open"] += count
        elif "analysis" in status or "anal" in status:
            result["analysis"] += count
        else:
            result["other"] += count
    return result


from dashboard_service import save_meta_report_bulk, ensure_meta_builds_table
from src.axiom_client import get_devices_by_chipset, get_devices_site_wise
from dashboard_common import get_bu_for_target, get_chip_name_for_target, is_axiom_enabled_for_target, get_business_units, get_targets_for_bu


# ---- Caching helpers for Axiom device JSON ----

def _axiom_cache_dir():

    root = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'axiom_cache')
    try:
        os.makedirs(root, exist_ok=True)
    except Exception:
        pass
    return root


def _cache_path_for(chipset, pdt_type):

    chip = (chipset or '').strip().upper() or 'UNKNOWN'
    pdt = (pdt_type or 'SWPDT').strip().upper()
    name = "{}_{}.json".format(chip, pdt)
    return os.path.join(_axiom_cache_dir(), name)


def _load_cached_devices(chipset, pdt_type):

    path = _cache_path_for(chipset, pdt_type)
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            return json.load(fh)
    except Exception:
        return None


def _save_cached_devices(chipset, pdt_type, devices):

    payload = {
        'chipset': (chipset or '').strip().upper(),
        'pdt_type': (pdt_type or 'SWPDT').strip().upper(),
        'saved_at': _dt.utcnow().isoformat() + 'Z',
        'count': len(devices or []),
        'devices': devices or [],
    }
    path = _cache_path_for(payload['chipset'], payload['pdt_type'])
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
    return path

# ---- Overrides: SW DEL and HWPDT metrics ----

def _overrides_dir():

    root = os.path.join(_axiom_cache_dir(), 'overrides')
    try:
        os.makedirs(root, exist_ok=True)
    except Exception:
        pass
    return root


def _sw_del_override_path(chipset):

    chip = (chipset or '').strip().upper() or 'UNKNOWN'
    return os.path.join(_overrides_dir(), "{}_SWPDT_DEL.json".format(chip))


def _hw_metrics_override_path(chipset):

    chip = (chipset or '').strip().upper() or 'UNKNOWN'
    return os.path.join(_overrides_dir(), "{}_HWPDT_metrics.json".format(chip))


def _load_sw_del_overrides(chipset):

    path = _sw_del_override_path(chipset)
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            return json.load(fh)
    except Exception:
        return None


def _save_sw_del_overrides(chipset, rows):

    payload = {
        'chipset': (chipset or '').strip().upper(),
        'saved_at': _dt.utcnow().isoformat() + 'Z',
        'rows': rows or [],
    }
    path = _sw_del_override_path(payload['chipset'])
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
    return path


def _load_hw_metrics_overrides(chipset):

    path = _hw_metrics_override_path(chipset)
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            return json.load(fh)
    except Exception:
        return None


def _save_hw_metrics_overrides(chipset, columns, rows):

    payload = {
        'chipset': (chipset or '').strip().upper(),
        'saved_at': _dt.utcnow().isoformat() + 'Z',
        'columns': columns or ["REV0","REV1","Part Type","Total"],
        'rows': rows or [],
    }
    path = _hw_metrics_override_path(payload['chipset'])
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
    return path

def _build_view_devices_url(target_name, pdt_type):

    pdt = str(pdt_type or 'SWPDT').strip().upper()
    return url_for('device_summary_api_bp.device_summary_devices_page', target_name=target_name, pdt=pdt)


def get_unified_device_summary(target_name, pdt_type, refresh=False):
    chip_name = get_chip_name_for_target(target_name) or ""
    if not chip_name:
        return None, "No chip name configured for this target."

    pdt_type = (pdt_type or "SWPDT").strip().upper()
    cache_path = os.path.join(_axiom_cache_dir(), f"{chip_name}_{pdt_type}_unified.json")

    if not refresh and os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                cached = json.load(f)
            return cached.get("devices", []), None
        except Exception:
            pass

    # No cache � return sentinel so the page renders the syncing banner immediately.
    # Live fetch is handled by /api/device_summary_data (triggered by the banner button).
    if not refresh:
        return [], "NO_CACHE"

    # refresh=True path: live fetch from Axiom + QDT
    try:
        axiom_devices = get_devices_by_chipset(
            chip_name,
            pdt_type=pdt_type,
            include_site_details=True,
        ) or []
    except Exception as e:
        return None, f"Axiom error: {e}"

    qdt_devices = []
    try:
        qdt_devices = get_rework_info_from_qdt(chip_name) or []
    except Exception:
        qdt_devices = []

    qdt_map = {}
    for item in qdt_devices:
        serial = str(item.get("SERIAL_NO") or item.get("serial_no") or item.get("serial") or "").strip().upper()
        if serial:
            qdt_map[serial] = item

    unified = []
    for dev in axiom_devices:
        merged = dict(dev)
        serial = str(dev.get("serial") or dev.get("serial_no") or dev.get("SERIAL_NO") or "").strip().upper()
        q_item = qdt_map.get(serial)
        if q_item:
            merged["rework_info"] = q_item.get("REWORK_INFO") or q_item.get("rework_info") or q_item.get("ReworkInfo") or ""
            merged["mcn"] = merged.get("mcn") or q_item.get("MCN") or q_item.get("mcn") or q_item.get("MCNNO") or ""
        unified.append(merged)

    payload = {
        "chipset": chip_name.upper(),
        "pdt_type": pdt_type,
        "saved_at": _dt.utcnow().isoformat() + "Z",
        "count": len(unified),
        "devices": unified,
    }
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
    except Exception:
        pass

    return unified, None


def _build_swpdt_deployment_table(devices):

    """
    Build the SW PDT Global Device Deployment table data.
    Uses _raw.taxonomyPath to determine site (CH / QIPL / SD).
    Excludes /PDT/QIPL/HW devices.
    Only shows sites that have at least 1 device.
        """
    import re as _re
    # MCN pattern: 10-XXXXX-XXX or 10-XXXXX-XXXX (followed by non-digit or end)
    _MCN_RE = _re.compile(r'\b(\d{2}-\d{5}-\d{3,4})(?!\d)')

    def _get_taxonomy(dev):
        raw = dev.get("_raw") or {}
        return str(
            raw.get("taxonomyPath")
            or dev.get("taxonomy_path")
            or ""
        ).strip().upper()

    def _get_site_abbr(dev):
        tx = _get_taxonomy(dev)
        if "/PDT/QIPL" in tx:  return "QIPL"
        if "/PDT/CHINA" in tx: return "CH"
        if "/PDT/SD" in tx or "/PDT/SAN DIEGO" in tx or "/PDT/SANDIEGO" in tx:
            return "SD"
        loc = str(dev.get("location") or "").upper()
        if "/AP/SHANGHAI" in loc or "/AP/CHINA" in loc:  return "CH"
        if "/AP/SAN DIEGO" in loc or "/AP/SD" in loc:    return "SD"
        if "/AP/HYDERABAD" in loc:                        return "QIPL"
        rack = str((dev.get("site_info") or {}).get("rack") or "").upper()
        if "SHANGHAI" in rack or "SHENZHEN" in rack:     return "CH"
        if "SAN DIEGO" in rack:                           return "SD"
        if "HYDERABAD" in rack:                           return "QIPL"
        return "UNKNOWN"

    def _get_mcn(dev):
        # Extract MCN (10-XXXXX-XXXX) from description only.
        # chipsetRev (V1.0) is a hardware revision, NOT an MCN.
        raw  = dev.get("_raw") or {}
        desc = str(raw.get("description") or dev.get("description") or "")
        m = _MCN_RE.search(desc)
        if m:
            return m.group(1)
        qdt_desc = str(dev.get("qdt_model_desc") or "")
        m2 = _MCN_RE.search(qdt_desc)
        if m2:
            return m2.group(1)
        return ""

    # Exclude /PDT/QIPL/HW � those belong to HWPDT only
    sw_only_devices = [
        d for d in (devices or [])
        if not _get_taxonomy(d).startswith("/PDT/QIPL/HW")
    ]

    table = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {"del": 0, "dep": 0}))))
    site_counts = defaultdict(int)

    for dev in sw_only_devices:
        raw  = dev.get("_raw") or {}
        deps = raw.get("dependencies") or {}

        ff        = str(dev.get("form_factor") or deps.get("formFactor") or "Unknown").strip() or "Unknown"
        mcn       = _get_mcn(dev)
        sto       = str(deps.get("storage Type") or dev.get("storage") or "LP4X").strip() or "LP4X"
        site_abbr = _get_site_abbr(dev)

        table[ff][mcn][sto][site_abbr]["del"] += 1
        table[ff][mcn][sto][site_abbr]["dep"] += 1
        site_counts[site_abbr] += 1

    # Only include sites that actually have devices � skip UNKNOWN and empty sites
    SITE_ORDER = ["QIPL", "CH", "SD"]
    active_sites = [s for s in SITE_ORDER if site_counts.get(s, 0) > 0]

    rows = []
    totals = {s: {"del": 0, "dep": 0} for s in active_sites}
    grand_del = grand_dep = 0

    for ff, mcn_map in sorted(table.items()):
        ff_first = True
        for mcn, sto_map in sorted(mcn_map.items()):
            for sto, site_map in sorted(sto_map.items()):
                row = {
                    "form_factor": ff if ff_first else "",
                    "mcn": mcn,
                    "storage": sto,
                    "sites": {},
                    "total_del": 0,
                    "total_dep": 0,
                }
                for s in active_sites:
                    d = site_map.get(s, {"del": 0, "dep": 0})
                    row["sites"][s] = d
                    row["total_del"] += d["del"]
                    row["total_dep"] += d["dep"]
                    totals[s]["del"] += d["del"]
                    totals[s]["dep"] += d["dep"]
                grand_del += row["total_del"]
                grand_dep += row["total_dep"]
                rows.append(row)
                ff_first = False

    return {
        "sites": active_sites,
        "rows": rows,
        "totals": totals,
        "grand_del": grand_del,
        "grand_dep": grand_dep,
    }


def _default_hwpdt_summary():

    columns = ["REV0", "REV1", "Total"]
    labels = [
        "Total No of MSMs received",
        "Total No of MSMs Screened",
        "No of MSMs Failed",
        "No of MSM's Under Debug",
        "No of MSM's Recovered",
        "No of CDPs Received",
        "No of CDPs Deployed",
    ]
    return {
        "columns": columns,
        "metrics": [{"label": label, "values": ["", "", ""]} for label in labels],
        "total_devices": 0,
    }


def _build_hwpdt_screening_summary(hw_devices):

    """
    Build the HW PDT MSM Screening Summary table.
    Returns editable structure with columns + values.
    For devices with taxonomy path /PDT/QIPL/HW, update CDP values.
    """
    if not hw_devices:
        return _default_hwpdt_summary()

    hw_qipl_devices = [d for d in hw_devices if str(d.get("taxonomy_path") or "").strip().upper() == "/PDT/QIPL/HW"]

    total = len(hw_devices)
    cdp_received = len(hw_qipl_devices)
    cdp_deployed = len(hw_qipl_devices)

    return {
        "columns": ["REV0", "REV1", "Total"],
        "metrics": [
            {"label": "Total No of MSMs received", "values": ["", "", ""]},
            {"label": "Total No of MSMs Screened", "values": ["", "", ""]},
            {"label": "No of MSMs Failed", "values": ["", "", ""]},
            {"label": "No of MSM's Under Debug", "values": ["", "", ""]},
            {"label": "No of MSM's Recovered", "values": ["", "", ""]},
            {"label": "No of CDPs Received", "values": ["", "", str(cdp_received)]},
            {"label": "No of CDPs Deployed", "values": ["", "", str(cdp_deployed)]},
        ],
        "total_devices": total,
    }


def device_summary(target_name, base_context, cursor, conn):
    """
    Device Summary section � only for non-AUTO BUs.
    SW DEL comes from saved override if present, else Axiom default.
    SW DEP always stays from Axiom.
    HW summary is read-only by default and becomes editable only after Edit is clicked.
    """
    from dashboard_common import get_bu_for_target

    bu = (get_bu_for_target(target_name) or "").upper()
    if bu == "AUTO":
        return render_template(
            "device_summary.html",
            target=target_name,
            axiom_enabled=False,
            axiom_error="Device Summary is not available for Automotive (AUTO) targets.",
            sw_table=None,
            hw_summary=None,
            chip_name="",
            total_sw_devices=0,
            total_hw_devices=0,
            hw_editable=False,
            **base_context,
        )

    chip_name = get_chip_name_for_target(target_name) or ""
    axiom_enabled = is_axiom_enabled_for_target(target_name)
    axiom_error = None
    sw_table = None
    hw_summary = None
    total_sw_devices = 0
    total_hw_devices = 0
    sw_override = _load_sw_del_overrides(chip_name) if chip_name else None
    hw_override = _load_hw_metrics_overrides(chip_name) if chip_name else None
    hw_available = bool(base_context.get("hwpdt_available"))

    try:
        sw_devices, sw_err = get_unified_device_summary(target_name, "SWPDT")
        if sw_err:
            axiom_error = sw_err
            sw_devices = []
        total_sw_devices = len(sw_devices)

        if sw_devices:
            sw_table = _build_swpdt_deployment_table(sw_devices)
            if sw_override and sw_override.get("rows"):
                del_map = {}
                for item in sw_override.get("rows") or []:
                    key = (
                        str(item.get("form_factor") or ""),
                        str(item.get("mcn") or ""),
                        str(item.get("storage") or ""),
                        str(item.get("site") or ""),
                    )
                    del_map[key] = int(item.get("del") or 0)

                for row in sw_table.get("rows", []):
                    row["total_del"] = 0
                    row["total_dep"] = 0
                    for site in sw_table.get("sites", []):
                        site_data = row["sites"].get(site) or {"del": 0, "dep": 0}
                        key = (
                            str(row.get("form_factor") or ""),
                            str(row.get("mcn") or ""),
                            str(row.get("storage") or ""),
                            str(site),
                        )
                        if key in del_map:
                            site_data["del"] = del_map[key]
                        row["sites"][site] = site_data
                        row["total_del"] += int(site_data.get("del") or 0)
                        row["total_dep"] += int(site_data.get("dep") or 0)

                totals = {s: {"del": 0, "dep": 0} for s in sw_table.get("sites", [])}
                grand_del = 0
                grand_dep = 0
                for row in sw_table.get("rows", []):
                    grand_del += int(row.get("total_del") or 0)
                    grand_dep += int(row.get("total_dep") or 0)
                    for site in sw_table.get("sites", []):
                        site_data = row["sites"].get(site) or {"del": 0, "dep": 0}
                        totals[site]["del"] += int(site_data.get("del") or 0)
                        totals[site]["dep"] += int(site_data.get("dep") or 0)
                sw_table["totals"] = totals
                sw_table["grand_del"] = grand_del
                sw_table["grand_dep"] = grand_dep

        hw_devices, hw_err = get_unified_device_summary(target_name, "HWPDT")
        if hw_err and not axiom_error:
            axiom_error = hw_err
            hw_devices = []
        total_hw_devices = len(hw_devices or [])

        if hw_override and (hw_override.get("rows") or hw_override.get("columns")):
            hw_summary = {
                "columns": hw_override.get("columns") or ["REV0", "REV1", "Part Type", "Total"],
                "metrics": hw_override.get("rows") or [],
                "total_devices": total_hw_devices,
            }
        elif hw_devices:
            hw_summary = _build_hwpdt_screening_summary(hw_devices)
        elif hw_available:
            hw_summary = _default_hwpdt_summary()

    except Exception as e:
        logger.info(f"[DEVICE SUMMARY] error for {target_name}: {e}")
        axiom_error = str(e)

    return render_template(
        "device_summary.html",
        target=target_name,
        axiom_enabled=axiom_enabled,
        axiom_error=axiom_error,
        chip_name=chip_name,
        sw_table=sw_table,
        hw_summary=hw_summary,
        total_sw_devices=total_sw_devices,
        total_hw_devices=total_hw_devices,
        sw_view_devices_url=_build_view_devices_url(target_name, "SWPDT"),
        hw_view_devices_url=_build_view_devices_url(target_name, "HWPDT"),
        hw_editable=False,
        **base_context,
    )


# ---------------------------------------------------------------------
# MAIN DASHBOARD VIEW
# ---------------------------------------------------------------------
@dashboard_bp.route("/dashboard/<string:target_name>/mtbf-json")
@dashboard_bp.route("/dashboard/<string:target_name>/mtbf-excel")
@login_required
def target_mtbf_excel_page(target_name):

    cfg            = (_get_target_excel_config(target_name) or {}).get('mtbf', {})
    excel_path     = cfg.get('excel_path', '')
    selected_sheet = (request.args.get('sheet') or '').strip()
    sheet_names    = []
    preview_columns = []
    preview_rows    = []
    total_rows      = 0
    chart_data      = []   # [{build, mtbf, week, hours, crashes}]
    error           = ''
    from dashboard_common import get_bu_for_target
    is_compute_mtbf = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'

    # MTBF is JSON-only for all targets. Keep Compute's historical Glymur/Mahua
    # view switch; non-compute targets use a single MTBF JSON table.
    selected_sheet = _mtbf_json_view_name(selected_sheet or cfg.get('sheet_name') or ('Glymur' if is_compute_mtbf else 'MTBF')) if is_compute_mtbf else 'MTBF'
    sheet_names = list(_MTBF_JSON_VIEW_NAMES) if is_compute_mtbf else ['MTBF']
    if is_compute_mtbf:
        try:
            _migrate_compute_mtbf_excel_to_json_if_needed(target_name, excel_path)
        except Exception:
            logger.debug('[MTBF JSON] Excel-to-JSON migration failed for %s', target_name, exc_info=True)
    json_payload = _load_mtbf_json_payload(target_name, selected_sheet)
    json_rows = json_payload.get('rows') or []
    preview_columns = ['S.No'] + _mtbf_json_headers(is_compute_mtbf)
    preview_rows = _mtbf_json_to_preview_rows(json_rows, is_compute=is_compute_mtbf)
    total_rows = len(json_rows)
    chart_data = _mtbf_json_to_chart_data(json_rows, is_compute=is_compute_mtbf)
    ctx = _build_sidebar_context(target_name, active_section='mtbf-excel')
    ctx['target_display_name'] = get_display_name_for_target(target_name).upper()
    return render_template(
        'target_mtbf_excel.html',
        excel_path='',
        managed_upload=False,
        original_filename=f'{selected_sheet}.json',
        selected_sheet=selected_sheet,
        sheet_names=sheet_names,
        preview_columns=preview_columns,
        preview_rows=preview_rows,
        total_rows=total_rows,
        chart_data=chart_data,
        page_error='',
        is_compute_mtbf=is_compute_mtbf,
        mtbf_storage_mode='json',
        **ctx,
    )

    if not selected_sheet and not is_compute_mtbf:
        selected_sheet = cfg.get('sheet_name', '')



    if excel_path:
        try:
            _, sheet_names = _read_excel_sheet_names(excel_path)
            # Keep workbook sheet names exactly as Excel reports them. Use strip/lower only
            # for matching, otherwise openpyxl can fail if a valid sheet has spaces/case.
            sheet_names = [str(s) for s in (sheet_names or []) if str(s).strip()]
            if selected_sheet:
                requested_sheet = str(selected_sheet)
                selected_sheet = requested_sheet
                if selected_sheet not in sheet_names:
                    selected_sheet = next(
                        (s for s in sheet_names if str(s).strip().lower() == requested_sheet.strip().lower()),
                        ''
                    )

                if not selected_sheet and len(sheet_names) == 1:
                    selected_sheet = sheet_names[0]
            if is_compute_mtbf and not selected_sheet and len(sheet_names) > 1:
                selected_sheet = sheet_names[0]

            if (not is_compute_mtbf) and not selected_sheet and len(sheet_names) > 1:
                cfg_sheet = str(cfg.get('sheet_name', '') or '')
                selected_sheet = cfg_sheet if cfg_sheet in sheet_names else next(
                    (s for s in sheet_names if str(s).strip().lower() == cfg_sheet.strip().lower()),
                    sheet_names[0]
                )

            if selected_sheet and sheet_names and selected_sheet not in sheet_names:
                matched_sheet = next(
                    (s for s in sheet_names if str(s).strip().lower() == str(selected_sheet).strip().lower()),
                    ''
                )
                selected_sheet = matched_sheet or selected_sheet



            if selected_sheet:
                _, preview_columns, preview_rows, total_rows = _read_excel_sheet_preview(
                    excel_path, selected_sheet, max_rows=200
                )
                # build chart data from full sheet
                import openpyxl
                from datetime import datetime as _dt
                _path = _normalize_excel_path(excel_path)
                _wb   = openpyxl.load_workbook(_path, data_only=True)
                _ws   = _wb[selected_sheet]
                _hdrs = [str(_ws.cell(1, c).value or '').strip().lower()
                         for c in range(1, _ws.max_column + 1)]
                def _ci(names):
                    for n in names:
                        for i, h in enumerate(_hdrs):
                            if n in h:
                                return i
                    return None
                wi = _ci(['date', 'week', 'run date'])
                hi = _ci(['tested hours', 'total hours', 'hours tested', 'hours'])
                ci = _ci(['total crashes', 'crash count', 'crashes', 'crash'])
                qci = _ci(['qc crashes'])
                sheet_key = (selected_sheet or target_name or '').strip().lower()
                pmi = _ci([f'{sheet_key} product mtbf', 'product mtbf']) if is_compute_mtbf else None
                qmi = _ci([f'{sheet_key} qc mtbf', 'qc mtbf']) if is_compute_mtbf else None
                mi = _ci(['mtbf (hrs)', 'mtbf hrs', 'mtbf hours', 'mtbf']) if not is_compute_mtbf else None
                fi = _ci(['product line'])

                                # Prefer the LAST build-like column as the display/full build source.
                # Many workbooks have multiple build columns where the later one is the
                # actual plotted build string users expect on the x-axis.
                _build_col_indices = [idx for idx, h in enumerate(_hdrs) if 'build' in h]
                bi = _build_col_indices[-1] if _build_col_indices else _ci(['meta id', 'meta'])
                full_bi = bi
                # Dedicated META-ID column (separate from build string)
                meta_id_bi = _ci(['meta id', 'meta-id', 'metaid', 'meta_id'])

                def _short_build_label(raw):
                    s = '' if raw is None else str(raw).strip()
                    if not s:
                        return ''
                    import re as _re
                    m = _re.search(r'\b(META-[\w.-]+)', s, _re.I)
                    if m:
                        return m.group(1)
                    if ',' in s:
                        s = s.split(',')[-1].strip()
                    return s
                for rn in range(2, _ws.max_row + 1):
                    def cv(i):
                        if i is None: return ''
                        v = _ws.cell(rn, i + 1).value
                        if v is None: return ''
                        if isinstance(v, _dt): return v.strftime('%Y-%m-%d')
                        return str(v).strip()
                    def _date_or_empty(i):
                        if i is None: return ''
                        v = _ws.cell(rn, i + 1).value
                        if v is None: return ''
                        if isinstance(v, _dt): return v.strftime('%Y-%m-%d')
                        s = str(v).strip()
                        return s if s else ''
                    build = cv(bi)

                    if not build: continue
                    try:    hours_val = int(float(_ws.cell(rn, hi + 1).value or 0)) if hi is not None else 0
                    except:
                        hours_val = 0
                    try:    crashes_val = int(float(_ws.cell(rn, ci + 1).value or 0)) if ci is not None else 0
                    except:
                        crashes_val = 0
                    try:    qc_crashes_val = int(float(_ws.cell(rn, qci + 1).value or 0)) if qci is not None else 0
                    except:
                        qc_crashes_val = 0
                    if is_compute_mtbf:
                        try:    product_mtbf_val = round(float(_ws.cell(rn, pmi + 1).value or 0), 2) if pmi is not None else 0
                        except:
                            product_mtbf_val = 0
                        try:    qc_mtbf_val = round(float(_ws.cell(rn, qmi + 1).value or 0), 2) if qmi is not None else 0
                        except:
                            qc_mtbf_val = 0
                            full_build = cv(full_bi) if full_bi is not None else build
                            chart_data.append({
                            'build': build,
                            'full_build': full_build,
                            'label': _short_build_label(full_build or build),
                            'product_line': cv(fi) if fi is not None else '',
                            'week': _date_or_empty(wi),
                            'hours': hours_val,
                            'crashes': crashes_val,
                            'qc_crashes': qc_crashes_val,
                            'product_mtbf': product_mtbf_val,
                            'qc_mtbf': qc_mtbf_val,
                            'mtbf': qc_mtbf_val,
                        })
                    else:
                        try:    mtbf_val = round(float(_ws.cell(rn, mi + 1).value or 0), 1) if mi is not None else 0
                        except:
                            mtbf_val = 0
                            full_build = cv(full_bi) if full_bi is not None else build
                            # META-ID: use dedicated column first, then extract from build string
                            raw_meta = cv(meta_id_bi) if meta_id_bi is not None else ''
                            meta_id = raw_meta or _short_build_label(build) or build
                            chart_data.append({
                                'build': meta_id,
                                'full_build': full_build,
                                'label': meta_id,
                                'product_line': cv(fi) if fi is not None else '',
                                'mtbf': mtbf_val,
                                'week': _date_or_empty(wi),
                                'hours': hours_val,
                                'crashes': crashes_val,
                            })


        except Exception as exc:
            error = str(exc)

        if selected_sheet and total_rows and not chart_data and not error:
            error = 'Selected sheet was loaded, but MTBF parser could not detect expected Build / MTBF columns. Please verify sheet headers.'
        if selected_sheet and sheet_names and selected_sheet not in sheet_names:
            error = f'Worksheet named {selected_sheet!r} not found'
        if is_compute_mtbf and len(sheet_names) > 1 and selected_sheet in sheet_names:
            error = ''

        # inject S.No if not present
    if preview_columns and preview_rows:
        sno_present = any('s.no' in str(c).lower() or str(c).strip() == '#'
                          for c in preview_columns)
        if not sno_present:
            preview_columns = ['S.No'] + list(preview_columns)
            preview_rows = [
                [{'v': str(i + 1), 'rs': 1, 'cs': 1, 'skip': False}] + list(r)
                for i, r in enumerate(preview_rows)
            ]

    ctx = _build_sidebar_context(target_name, active_section='mtbf-excel')
    ctx['target_display_name'] = get_display_name_for_target(target_name).upper()
    return render_template(
        'target_mtbf_excel.html',
                excel_path=excel_path,
        managed_upload=bool(cfg.get('managed_upload')),
        original_filename=cfg.get('original_filename', ''),
        selected_sheet=selected_sheet,
        sheet_names=sheet_names,

        preview_columns=preview_columns,
        preview_rows=preview_rows,
        total_rows=total_rows,
        chart_data=chart_data,
                page_error=error,
        is_compute_mtbf=is_compute_mtbf,
        **ctx,
    )


@dashboard_bp.route("/dashboard/<string:target_name>/mtbf-excel/sample")
@login_required
def download_mtbf_excel_sample(target_name):
    """Download a sample MTBF workbook. Extra columns are allowed on upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MTBF'
    headers = [
        'Target(s)', 'Product Line(s)', 'Device(s)', 'Device Count', 'Week',
        'Meta-ID', 'Build(s) Full ID', 'Tested Hours', 'Reduction %', 'Total Crashes',
        'MTBF', 'Product MTBF', 'QC MTBF', 'MTBF Details', 'Notes'
    ]

    ws.append(headers)
    ws.append([
        target_name, 'Example Product', 'Device-A / Device-B', 2,

        date.today().strftime('%Y-%m-%d'), 'META-00250',
        'Example.LA.1.0-META-00250', 120, 0, 1, 120.0, 125.0, 120.0,
        '120 tested hours / 1 crash',
        'Extra columns are fine; keep required Build and MTBF headers.'
    ])

    fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
    ws.freeze_panes = 'A2'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=f'{_safe_target_slug(target_name)}_MTBF_sample.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/upload", methods=['POST'])
@login_required
def api_upload_mtbf_excel(target_name):
    try:
        page_key = (request.form.get('page_key') or 'mtbf').strip().lower()
        if page_key != 'mtbf':
            return jsonify({'success': False, 'message': 'Unsupported page key.'}), 400
        upload = request.files.get('file')
        if not upload or not upload.filename:
            return jsonify({'success': False, 'message': 'Please choose an Excel file to upload.'}), 400
        original_name = secure_filename(upload.filename) or 'mtbf.xlsx'
        ext = os.path.splitext(original_name)[1].lower()
        if ext not in ('.xlsx', '.xlsm'):
            return jsonify({'success': False, 'message': 'Please upload an .xlsx or .xlsm file.'}), 400

        upload_dir = _managed_mtbf_upload_dir(target_name)
        timestamp = _dt.utcnow().strftime('%Y%m%d_%H%M%S')
        stored_name = f'{_safe_target_slug(target_name)}_{timestamp}_{original_name}'
        stored_path = os.path.join(upload_dir, stored_name)
        upload.save(stored_path)

        from dashboard_common import get_bu_for_target
        is_compute_mtbf = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
        selected_sheet, missing = _validate_mtbf_workbook(stored_path, is_compute_mtbf)
        if not selected_sheet:
            try:
                os.remove(stored_path)
            except Exception:
                pass
            return jsonify({
                'success': False,
                'message': 'Uploaded file is missing required column(s): ' + ', '.join(missing)
            }), 400

        _, sheet_names = _read_excel_sheet_names(stored_path)
        saved = _update_target_excel_config(target_name, page_key, {
            'excel_path': stored_path,
            'sheet_name': selected_sheet,
            'managed_upload': True,
            'original_filename': original_name,
            'stored_filename': stored_name,
        })
        return jsonify({
            'success': True,
            'message': f'Uploaded {original_name}. Created MTBF table from sheet {selected_sheet}.',
            'config': saved,
            'sheet_names': sheet_names,
            'selected_sheet': selected_sheet,
        })
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500


@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/create_blank", methods=['POST'])
@login_required
def api_create_blank_mtbf_excel(target_name):
    try:
        from dashboard_common import get_bu_for_target
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        is_compute_mtbf = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
        upload_dir = _managed_mtbf_upload_dir(target_name)
        stored_name = f'{_safe_target_slug(target_name)}_MTBF.xlsx'
        stored_path = os.path.join(upload_dir, stored_name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'MTBF'
        headers = [
            'Target(s)', 'Product Line(s)', 'Device(s)', 'Device Count', 'Week',
            'Meta-ID', 'Build(s) Full ID', 'Tested Hours', 'Reduction %', 'Total Crashes',
            'MTBF', 'Product MTBF', 'QC MTBF', 'MTBF Details', 'Notes'
        ]
        ws.append(headers)
        fill = PatternFill('solid', fgColor='D9EAF7')
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
        ws.freeze_panes = 'A2'
        wb.save(stored_path)

        selected_sheet, missing = _validate_mtbf_workbook(stored_path, is_compute_mtbf)
        if not selected_sheet:
            return jsonify({'success': False, 'message': 'Blank workbook validation failed: ' + ', '.join(missing)}), 500
        saved = _update_target_excel_config(target_name, 'mtbf', {
            'excel_path': stored_path,
            'sheet_name': selected_sheet,
            'managed_upload': True,
            'original_filename': stored_name,
            'stored_filename': stored_name,
        })
        return jsonify({
            'success': True,
            'message': 'Blank MTBF table created. Use Add Build or Edit Table to enter device and MTBF details.',
            'config': saved,
            'selected_sheet': selected_sheet,
        })
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500


@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/sheets", methods=['POST'])
@login_required
def api_excel_sheet_names(target_name):


    try:
        payload = request.get_json(force=True) or {}
        excel_path = payload.get('excel_path')
        _, sheets = _read_excel_sheet_names(excel_path)
        return jsonify({'success': True, 'sheets': sheets})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc), 'sheets': []}), 400


@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/preview", methods=['POST'])
@login_required
def api_excel_sheet_preview(target_name):
    try:
        payload = request.get_json(force=True) or {}
        excel_path = payload.get('excel_path')
        sheet_name = payload.get('sheet_name')
        _, columns, rows, total_rows = _read_excel_sheet_preview(excel_path, sheet_name)
        return jsonify({'success': True, 'columns': columns, 'rows': rows, 'total_rows': total_rows})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc), 'columns': [], 'rows': []}), 400






@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/config/<string:page_key>", methods=['POST'])
@login_required
def api_save_excel_page_config(target_name, page_key):
    if page_key != 'mtbf':
        return jsonify({'success': False, 'message': 'Unsupported page key.'}), 400
    try:
        payload = request.get_json(force=True) or {}
        excel_path = _normalize_excel_path(payload.get('excel_path'))
        sheet_name = str(payload.get('sheet_name') or '')
        saved = _update_target_excel_config(target_name, page_key, {
            'excel_path': excel_path,
            'sheet_name': sheet_name,
        })
        return jsonify({'success': True, 'config': saved})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400



# -- pending build rows waiting for Excel unlock --
_pending_builds = {}   # key: target_name ? list of row dicts

def _get_excel_lock_info(path):
    """Return (is_locked, locked_by) for an Excel file."""
    import os, struct
    lock_path = path + '.lock'
    # openpyxl lock file approach
    owner_path = os.path.join(os.path.dirname(path), '~$' + os.path.basename(path))
    if os.path.exists(owner_path):
        try:
            with open(owner_path, 'rb') as f:
                data = f.read()
            # username starts at byte 8, null-terminated UTF-16
            raw = data[8:]
            name = raw.split(b'\x00\x00')[0].decode('utf-16-le', errors='ignore').strip()
            return True, name or 'another user'
        except Exception:
            return True, 'another user'
    return False, None

def _excel_cell_value_from_string(value):
    """Convert browser form values to Excel-friendly cell values where obvious."""
    if value is None:
        return ''
    if isinstance(value, (int, float, date, datetime)):
        return value
    s = str(value).strip()
    if s == '':
        return ''
    try:
        if s.isdigit() or (s.startswith('-') and s[1:].isdigit()):
            return int(s)
        if any(ch in s for ch in ('.', 'e', 'E')):
            return float(s)
    except Exception:
        return s
    return s


def _append_build_to_excel(excel_path, sheet_name, row_data, hdrs):
    """
    Add a new row to the Excel sheet by:
    1. Reading ALL existing data rows (skipping blanks)
    2. Building the new row mapped to every header column
    3. Clearing the sheet and rewriting header + all rows + new row
    This avoids empty row gaps and column mismatches.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from copy import copy

    path = _normalize_excel_path(excel_path)
    wb   = openpyxl.load_workbook(path)
    ws   = wb[sheet_name]

    # -- 1. Read header row (row 1) --
    max_col = ws.max_column
    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, max_col + 1)]
    # Ensure legacy configured MTBF sheets can persist the new Hours Reduction value.
    if not any('reduction' in h.lower() and ('%' in h or 'percent' in h.lower()) for h in headers):
        max_col += 1
        ws.cell(1, max_col, 'Reduction %')
        headers.append('Reduction %')

    # -- 2. Read all existing data rows (skip fully-blank rows) --
    existing_rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if any(v not in (None, '', 0) for v in row_vals):
            existing_rows.append(row_vals)

    # -- 3. Map new row_data to header columns --
    def _map_val(header, row_data):
        hl = header.lower().strip()
        # Target(s)
        if 'target' in hl:
            return row_data.get('target', '')
                        # Product Line(s) (avoid Product MTBF columns)
        if 'product line' in hl or ('product' in hl and 'mtbf' not in hl):
            return row_data.get('product', '')


        # Device details
        if 'device count' in hl or ('device' in hl and 'count' in hl):
            return row_data.get('device_count', '')
        if 'device' in hl:
            return row_data.get('devices', '')
        # Build(s).1 � full build string (second build column)
        if hl in ('build(s).1', 'build.1', 'full build', 'build id') or ('build' in hl and 'full' in hl):
            return row_data.get('build_full', row_data.get('build', ''))
        # Build(s) � short META id (first build column)
        if 'meta' in hl:
            return row_data.get('build', '')
        if 'build' in hl:
            return row_data.get('build', '')
        # Hours / reduction percent
        if 'reduction' in hl and ('%' in hl or 'percent' in hl):
            return row_data.get('reduction_percent', '')
        if 'hour' in hl:
            return row_data.get('hours', '')
        # Crashes
        if 'crash' in hl:
            return row_data.get('crashes', '')
        # MTBF
        if 'mtbf' in hl:
            return row_data.get('mtbf', '')
                # Week
        if 'week' in hl:
            return row_data.get('week', '')

        # MTBF details / notes
        if 'detail' in hl:
            return row_data.get('mtbf_details', '')
        if 'note' in hl:
            return row_data.get('notes', '') or row_data.get('mtbf_details', '')
        return ''


    new_row = [_map_val(h, row_data) for h in headers]

    # -- 4. Save header row styles --
    hdr_styles = []
    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        hdr_styles.append({
            'font':      copy(cell.font),
            'fill':      copy(cell.fill),
            'alignment': copy(cell.alignment),
            'border':    copy(cell.border),
        })

        # -- 5. Clear sheet and rewrite --
    ws.delete_rows(1, ws.max_row)

    # Write header
    for c, (val, style) in enumerate(zip(headers, hdr_styles), 1):
        cell = ws.cell(1, c, val)
        cell.font      = style['font']
        cell.fill      = style['fill']
        cell.alignment = style['alignment']
        cell.border    = style['border']

    # Write existing data rows
    for ri, row_vals in enumerate(existing_rows, 2):
        for c, val in enumerate(row_vals, 1):
            ws.cell(ri, c, val)

    # Write new row at end
    new_r = len(existing_rows) + 2
    for c, val in enumerate(new_row, 1):
        if val not in (None, ''):
            ws.cell(new_r, c, val)

    wb.save(path)
    return True


def _update_excel_row(excel_path, sheet_name, row_number, values_by_header):
    """Update an existing Excel data row using header names from row 1."""
    import openpyxl

    path = _normalize_excel_path(excel_path)
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found.')
    ws = wb[sheet_name]

    row_number = int(row_number or 0)
    if row_number < 2 or row_number > ws.max_row:
        raise ValueError(f'Invalid Excel row number: {row_number}')

        values_by_header = values_by_header or {}
    normalized_values = {
        str(k or '').strip().lower(): v
        for k, v in values_by_header.items()
    }

    updated = 0
    header_seen = {}
    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or '').strip()
        if not header:
            continue
        count = header_seen.get(header, 0)
        header_seen[header] = count + 1
        display_header = header if count == 0 else f'{header}.{count}'
        if display_header in values_by_header:
            new_val = values_by_header[display_header]
        elif display_header.lower() in normalized_values:
            new_val = normalized_values[display_header.lower()]
        elif header in values_by_header:
            new_val = values_by_header[header]
        elif header.lower() in normalized_values:
            new_val = normalized_values[header.lower()]
        else:
            continue
        ws.cell(row_number, col_idx).value = _excel_cell_value_from_string(new_val)
        updated += 1

    if updated == 0:
        raise ValueError('No matching Excel columns found to update.')

    wb.save(path)
    return updated


@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/add_build", methods=['POST'])
@login_required
def api_excel_add_build(target_name):
    try:
        payload    = request.get_json(force=True) or {}
        from dashboard_common import get_bu_for_target
        is_compute = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
        view_name = _mtbf_json_view_name(payload.get('view') or payload.get('sheet_name') or request.args.get('sheet') or ('Glymur' if is_compute else 'MTBF')) if is_compute else 'MTBF'
        row = _mtbf_json_row_from_payload(payload)
        if not row.get('build'):
            return jsonify({'success': False, 'message': 'Build is required.'}), 400
        data = _load_mtbf_json_payload(target_name, view_name)
        rows = data.get('rows') or []
        rows.append(row)
        data['rows'] = rows
        _save_mtbf_json_payload(target_name, view_name, data)
        return jsonify({'success': True, 'message': f'Build saved to {view_name} JSON.', 'storage': 'json', 'view': view_name})

        build      = str(payload.get('build') or '').strip()

        if not build:
            return jsonify({'success': False, 'message': 'Build ID is required.'}), 400
        target     = str(payload.get('target') or '').strip()
        product    = str(payload.get('product') or '').strip()
        build_full = str(payload.get('build_full') or build).strip()
        devices    = str(payload.get('devices') or '').strip()
        device_count = payload.get('device_count')
        mtbf_details = str(payload.get('mtbf_details') or '').strip()
        hours      = payload.get('hours')
        reduction_percent = payload.get('reduction_percent')

        crashes    = payload.get('crashes')
        mtbf       = payload.get('mtbf')
        week       = payload.get('week')
        row_data   = {
            'target':      target,
            'product':     product,
                        'build':       build,
            'build_full':  build_full,
            'devices':     devices,
            'device_count': int(float(device_count)) if device_count not in (None,'') else '',
            'mtbf_details': mtbf_details,
            'notes':       mtbf_details,
            'hours':       float(hours)        if hours   not in (None,'') else '',
            'reduction_percent': float(reduction_percent) if reduction_percent not in (None,'') else '',

            'crashes':     int(float(crashes)) if crashes not in (None,'') else '',
            'mtbf':        round(float(mtbf),2) if mtbf   not in (None,'') else '',
            'week':        str(week) if week else '',
        }
        cfg = (_get_target_excel_config(target_name) or {}).get('mtbf', {})
        excel_path = cfg.get('excel_path', '')
        sheet_name = cfg.get('sheet_name', '')
        if not excel_path or not sheet_name:
            return jsonify({'success': False, 'message': 'Excel not configured.'}), 400
        path = _normalize_excel_path(excel_path)
        # check lock before opening
        is_locked, locked_by = _get_excel_lock_info(path)
        if is_locked:
            _pending_builds.setdefault(target_name, []).append({'row_data': row_data})
            return jsonify({'success': False, 'locked': True, 'locked_by': locked_by,
                            'message': f'Excel locked by {locked_by}. Saved locally.'})
                # write to Excel � full table replace
        _append_build_to_excel(excel_path, sheet_name, row_data, [])
        return jsonify({'success': True, 'message': 'Build added to Excel.'})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500


@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/update_row", methods=['POST'])
@login_required
def api_excel_update_row(target_name):
    try:
        payload = request.get_json(force=True) or {}
        row_number = int(payload.get('row_number') or 0)
        values = payload.get('values') or {}
        if row_number < 2:
            return jsonify({'success': False, 'message': 'Invalid row number.'}), 400
        if not isinstance(values, dict) or not values:
            return jsonify({'success': False, 'message': 'No row values provided.'}), 400

        cfg = (_get_target_excel_config(target_name) or {}).get('mtbf', {})
        excel_path = cfg.get('excel_path', '')
        sheet_name = cfg.get('sheet_name', '')
        if not excel_path or not sheet_name:
            return jsonify({'success': False, 'message': 'Excel not configured.'}), 400

        path = _normalize_excel_path(excel_path)
        is_locked, locked_by = _get_excel_lock_info(path)
        if is_locked:
            return jsonify({
                'success': False,
                'locked': True,
                'locked_by': locked_by,
                'message': f'Excel locked by {locked_by}. Please close it and retry.'
            })

        updated = _update_excel_row(excel_path, sheet_name, row_number, values)
        return jsonify({'success': True, 'message': f'Updated row {row_number} in Excel.', 'updated': updated})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500


# -- Full table read (flat, no merge info) for the edit page ------------------
@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/full_table", methods=['GET'])
@login_required
def api_excel_full_table(target_name):
    """Return every data row as a flat list of string values for the edit page."""
    try:
        from dashboard_common import get_bu_for_target
        is_compute = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
        sheet_name = _mtbf_json_view_name(request.args.get('sheet') or ('Glymur' if is_compute else 'MTBF')) if is_compute else 'MTBF'
        data = _load_mtbf_json_payload(target_name, sheet_name)
        rows = []
        for idx, r in enumerate(data.get('rows') or [], start=1):
            if is_compute:
                values = [
                    str(r.get('build') or ''),
                    str(r.get('build_full') or r.get('full_build') or r.get('build') or ''),
                    str(r.get('date') or ''),
                    str(r.get('hours') or ''),
                    str(r.get('total_crashes') or ''),
                    str(r.get('qc_crashes') or ''),
                    str(r.get('product_mtbf') or ''),
                    str(r.get('qc_mtbf') or ''),
                    str(r.get('comments') or ''),
                ]
            else:
                values = [
                    str(r.get('build') or ''),
                    str(r.get('build_full') or r.get('full_build') or r.get('build') or ''),
                    str(r.get('date') or ''),
                    str(r.get('hours') or ''),
                    str(r.get('total_crashes') or ''),
                    str(r.get('mtbf')  or ''),
                    str(r.get('comments') or ''),
                ]
            rows.append({'excel_row': idx + 1, 'values': values})
        return jsonify({'success': True, 'headers': _mtbf_json_headers(is_compute), 'rows': rows, 'sheet_name': sheet_name, 'storage': 'json'})

        cfg        = (_get_target_excel_config(target_name) or {}).get('mtbf', {})

        excel_path = cfg.get('excel_path', '')
        sheet_name = (request.args.get('sheet') or cfg.get('sheet_name', '')).strip()
        if not excel_path:
            return jsonify({'success': False, 'message': 'Excel not configured.'}), 400
        import openpyxl
        from datetime import datetime as _dtt, date as _ddate
        path = _normalize_excel_path(excel_path)
        if not os.path.exists(path):
            return jsonify({'success': False, 'message': f'File not found: {path}'}), 404
        wb = openpyxl.load_workbook(path, data_only=True)
        # Resolve sheet robustly: requested -> config -> single/first sheet
        actual_sheet = sheet_name
        if actual_sheet and actual_sheet not in wb.sheetnames:
            actual_sheet = next(
                (s for s in wb.sheetnames if s.strip().lower() == sheet_name.strip().lower()),
                None
            )
        if not actual_sheet and cfg.get('sheet_name'):
            cfg_sheet = str(cfg.get('sheet_name') or '').strip()
            actual_sheet = next(
                (s for s in wb.sheetnames if s.strip().lower() == cfg_sheet.lower()),
                None
            )
        if not actual_sheet and len(wb.sheetnames) == 1:
            actual_sheet = wb.sheetnames[0]
        if not actual_sheet and wb.sheetnames:
            actual_sheet = wb.sheetnames[0]
        if not actual_sheet:
            return jsonify({'success': False, 'message': 'No sheets found in workbook.'}), 404
        sheet_name = actual_sheet
        ws = wb[sheet_name]
        # forward-fill merged cells for flat edit view
        merge_map = {}
        for mr in list(ws.merged_cells.ranges):
            val = ws.cell(mr.min_row, mr.min_col).value
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merge_map[(r, c)] = val
        def _cv(r, c):
            v = merge_map.get((r, c), ws.cell(r, c).value)
            if isinstance(v, _dtt): return v.strftime('%Y-%m-%d')
            if isinstance(v, _ddate): return v.strftime('%Y-%m-%d')
            return '' if v is None else str(v).strip()
        headers = [_cv(1, c) for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            vals = [_cv(r, c) for c in range(1, ws.max_column + 1)]
            if any(v.strip() for v in vals):
                rows.append({'excel_row': r, 'values': vals})
        return jsonify({'success': True, 'headers': headers, 'rows': rows,
                        'sheet_name': sheet_name, 'excel_path': excel_path})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500


# -- Save entire table back to Excel ------------------------------------------
@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/save_table", methods=['POST'])
@login_required
def api_excel_save_table(target_name):
    """Overwrite every data row in the sheet with the values sent from the edit page.
    Body: {sheet_name, rows: [{excel_row, values: [...]}]}
    """
    try:
        payload    = request.get_json(force=True) or {}
        sheet_name = str(payload.get('sheet_name') or '').strip()
        rows_data  = payload.get('rows') or []
        from dashboard_common import get_bu_for_target
        is_compute = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
        view_name = _mtbf_json_view_name(sheet_name or ('Glymur' if is_compute else 'MTBF')) if is_compute else 'MTBF'
        parsed_rows = []
        for item in rows_data:
            vals = list(item.get('values') or []) if isinstance(item, dict) else []
            if is_compute:
                parsed = {
                    'id': str(item.get('id') or '').strip() if isinstance(item, dict) else '',
                    'build':         str(vals[0] if len(vals) > 0 else '').strip(),
                    'build_full':    str(vals[1] if len(vals) > 1 else (vals[0] if len(vals) > 0 else '')).strip(),
                    'date':          str(vals[2] if len(vals) > 2 else '').strip(),
                    'hours':         _num_or_blank(vals[3] if len(vals) > 3 else ''),
                    'total_crashes': _num_or_blank(vals[4] if len(vals) > 4 else '', integer=True),
                    'qc_crashes':    _num_or_blank(vals[5] if len(vals) > 5 else '', integer=True),
                    'product_mtbf':  _num_or_blank(vals[6] if len(vals) > 6 else ''),
                    'qc_mtbf':       _num_or_blank(vals[7] if len(vals) > 7 else ''),
                    'comments':      str(vals[8] if len(vals) > 8 else '').strip(),
                }
            else:
                parsed = {
                    'id': str(item.get('id') or '').strip() if isinstance(item, dict) else '',
                    'build':         str(vals[0] if len(vals) > 0 else '').strip(),
                    'build_full':    str(vals[1] if len(vals) > 1 else (vals[0] if len(vals) > 0 else '')).strip(),
                    'date':          str(vals[2] if len(vals) > 2 else '').strip(),
                    'hours':         _num_or_blank(vals[3] if len(vals) > 3 else ''),
                    'total_crashes': _num_or_blank(vals[4] if len(vals) > 4 else '', integer=True),
                    'mtbf':          _num_or_blank(vals[5] if len(vals) > 5 else ''),
                    'comments':      str(vals[6] if len(vals) > 6 else '').strip(),
                }
            parsed_rows.append(parsed)
        parsed_rows = [r for r in parsed_rows if any(str(v).strip() for v in r.values())]
        existing_payload = _load_mtbf_json_payload(target_name, view_name)
        existing_payload['rows'] = parsed_rows
        existing_payload['migrated_from_excel'] = True
        _save_mtbf_json_payload(target_name, view_name, existing_payload)
        return jsonify({'success': True, 'message': f'Saved {len(parsed_rows)} row(s) to {view_name} JSON.', 'updated': len(parsed_rows), 'storage': 'json'})


        cfg        = (_get_target_excel_config(target_name) or {}).get('mtbf', {})

        excel_path = cfg.get('excel_path', '')
        if not excel_path:
            return jsonify({'success': False, 'message': 'Excel not configured.'}), 400
        path = _normalize_excel_path(excel_path)
        is_locked, locked_by = _get_excel_lock_info(path)
        if is_locked:
            return jsonify({'success': False, 'locked': True, 'locked_by': locked_by,
                            'message': f'Excel locked by {locked_by}. Please close it and retry.'})
        import openpyxl
        from openpyxl.cell import MergedCell as _MergedCell
        wb = openpyxl.load_workbook(path)
        # Resolve sheet robustly: payload -> config -> single/first sheet
        actual_sheet = sheet_name
        if actual_sheet and actual_sheet not in wb.sheetnames:
            actual_sheet = next(
                (s for s in wb.sheetnames if s.strip().lower() == sheet_name.strip().lower()),
                None
            )
        if not actual_sheet and cfg.get('sheet_name'):
            cfg_sheet = str(cfg.get('sheet_name') or '').strip()
            actual_sheet = next(
                (s for s in wb.sheetnames if s.strip().lower() == cfg_sheet.lower()),
                None
            )
        if not actual_sheet and len(wb.sheetnames) == 1:
            actual_sheet = wb.sheetnames[0]
        if not actual_sheet and wb.sheetnames:
            actual_sheet = wb.sheetnames[0]
        if not actual_sheet:
            return jsonify({'success': False, 'message': 'No sheets found in workbook.'}), 404
        sheet_name = actual_sheet
        ws = wb[sheet_name]

        # Unmerge ALL merged ranges so every cell becomes writable.
        # We must iterate over a copy because unmerge_cells mutates the collection.
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))

        max_col = ws.max_column
        max_col = ws.max_column
        start_row = 2
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, _MergedCell):
                    cell.value = None
        updated = 0
        for row_offset, row_info in enumerate(rows_data, start=0):
            values = row_info.get('values') or []
            excel_row = start_row + row_offset
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(excel_row, col_idx)
                if isinstance(cell, _MergedCell):
                    continue
                val = values[col_idx - 1] if (col_idx - 1) < len(values) else ''
                cell.value = _excel_cell_value_from_string(val)
            updated += 1
        wb.save(path)
        return jsonify({'success': True, 'message': f'Saved {updated} row(s) to Excel.',
                        'updated': updated, 'sheet_name': sheet_name, 'reload': True})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500

# -- Full-table edit page ------------------------------------------------------
@dashboard_bp.route("/dashboard/<string:target_name>/mtbf-excel/edit")
@login_required
def target_mtbf_excel_edit_page(target_name):
    """Dedicated page to edit the entire JSON MTBF build table inline."""

    cfg        = (_get_target_excel_config(target_name) or {}).get('mtbf', {})

    excel_path = cfg.get('excel_path', '')
    sheet_name = (request.args.get('sheet') or cfg.get('sheet_name', '')).strip()
    ctx = _build_sidebar_context(target_name, active_section='mtbf-excel')
    ctx['target_display_name'] = get_display_name_for_target(target_name).upper()
    return render_template(
        'target_mtbf_excel_edit.html',
        excel_path=excel_path,
        selected_sheet=sheet_name,
        **ctx,
    )


@dashboard_bp.route("/api/dashboard/<string:target_name>/excel/retry_sync", methods=['POST'])
@login_required
def api_excel_retry_sync(target_name):
    try:
        pending = _pending_builds.get(target_name, [])
        if not pending:
            return jsonify({'success': True, 'message': 'Nothing to sync.'})
        cfg        = (_get_target_excel_config(target_name) or {}).get('mtbf', {})
        excel_path = cfg.get('excel_path', '')
        sheet_name = cfg.get('sheet_name', '')
        path       = _normalize_excel_path(excel_path)
        is_locked, locked_by = _get_excel_lock_info(path)
        if is_locked:
            return jsonify({
                'success':   False,
                'locked':    True,
                'locked_by': locked_by,
                'message':   f'Still locked by {locked_by}.'
            })
                # write all pending rows
        for item in pending:
            _append_build_to_excel(excel_path, sheet_name, item['row_data'], [])
            _pending_builds[target_name] = []
            return jsonify({'success': True, 'message': f'Synced {len(pending)} row(s) to Excel.'})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500



@dashboard_bp.route("/api/browse_files", methods=['POST'])
@login_required
def api_browse_files():
    """Windows-style file browser: drives, folders, Excel files."""
    import os, string
    try:
        payload = request.get_json(force=True) or {}
        path    = str(payload.get('path') or '').strip().strip('"').strip("'")

        # -- Special root: list all drives + quick-access --
        if not path or path in ('/', 'root', 'ROOT'):
            items = []
            # Windows drives
            for drive in string.ascii_uppercase:
                dp = f'{drive}:\\'
                if os.path.exists(dp):
                    try:
                        import shutil
                        total, used, free = shutil.disk_usage(dp)
                        label = f'{drive}:  ({_fmt_size(free)} free)'
                    except Exception:
                        label = f'{drive}:'
                    items.append({'name': label, 'path': dp, 'type': 'drive'})
            # Network shares shortcut
            items.append({'name': 'Network (\\\\)', 'path': '\\\\', 'type': 'network'})
            # Quick access
            for name, p in [
                ('Desktop',   os.path.expanduser('~\\Desktop')),
                ('Documents', os.path.expanduser('~\\Documents')),
                ('Downloads', os.path.expanduser('~\\Downloads')),
            ]:
                if os.path.exists(p):
                    items.append({'name': f'?? {name}', 'path': p, 'type': 'quick'})
            return jsonify({'success': True, 'path': 'This PC', 'items': items, 'is_root': True})

        # -- Network path: list shares or UNC path --
        if path.startswith('\\\\'):
            path = os.path.normpath(path)

        # -- Normal path --
        path = os.path.normpath(path)

        # If path is a file, go to its parent
        if os.path.isfile(path):
            path = os.path.dirname(path)

        if not os.path.exists(path):
            parent = os.path.dirname(path)
            if os.path.exists(parent):
                path = parent
            else:
                return jsonify({'success': False, 'message': f'Path not found: {path}'}), 404

        try:
            entries = sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower()))
        except PermissionError:
            return jsonify({'success': False, 'message': f'Access denied: {path}'}), 403

        items = []
        # Parent navigation
        parent = os.path.dirname(path)
        if parent and parent != path:
            # Check if parent is a drive root ? go to 'This PC'
            if os.path.splitdrive(path)[1] in ('\\', '/'):
                items.append({'name': '? This PC', 'path': 'root', 'type': 'up'})
            else:
                items.append({'name': f'? {os.path.basename(parent) or parent}', 'path': parent, 'type': 'up'})

        for e in entries:
            try:
                if e.name.startswith('$') or e.name.startswith('.'):
                    continue  # skip system/hidden
                if e.is_dir(follow_symlinks=False):
                    items.append({'name': e.name, 'path': e.path, 'type': 'dir'})
                elif e.name.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                    try:
                        sz = _fmt_size(e.stat().st_size)
                    except Exception:
                        sz = ''
                    items.append({'name': e.name, 'path': e.path, 'type': 'file', 'size': sz})
            except (PermissionError, OSError):
                continue

        return jsonify({'success': True, 'path': path, 'items': items, 'is_root': False})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500


def _fmt_size(b):
    for u in ['B','KB','MB','GB']:
        if b < 1024: return f'{b:.0f} {u}'
        b /= 1024
    return f'{b:.1f} GB'








@dashboard_bp.route("/device-summary/<string:target_name>")
@login_required
def device_summary_page(target_name):
    import device_summary_service as ds_svc
    # Ensure each BU/target has a managed Excel workbook for direct Add/Edit/Remove.
    # If a user configured a network Excel, this keeps using it; otherwise it creates:
        # \\sphere\pdtqipl_internal\PDTBuddy\managed_excel\<BU>\<TARGET>\Devices\<TARGET>_device_summary.xlsx

    ds_svc.get_or_create_device_excel_config(target_name)
    project_filter = (request.args.get('project') or 'All').strip() or 'All'
    data = ds_svc.load_page_data(target_name, project_filter=project_filter)
    ctx  = _build_sidebar_context(target_name, active_section='device-summary')

    ctx['target_display_name'] = get_display_name_for_target(target_name).upper()
    return render_template('device_summary_page.html', **data, **ctx)




def normalize_dashboard_section(section):
    section = (section or 'dashboard').strip().lower()
    mapping = {
        'overview'   : 'dashboard',
        'home'       : 'dashboard',
        'mtbf'       : 'mtbf-table',
        'mtbftable'  : 'mtbf-table',
        'mtbftrend'  : 'mtbf-trend',
        'weekly'     : 'weekly-report',
    }
    return mapping.get(section, section)


@dashboard_bp.route("/dashboard/<string:target_name>/<string:section>", methods=["GET", "POST"])
@login_required
def dashboard(target_name, section="dashboard"):


    perf_total_start = _perf_now()
    perf_marks = []

    conn = None
    cursor = None
    section = normalize_dashboard_section(section)
    toggle_mode = request.args.get("toggle_mode", "CRM")
    pdt_type = request.args.get("pdt_type", "SWPDT")
    compute_bu_flag = request.args.get("compute_bu", "0")
    compute_bu = compute_bu_flag in ("1", "true", "True", "TRUE")

    try:
        if VALID_TARGETS and target_name not in VALID_TARGETS:
            return f"Invalid target '{target_name}'", 404

        # Always fetch latest display name from DB so UI reflects DB changes
        target_display_name = get_display_name_for_target(target_name).upper()
        perf_marks.append(("target_display", _perf_elapsed_ms(perf_total_start)))

        conn = get_mysql_connection_db()
        cursor = conn.cursor(dictionary=True)
        perf_marks.append(("db_connect", _perf_elapsed_ms(perf_total_start)))


        # 1) KPI / Glance data
        tables = {
            "u": fq_table_for_target(target_name, "unique_crs"),
            "o": fq_table_for_target(target_name, "openjiras"),
            "j": fq_table_for_target(target_name, "jiras"),
            "c": fq_table_for_target(target_name, "closed_jiras"),
        }

        def table_exists(fq_name):
            """

            Check if a fully qualified table exists.
            fq_name should look like:
            `schema`.`table` or schema.table
            """
            name = fq_name.replace("`", "")
            try:
                schema, table = name.split(".", 1)
            except ValueError:
                # Not fully qualified; assume it exists and let the query fail if not
                return True

            cursor.execute(
                """
                SELECT 1
                FROM information_schema.tables
                WHERE table_schema = %s AND table_name = %s
                LIMIT 1
                """,
                (schema, table),
                        )
            return cursor.fetchone() is not None

        def get_count(tbl_name):
            if not table_exists(tbl_name):

                logger.info(f"[WARN] get_count: table missing {tbl_name} (target={target_name})")
                return 0
            cursor.execute(f"SELECT COUNT(*) AS cnt FROM {tbl_name}")
            row = cursor.fetchone() or {}
            return int(row.get("cnt", 0) or 0)

        mapped_jiras = get_count(tables["j"])
        open_jiras = get_count(tables["o"])
        closed_jiras = get_count(tables["c"])
        total_crs = get_count(tables["u"])

        glance = {
            "mapped_jiras": mapped_jiras,
            "open_jiras": open_jiras,
            "closed_jiras": closed_jiras,
            "total_jiras": mapped_jiras + open_jiras + closed_jiras,
            "total_crs": total_crs,
        }
        perf_marks.append(("kpi_counts", _perf_elapsed_ms(perf_total_start)))

        dash_meta = get_dashboard_meta_for_target(target_name)

        raw_dt = dash_meta.get("dashboard_latest_update") or dash_meta.get("unique_cr_last_update")

        if isinstance(raw_dt, datetime):
            target_update = raw_dt.strftime("%Y-%m-%d %H:%M:%S")
        elif raw_dt:
            target_update = str(raw_dt)
        else:
            target_update = "N/A"

        milestones = {
            "ES": dash_meta.get("ES") or "TBD",
            "FC": dash_meta.get("FC") or "TBD",
            "CS": dash_meta.get("CS") or "TBD",
        }
        perf_marks.append(("dashboard_meta", _perf_elapsed_ms(perf_total_start)))

        # 2) CR Age analysis � separate Built / Undisposed charts

        # Check which optional columns exist before querying
        cursor.execute(f"SHOW COLUMNS FROM {tables['u']}")
        _u_cols = {row['Field'] for row in (cursor.fetchall() or [])}
        _sel_cr_raw = "cr AS cr_raw" if 'cr' in _u_cols else "mapped_cr AS cr_raw"
        _sel_cr_subsystem = "cr_subsystem" if 'cr_subsystem' in _u_cols else "NULL AS cr_subsystem"
        _sel_cr_functionality = "cr_functionality" if 'cr_functionality' in _u_cols else "NULL AS cr_functionality"

        cursor.execute(f"""
            SELECT
                mapped_cr AS cr_id,
                {_sel_cr_raw},
                cr_title,
                cr_status,
                cr_area,
                {_sel_cr_subsystem},
                {_sel_cr_functionality},
                cr_category,
                CAST(NULLIF(cr_age, '') AS UNSIGNED) AS cr_age
            FROM {tables['u']}
            WHERE LOWER(TRIM(cr_category)) = 'undisposed'
            ORDER BY cr_area, cr_age DESC
        """)
        cr_rows = cursor.fetchall() or []

        def bucket_counts(rows, age_key="avg_cr_age"):
            b5_20 = b20_40 = bover_40 = 0
            for r in rows:
                age = float(r[age_key] or 0)
                cnt = int(r["cr_count"] or 0)
                if 5 <= age < 20:
                    b5_20 += cnt
                elif 20 <= age < 40:
                    b20_40 += cnt
                elif age >= 40:
                    bover_40 += cnt
            return {"5_20": b5_20, "20_40": b20_40, "over_40": bover_40}

        # Built CRs (area aggregate)
        built_query = build_cr_area_age_query(tables["u"], "Built")
        cursor.execute(built_query)
        built_rows = cursor.fetchall() or []
        cr_age_chart_built = {
            "categories": [r["cr_area"] or "Unknown" for r in built_rows],
            "cr_count": [int(r["cr_count"] or 0) for r in built_rows],
            "avg_cr_age": [round(float(r["avg_cr_age"] or 0), 1) for r in built_rows],
        }

        # Undisposed (open) CRs (area aggregate)
        undisp_query = build_cr_area_age_query(tables["u"], "Undisposed")
        cursor.execute(undisp_query)
        undisp_rows = cursor.fetchall() or []
        cr_age_chart_undisposed = {
            "categories": [r["cr_area"] or "Unknown" for r in undisp_rows],
            "cr_count": [int(r["cr_count"] or 0) for r in undisp_rows],
            "avg_cr_age": [round(float(r["avg_cr_age"] or 0), 1) for r in undisp_rows],
        }

        cr_age_buckets = bucket_counts(undisp_rows)

        # Detailed rows are only used for client-side modal previews.
        # Keep them bounded so large targets do not make the initial dashboard HTML huge.
        _DASHBOARD_ROW_PREVIEW_LIMIT = 500
        cursor.execute(fetch_undisposed_crs_in_age_band(tables["u"], 5, 20) + f"\nLIMIT {_DASHBOARD_ROW_PREVIEW_LIMIT}")
        cr_age_list_5_20 = cursor.fetchall() or []

        cursor.execute(fetch_undisposed_crs_in_age_band(tables["u"], 20, 40) + f"\nLIMIT {_DASHBOARD_ROW_PREVIEW_LIMIT}")
        cr_age_list_20_40 = cursor.fetchall() or []

        cursor.execute(fetch_undisposed_crs_in_age_band(tables["u"], 40, None) + f"\nLIMIT {_DASHBOARD_ROW_PREVIEW_LIMIT}")
        cr_age_list_over_40 = cursor.fetchall() or []

                                                                # Open / Analysis / Other counts for undisposed CRs (for pie)
        undisp_status_counts = fetch_undisposed_status_counts(cursor, tables["u"])

        # Category breakdown from cr_category:
        # Valid = built + undisposed, Invalid = invalid variants, Dup = dup.
        try:
            cursor.execute(f"""
                SELECT
                    SUM(CASE WHEN LOWER(TRIM(cr_category)) IN ('built','undisposed') THEN 1 ELSE 0 END) AS valid_count,
                    SUM(CASE WHEN LOWER(TRIM(cr_category)) IN ('invalid','invalid_dup','nosir') THEN 1 ELSE 0 END) AS invalid_count,
                    SUM(CASE WHEN LOWER(TRIM(cr_category)) = 'dup' THEN 1 ELSE 0 END) AS dup_count
                FROM {tables['u']}
            """)
            _inv_row = cursor.fetchone() or {}
            cr_valid_count   = int(_inv_row.get('valid_count') or 0)
            cr_invalid_count = int(_inv_row.get('invalid_count') or 0)
            cr_dup_count     = int(_inv_row.get('dup_count') or 0)
            glance['cr_valid_count'] = cr_valid_count
            glance['cr_invalid_count'] = cr_invalid_count
            glance['cr_dup_count'] = cr_dup_count
        except Exception:
            cr_valid_count   = 0
            cr_invalid_count = 0
            cr_dup_count     = 0
            glance['cr_valid_count'] = 0
            glance['cr_invalid_count'] = 0
            glance['cr_dup_count'] = 0

        # ALL CR rows (built + undisposed) with cr_status for dynamic checkbox filter
        try:
            cursor.execute(f"""
                SELECT
                    mapped_cr AS cr_id,
                    cr_title,
                    cr_status,
                    cr_area,
                {_sel_cr_subsystem},
                {_sel_cr_functionality},
                    cr_category,
                    CAST(NULLIF(cr_age, '') AS UNSIGNED) AS cr_age
                FROM {tables['u']}
                ORDER BY cr_area, cr_age DESC
                LIMIT 10000
            """)
            all_cr_rows = cursor.fetchall() or []
        except Exception:
            all_cr_rows = cr_rows  # fallback to undisposed only

            perf_marks.append(("cr_age_and_status", _perf_elapsed_ms(perf_total_start)))

        # 3) MTBF / build report data

        schema_name = get_schema_for_target(target_name) or "pdt_stats_mobile"
        is_compute_bu = (schema_name == "pdt_stats_compute")

        # -- CR Title Exclude (Compute only) ------------------------------
        if is_compute_bu:
            _excl = _get_cr_title_exclude(target_name)
            if _excl['enabled'] and _excl['keywords']:
                _kws = [k.lower() for k in _excl['keywords']]
                def _excl_filter(rows):
                    return [r for r in (rows or []) if not any(kw in str(r.get('cr_title') or '').lower() for kw in _kws)]
                cr_rows     = _excl_filter(cr_rows)
                all_cr_rows = _excl_filter(all_cr_rows)

        # SWPDT dataset for the main build table / dashboard
        sw_build_report_data = get_build_report_for_target(
            cursor=cursor,
            target_name=target_name,
            schema_name=schema_name,
            pdt_type="SWPDT",
            toggle_mode=toggle_mode,
        ) or {}
        base_rows = sw_build_report_data.get("rows", [])

        # Separate HWPDT dataset so HW is treated as a different feature
        hw_build_report_data = get_build_report_for_target(
            cursor=cursor,
            target_name=target_name,
            schema_name=schema_name,
            pdt_type="HWPDT",
            toggle_mode=toggle_mode,
        ) or {}
        hw_base_rows = hw_build_report_data.get("rows", [])
        hwpdt_available = bool(hw_build_report_data.get("hwpdt_available"))
        perf_marks.append(("build_report_fetch", _perf_elapsed_ms(perf_total_start)))

        # For compute BU, preserve stored QC/Product MTBF when available.

        # Only fall back to a split if neither value was saved.
        if is_compute_bu:
            for r in base_rows:
                saved_qc = r.get("qc_mtbf")
                saved_product = r.get("product_mtbf")
                row_mtbf = float(r.get("mtbf") or 0.0)

                if saved_qc is None and saved_product is None:
                    split_mtbf = round(row_mtbf / 2.0, 2)
                    r["qc_mtbf"] = split_mtbf
                    r["product_mtbf"] = split_mtbf
                else:
                    if saved_qc is None:
                        r["qc_mtbf"] = round(row_mtbf / 2.0, 2)
                    if saved_product is None:
                        r["product_mtbf"] = round(row_mtbf / 2.0, 2)

                for b in (r.get("builds") or []):
                    saved_b_qc = b.get("qc_mtbf")
                    saved_b_product = b.get("product_mtbf")
                    b_mtbf = float(b.get("mtbf") or 0.0)

                    if saved_b_qc is None and saved_b_product is None:
                        b_split_mtbf = round(b_mtbf / 2.0, 2)
                        b["qc_mtbf"] = b_split_mtbf
                        b["product_mtbf"] = b_split_mtbf
                    else:
                        if saved_b_qc is None:
                            b["qc_mtbf"] = round(b_mtbf / 2.0, 2)
                        if saved_b_product is None:
                            b["product_mtbf"] = round(b_mtbf / 2.0, 2)
        # Keep raw crashes for excluded-count computation
        for r in base_rows:
            if "crashes" in r and "raw_crashes" not in r:
                r["raw_crashes"] = r["crashes"]

        # Ensure meta_builds table exists and get its name
        meta_builds_table_name = ensure_meta_builds_table(cursor, schema_name, target_name)
        meta_builds_table = f"`{schema_name}`.`{meta_builds_table_name}`"
        AGG_BUILD_ID = "__META__"

        # --- Load META-level crash/MTBF overrides (__META__ rows) ---
        cursor.execute(
            f"""
            SELECT meta_id, swpdt_crashes, mtbf, product_mtbf, qc_mtbf
            FROM {meta_builds_table}
            WHERE build_id = %s
              AND pdt_type = %s
              AND is_active = 1
            """,
            (AGG_BUILD_ID, pdt_type),
        )
        override_rows = cursor.fetchall() or []
        meta_override_map = {
            r["meta_id"]: {
                "swpdt_crashes": r.get("swpdt_crashes"),
                "mtbf": r.get("mtbf"),
                "product_mtbf": r.get("product_mtbf"),
                "qc_mtbf": r.get("qc_mtbf"),
            }
            for r in override_rows
            if r.get("meta_id") is not None
        }
        meta_mtbf_map = {
            r["meta_id"]: _round_if_number(r.get("mtbf"), 2)
            for r in override_rows
            if r.get("meta_id") is not None and r.get("mtbf") is not None
        }

        # --- Load per-build selection state ---
        cursor.execute(
            f"""
            SELECT meta_id, build_id, is_selected, is_active
            FROM {meta_builds_table}
            WHERE build_id <> %s
              AND pdt_type = %s
            """,
            (AGG_BUILD_ID, pdt_type),
        )
        sel_rows = cursor.fetchall() or []
        sel_map = {
            (r["meta_id"], r["build_id"]): (r["is_selected"], r["is_active"])
            for r in sel_rows
            if r.get("meta_id") and r.get("build_id")
        }

        # Apply selection/active state to builds and crash overrides to META rows
        for r in base_rows:
            meta_id = r.get("meta_id")
            raw = r.get("raw_crashes") or 0

            # Adjust per-build selection
            builds = r.get("builds") or []
            filtered_builds = []
            for b in builds:
                key = (meta_id, b.get("build_id"))
                sel = sel_map.get(key)
                if sel is not None:
                    is_sel, is_act = sel
                    b["is_selected"] = bool(is_sel)
                    if is_act:
                        filtered_builds.append(b)
                else:
                    filtered_builds.append(b)
            r["builds"] = filtered_builds

            override = meta_override_map.get(meta_id, {})

            # crashes
            if override.get("swpdt_crashes") is not None:
                r["crashes"] = override.get("swpdt_crashes")
            else:
                r["crashes"] = raw

            r["excluded_crashes"] = max(raw - (r["crashes"] or 0), 0)

            # MTBF handling: prefer saved META row MTBF if present
            if meta_id in meta_mtbf_map and meta_mtbf_map[meta_id] is not None:
                r["mtbf"] = meta_mtbf_map[meta_id]

            if is_compute_bu:
                if override.get("qc_mtbf") is not None:
                    r["qc_mtbf"] = _round_if_number(override.get("qc_mtbf"), 2)
                else:
                    r["qc_mtbf"] = _round_if_number(float(r.get("mtbf") or 0.0) / 2.0, 2)

                if override.get("product_mtbf") is not None:
                    r["product_mtbf"] = _round_if_number(override.get("product_mtbf"), 2)
                else:
                    r["product_mtbf"] = _round_if_number(float(r.get("mtbf") or 0.0) / 2.0, 2)
        mtbf_dashboard = build_mtbf_dashboard_payload(base_rows, pdt_type="SWPDT")
        mtbf_series = mtbf_dashboard.get("mtbf_series", [])
        mtbf_build_table = mtbf_dashboard.get("mtbf_build_table", [])
        build_report_rows = base_rows

        hwpdt_dashboard = build_mtbf_dashboard_payload(hw_base_rows, pdt_type="HWPDT")
        hwpdt_build_table = hwpdt_dashboard.get("mtbf_build_table", [])
        perf_marks.append(("mtbf_payloads", _perf_elapsed_ms(perf_total_start)))

        # If HW availability was not detected at source level, infer it from

        # the flattened build table: any non-zero hwpdt_crashes means HWPDT
        # should be available.
        if not hwpdt_available:
            for r in hwpdt_build_table:
                try:
                    if int(r.get("crashes") or 0) > 0:
                        hwpdt_available = True
                        break
                except Exception:
                    continue

        # 4) Glance URLs (blueprint endpoints)
        hw_rows = sum(1 for r in hwpdt_build_table if int(r.get("crashes") or 0) > 0)
        # logger.info("[HWPDT DEBUG] target:")

        def safe_url(endpoint, **values):
            try:

                return url_for(endpoint, **values)
            except Exception:
                return "#"

        mapped_jiras_url = safe_url("dashboard_bp.view_all_jiras", target_name=target_name)
        open_jiras_url = safe_url("dashboard_bp.view_all_open_jiras", target_name=target_name)
        closed_jiras_url = safe_url("dashboard_bp.view_all_closed_jiras", target_name=target_name)
        total_crs_url = safe_url("dashboard_bp.view_all_unique_crs", target_name=target_name)

        # 5) Common context
        active_bu_key = (get_bu_for_target(target_name) or '').upper()
        _tinfo = get_target_info(target_name) or {}
        _sp_name   = str(_tinfo.get("sp_name",   "") or "")
        _chip_name = str(_tinfo.get("chip_name", "") or "")
        _pdt_poc   = str(_tinfo.get("pdt_poc",   "") or "")
        _mailing   = str(_tinfo.get("mailing_list", "") or _tinfo.get("mailing", "") or "")
        base_context = {
            "target_name": target_name,
            "target_display_name": target_display_name,
            "sp_name":   _sp_name,
            "chip_name": _chip_name,
            "pdt_poc":   _pdt_poc,
            "mailing_list": _mailing,
            "glance": glance,
            "active_section": section,
            **_build_bu_shell_context(active_bu_key),

            "toggle_mode": toggle_mode,
            "pdt_type": pdt_type,
            "schema_name": schema_name,
            "is_compute_bu": is_compute_bu,
            "compute_bu": compute_bu,

            "cr_age_chart_built": cr_age_chart_built,
            "cr_age_chart_undisposed": cr_age_chart_undisposed,
            "cr_age_buckets": cr_age_buckets,
            "cr_age_list_5_20": cr_age_list_5_20,
            "cr_age_list_20_40": cr_age_list_20_40,
            "cr_age_list_over_40": cr_age_list_over_40,
            "undisp_status_counts": undisp_status_counts,
            "cr_valid_count": cr_valid_count,
            "cr_dup_count": cr_dup_count,
            "cr_invalid_count": cr_invalid_count,
            "cr_rows": cr_rows,
            "all_cr_rows": all_cr_rows,
            "mapped_jiras_url": mapped_jiras_url,
            "open_jiras_url": open_jiras_url,
            "closed_jiras_url": closed_jiras_url,
            "total_crs_url": total_crs_url,
            "target_update": target_update,
            "milestones": milestones,
            "milestone_phase": build_milestone_phase_context(target_name),
            "hwpdt_available": hwpdt_available,
            "pv_hidden_tabs": _get_pv_hidden_tabs(target_name),
            "unique_cr_path": (_tinfo.get("unique_cr_path") or ""),
        }

                                # Tab visibility:
        # - unique_crs: hide if no unique_cr_path OR sp-level target
        # - overall_crs: removed from dashboard (now standalone page via /overall_crs/<target>)
        try:
            is_sp_target = '.' in str(target_name)
            hidden = set(base_context.get("pv_hidden_tabs") or [])
            # Hide unique_crs tab if no path or SP target
            if not str(base_context.get("unique_cr_path") or "").strip() or is_sp_target:
                hidden.add("unique_crs")
            # Always hide overall_crs from dashboard � it is a standalone page now
            hidden.add("overall_crs")
            base_context["pv_hidden_tabs"] = sorted(hidden)
        except Exception:
            pass


        perf_marks.append(("context_ready", _perf_elapsed_ms(perf_total_start)))

        # 6) Section routing
        if section == "mtbf-table":
            _perf_log_dashboard(target_name, section, perf_marks, {
                "rows": len(build_report_rows or []),
                "open_crs": len(cr_rows or []),
                "hwpdt": int(bool(hwpdt_available)),
                "total": _perf_elapsed_ms(perf_total_start),
            })
            return render_template(

                "mtbf_table.html",
                target=target_name,
                page_heading=f"{target_display_name} - SWPDT",
                build_report_rows=build_report_rows,
                mtbf_series=mtbf_series,
                mtbf_build_table=mtbf_build_table,
                **base_context,
            )
        elif section == "mtbf-trend":
            _perf_log_dashboard(target_name, section, perf_marks, {
                "rows": len(base_rows or []),
                "total": _perf_elapsed_ms(perf_total_start),
            })
            current_toggle_mode = (request.args.get("toggle_mode") or "ALL").upper()

            current_mtbf_mode = (
                request.args.get("mtbf_mode")
                or ("product_qc" if (is_compute_bu and compute_bu) else "normal")
            ).lower()
            trend_map = {}
            for r in base_rows:
                mid = str(r.get("meta_id") or "").strip()
                if not mid:
                    continue
                trend_map[mid] = {
                    "meta_id": mid,
                    "crashes": r.get("crashes") or 0,
                    "hours": r.get("total_hours") or r.get("hours") or 0,
                    "total_hours": r.get("total_hours") or r.get("hours") or 0,
                    "mtbf": _round_if_number(r.get("mtbf"), 2),
                    "product_mtbf": _round_if_number(r.get("product_mtbf"), 2),
                    "qc_mtbf": _round_if_number(r.get("qc_mtbf"), 2),
                }
            trend_rows = list(trend_map.values())
            trend_rows.sort(key=lambda x: str(x.get("meta_id") or ""))
            return render_template(
                "mtbf_trend.html",
                target=target_name,
                page_heading=f"{target_display_name} - MTBF Trend",
                page_subtitle="Trend view for MTBF, crashes and hours.",
                trend_rows=trend_rows,
                current_toggle_mode=current_toggle_mode,
                current_mtbf_mode=current_mtbf_mode,
                **base_context,
            )
        elif section == "hwpdt":
            _perf_log_dashboard(target_name, section, perf_marks, {
                "hw_rows": len(hwpdt_build_table or []),
                "total": _perf_elapsed_ms(perf_total_start),
            })
            if not hwpdt_available:

                return redirect(url_for(
                    "dashboard_bp.dashboard",
                    target_name=target_name,
                    section="mtbf-table",
                    toggle_mode=toggle_mode,
                    pdt_type="SWPDT",
                ))
            flat_rows = hwpdt_build_table or []
            grouped = defaultdict(list)
            for r in flat_rows:
                mid = r.get("meta_id")
                if not mid:
                    continue
                hw_cr = int(r.get("crashes") or 0)
                if hw_cr <= 0:
                    continue
                grouped[mid].append(r)
            hwpdt_rows = []
            for meta_id, builds in grouped.items():
                total_crashes = sum(int(b.get("crashes") or 0) for b in builds)
                mode = builds[0].get("mode") or builds[0].get("build_mode") or toggle_mode
                chipmd = builds[0].get("chipmd_ticket") or builds[0].get("chipmd") or ""
                hwpdt_rows.append({
                    "meta_id": meta_id,
                    "builds": builds,
                    "hwpdt_crashes": total_crashes,
                    "mode": mode,
                    "chipmd_ticket": chipmd,
                })
            try:
                hwpdt_rows.sort(
                    key=lambda r: int(str(r.get("meta_id") or "0").split("-")[-1]),
                    reverse=True,
                )
            except ValueError:
                hwpdt_rows.sort(key=lambda r: r.get("meta_id") or "")
            for idx, r in enumerate(hwpdt_rows, start=1):
                r["s_no"] = idx
            return render_template(
                "hwpdt.html",
                target=target_name,
                page_heading=f"{target_display_name} - HWPDT Build Table",
                page_subtitle="Hardware PDT metrics and CHIPMD tickets",
                hwpdt_rows=hwpdt_rows,
                **base_context,
            )
        elif section == "open-cr-analysis":
            _perf_log_dashboard(target_name, section, perf_marks, {
                "open_crs": len(cr_rows or []),
                "total": _perf_elapsed_ms(perf_total_start),
            })
            return render_template(

                "open_cr_analysis.html",
                target=target_name,
                page_heading=f"{target_display_name} - Open_Analysis CR",
                page_subtitle="Open/Analysis CRs with debug notes and AI insights",
                **base_context,
            )
        elif section == "test-analysis":

            return render_template(
                "test_analysis.html",
                target=target_name,
                page_heading=f"{target_display_name} - Test Analysis",
                page_subtitle="Test pass/fail, coverage, stability",
                **base_context,
            )
        elif section == "rca":
            return render_template(
                "rca.html",
                target=target_name,
                page_heading=f"{target_display_name} - RCA",
                page_subtitle="Root cause analysis by CR / JIRA / META",
                **base_context,
            )
        elif section == "unique-crs":
            base_context["active_section"] = "unique-crs"
            return render_template(
                "target_layout.html",
                target=target_name,
                page_heading=f"{target_display_name} - Unique CRs",
                page_subtitle="All unique CRs with area, status and priority breakdown",
                **base_context,
            )
        elif section == "pdt-crs":
            is_compute_target = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
            cr_tag_enabled = _is_compute_cr_tag_enabled_target(target_name)

            return render_template(
                "pdt_crs_section.html",
                target=target_name,
                page_heading=f"{target_display_name} - PDT CRs",
                page_subtitle="Top offenders by JIRA hits, area & status breakdown",
                is_compute_target=is_compute_target,
                cr_tag_enabled=cr_tag_enabled,
                cr_tag_alias_groups=_load_compute_cr_tag_alias_config() if cr_tag_enabled else [],
                **base_context,
            )

        elif section == "weekly-report":
            report_summary, from_date, to_date, error = build_weekly_report_context(target_name, request)
            _perf_log_dashboard(target_name, section, perf_marks, {
                "weekly_rows": len((report_summary or {}).get("cr_rows", []) or []),
                "total": _perf_elapsed_ms(perf_total_start),
            })
            return render_template(

                "weekly_data.html",
                target=target_name,
                page_heading=f"{target_display_name} - Weekly Report",
                from_date=from_date,
                to_date=to_date,
                report_summary=report_summary,
                error=error,
                **base_context,
            )
        elif section == "open-jiras":
            _perf_log_dashboard(target_name, section, perf_marks, {
                "total": _perf_elapsed_ms(perf_total_start),
            })
            return render_template(

                "open_jiras_section.html",
                target=target_name,
                page_heading=f"{target_display_name} - Open JIRAs",
                page_subtitle="Open JIRA tickets grouped by area with date range filter",
                **base_context,
            )
        elif section == "cr-info":
            search_query = (request.args.get("cr") or "").strip()
            cr_info = None
            jira_info = None
            cr_jiras = []
            cr_quick_summary = {}
            import re
            is_jira_key = re.match(r"^(QSTABILITY|CHIPMD|DROIDBUG)-\d+$", search_query, flags=re.I) is not None
            if is_jira_key:
                j_table = tables["j"]
                o_table = tables["o"]
                cursor.execute(
                    f"SELECT * FROM {j_table} WHERE stability_ticket = %s ORDER BY jira_date DESC",
                    (search_query,),
                )
                j_results = cursor.fetchall() or []
                cursor.execute(
                    f"SELECT * FROM {o_table} WHERE stability_ticket = %s ORDER BY jira_date DESC",
                    (search_query,),
                )
                o_results = cursor.fetchall() or []
                cr_jiras = sorted(j_results + o_results, key=_jira_sort_key, reverse=True)
                if cr_jiras:
                    latest_jira = cr_jiras[0]
                    jira_info = {
                        "stability_ticket": latest_jira.get("stability_ticket") or search_query,
                        "jira_title": latest_jira.get("jira_title") or latest_jira.get("title") or "",
                        "mcn": latest_jira.get("mcn") or "",
                        "serial_no": latest_jira.get("serial_no") or "",
                        "device_id": latest_jira.get("device_id") or latest_jira.get("serial_no") or "",
                        "jira_date": latest_jira.get("jira_date") or latest_jira.get("date") or "",
                        "test_team": latest_jira.get("test_team") or "",
                        "cr": latest_jira.get("cr") or "",
                        "mapped_crs": latest_jira.get("mapped_crs") or "",
                    }
                    devices    = sorted({(r.get("serial_no") or "").strip() for r in cr_jiras if r.get("serial_no")})
                    test_teams = sorted({(r.get("test_team") or "").strip() for r in cr_jiras if r.get("test_team")})
                    mcn_types  = sorted({(r.get("mcn") or "").strip() for r in cr_jiras if r.get("mcn")})
                    from collections import Counter
                    # Latest Meta Seen = metabuild from the most recently reported JIRA row.
                    # Do not use highest meta ID; jira_date decides what was seen latest.
                    _latest_meta_jira  = (latest_jira.get("metabuild") or latest_jira.get("build_id") or "").strip()
                    _latest_meta_count = sum(1 for r in cr_jiras if str(r.get("metabuild") or r.get("build_id") or "").strip() == _latest_meta_jira)
                    mcn_counts    = Counter((r.get("mcn") or "").strip()       for r in cr_jiras if (r.get("mcn") or "").strip())
                    device_counts = Counter((r.get("serial_no") or "").strip() for r in cr_jiras if (r.get("serial_no") or "").strip())
                    build_counts  = Counter((r.get("metabuild") or r.get("build_id") or "").strip() for r in cr_jiras if (r.get("metabuild") or r.get("build_id") or "").strip())
                    team_device_map = defaultdict(set)
                    team_mcn_map = defaultdict(set)
                    for _r in cr_jiras:
                        _team = (_r.get("test_team") or "").strip()
                        if not _team:
                            continue
                        _serial = (_r.get("serial_no") or "").strip()
                        _mcn = (_r.get("mcn") or "").strip()
                        if _serial:
                            team_device_map[_team].add(_serial)
                        if _mcn:
                            team_mcn_map[_team].add(_mcn)
                    team_device_map = {k: sorted(v) for k, v in team_device_map.items()}
                    team_mcn_map = {k: sorted(v) for k, v in team_mcn_map.items()}
                    cr_num_from_jira = (latest_jira.get("cr") or latest_jira.get("mapped_cr") or "").strip()
                    if cr_num_from_jira:
                        u_table = tables["u"]
                        cursor.execute(f"SELECT * FROM {u_table} WHERE mapped_cr = %s LIMIT 1", (cr_num_from_jira,))
                        cr_row = cursor.fetchone()
                        if cr_row:
                            cr_info = cr_row
                            cr_quick_summary = {
                                "occurrences":        len(cr_jiras),
                                "devices":            len(devices),
                                "serials":            devices,
                                "test_teams":         test_teams,
                                "mcn_types":          mcn_types,
                                "mcn_counts":         dict(mcn_counts.most_common()),
                                "device_counts":      dict(device_counts.most_common()),
                                "build_counts":       dict(build_counts.most_common(20)),
                                "team_device_map":    team_device_map,
                                "team_mcn_map":       team_mcn_map,
                                                                "searched_cr":        cr_num_from_jira or search_query,
                                "latest_meta":        _latest_meta_jira,
                                "latest_meta_count":  _latest_meta_count,
                                "cr_age":            _resolve_cr_info_group_age(cr_row, [cr_row], cr_row.get("mapped_cr") or cr_num_from_jira),
                            }

                else:
                    cr_quick_summary = {"error": f"JIRA '{search_query}' not found for {target_name}."}
            elif search_query:
                grouped_ctx = _fetch_grouped_cr_jira_context(cursor, target_name, search_query)
                if grouped_ctx:
                    cr_info  = grouped_ctx["primary_cr_info"]
                    cr_jiras = grouped_ctx["cr_jiras"]
                    latest_meta_jira = grouped_ctx.get("latest_meta_jira") or {}
                    from collections import Counter
                    mcn_counts    = Counter((r.get("mcn") or "").strip()       for r in cr_jiras if (r.get("mcn") or "").strip())
                    device_counts = Counter((r.get("serial_no") or "").strip() for r in cr_jiras if (r.get("serial_no") or "").strip())
                    build_counts  = Counter((r.get("metabuild") or r.get("build_id") or "").strip() for r in cr_jiras if (r.get("metabuild") or r.get("build_id") or "").strip())
                    latest_meta      = grouped_ctx.get("latest_meta") or ""
                    latest_meta_rows_local = [r for r in cr_jiras if str(r.get("metabuild") or "").strip() == latest_meta] if latest_meta else []
                    lm_mcn_counts    = Counter((r.get("mcn") or "").strip()       for r in latest_meta_rows_local if (r.get("mcn") or "").strip())
                    lm_device_counts = Counter((r.get("serial_no") or "").strip() for r in latest_meta_rows_local if (r.get("serial_no") or "").strip())
                    team_device_map = defaultdict(set)
                    team_mcn_map = defaultdict(set)
                    for _r in cr_jiras:
                        _team = (_r.get("test_team") or "").strip()
                        if not _team:
                            continue
                        _serial = (_r.get("serial_no") or "").strip()
                        _mcn = (_r.get("mcn") or "").strip()
                        if _serial:
                            team_device_map[_team].add(_serial)
                        if _mcn:
                            team_mcn_map[_team].add(_mcn)
                    team_device_map = {k: sorted(v) for k, v in team_device_map.items()}
                    team_mcn_map = {k: sorted(v) for k, v in team_mcn_map.items()}
                    cr_quick_summary = {
                        "occurrences": len(cr_jiras),
                        "cr_instances": len(grouped_ctx.get("cr_group_rows") or []),
                        "jira_instances": len(grouped_ctx.get("jira_instances") or []),
                        "devices": len(grouped_ctx.get("devices") or []),
                        "serials": grouped_ctx.get("devices") or [],
                        "test_teams": grouped_ctx.get("test_teams") or [],
                        "mcn_types": grouped_ctx.get("mcn_types") or [],
                        "latest_meta": grouped_ctx.get("latest_meta") or "",
                        "latest_meta_count": grouped_ctx.get("latest_meta_count") or 0,
                        "linked_crs": grouped_ctx.get("linked_crs") or [],
                        "searched_cr": grouped_ctx.get("searched_cr") or search_query,
                        "canonical_mapped_cr": grouped_ctx.get("canonical_mapped_cr") or "",
                        "cr_group_rows": grouped_ctx.get("cr_group_rows") or [],
                        "mcn_counts":             dict(mcn_counts.most_common()),
                        "device_counts":          dict(device_counts.most_common()),
                        "build_counts":           dict(build_counts.most_common(20)),
                        "team_device_map":        team_device_map,
                        "team_mcn_map":           team_mcn_map,
                        "latest_meta_mcn_counts":    dict(lm_mcn_counts.most_common()),
                        "latest_meta_device_counts": dict(lm_device_counts.most_common()),
                                                "cr_age": _resolve_cr_info_group_age(
                            cr_info,
                            grouped_ctx.get("cr_group_rows") or [],
                            grouped_ctx.get("canonical_mapped_cr") or "",
                        ),
                    }
                    if latest_meta_jira:
                        jira_info = {
                            "stability_ticket": latest_meta_jira.get("stability_ticket") or "",
                            "jira_title": latest_meta_jira.get("jira_title") or latest_meta_jira.get("title") or "",
                            "mcn": latest_meta_jira.get("mcn") or "",
                            "serial_no": latest_meta_jira.get("serial_no") or "",
                            "device_id": latest_meta_jira.get("device_id") or latest_meta_jira.get("serial_no") or "",
                            "jira_date": latest_meta_jira.get("jira_date") or "",
                            "test_team": latest_meta_jira.get("test_team") or "",
                            "cr": latest_meta_jira.get("cr") or "",
                            "mapped_crs": latest_meta_jira.get("mapped_crs") or "",
                        }
                        _perf_log_dashboard(target_name, section, perf_marks, {
                "search": search_query or "-",
                "jira_rows": len(cr_jiras or []),
                "total": _perf_elapsed_ms(perf_total_start),
            })
            return render_template(
                "cr_info.html",

                target=target_name,
                page_heading=f"{target_display_name} - CR Info",
                search_query=search_query,
                cr_info=cr_info,
                jira_info=jira_info,
                cr_jiras=cr_jiras,
                cr_quick_summary=cr_quick_summary,
                **base_context,
            )

                # Default: render the main dashboard section
        _perf_log_dashboard(target_name, section, perf_marks, {
            "rows": len(build_report_rows or []),
            "open_crs": len(cr_rows or []),
            "total": _perf_elapsed_ms(perf_total_start),
        })
        return render_template(

            "dashboard_overview.html",
            target=target_name,
            page_heading=f"{target_display_name} - Dashboard",
            **base_context,
        )

    except Exception as e:
        logger.info(f"[DASHBOARD PERF] target={target_name} section={section} failed_after={_perf_elapsed_ms(perf_total_start)}ms")
        return f"Error loading dashboard for {target_name}: {e}", 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# ---------------------------------------------------------------------
# FULL TABLE ROUTES
# ---------------------------------------------------------------------
@dashboard_bp.route("/dashboard/<target_name>/view_all_unique_crs")
@login_required
def view_all_unique_crs(target_name):
    info = get_target_info(target_name)
    if not info:
        return redirect(url_for("bu_selection"))
    query = f"SELECT * FROM {fq_table_for_target(target_name, 'unique_crs')} ORDER BY `cr` DESC"
    return redirect_to_full_table(query, target_name, table_name=f"{target_name} - unique_crs")


@dashboard_bp.route("/dashboard/<target_name>/view_all_open_jiras")
@login_required
def view_all_open_jiras(target_name):
    info = get_target_info(target_name)
    if not info:
        return redirect(url_for("bu_selection"))
    try:
        tbl = fq_table_for_target(target_name, 'openjiras')
    except Exception:
        flash("Open JIRAs table not available for this target.", "info")
        return redirect(url_for("dashboard_bp.dashboard", target_name=target_name, section="dashboard"))
    _conn = get_mysql_connection_db()
    _cur = _conn.cursor()
    try:
        _name = tbl.replace("`", "")
        _sch, _tname = _name.split(".", 1)
        _cur.execute("SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1", (_sch, _tname))
        if not _cur.fetchone():
            flash("Open JIRAs table not available for this target.", "info")
            return redirect(url_for("dashboard_bp.dashboard", target_name=target_name, section="dashboard"))
    finally:
        _cur.close(); _conn.close()
    query = f"SELECT * FROM {tbl} ORDER BY `stability_ticket` DESC"
    return redirect_to_full_table(query, target_name, table_name=f"{target_name} - openjiras")


@dashboard_bp.route("/dashboard/<target_name>/view_all_closed_jiras")
@login_required
def view_all_closed_jiras(target_name):
    info = get_target_info(target_name)
    if not info:
        return redirect(url_for("bu_selection"))
    try:
        tbl = fq_table_for_target(target_name, 'closed_jiras')
    except Exception:
        flash("Closed JIRAs table not available for this target.", "info")
        return redirect(url_for("dashboard_bp.dashboard", target_name=target_name, section="dashboard"))
    _conn = get_mysql_connection_db()
    _cur = _conn.cursor()
    try:
        _name = tbl.replace("`", "")
        _sch, _tname = _name.split(".", 1)
        _cur.execute("SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1", (_sch, _tname))
        if not _cur.fetchone():
            flash("Closed JIRAs table not available for this target.", "info")
            return redirect(url_for("dashboard_bp.dashboard", target_name=target_name, section="dashboard"))
    finally:
        _cur.close(); _conn.close()
    query = f"SELECT * FROM {tbl} ORDER BY `stability_ticket` DESC"
    return redirect_to_full_table(query, target_name, table_name=f"{target_name} - closed_jiras")


@dashboard_bp.route("/dashboard/<target_name>/view_all_jiras")
@login_required
def view_all_jiras(target_name):
    info = get_target_info(target_name)
    if not info:
        return redirect(url_for("bu_selection"))
    query = f"SELECT * FROM {fq_table_for_target(target_name, 'jiras')} ORDER BY `stability_ticket` DESC"
    return redirect_to_full_table(query, target_name, table_name=f"{target_name} - jiras")


@dashboard_bp.route("/view_all_undiposed_cr/<target_name>", methods=["GET"])
@login_required
def view_all_undiposed_cr(target_name):
    flash("Undisposed CR report not implemented in this version.", "info")
    return redirect(url_for("dashboard_bp.dashboard", target_name=target_name, section="dashboard"))


# ---------------------------------------------------------------------
# TARGET WORKSPACE
# ---------------------------------------------------------------------
@dashboard_bp.route("/target_workspace/<string:target_name>")
@login_required
def target_workspace(target_name):
    if (target_name or "").lower() in ("auto", "automotive"):
        return redirect(url_for("auto_select_gen"))
    ctx = _build_sidebar_context(target_name, active_section='customer-issues')
    return render_template(
        "coming_soon_template.html",
        target_name=target_name,
        page_heading="Customer Issues",
        page_subtitle="Coming soon � this section is under development.",
        **ctx,
    )


# ---------------------------------------------------------------------
# PDT PLANNING
# ---------------------------------------------------------------------
@dashboard_bp.route("/pdt/planning/<string:target_name>")
@login_required
def pdt_planning_empty(target_name):
    import time as _time
    from dashboard_common import get_bu_for_target
    active_bu_key = (get_bu_for_target(target_name) or '').upper()
    _ctx = _build_bu_shell_context(active_bu_key)
    _ctx.pop('active_section', None)
    # Add unique_cr_path so Unique CRs tab shows in panel
    try:
        from dashboard_common import get_targets_config
        _tinfo = (get_targets_config() or {}).get(target_name) or {}
        _ctx['unique_cr_path'] = _tinfo.get('unique_cr_path') or ''
    except Exception:
        pass
    return render_template(
        "pdt_planning.html",
        target=target_name,
        target_name=target_name,
        active_section='pdt-planning',
        milestone_phase=build_milestone_phase_context(target_name),
        cache_buster=int(_time.time()),
        **_ctx,
    )


# ---------------------------------------------------------------------
# API: CR/JIRA search helpers
# ---------------------------------------------------------------------
def _sort_meta_ids(meta_list):
    try:
        import re as _re
        def key_fn(x):
            s = str(x or "")
            m = _re.search(r"(\d+)$", s)
            return int(m.group(1)) if m else s
        return sorted({str(m or "").strip() for m in (meta_list or [])}, key=key_fn)
    except Exception:
        return sorted({str(m or "").strip() for m in (meta_list or [])})


def _normalize_cr_search_tokens(search_value):
    raw = (search_value or "").strip().upper()
    if not raw:
        return []
    digits = raw.replace("CR", "").strip()
    vals = []
    for v in (raw, digits, f"CR{digits}" if digits else ""):
        v = str(v or "").strip()
        if v and v not in vals:
            vals.append(v)
    return vals


def _cr_info_valid_number(value):
    """Return a positive integer for valid CR info numeric fields; ignore NA/Dup blanks."""
    import re as _re
    text = str(value or "").strip()
    if not text or text.upper() in {"NA", "N/A", "NONE", "NULL", "NAN", "DUP", "-", "--"}:
        return None
    match = _re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        number = int(float(match.group(0)))
    except Exception:
        return None
    return number if number > 0 else None


def _cr_info_key(value):
    text = str(value or "").strip().upper().replace(" ", "")
    return text[2:] if text.startswith("CR") else text


def _resolve_cr_info_group_age(primary_row, group_rows=None, canonical_mapped_cr=None):
    """
    Resolve CR age for the CR Info hero card.

    Some Excel rows in a mapped-CR family have occurrence='Dup' and cr_age='NA'.
    In that case, use the corresponding mapped/master row that has a real
    occurrence count and a valid age.
    """
    rows = list(group_rows or [])
    if primary_row and primary_row not in rows:
        rows.insert(0, primary_row)

    canonical_key = _cr_info_key(canonical_mapped_cr)
    candidates = []
    for idx, row in enumerate(rows):
        if not row:
            continue
        age = _cr_info_valid_number(row.get("effective_cr_age")) or _cr_info_valid_number(row.get("cr_age"))
        if age is None:
            continue
        occurrence = (
            _cr_info_valid_number(row.get("effective_jira_count"))
            or _cr_info_valid_number(row.get("cr_occurrence"))
            or _cr_info_valid_number(row.get("jira_count"))
        )
        row_cr_key = _cr_info_key(row.get("cr") or row.get("cr_number"))
        row_mapped_key = _cr_info_key(row.get("mapped_cr"))
        score = (
            1 if canonical_key and row_cr_key == canonical_key else 0,
            1 if row_mapped_key and row_cr_key == row_mapped_key else 0,
            1 if occurrence is not None else 0,
            occurrence or 0,
            -idx,
        )
        candidates.append((score, age))

    if candidates:
        candidates.sort(reverse=True)
        return candidates[0][1]
    return None


# ---------------------------------------------------------------------
# PDT ANALYSIS
# ---------------------------------------------------------------------
@dashboard_bp.route("/pdt/analysis/<string:target_name>")
@login_required
def pdt_analysis_empty(target_name):
    import time as _time
    from dashboard_common import get_bu_for_target
    active_bu_key = (get_bu_for_target(target_name) or '').upper()
    _ctx = _build_bu_shell_context(active_bu_key)
    _ctx.pop('active_section', None)
    # Add unique_cr_path so Unique CRs tab shows in panel
    try:
        from dashboard_common import get_targets_config
        _tinfo = (get_targets_config() or {}).get(target_name) or {}
        _ctx['unique_cr_path'] = _tinfo.get('unique_cr_path') or ''
    except Exception:
        pass
    return render_template(
        "pdt_analysis.html",
        target=target_name,
        target_name=target_name,
                active_section='pdt-analysis',
        milestone_phase=build_milestone_phase_context(target_name),
        cache_buster=int(_time.time()),
        **_ctx,
    )


# ---------------------------------------------------------------------
# LIVE STATUS VIEW  - standalone page (opens in new tab from dashboard)
# ---------------------------------------------------------------------
@dashboard_bp.route("/pdt/live-status-view/<string:target_name>")
@login_required
def live_status_view(target_name):
    import time as _time
    prefill_build = request.args.get('buildid', '').strip()
    return render_template(
        "live_status_view.html",
        target_name=target_name,
        prefill_build=prefill_build,
        cache_buster=int(_time.time()),
    )


# ---------------------------------------------------------------------
# BUILD TRACKER - standalone page (opens in new tab from dashboard)
# ---------------------------------------------------------------------
@dashboard_bp.route("/pdt/build-tracker/<string:target_name>")
@login_required
def build_tracker(target_name):
    import time as _time
    return render_template(
        "build_tracker.html",
        target_name=target_name,
        cache_buster=int(_time.time()),
    )


@dashboard_bp.route("/api/jira_build_lookup")
@login_required
def api_jira_build_lookup():
    """
    Fetch raw JIRA info for a given Build ID.
    Query params:
      buildid = Build ID  e.g. Skyros.LA.1.0-00321-PERF.INT-1
      jql     = Raw JQL (overrides buildid)
    Returns JSON: { meta, summary, jiras[] }
    """
    import time as _time
    import sys as _sys
    import os as _os

    build_id = request.args.get('buildid', '').strip()
    raw_jql  = request.args.get('jql',     '').strip()

    if not build_id and not raw_jql:
        return jsonify({'error': 'buildid or jql parameter is required'}), 400

        # import helpers from scripts/
    try:
        import os as _os, sys as _sys
        _scripts_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'scripts')
        if _scripts_dir not in _sys.path:
            _sys.path.insert(0, _scripts_dir)
        import urllib3 as _urllib3
        _urllib3.disable_warnings(_urllib3.exceptions.InsecureRequestWarning)
    except Exception:
        pass

    try:
        from fetch_jira_by_build import (
            connect_jira, run_query, issue_to_dict,
            make_summary, build_jql_from_buildid,
        )
        from config import JIRA_USER, JIRA_PASSWORD, JIRA_SERVER_ENDPOINT, JIRA_PDT_FILTER_ID
    except ImportError as ie:
        logger.error(f'[LIVE STATUS VIEW] Import error: {ie}')
        return jsonify({'error': f'Server config error: {ie}'}), 500

    if not JIRA_USER or not JIRA_PASSWORD:
        return jsonify({'error': 'JIRA credentials not configured. Set JIRA_USER and JIRA_PASSWORD in .env'}), 500

    jql = raw_jql if raw_jql else build_jql_from_buildid(build_id, JIRA_PDT_FILTER_ID)

    try:
        t0       = _time.time()
        jira_obj = connect_jira(JIRA_USER, JIRA_PASSWORD, JIRA_SERVER_ENDPOINT)

        # count first (no data fetch)
        count_result = jira_obj.search_issues(jql, startAt=0, maxResults=0, fields='summary')
        total = count_result.total if hasattr(count_result, 'total') else 0
        logger.info(f'[LIVE STATUS VIEW] build={build_id} total_available={total}')

        if total == 0:
            return jsonify({
                'meta'   : {'jql': jql, 'jira_server': JIRA_SERVER_ENDPOINT,
                            'fetch_time_sec': 0, 'total_available': 0, 'total_fetched': 0},
                'summary': {'total_jiras': 0, 'by_project': {}, 'with_resolution_notes': 0, 'with_cr_number_field': 0},
                'jiras'  : []
            })

        issues       = run_query(jira_obj, jql, max_results=total)
        elapsed      = round(_time.time() - t0, 2)
        issues_dicts = [issue_to_dict(i) for i in issues]

        return jsonify({
            'meta': {
                'jql'            : jql,
                'jira_server'    : JIRA_SERVER_ENDPOINT,
                'fetch_time_sec' : elapsed,
                'total_available': total,
                                  'total_fetched'  : len(issues_dicts),
              },
              'summary': make_summary(issues_dicts),
              'jiras'  : issues_dicts,
          })

    except Exception as e:
        import traceback
        logger.error(f'[jira_build_lookup] {e}\n{traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500


import hashlib as _hashlib
import uuid as _uuid

_CONSOLIDATED_REPORT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), 'data', 'consolidated_reports'
)
os.makedirs(_CONSOLIDATED_REPORT_DIR, exist_ok=True)


def _consolidated_report_path(target, builds, custom_jql=None):
    key_parts = [(target or 'notarget'), '_'.join(sorted(str(b) for b in (builds or [])))]
    if custom_jql:
        key_parts.append('jql:' + str(custom_jql).strip())
    key   = '_'.join(key_parts)
    fname = _hashlib.md5(key.encode()).hexdigest()[:12] + '.json'
    return os.path.join(_CONSOLIDATED_REPORT_DIR, fname), key


# job_id ? final report dict (populated by background thread)
_JOB_RESULTS: dict = {}


# -- SSE progress endpoint -----------------------------------------------------

def _br_tail(value):
    text = str(value or '').strip().replace('/', '\\')
    parts = [p for p in text.split('\\') if p]
    return parts[-1] if parts else text


def _br_norm_name(value):
    import re as _re
    return _re.sub(r'[^a-z0-9]+', '', str(value or '').lower())


def _br_first_col(cols, names):
    by_norm = {_br_norm_name(c): c for c in (cols or [])}
    for name in names:
        hit = by_norm.get(_br_norm_name(name))
        if hit:
            return hit
    return None


def _br_table_exists(cur, schema, table):
    cur.execute('SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1', (schema, table))
    return cur.fetchone() is not None


def _br_cols(cur, schema, table):
    try:
        cur.execute(f'SHOW COLUMNS FROM `{schema}`.`{table}`')
        return {r.get('Field') for r in (cur.fetchall() or []) if r.get('Field')}
    except Exception:
        return set()


def _br_is_real_cr(value):
    import re as _re
    text = str(value or '').strip().upper()
    return bool(text and text not in ('NO_CR', '--', 'NONE', 'NULL', 'N/A', 'NA') and (_re.match(r'^(CR)?\d{3,}$', text) or _re.match(r'^CR[-_A-Z0-9]+$', text)))


def _br_build_like_values(builds):
    import re as _re
    vals = []
    for build in builds or []:
        tail = _br_tail(build)
        if not tail:
            continue
        vals.append(f'%{tail}%')
        m = _re.search(r'-(\d{3,6})(?:\.\d+)?-(?:STD|PERF|SAFE|USER|ENG)', tail, _re.I)
        if m:
            n = str(int(m.group(1))); p = n.zfill(3)
            vals.extend([f'%-{p}-%', f'%-{n}-%', f'%Meta-{p}%', f'%Meta-{n}%'])
    return list(dict.fromkeys(vals))


def _br_domain_sources(target, domain):
    try:
        from core_deck_routes import _load_state, _db_source_tables_for_targets
        state = _load_state(target) or {}
        preview = state.get('saved_preview') if isinstance(state.get('saved_preview'), dict) else {}
        cfg = state.get('deck_config') or preview.get('deck_config') or {}
        entries = cfg.get(domain) or cfg.get(str(domain).lower()) or []
        if isinstance(entries, str):
            import re as _re
            entries = [v.strip() for v in _re.split(r'[,;\n]+', entries) if v.strip()]
        return _db_source_tables_for_targets(entries), cfg
    except Exception:
        logger.debug('[build_report] Core Deck Config source load failed', exc_info=True)
        return [], {}


def _build_domain_table_report(target, domain, builds):
    """Fetch Build Report rows from Core Deck Config domain JIRA/OpenJIRA tables."""
    domain = str(domain or '').strip().upper()
    builds = [str(b or '').strip() for b in (builds or []) if str(b or '').strip()]
    sources, cfg = _br_domain_sources(target, domain)
    like_vals = _br_build_like_values(builds)
    if not sources:
        raise ValueError(f'No Core Slide Config table details found for {domain}. Save Config in Core Slide page first.')
    if not like_vals:
        raise ValueError('Select at least one valid build.')
    crs, open_jiras, total = {}, [], 0
    for src in sources:
        schema = (src.get('schema') or '').strip('`')
        if not schema:
            continue
        conn = get_mysql_connection_db(bu_key=schema)
        if not conn:
            continue
        cur = conn.cursor(dictionary=True)
        try:
            for exists_key, table_key, kind in (('jiras', 'jiras_table_name', 'jira'), ('openjiras', 'openjiras_table_name', 'open_jira')):
                table = src.get(table_key) or ''
                if not table or not (src.get('exists') or {}).get(exists_key) or not _br_table_exists(cur, schema, table):
                    continue
                cols = _br_cols(cur, schema, table)
                mb_col = _br_first_col(cols, ['metabuild', 'meta_build', 'build_id', 'build', 'MetaBuild', 'Meta Build'])
                if not mb_col:
                    continue
                cr_col = _br_first_col(cols, ['mapped_cr', 'mapped_crs', 'Mapped CRs', 'Mapped CR', 'cr', 'cr_number'])
                ticket_col = _br_first_col(cols, ['stability_ticket', 'jira_id', 'ticket', 'key'])
                title_col = _br_first_col(cols, ['jira_title', 'summary', 'title'])
                status_col = _br_first_col(cols, ['cr_status', 'status', 'final_status'])
                area_col = _br_first_col(cols, ['cr_area', 'area'])
                sub_col = _br_first_col(cols, ['cr_subsystem', 'subsystem'])
                func_col = _br_first_col(cols, ['cr_functionality', 'functionality', 'cr_function'])
                age_col = _br_first_col(cols, ['cr_age', 'age', 'overall_age'])
                select_cols = [
                    f'`{mb_col}` AS metabuild',
                    f'`{cr_col}` AS cr' if cr_col else "'' AS cr",
                    f'`{ticket_col}` AS stability_ticket' if ticket_col else "'' AS stability_ticket",
                    f'`{title_col}` AS jira_title' if title_col else "'' AS jira_title",
                    f'`{status_col}` AS status' if status_col else "'' AS status",
                    f'`{area_col}` AS area' if area_col else "'' AS area",
                    f'`{sub_col}` AS subsystem' if sub_col else "'' AS subsystem",
                    f'`{func_col}` AS functionality' if func_col else "'' AS functionality",
                    f'`{age_col}` AS age' if age_col else "'' AS age",
                ]
                where_like = ' OR '.join([f'`{mb_col}` LIKE %s' for _ in like_vals])
                cur.execute(f"SELECT {', '.join(select_cols)} FROM `{schema}`.`{table}` WHERE ({where_like})", tuple(like_vals))
                rows = [dict(r) for r in (cur.fetchall() or [])]
                total += len(rows)
                for r in rows:
                    cr = str(r.get('cr') or '').strip()
                    if _br_is_real_cr(cr):
                        cr_key = cr if cr.upper().startswith('CR') else 'CR' + cr
                        row = crs.setdefault(cr_key, {
                            'cr': cr_key, 'mapped_cr': cr_key,
                            'cr_title': str(r.get('jira_title') or '').strip(),
                            'cr_area': str(r.get('area') or '').strip(),
                            'cr_subsystem': str(r.get('subsystem') or '').strip(),
                            'cr_functionality': str(r.get('functionality') or '').strip(),
                            'cr_age': str(r.get('age') or '').strip(),
                            'cr_status': str(r.get('status') or '').strip(),
                            'jira_count': 0, 'domain': domain,
                            'source_table': f'{schema}.{table}',
                        })
                        row['jira_count'] += 1
                    if kind == 'open_jira':
                        open_jiras.append({
                            'key': str(r.get('stability_ticket') or '').strip(),
                            'stability_ticket': str(r.get('stability_ticket') or '').strip(),
                            'summary': str(r.get('jira_title') or '').strip(),
                            'jira_title': str(r.get('jira_title') or '').strip(),
                            'status': str(r.get('status') or 'Open').strip(),
                            'metabuild': str(r.get('metabuild') or '').strip(),
                            'domain': domain, 'source_table': f'{schema}.{table}',
                        })
        finally:
            try:
                cur.close(); conn.close()
            except Exception:
                pass
    cr_rows = sorted(crs.values(), key=lambda r: int(r.get('jira_count') or 0), reverse=True)
    return {
        'meta': {'target_name': target, 'domain': domain, 'build_ids': builds, 'source': 'core_slide_config_domain_tables', 'generated_at': time.strftime('%Y-%m-%dT%H:%M:%S')},
        'summary': {'total_jiras': total, 'with_cr': sum(int(r.get('jira_count') or 0) for r in cr_rows), 'open_jiras': len(open_jiras)},
        'cr_rows': cr_rows,
        'open_jiras': open_jiras,
        'db_sources': sources,
        'deck_config': cfg,
    }
@dashboard_bp.route("/api/consolidated_report/progress/<job_id>")
@login_required
def api_consolidated_report_progress(job_id):
    """
    SSE stream: GET /api/consolidated_report/progress/<job_id>
    Streams { stage, total, done, pct, message, log } events until done/error.
    """
    import sys as _sys, os as _os
    _scripts_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'scripts')
    if _scripts_dir not in _sys.path:
        _sys.path.insert(0, _scripts_dir)

    from flask import Response, stream_with_context

    def _generate():
        try:
            from fetch_consolidated_report import get_progress
        except Exception:
            yield 'data: {"error": "module not found"}\n\n'
            return

        last_done = -1
        deadline  = time.time() + 600   # 10 min max
        while time.time() < deadline:
            pt = get_progress(job_id)
            if pt is None:
                yield f'data: {{"stage":"error","message":"job not found","done":0,"total":0,"pct":0}}\n\n'
                return
            snap = pt.snapshot()
            if snap['done'] != last_done or snap['stage'] in ('done', 'error'):
                last_done = snap['done']
                yield f'data: {json.dumps(snap)}\n\n'
            if snap['stage'] in ('done', 'error'):
                return
            time.sleep(0.5)

    return Response(
        stream_with_context(_generate()),
        mimetype='text/event-stream',
        headers={
            'Cache-Control'  : 'no-cache',
            'X-Accel-Buffering': 'no',
        }
    )


@dashboard_bp.route("/api/consolidated_report/result/<job_id>")
@login_required
def api_consolidated_report_result(job_id):
    """Pick up the finished report by job_id."""
    report = _JOB_RESULTS.pop(job_id, None)
    if report is None:
        return jsonify({'status': 'pending'}), 202
    if 'error' in report:
        return jsonify(report), 500
    return jsonify(report)


@dashboard_bp.route("/api/consolidated_report/load", methods=["GET", "POST"])
@login_required
def api_consolidated_report_load():
    """
    Load a previously saved consolidated report from disk (no JIRA/Orbit call).
    Returns { found: true, report: {...}, saved_at: '...' } or { found: false }
    """
    if request.method == 'POST':
        body   = request.get_json(force=True, silent=True) or {}
        builds = body.get('builds', [])
        target = (body.get('target') or '').strip()
    else:
        raw    = (request.args.get('builds') or '').strip()
        builds = [b.strip() for b in raw.split(',') if b.strip()]
        target = (request.args.get('target') or '').strip()

    if not builds:
        return jsonify({'found': False, 'reason': 'no builds'})

    path, _ = _consolidated_report_path(target, builds)
    if not os.path.exists(path):
        return jsonify({'found': False})

    try:
        with open(path, encoding='utf-8') as fh:
            data = json.load(fh)
        saved_at = data.get('meta', {}).get('saved_at') or data.get('meta', {}).get('generated_at', '')
        return jsonify({'found': True, 'report': data, 'saved_at': saved_at})
    except Exception as e:
        return jsonify({'found': False, 'reason': str(e)})


@dashboard_bp.route("/api/consolidated_report/save", methods=["POST"])
@login_required
def api_consolidated_report_save():
    """Persist the currently displayed consolidated report JSON to the static cache."""
    body   = request.get_json(force=True, silent=True) or {}
    report = body.get('report') or {}
    if not isinstance(report, dict) or not report:
        return jsonify({'ok': False, 'error': 'report payload required'}), 400

    meta       = report.setdefault('meta', {})
    builds     = body.get('builds') or meta.get('build_ids') or []
    target     = (body.get('target') or meta.get('target_name') or '').strip()
    custom_jql = (body.get('custom_jql') or meta.get('custom_jql') or '').strip()
    if not builds and not custom_jql:
        return jsonify({'ok': False, 'error': 'builds or custom_jql required'}), 400

    meta['target_name'] = target or meta.get('target_name')
    meta['saved_at'] = time.strftime("%Y-%m-%dT%H:%M:%S")
    path, _ = _consolidated_report_path(target, builds, custom_jql or None)
    try:
        with open(path, 'w', encoding='utf-8') as fh:
            json.dump(report, fh, ensure_ascii=False)
        return jsonify({'ok': True, 'path': path, 'saved_at': meta['saved_at']})
    except Exception as e:
        logger.warning(f"[consolidated_report] manual save failed: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@dashboard_bp.route("/api/consolidated_report/status")
@login_required
def api_consolidated_report_status():
    """Preflight status for the JIRA-backed consolidated-report runner."""
    try:
        from config import JIRA_USER, JIRA_PASSWORD, JIRA_SERVER_ENDPOINT
        configured = bool(str(JIRA_USER or '').strip() and str(JIRA_PASSWORD or '').strip())
        return jsonify({
            'ok': True,
            'configured': configured,
            'jira_user_set': bool(str(JIRA_USER or '').strip()),
            'jira_password_set': bool(str(JIRA_PASSWORD or '').strip()),
            'jira_server': JIRA_SERVER_ENDPOINT,
            'message': 'JIRA credentials configured.' if configured else 'JIRA credentials missing. Set JIRA_USER and JIRA_PASSWORD in .env, or LDAP_USER and LDAP_PASSWORD aliases.',
        })
    except Exception as exc:
        return jsonify({'ok': False, 'configured': False, 'error': str(exc)}), 500


def _extract_axiom_meta_ids_from_jql(jql: str):
    """Extract likely meta/build IDs from direct JQL summary clauses.

    Handles examples like:
      summary ~ "Glymur.WP.1.0.r0-05125.13-MAH.INT-1"
      summary ~ Maili.LA.1.0-00129-STD.INT-1
    """
    import re as _re
    text = str(jql or "")
    out, seen = [], set()

    def _add(value):
        v = str(value or "").strip().strip('"').strip("'")
        if not v or v.lower() in ("tombstone", "target stability"):
            return
        # Meta/build IDs normally contain both dots and dashes and at least one digit.
        if not ("." in v and "-" in v and _re.search(r"\d", v)):
            return
        key = v.upper()
        if key not in seen:
            seen.add(key); out.append(v)

    for m in _re.finditer(r"summary\s*~\s*(['\"])(.*?)\1", text, flags=_re.I):
        _add(m.group(2))
    for m in _re.finditer(r"summary\s*~\s*([^\s\)]+)", text, flags=_re.I):
        _add(m.group(1))

    # Fallback: scan all tokens for typical build/meta shape.
    token_re = r"\b[A-Za-z0-9][A-Za-z0-9_]*(?:\.[A-Za-z0-9_]+)+(?:-[A-Za-z0-9_.]+)+\b"
    for m in _re.finditer(token_re, text):
        _add(m.group(0))
    return out


def _fetch_axiom_stability_metrics_for_meta_ids(meta_ids, taxonomy_path=None, max_workers=8, start_date=None):
    """Create an Axiom stability report instance and fetch metrics for meta IDs.

    Correct Axiom sequence:
      POST /stabilityreport -> reportId
      POST /stabilityreport/{reportId}/instances -> instanceId
      poll GET /stabilityreport/{reportId}/instances until Completed
      GET /stabilityreport/{reportId}/instances/{instanceId}/metrics
    """
    import base64 as _base64, http.client as _http_client, json as _json
    import os as _os, re as _re, ssl as _ssl, time as _time
    import urllib.parse as _urlparse, uuid as _uuid
    from datetime import datetime as _datetime, timezone as _timezone, timedelta as _timedelta

    unique = []
    seen = set()
    for m in (meta_ids or []):
        s = str(m or "").strip()
        if s and s.upper() not in seen:
            seen.add(s.upper())
            unique.append(s)
    if not unique:
        return {}

    host = (_os.getenv("AXIOM_API_HOST") or "api-int.qualcomm.com").replace("https://", "").replace("http://", "").strip("/")
    if not host or "qualcomm" not in host.lower():
        host = "api-int.qualcomm.com"
    app_name = _os.getenv("AXIOM_APP_NAME", "Axiom_public-pdt-pcie").strip() or "Axiom_public-pdt-pcie"
    client_id = _os.getenv("AXIOM_CLIENT_ID", "").strip()
    client_secret = _os.getenv("AXIOM_CLIENT_SECRET", "").strip()
    taxonomy = str(taxonomy_path or _os.getenv("AXIOM_STABILITY_TAXONOMY_PATH") or _os.getenv("AXIOM_TAXONOMY_PATH_SW") or "/PDT").strip() or "/PDT"

    if not client_id or not client_secret:
        return {m: {"ok": False, "matched": False, "taxonomyPath": taxonomy, "error": "AXIOM_CLIENT_ID / AXIOM_CLIENT_SECRET missing"} for m in unique}

    def _default_start_date():
        configured = str(start_date or _os.getenv("AXIOM_STABILITY_START_DATE") or "").strip()
        if configured:
            return configured
        now = _datetime.now(_timezone.utc)
        if now.day >= 24:
            return now.replace(day=24, hour=0, minute=0, second=0, microsecond=0).isoformat(timespec="milliseconds").replace("+00:00", "Z")
        return (now - _timedelta(days=27)).replace(hour=0, minute=0, second=0, microsecond=0).isoformat(timespec="milliseconds").replace("+00:00", "Z")

    def _runtime_hours_label(value):
        raw = str(value or "").strip().lower()
        if not raw:
            return "--"
        days = hours = minutes = 0.0
        m = _re.search(r"(\d+(?:\.\d+)?)\s*day", raw)
        if m: days = float(m.group(1))
        m = _re.search(r"(\d+(?:\.\d+)?)\s*hr", raw)
        if m: hours = float(m.group(1))
        m = _re.search(r"(\d+(?:\.\d+)?)\s*min", raw)
        if m: minutes = float(m.group(1))
        total = days * 24.0 + hours + minutes / 60.0
        if total <= 0:
            return raw
        return f"{total:.1f} hr" if abs(total - round(total)) >= 0.05 else f"{int(round(total))} hr"

    ctx = _ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = _ssl.CERT_NONE

    def _request(method, path, token=None, json_body=None):
        if token:
            headers = {
                "Authorization": f"Bearer {token}", "Accept": "application/json",
                "X-QCOM-AppName": app_name, "X-QCOM-TokenType": "OAuth",
                "X-QCOM-ClientType": "Python", "X-QCOM-TracingID": _uuid.uuid4().hex,
            }
            body = ""
            if json_body is not None:
                headers["Content-Type"] = "application/json"
                body = _json.dumps(json_body).encode("utf-8")
        else:
            headers = {"Authorization": "Basic " + _base64.b64encode(f"{client_id}:{client_secret}".encode()).decode()}
            body = ""
        conn = _http_client.HTTPSConnection(host, timeout=180, context=ctx)
        try:
            conn.request(method, path, body=body, headers=headers)
            resp = conn.getresponse()
            text = resp.read().decode("utf-8", errors="ignore")
            return resp.status, text
        finally:
            conn.close()

    def _data_list(payload):
        data = payload.get("data") if isinstance(payload, dict) else None
        if isinstance(data, list): return data
        if isinstance(data, dict): return [data]
        return []

    def _fail_all(error, **extra):
        return {m: {"ok": False, "matched": False, "metaId": m, "taxonomyPath": taxonomy, "metrics": [], "error": error, **extra} for m in unique}

    st, body = _request("POST", "/ent/oauth/v1/accesstoken?grant_type=client_credentials")
    if st != 200:
        return _fail_all(f"Axiom token HTTP {st}: {body[:250]}")
    try:
        token = (_json.loads(body) or {}).get("access_token")
    except Exception:
        token = None
    if not token:
        return _fail_all("Axiom token response missing access_token")

    report_body = {
        "reportType": "ByBuilds",
        "buildInfo": {"buildType": "MetaId", "metaIdBuilds": unique},
        "taxonomy": taxonomy,
        "startDate": _default_start_date(),
        "published": "All",
        "typesOfCrash": "All",
        "buildComposition": "All",
        "softwareImages": [],
    }
    st, body = _request("POST", "/axiom/v1/public/stabilityreport", token, report_body)
    if st not in (200, 201, 202):
        return _fail_all(f"stabilityreport POST HTTP {st}: {body[:300]}", requestBody=report_body)
    try:
        payload = _json.loads(body) or {}
        report_id = payload.get("reportId") or ((payload.get("data") or {}).get("reportId") if isinstance(payload.get("data"), dict) else None)
    except Exception:
        report_id = None
    if not report_id:
        return _fail_all("stabilityreport POST did not return reportId", requestBody=report_body, response=body[:300])

    st, body = _request("POST", f"/axiom/v1/public/stabilityreport/{_urlparse.quote(str(report_id), safe='')}/instances", token, None)
    if st not in (200, 201, 202):
        return _fail_all(f"instances POST HTTP {st}: {body[:300]}", reportId=str(report_id), requestBody=report_body)
    try:
        payload = _json.loads(body) or {}
        instance_id = payload.get("instanceId") or ((payload.get("data") or {}).get("instanceId") if isinstance(payload.get("data"), dict) else None)
    except Exception:
        instance_id = None
    if not instance_id:
        return _fail_all("instances POST did not return instanceId", reportId=str(report_id), response=body[:300])

    poll_seconds = int(_os.getenv("AXIOM_STABILITY_POLL_SECONDS", "120") or "120")
    interval = max(2, int(_os.getenv("AXIOM_STABILITY_POLL_INTERVAL", "5") or "5"))
    deadline = _time.time() + max(interval, poll_seconds)
    instance_status = "InProgress"
    while _time.time() < deadline:
        st, body = _request("GET", f"/axiom/v1/public/stabilityreport/{_urlparse.quote(str(report_id), safe='')}/instances?pageNumber=0&pageSize=50", token)
        if st == 200:
            try:
                instances = _data_list(_json.loads(body))
            except Exception:
                instances = []
            for inst in instances:
                iid = inst.get("instanceId") or inst.get("id") or inst.get("instance_id")
                if str(iid) == str(instance_id):
                    instance_status = str(inst.get("status") or instance_status)
                    break
            if instance_status.lower() == "completed":
                break
            if instance_status.lower() in ("failed", "error"):
                return _fail_all(f"Axiom instance status: {instance_status}", reportId=str(report_id), instanceId=str(instance_id), requestBody=report_body)
        _time.sleep(interval)

    if instance_status.lower() != "completed":
        return _fail_all(f"Axiom instance not completed yet: {instance_status}", reportId=str(report_id), instanceId=str(instance_id), requestBody=report_body)

    q = _urlparse.urlencode({"pageNumber": 0, "pageSize": 500})
    st, body = _request("GET", f"/axiom/v1/public/stabilityreport/{_urlparse.quote(str(report_id), safe='')}/instances/{_urlparse.quote(str(instance_id), safe='')}/metrics?{q}", token)
    if st != 200:
        return _fail_all(f"metrics HTTP {st}: {body[:300]}", reportId=str(report_id), instanceId=str(instance_id), requestBody=report_body)
    try:
        metrics = _data_list(_json.loads(body))
    except Exception as exc:
        return _fail_all(f"Unable to parse metrics response: {exc}", reportId=str(report_id), instanceId=str(instance_id), requestBody=report_body)

    by_meta = {}
    for metric in metrics:
        key = str(metric.get("meta") or "").strip().upper()
        if not key:
            continue
        enriched = dict(metric)
        enriched["runtimeHours"] = _runtime_hours_label(metric.get("runtime"))
        enriched["deviceCount"] = metric.get("uniqueDevices")
        by_meta.setdefault(key, []).append(enriched)

    results = {}
    for meta_id in unique:
        matches = by_meta.get(meta_id.upper(), [])
        results[meta_id] = {
            "ok": bool(matches), "matched": bool(matches), "metaId": meta_id,
            "taxonomyPath": taxonomy, "reportId": str(report_id), "instanceId": str(instance_id),
            "instanceStatus": instance_status, "requestBody": report_body, "metrics": matches,
        }
        if not matches:
            results[meta_id]["error"] = "No Axiom metric row matched requested metaId"
    return results


@dashboard_bp.route("/api/consolidated_report", methods=["GET", "POST"])
@login_required
def api_consolidated_report():
    """
    Single endpoint � pass one or more build IDs, get back one complete JSON.
    Saves result to disk so subsequent loads are instant.

    POST body (JSON):
      { "builds": ["Build1", "Build2"], "traverse": true, "orbit": true,
        "target": "aldabra", "force": false }

    GET params:
      builds=Build1,Build2  traverse=1  orbit=1  target=aldabra  force=0
    """
    import time as _time

    # -- parse params ----------------------------------------------------------
    if request.method == "POST":
        body      = request.get_json(force=True, silent=True) or {}
        raw       = body.get("builds", [])
        traverse  = bool(body.get("traverse", True))
        do_orbit  = bool(body.get("orbit",    True))
        target    = (body.get("target") or "").strip()
        force      = bool(body.get("force",    False))
        custom_jql = (body.get("custom_jql") or body.get("jql") or "").strip()
        axiom_taxonomy_path = (body.get("axiom_taxonomy_path") or body.get("taxonomyPath") or body.get("taxonomy_path") or "").strip()
        include_axiom_metrics = body.get("include_axiom_metrics", True)
        domain = (body.get("domain") or "").strip().upper()
        use_domain_tables = bool(body.get("use_domain_tables") or body.get("domain_tables") or body.get("core_deck_domain_tables"))
    else:
        raw_str   = (request.args.get("builds") or "").strip()
        raw       = [b.strip() for b in raw_str.split(",") if b.strip()]
        traverse  = request.args.get("traverse", "1") != "0"
        do_orbit  = request.args.get("orbit",    "1") != "0"
        target    = (request.args.get("target") or "").strip()
        force      = request.args.get("force",    "0") != "0"
        custom_jql = (request.args.get("custom_jql") or request.args.get("jql") or "").strip()
        axiom_taxonomy_path = (request.args.get("axiom_taxonomy_path") or request.args.get("taxonomyPath") or request.args.get("taxonomy_path") or "").strip()
        include_axiom_metrics = request.args.get("include_axiom_metrics", "1") != "0"
        domain = (request.args.get("domain") or "").strip().upper()
        use_domain_tables = request.args.get("use_domain_tables", "0") != "0"

    if isinstance(raw, str):
        raw = [b.strip() for b in raw.replace("\n", ",").split(",") if b.strip()]
    # If user pasted direct JQL only, extract build/meta IDs from summary clauses
    # so Axiom stability metrics can be fetched for those builds.
    extracted_from_jql = _extract_axiom_meta_ids_from_jql(custom_jql) if custom_jql else []
    if not raw and extracted_from_jql:
        raw = extracted_from_jql[:]
    if not raw and not custom_jql:
        return jsonify({"error": "'builds' or 'custom_jql' param required"}), 400

    if use_domain_tables:
        try:
            return jsonify(_build_domain_table_report(target, domain, raw))
        except Exception as e:
            logger.error(f"[consolidated_report] domain table report failed: {e}", exc_info=True)
            return jsonify({'error': str(e)}), 500

    try:
        from config import JIRA_USER, JIRA_PASSWORD
        if not (str(JIRA_USER or '').strip() and str(JIRA_PASSWORD or '').strip()):
            return jsonify({
                'error': 'JIRA credentials missing. Set JIRA_USER and JIRA_PASSWORD in .env, or LDAP_USER and LDAP_PASSWORD aliases.',
                'missing_credentials': True,
            }), 500
    except Exception as e:
        return jsonify({'error': f'Unable to read JIRA config: {e}', 'missing_credentials': True}), 500

    # -- check static cache unless force=true ---------------------------------


    cache_path, _ = _consolidated_report_path(target, raw, custom_jql or None)
    if not force and os.path.exists(cache_path):
        try:
            with open(cache_path, encoding='utf-8') as fh:
                cached = json.load(fh)
            cached.setdefault('meta', {})['from_cache'] = True
            logger.info(f"[consolidated_report] serving from cache: {cache_path}")
            return jsonify(cached)
        except Exception:
            pass

    # -- start background job, return job_id for SSE polling -----------------
    job_id = _uuid.uuid4().hex[:16]

    def _run_job():
        import os as _os, sys as _sys
        _scripts_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'scripts')
        if _scripts_dir not in _sys.path:
            _sys.path.insert(0, _scripts_dir)
        try:
            from fetch_consolidated_report import (
                run_consolidated_report, register_progress, unregister_progress
            )
            from config import JIRA_PDT_FILTER_ID

            pt = register_progress(job_id)
            report = run_consolidated_report(
                build_ids    = raw,
                filter_id    = JIRA_PDT_FILTER_ID,
                traverse     = traverse,
                enrich_orbit = do_orbit,
                target_name  = target or None,
                progress     = pt,
                custom_jql   = custom_jql or None,
            )
            report_meta = report.setdefault('meta', {})
            report_meta['build_ids'] = raw
            if extracted_from_jql:
                report_meta['build_ids_extracted_from_jql'] = extracted_from_jql
            if include_axiom_metrics:
                try:
                    metrics = _fetch_axiom_stability_metrics_for_meta_ids(raw, taxonomy_path=axiom_taxonomy_path or None)
                    report['axiom_metrics'] = metrics
                    report.setdefault('summary', {})['axiom_metrics_found'] = sum(1 for v in metrics.values() if isinstance(v, dict) and v.get('matched'))
                    report.setdefault('summary', {})['axiom_metrics_requested'] = len(metrics)
                except Exception as me:
                    logger.warning(f"[consolidated_report] Axiom metrics enrichment failed: {me}")
                    report['axiom_metrics_error'] = str(me)
            # save to disk
            try:
                with open(cache_path, 'w', encoding='utf-8') as fh:
                    json.dump(report, fh, ensure_ascii=False)
                logger.info(f"[consolidated_report] saved: {cache_path}")
            except Exception as se:
                logger.warning(f"[consolidated_report] cache save failed: {se}")

            # store result for pickup
            report.setdefault('meta', {})['from_cache'] = False
            report.setdefault('meta', {})['job_id']     = job_id
            _JOB_RESULTS[job_id] = report

        except Exception as e:
            import traceback
            logger.error(f"[consolidated_report] job {job_id} error: {e}\n{traceback.format_exc()}")
            _JOB_RESULTS[job_id] = {'error': str(e)}
            try:
                from fetch_consolidated_report import get_progress
                pt = get_progress(job_id)
                if pt: pt.update(stage='error', message=str(e))
            except Exception:
                pass
        finally:
            pass  # keep progress alive for SSE to read final state

    import threading as _threading
    t = _threading.Thread(target=_run_job, daemon=True)
    t.start()
    return jsonify({'job_id': job_id, 'status': 'running'})




@login_required
def api_jira_traverse():
    """
    Traverse a single JIRA ticket through resolution notes and inward links
    to find the final ticket (with CR or dead end).
    Query params:
      key   - JIRA key to start from e.g. QSTABILITY-1234567
    """
    import time as _time
    key = (request.args.get('key') or '').strip()
    if not key:
        return jsonify({'error': 'key param required'}), 400
    try:
        import os as _os, sys as _sys
        _scripts_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'scripts')
        if _scripts_dir not in _sys.path:
            _sys.path.insert(0, _scripts_dir)
        from fetch_jira_by_build import (
            connect_jira, traverse_to_final_ticket,
        )
        from config import JIRA_USER, JIRA_PASSWORD, JIRA_SERVER_ENDPOINT
        if not JIRA_USER or not JIRA_PASSWORD:
            return jsonify({'error': 'JIRA credentials not configured'}), 500
        jira_obj = connect_jira(JIRA_USER, JIRA_PASSWORD, JIRA_SERVER_ENDPOINT)
        result   = traverse_to_final_ticket(jira_obj, key)
        return jsonify(result)
    except Exception as e:
        import traceback
        logger.error(f'[jira_traverse] {e}\n{traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500


    except Exception as e:
        logger.error(f'[LIVE STATUS VIEW] Error: {e}')
        return jsonify({'error': str(e)}), 500


def _fetch_grouped_cr_jira_context(cursor, target_name, search_value):
    u_table = fq_table_for_target(target_name, "unique_crs")
    j_table = fq_table_for_target(target_name, "jiras")
    o_table = fq_table_for_target(target_name, "openjiras")
    j_cr_exists = False
    try:
        cursor.execute(f"SHOW COLUMNS FROM {j_table} LIKE 'cr'")
        j_cr_exists = cursor.fetchone() is not None
    except Exception:
        pass
    search_tokens = _normalize_cr_search_tokens(search_value)
    if not search_tokens:
        return None
    where_one = " OR ".join(["cr = %s", "mapped_cr = %s"] * len(search_tokens))
    params_one = []
    for token in search_tokens:
        params_one.extend([token, token])
    cursor.execute(f"SELECT * FROM {u_table} WHERE {where_one} LIMIT 1", tuple(params_one))
    first_match = cursor.fetchone()
    if not first_match:
        return None
    canonical_mapped_cr = (first_match.get("mapped_cr") or first_match.get("cr") or "").strip()
    if not canonical_mapped_cr:
        canonical_mapped_cr = search_tokens[0]
    cursor.execute(f"SELECT * FROM {u_table} WHERE mapped_cr = %s ORDER BY cr", (canonical_mapped_cr,))
    cr_group_rows = cursor.fetchall() or []
    if not cr_group_rows:
        cr_group_rows = [first_match]
    linked_crs = []
    for row in cr_group_rows:
        cr_val = str(row.get("cr") or "").strip()
        if cr_val and cr_val not in linked_crs:
            linked_crs.append(cr_val)
    if not linked_crs:
        fallback_cr = str(first_match.get("cr") or "").strip()
        if fallback_cr:
            linked_crs.append(fallback_cr)
    j_mapped_crs_exists = False
    try:
        cursor.execute(f"SHOW COLUMNS FROM {j_table} LIKE 'mapped_crs'")
        j_mapped_crs_exists = cursor.fetchone() is not None
    except Exception:
        pass
    j_queries = []
    j_params = []
    if j_cr_exists:
        for cr_val in linked_crs:
            j_queries.append("cr = %s")
            j_params.append(cr_val)
            alt = cr_val.replace("CR", "") if cr_val.upper().startswith("CR") else f"CR{cr_val}"
            j_queries.append("cr = %s")
            j_params.append(alt)
    if j_mapped_crs_exists:
        for cr_val in linked_crs:
            j_queries.append("mapped_crs LIKE %s")
            j_params.append(f"%{cr_val}%")
        j_queries.append("mapped_crs LIKE %s")
        j_params.append(f"%{canonical_mapped_cr}%")
    if j_queries:
        j_where = " OR ".join(j_queries)
        cursor.execute(f"SELECT * FROM {j_table} WHERE {j_where}", tuple(j_params))
        j_rows = cursor.fetchall() or []
    else:
        j_rows = []
        o_rows = []
    if j_rows:
        tickets = list({str(r.get("stability_ticket") or "").strip() for r in j_rows if r.get("stability_ticket")})
        if tickets:
            placeholders = ", ".join(["%s"] * len(tickets))
            try:
                # Guard: openjiras may not exist for this target
                _o_name = o_table.replace("`", "")
                try:
                    _o_sch, _o_tbl = _o_name.split(".", 1)
                    cursor.execute("SELECT 1 FROM information_schema.tables WHERE table_schema=%s AND table_name=%s LIMIT 1", (_o_sch, _o_tbl))
                    _o_exists = cursor.fetchone() is not None
                except Exception:
                    _o_exists = True
                if _o_exists:
                    cursor.execute(f"SELECT * FROM {o_table} WHERE stability_ticket IN ({placeholders})", tuple(tickets))
                    o_rows = cursor.fetchall() or []
            except Exception:
                pass
    all_jiras = j_rows + o_rows
    unique_jiras = []
    seen_jira_keys = set()
    for row in all_jiras:
        key = (
            str(row.get("stability_ticket") or "").strip(),
            str(row.get("jira_date") or "").strip(),
            str(row.get("serial_no") or "").strip(),
            str(row.get("metabuild") or "").strip(),
            str(row.get("cr") or "").strip(),
                )
        if key in seen_jira_keys:
            continue
        seen_jira_keys.add(key)
        unique_jiras.append(row)
    # Sort by jira_date DESC so [:45] in template always gives the LATEST 45
    unique_jiras = sorted(unique_jiras, key=_jira_sort_key, reverse=True)
    latest_meta_jira = unique_jiras[0] if unique_jiras else None
    # Latest Meta Seen = metabuild from the most recently reported JIRA row.
    # Do not use highest meta ID; jira_date decides what was seen latest.
    latest_meta = (latest_meta_jira.get("metabuild") or latest_meta_jira.get("build_id") or "").strip() if latest_meta_jira else ""
    latest_meta_rows = [r for r in unique_jiras if str(r.get("metabuild") or r.get("build_id") or "").strip() == latest_meta] if latest_meta else []

    primary_cr_info = first_match
    devices    = sorted({(r.get("serial_no") or "").strip() for r in unique_jiras if r.get("serial_no")})
    test_teams = sorted({(r.get("test_team") or "").strip() for r in unique_jiras if r.get("test_team")})
    mcn_types  = sorted({(r.get("mcn") or "").strip() for r in unique_jiras if r.get("mcn")})
    jira_instances = sorted({(r.get("stability_ticket") or "").strip() for r in unique_jiras if r.get("stability_ticket")})
    return {
        "searched_cr": search_value.strip().upper(),
        "canonical_mapped_cr": canonical_mapped_cr,
        "primary_cr_info": primary_cr_info,
        "cr_group_rows": cr_group_rows,
        "linked_crs": linked_crs,
        "cr_jiras": unique_jiras,
        "latest_meta": latest_meta,
        "latest_meta_count": len(latest_meta_rows),
        "latest_meta_jira": latest_meta_jira,
        "devices": devices,
        "test_teams": test_teams,
        "mcn_types": mcn_types,
        "jira_instances": jira_instances,
    }


# ---------------------------------------------------------------------
# CR Title Exclude Keywords � save/load per Compute target
# ---------------------------------------------------------------------
def _get_cr_title_exclude(target_name):
    cfg = (_get_target_excel_config(target_name) or {}).get('cr_title_exclude', {})
    return {
        'enabled':  bool(cfg.get('enabled', False)),
        'keywords': [str(k).strip() for k in (cfg.get('keywords') or []) if str(k).strip()],
    }

@dashboard_bp.route('/api/dashboard/<string:target_name>/cr_title_exclude', methods=['GET'])
@login_required
def api_get_cr_title_exclude(target_name):
    return jsonify({'success': True, **_get_cr_title_exclude(target_name)})

@dashboard_bp.route('/api/dashboard/<string:target_name>/cr_title_exclude', methods=['POST'])
@login_required
def api_save_cr_title_exclude(target_name):
    try:
        payload  = request.get_json(force=True) or {}
        enabled  = bool(payload.get('enabled', False))
        keywords = [str(k).strip() for k in (payload.get('keywords') or []) if str(k).strip()]
        _update_target_excel_config(target_name, 'cr_title_exclude', {
            'enabled':  enabled,
            'keywords': keywords,
        })
        return jsonify({'success': True, 'enabled': enabled, 'keywords': keywords})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ---------------------------------------------------------------------
# API: PDT CRs
# ---------------------------------------------------------------------
@dashboard_bp.route("/api/dashboard/<string:target_name>/pdt_crs")
@login_required
def api_pdt_crs(target_name):
    conn = None
    cursor = None
    try:
        conn = get_mysql_connection_db()
        cursor = conn.cursor(dictionary=True)
        u_table = fq_table_for_target(target_name, "unique_crs")
        cursor.execute(f"SHOW COLUMNS FROM {u_table}")
        cols = {r["Field"] for r in (cursor.fetchall() or [])}

        def _col(name, alias=None):
            a = alias or name
            return f"`{name}` AS `{a}`" if name in cols else f"NULL AS `{a}`"

        def _coalesce_col(names, alias):
            available = [n for n in names if n in cols]
            if not available:
                return f"NULL AS `{alias}`"
            if len(available) == 1:
                return f"`{available[0]}` AS `{alias}`"
            return "COALESCE(" + ", ".join(f"`{n}`" for n in available) + f") AS `{alias}`"



        last_jira_col = next(
            (c for c in ("jira_date__last_instance", "qstability__last_instance", "jira_date_last_instance", "jira_date_last") if c in cols),
            None,
        )
        last_jira_sel = f"`{last_jira_col}` AS `jira_date_last`" if last_jira_col else "NULL AS `jira_date_last`"

        select_parts = ", ".join([
            _col("mapped_cr", "cr_id"),
            _col("cr_title"),
            _col("cr_area"),
            _col("cr_subsystem"),
            _col("cr_functionality"),
            _col("cr_occurrence"),
            _col("cr_age"),
            _col("cr_status"),
            _col("image", "cr_si"),
            _coalesce_col(("built_date", "build_date", "cr_built_date", "cr_date", "date_added__created"), "built_date"),
            _coalesce_col(("pdt_priority_tag", "pdt_tag", "pdt_priority"), "pdt_priority_tag"),
            last_jira_sel,
            _col("cr_category"),
        ])

        cursor.execute(
            f"""
            SELECT {select_parts}
            FROM {u_table}
            WHERE (cr_occurrence IS NULL OR LOWER(TRIM(cr_occurrence)) <> 'dup')
              AND (cr_category   IS NULL OR LOWER(TRIM(cr_category))   <> 'dup')
              AND (cr_category   IS NULL OR LOWER(TRIM(cr_category)) NOT LIKE '%invalid%')
            ORDER BY
                CAST(NULLIF(cr_occurrence, '') AS UNSIGNED) DESC,
                CAST(NULLIF(cr_age, '') AS UNSIGNED) DESC
            """
        )
        raw_rows = cursor.fetchall() or []

        def _num(v):
            try:
                return int(str(v or '').strip())
            except Exception:
                return 0

        seen = {}

        for r in raw_rows:
            cr_id = str(r.get('cr_id') or '').strip()
            if not cr_id:
                continue
            existing = seen.get(cr_id)
            if existing is None:
                seen[cr_id] = r
                continue
            curr_occ = _num(r.get('cr_occurrence'))
            prev_occ = _num(existing.get('cr_occurrence'))
            curr_age = _num(r.get('cr_age'))
            prev_age = _num(existing.get('cr_age'))
            if curr_occ > prev_occ or (curr_occ == prev_occ and curr_age > prev_age):
                seen[cr_id] = r

        rows = sorted(
            seen.values(),
            key=lambda x: (_num(x.get('cr_occurrence')), _num(x.get('cr_age'))),
            reverse=True,
        )


        is_compute_target = (get_bu_for_target(target_name) or '').upper() == 'COMPUTE'
        cr_tag_enabled = _is_compute_cr_tag_enabled_target(target_name)

        cr_tag_alias_groups = _load_compute_cr_tag_alias_config() if cr_tag_enabled else []
        include_cr_tags = cr_tag_enabled and str(request.args.get('include_cr_tags') or '').lower() in ('1', 'true', 'yes')

        cr_tag_cache_rows, cr_tag_cache_updated_at = _load_compute_cr_tag_cache(target_name) if cr_tag_enabled else ({}, None)
        tags_by_cr = {str(k).replace('CR','').replace('cr','').strip(): (v.get('cr_tags') or []) for k, v in (cr_tag_cache_rows or {}).items() if isinstance(v, dict)}

        if include_cr_tags and rows:

            try:
                from orbit_client import bulk_get_cr_tags
                cr_list = []
                for r in rows:
                    cr_digits = str(r.get('cr_id') or '').replace('CR', '').replace('cr', '').replace('-', '').strip()
                    if cr_digits.isdigit() and cr_digits not in cr_list:
                        cr_list.append(cr_digits)
                tags_by_cr = bulk_get_cr_tags(cr_list) if cr_list else {}
            except Exception:
                logger.warning('[PDT CR TAG] bulk tag fetch failed for %s', target_name, exc_info=True)
                tags_by_cr = {}


        # Resolve the most recent reported JIRA per CR from the mapped JIRA table.

        # unique_crs contains first/last dates, but the actual last JIRA key lives
        # in the jiras table. Handle both `cr` and comma-separated `mapped_crs`.
        last_jira_by_cr = {}
        try:
            j_table = fq_table_for_target(target_name, "jiras")
            cursor.execute(f"SHOW COLUMNS FROM {j_table}")
            j_cols = {r["Field"] for r in (cursor.fetchall() or [])}
            cr_col = "mapped_crs" if "mapped_crs" in j_cols else ("cr" if "cr" in j_cols else None)
            if cr_col and "stability_ticket" in j_cols:
                date_expr = "`jira_date`" if "jira_date" in j_cols else "NULL"
                cursor.execute(
                    f"""
                    SELECT `{cr_col}` AS cr_key, `stability_ticket`, {date_expr} AS jira_date
                    FROM {j_table}
                    WHERE `{cr_col}` IS NOT NULL AND TRIM(`{cr_col}`) <> ''
                    ORDER BY {date_expr} DESC, `stability_ticket` DESC
                    """
                )
                def _cr_keys(raw):
                    import re as _re
                    out = []
                    for part in _re.split(r"[,;\s]+", str(raw or "")):
                        val = part.strip()
                        if not val:
                            continue
                        compact = val.replace("-", "").upper()
                        variants = [compact]
                        if compact.startswith("CR"):
                            variants.append(compact[2:])
                        else:
                            variants.append("CR" + compact)
                        for item in variants:
                            if item and item not in out:
                                out.append(item)
                    return out
                for jr in cursor.fetchall() or []:
                    ticket = str(jr.get("stability_ticket") or "").strip()
                    if not ticket:
                        continue
                    for key in _cr_keys(jr.get("cr_key")):
                        if key not in last_jira_by_cr:
                            last_jira_by_cr[key] = {
                                "last_reported_jira": ticket,
                                "last_reported_date": jr.get("jira_date") or "",
                            }
        except Exception:
            pass
        import datetime as _dt

        from email.utils import parsedate
        MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        def fmt_date(v):
            if not v: return ""
            if isinstance(v, (_dt.date, _dt.datetime)):
                return f"{v.day} {MONTHS[v.month - 1]} {v.year}"
            s = str(v).strip()
            import re
            if re.match(r'^\d{1,2}\s[A-Za-z]{3}\s\d{4}$', s): return s
            try:
                t = parsedate(s)
                if t: return f"{t[2]} {MONTHS[t[1] - 1]} {t[0]}"
            except Exception: pass
            try:
                date_part = s.split(' ')[0].split('T')[0]
                parts = date_part.split('-')
                if len(parts) == 3:
                    d, m, y = int(parts[2]), int(parts[1]) - 1, parts[0]
                    if 0 <= m <= 11:
                        return f"{d} {MONTHS[m]} {y}"
            except Exception:
                pass
            return s


        # -- CR Title Exclude (Compute only) ------------------------------
        if is_compute_target:
            _excl_cfg = _get_cr_title_exclude(target_name)
            if _excl_cfg['enabled'] and _excl_cfg['keywords']:
                _kws = [k.lower() for k in _excl_cfg['keywords']]
                rows = [
                    r for r in rows
                    if not any(kw in str(r.get('cr_title') or '').lower() for kw in _kws)
                ]

        clean = []
        for r in rows:

            cr_key_raw = str(r.get("cr_id") or "").replace("-", "").upper()
            lookup_keys = [cr_key_raw]
            if cr_key_raw.startswith("CR"):
                lookup_keys.append(cr_key_raw[2:])
            elif cr_key_raw:
                lookup_keys.append("CR" + cr_key_raw)
            last_info = next((last_jira_by_cr.get(k) for k in lookup_keys if last_jira_by_cr.get(k)), {})

            row = {}
            for k, v in r.items():
                if k in ("built_date", "jira_date_last"):
                    row[k] = fmt_date(v)
                elif v is None:
                    row[k] = ""
                elif isinstance(v, (_dt.date, _dt.datetime)):
                    row[k] = str(v)
                else:
                    row[k] = v
            row["last_reported_jira"] = last_info.get("last_reported_jira") or ""
            row["last_reported_date"] = fmt_date(last_info.get("last_reported_date")) or row.get("jira_date_last", "")

            if cr_tag_enabled:
                cr_digits = str(row.get('cr_id') or '').replace('CR', '').replace('cr', '').replace('-', '').strip()
                cr_tags = tags_by_cr.get(cr_digits) or tags_by_cr.get('CR' + cr_digits) or []
                cr_tag_group, cr_tag_alias = _match_compute_cr_tag_aliases(cr_tags, cr_tag_alias_groups)
                row['cr_tags'] = cr_tags
                row['cr_tag'] = cr_tag_group
                row['cr_tag_alias'] = cr_tag_alias
            clean.append(row)

        return jsonify({"success": True, "rows": clean, "is_compute": is_compute_target, "cr_tag_enabled": cr_tag_enabled, "cr_tags_included": include_cr_tags, "cr_tag_cache_loaded": bool(cr_tag_cache_rows), "cr_tag_cache_updated_at": cr_tag_cache_updated_at, "cr_tag_alias_groups": cr_tag_alias_groups})




    except Exception as e:
        return jsonify({"success": False, "message": str(e), "rows": []}), 500
    finally:
        if cursor: cursor.close()
        if conn:   conn.close()




# -- PDT CR Software Image Ready Date API -----------------------------------
@dashboard_bp.route("/api/dashboard/<string:target_name>/cr_si_ready_dates", methods=["POST"])
@login_required
def api_pdt_cr_si_ready_dates(target_name):
    """Return Orbit Software Image ReadyDate values for CR/SI pairs.

    Request body: {items:[{cr:"4574261", si:"LPAICP.FW.1.0"}, ...]}
    Response: {success:true, rows:{"4574261|LPAICP.FW.1.0":{ready_date:"19 Jun 2026", ...}}}
    Missing/empty ReadyDate is returned as "NA".
    """
    import re as _re
    from datetime import datetime as _dt

    def _norm_cr(v):
        return _re.sub(r"\D", "", str(v or ""))

    def _norm_si(v):
        return _re.sub(r"[^A-Z0-9]+", "", str(v or "").upper())

    def _fmt_ready(v):
        if not v:
            return "NA"
        s = str(v).strip()
        if not s:
            return "NA"
        for fmt in ("%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
            try:
                d = _dt.strptime(s.split('.')[0], fmt)
                return d.strftime("%d %b %Y")
            except Exception:
                pass
        return s

    try:
        payload = request.get_json(silent=True) or {}
        items = payload.get("items") or []
        if not isinstance(items, list):
            items = []
        cleaned = []
        for item in items[:100]:
            if not isinstance(item, dict):
                continue
            cr = _norm_cr(item.get("cr") or item.get("cr_id"))
            si = str(item.get("si") or item.get("cr_si") or "").strip()
            if cr:
                cleaned.append({"cr": cr, "si": si})
        if not cleaned:
            return jsonify({"success": True, "rows": {}})

        unique_crs = sorted({item["cr"] for item in cleaned})
        try:
            from orbit_client import bulk_query_cr_software_images
            cache = bulk_query_cr_software_images(unique_crs, batch_size=100) if unique_crs else {}
        except Exception:
            logger.warning("[PDT CR SI ReadyDate] bulk Orbit SIR query failed; falling back to empty ReadyDate", exc_info=True)
            cache = {cr: [] for cr in unique_crs}

        out = {}

        for item in cleaned:
            cr, si = item["cr"], item["si"]
            key = f"{cr}|{si}"
            sirs = [x for x in cache.get(cr, []) if isinstance(x, dict)]

            wanted = _norm_si(si)
            matched = None
            if wanted:
                for sir in sirs:
                    name = str(sir.get("SoftwareImageName") or sir.get("Name") or sir.get("SoftwareImage") or "").strip()
                    if _norm_si(name) == wanted:
                        matched = sir
                        break
            if matched is None and not wanted and len(sirs) == 1:
                matched = sirs[0]
            all_ready = []
            for sir in sirs:
                name = str(sir.get("SoftwareImageName") or sir.get("Name") or sir.get("SoftwareImage") or "").strip()
                all_ready.append({
                    "software_image": name,
                    "ready_date": _fmt_ready(sir.get("ReadyDate")),
                    "raw_ready_date": sir.get("ReadyDate"),
                })
            out[key] = {
                "cr": cr,
                "si": si,
                "ready_date": _fmt_ready(matched.get("ReadyDate") if matched else None),
                "raw_ready_date": matched.get("ReadyDate") if matched else None,
                "matched_si": str((matched or {}).get("SoftwareImageName") or (matched or {}).get("Name") or ""),
                "all_ready_dates": all_ready,
            }
        return jsonify({"success": True, "rows": out})
    except Exception as e:
        logger.warning("[PDT CR SI ReadyDate] failed for target=%s", target_name, exc_info=True)
        return jsonify({"success": False, "message": str(e), "rows": {}}), 500


# -- Open JIRAs API ----------------------------------------------------------
@dashboard_bp.route("/api/dashboard/<string:target_name>/open_jiras")

@login_required
def api_open_jiras(target_name):
    """Return open jiras with area bucketing and date range filter."""
    conn = None; cursor = None
    try:
        date_from = (request.args.get("date_from") or "").strip()
        date_to   = (request.args.get("date_to")   or "").strip()

        conn   = get_mysql_connection_db()
        cursor = conn.cursor(dictionary=True)
        # Try fq_table_for_target first; fall back to direct DB search
        o_table = None
        try:
            o_table = fq_table_for_target(target_name, "openjiras")
        except Exception as _fq_err:
            pass

        if not o_table:
            tbl_pattern = target_name.lower().replace('-', '_').replace(' ', '_') + '_openjiras'
            cursor.execute(
                "SELECT TABLE_SCHEMA, TABLE_NAME FROM information_schema.TABLES "
                "WHERE TABLE_NAME = %s LIMIT 1",
                (tbl_pattern,)
            )
            _row = cursor.fetchone()
            if _row:
                _s = _row.get('TABLE_SCHEMA') or _row.get('table_schema', '')
                _n = _row.get('TABLE_NAME')   or _row.get('table_name', '')
                o_table = f"`{_s}`.`{_n}`"
            else:
                return jsonify({"success": True, "rows": [], "area_summary": [],
                                "notice": "Open JIRAs table not available for this target."})

        # Guard: openjiras table may not exist for all targets
        def _tbl_exists(fq_name):
            name = fq_name.replace("`", "")
            try:
                schema, table = name.split(".", 1)
            except ValueError:
                return True
            cursor.execute(
                "SELECT 1 FROM information_schema.tables "
                "WHERE table_schema=%s AND table_name=%s LIMIT 1",
                (schema, table),
            )
            return cursor.fetchone() is not None

        if not _tbl_exists(o_table):
            logger.info(f"[OPEN JIRAS API] table missing for target={target_name}, returning empty")
            return jsonify({"success": True, "rows": [], "area_summary": [],
                            "notice": "Open JIRAs table not available for this target."})

        where_clauses = []
        params = []
        if date_from:
            where_clauses.append("`jira_date` >= %s")
            params.append(date_from)
        if date_to:
            where_clauses.append("`jira_date` <= %s")
            params.append(date_to + " 23:59:59")
        where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

        cursor.execute(
            f"SELECT stability_ticket, jira_date, cr_area, cr_current_ticket, "
            f"jira_title, jira_category, status, test_team, jira_reporter, "
            f"metabuild, jira_component "
            f"FROM {o_table} {where_sql} ORDER BY jira_date DESC",
            params or None
        )
        raw = cursor.fetchall() or []

        import datetime as _dt, re as _re
        MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        def fmt_dt(v):
            if not v: return ""
            if isinstance(v, (_dt.date, _dt.datetime)):
                return f"{v.day} {MONTHS[v.month-1]} {v.year}"
            return str(v)[:10]

        # -- Area bucketing logic ------------------------------------------
        # Priority: cr_area field ? title keyword match ? jira_category ? 'Other'
        # Area bucketing logic
        # Step 1: cr_current_ticket prefix
        # Step 2: cr_area field from DB
        # Step 3: jira_title / component / category keyword scan
        # Step 4: cr_current_ticket empty + no keyword match -> APPS
        # NOTE: WCNSS/CNSS/CNSSDEBUG/BT/BTFM all -> WLAN. BOOT/XBL/UEFI -> APPS.
        CR_TICKET_PREFIXES = [
            ("ADSP",       "ADSP"),
            ("CDSP",       "CDSP"),
            ("TZ",         "TZ"),
            ("MODEM",      "Modem"),
            ("VIDEO",      "Video"),
            ("CNSSDEBUG",  "WConnect"),
            ("CNSS",       "WConnect"),
            ("WCNSS",      "WConnect"),
            ("WLAN",       "WConnect"),
            ("BTFM",       "WConnect"),
            ("BT",         "WConnect"),
            ("APPS",       "APPS"),
        ]
        AREA_KEYWORDS = {
            "WConnect": ["wconnect", "wcnss", "cnss", "cnssdebug", "wlan", "wifi", "wi-fi", "btfm", "bluetooth", "wireless"],
            "Modem":   ["modem", "mpss", "ril", "data call", "lte", "5g", "nr", "ims", "qmi"],
            "Video":   ["video", "venc", "vdec", "venus", "codec"],
            "TZ":      ["trustzone", "trust zone", "qsee"],
            "ADSP":    ["adsp", "audio", "qdsp"],
            "CDSP":    ["cdsp", "compute dsp"],
            "Camera":  ["camera", "csiphy", "csid", "ife", "isp"],
            "Display": ["display", "mdss", "dpu", "dsi", "panel"],
            "Sensors": ["sensor", "sensors", "ssc", "slpi"],
            "BOOT":    ["boot", "xbl", "uefi", "abl"],
            "APPS":    ["apps", "apss", "gcc", "kernel", "android", "framework", "userspace"],
        }

        def bucket_area(row):
            cr_ticket = (row.get("cr_current_ticket") or "").strip().upper()

            # Step 1: cr_current_ticket prefix is most reliable
            if cr_ticket:
                for prefix, bucket in CR_TICKET_PREFIXES:
                    if cr_ticket.startswith(prefix):
                        return bucket

            # Step 2: keyword scan on title + component + category. Do this before
            # DB cr_area so WCNSS/Modem/Video/TZ titles are not hidden under APPS.
            title    = (row.get("jira_title")     or "").lower()
            comp     = (row.get("jira_component")  or "").lower()
            cat      = (row.get("jira_category")   or "").lower()
            combined = title + " " + comp + " " + cat
            for bucket, keywords in AREA_KEYWORDS.items():
                if any(k in combined for k in keywords):
                    return bucket

            # Step 3: cr_area field from DB as fallback, normalized to display buckets
            area = (row.get("cr_area") or "").strip()
            area_upper = area.upper().replace('-', '_').replace(' ', '_')
            if area and area.lower() not in ("", "none", "null", "n/a"):
                if any(t in area_upper for t in ("WCONNECT", "WCNSS", "CNSS", "WLAN", "WIFI", "BT", "BTFM")):
                    return "WConnect"
                if "MODEM" in area_upper or area_upper == "MPSS":
                    return "Modem"
                if "VIDEO" in area_upper:
                    return "Video"
                if area_upper in ("TZ", "TRUSTZONE", "TRUST_ZONE"):
                    return "TZ"
                if area_upper in ("APPS", "APSS"):
                    return "APPS"
                return area

            # Step 4: unknown open JIRAs should be explicit instead of inflating APPS.
            return "Other"


        rows = []
        for r in raw:
            area = bucket_area(r)
            rows.append({
                "stability_ticket":  r.get("stability_ticket") or "",
                "jira_date":         fmt_dt(r.get("jira_date")),
                "area":              area,
                "cr_current_ticket": r.get("cr_current_ticket") or "",
                "jira_title":        r.get("jira_title") or "",
                "jira_category":     r.get("jira_category") or "",
                "status":            r.get("status") or "",
                "test_team":         r.get("test_team") or "",
                "jira_reporter":     r.get("jira_reporter") or "",
                "metabuild":         r.get("metabuild") or "",
            })

        # Area summary
        from collections import Counter
        area_counts = Counter(r["area"] for r in rows)
        area_summary = sorted(
            [{"area": a, "count": c} for a, c in area_counts.items()],
            key=lambda x: -x["count"]
        )
        area_summary.append({"area": "Total", "count": len(rows)})

                # AUTO-only grouping helpers for the redesigned Open JIRAs tab.
        bu_key = (get_bu_for_target(target_name) or '').upper()
        is_auto_bu = bu_key in ('AUTO', 'AUTOMOTIVE')

        def _crash_type(title):
            t = (title or '').lower()
            # Process: contains ProcessDump OR ProcessCrash OR QNX OR Undetermined
            if any(k in t for k in ('processdump', 'processcrash', 'process_crash', 'qnx', 'undetermined')):
                return 'Process'
            # SSR: does NOT contain above AND contains sleep OR ssr
            if any(k in t for k in ('sleep', 'ssr', 'subsystem restart')):
                return 'SSR'
            # System: everything else (no ProcessDump/ProcessCrash/QNX/sleep/ssr/Undetermined)
            return 'System'

        def _domain(row):
            hay = ' '.join(str(row.get(k) or '') for k in (
                'jira_title', 'jira_component', 'jira_category', 'test_team', 'metabuild', 'cr_area'
            )).upper()
            # Prefer explicit domain tokens; keep this heuristic conservative.
            if any(k in hay for k in ('ADAS', 'ADP', 'RIDE', 'VISION', 'CAMERA', 'CVP')):
                return 'ADAS'
            if any(k in hay for k in ('FLEX', 'PVM', 'SURROUND', 'PARK', 'VIP')):
                return 'FLEX'
            if any(k in hay for k in ('IVI', 'COCKPIT', 'INFOTAINMENT', 'DISPLAY', 'AUDIO')):
                return 'IVI'
            # AUTO tickets without explicit ADAS/FLEX tokens are IVI by default.
            return 'IVI'

        for row in rows:
            row['crash_type'] = _crash_type(row.get('jira_title', ''))
            row['domain'] = _domain(row) if is_auto_bu else ''

        crash_type_counts = {}
        domain_counts = {}
        domain_crash_counts = {}
        for row in rows:
            ct = row['crash_type']
            crash_type_counts[ct] = crash_type_counts.get(ct, 0) + 1
            if is_auto_bu:
                dom = row.get('domain') or 'Unassigned'
                domain_counts[dom] = domain_counts.get(dom, 0) + 1
                domain_crash_counts.setdefault(dom, {})[ct] = domain_crash_counts.setdefault(dom, {}).get(ct, 0) + 1

        crash_summary = [{"type": t, "count": c} for t, c in sorted(crash_type_counts.items(), key=lambda x: -x[1])]
        domain_summary = [
            {"domain": d, "count": c, "crash_types": domain_crash_counts.get(d, {})}
            for d, c in sorted(domain_counts.items(), key=lambda x: (x[0] == 'Unassigned', x[0]))
        ]

        return jsonify({
            "success": True,
            "rows": rows,
            "area_summary": area_summary,
            "crash_summary": crash_summary,
            "domain_summary": domain_summary,
            "is_auto_bu": is_auto_bu,
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e), "rows": [], "area_summary": []}), 500
    finally:
        if cursor: cursor.close()
        if conn:   conn.close()



@dashboard_bp.route("/api/dashboard/<string:target_name>/pdt_cr_tags/start", methods=["POST"])
@login_required
def api_pdt_cr_tags_start(target_name):
    if not _is_compute_cr_tag_enabled_target(target_name):
        return jsonify({'success': False, 'message': 'CR TAG fetch is disabled for this target.'}), 403
    body = request.get_json(force=True) or {}
    crs = []
    seen = set()
    for item in body.get('crs') or []:
        cr = str(item or '').upper().replace('CR', '').replace('-', '').strip()
        if cr.isdigit() and cr not in seen:
            crs.append(cr)
            seen.add(cr)
    if not crs:
        return jsonify({'success': False, 'message': 'No CRs provided.'}), 400

    job_id = uuid.uuid4().hex
    _PDT_CR_TAG_JOBS[job_id] = {
        'success': True,
        'state': 'queued',
        'done': 0,
        'total': len(crs),
        'batch': 0,
        'rows': {},
        'message': 'Queued CR TAG fetch...',
        'updated_at': time.time(),
    }

    def _run():
        try:
            from orbit_client import bulk_query_cr_tags, bulk_get_cr_tags
            groups = _load_compute_cr_tag_alias_config()
            def _progress(done, total, batch):
                job = _PDT_CR_TAG_JOBS.get(job_id) or {}
                job.update({
                    'state': 'running',
                    'done': int(done),
                    'total': int(total),
                    'batch': int(batch),
                    'message': f'Fetching Orbit CR TAGs... {done}/{total} CRs',
                    'updated_at': time.time(),
                })
                _PDT_CR_TAG_JOBS[job_id] = job
            try:
                # Estimate from CR count only. Query API batches up to 100 CRs, typically ~8-12s per batch.
                est_total = max(10, int(((len(crs) + 99) // 100) * 10))
                job = _PDT_CR_TAG_JOBS.get(job_id) or {}
                job.update({
                    'state': 'running',
                    'done': 0,
                    'total': len(crs),
                    'message': f'Fetching {len(crs)} CR TAGs from Orbit. Estimated time: ~{est_total}s.',
                    'estimated_seconds': est_total,
                    'updated_at': time.time(),
                })
                _PDT_CR_TAG_JOBS[job_id] = job
                def _progress_est(done, total, batch):
                    pct_done = float(done) / max(1, total)
                    remaining = max(0, int(round(est_total * (1 - pct_done))))
                    job = _PDT_CR_TAG_JOBS.get(job_id) or {}
                    job.update({
                        'state': 'running',
                        'done': int(done),
                        'total': int(total),
                        'batch': int(batch),
                        'message': f'Fetching Orbit CR TAGs... {done}/{total} CRs. ETA ~{remaining}s',
                        'estimated_seconds': est_total,
                        'updated_at': time.time(),
                    })
                    _PDT_CR_TAG_JOBS[job_id] = job
                tags_by_cr = bulk_query_cr_tags(crs, batch_size=100, progress_callback=_progress_est)

            except Exception:
                logger.warning('[PDT CR TAG JOB] query/run failed, using fallback', exc_info=True)
                tags_by_cr = bulk_get_cr_tags(crs)
                _progress(len(crs), len(crs), 1)

            rows = {}
            for cr in crs:
                tags = tags_by_cr.get(cr) or tags_by_cr.get('CR' + cr) or []
                group, alias = _match_compute_cr_tag_aliases(tags, groups)
                rows[cr] = {'cr_tags': tags, 'cr_tag': group, 'cr_tag_alias': alias}
            cache_payload = _save_compute_cr_tag_cache(target_name, rows)
            _PDT_CR_TAG_JOBS[job_id].update({

                'state': 'done',
                'done': len(crs),
                'total': len(crs),
                'rows': rows,
                'cache_path': _compute_cr_tag_cache_path(target_name),
                'cache_updated_at': cache_payload.get('updated_at'),
                'message': 'CR TAG fetch completed and saved.',
                'updated_at': time.time(),
            })

        except Exception as exc:
            logger.warning('[PDT CR TAG JOB] failed', exc_info=True)
            _PDT_CR_TAG_JOBS[job_id].update({
                'success': False,
                'state': 'error',
                'message': str(exc),
                'updated_at': time.time(),
            })

    import threading
    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'success': True, 'job_id': job_id, 'total': len(crs)})


@dashboard_bp.route("/api/dashboard/<string:target_name>/pdt_cr_tags/status/<string:job_id>")
@login_required
def api_pdt_cr_tags_status(target_name, job_id):
    job = _PDT_CR_TAG_JOBS.get(job_id)
    if not job:
        return jsonify({'success': False, 'state': 'missing', 'message': 'CR TAG job not found.'}), 404
    return jsonify(job)


@dashboard_bp.route("/api/dashboard/<string:target_name>/cr_tag_aliases", methods=["GET", "POST"])
@login_required
def api_compute_cr_tag_aliases(target_name):

    if not _is_compute_cr_tag_enabled_target(target_name):
        return jsonify({'success': False, 'message': 'CR TAG aliases are disabled for this target.'}), 403
    if request.method == 'GET':
        return jsonify({'success': True, 'groups': _load_compute_cr_tag_alias_config()})
    data = request.get_json(force=True) or {}
    groups = _save_compute_cr_tag_alias_config(data.get('groups') or [])
    return jsonify({'success': True, 'groups': groups, 'message': 'CR TAG aliases saved.'})


# =============================================================================
# PDT TAGGING
#   GET  /api/dashboard/<target>/pdt_tags  - read existing PDT tags via orbit_client

#   POST /api/dashboard/<target>/pdt_tag   - add/remove tags via PDT_Stats.exe subprocess
# =============================================================================

PDT_STATS_EXE = r"C:\Dropbox\DATA_MINING\PDT_Stats.exe"


@dashboard_bp.route("/api/dashboard/<string:target_name>/pdt_tags", methods=["GET", "POST"])
@login_required
def api_pdt_tags_get(target_name):
    """
    Return existing PDT tags on the given CRs.
    GET  ?crs=1234,5678   (small lists)
    POST {crs: [1234, 5678]}  (large lists � avoids URL length limit)
    Response: { tags: ["PDT_P1", ...] }
    """
    from orbit_client import bulk_get_cr_tags
    if request.method == "POST":
        body    = request.get_json(force=True) or {}
        cr_list = [str(c).strip() for c in (body.get("crs") or []) if str(c).strip().isdigit()]
    else:
        crs_raw = (request.args.get("crs") or "").strip()
        cr_list = [c.strip() for c in crs_raw.split(",") if c.strip().isdigit()]
    if not cr_list:
        return jsonify({"tags": []})
    # Cap at 200 CRs to avoid overloading Orbit API
    cr_list = cr_list[:200]
    try:
        all_tags_map = bulk_get_cr_tags(cr_list)
        tags_set = set()
        for tag_list in all_tags_map.values():
            for t in tag_list:
                if t.strip():
                    tags_set.add(t.strip())
        all_tags  = sorted(tags_set)
        pdt_tags  = [t for t in all_tags if "PDT" in t.upper()]
        return jsonify({"tags": all_tags, "pdt_tags": pdt_tags})
    except Exception as e:
        logger.warning(f"[PDT TAGS GET] {e}")
        return jsonify({"tags": [], "error": str(e)})


@dashboard_bp.route("/api/dashboard/<string:target_name>/pdt_tag", methods=["POST"])
@login_required
def api_pdt_tag_post(target_name):
    """
    Add / remove PDT tags on a list of CRs by calling PDT_Stats.exe via subprocess.
    PDT_Stats.exe handles its own Orbit auth internally.

    Body: { crs: ["4520954", ...], add_tag: "PDT_P1", remove_tags: ["PDT_P2"] }
    Response: { success: bool, message: str }

    PDT_Stats.exe command:
      PDT_Stats.exe  CRS="CR111,CR222"  CrTags="PDT_P1"  RemoveCrTags="PDT_P2"
    """
    import subprocess, re as _re

    data        = request.get_json(force=True) or {}
    cr_list     = [str(c).strip() for c in (data.get("crs") or []) if str(c).strip().isdigit()]
    add_tag     = str(data.get("add_tag") or "").strip()
    remove_tags = [str(t).strip() for t in (data.get("remove_tags") or []) if str(t).strip()]

    if not cr_list:
        return jsonify({"success": False, "message": "No valid CR numbers provided."})
    if not add_tag and not remove_tags:
        return jsonify({"success": False, "message": "Specify at least one tag to Add or Remove."})

    tag_re = _re.compile(r"^[A-Za-z0-9_]{3,}$")
    if add_tag and not tag_re.match(add_tag):
        return jsonify({"success": False, "message": f"Invalid add tag: '{add_tag}'"})
    for rt in remove_tags:
        if not tag_re.match(rt):
            return jsonify({"success": False, "message": f"Invalid remove tag: '{rt}'"})

    # Build CRS= param: PDT_Stats.exe expects "CR1234,CR5678" format
    crs_param = ",".join("CR" + c for c in cr_list)

    # Build command
    cmd = [PDT_STATS_EXE, f'CRS="{crs_param}"']
    if add_tag:
        cmd.append(f'CrTags="{add_tag}"')
    if remove_tags:
        cmd.append(f'RemoveCrTags="{",".join(remove_tags)}"')

    logger.info(f"[PDT TAG] Running: {' '.join(cmd)}")

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=120,
            cwd=r"C:\Dropbox\DATA_MINING"
        )
        output = (result.stdout or "") + (result.stderr or "")
        logger.info(f"[PDT TAG] Output: {output[:500]}")

        # PDT_Stats.exe prints "Added tag(s) successfully" / "Removed tag(s) successfully"
        added_count   = output.count("Added tag(s) successfully")
        removed_count = output.count("Removed tag(s) successfully")

        if result.returncode != 0 and added_count == 0 and removed_count == 0:
            return jsonify({"success": False, "message": f"PDT_Stats.exe error: {output[:300]}"})

        parts = []
        if add_tag:    parts.append(f"'{add_tag}' added to {added_count}/{len(cr_list)} CR(s)")
        if remove_tags: parts.append(f"{remove_tags} removed from {removed_count}/{len(cr_list)} CR(s)")
        msg = "  |  ".join(parts) or "Done"

        return jsonify({"success": True, "message": msg})

    except subprocess.TimeoutExpired:
        return jsonify({"success": False, "message": "PDT_Stats.exe timed out (>120s)"})
    except FileNotFoundError:
        return jsonify({"success": False, "message": f"PDT_Stats.exe not found at {PDT_STATS_EXE}"})
    except Exception as e:
        logger.error(f"[PDT TAG POST] {e}")
        return jsonify({"success": False, "message": str(e)})
