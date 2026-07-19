"""
mcp_mtbf_server.py
------------------
Standalone MCP server that exposes MTBF trend data from the live_status_view
page to external apps (e.g. test automation, dashboards, CI pipelines).

Data source: same JSON files used by live_status_view_api.py
  <DATA_ROOT>/managed_excel/AUTO/MTBF/<TARGET>/mtbf_<view>.json

Transport: stdio (default) - works with any MCP client.

Usage:
    py -3 mcp_mtbf_server.py                        # stdio transport
    py -3 mcp_mtbf_server.py --transport sse --port 8765   # SSE transport

Tools exposed:
    get_mtbf_trend      - Get MTBF trend rows for a target + view
    get_mtbf_chart_data - Get chart-ready series (label/hours/crashes/mtbf)
    list_mtbf_targets   - List all targets that have MTBF data
    get_mtbf_summary    - Get latest MTBF, total hours, total crashes for a target
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Bootstrap .env
# ---------------------------------------------------------------------------
try:
    from dotenv import load_dotenv as _ld
    _ld(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"), override=True)
except Exception:
    pass

# ---------------------------------------------------------------------------
# MCP SDK
# ---------------------------------------------------------------------------
try:
    from mcp.server.fastmcp import FastMCP
except ImportError:
    print(
        "ERROR: mcp package not installed. Run: pip install mcp",
        file=sys.stderr,
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Config - same paths as live_status_view_api.py + Dropbox Excel source
# ---------------------------------------------------------------------------
_DATA_ROOT = os.environ.get(
    "PDTBUDDY_DATA_ROOT",
    r"\\Sphere\pdtqipl_internal\PDTBuddy",
)
_MTBF_BASE = os.path.join(_DATA_ROOT, "managed_excel", "AUTO", "MTBF")
_LOCAL_FALLBACK = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "adas_mtbf")

_DROPBOX_ROOT = os.environ.get("PDTBUDDY_DROPBOX_ROOT", r"C:\Dropbox")
_AUTO_EXCEL_PATH = os.environ.get(
    "PDTBUDDY_AUTO_EXCEL_PATH",
    os.path.join(_DROPBOX_ROOT, "4.8.0.9_Auto.xlsx"),
)
_AUTO_JSON_DIR = os.environ.get(
    "PDTBUDDY_AUTO_JSON_DIR",
    os.path.join(_DATA_ROOT, "managed_excel", "AUTO", "Automotive", "Gen4.5"),
)
_AUTO_JSON_PATH = os.environ.get(
    "PDTBUDDY_AUTO_JSON_PATH",
    os.path.join(_AUTO_JSON_DIR, "4.8.0.9_Auto.json"),
)

_VIEWS = ["ADAS", "IVI", "FLEX"]

# Target slug -> folder name mapping (matches live_status_view_api.py)
_FOLDER_MAP = {
    "NORD_HQX": "Nord_HQX",
    "NORD_HGY": "Nord_HGY",
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _target_slug(target_name: str) -> str:
    return str(target_name or "").strip().upper().replace(".", "_")


def _mtbf_folder(target_name: str) -> str:
    slug = _target_slug(target_name)
    folder = _FOLDER_MAP.get(slug, slug)
    return os.path.join(_MTBF_BASE, folder)


def _mtbf_json_path(target_name: str, view: str) -> str:
    view_clean = str(view or "ADAS").strip().upper()
    if view_clean not in _VIEWS:
        view_clean = "ADAS"
    # Primary: network share
    primary = os.path.join(_mtbf_folder(target_name), f"mtbf_{view_clean.lower()}.json")
    if os.path.exists(primary):
        return primary
    # Fallback: local data dir
    local_dir = os.path.join(_LOCAL_FALLBACK, target_name.lower())
    return os.path.join(local_dir, f"mtbf_{view_clean.lower()}.json")


def _load_mtbf(target_name: str, view: str) -> Dict[str, Any]:
    path = _mtbf_json_path(target_name, view)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if isinstance(data, dict):
                data.setdefault("target", target_name)
                data.setdefault("view", view.upper())
                data.setdefault("rows", [])
                return data
        except Exception as exc:
            return {"error": f"Failed to read {path}: {exc}", "rows": []}
    return {"target": target_name, "view": view.upper(), "rows": [], "note": "No data file found"}


def _num(v: Any) -> float:
    try:
        return float(str(v or "0").replace(",", "").strip())
    except Exception:
        return 0.0


def _rows_to_chart(rows: List[Dict], crash_types: Optional[List[str]] = None) -> List[Dict]:
    """Convert raw MTBF rows to chart-ready series."""
    if crash_types is None:
        crash_types = ["system", "ssr", "process"]
    out = []
    for r in rows or []:
        meta_id = str(r.get("meta_id") or "").strip()
        if not meta_id:
            continue
        hours = _num(r.get("hours"))
        sys_c = int(_num(r.get("system_crashes")))
        ssr_c = int(_num(r.get("ssr_crashes")))
        proc_c = int(_num(r.get("process_crashes")))
        total_c = 0
        if "system" in crash_types:
            total_c += sys_c
        if "ssr" in crash_types:
            total_c += ssr_c
        if "process" in crash_types:
            total_c += proc_c
        mtbf = _num(r.get("mtbf"))
        if not mtbf and hours and total_c:
            mtbf = round(hours / total_c, 2)
        out.append({
            "meta_id":         meta_id,
            "date":            str(r.get("date") or ""),
            "hours":           round(hours, 2),
            "system_crashes":  sys_c,
            "ssr_crashes":     ssr_c,
            "process_crashes": proc_c,
            "crashes":         total_c,
            "mtbf":            round(mtbf, 2),
            "s_no":            int(_num(r.get("s_no"))),
        })
    return out


def _list_available_targets() -> List[Dict[str, Any]]:
    """Scan MTBF base folder for targets that have data files."""
    targets = []
    base = _MTBF_BASE
    if not os.path.isdir(base):
        # Try local fallback
        base = _LOCAL_FALLBACK
    if not os.path.isdir(base):
        return []
    for folder in sorted(os.listdir(base)):
        folder_path = os.path.join(base, folder)
        if not os.path.isdir(folder_path):
            continue
        views_found = []
        for view in _VIEWS:
            fpath = os.path.join(folder_path, f"mtbf_{view.lower()}.json")
            if os.path.exists(fpath):
                views_found.append(view)
        if views_found:
            targets.append({
                "target":      folder,
                "views":       views_found,
                "folder_path": folder_path,
            })
    return targets


def _resolve_excel_path(excel_path: str = "") -> str:
    """Return a safe Excel path. Defaults to C:\\Dropbox\\4.8.0.9_Auto.xlsx."""
    path = str(excel_path or "").strip() or _AUTO_EXCEL_PATH
    path = os.path.abspath(os.path.expanduser(path))
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")
    if not path.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("Only .xlsx/.xlsm Excel files are supported")
    return path


def _clean_text(value: str) -> str:
    text = str(value or "").replace("\u00a0", " ").replace("\t", " ").strip()
    return " ".join(text.split())


def _normalize_header(value: Any, fallback_index: int) -> str:
    header = _clean_text(str(value or ""))
    if not header:
        header = f"Column {fallback_index}"
    header = header.replace(".", "")
    header = re.sub(r"[^A-Za-z0-9]+", "_", header).strip("_").lower()
    return header or f"column_{fallback_index}"


def _program_key(sheet_name: str) -> str:
    """Convert Excel sheet names like '8775(Flex)' to '8775 (Flex)'."""
    name = _clean_text(sheet_name)
    match = re.match(r"^(\d+)\s*\(([^)]+)\)$", name)
    if match:
        return f"{match.group(1)} ({match.group(2).strip()})"
    return name


def _normalize_excel_value(value: Any, column_name: str = "") -> Any:
    """Normalize Excel cell values for JSON/API consumers.

    Required mappings:
      '-'       -> 'NA'
      '>600'    -> 600, '>260' -> 260, etc.
      '200*'    -> 200, '150*' -> 150
      Devices '-' / 'devices:-' -> null
    """
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return int(value) if isinstance(value, float) and value.is_integer() else value

    text = _clean_text(str(value))
    if not text:
        return None

    column_clean = _clean_text(column_name).lower().replace(" ", "_")
    text_lower = text.lower().replace(" ", "")
    if column_clean == "devices" and text in {"-", "NA", "N/A"}:
        return None
    if text_lower in {"devices:-", "device:-"}:
        return None
    if text == "-":
        return "NA"

    numeric_match = re.fullmatch(r">\s*(-?\d+(?:\.\d+)?)", text)
    if not numeric_match:
        numeric_match = re.fullmatch(r"(-?\d+(?:\.\d+)?)\*", text)
    if numeric_match:
        number = float(numeric_match.group(1))
        return int(number) if number.is_integer() else number

    plain_number = re.fullmatch(r"-?\d+(?:\.\d+)?", text)
    if plain_number:
        number = float(text)
        return int(number) if number.is_integer() else number

    return text


def _resolve_json_path(json_path: str = "") -> str:
    """Return the Auto program JSON path. Defaults to Sphere Automotive Gen4.5."""
    path = str(json_path or "").strip() or _AUTO_JSON_PATH
    return os.path.abspath(os.path.expanduser(path))


def _load_auto_payload_from_json(json_path: str = "") -> Dict[str, Any]:
    """Load the already-generated Auto JSON. Does not read Excel."""
    path = _resolve_json_path(json_path)
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Auto JSON not found: {path}. Run refresh_auto_program_json once to create it."
        )
    with open(path, "r", encoding="utf-8") as fh:
        payload = json.load(fh)
    if not isinstance(payload, dict) or not isinstance(payload.get("programs"), dict):
        raise ValueError(f"Invalid Auto JSON format in {path}; expected top-level 'programs' object")
    payload.setdefault("metadata", {})
    payload["metadata"].setdefault("json_path", path)
    return payload


def _drop_rows_with_null_date(programs: Dict[str, List[Dict[str, Any]]]) -> Dict[str, List[Dict[str, Any]]]:
    """Remove rows where the canonical date field is null/blank."""
    cleaned: Dict[str, List[Dict[str, Any]]] = {}
    for program, rows in (programs or {}).items():
        cleaned[program] = [
            row for row in (rows or [])
            if isinstance(row, dict) and row.get("date") not in (None, "")
        ]
    return cleaned


def _load_auto_programs_from_json(json_path: str = "") -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, Any]]:
    """Load Auto programs from JSON only."""
    payload = _load_auto_payload_from_json(json_path)
    programs = _drop_rows_with_null_date(payload.get("programs") or {})
    metadata = payload.get("metadata") or {}
    metadata.setdefault("json_path", _resolve_json_path(json_path))
    return programs, metadata


def _load_auto_programs_from_excel(excel_path: str = "") -> Tuple[Dict[str, List[Dict[str, Any]]], Dict[str, Any]]:
    """Read the Auto Excel workbook and return {'8775 (Flex)': [rows...], ...}."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl") from exc

    path = _resolve_excel_path(excel_path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    programs: Dict[str, List[Dict[str, Any]]] = {}
    sheets: List[Dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        program = _program_key(sheet_name)
        header_values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
        headers = [_normalize_header(value, idx) for idx, value in enumerate(header_values, start=1)]
        rows: List[Dict[str, Any]] = []

        for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not values or all(v is None or _clean_text(str(v)) == "" for v in values):
                continue
            row: Dict[str, Any] = {"excel_row": row_number}
            for idx, header in enumerate(headers):
                value = values[idx] if idx < len(values) else None
                row[header] = _normalize_excel_value(value, header)
            if row.get("date") in (None, ""):
                continue
            rows.append(row)

        programs[program] = rows
        sheets.append({"sheet_name": sheet_name, "program": program, "rows": len(rows)})

    metadata = {
        "source_excel": path,
        "sheet_count": len(workbook.sheetnames),
        "sheets": sheets,
        "updated_at": datetime.now().isoformat(timespec="seconds"),
    }
    try:
        workbook.close()
    except Exception:
        pass
    return programs, metadata


def _find_program_key(programs: Dict[str, List[Dict[str, Any]]], program: str) -> Optional[str]:
    query = _clean_text(program).lower()
    if not query:
        return None
    for key in programs:
        key_lower = key.lower()
        digits = "".join(re.findall(r"\d+", key))
        if query == key_lower or query == digits or query in key_lower:
            return key
    return None


def _available_sp_response(programs: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, str]]:
    out = []
    for key in programs:
        digits = "".join(re.findall(r"\d+", key))
        domain_match = re.search(r"\(([^)]+)\)", key)
        out.append({
            "sp": digits or key,
            "program": key,
            "domain": domain_match.group(1) if domain_match else "",
        })
    return out


def _program_summary(rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    total_hours = round(sum(_num(r.get("hours")) for r in rows or []), 2)
    total_crashes = int(sum(_num(r.get("crashes")) for r in rows or []))
    published_mtbf = [
        _num(r.get("mtbf")) for r in (rows or [])
        if isinstance(r.get("mtbf"), (int, float)) or str(r.get("mtbf") or "").strip().replace(".", "", 1).isdigit()
    ]
    latest = rows[-1] if rows else {}
    return {
        "row_count": len(rows or []),
        "latest_date": latest.get("date") or "",
        "latest_build": latest.get("build_s") or latest.get("builds") or latest.get("build") or "",
        "latest_mtbf": latest.get("mtbf"),
        "total_hours": total_hours,
        "total_crashes": total_crashes,
        "calculated_overall_mtbf": round(total_hours / total_crashes, 2) if total_crashes else total_hours,
        "max_published_mtbf": max(published_mtbf) if published_mtbf else None,
    }


# ---------------------------------------------------------------------------
# MCP Server
# ---------------------------------------------------------------------------
mcp = FastMCP(
    name="PDTBuddy MTBF + Auto Excel Server",
    instructions=(
        "Provides MTBF trend data and Auto Excel program data from C:\\Dropbox. "
        "Use get_auto_program_json or get_auto_program_info to fetch Excel-backed "
        "program information for 8775, 7255, 8255, 8650, and 8620. Existing MTBF "
        "tools are also available."
    ),
)


@mcp.tool()
def list_available_sps(json_path: str = "") -> Dict[str, Any]:
    """List all SPs available in the Auto Gen4.5 JSON.

    Request params:
        json_path: optional override path; usually leave blank.

    Response has available_sps like 8775/Flex, 7255/IVI, etc.
    """
    try:
        programs, metadata = _load_auto_programs_from_json(json_path)
        return {
            "ok": True,
            "message": "Available SPs fetched successfully.",
            "available_sps": _available_sp_response(programs),
            "count": len(programs),
            "json_path": metadata.get("json_path") or _resolve_json_path(json_path),
        }
    except Exception as exc:
        return {"ok": False, "message": f"Unable to list available SPs: {exc}", "available_sps": []}


@mcp.tool()
def get_sp_info(sp: str, json_path: str = "", last_n: int = 0, include_summary: bool = True) -> Dict[str, Any]:
    """Get Auto Gen4.5 information for a requested SP.

    Request params:
        sp: SP/program requested by external tool. Examples: '8775', '7255', '8255', '8650', '8620'.
        json_path: optional override path; usually leave blank.
        last_n: optional latest N rows only. 0 means all rows.
        include_summary: include summary metrics.

    If SP is not available, response ok=false with message and available_sps.
    """
    try:
        programs, metadata = _load_auto_programs_from_json(json_path)
        key = _find_program_key(programs, sp)
        if not key:
            return {
                "ok": False,
                "message": f"Requested SP '{sp}' is not available in Auto Gen4.5 data.",
                "requested_sp": sp,
                "available_sps": _available_sp_response(programs),
                "rows": [],
                "row_count": 0,
            }
        rows = programs[key]
        if last_n and last_n > 0:
            rows = rows[-last_n:]
        response = {
            "ok": True,
            "message": f"SP '{sp}' data fetched successfully.",
            "requested_sp": sp,
            "resolved_program": key,
            "rows": rows,
            "row_count": len(rows),
            "json_path": metadata.get("json_path") or _resolve_json_path(json_path),
            "generated_at": metadata.get("generated_at") or metadata.get("updated_at") or "",
        }
        if include_summary:
            response["summary"] = _program_summary(rows)
        return response
    except Exception as exc:
        return {"ok": False, "message": f"Unable to fetch SP '{sp}': {exc}", "requested_sp": sp, "rows": [], "row_count": 0}


@mcp.tool()
def search_sp_info(sp: str, query: str = "", json_path: str = "", limit: int = 50) -> Dict[str, Any]:
    """Search rows for a requested SP.

    Request params:
        sp: SP/program requested by external tool, e.g. '8775'.
        query: text to search in that SP rows.
        json_path: optional override path; usually leave blank.
        limit: maximum rows to return.
    """
    try:
        programs, metadata = _load_auto_programs_from_json(json_path)
        key = _find_program_key(programs, sp)
        if not key:
            return {
                "ok": False,
                "message": f"Requested SP '{sp}' is not available in Auto Gen4.5 data.",
                "requested_sp": sp,
                "available_sps": _available_sp_response(programs),
                "matches": [],
                "match_count": 0,
            }
        q = _clean_text(query).lower()
        matches: List[Dict[str, Any]] = []
        max_rows = max(1, int(limit or 50))
        for row in programs.get(key, []):
            haystack = json.dumps(row, ensure_ascii=False, default=str).lower()
            if not q or q in haystack:
                matches.append({"program": key, **row})
                if len(matches) >= max_rows:
                    break
        return {
            "ok": True,
            "message": f"Search completed for SP '{sp}'.",
            "requested_sp": sp,
            "resolved_program": key,
            "query": query,
            "matches": matches,
            "match_count": len(matches),
            "truncated": len(matches) >= max_rows,
            "json_path": metadata.get("json_path") or _resolve_json_path(json_path),
        }
    except Exception as exc:
        return {"ok": False, "message": f"Unable to search SP '{sp}': {exc}", "requested_sp": sp, "matches": [], "match_count": 0}


@mcp.tool()
def list_auto_programs(json_path: str = "") -> Dict[str, Any]:
    """List available Auto programs from the already-generated JSON.

    Args:
        json_path: Optional JSON path. Defaults to Sphere managed_excel/AUTO/Automotive/Gen4.5/4.8.0.9_Auto.json.
    """
    try:
        programs, metadata = _load_auto_programs_from_json(json_path)
        return {
            "ok": True,
            "programs": list(programs.keys()),
            "count": len(programs),
            "json_path": metadata.get("json_path") or _resolve_json_path(json_path),
            "source_excel": metadata.get("source_excel") or "",
            "generated_at": metadata.get("generated_at") or metadata.get("updated_at") or "",
        }
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


@mcp.tool()
def refresh_auto_program_json(excel_path: str = "", output_path: str = "") -> Dict[str, Any]:
    """One-time/explicit refresh: convert Dropbox Excel workbook into JSON.

    Use this only when the Excel has changed and you want to regenerate the JSON.
    Normal public MCP tools read JSON and do not touch Excel.

    The JSON has this top-level shape:
        {'programs': {'8775 (Flex)': [...], '7255 (IVI)': [...], ...}}

    Value normalization includes '-' -> 'NA', '>600' -> 600, '200*' -> 200,
    and Devices '-' / 'devices:-' -> null.
    """
    try:
        programs, metadata = _load_auto_programs_from_excel(excel_path)
        programs = _drop_rows_with_null_date(programs)
        out_path = _resolve_json_path(output_path)
        metadata["generated_at"] = datetime.now().isoformat(timespec="seconds")
        metadata["json_path"] = out_path
        payload = {"programs": programs, "metadata": metadata}
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        tmp_path = out_path + ".tmp"
        with open(tmp_path, "w", encoding="utf-8") as fh:
            json.dump(payload, fh, ensure_ascii=False, indent=2)
        os.replace(tmp_path, out_path)
        return {
            "ok": True,
            "json_path": out_path,
            "source_excel": metadata.get("source_excel"),
            "programs": list(programs.keys()),
            "program_count": len(programs),
            "metadata": metadata,
        }
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


@mcp.tool()
def get_auto_program_json(json_path: str = "") -> Dict[str, Any]:
    """Return the already-generated Auto program JSON without reading Excel.

    Args:
        json_path: Optional JSON path. Defaults to Sphere managed_excel/AUTO/Automotive/Gen4.5/4.8.0.9_Auto.json.
    """
    try:
        payload = _load_auto_payload_from_json(json_path)
        payload["programs"] = _drop_rows_with_null_date(payload.get("programs") or {})
        return {"ok": True, **payload}
    except Exception as exc:
        return {"ok": False, "error": str(exc), "programs": {}}


@mcp.tool()
def get_auto_program_info(program: str, json_path: str = "", last_n: int = 0) -> Dict[str, Any]:
    """Fetch related rows for one Auto program from the generated JSON.

    Args:
        program: Program identifier. Examples: '8775', '8775 (Flex)', '7255'.
        json_path: Optional JSON path. Defaults to Sphere managed_excel/AUTO/Automotive/Gen4.5/4.8.0.9_Auto.json.
        last_n: Return only last N rows (0 = all).
    """
    try:
        programs, metadata = _load_auto_programs_from_json(json_path)
        key = _find_program_key(programs, program)
        if not key:
            return {
                "ok": False,
                "error": f"Program '{program}' not found",
                "available_programs": list(programs.keys()),
            }
        rows = programs[key]
        if last_n and last_n > 0:
            rows = rows[-last_n:]
        return {
            "ok": True,
            "program": key,
            "rows": rows,
            "row_count": len(rows),
            "json_path": metadata.get("json_path") or _resolve_json_path(json_path),
            "source_excel": metadata.get("source_excel") or "",
            "generated_at": metadata.get("generated_at") or metadata.get("updated_at") or "",
        }
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


@mcp.tool()
def search_auto_program_info(
    program: str = "",
    query: str = "",
    json_path: str = "",
    limit: int = 50,
) -> Dict[str, Any]:
    """Search Auto JSON rows by program and text query.

    Args:
        program: Optional program identifier, e.g. '8775'. Empty searches all programs.
        query: Case-insensitive text to search across row values.
        json_path: Optional JSON path. Defaults to Sphere managed_excel/AUTO/Automotive/Gen4.5/4.8.0.9_Auto.json.
        limit: Maximum matched rows to return.
    """
    try:
        programs, metadata = _load_auto_programs_from_json(json_path)
        keys = [program]
        if program:
            found = _find_program_key(programs, program)
            if not found:
                return {"ok": False, "error": f"Program '{program}' not found", "available_programs": list(programs.keys())}
            keys = [found]
        else:
            keys = list(programs.keys())

        q = _clean_text(query).lower()
        matches: List[Dict[str, Any]] = []
        max_rows = max(1, int(limit or 50))
        for key in keys:
            for row in programs.get(key, []):
                haystack = json.dumps(row, ensure_ascii=False, default=str).lower()
                if not q or q in haystack:
                    matches.append({"program": key, **row})
                    if len(matches) >= max_rows:
                        return {
                            "ok": True,
                            "matches": matches,
                            "match_count": len(matches),
                            "truncated": True,
                            "json_path": metadata.get("json_path") or _resolve_json_path(json_path),
                            "source_excel": metadata.get("source_excel") or "",
                        }
        return {
            "ok": True,
            "matches": matches,
            "match_count": len(matches),
            "truncated": False,
            "json_path": metadata.get("json_path") or _resolve_json_path(json_path),
            "source_excel": metadata.get("source_excel") or "",
        }
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


@mcp.tool()
def list_mtbf_targets() -> Dict[str, Any]:
    """List all targets that have MTBF trend data available.

    Returns a list of target names with their available views (ADAS/IVI/FLEX).
    """
    targets = _list_available_targets()
    return {
        "ok":      True,
        "count":   len(targets),
        "targets": targets,
        "base_path": _MTBF_BASE,
    }


@mcp.tool()
def get_mtbf_trend(
    target_name: str,
    view: str = "ADAS",
    last_n: int = 0,
) -> Dict[str, Any]:
    """Get MTBF trend rows for a target and view.

    Args:
        target_name: Target name, e.g. 'Nord_HQX' or 'Nord_HGY'.
        view:        MTBF view - one of ADAS, IVI, FLEX. Default: ADAS.
        last_n:      Return only the last N rows (0 = all rows).

    Returns:
        ok, target, view, rows (list of dicts with s_no/date/meta_id/hours/
        system_crashes/ssr_crashes/process_crashes/total_crashes/mtbf),
        row_count, updated_at.
    """
    view_clean = str(view or "ADAS").strip().upper()
    if view_clean not in _VIEWS:
        return {"ok": False, "error": f"Invalid view '{view}'. Must be one of {_VIEWS}"}

    data = _load_mtbf(target_name, view_clean)
    rows = data.get("rows") or []

    if last_n and last_n > 0:
        rows = rows[-last_n:]

    return {
        "ok":         True,
        "target":     target_name,
        "view":       view_clean,
        "views":      _VIEWS,
        "rows":       rows,
        "row_count":  len(rows),
        "updated_at": data.get("updated_at") or "",
        "note":       data.get("note") or "",
    }


@mcp.tool()
def get_mtbf_chart_data(
    target_name: str,
    view: str = "ADAS",
    crash_types: str = "system,ssr,process",
    last_n: int = 0,
) -> Dict[str, Any]:
    """Get chart-ready MTBF trend series for a target and view.

    Args:
        target_name:  Target name, e.g. 'Nord_HQX'.
        view:         MTBF view - ADAS, IVI, or FLEX. Default: ADAS.
        crash_types:  Comma-separated crash types to include in MTBF calculation.
                      Options: system, ssr, process. Default: 'system,ssr,process'.
        last_n:       Return only the last N data points (0 = all).

    Returns:
        ok, target, view, chart_data (list of dicts with meta_id/date/hours/
        crashes/system_crashes/ssr_crashes/process_crashes/mtbf), updated_at.
    """
    view_clean = str(view or "ADAS").strip().upper()
    if view_clean not in _VIEWS:
        return {"ok": False, "error": f"Invalid view '{view}'. Must be one of {_VIEWS}"}

    ct_list = [c.strip().lower() for c in str(crash_types or "system,ssr,process").split(",") if c.strip()]
    if not ct_list:
        ct_list = ["system", "ssr", "process"]

    data = _load_mtbf(target_name, view_clean)
    rows = data.get("rows") or []

    if last_n and last_n > 0:
        rows = rows[-last_n:]

    chart_data = _rows_to_chart(rows, ct_list)

    return {
        "ok":          True,
        "target":      target_name,
        "view":        view_clean,
        "crash_types": ct_list,
        "chart_data":  chart_data,
        "point_count": len(chart_data),
        "updated_at":  data.get("updated_at") or "",
    }


@mcp.tool()
def get_mtbf_summary(
    target_name: str,
    view: str = "ADAS",
) -> Dict[str, Any]:
    """Get a quick MTBF summary for a target - latest MTBF, total hours, total crashes.

    Args:
        target_name: Target name, e.g. 'Nord_HQX'.
        view:        MTBF view - ADAS, IVI, or FLEX. Default: ADAS.

    Returns:
        ok, target, view, latest_meta_id, latest_mtbf, latest_date,
        total_hours, total_crashes, row_count, trend (list of meta_id+mtbf pairs),
        updated_at.
    """
    view_clean = str(view or "ADAS").strip().upper()
    if view_clean not in _VIEWS:
        return {"ok": False, "error": f"Invalid view '{view}'. Must be one of {_VIEWS}"}

    data = _load_mtbf(target_name, view_clean)
    rows = data.get("rows") or []
    chart = _rows_to_chart(rows)

    latest = chart[-1] if chart else {}
    total_hours   = round(sum(_num(r.get("hours"))   for r in chart), 2)
    total_crashes = int(sum(_num(r.get("crashes"))   for r in chart))

    trend = [
        {"meta_id": r["meta_id"], "mtbf": r["mtbf"], "date": r["date"]}
        for r in chart
    ]

    return {
        "ok":            True,
        "target":        target_name,
        "view":          view_clean,
        "latest_meta_id": latest.get("meta_id") or "",
        "latest_mtbf":   latest.get("mtbf") or 0,
        "latest_date":   latest.get("date") or "",
        "total_hours":   total_hours,
        "total_crashes": total_crashes,
        "row_count":     len(rows),
        "trend":         trend,
        "updated_at":    data.get("updated_at") or "",
    }


@mcp.tool()
def get_mtbf_all_views(
    target_name: str,
) -> Dict[str, Any]:
    """Get MTBF summary for all views (ADAS, IVI, FLEX) for a target in one call.

    Args:
        target_name: Target name, e.g. 'Nord_HQX'.

    Returns:
        ok, target, views dict with ADAS/IVI/FLEX each containing
        row_count, latest_mtbf, latest_meta_id, total_hours, total_crashes, updated_at.
    """
    result: Dict[str, Any] = {"ok": True, "target": target_name, "views": {}}
    for view in _VIEWS:
        data  = _load_mtbf(target_name, view)
        rows  = data.get("rows") or []
        chart = _rows_to_chart(rows)
        latest = chart[-1] if chart else {}
        result["views"][view] = {
            "row_count":      len(rows),
            "latest_meta_id": latest.get("meta_id") or "",
            "latest_mtbf":    latest.get("mtbf") or 0,
            "latest_date":    latest.get("date") or "",
            "total_hours":    round(sum(_num(r.get("hours"))   for r in chart), 2),
            "total_crashes":  int(sum(_num(r.get("crashes"))   for r in chart)),
            "updated_at":     data.get("updated_at") or "",
            "has_data":       len(rows) > 0,
        }
    return result


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="PDTBuddy MTBF + Auto Excel MCP Server")
    parser.add_argument(
        "--transport",
        choices=["stdio", "sse"],
        default="stdio",
        help="MCP transport (default: stdio)",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=8765,
        help="Port for SSE transport (default: 8765)",
    )
    parser.add_argument(
        "--host",
        default="0.0.0.0",
        help="Host for SSE transport (default: 0.0.0.0)",
    )
    parser.add_argument(
        "--refresh-auto-json",
        action="store_true",
        help="Generate/refresh Auto JSON from Excel once, then exit unless --serve-after-refresh is set.",
    )
    parser.add_argument(
        "--serve-after-refresh",
        action="store_true",
        help="Start the MCP server after --refresh-auto-json completes.",
    )
    args = parser.parse_args()

    if args.refresh_auto_json:
        result = refresh_auto_program_json()
        print(json.dumps(result, ensure_ascii=False, indent=2), file=sys.stderr)
        if not result.get("ok"):
            sys.exit(1)
        if not args.serve_after_refresh:
            sys.exit(0)

    if args.transport == "sse":
        # host/port are constructor-level settings in this version of FastMCP
        # Re-create the server with the requested host/port before running.
        print(f"Starting PDTBuddy MTBF + Auto Excel MCP server (SSE) on {args.host}:{args.port}", file=sys.stderr)
        sse_mcp = FastMCP(
            name=mcp.name,
            instructions=mcp.instructions if hasattr(mcp, 'instructions') else None,
            host=args.host,
            port=args.port,
        )
        # Re-register all tools from the original mcp instance
        for tool in mcp._tool_manager.list_tools():
            sse_mcp._tool_manager.add_tool(tool)
        sse_mcp.run(transport="sse")
    else:
        print("Starting PDTBuddy MTBF + Auto Excel MCP server (stdio)", file=sys.stderr)
        mcp.run(transport="stdio")
