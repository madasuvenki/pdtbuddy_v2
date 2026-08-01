import json
import logging
import os
import re
import threading
import time
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from dashboard_routes import _get_target_excel_config, _normalize_excel_path
from dashboard_common import get_bu_for_target

logger = logging.getLogger(__name__)

_PDTBUDDY_DATA_ROOT = os.environ.get(
    'PDTBUDDY_DATA_ROOT',
    r'\\Sphere\pdtqipl_internal\PDTBuddy'
)
_LIVE_STATUS_ROOT = os.path.join(_PDTBUDDY_DATA_ROOT, 'live_status_publish')
_INDEX_FILE       = os.path.join(_LIVE_STATUS_ROOT, 'jobs_index.json') # lightweight listing
_JOBS_FILE        = os.path.join(_LIVE_STATUS_ROOT, 'jobs.json')       # legacy compat alias

# Local fallback when network share is read-only
_LOCAL_ROOT       = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'live_status_publish')
_LOCAL_INDEX_FILE = os.path.join(_LOCAL_ROOT, 'jobs_index.json')

def _safe_path_part(value: str) -> str:
    text = str(value or '').strip() or 'UNKNOWN'
    return re.sub(r'[^A-Za-z0-9._-]+', '_', text).strip('._') or 'UNKNOWN'


def _target_dir_parts(target_name: str) -> Tuple[str, str]:
    target = str(target_name or '').strip() or 'UNKNOWN_TARGET'
    try:
        bu = str(get_bu_for_target(target) or '').strip().upper() or 'UNKNOWN_BU'
    except Exception:
        bu = 'UNKNOWN_BU'
    return _safe_path_part(bu), _safe_path_part(target)


def target_live_status_dir(target_name: str) -> str:
    """Dedicated Live Status folder: <root>/<BU>/<TARGET>."""
    root, _ = _storage_root()
    bu, target = _target_dir_parts(target_name)
    return os.path.join(root, bu, target)


# -- Sidecar files ------------------------------------------------------------
#
# Volatile data (JQL, JIRA cache, excluded JIRAs, SWPDT builds) lives in
# small sidecar files keyed by BUILD-ID so each build/merge has its own file.
# The main job JSON is NEVER written for these operations.
#
# Build-ID key rules:
#   Single build : short numeric part of build name  e.g. "01792"
#                  extracted from SA8797P_ADAS.HQX.5.1.7.0-01792-STD.INT-1
#   Merged builds: sorted short IDs joined with "+"  e.g. "01792+01800"
#   Fallback     : sanitised full build name (no path separators)
#
# File paths:
#   Auto BU  : <root>/<BU>/<TARGET>/builds/<BUILD_KEY>_<DOMAIN>_report.json
#              <root>/<BU>/<TARGET>/builds/<BUILD_KEY>_<DOMAIN>_swpdt.json
#              e.g.  AUTO/NORD_HQX/builds/01792_ADAS_report.json
#                    AUTO/NORD_HQX/builds/01792+01800_ADAS_report.json  (merged)
#
#   Non-Auto : <root>/<BU>/<TARGET>/builds/<BUILD_KEY>_report.json
#              <root>/<BU>/<TARGET>/builds/<BUILD_KEY>_swpdt.json
#              e.g.  COMPUTE/KONA/builds/00270_report.json
#
# On job delete: entire builds/ directory is wiped for that target.

_AUTO_BU_KEYS = {"AUTO", "AUTOMOTIVE"}
_AUTO_DOMAINS = ("ADAS", "FLEX", "IVI")


def _is_auto_bu_target(target_name: str) -> bool:
    """True for Automotive/NORD targets that have ADAS/FLEX/IVI domains."""
    try:
        bu = str(get_bu_for_target(target_name) or "").strip().upper()
        return bu in _AUTO_BU_KEYS
    except Exception:
        return False


def _builds_dir(target_name: str) -> str:
    return os.path.join(target_live_status_dir(target_name), "builds")


def _short_build_id(build_full: str) -> str:
    """
    Extract the short numeric build number from a full build path/name.
    SA8797P_ADAS.HQX.5.1.7.0-01792-STD.INT-1  ->  01792
    Kona.LA.1.0-00270-STD.INT-1               ->  00270
    Falls back to the last path segment sanitised for use in a filename.
    """
    if not build_full:
        return "UNKNOWN"
    # Take last path segment (strip UNC/share prefix)
    seg = str(build_full).replace("\\", "/").rstrip("/")
    seg = seg.split("/")[-1] if "/" in seg else seg
    # Extract leading zero-padded number after a dash: -01792-
    m = re.search(r"-0*(\d{3,6})-", seg)
    if m:
        return m.group(1).zfill(5)
    # Fallback: sanitise the segment so it is safe as a filename
    return re.sub(r"[^A-Za-z0-9._+-]", "_", seg)[:60]


def _build_key(build_full_or_list) -> str:
    """
    Return a stable, filename-safe key for a single build or a merged list.
    Single : "01792"
    Merged : "01792+01800"  (sorted ascending)
    """
    if isinstance(build_full_or_list, (list, tuple)):
        ids = sorted({_short_build_id(b) for b in build_full_or_list if b})
        return "+".join(ids) if ids else "UNKNOWN"
    return _short_build_id(str(build_full_or_list or ""))


def _report_sidecar_path(target_name: str, build_key: str,
                         domain: Optional[str] = None) -> str:
    """
    Per-build JIRA report sidecar.
    build_key : "01792" | "01792+01800"  (from _build_key())
    domain    : "ADAS"|"FLEX"|"IVI" for Auto BU; None for non-Auto

    Auto BU  : builds/01792_ADAS_report.json
    Non-Auto : builds/00270_report.json
    """
    bdir = _builds_dir(target_name)
    key  = re.sub(r"[^A-Za-z0-9._+-]", "_", build_key or "UNKNOWN")
    if _is_auto_bu_target(target_name) and domain:
        dom = str(domain).strip().upper()
        if dom not in _AUTO_DOMAINS:
            dom = "ADAS"
        return os.path.join(bdir, f"{key}_{dom}_report.json")
    return os.path.join(bdir, f"{key}_report.json")


def _swpdt_sidecar_path(target_name: str,
                        domain: Optional[str] = None) -> str:
    """
    Domain-level SWPDT build list sidecar (one per domain, not per build).
    Written by background refresh thread every 2-3 h.

    Auto BU  : builds/NORD_HQX_ADAS_swpdt.json
    Non-Auto : builds/KONA_swpdt.json
    """
    bdir = _builds_dir(target_name)
    _, tgt = _target_dir_parts(target_name)
    if _is_auto_bu_target(target_name) and domain:
        dom = str(domain).strip().upper()
        if dom not in _AUTO_DOMAINS:
            dom = "ADAS"
        return os.path.join(bdir, f"{tgt}_{dom}_swpdt.json")
    return os.path.join(bdir, f"{tgt}_swpdt.json")


def _safe_write(path: str, data: Dict[str, Any]) -> None:
    """Atomic-ish write with retry + direct-write fallback (network share safe)."""
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp      = path + ".tmp"
    lock_key = "sc:" + path
    with _job_lock(lock_key):
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump(data, fh, indent=2)
        last_err = None
        for attempt in range(3):
            try:
                os.replace(tmp, path)
                last_err = None
                break
            except (PermissionError, OSError) as exc:
                last_err = exc
                time.sleep(0.05 * (attempt + 1))
        if last_err is not None:
            logger.warning("[SIDECAR] os.replace failed (%s); direct-write fallback", last_err)
            try:
                with open(path, "w", encoding="utf-8") as fh:
                    json.dump(data, fh, indent=2)
            finally:
                try:
                    if os.path.exists(tmp):
                        os.remove(tmp)
                except OSError:
                    pass


def _read_sidecar(path: str) -> Dict[str, Any]:
    """Read any sidecar by absolute path; return {} if missing or corrupt."""
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as fh:
                return json.load(fh)
    except Exception:
        pass
    return {}


def _write_sidecar(path: str, updates: Dict[str, Any],
                   extra_meta: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Merge updates into existing sidecar at path and write."""
    data = _read_sidecar(path)
    data.update(updates)
    if extra_meta:
        data.update(extra_meta)
    data["updated_at"] = _utc_now()
    _safe_write(path, data)
    return data


def _delete_sidecars(target_name: str) -> None:
    """Delete entire builds/ sidecar directory for a target (called on job delete)."""
    bdir = _builds_dir(target_name)
    if not os.path.isdir(bdir):
        return
    import shutil
    try:
        shutil.rmtree(bdir)
        logger.info("[SIDECAR] Deleted builds dir %s", bdir)
    except OSError as exc:
        # rmtree failed - try file-by-file
        logger.warning("[SIDECAR] rmtree failed (%s); trying file-by-file", exc)
        try:
            for fname in os.listdir(bdir):
                if fname.endswith(".json") or fname.endswith(".tmp"):
                    try:
                        os.remove(os.path.join(bdir, fname))
                    except OSError:
                        pass
        except OSError:
            pass


# -- Public sidecar API (called by routes) ------------------------------------

def get_report_sidecar(target_name: str,
                       build_key: str,
                       domain: Optional[str] = None) -> Dict[str, Any]:
    """Read per-build JIRA report sidecar {jql, report_cache, excluded_jiras}."""
    return _read_sidecar(_report_sidecar_path(target_name, build_key, domain))


def set_sidecar_jql(target_name: str,
                    build_key: str,
                    domain: Optional[str],
                    jql: str) -> Dict[str, Any]:
    """Persist JQL to per-build report sidecar -- never touches main job JSON."""
    path = _report_sidecar_path(target_name, build_key, domain)
    return _write_sidecar(path, {"jql": (jql or "").strip()},
                          extra_meta={"build_key": build_key})


def set_sidecar_report_cache(target_name: str,
                              build_key: str,
                              domain: Optional[str],
                              cache: Dict[str, Any]) -> Dict[str, Any]:
    """Persist JIRA report cache -- never touches main job JSON."""
    path = _report_sidecar_path(target_name, build_key, domain)
    return _write_sidecar(path, {"report_cache": cache or {}},
                          extra_meta={"build_key": build_key})


def set_sidecar_exclusions(target_name: str,
                            build_key: str,
                            domain: Optional[str],
                            excluded_keys: List[str]) -> Dict[str, Any]:
    """Persist excluded JIRA keys -- never touches main job JSON."""
    cleaned = sorted({str(k or "").strip().upper()
                      for k in (excluded_keys or []) if str(k or "").strip()})
    path = _report_sidecar_path(target_name, build_key, domain)
    return _write_sidecar(path, {"excluded_jiras": cleaned},
                          extra_meta={"build_key": build_key})


def set_sidecar_swpdt_builds(target_name: str,
                              domain: Optional[str],
                              builds: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Persist domain-level SWPDT build list.
    Written by background refresh thread -- no build_key, no per-build file.
    Auto BU  : NORD_HQX_ADAS_swpdt.json
    Non-Auto : KONA_swpdt.json
    """
    path = _swpdt_sidecar_path(target_name, domain)
    return _write_sidecar(path, {"builds": builds or {}})


def get_swpdt_sidecar(target_name: str,
                      domain: Optional[str] = None) -> Dict[str, Any]:
    """Read domain-level SWPDT sidecar {builds, updated_at}."""
    return _read_sidecar(_swpdt_sidecar_path(target_name, domain))


def _job_primary_target(job: Dict[str, Any]) -> str:
    return str(((job or {}).get('targets') or ['UNKNOWN_TARGET'])[0] or 'UNKNOWN_TARGET').strip() or 'UNKNOWN_TARGET'


def _storage_root() -> tuple:
    """
    Returns the preferred write root as (live_status_root, index_file).

    Important: do not switch all reads to local just because a write probe on
    the Sphere share fails. Published Live Status jobs are shared from Sphere;
    a read-only/process-lock issue should not make the landing page look empty.
    """
    cached = getattr(_storage_root, '_cached', None)
    if cached:
        return cached
    try:
        os.makedirs(_LIVE_STATUS_ROOT, exist_ok=True)
        if os.path.isdir(_LIVE_STATUS_ROOT):
            _storage_root._cached = (_LIVE_STATUS_ROOT, _INDEX_FILE)
            return _storage_root._cached
    except Exception as exc:
        logger.warning('[LIVE STATUS] Sphere storage unavailable for write root: %s', exc)
    os.makedirs(_LOCAL_ROOT, exist_ok=True)
    _storage_root._cached = (_LOCAL_ROOT, _LOCAL_INDEX_FILE)
    return _storage_root._cached


def _storage_roots_for_read() -> List[str]:
    """Return all readable Live Status roots, Sphere first, then local fallback."""
    roots: List[str] = []
    for root in (_LIVE_STATUS_ROOT, _LOCAL_ROOT):
        try:
            if root and os.path.isdir(root) and root not in roots:
                roots.append(root)
        except Exception:
            continue
    active_root, _ = _storage_root()
    if active_root and active_root not in roots:
        roots.append(active_root)
    return roots


def _index_file_for_root(root: str) -> str:
    if os.path.abspath(root) == os.path.abspath(_LIVE_STATUS_ROOT):
        return _INDEX_FILE
    if os.path.abspath(root) == os.path.abspath(_LOCAL_ROOT):
        return _LOCAL_INDEX_FILE
    return os.path.join(root, 'jobs_index.json')


def _target_live_status_dir_for_root(root: str, target_name: str) -> str:
    bu, target = _target_dir_parts(target_name)
    return os.path.join(root, bu, target)

# SWPDT JSON paths
_SWPDT_JSON_NETWORK = os.path.join(_PDTBUDDY_DATA_ROOT, 'SWPDT', 'SWPDT_job_summary.json')
_SWPDT_JSON_LOCAL   = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'SWPDT_job_summary.json')

def _get_swpdt_json_path() -> str:
    """
    Always return the FRESHEST available SWPDT JSON path.
    Prefers network path if it exists AND is newer than local.
    Falls back to local backup if network is unavailable.
    """
    net_exists   = os.path.exists(_SWPDT_JSON_NETWORK)
    local_exists = os.path.exists(_SWPDT_JSON_LOCAL)
    if net_exists and local_exists:
        net_mtime   = os.path.getmtime(_SWPDT_JSON_NETWORK)
        local_mtime = os.path.getmtime(_SWPDT_JSON_LOCAL)
        return _SWPDT_JSON_NETWORK if net_mtime >= local_mtime else _SWPDT_JSON_LOCAL
    if net_exists:
        return _SWPDT_JSON_NETWORK
    if local_exists:
        return _SWPDT_JSON_LOCAL
    return _SWPDT_JSON_NETWORK  # default even if missing

# Keep _SWPDT_JSON as a compat alias pointing to network (routes import it)
_SWPDT_JSON = _SWPDT_JSON_NETWORK


def _extract_meta_from_build_path(build_path: str) -> str:
    """
    Extract META-ID from a build path like:
    //server/path/Aldabra.LA.1.0-00270-STD.INT-1
    -> META-00270
    """
    if not build_path:
        return ''
    m = re.search(r'-0*(\d{3,6})[-_]', str(build_path))
    if m:
        return f"META-{m.group(1).zfill(5)}"
    return ''


def _extract_build_name_from_path(build_path: str) -> str:
    """
    Extract clean build name from UNC path.
    e.g. \\\\server\\path\\Aldabra.LA.1.0-00270-STD.INT-1 -> Aldabra.LA.1.0-00270-STD.INT-1
    """
    if not build_path:
        return ''
    parts = [p for p in str(build_path).replace('\\', '/').split('/') if p]
    return parts[-1] if parts else build_path
    if not build_path:
        return ''
    # match pattern like -00270- or -00270_ in the build path
    m = re.search(r'-0*(\d{3,6})[-_]', str(build_path))
    if m:
        return f"META-{m.group(1).zfill(5)}"
    return ''


def load_swpdt_running_builds(software_product_prefix: str) -> List[Dict[str, Any]]:
    """
    Read SWPDT_job_summary.json (new build-keyed format from fetch_axiom_combined).
    Returns list of { build_id, software_product, device_count, chip_ids, submitted,
                      meta_id, build_name, run_status, first_submitted }
    Falls back gracefully to old job-list format if new format not yet present.
    """
    def _make_row(build_id, sp, device_count, chip_ids, submitted):
        return {
            'meta_id':          _extract_meta_from_build_path(build_id),
            'build_name':       build_id,
            'build_full':       build_id,
            'software_product': sp,
            'device_count':     device_count,
            'chip_ids':         chip_ids or [],
            'run_status':       'completed',
            'job_count':        1,
            'first_submitted':  str(submitted or '')[:10],
        }

    prefix_lower = software_product_prefix.strip().lower()
    prefix_code  = prefix_lower.split('.')[0]

    def _sp_matches(sp):
        sp_l = str(sp or '').lower()
        return sp_l.startswith(prefix_lower) or sp_l.split('.')[0] == prefix_code

    # -- Try new build-keyed format first --
    try:
        swpdt_path = _get_swpdt_json_path()
        if os.path.exists(swpdt_path):
            with open(swpdt_path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            if 'builds' in data:
                result = []
                for b in (data.get('builds') or {}).values():
                    if _sp_matches(b.get('software_product') or b.get('build_id')):
                        result.append(_make_row(
                            b.get('build_id', ''),
                            b.get('software_product', ''),
                            b.get('device_count', 0),
                            b.get('chip_ids', []),
                            b.get('submitted', ''),
                        ))
                return result
    except Exception:
        pass

    # -- Fallback: old job-list format --
    try:
        swpdt_path = _get_swpdt_json_path()
        if not os.path.exists(swpdt_path):
            return []
        with open(swpdt_path, 'r', encoding='utf-8') as fh:
            data = json.load(fh)
        jobs = data.get('jobs') or []
        build_map: Dict[str, Dict[str, Any]] = {}
        for job in jobs:
            if not _sp_matches(job.get('software_product')):
                continue
            state = str(job.get('state') or '').lower()
            if state not in ('running', 'completed', 'submitted', 'dispatched'):
                continue
            build_path = str(job.get('build') or '')
            build_name = _extract_build_name_from_path(build_path)
            meta_id    = _extract_meta_from_build_path(build_path)
            if not build_name or not meta_id:
                continue
            submitted = str(job.get('submitted') or '')[:10]
            key = build_name.upper()
            if key not in build_map:
                build_map[key] = {
                    'meta_id': meta_id, 'build_name': build_name,
                    'build_full': build_name, 'run_status': 'completed',
                    'job_count': 0, 'device_count': 0,
                    'chip_ids': [], 'first_submitted': submitted,
                }
            build_map[key]['job_count']    += 1
            build_map[key]['device_count'] += int(job.get('device_count') or 0)
            if submitted and (not build_map[key]['first_submitted'] or
                              submitted < build_map[key]['first_submitted']):
                build_map[key]['first_submitted'] = submitted
        return list(build_map.values())
    except Exception:
        return []

    os.makedirs(_LIVE_STATUS_ROOT, exist_ok=True)
    if not os.path.exists(_JOBS_FILE):
        with open(_JOBS_FILE, 'w', encoding='utf-8') as fh:
            json.dump({'jobs': []}, fh, indent=2)


def _utc_now() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + 'Z'


# - per-job write locks -
# One threading.Lock per job_id. Prevents two Flask threads from writing
# the same <job_id>.json simultaneously (auto-refresh + manual save, etc.).
_JOB_LOCKS: Dict[str, threading.Lock] = {}
_JOB_LOCKS_GUARD = threading.Lock()

def _job_lock(job_id: str) -> threading.Lock:
    with _JOB_LOCKS_GUARD:
        if job_id not in _JOB_LOCKS:
            _JOB_LOCKS[job_id] = threading.Lock()
        return _JOB_LOCKS[job_id]

# - per-job file helpers -
    jobs_dir, _ = _storage_root()
    return os.path.join(jobs_dir, f'{job_id}.json')


def _candidate_job_files(job_id: str) -> List[str]:
    paths: List[str] = []
    roots = _storage_roots_for_read()
    # New target-scoped path from any available index, across Sphere + local.
    for e in _read_index():
        if str(e.get('id')) != str(job_id):
            continue
        target = str(((e.get('targets') or []) + ['UNKNOWN_TARGET'])[0] or 'UNKNOWN_TARGET')
        for root in roots:
            paths.append(os.path.join(_target_live_status_dir_for_root(root, target), 'jobs', f'{job_id}.json'))
    # Legacy flat paths in both locations.
    for root in roots:
        paths.append(os.path.join(root, 'jobs', f'{job_id}.json'))
    paths.append(os.path.join(_LIVE_STATUS_ROOT, 'jobs', f'{job_id}.json'))
    paths.append(os.path.join(_LOCAL_ROOT, 'jobs', f'{job_id}.json'))
    # Last-resort recursive search under all readable roots when indexed/legacy
    # paths do not exist.
    if not any(os.path.exists(p) for p in paths):
        for root in roots:
            try:
                for base, _dirs, files in os.walk(root):
                    if f'{job_id}.json' in files:
                        paths.append(os.path.join(base, f'{job_id}.json'))
                        break
            except Exception:
                pass
    out = []
    for p in paths:
        if p and p not in out:
            out.append(p)
    return out


def _job_file_for_job(job: Dict[str, Any]) -> str:
    return os.path.join(target_live_status_dir(_job_primary_target(job)), 'jobs', f'{job["id"]}.json')


def get_job_for_target(target_name: str) -> Optional[Dict[str, Any]]:
    """Directly load the latest job for a target from Sphere or local fallback."""
    best = None
    for root in _storage_roots_for_read():
        jobs_dir = os.path.join(_target_live_status_dir_for_root(root, target_name), 'jobs')
        if not os.path.isdir(jobs_dir):
            continue
        for fname in os.listdir(jobs_dir):
            if not fname.endswith('.json'):
                continue
            fpath = os.path.join(jobs_dir, fname)
            try:
                with open(fpath, 'r', encoding='utf-8') as fh:
                    job = json.load(fh)
                if not isinstance(job, dict) or not job.get('id'):
                    continue
                if best is None or str(job.get('updated_at') or '') > str(best.get('updated_at') or ''):
                    best = job
            except Exception:
                continue
    return best


def _read_job_file(job_id: str) -> Optional[Dict[str, Any]]:
    for path in _candidate_job_files(job_id):
        try:
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as fh:
                    return json.load(fh)
        except Exception:
            continue
    return None


def _write_job_file(job: Dict[str, Any], refresh_index: bool = True) -> None:
    """
    Write one job under a per-job threading.Lock.

    Why this fixes WinError 32 / WinError 5:
      On Windows network shares os.replace() can fail with:
        - WinError 32: file in use by another process (antivirus, indexer)
        - WinError 5:  access denied during rename (share-level lock)
      Strategy:
        1. Write to .tmp under the per-job lock
        2. Try os.replace() up to 3 times with short back-off
        3. If all retries fail, fall back to direct overwrite (non-atomic
           but safe because the lock serialises concurrent Flask threads)
    """
    import time
    path = _job_file_for_job(job)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + '.tmp'
    with _job_lock(str(job['id'])):
        with open(tmp, 'w', encoding='utf-8') as fh:
            json.dump(job, fh, indent=2)
        # Try atomic rename with retries
        last_err = None
        for attempt in range(3):
            try:
                os.replace(tmp, path)
                last_err = None
                break
            except (PermissionError, OSError) as exc:
                last_err = exc
                time.sleep(0.05 * (attempt + 1))  # 50ms, 100ms, 150ms
        if last_err is not None:
            # Atomic rename failed - fall back to direct write
            # (safe: we hold the per-job lock so no other thread is writing)
            logger.warning(
                '[LIVE STATUS] os.replace failed for job %s (%s); '
                'falling back to direct write', job.get('id'), last_err
            )
            try:
                with open(path, 'w', encoding='utf-8') as fh:
                    json.dump(job, fh, indent=2)
            finally:
                try:
                    if os.path.exists(tmp):
                        os.remove(tmp)
                except OSError:
                    pass
    if refresh_index:
        try:
            _update_index(job)
        except Exception as exc:
            logger.warning('[LIVE STATUS] Job %s saved, but index update failed: %s', job.get('id'), exc)


def _job_file(job_id: str) -> str:
    """Legacy helper: return first existing job file, else flat legacy path."""
    for path in _candidate_job_files(job_id):
        if os.path.exists(path):
            return path
    root, _ = _storage_root()
    return os.path.join(root, 'jobs', f'{job_id}.json')


def _delete_job_file(job_id: str) -> None:
    for path in _candidate_job_files(job_id):
        try:
            if os.path.exists(path):
                os.remove(path)
        except OSError:
            pass
    _remove_from_index(job_id)


# - lightweight index (id/name/status/updated_at only) -
_INDEX_FIELDS = ('id', 'name', 'status', 'job_type', 'targets', 'targets_display',
                 'public_token', 'updated_at', 'created_at', 'published_at', 'published_by',
                 'revoked_at', 'revoked_by')


def _read_index() -> List[Dict[str, Any]]:
    """Read and merge jobs_index.json from Sphere and local fallback."""
    by_id: Dict[str, Dict[str, Any]] = {}
    for root in _storage_roots_for_read():
        index_file = _index_file_for_root(root)
        try:
            with open(index_file, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            entries = data if isinstance(data, list) else []
        except Exception:
            entries = []
        for entry in entries:
            if not isinstance(entry, dict) or not entry.get('id'):
                continue
            jid = str(entry.get('id'))
            old = by_id.get(jid)
            if old is None or str(entry.get('updated_at') or '') >= str(old.get('updated_at') or ''):
                by_id[jid] = entry
    return sorted(by_id.values(), key=lambda row: str(row.get('updated_at') or ''), reverse=True)


def _write_index(entries: List[Dict[str, Any]]) -> None:
    import time
    _, index_file = _storage_root()
    os.makedirs(os.path.dirname(index_file), exist_ok=True)
    tmp = index_file + '.tmp'
    with _job_lock('__index__'):
        with open(tmp, 'w', encoding='utf-8') as fh:
            json.dump(entries, fh, indent=2)
        last_err = None
        for attempt in range(3):
            try:
                os.replace(tmp, index_file)
                last_err = None
                break
            except (PermissionError, OSError) as exc:
                last_err = exc
                time.sleep(0.05 * (attempt + 1))
        if last_err is not None:
            logger.warning(
                '[LIVE STATUS] os.replace failed for index (%s); '
                'falling back to direct write', last_err
            )
            try:
                with open(index_file, 'w', encoding='utf-8') as fh:
                    json.dump(entries, fh, indent=2)
            finally:
                try:
                    if os.path.exists(tmp):
                        os.remove(tmp)
                except OSError:
                    pass


def _update_index(job: Dict[str, Any]) -> None:
    entries = _read_index()
    stub = {k: job.get(k) for k in _INDEX_FIELDS}
    for i, e in enumerate(entries):
        if str(e.get('id')) == str(job['id']):
            entries[i] = stub
            _write_index(entries)
            return
    entries.append(stub)
    _write_index(entries)


def _remove_from_index(job_id: str) -> None:
    _write_index([e for e in _read_index() if str(e.get('id')) != str(job_id)])


# - one-time migration from old monolithic jobs.json -

def _migrate_legacy() -> None:
    root, _ = _storage_root()
    legacy = os.path.join(root, 'jobs.json')
    done   = legacy + '.migrated'
    if not os.path.exists(legacy) or os.path.exists(done):
        return
    try:
        with open(legacy, 'r', encoding='utf-8') as fh:
            data = json.load(fh)
        for job in (data.get('jobs') or []):
            if job.get('id') and not _read_job_file(str(job['id'])):
                _write_job_file(job)
        with open(done, 'w') as fh:
            fh.write(_utc_now())
    except Exception:
        pass


# - legacy shims so any remaining internal callers still work -

def _read_store() -> Dict[str, Any]:
    return {'jobs': list_jobs()}


def _write_store(data: Dict[str, Any]) -> None:
    for job in (data.get('jobs') or []):
        if job.get('id'):
            _write_job_file(job)


def update_job_fields(job_id: str, updates: Dict[str, Any], refresh_index: bool = False) -> Optional[Dict[str, Any]]:
    """Read + write only the single job file.

    Frequent background writes such as rows/cache/JQL/viewer heartbeat should
    not touch jobs_index.json. The shared index is refreshed only for lifecycle
    and metadata changes that affect job listing/navigation.
    """
    job = _read_job_file(job_id)
    if job is None:
        return None
    job.update(updates or {})
    job['updated_at'] = _utc_now()
    _write_job_file(job, refresh_index=refresh_index)
    return job


def set_current_report_cache(job_id: str, cache: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    return update_job_fields(job_id, {'current_report_cache': cache or {}})


def update_current_report_fresh_check(job_id: str, checked_at: str) -> Optional[Dict[str, Any]]:
    job = get_job(job_id)
    if not job:
        return None
    cache = dict(job.get('current_report_cache') or {})
    cache['last_fresh_check_at'] = checked_at
    return set_current_report_cache(job_id, cache)


def set_current_report_exclusions(job_id: str, excluded_keys: List[str]) -> Optional[Dict[str, Any]]:
    cleaned = sorted({str(k or '').strip().upper() for k in (excluded_keys or []) if str(k or '').strip()})
    return update_job_fields(job_id, {'current_report_excluded_jiras': cleaned})


def set_current_report_jql(job_id: str, jql: str) -> Optional[Dict[str, Any]]:
    return update_job_fields(job_id, {'current_report_jql': (jql or '').strip()})


def set_weekly_report_selection(job_id: str, selection: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Persist the editor-selected weekly build subset for published viewers."""
    return update_job_fields(job_id, {'weekly_report_selection': selection or {}})

def delete_job(job_id: str) -> bool:
    if not _read_job_file(job_id):
        return False
    _delete_job_file(job_id)
    return True


def list_jobs() -> List[Dict[str, Any]]:
    """
    Return all jobs sorted by updated_at desc.
    Reads both the shared Sphere location and the local fallback so old
    published jobs do not disappear when the process temporarily falls back.
    """
    _migrate_legacy()   # no-op after first run
    index = _read_index()
    if index:
        ordered_ids = [e['id'] for e in index if e.get('id')]
    else:
        ordered_ids = []
        for root in _storage_roots_for_read():
            try:
                for base, _dirs, files in os.walk(root):
                    for f in files:
                        if f.endswith('.json') and not f.startswith('.') and f not in ('jobs_index.json', 'jobs.json'):
                            ordered_ids.append(f[:-5])
            except Exception:
                pass
    jobs_by_id: Dict[str, Dict[str, Any]] = {}
    for jid in ordered_ids:
        job = _read_job_file(jid)
        if job and isinstance(job, dict) and job.get('id'):
            old = jobs_by_id.get(str(job.get('id')))
            if old is None or str(job.get('updated_at') or '') >= str(old.get('updated_at') or ''):
                jobs_by_id[str(job.get('id'))] = job
    jobs = list(jobs_by_id.values())
    jobs.sort(key=lambda row: str(row.get('updated_at') or ''), reverse=True)
    return jobs


def get_job(job_id: str) -> Optional[Dict[str, Any]]:
    return _read_job_file(job_id)


def create_job(name: str, targets: List[str], username: str, job_type: str = 'CRM') -> Dict[str, Any]:
    now = _utc_now()
    clean_type = str(job_type or 'CRM').strip().upper()
    if clean_type not in ('CRM', 'ENG'):
        clean_type = 'CRM'
    job = {
        'id': str(uuid.uuid4()),
        'name': name.strip() or 'Untitled Live Status View',
        'targets': [t for t in targets if str(t).strip()],
        'job_type': clean_type,
        'status': 'draft',
        'created_by': username,
        'created_at': now,
        'updated_at': now,
        'published_at': None,
        'published_by': None,
        'revoked_at': None,
        'revoked_by': None,
        'revoke_reason': '',
        'public_token': str(uuid.uuid4()).replace('-', ''),
        'hours_mode': 'disabled' if clean_type == 'ENG' else 'enabled',
        'mtbf_mode': 'disabled' if clean_type == 'ENG' else 'enabled',
        'column_config': {
            'show_build_job_left_column': True,
            'visible_columns': ['build_name', 'target', 'status', 'updated_at'],
            'hidden_columns': [],
        },
        'published_comments_draft': '',
        'published_comments_snapshot': '',
        'internal_comments': '',
        'targets_display': ', '.join([t for t in targets if str(t).strip()]),
        'draft_rows': [],
        'published_rows': [],
        'current_report_cache': {},
        'current_report_excluded_jiras': [],
    }
    _write_job_file(job, refresh_index=True)   # creates <job_id>.json and list index
    return job


def save_job_meta(job_id: str, updates: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    job = _read_job_file(job_id)
    if job is None:
        return None
    job_type = str(updates.get('job_type', job.get('job_type') or 'CRM')).strip().upper()
    if job_type not in ('CRM', 'ENG'):
        job_type = 'CRM'
    job.update({
        'name':                     updates.get('name',                     job.get('name')),
        'targets':                  updates.get('targets',                  job.get('targets', [])),
        'job_type':                 job_type,
        'hours_mode':               'disabled' if job_type == 'ENG' else updates.get('hours_mode', job.get('hours_mode', 'enabled')),
        'mtbf_mode':                'disabled' if job_type == 'ENG' else updates.get('mtbf_mode', job.get('mtbf_mode', 'enabled')),
        'internal_comments':        updates.get('internal_comments',        job.get('internal_comments', '')),
        'published_comments_draft': updates.get('published_comments_draft', job.get('published_comments_draft', '')),
        'test_eng_comments':        updates.get('test_eng_comments',        job.get('test_eng_comments', '')),
        'updated_at': _utc_now(),
    })
    # Persist SP configs when provided
    if 'sp_configs' in updates:
        existing = job.get('sp_configs') or {}
        incoming = updates['sp_configs'] or {}
        existing.update(incoming)
        job['sp_configs'] = existing
    job['targets_display'] = ', '.join(job.get('targets') or [])
    _write_job_file(job, refresh_index=True)
    return job


def save_job_rows(job_id: str, rows: List[Dict[str, Any]], username: str = '') -> Optional[Dict[str, Any]]:
    """Save draft rows only.

    Viewers should see data only after the editor explicitly publishes.  Save
    updates draft_rows for editing, while publish_job copies draft_rows to the
    live published_rows snapshot.
    """
    job = _read_job_file(job_id)
    if job is None:
        return None
    clean_rows = [
        r for r in (rows or [])
        if str((r or {}).get('source') or '').strip().lower() not in ('excel', 'excel+json')
    ]
    job['draft_rows'] = clean_rows
    job['updated_at'] = _utc_now()
    _write_job_file(job, refresh_index=False)
    return job



# ----?-------?--- helpers ----?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?---

def _find_header_index(headers: List[str], candidates: List[str]) -> int:
    norm = [str(h or '').strip().lower() for h in (headers or [])]
    for cand in candidates:
        c = str(cand or '').strip().lower()
        for idx, header in enumerate(norm):
            if header == c or c in header:
                return idx
    return -1


def _cell(values: List[Any], idx: int) -> str:
    if idx < 0 or idx >= len(values or []):
        return ''
    value = values[idx]
    return '' if value is None else str(value).strip()


def _to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except Exception:
        return 0.0


def _row_key(row: Dict[str, Any]) -> str:
    """Stable merge key: build_full first (unique per variant), fallback to meta_id."""
    bf = str(row.get('build_full') or '').strip().upper()
    if bf:
        return bf
    return str(row.get('meta_id') or '').strip().upper()


# ----?-------?--- Excel reader ----?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?---

def _read_excel_rows(primary_target: str) -> Tuple[List[Dict[str, Any]], str, str, Optional[str]]:
    """
    Returns (all_rows, sheet_name, excel_path, error_or_None).
    Each row has: excel_row, target, product_line, meta_id, build_full,
                  hours, crashes, mtbf, week, run_status, comments,
                  devices, device_count, source='excel'
    """
    try:
        cfg = (_get_target_excel_config(primary_target) or {}).get('mtbf', {})
        excel_path = _normalize_excel_path(cfg.get('excel_path', ''))
        sheet_name = str(cfg.get('sheet_name') or '').strip()
        if not excel_path:
            return [], '', '', 'MTBF Excel not configured for this target.'
        if not os.path.exists(excel_path):
            return [], '', excel_path, f'MTBF Excel file not found: {excel_path}'

        import openpyxl
        from datetime import datetime as _dtt, date as _ddate

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        actual_sheet = sheet_name
        if actual_sheet and actual_sheet not in wb.sheetnames:
            actual_sheet = next(
                (s for s in wb.sheetnames if s.strip().lower() == actual_sheet.lower()), ''
            )
        if not actual_sheet and wb.sheetnames:
            actual_sheet = wb.sheetnames[0]
        if not actual_sheet:
            return [], '', excel_path, 'No sheets found in workbook.'

        ws = wb[actual_sheet]
        merge_map: Dict[Tuple[int, int], Any] = {}
        for mr in list(ws.merged_cells.ranges):
            val = ws.cell(mr.min_row, mr.min_col).value
            for r in range(mr.min_row, mr.max_row + 1):
                for c in range(mr.min_col, mr.max_col + 1):
                    merge_map[(r, c)] = val

        def _cv(r: int, c: int) -> str:
            v = merge_map.get((r, c), ws.cell(r, c).value)
            if isinstance(v, _dtt):
                return v.strftime('%Y-%m-%d')
            if isinstance(v, _ddate):
                return v.strftime('%Y-%m-%d')
            return '' if v is None else str(v).strip()

        headers = [_cv(1, c) for c in range(1, ws.max_column + 1)]
        hm = {
            'target':       _find_header_index(headers, ['Target(s)', 'Target']),
            'product':      _find_header_index(headers, ['Product Line(s)', 'Product Line', 'Product']),
            'devices':      _find_header_index(headers, ['Device(s)', 'Devices']),
            'device_count': _find_header_index(headers, ['Device Count']),
            'week':         _find_header_index(headers, ['Week', 'Date']),
            'meta':         _find_header_index(headers, ['Meta-ID', 'META-ID', 'Meta ID', 'META']),
            'build_full':   _find_header_index(headers, ['Build(s) Full ID', 'Build(s)', 'Full Build', 'Build']),
            'hours':        _find_header_index(headers, ['Tested Hours', 'Hours', 'Total Hours']),
            'crashes':      _find_header_index(headers, ['Total Crashes', 'Crashes', 'Crash Count']),
            'mtbf':         _find_header_index(headers, ['MTBF', 'MTBF (hrs)', 'MTBF Hrs']),
            'run_status':   _find_header_index(headers, ['Build Status', 'Run Status', 'Status', 'Running Status']),
            'comments':     _find_header_index(headers, ['Notes', 'MTBF Details', 'Comments']),
        }
        has_status_col = hm['run_status'] >= 0

        all_rows: List[Dict[str, Any]] = []
        for r in range(2, ws.max_row + 1):
            values = [_cv(r, c) for c in range(1, ws.max_column + 1)]
            if not any(v.strip() for v in values):
                continue
            raw_status = _cell(values, hm['run_status']).lower().strip()
            # if no status column in Excel, default all to 'stopped'
            # (user can override via JSON draft_rows layer)
            if not has_status_col:
                run_status = 'stopped'
            else:
                run_status = 'running' if raw_status == 'running' else 'stopped'
            meta_id    = _cell(values, hm['meta'])
            build_full = _cell(values, hm['build_full'])
            all_rows.append({
                'excel_row':    r,
                'source':       'excel',
                'target':       _cell(values, hm['target']) or primary_target,
                'product_line': _cell(values, hm['product']),
                'meta_id':      meta_id,
                'build_full':   build_full,
                'display_build': meta_id or build_full,
                'hours':        _cell(values, hm['hours']),
                'crashes':      _cell(values, hm['crashes']),
                'mtbf':         _cell(values, hm['mtbf']),
                'week':         _cell(values, hm['week']),
                'run_status':   run_status,
                'comments':     _cell(values, hm['comments']),
                'devices':      _cell(values, hm['devices']),
                'device_count': _cell(values, hm['device_count']),
            })
        return all_rows, actual_sheet, excel_path, None
    except Exception as exc:
        return [], '', '', str(exc)


# ----?-------?--- JSON row normalizer ----?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?---

def _merge_rows(
    excel_rows: List[Dict[str, Any]],
    json_rows: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    Keep Excel completely separate from the Live Status JSON layer.

    The Live Status running/stopped workspace is backed only by draft_rows and
    SWPDT-added rows. Excel rows are read elsewhere only for the dedicated
    inline Excel editor / Excel save APIs.
    """
    del excel_rows  # explicitly unused: do not merge Excel into JSON state
    normalized: List[Dict[str, Any]] = []
    for jr in (json_rows or []):
        if str((jr or {}).get('source') or '').strip().lower() in ('excel', 'excel+json'):
            continue
        row = dict(jr)
        row.setdefault('source', 'json')
        row.setdefault('target', '')
        row.setdefault('excel_row', None)
        row['display_build'] = row.get('meta_id') or row.get('build_full') or ''
        normalized.append(row)
    return normalized


# ----?-------?--- collapsed view selection ----?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?---

def _build_collapsed(
    merged_rows: List[Dict[str, Any]],
) -> Tuple[List[Dict[str, Any]], int]:
    running = [r for r in merged_rows if str(r.get('run_status') or '').lower() == 'running']

    # latest 3 metas (or 5 if no running rows)
    max_metas = 3 if running else 5
    latest_keys: List[str] = []
    seen: set = set()
    for row in reversed(merged_rows):
        k = _row_key(row)
        if not k or k in seen:
            continue
        seen.add(k)
        latest_keys.append(k)
        if len(latest_keys) >= max_metas:
            break
    latest_set = set(latest_keys)

    used: set = set()
    collapsed: List[Dict[str, Any]] = []
    for row in running + [r for r in merged_rows if _row_key(r) in latest_set]:
        k = _row_key(row)
        if k in used:
            continue
        used.add(k)
        collapsed.append(row)

    return collapsed, len(latest_set)


# ----?-------?--- public API ----?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?-------?---

def load_job_workspace_data(job: Dict[str, Any]) -> Dict[str, Any]:
    targets = [str(t).strip() for t in (job.get('targets') or []) if str(t).strip()]
    primary_target = targets[0] if targets else ''
    workspace: Dict[str, Any] = {
        'job_id':         job.get('id'),
        'primary_target': primary_target,
        'targets':        targets,
        'stats': {
            'target_count':      len(targets),
            'meta_count':        0,
            'build_count':       0,
            'running_count':     0,
            'stopped_count':     0,
            'json_only_count':   0,
            'excel_count':       0,
            'merged_count':      0,
            'total_hours':       0.0,
            'total_crashes':     0,
            'latest_meta_count': 0,
        },
        'rows':       [],   # collapsed view
        'all_rows':   [],   # full merged sheet
        'sheet_name': '',
        'excel_path': '',
        'error':      None,
    }
    if not primary_target:
        return workspace

    # 1. read Excel
    excel_rows, sheet_name, excel_path, excel_err = _read_excel_rows(primary_target)
    workspace['sheet_name'] = sheet_name
    workspace['excel_path'] = excel_path
    if excel_err and not excel_rows:
        workspace['error'] = excel_err
        # still continue -------? JSON rows may exist

    # 2. read JSON draft_rows from job
    json_rows: List[Dict[str, Any]] = list(job.get('draft_rows') or [])

    # 3. load SWPDT running data, grouped by build name.
    # Completed SWPDT rows are not inferred from Excel; Excel must not seed JSON.
    swpdt_rows = load_swpdt_running_builds(primary_target.capitalize())
    # key for SWPDT dedup = build_name (unique per variant)
    json_build_keys  = {str(r.get('build_full') or r.get('meta_id') or '').strip().upper() for r in json_rows}

    for sr in swpdt_rows:
        build_name   = str(sr.get('build_name') or '').strip()
        meta_id      = str(sr.get('meta_id') or '').strip()
        run_status   = sr.get('run_status', 'completed')
        bkey         = build_name.upper()
        if not build_name or not meta_id:
            continue

        if run_status in ('running', 'submitted', 'dispatched'):
            run_status = 'running'
            if bkey not in json_build_keys:
                json_rows.append({
                    'meta_id':        meta_id,
                    'build_full':     build_name,
                    'run_status':     'running',
                    'hours':          '',
                    'crashes':        '',
                    'mtbf':           '',
                    'week':           sr.get('first_submitted',''),
                    'comments':       '',
                    'product_line':   '',
                    'target':         primary_target,
                    'source':         'swpdt',
                    'job_count':      sr['job_count'],
                    'device_count':   sr['device_count'],
                    'first_submitted':sr.get('first_submitted',''),
                })
                json_build_keys.add(bkey)
            else:
                for jr in json_rows:
                    jbkey = str(jr.get('build_full') or jr.get('meta_id') or '').strip().upper()
                    if jbkey == bkey:
                        jr['run_status']   = 'running'
                        jr['job_count']    = sr['job_count']
                        jr['device_count'] = sr['device_count']
                        break

    # 4. Normalize JSON(+SWPDT) rows only. Do not merge Excel rows here.
    merged = _merge_rows(excel_rows, json_rows)

    # 4. collapsed view
    collapsed, latest_meta_count = _build_collapsed(merged)

    # 5. stats
    running_rows  = [r for r in merged if str(r.get('run_status') or '').lower() == 'running']
    json_only     = [r for r in merged if r.get('source') == 'json']
    excel_only    = [r for r in merged if r.get('source') == 'excel']
    merged_both   = [r for r in merged if r.get('source') == 'excel+json']
    unique_metas  = {_row_key(r) for r in merged if _row_key(r)}

    workspace['rows']     = collapsed
    workspace['all_rows'] = merged
    workspace['stats'].update({
        'meta_count':        len(unique_metas),
        'build_count':       len(merged),
        'running_count':     len(running_rows),
        'stopped_count':     len(merged) - len(running_rows),
        'json_only_count':   len(json_only),
        'excel_count':       len(excel_only),
        'merged_count':      len(merged_both),
        'total_hours':       round(sum(_to_float(r.get('hours')) for r in collapsed), 2),
        'total_crashes':     int(sum(_to_float(r.get('crashes')) for r in collapsed)),
        'latest_meta_count': latest_meta_count,
    })
    return workspace


def publish_job(job_id: str, username: str) -> Optional[Dict[str, Any]]:
    job = _read_job_file(job_id)
    if job is None:
        return None
    now = _utc_now()
    job['status']                      = 'published'
    job['published_at']                = now
    job['published_by']                = username or job.get('created_by') or 'unknown'
    job['updated_at']                  = now
    job['published_comments_snapshot'] = job.get('published_comments_draft', '')
    job['published_rows']              = list(job.get('draft_rows') or [])
    _write_job_file(job, refresh_index=True)
    return job



def revoke_job(job_id: str, username: str, reason: str = '') -> Optional[Dict[str, Any]]:
    job = _read_job_file(job_id)
    if job is None:
        return None
    now = _utc_now()
    job['status'] = 'revoked'
    job['revoked_at'] = now
    job['revoked_by'] = username
    job['revoke_reason'] = str(reason or '').strip()
    job['updated_at'] = now
    _write_job_file(job, refresh_index=True)
    return job


def update_viewer_heartbeat(job_id: str, username: str) -> Optional[Dict[str, Any]]:
    # No write - viewer tracking removed. Was causing WinError 5 on network
    # share by writing job JSON on every page load from every viewer.
    return _read_job_file(job_id)

