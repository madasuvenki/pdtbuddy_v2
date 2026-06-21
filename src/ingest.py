import logging
logger = logging.getLogger(__name__)
import openpyxl
import pandas as pd
import datetime
import re
import traceback
import os
import glob

from mysql.connector import Error

from config import SHEET_CONFIG, TRUNCATE_EXISTING_TABLES
from src.utils import sanitize_column_name, get_mysql_connection_db


EXPECTED_EXCEL_SUFFIX = "_Overall_PDT_Stats"
# Unique CR workbook name patterns can vary by BU/source.
# Keep the old pattern, but also accept newer naming like:
#   OVERALL_NORD_HQX_2026_05_04__Unique_CRs-2026y_05m_04d_15h02m35s.xlsx
UNIQUE_CR_FILENAME_PATTERNS = [
    "*EXCLUSIVE__Unique_CRs*.xlsx",   # double underscore standard
    "*EXCLUSIVE_Unique_CRs*.xlsx",    # single underscore variant
    "*EXCLUSIVE*.xlsx",               # any EXCLUSIVE file
    "*Unique_CRs-*.xlsx",
    "*Unique_CRs*.xlsx",
    # CSV variants (e.g. Maili: OVERALL_MAILI_PDTUNIQUE_06_04__Unique_CRs-*.csv)
    "*EXCLUSIVE__Unique_CRs*.csv",
    "*EXCLUSIVE_Unique_CRs*.csv",
    "*Unique_CRs-*.csv",
    "*Unique_CRs*.csv",
]
OVERALL_CRS_SHEET_NAME = "OverallCrs"
PDT_UNIQUE_CRS_SHEET_NAME = "PDT_UniqueCrs"
PDT_REPORTED_CRS_SHEET_NAME = "PDT_ReportedCrs"
REPORTED_TEAM_COLUMN = "reported_team"
CRID_COLUMN = "crid"


def build_create_table_sql(table_name, sanitized_headers, primary_key_column=None):
    """
    Generates a CREATE TABLE SQL statement with basic type guessing.
    Includes primary key constraint if specified.
    """
    columns_defs = []

    for header in sanitized_headers:
        col_type = "VARCHAR(255)"

        if any(keyword in header.lower() for keyword in [
            "labels", "description", "comment", "summary", "test_tas", "details",
            "scenario", "notes", "error_message", "stack_trace", "logs", "title"
        ]):
            col_type = "TEXT"
        elif any(keyword in header.lower() for keyword in [
            "date", "time", "added", "fetched", "built", "created_at", "updated_at"
        ]):
            col_type = "DATETIME"
        elif any(keyword in header.lower() for keyword in ["id", "cr", "ticket", "jira", "key", "pl_id"]):
            col_type = "VARCHAR(255)"
        elif any(keyword in header.lower() for keyword in ["count", "num", "occurrence", "version_int"]):
            col_type = "INT"
        elif "image" in header.lower():
            col_type = "TEXT"

        col_def = f"`{header}` {col_type} NULL"

        if primary_key_column and header.lower() == sanitize_column_name(primary_key_column).lower():
            col_def = f"`{header}` {col_type} NOT NULL"

        columns_defs.append(col_def)

    pk_constraint = ""
    if primary_key_column and sanitize_column_name(primary_key_column).lower() in [h.lower() for h in sanitized_headers]:
        pk_sanitized = sanitize_column_name(primary_key_column)
        pk_constraint = f", PRIMARY KEY (`{pk_sanitized}`)"

    return f"CREATE TABLE IF NOT EXISTS `{table_name}` ({', '.join(columns_defs)}{pk_constraint})"


def _exclude_temp_excel_files(files):
    return [f for f in files if not os.path.basename(f).startswith('~$')]


def _resolve_actual_excel_file(excel_file_path: str) -> str:
    """
    If excel_file_path is a directory, pick a matching *_Overall_PDT_Stats.(xlsx|xls).
    If multiple matches, picks the newest by mtime.
    Falls back to any .xlsx if no suffix match found.
    Excludes Excel temp/lock files starting with ~$ (open-file locks).
    """
    try:
        if os.path.isdir(excel_file_path):
            matching_files = _exclude_temp_excel_files(
                glob.glob(os.path.join(excel_file_path, f"*{EXPECTED_EXCEL_SUFFIX}.xlsx")) +
                glob.glob(os.path.join(excel_file_path, f"*{EXPECTED_EXCEL_SUFFIX}.xls"))
            )
            if not matching_files:
                matching_files = _exclude_temp_excel_files(
                    glob.glob(os.path.join(excel_file_path, "*.xlsx")) +
                    glob.glob(os.path.join(excel_file_path, "*.xls"))
                )
            if not matching_files:
                raise FileNotFoundError(f"No Excel file found in '{excel_file_path}'.")
            try:
                return max(matching_files, key=os.path.getmtime)
            except OSError:
                return matching_files[0]

        if os.path.isfile(excel_file_path) and (excel_file_path.endswith(".xlsx") or excel_file_path.endswith(".xls")):
            return excel_file_path

        raise FileNotFoundError(f"Provided path '{excel_file_path}' is neither a directory nor a valid Excel file.")
    except OSError as e:
        raise FileNotFoundError(f"Network error accessing '{excel_file_path}': {e}") from e


def _resolve_latest_unique_cr_workbook(unique_cr_path: str) -> str:
    """
    Resolve the correct Unique CR workbook.

    Rules:
    1. If a direct .xlsx file path is provided, use it.
    2. If a directory is provided:
       a. Look for date-named subfolders (e.g. 2026, 2026_05_04).
          If found, pick the latest subfolder and search for matching xlsx inside.
       b. If NO subfolders exist, search the directory itself for any .xlsx file.
    3. Filename patterns tried in order: EXCLUSIVE__Unique_CRs, Unique_CRs-, Unique_CRs, then any .xlsx.
    4. Ignore Excel temp files starting with '~$'.
    """
    if not unique_cr_path:
        raise FileNotFoundError("Unique CR path is empty.")

    if os.path.isfile(unique_cr_path):
        name = os.path.basename(unique_cr_path)
        if name.startswith('~$'):
            raise FileNotFoundError(f"Unique CR workbook is a temp file: '{unique_cr_path}'")
        if unique_cr_path.lower().endswith((".xlsx", ".csv")):
            return unique_cr_path
        raise FileNotFoundError(f"Unique CR workbook is not a valid .xlsx/.csv file: '{unique_cr_path}'")

    if not os.path.isdir(unique_cr_path):
        raise FileNotFoundError(f"Unique CR path does not exist: '{unique_cr_path}'")


    try:
        dir_contents = os.listdir(unique_cr_path)
    except OSError as e:
        raise FileNotFoundError(f"Network error listing '{unique_cr_path}': {e}") from e

        # ── Only consider date-pattern folders: YYYY_MM_DD (e.g. 2026_05_02) ────────
    import re as _re
    _DATE_FOLDER_RE = _re.compile(r'^\d{4}_\d{2}_\d{2}$')
    date_folders = [
        os.path.join(unique_cr_path, name)
        for name in dir_contents
        if os.path.isdir(os.path.join(unique_cr_path, name))
        and _DATE_FOLDER_RE.match(name)
    ]

    if not date_folders:
        # No date subfolders — search the directory itself for xlsx/csv files directly
        all_patterns = list(UNIQUE_CR_FILENAME_PATTERNS) + ["*.xlsx", "*.csv"]
        direct_candidates = []
        for pat in all_patterns:
            try:
                found = _exclude_temp_excel_files(glob.glob(os.path.join(unique_cr_path, pat)))
            except OSError:
                continue
            direct_candidates.extend(found)
            if direct_candidates:
                break
        if direct_candidates:
            try:
                chosen = max(direct_candidates, key=os.path.getmtime)
            except OSError:
                chosen = direct_candidates[0]
            logger.info(f"_resolve_latest_unique_cr_workbook: no date folders, using direct file '{chosen}'")
            return chosen
        raise FileNotFoundError(
            f"No date folders (YYYY_MM_DD) and no .xlsx/.csv files found directly under '{unique_cr_path}'. "
            f"Contents: {dir_contents}"
        )
    # Pick the latest date folder that actually contains a matching xlsx or csv.
    # Walk from newest to oldest until a file is found.
    search_dirs = sorted(date_folders, key=lambda p: os.path.basename(p), reverse=True)

    # ── Search each date folder newest→oldest, xlsx preferred over csv ──
    all_patterns = list(UNIQUE_CR_FILENAME_PATTERNS) + ["*.xlsx", "*.csv"]
    for search_dir in search_dirs:
        workbook_candidates = []
        for pat in all_patterns:
            try:
                found = _exclude_temp_excel_files(glob.glob(os.path.join(search_dir, pat)))
            except OSError:
                continue
            workbook_candidates.extend(found)
            if workbook_candidates:
                break
        if workbook_candidates:
            # Prefer xlsx over csv when both exist in the same folder
            xlsx_files = [f for f in workbook_candidates if f.lower().endswith('.xlsx')]
            chosen_list = xlsx_files if xlsx_files else workbook_candidates
            try:
                chosen = max(chosen_list, key=os.path.getmtime)
            except OSError:
                chosen = chosen_list[0]
            logger.info(f"_resolve_latest_unique_cr_workbook: using '{chosen}'")
            return chosen

    raise FileNotFoundError(
        f"No Unique CR workbook (.xlsx/.csv) found under '{unique_cr_path}' "
        f"(searched {len(search_dirs)} date folders: "
        f"{', '.join(os.path.basename(d) for d in search_dirs)})"
    )

def _get_table_col_maxlens(cursor, table_name):
    """
    Returns dict: {col_name_lower: max_len or None}
    max_len is int for VARCHAR/CHAR, None for TEXT/BLOB/others.
    """
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
    out = {}
    for row in cursor.fetchall():
        field = row[0]
        col_type = (row[1] or "").lower()
        m = re.match(r"^(varchar|char)\((\d+)\)", col_type)
        if m:
            out[field.lower()] = int(m.group(2))
        else:
            out[field.lower()] = None
    return out


def _get_table_columns(cursor, table_name):
    """
    Returns actual table column names in order.
    """
    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
    return [row[0] for row in cursor.fetchall()]


def _clean_jira_traversal(s):
    """
    Keep only till '[' and remove '[' and everything after.
    Example:
      'QSTABILITY-20741170->QWINBUG-411242[ something]' -> 'QSTABILITY-20741170->QWINBUG-411242'
    """
    if s is None:
        return None
    s = str(s)
    if "[" in s:
        s = s.split("[", 1)[0]
    return s.strip()


_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
    "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
    "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
    "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
    "%H:%M:%S", "%H:%M",
]


def _coerce_value_for_mysql(header, value, col_maxlens=None):
    """
    Convert pandas/openpyxl cell values to MySQL-safe values using the same
    type heuristics as table creation.
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    n = str(header or "").lower()

    if any(k in n for k in ["date", "time", "added", "fetched", "built", "created_at", "updated_at"]):
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.strftime("%Y-%m-%d %H:%M:%S")

        val_str = str(value).strip()
        if not val_str or val_str.upper() in {"NAN", "NONE", "NULL"}:
            return None

        for fmt in _DATE_FORMATS:
            try:
                parsed = datetime.datetime.strptime(val_str, fmt)
                return parsed.strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass

        # Last-resort pandas parser for variants such as Excel serial/date strings.
        try:
            parsed = pd.to_datetime(val_str, errors="coerce")
            if pd.notna(parsed):
                return parsed.to_pydatetime().strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
        return None

    if any(k in n for k in ["count", "num", "occurrence", "version_int"]):
        val_str = re.sub(r"[^0-9\.\-]", "", str(value))
        try:
            return int(float(val_str)) if val_str else None
        except Exception:
            return None

    out_val = str(value)
    if header and str(header).lower() == "jira_traversal":
        out_val = _clean_jira_traversal(out_val)

    if col_maxlens:
        max_len = col_maxlens.get(str(header).lower())
        if max_len is not None and len(out_val) > max_len:
            out_val = out_val[:max_len]
    return out_val





def _pick_existing_column(table_columns, preferred_name=None):
    """
    Returns actual column name from table if present, case-insensitive.
    """
    lower_to_actual = {c.lower(): c for c in table_columns}

    if preferred_name and preferred_name.lower() in lower_to_actual:
        return lower_to_actual[preferred_name.lower()]

    return None


def _pick_updated_column(table_columns, preferred_name=None):
    """
    Returns the actual updated/timestamp column name if found.
    First tries preferred_name, then common fallback names.
    """
    lower_to_actual = {c.lower(): c for c in table_columns}

    if preferred_name and preferred_name.lower() in lower_to_actual:
        return lower_to_actual[preferred_name.lower()]

    for candidate in [
        "updated_at",
        "last_updated",
        "modified_at",
        "modified_time",
        "updated_time",
        "timestamp",
        "created_at"
    ]:
        if candidate in lower_to_actual:
            return lower_to_actual[candidate]

    return None


def _remove_old_duplicates_keep_latest(cursor, table_name, sanitized_pk, updated_at_column=None):
    """
    Removes old duplicate rows from DB and keeps only the latest row
    for each primary-key value using ROW_NUMBER() over updated_at_column.

    Notes:
    - Keeps rows where PK is NULL as-is
    - Requires MySQL 8+ because ROW_NUMBER() is used
    - If updated column is missing, this function skips safely
    """
    try:
        table_columns = _get_table_columns(cursor, table_name)
        actual_pk = _pick_existing_column(table_columns, sanitized_pk)

        if not actual_pk:
            logger.info(
                    f"WARN_INGEST_EXCEL: Cannot deduplicate `{table_name}`. "
                    f"PK column `{sanitized_pk}` not found."
                )

            return

        actual_updated_col = _pick_updated_column(table_columns, updated_at_column)

        if not actual_updated_col:
            logger.info(
                f"WARN_INGEST_EXCEL: Cannot remove older duplicates from `{table_name}` "
                f"because updated column '{updated_at_column or 'updated_at'}' was not found. "
                f"Skipping old-duplicate cleanup."
            )
            return

        safe_tmp_name = re.sub(r"[^a-zA-Z0-9_]", "_", f"tmp_{table_name}_dedup")
        safe_tmp_name = safe_tmp_name[:60]

        columns_sql = ", ".join([f"`{c}`" for c in table_columns])

        cursor.execute(f"DROP TEMPORARY TABLE IF EXISTS `{safe_tmp_name}`")
        cursor.execute(f"CREATE TEMPORARY TABLE `{safe_tmp_name}` LIKE `{table_name}`")

        dedupe_insert_sql = f"""
        INSERT INTO `{safe_tmp_name}` ({columns_sql})
        SELECT {columns_sql}
        FROM (
            SELECT
                {columns_sql},
                ROW_NUMBER() OVER (
                    PARTITION BY `{actual_pk}`
                    ORDER BY
                        CASE WHEN `{actual_updated_col}` IS NULL THEN 1 ELSE 0 END,
                        `{actual_updated_col}` DESC
                ) AS rn
            FROM `{table_name}`
            WHERE `{actual_pk}` IS NOT NULL
        ) ranked
        WHERE rn = 1
        """

        cursor.execute(dedupe_insert_sql)

        # Keep rows with NULL PK untouched
        null_pk_insert_sql = f"""
        INSERT INTO `{safe_tmp_name}` ({columns_sql})
        SELECT {columns_sql}
        FROM `{table_name}`
        WHERE `{actual_pk}` IS NULL
        """
        cursor.execute(null_pk_insert_sql)

        cursor.execute(f"DELETE FROM `{table_name}`")
        cursor.execute(f"INSERT INTO `{table_name}` ({columns_sql}) SELECT {columns_sql} FROM `{safe_tmp_name}`")
        cursor.execute(f"DROP TEMPORARY TABLE IF EXISTS `{safe_tmp_name}`")

        logger.info(
            f"DEBUG_INGEST_EXCEL: Older duplicates removed from `{table_name}` using "
            f"PK `{actual_pk}` and updated column `{actual_updated_col}`. Latest row kept."
        )

    except Exception as e:
        logger.info(
            f"WARN_INGEST_EXCEL: Could not remove old duplicates from `{table_name}`. Error: {e}"
        )
        logger.debug(traceback.format_exc())


def _ensure_unique_index(cursor, table_name, sanitized_pk, updated_at_column=None):
    """
    Ensure UNIQUE index exists on sanitized_pk.
    Before creating it, remove older duplicates and keep latest row.
    """
    try:
        cursor.execute(f"SHOW INDEX FROM `{table_name}`")
        rows = cursor.fetchall()
        for r in rows:
            # Non-unique = 0 means UNIQUE / PRIMARY
            if len(r) >= 5 and str(r[4]).lower() == sanitized_pk.lower() and int(r[1]) == 0:
                logger.info(
                    f"DEBUG_INGEST_EXCEL: UNIQUE/PRIMARY index already exists on `{table_name}`.`{sanitized_pk}`"
                )
                return

        # Remove old duplicates before adding unique index
        _remove_old_duplicates_keep_latest(cursor, table_name, sanitized_pk, updated_at_column)

        idx_name = f"uk_{sanitized_pk}"
        cursor.execute(f"ALTER TABLE `{table_name}` ADD UNIQUE KEY `{idx_name}` (`{sanitized_pk}`)")
        logger.info(
            f"DEBUG_INGEST_EXCEL: Added UNIQUE index `{idx_name}` on `{table_name}`.`{sanitized_pk}`"
        )

    except Exception as e:
        logger.info(
            f"WARN_INGEST_EXCEL: Could not create unique index on `{table_name}`.`{sanitized_pk}`. Error: {e}"
        )
        logger.debug(traceback.format_exc())


def _normalize_crid(value):
    if value is None:
        return ""
    return str(value).strip().upper()


def _title_has_team_marker(title: str, marker: str) -> bool:
    text = str(title or "").strip()
    if not text:
        return False
    pattern = rf'(?<![A-Z0-9]){re.escape(marker)}(?![A-Z0-9])'
    return bool(re.search(pattern, text, re.IGNORECASE))


def _derive_bonsai_test_team(stability_ticket, current_team, jira_title):
    ticket_val = str(stability_ticket or "").strip().upper()
    if not ticket_val.startswith("CHIPMD-"):
        return current_team

    if _title_has_team_marker(jira_title, "SD"):
        return "PDT_SD_HWPDT"
    if _title_has_team_marker(jira_title, "CH"):
        return "PDT_CH_HWPDT"
    return "PDT_QIPL_HWPDT"


def _apply_jira_team_rule_df(df: 'pd.DataFrame', target_name: str, sheet_name: str) -> 'pd.DataFrame':
    """
    Apply CHIPMD → PDT_QIPL_HWPDT test_team rule on a full DataFrame.
    For ALL targets (not just Bonsai).
    If stability_ticket starts with CHIPMD-, derive correct test_team.
    """
    sheet_val = sanitize_column_name(sheet_name).lower()
    if sheet_val not in {"jiras", "openjiras", "closed_jiras"}:
        return df

    # find columns case-insensitively
    col_lower = {c.lower(): c for c in df.columns}
    ticket_col = col_lower.get("stability_ticket") or col_lower.get("jira_id")
    team_col   = col_lower.get("test_team")
    title_col  = col_lower.get("jira_title") or col_lower.get("title")

    if not ticket_col or not team_col or not title_col:
        return df

    changed = 0
    for idx, row in df.iterrows():
        ticket_val = str(row[ticket_col] or "").strip().upper()
        if not ticket_val.startswith("CHIPMD-"):
            continue
        old_team = row[team_col]
        new_team = _derive_bonsai_test_team(row[ticket_col], row[team_col], row[title_col])
        if str(old_team or "") != str(new_team or ""):
            df.at[idx, team_col] = new_team
            changed += 1

    if changed:
        logger.info(
            f"JIRA_TEAM_RULE: target={target_name} sheet={sheet_name} "
            f"updated test_team for {changed} CHIPMD rows → PDT_QIPL_HWPDT"
        )
    return df


    sheet_val = sanitize_column_name(sheet_name).lower()
    if sheet_val not in {"jiras", "openjiras", "closed_jiras"}:
        return processed_values

    lower_to_index = {str(h).lower(): i for i, h in enumerate(headers)}
    ticket_idx = lower_to_index.get("stability_ticket")
    if ticket_idx is None:
        ticket_idx = lower_to_index.get("jira_id")
    team_idx = lower_to_index.get("test_team")
    title_idx = lower_to_index.get("jira_title")
    if title_idx is None:
        title_idx = lower_to_index.get("title")

    if ticket_idx is None or team_idx is None or title_idx is None:
        return processed_values

    updated = list(processed_values)
    old_team = updated[team_idx]
    new_team = _derive_bonsai_test_team(
        updated[ticket_idx],
        updated[team_idx],
        updated[title_idx],
    )
    updated[team_idx] = new_team

    ticket_val = str(updated[ticket_idx] or "").strip()
    if ticket_val.upper().startswith("CHIPMD-") and str(old_team or "") != str(new_team or ""):
        logger.info(
            "JIRA_TEAM_RULE: target=%s sheet=%s ticket=%s old_team=%s new_team=%s title=%s",
            target_name,
            sheet_name,
            ticket_val,
            old_team,
            new_team,
            str(updated[title_idx] or "")[:160],
        )
    return updated





def _build_reported_team_map(workbook):
    """
    Build CRID -> Reported Team mapping using the 3 special sheets:
      - OverallCrs
      - PDT_UniqueCrs
      - PDT_ReportedCrs

    Priority:
      1. PDT_Unique
      2. PDT_Reported
      3. OtherTeam Reported
    """
    required_sheets = [
        OVERALL_CRS_SHEET_NAME,
        PDT_UNIQUE_CRS_SHEET_NAME,
        PDT_REPORTED_CRS_SHEET_NAME,
    ]
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            return None

        def _extract_crids(sheet_name):
            sheet = workbook[sheet_name]
            first_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not first_row:
            raise ValueError(f"Sheet '{sheet_name}' has no header row")

        headers = [sanitize_column_name(cell.value) for cell in first_row if cell.value is not None and str(cell.value).strip() != ""]
        try:
            crid_idx = headers.index(CRID_COLUMN)
        except ValueError as exc:
            raise ValueError(f"Sheet '{sheet_name}' does not contain required '{CRID_COLUMN}' column") from exc

        crids = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row is None or crid_idx >= len(row):
                continue
            crid = _normalize_crid(row[crid_idx])
            if crid:
                crids.add(crid)
        return crids

    overall_crids = _extract_crids(OVERALL_CRS_SHEET_NAME)
    unique_crids = _extract_crids(PDT_UNIQUE_CRS_SHEET_NAME)
    reported_crids = _extract_crids(PDT_REPORTED_CRS_SHEET_NAME)

    mapping = {}
    for crid in overall_crids:
        if crid in unique_crids:
            mapping[crid] = "PDT_Unique"
        elif crid in reported_crids:
            mapping[crid] = "PDT_Reported"
        else:
            mapping[crid] = "OtherTeam Reported"

    return mapping


# Sheets that are read from Excel ONLY for building the reported_team map.
# They are NEVER written to DB as tables.
_MAP_ONLY_SHEETS = {PDT_UNIQUE_CRS_SHEET_NAME, PDT_REPORTED_CRS_SHEET_NAME}

# Column name written to overallcrs table
SEEN_IN_TARGETS_COLUMN = "seen_in_targets"


def _resolve_crinfo_file(unique_cr_file: str) -> str:
    """
    Given the path to the Unique CR workbook (xlsx or csv), find the
    matching *_CR_TAT_CrInfo_*.csv in the SAME folder.

    The CrInfo file lives alongside the Unique CR file:
      e.g.  .../2026_06_15/OVERALL_NORD_HQX_2026_06_15__CR_TAT_CrInfo_*.csv
            .../2026_06_15/OVERALL_NORD_HQX_2026_06_15__Unique_CRs-*.xlsx

    Returns the path to the latest CrInfo CSV found, or empty string if none.
    """
    try:
        folder = os.path.dirname(os.path.abspath(unique_cr_file))
        candidates = glob.glob(os.path.join(folder, "*_CR_TAT_CrInfo_*.csv"))
        if not candidates:
            # Also try one level up (in case unique_cr_file is directly in root)
            parent = os.path.dirname(folder)
            candidates = glob.glob(os.path.join(parent, "*_CR_TAT_CrInfo_*.csv"))
        if not candidates:
            return ""
        # Pick the newest
        try:
            return max(candidates, key=os.path.getmtime)
        except OSError:
            return candidates[0]
    except Exception as e:
        logger.warning(f"CRINFO: Could not resolve CrInfo file near '{unique_cr_file}': {e}")
        return ""


def _build_seen_in_targets_map(crinfo_file: str) -> dict:
    """
    Read *_CR_TAT_CrInfo_*.csv and build:
        { 'CR4544697': 'SA8797P.HQX;SA8797P_ADAS.HQX;SA8797P_FLEX.HQX', ... }

    Rules:
    - Group by CR column (case-insensitive header match)
    - Collect unique, sorted Target values per CR
    - Join with ';'
    - CRs with no target rows → not in map (will get NULL in DB)

    Returns empty dict on any failure (non-fatal).
    """
    if not crinfo_file or not os.path.isfile(crinfo_file):
        return {}
    try:
        # Read into memory to avoid SMB locking issues
        import io as _io
        _bytes = _io.BytesIO(open(crinfo_file, 'rb').read())
        df = pd.read_csv(_bytes, dtype=str, on_bad_lines='skip')

        if df.empty:
            return {}

        # Find CR and Target columns case-insensitively
        col_lower = {c.strip().lower(): c for c in df.columns}
        cr_col     = col_lower.get('cr')
        target_col = col_lower.get('target')

        if not cr_col or not target_col:
            logger.warning(
                f"CRINFO: Missing 'CR' or 'Target' column in '{crinfo_file}'. "
                f"Found columns: {list(df.columns)}"
            )
            return {}

        # Normalise: strip whitespace, drop empty
        df[cr_col]     = df[cr_col].astype(str).str.strip()
        df[target_col] = df[target_col].astype(str).str.strip()
        df = df[(df[cr_col] != '') & (df[cr_col].str.upper() != 'NAN')]
        df = df[(df[target_col] != '') & (df[target_col].str.upper() != 'NAN')]

        # Group CR → sorted unique targets joined by ';'
        seen_map = (
            df.groupby(cr_col)[target_col]
            .apply(lambda s: ';'.join(sorted(s.dropna().unique())))
            .to_dict()
        )

        logger.info(
            f"CRINFO: Built seen_in_targets map from '{os.path.basename(crinfo_file)}' "
            f"— {len(seen_map)} unique CRs, "
            f"{df[target_col].nunique()} unique targets: "
            f"{sorted(df[target_col].unique())}"
        )
        return seen_map

    except Exception as e:
        logger.warning(f"CRINFO: Failed to build seen_in_targets map from '{crinfo_file}': {e}")
        logger.debug(traceback.format_exc())
        return {}


def _ingest_overallcrs_pandas(
    unique_cr_file: str,
    target_db_prefix: str,
    conn,
    target_name: str,
) -> bool:
    """
    Fast pandas-based ingest for OverallCrs.
    Steps:
      1. Read OverallCrs, PDT_UniqueCrs, PDT_ReportedCrs from unique_cr_file.
      2. Build reported_team map from the two helper sheets.
      3. Append reported_team column to OverallCrs dataframe.
      4. Write to <target_db_prefix>_overallcrs (replace).
    Returns True on success, False on failure.
    """
    try:
        logger.info(f"INGEST_OVERALLCRS: Reading workbook '{unique_cr_file}' for '{target_name}'")

                        # --- read all 3 sheets --- read into memory first to avoid network temp-file issues
        try:
            import io as _io
            _ucr_bytes = _io.BytesIO(open(unique_cr_file, 'rb').read())
        except OSError as e:
            logger.error(f"INGEST_OVERALLCRS: Cannot read file (network/disk issue) '{unique_cr_file}': {e}")
            return False

        try:
            df_overall  = pd.read_excel(_ucr_bytes, sheet_name=OVERALL_CRS_SHEET_NAME,  dtype=str)
        except Exception as e:
            logger.warning(f"INGEST_OVERALLCRS: Cannot read sheet '{OVERALL_CRS_SHEET_NAME}': {e}")
            return False

        try:
            _ucr_bytes.seek(0)
            df_unique   = pd.read_excel(_ucr_bytes, sheet_name=PDT_UNIQUE_CRS_SHEET_NAME,   dtype=str)
        except Exception:
            df_unique = pd.DataFrame()

        try:
            _ucr_bytes.seek(0)
            df_reported = pd.read_excel(_ucr_bytes, sheet_name=PDT_REPORTED_CRS_SHEET_NAME, dtype=str)
        except Exception:
            df_reported = pd.DataFrame()

        if df_overall.empty:
            logger.warning(f"INGEST_OVERALLCRS: '{OVERALL_CRS_SHEET_NAME}' sheet is empty for '{target_name}'")
            return False

        # --- sanitize column names ---
        df_overall.columns  = [sanitize_column_name(c) for c in df_overall.columns]

                # --- find crid column (case-insensitive) ---
        crid_col = next((c for c in df_overall.columns if c.lower() == CRID_COLUMN), None)
        if not crid_col:
            logger.warning(f"INGEST_OVERALLCRS: No '{CRID_COLUMN}' column in OverallCrs for '{target_name}'")
            return False

        # --- build seen_in_targets map from CrInfo CSV (same folder as unique_cr_file) ---
        _crinfo_file = _resolve_crinfo_file(unique_cr_file)
        if _crinfo_file:
            logger.info(f"INGEST_OVERALLCRS: Found CrInfo file '{os.path.basename(_crinfo_file)}'")
            _seen_map = _build_seen_in_targets_map(_crinfo_file)
        else:
            f"INGEST_OVERALLCRS: seen_in_targets - "
            _seen_map = {}

        # --- build reported_team map ---
        def _get_crids(df):
            if df.empty:
                return set()
            col = next((c for c in df.columns if str(c).strip().lower() == CRID_COLUMN), None)
            if not col:
                return set()
            return set(df[col].dropna().astype(str).str.strip().str.upper())

        unique_crids   = _get_crids(df_unique)
        reported_crids = _get_crids(df_reported)

        def _classify(crid):
            v = str(crid).strip().upper() if crid and str(crid).strip() not in ('', 'NAN', 'NONE') else ''
            if not v:
                return 'OtherTeam Reported'
            if v in unique_crids:
                return 'PDT_Unique'
            if v in reported_crids:
                return 'PDT_Reported'
            return 'OtherTeam Reported'

        df_overall[REPORTED_TEAM_COLUMN] = df_overall[crid_col].apply(_classify)

        # --- append seen_in_targets column ---
        # Map each CRID to its semicolon-separated target list from CrInfo.
        # CRs not present in CrInfo get NULL (None).
        def _get_seen(crid):
            v = str(crid).strip() if crid and str(crid).strip() not in ('', 'NAN', 'NONE') else ''
            if not v:
                return None
            # Try exact match first, then upper-cased
            return _seen_map.get(v) or _seen_map.get(v.upper()) or None

        df_overall[SEEN_IN_TARGETS_COLUMN] = df_overall[crid_col].apply(_get_seen)

        _matched = df_overall[SEEN_IN_TARGETS_COLUMN].notna().sum()
        logger.info(
            f"INGEST_OVERALLCRS: seen_in_targets - "
            f"{_matched}/{len(df_overall)} CRs matched from CrInfo"
        )

        logger.info(
            f"INGEST_OVERALLCRS: '{target_name}' - {len(df_overall)} rows, "
            f"PDT_Unique={( df_overall[REPORTED_TEAM_COLUMN]=='PDT_Unique').sum()}, "
            f"PDT_Reported={(df_overall[REPORTED_TEAM_COLUMN]=='PDT_Reported').sum()}, "
            f"Other={(df_overall[REPORTED_TEAM_COLUMN]=='OtherTeam Reported').sum()}"
        )

        # --- replace NaN with None for MySQL ---
        df_overall = df_overall.where(pd.notnull(df_overall), None)

                # --- write to DB ---
        table_name = f"{target_db_prefix}_overallcrs"
        cursor = conn.cursor()
        try:
            # Drop & recreate for clean replace
            cursor.execute(f"DROP TABLE IF EXISTS `{table_name}`")

            # Build CREATE TABLE from df columns.
            # Use TEXT for known long-text columns AND any column whose actual
            # max value length exceeds 512 chars (prevents 1406 truncation errors).
            col_defs = []
            for col in df_overall.columns:
                n = col.lower()
                max_len = df_overall[col].dropna().astype(str).str.len().max() if len(df_overall) else 0
                if any(k in n for k in ["label","labels","description","comment","summary","details","notes","title"]) \
                        or (max_len and max_len > 512):
                    t = "TEXT"
                elif any(k in n for k in ["date","time","added","fetched","built","created_at","updated_at"]):
                    t = "DATETIME"
                elif any(k in n for k in ["count","num","occurrence"]):
                    t = "INT"
                elif n == SEEN_IN_TARGETS_COLUMN:
                    # Store as TEXT - can hold many target names separated by semicolon
                    t = "TEXT"
                else:
                    t = "VARCHAR(512)"
                col_defs.append(f"`{col}` {t} NULL")
            create_sql = f"CREATE TABLE `{table_name}` ({', '.join(col_defs)})"
            # Always CREATE the table first so the API never hits 1146
            # even if the subsequent INSERT fails.
            cursor.execute(create_sql)
            conn.commit()

            # Batch insert
            cols_sql = ", ".join([f"`{c}`" for c in df_overall.columns])
            placeholders = ", ".join(["%s"] * len(df_overall.columns))
            insert_sql = f"INSERT INTO `{table_name}` ({cols_sql}) VALUES ({placeholders})"

            batch_size = 500
            rows = df_overall.values.tolist()
            for i in range(0, len(rows), batch_size):
                batch = rows[i:i+batch_size]
                # convert any remaining non-None non-str to str
                clean = []
                for row in batch:
                    clean.append([
                        None if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)
                        for v in row
                    ])
                cursor.executemany(insert_sql, clean)
                conn.commit()

            logger.info(f"INGEST_OVERALLCRS: Inserted {len(rows)} rows into `{table_name}`")
            return True

        except Exception as e:
            logger.error(f"INGEST_OVERALLCRS: DB error for `{table_name}`: {e}")
            logger.debug(traceback.format_exc())
            try:
                conn.rollback()
            except Exception:
                pass
            return False
        finally:
            try:
                cursor.close()
            except Exception:
                pass

    except Exception as e:
        logger.error(f"INGEST_OVERALLCRS: General failure for '{target_name}': {e}")
        logger.debug(traceback.format_exc())
        return False


def upsert_dashboard_status(cursor, target_name, dashboard_latest_update, unique_cr_last_update=None):
    """
    Updates dashboard_latest_update (and optionally unique_cr_last_update)
    for an existing target row in pdt_stats_dashboard.dashboard_status.

    Assumes the row was already created by add_target_to_dashboard_status.
    If no row is found, it logs a warning and returns without inserting.
    """
    # No CREATE TABLE here; table is managed by dashboard_common
    # and has many more columns (bu, platform, etc.)
    if unique_cr_last_update is not None:
        update_sql = """
            UPDATE pdt_stats_dashboard.dashboard_status
            SET dashboard_latest_update = %s,
                unique_cr_last_update   = %s
            WHERE target_name = %s AND is_active = 1
        """
        cursor.execute(update_sql, (dashboard_latest_update, unique_cr_last_update, target_name))
    else:
        update_sql = """
            UPDATE pdt_stats_dashboard.dashboard_status
            SET dashboard_latest_update = %s
            WHERE target_name = %s AND is_active = 1
        """
        cursor.execute(update_sql, (dashboard_latest_update, target_name))
    if cursor.rowcount == 0:
        logger.info(
            f"WARN_INGEST_EXCEL: upsert_dashboard_status - "
            f"No active dashboard_status row found for target '{target_name}' to update dashboard_latest_update."
        )



def ingest_excel_data(excel_file_path, target_db_prefix, bu_key, target_name, unique_cr_path=None):
    """
    Loads Excel sheets into MySQL tables.
    OverallCrs is handled via pandas (_ingest_overallcrs_pandas) from the unique_cr workbook.
    PDT_UniqueCrs / PDT_ReportedCrs are NEVER written to DB (map-only).
    """
    conn = get_mysql_connection_db(bu_key=bu_key)
    if conn is None:
        logger.error(f"INGEST_EXCEL: Failed to connect to DB for BU '{bu_key}'.")
        return False

    cursor = conn.cursor()
    overall_success = True
    workbook = None

        
    try:
        unique_only_mode = not excel_file_path
        actual_excel_file = None if unique_only_mode else _resolve_actual_excel_file(excel_file_path)

        actual_unique_cr_file = actual_excel_file
        if unique_cr_path:
            try:
                actual_unique_cr_file = _resolve_latest_unique_cr_workbook(unique_cr_path)
            except FileNotFoundError as nf:
                logger.warning(f"WARN_INGEST_EXCEL: Unique CR workbook not found for '{target_name}': {nf}")
                if unique_only_mode:
                    raise
                actual_unique_cr_file = actual_excel_file

        # Drop stale map-only tables from previous runs
        try:
            for _stale in [f"{target_db_prefix}_pdt_uniquecrs", f"{target_db_prefix}_pdt_reportedcrs"]:
                cursor.execute(f"DROP TABLE IF EXISTS `{_stale}`")
            conn.commit()
        except Exception:
            pass

                
                # OverallCrs: pandas ingest from unique_cr workbook
        # Always run when actual_unique_cr_file is available — even if it is the
        # same file as excel_file_path (the OverallCrs sheet lives in both).
        if actual_unique_cr_file:
            logger.info(f"INFO_INGEST_EXCEL: Ingesting OverallCrs via pandas from '{actual_unique_cr_file}'")
            _ok = _ingest_overallcrs_pandas(actual_unique_cr_file, target_db_prefix, conn, target_name)
            if not _ok:
                logger.warning(f"INGEST_EXCEL: OverallCrs ingest failed for '{target_name}'")
                if unique_only_mode:
                    overall_success = False

        if unique_only_mode:
            if overall_success and actual_unique_cr_file:
                unique_cr_mtime = datetime.datetime.fromtimestamp(os.path.getmtime(actual_unique_cr_file))
                upsert_dashboard_status(
                    cursor, target_name,
                    dashboard_latest_update=unique_cr_mtime,
                    unique_cr_last_update=unique_cr_mtime,
                )
                conn.commit()
            return overall_success

                        # Get mtimes safely
        try:
            excel_last_updated     = datetime.datetime.fromtimestamp(os.path.getmtime(actual_excel_file))
            unique_cr_last_updated = datetime.datetime.fromtimestamp(os.path.getmtime(actual_unique_cr_file))
        except OSError as e:
            logger.error(f"INGEST_EXCEL: Cannot stat file (network/disk issue) for '{target_name}': {e}")
            return False
        dashboard_latest_update = max(excel_last_updated, unique_cr_last_updated)

        # Read excel file into memory (avoids Windows SMB temp-space issues)
        try:
            import io as _io
            _excel_bytes = _io.BytesIO(open(actual_excel_file, 'rb').read())
        except OSError as e:
            logger.error(f"INGEST_EXCEL: Cannot read excel file (network/disk issue) for '{target_name}': {e}")
            return False

        workbook = openpyxl.load_workbook(_excel_bytes, read_only=True)
        sheet_configs_map = {cfg["sheet_name"]: cfg for cfg in SHEET_CONFIG} if SHEET_CONFIG else {}

        # Exclude OverallCrs and map-only sheets from the main openpyxl loop
        sheets_to_process = [
            s for s in workbook.sheetnames
            if (not SHEET_CONFIG or s in sheet_configs_map)
            and s not in _MAP_ONLY_SHEETS
            and s != OVERALL_CRS_SHEET_NAME
        ]

        if not sheets_to_process:
            logger.info(f"WARN_INGEST_EXCEL: No sheets to process in '{actual_excel_file}' (check SHEET_CONFIG).")

        for raw_sheet_name in sheets_to_process:
            sheet_cfg          = sheet_configs_map.get(raw_sheet_name, {"sheet_name": raw_sheet_name})
            sheet_name         = sheet_cfg["sheet_name"]
            primary_key_column = sheet_cfg.get("primary_key_column")
            updated_at_column  = sheet_cfg.get("updated_at_column", "updated_at")

            if sheet_name not in workbook.sheetnames:
                logger.warning(f"INGEST_EXCEL: Sheet '{sheet_name}' not found for '{target_name}'. Skipping.")
                overall_success = False
                continue

            table_name = f"{target_db_prefix}_{sanitize_column_name(sheet_name)}"

            # ── Jiras sheets: read full DataFrame, fix test_team, then write ──
            is_jiras_sheet = sanitize_column_name(sheet_name).lower() in {"jiras", "openjiras", "closed_jiras"}
            if is_jiras_sheet:
                try:
                    _excel_bytes.seek(0)
                    df_jiras = pd.read_excel(_excel_bytes, sheet_name=sheet_name, dtype=str)
                    if df_jiras.empty:
                        logger.warning(f"INGEST_EXCEL: Sheet '{sheet_name}' is empty for '{target_name}'. Skipping.")
                        continue

                    # Sanitize column names
                    df_jiras.columns = [sanitize_column_name(c) for c in df_jiras.columns]

                    # ── Apply test_team rule on full DataFrame BEFORE writing ──
                    df_jiras = _apply_jira_team_rule_df(df_jiras, target_name, sheet_name)

                    # ── Verify CHIPMD rows got correct test_team ──
                    col_lower = {c.lower(): c for c in df_jiras.columns}
                    ticket_col = col_lower.get("stability_ticket") or col_lower.get("jira_id")
                    team_col   = col_lower.get("test_team")
                    if ticket_col and team_col:
                        chipmd_mask = df_jiras[ticket_col].astype(str).str.upper().str.startswith("CHIPMD-")
                        chipmd_count = chipmd_mask.sum()
                        hwpdt_count  = (df_jiras.loc[chipmd_mask, team_col] == "PDT_QIPL_HWPDT").sum() if chipmd_count else 0
                        logger.info(
                            f"INGEST_EXCEL: '{target_name}' sheet='{sheet_name}' "
                            f"CHIPMD rows={chipmd_count} with PDT_QIPL_HWPDT={hwpdt_count}"
                        )

                    # Replace NaN with None
                    df_jiras = df_jiras.where(pd.notnull(df_jiras), None)

                    # Create/truncate table
                    excel_sanitized_headers = list(df_jiras.columns)
                    create_table_sql = build_create_table_sql(table_name, excel_sanitized_headers, primary_key_column)
                    try:
                        cursor.execute(create_table_sql)
                        # Auto-add missing columns
                        cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
                        existing_db_cols = {row[0].lower() for row in cursor.fetchall()}
                        for h in excel_sanitized_headers:
                            if h.lower() not in existing_db_cols:
                                n = h.lower()
                                if any(k in n for k in ["labels","description","comment","summary","test_tas","details","scenario","notes","error_message","stack_trace","logs","title"]):
                                    nct = "TEXT"
                                elif any(k in n for k in ["date","time","added","fetched","built","created_at","updated_at"]):
                                    nct = "DATETIME"
                                elif any(k in n for k in ["count","num","occurrence","version_int"]):
                                    nct = "INT"
                                else:
                                    nct = "VARCHAR(255)"
                                cursor.execute(f"ALTER TABLE `{table_name}` ADD COLUMN `{h}` {nct} NULL")
                        if TRUNCATE_EXISTING_TABLES:
                            cursor.execute(f"TRUNCATE TABLE `{table_name}`")
                    except Error as e:
                        logger.error(f"INGEST_EXCEL: Could not create/verify table `{table_name}`: {e}")
                        overall_success = False
                        continue

                    # Batch insert
                    cols_sql     = ", ".join([f"`{h}`" for h in excel_sanitized_headers])
                    placeholders = ", ".join(["%s"] * len(excel_sanitized_headers))
                    insert_sql   = f"INSERT INTO `{table_name}` ({cols_sql}) VALUES ({placeholders})"
                    if primary_key_column:
                        sanitized_pk = sanitize_column_name(primary_key_column)
                        if any(h.lower() == sanitized_pk.lower() for h in excel_sanitized_headers):
                            _ensure_unique_index(cursor, table_name, sanitized_pk, updated_at_column)
                            updates = [f"`{h}` = VALUES(`{h}`)" for h in excel_sanitized_headers if h.lower() != sanitized_pk.lower()]
                            if updates:
                                insert_sql += " ON DUPLICATE KEY UPDATE " + ", ".join(updates)

                    col_maxlens = _get_table_col_maxlens(cursor, table_name)
                    rows = df_jiras.values.tolist()
                    batch_size = 500
                    for i in range(0, len(rows), batch_size):
                        batch = []
                        for row in rows[i:i+batch_size]:
                            clean_row = []
                            for j, v in enumerate(row):
                                h = excel_sanitized_headers[j]
                                clean_row.append(_coerce_value_for_mysql(h, v, col_maxlens))
                            batch.append(clean_row)
                        try:
                            cursor.executemany(insert_sql, batch)
                        except Error:
                            overall_success = False
                            for one in batch:
                                try:
                                    cursor.execute(insert_sql, one)
                                except Error as e2:
                                    logger.error(f"INGEST_EXCEL: Row insert error `{table_name}`: {e2}")
                    conn.commit()
                    logger.info(f"INGEST_EXCEL: Sheet '{sheet_name}' done for '{target_name}' ({len(rows)} rows).")

                except Exception as e:
                    logger.error(f"INGEST_EXCEL: Failed jiras sheet '{sheet_name}' for '{target_name}': {e}")
                    logger.debug(traceback.format_exc())
                    overall_success = False
                continue  # skip the openpyxl path below for jiras sheets

            # ── Non-jiras sheets: original openpyxl row-by-row path ──
            sheet = workbook[sheet_name]
            first_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
            if not first_row:
                logger.warning(f"INGEST_EXCEL: Sheet '{sheet_name}' has no header. Skipping.")
                overall_success = False
                continue

            excel_headers_raw          = [cell.value for cell in first_row]
            excel_headers_raw_filtered = [h for h in excel_headers_raw if h is not None and str(h).strip() != ""]
            if not excel_headers_raw_filtered:
                logger.warning(f"INGEST_EXCEL: No valid headers in sheet '{sheet_name}'. Skipping.")
                overall_success = False
                continue

            excel_sanitized_headers = [sanitize_column_name(h) for h in excel_headers_raw_filtered]
            table_name              = f"{target_db_prefix}_{sanitize_column_name(sheet_name)}"
            create_table_sql        = build_create_table_sql(table_name, excel_sanitized_headers, primary_key_column)

            try:
                cursor.execute(create_table_sql)
                try:
                    cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
                    existing_db_cols = {row[0].lower() for row in cursor.fetchall()}
                    for h in excel_sanitized_headers:
                        if h.lower() not in existing_db_cols:
                            n = h.lower()
                            if any(k in n for k in ["labels","description","comment","summary","test_tas","details","scenario","notes","error_message","stack_trace","logs","title"]):
                                nct = "TEXT"
                            elif any(k in n for k in ["date","time","added","fetched","built","created_at","updated_at"]):
                                nct = "DATETIME"
                            elif any(k in n for k in ["count","num","occurrence","version_int"]):
                                nct = "INT"
                            elif "image" in n:
                                nct = "TEXT"
                            else:
                                nct = "VARCHAR(255)"
                            cursor.execute(f"ALTER TABLE `{table_name}` ADD COLUMN `{h}` {nct} NULL")
                except Error as ae:
                    logger.info(f"WARN_INGEST_EXCEL: Could not auto-add columns to `{table_name}`: {ae}")

                if TRUNCATE_EXISTING_TABLES:
                    cursor.execute(f"TRUNCATE TABLE `{table_name}`")

            except Error as e:
                logger.error(f"INGEST_EXCEL: Could not create/verify table `{table_name}`. Error: {e}")
                logger.debug(traceback.format_exc())
                overall_success = False
                continue

            col_maxlens      = _get_table_col_maxlens(cursor, table_name)
            insert_cols_sql  = ", ".join([f"`{h}`" for h in excel_sanitized_headers])
            placeholders_sql = ", ".join(["%s"] * len(excel_sanitized_headers))
            final_insert_sql = f"INSERT INTO `{table_name}` ({insert_cols_sql}) VALUES ({placeholders_sql})"

            if primary_key_column:
                sanitized_pk = sanitize_column_name(primary_key_column)
                if any(h.lower() == sanitized_pk.lower() for h in excel_sanitized_headers):
                    _ensure_unique_index(cursor, table_name, sanitized_pk, updated_at_column)
                    updates = [f"`{h}` = VALUES(`{h}`)" for h in excel_sanitized_headers if h.lower() != sanitized_pk.lower()]
                    if updates:
                        final_insert_sql += " ON DUPLICATE KEY UPDATE " + ", ".join(updates)

            batch = []

            def _flush_batch(batch_rows):
                nonlocal overall_success
                if not batch_rows:
                    return
                try:
                    cursor.executemany(final_insert_sql, batch_rows)
                except Error:
                    overall_success = False
                    for one in batch_rows:
                        try:
                            cursor.execute(final_insert_sql, one)
                        except Error as e2:
                            logger.error(f"INGEST_EXCEL: MySQL error inserting row into `{table_name}`: {e2}")

            for row in sheet.iter_rows(min_row=2):
                row_values_raw = [cell.value for cell in row]
                if not any(v is not None for v in row_values_raw):
                    continue

                src_len = len(excel_headers_raw_filtered)
                if src_len > len(row_values_raw):
                    row_values_raw.extend([None] * (src_len - len(row_values_raw)))
                elif len(row_values_raw) > src_len:
                    row_values_raw = row_values_raw[:src_len]

                processed_values = []
                for i, val in enumerate(row_values_raw):
                    header = excel_sanitized_headers[i]
                    n = header.lower()
                    if any(k in n for k in ["labels","description","comment","summary","details","scenario","notes","error_message","stack_trace","logs","title"]):
                        ctg = "TEXT"
                    elif any(k in n for k in ["date","time","added","fetched","built","created_at","updated_at"]):
                        ctg = "DATETIME"
                    elif any(k in n for k in ["id","cr","ticket","jira","key","pl_id"]):
                        ctg = "VARCHAR(255)"
                    elif any(k in n for k in ["count","num","occurrence","version_int"]):
                        ctg = "INT"
                    elif "image" in n:
                        ctg = "TEXT"
                    else:
                        ctg = "VARCHAR(255)"

                    try:
                        if val is None:
                            out_val = None
                        elif ctg == "DATETIME":
                            if isinstance(val, (datetime.datetime, datetime.date)):
                                out_val = val.strftime("%Y-%m-%d %H:%M:%S")
                            else:
                                val_str = str(val).strip()
                                parsed  = None
                                for fmt in ["%Y-%m-%d %H:%M:%S","%Y-%m-%d %H:%M","%Y-%m-%d","%m/%d/%Y %H:%M:%S","%m/%d/%Y %H:%M","%m/%d/%Y","%d-%m-%Y %H:%M:%S","%d-%m-%Y %H:%M","%d-%m-%Y","%Y/%m/%d %H:%M:%S","%Y/%m/%d %H:%M","%Y/%m/%d","%H:%M:%S","%H:%M"]:
                                    try:
                                        parsed = datetime.datetime.strptime(val_str, fmt)
                                        break
                                    except ValueError:
                                        pass
                                out_val = parsed.strftime("%Y-%m-%d %H:%M:%S") if parsed else None
                        elif ctg == "INT":
                            val_str = re.sub(r"[^0-9\.\-]", "", str(val))
                            out_val = int(float(val_str)) if val_str else None
                        else:
                            out_val = str(val)

                        if header.lower() == "jira_traversal" and out_val is not None:
                            out_val = _clean_jira_traversal(out_val)
                        if isinstance(out_val, str):
                            max_len = col_maxlens.get(header.lower())
                            if max_len is not None and len(out_val) > max_len:
                                out_val = out_val[:max_len]
                        processed_values.append(out_val)
                    except Exception:
                        processed_values.append(None)

                batch.append(processed_values)

                if len(batch) >= 1000:

                    _flush_batch(batch)
                    batch = []

            _flush_batch(batch)
            conn.commit()
            logger.info(f"INGEST_EXCEL: Sheet '{sheet_name}' done for '{target_name}'.")

        if overall_success:
            upsert_dashboard_status(
                cursor, target_name,
                dashboard_latest_update=dashboard_latest_update,
                unique_cr_last_update=unique_cr_last_updated if actual_unique_cr_file else None,
                        )
            conn.commit()

    except OSError as e:
        logger.error(f"INGEST_EXCEL: Network/IO error for '{target_name}': {e}")
        logger.debug(traceback.format_exc())
        overall_success = False
        try:
            conn.rollback()
        except Exception:
            pass
    except Exception as e:
        logger.error(f"INGEST_EXCEL: General failure: {e}")
        logger.debug(traceback.format_exc())
        overall_success = False
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        try:
            workbook.close()
        except Exception:
            pass
        try:
            cursor.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass

    try:
        if not overall_success:
            logger.info("SYNC SKIPPED - ingest had errors.")
        else:
            from src.sync_central import (
                ensure_all_central_tables,
                sync_cr_master_for_target,
                sync_cr_relationships_for_target,
                sync_target_summary_for_target,
            )
            from src.cr_master_search import sync_cr_master_search_for_target
            table_results = ensure_all_central_tables()
            for tbl, ok in table_results.items():
                if not ok:
                    logger.warning(f"SYNC: ensure failed for {tbl}; continuing.")
            sync_cr_master_for_target(target_name)
            sync_cr_master_search_for_target(target_name)
            sync_cr_relationships_for_target(target_name)
            sync_target_summary_for_target(target_name)
    except Exception as e:
        logger.error(f"SYNC: Central sync failure for '{target_name}': {e}")

    return overall_success
