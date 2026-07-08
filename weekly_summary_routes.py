import json
import os
import tempfile
import csv as _csv_mod
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path

from flask import (
    Blueprint, jsonify, request, render_template,
    redirect, url_for, flash,
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename

from weekly_summary_service import current_monday_sunday, normalize_to_monday_sunday, write_target_weekly_summary, _target_weekly_path
from src.utils import get_mysql_connection_db

try:
    from config import BU_DATABASE_MAPPING, STATIC_BUSINESS_UNITS
except Exception:
    BearU_DATABASE_MAPPING = {}
    STATIC_BUSINESS_UNITS = {}

try:
    from dashboard_common import fetch_milestones_for_sp
except Exception:
    fetch_milestones_for_sp = None

weekly_summary_bp = Blueprint('weekly_summary_bp', __name__)

_QIPL_WEEKLY_EXCEL_DIR   = r'\\sphere\pdtstats\WeeklyQIPL_PDT_CR_TAT'
_QIPL_DB                 = 'pdt_stats_dashboard'
_QIPL_TABLE              = 'weekly_qipl_data'
_QIPL_IMPORT_AUDIT_TABLE = 'weekly_qipl_import_audit'
_SHAREPOINT_SUMMARY_TABLE = 'weekly_sharepoint_build_summary'
_CONSOLIDATE_SUMMARY_TABLE = 'weekly_sharepoint_consolidate_summary'
_UNIQUE_CR_RAW_DIR       = r'\\sphere\pdtstats\WeeklyUniqueCRs\RawData'
_UNIQUE_CR_EXCEL_BASE    = r'\\sphere\pdtstats\WeeklyUniqueCRs'
_FARM_KPI_DIR            = r'\\sphere\pdtstats\Farm_KPI'
_SWPDT_JSON_NETWORK      = r'\\sphere\pdtqipl_internal\PDTBuddy\SWPDT\qipl_SWPDT_job_summary.json'
_SWPDT_JSON_LOCAL        = str(Path(__file__).with_name('qipl_SWPDT_job_summary_local.json'))

# Smart Build Report - per-week Axiom build consolidate (separate from sharepoint consolidate)
_SP2_BUILD_CONSOLIDATE_TABLE = 'sp2_build_consolidate'
_SP2_BUILD_TYPE_OVERRIDES_TABLE = 'sp2_build_type_overrides'

# Consolidate report JSON snapshot (one file per week_end, static after save)
_CONSOLIDATE_JSON_NET   = r'\\sphere\pdtqipl_internal\PDTBuddy\consolidate'
_CONSOLIDATE_JSON_LOCAL = str(Path(__file__).parent / 'consolidate_snapshots')

# Consolidate report JSON snapshot (one file per week_end, static after save)
_CONSOLIDATE_JSON_NET   = r"\\sphere\pdtqipl_internal\PDTBuddy\consolidate"
_CONSOLIDATE_JSON_LOCAL = str(Path(__file__).parent / "consolidate_snapshots")

_QIPL_MIN_DATE           = date(2026, 5, 18)
_QIPL_EXE_OUTPUT_DONE_MARKER = 'Please check following files for output/report'



# Lightweight in-process cache for the Weekly landing Unique CR summary.
# Keyed by workbook path + size + mtime, so it refreshes automatically when the
# generated Excel changes. This avoids opening/scanning network Excel files on
# every page load.
_UCR_LANDING_COUNTS_CACHE = {}
_QIPL_EXCEL_FILES_CACHE = {'ts': 0.0, 'value': []}
_QIPL_SOURCE_FILES_CACHE = {'ts': 0.0, 'value': []}
_UCR_EXCEL_FILES_CACHE = {'ts': 0.0, 'value': []}
_UCR_RAW_FILES_CACHE = {'ts': 0.0, 'value': []}
_FARM_STATION_MAP_CACHE = {'ts': 0.0, 'value': {}}
_SHARE_LIST_TTL_SECONDS = 60
_FARM_MAP_TTL_SECONDS = 300

_CARDS = [
    {'key': 'cr_age',        'title': 'CR Age',               'icon': '\U0001f4c5'},
    {'key': 'cr_pie',        'title': 'CR Pie Chart',         'icon': '\U0001f967'},
    {'key': 'smart_build',   'title': 'Smart Build Report',   'icon': '\U0001f4ca'},
    {'key': 'unique_report', 'title': 'Unique Weekly Report', 'icon': '\U0001f4cb'},
    {'key': 'farm_testing',  'title': 'Farm Testing',         'icon': '\U0001f9ea'},
]


def _load_swpdt_json_payload() -> tuple[dict, str]:
    """
    Load SWPDT Axiom build summary.
    PRIMARY  : DB table pdt_stats_dashboard.axiom_job_summary
               (team IN QIPL/PDT/SD/CH, excludes /PDT/QIPL/HW)
    FALLBACK : JSON files (network share -> local)
    Returns (payload_dict, source_label) â€” same shape as before so all
    callers (_flatten_swpdt_build_entries etc.) work without any change.
    """
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if conn:
            cur = conn.cursor(dictionary=True)
            cur.execute("""
                SELECT job_id, build_id, build_name, software_product,
                       taxonomy_path, team, state, device_count, chip_ids,
                       submitted_at, started_at, ended_at,
                       axiom_hours, hours, product_flavor, submitter, site
                FROM `pdt_stats_dashboard`.`axiom_job_summary`
                WHERE team IN ('QIPL','PDT','SD','CH')
                  AND taxonomy_path NOT LIKE '/PDT/QIPL/HW%'
                ORDER BY submitted_at DESC
            """)
            rows = cur.fetchall() or []
            cur.execute(
                "SELECT MAX(updated_at) AS ts "
                "FROM `pdt_stats_dashboard`.`axiom_job_summary`"
            )
            ts_row = cur.fetchone() or {}
            cur.close()
            conn.close()
            if rows:
                builds = {}
                for r in rows:
                    jid = str(r.get('job_id') or '')
                    if not jid:
                        continue
                    # chip_ids stored as JSON string in DB
                    chips = r.get('chip_ids') or '[]'
                    if isinstance(chips, str):
                        try:
                            chips = json.loads(chips)
                        except Exception:
                            chips = []
                    # normalise submitted_at -> ISO string with Z suffix
                    submitted = str(r.get('submitted_at') or '').replace(' ', 'T')
                    if submitted and not submitted.endswith('Z'):
                        submitted += 'Z'
                    builds[jid] = {
                        'job_id':           jid,
                        'build_id':         str(r.get('build_id')         or ''),
                        'build_name':       str(r.get('build_name')       or ''),
                        'software_product': str(r.get('software_product') or ''),
                        'taxonomy_path':    str(r.get('taxonomy_path')    or ''),
                        'team':             str(r.get('team')             or ''),
                        'state':            str(r.get('state')            or ''),
                        'device_count':     int(r.get('device_count')     or 0),
                        'chip_ids':         chips if isinstance(chips, list) else [],
                        'submitted':        submitted,
                        'started_at':       str(r.get('started_at')       or ''),
                        'completed_at':     str(r.get('ended_at')         or ''),
                        'axiom_hours':      str(r.get('axiom_hours')      or ''),
                        'hours':            float(r['hours']) if r.get('hours') else None,
                        'product_flavor':   str(r.get('product_flavor')   or ''),
                        'submitter':        str(r.get('submitter')        or ''),
                        'site':             str(r.get('site')             or ''),
                    }
                generated_at = str(ts_row.get('ts') or '').replace(' ', 'T')
                if generated_at and not generated_at.endswith('Z'):
                    generated_at += 'Z'
                return {
                    'generated_at': generated_at,
                    'source':       'db:axiom_job_summary',
                    'total_builds': len(builds),
                    'builds':       builds,
                }, 'db:axiom_job_summary'
    except Exception as _exc:
        import logging as _log
        _log.getLogger('weekly_summary_routes').warning(
            '[SWPDT] DB read failed, falling back to JSON: %s', _exc
        )
    # â”€â”€ JSON fallback: network share first, then local â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    for path in (_SWPDT_JSON_NETWORK, _SWPDT_JSON_LOCAL):
        try:
            if path and os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as fh:
                    payload = json.load(fh)
                if isinstance(payload, dict):
                    return payload, path
        except Exception:
            continue
    return {}, ''


def _axiom_date_from_value(value):
    if not value:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text.replace('Z', '+00:00')).date()
    except Exception:
        return _safe_date(text[:10])


def _flatten_swpdt_build_entries(payload: dict) -> list:
    """Return normalized SWPDT build entries from either build-keyed or job-keyed JSON."""
    raw = payload.get('builds') if isinstance(payload, dict) else None
    if isinstance(raw, dict):
        values = list(raw.values())
    elif isinstance(raw, list):
        values = raw
    elif isinstance(payload.get('jobs'), list):
        values = payload.get('jobs') or []
    else:
        values = []

    by_build = {}
    for item in values:
        if not isinstance(item, dict):
            continue
        build_id = str(item.get('build_id') or item.get('build') or item.get('meta_build') or '').strip()
        if not build_id:
            continue
        submitted = str(item.get('submitted') or item.get('submitted_at') or item.get('created_at') or item.get('start_time') or '').strip()
        chip_ids = item.get('chip_ids') or item.get('chips') or []
        if isinstance(chip_ids, str):
            chip_ids = [c.strip() for c in chip_ids.replace(';', ',').split(',') if c.strip()]
        if not isinstance(chip_ids, list):
            chip_ids = []
        device_count = _safe_int(item.get('device_count') or item.get('devices') or item.get('number_of_devices')) or 0
        key = build_id.upper()
        rec = by_build.setdefault(key, {
            'build_id': build_id,
            'software_product': str(item.get('software_product') or item.get('pl_id') or '').strip(),
            'submitted': submitted,
            'device_count': 0,
            'chip_ids': set(),
            'job_ids': set(),
        })
        if submitted and (not rec.get('submitted') or submitted > rec.get('submitted')):
            rec['submitted'] = submitted
        if not rec.get('software_product') and item.get('software_product'):
            rec['software_product'] = str(item.get('software_product') or '').strip()
        rec['device_count'] = max(int(rec.get('device_count') or 0), int(device_count or 0))
        rec['chip_ids'].update(str(c).strip() for c in chip_ids if str(c).strip())
        if item.get('job_id'):
            rec['job_ids'].add(str(item.get('job_id')))

    out = []
    for rec in by_build.values():
        chips = sorted(rec.pop('chip_ids', set()))
        jobs = sorted(rec.pop('job_ids', set()))
        rec['chip_ids'] = chips
        rec['job_ids'] = jobs
        rec['device_count'] = max(int(rec.get('device_count') or 0), len(chips))
        out.append(rec)
    return out


def _swpdt_target_from_product(product: str) -> str:
    text = str(product or '').strip().strip('[](){}')
    # Some Axiom rows contain combined products like
    # "[Kailua.LA.1.10 + Sariska.MN.1.0]". Use the first product as the
    # display target for the picker; the full product remains as PL-ID.
    text = text.split('+', 1)[0].split(',', 1)[0].strip().strip('[](){}')
    parts = [p for p in text.split('.') if p]
    if len(parts) >= 2:
        return '.'.join(parts[:2])
    return text



def _swpdt_weekly_target_pl_options(week_start: date, week_end: date) -> dict:
    payload, path = _load_swpdt_json_payload()
    grouped = {}
    for build in _flatten_swpdt_build_entries(payload):
        submitted_date = _axiom_date_from_value(build.get('submitted'))
        if week_start and submitted_date and submitted_date < week_start:
            continue
        if week_end and submitted_date and submitted_date > week_end:
            continue
        pl_id = str(build.get('software_product') or '').strip()
        target = _swpdt_target_from_product(pl_id) or pl_id
        if not target and not pl_id:
            continue
        key = (target, pl_id)
        rec = grouped.setdefault(key, {'target': target, 'pl_id': pl_id, 'build_count': 0, 'device_count': 0, 'latest_submitted': ''})
        rec['build_count'] += 1
        rec['device_count'] += int(build.get('device_count') or 0)
        if str(build.get('submitted') or '') > str(rec.get('latest_submitted') or ''):
            rec['latest_submitted'] = str(build.get('submitted') or '')
    rows = sorted(grouped.values(), key=lambda r: (str(r.get('target') or '').lower(), str(r.get('pl_id') or '').lower()))
    return {
        'rows': rows,
        'source_path': path,
        'generated_at': payload.get('generated_at') if isinstance(payload, dict) else '',
        'total_available': len(rows),
    }


def _text_match_score(build: dict, target: str, pl_id: str) -> int:

    import re as _local_re
    hay = ' '.join([str(build.get('build_id') or ''), str(build.get('software_product') or '')]).lower()
    score = 0
    for raw in (pl_id, target):
        txt = str(raw or '').strip().lower()
        if not txt:
            continue
        candidates = {txt}
        candidates.add(txt.split('.')[0])
        candidates.update(t for t in _local_re.split(r'[^a-z0-9.]+', txt) if len(t) >= 3)
        if any(c and c in hay for c in candidates):
            score += 1
    return score


def _build_key_variants(value: str) -> set:
    """Return comparable keys for a build path/name.

    Axiom JSON often has a full UNC path while weekly_qipl_data usually stores
    only the final build folder/name. Compare both so already-reported builds
    are excluded from the no-crash candidate list.
    """
    raw = str(value or '').strip()
    if not raw:
        return set()
    base = raw.replace('\\', '/').rstrip('/').split('/')[-1]
    return {raw.upper(), base.upper()}


def _weekly_qipl_existing_build_keys(week_start: date, week_end: date, target: str = '', pl_id: str = '') -> set:
    """Build IDs already present in weekly_qipl_data for this week/target/PL."""
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return set()
    cur = conn.cursor(dictionary=True)
    keys = set()
    try:
        where = ['week_start=%s', 'week_end=%s']
        params = [week_start.isoformat(), week_end.isoformat()]
        target = str(target or '').strip()
        pl_id = str(pl_id or '').strip()
        if target or pl_id:
            vals = [v for v in (target, pl_id) if v]
            ph = ','.join(['%s'] * len(vals))
            where.append(f'(target IN ({ph}) OR pl_id IN ({ph}))')
            params.extend(vals)
            params.extend(vals)
        expr = _sp_build_match_sql_expr()
        cur.execute(f"""
            SELECT DISTINCT {expr} AS build_id
            FROM `{_QIPL_DB}`.`{_QIPL_TABLE}`
            WHERE {' AND '.join(where)}
        """, tuple(params))
        for rec in cur.fetchall() or []:
            keys.update(_build_key_variants(rec.get('build_id')))
    except Exception:
        return keys
    finally:
        cur.close(); conn.close()
    return keys


def _find_swpdt_no_crash_builds(week_start: date, week_end: date, target: str = '', pl_id: str = '', limit: int = 200) -> dict:
    payload, path = _load_swpdt_json_payload()
    existing_keys = _weekly_qipl_existing_build_keys(week_start, week_end, target=target, pl_id=pl_id)
    rows = []
    excluded_existing = 0
    for build in _flatten_swpdt_build_entries(payload):
        submitted_date = _axiom_date_from_value(build.get('submitted'))
        if week_start and submitted_date and submitted_date < week_start:
            continue
        if week_end and submitted_date and submitted_date > week_end:
            continue
        if (target or pl_id) and _text_match_score(build, target, pl_id) <= 0:
            continue
        if existing_keys and (_build_key_variants(build.get('build_id')) & existing_keys):
            excluded_existing += 1
            continue
        build['submitted_date'] = submitted_date.isoformat() if submitted_date else ''
        build['match_score'] = _text_match_score(build, target, pl_id)
        rows.append(build)
    rows.sort(key=lambda r: (int(r.get('match_score') or 0), str(r.get('submitted') or ''), int(r.get('device_count') or 0)), reverse=True)
    return {
        'rows': rows[:limit],
        'source_path': path,
        'generated_at': payload.get('generated_at') if isinstance(payload, dict) else '',
        'total_available': len(rows),
        'excluded_existing_weekly_qipl': excluded_existing,
    }



def _weekly_trend_completed_rows(rows: list) -> list:

    """Clean weekly trend JSON before sending it to the chart.

    Removes current/future weeks, normalizes old rolling/bad dates to the
    Monday-Sunday bucket, and keeps one row per week. This prevents bad labels
    such as adjacent 05/24 and 05/25 bars from appearing together.
    """
    today = date.today()
    by_week = {}
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        end_dt = _safe_date(row.get('week_end')) or _safe_date(row.get('week_start'))
        if not end_dt:
            continue
        ws, we = normalize_to_monday_sunday(week_end=end_dt)
        if we >= today:
            continue
        cleaned = dict(row)
        cleaned['week_start'] = ws.isoformat()
        cleaned['week_end'] = we.isoformat()
        cleaned['week_end_display'] = we.strftime('%m/%d')
        key = (cleaned['week_start'], cleaned['week_end'])
        score = sum(_safe_int(cleaned.get(k)) or 0 for k in ('overall_total_cr', 'total_cr', 'built', 'undisposed', 'invalid', 'dup'))
        prev = by_week.get(key)
        prev_score = sum(_safe_int(prev.get(k)) or 0 for k in ('overall_total_cr', 'total_cr', 'built', 'undisposed', 'invalid', 'dup')) if prev else -1
        if not prev or score > prev_score or (score == prev_score and str(cleaned.get('generated_at') or '') >= str(prev.get('generated_at') or '')):
            by_week[key] = cleaned
    out = list(by_week.values())
    out.sort(key=lambda r: str(r.get('week_start') or ''))
    for idx, row in enumerate(out, start=1):
        row['sr_no'] = idx
    return out


def _safe_date(val):
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()[:10]
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%d-%b-%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def _safe_int(val):
    try:
        return int(float(str(val).strip()))
    except Exception:
        return None


def _norm(v):
    if not v:
        return ''
    return str(v).strip().lower().replace(' ', '_').replace('-', '_').replace('/', '_')


def _is_snapdragon_auto_target(target: str) -> bool:
    return str(target or '').strip().lower().startswith('snapdragon_auto')


def _apply_snapdragon_pl_alias(target: str, pl_id: str, fallback_raw_pl: str = '') -> str:
    target = str(target or '').strip()
    pl_id = str(pl_id or '').strip()
    fallback_raw_pl = str(fallback_raw_pl or '').strip()
    if _is_snapdragon_auto_target(target):
        return pl_id or fallback_raw_pl
    return pl_id


def _jira_week(d: date):

    ws = d - timedelta(days=d.weekday())
    return ws, ws + timedelta(days=6)


# Ã¢Â”Â€Ã¢Â”Â€ DB Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€

def _ensure_weekly_qipl_table():
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return
    cur = conn.cursor()
    try:
        # only create if not exists Ã¢Â€Â” never drop
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{_QIPL_DB}`.`{_QIPL_TABLE}` (
                row_data          JSON         NOT NULL,
                week_start        DATE         NULL,
                week_end          DATE         NULL,
                jira_date         DATE         NULL,
                cr_date           DATE         NULL,
                jira_category     VARCHAR(128) NULL,
                cr_current_ticket VARCHAR(128) NULL,
                cr_si             VARCHAR(128) NULL,
                cr_title          TEXT         NULL,
                jira_title        TEXT         NULL,
                ticket_status     VARCHAR(128) NULL,
                resolution        VARCHAR(128) NULL,
                jira_reporter     VARCHAR(255) NULL,
                fetched_date      DATE         NULL,
                target            VARCHAR(255) NULL,
                jira_component    VARCHAR(255) NULL,
                pl_id             VARCHAR(128) NULL,
                host_name         VARCHAR(255) NULL,
                type_of_farm      VARCHAR(128) NULL,
                cr_status         VARCHAR(128) NULL,
                cr_area           VARCHAR(255) NULL,
                cr_age            INT          NULL,
                INDEX idx_week  (week_start, week_end),
                INDEX idx_jcat  (jira_category),
                INDEX idx_jdate (jira_date)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{_QIPL_DB}`.`{_QIPL_IMPORT_AUDIT_TABLE}` (
                id              BIGINT AUTO_INCREMENT PRIMARY KEY,
                file_key        CHAR(40)     NOT NULL,
                file_name       VARCHAR(512) NOT NULL,
                file_path       TEXT         NOT NULL,
                file_size       BIGINT       NOT NULL DEFAULT 0,
                file_mtime      DATETIME     NULL,
                week_start      DATE         NULL,
                week_end        DATE         NULL,
                status          VARCHAR(32)  NOT NULL,
                row_count       INT          NOT NULL DEFAULT 0,
                message         TEXT         NULL,
                started_by      VARCHAR(128) NULL,
                started_at      DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                finished_at     DATETIME     NULL,
                updated_at      DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_file_key (file_key),
                INDEX idx_week_status (week_start, week_end, status)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` (
                id                  BIGINT AUTO_INCREMENT PRIMARY KEY,
                week_start          DATE         NOT NULL,
                week_end            DATE         NOT NULL,
                target              VARCHAR(255) NOT NULL,
                pl_id               VARCHAR(255) NULL,
                build_type          VARCHAR(16)  NULL,
                build_label         VARCHAR(64)  NOT NULL,
                selected_items_json TEXT         NULL,
                hours               DECIMAL(12,2) NULL,
                devices             INT          NULL,
                mtbf                DECIMAL(12,2) NULL,
                crash_count         INT          NULL,
                crash_details       LONGTEXT     NULL,
                bu                  VARCHAR(64)  NULL,
                created_by          VARCHAR(128) NULL,
                created_at          DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at          DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_week_target_build (week_start, week_end, target, pl_id, build_type, build_label),
                INDEX idx_week_target (week_start, week_end, target)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` (
                id BIGINT AUTO_INCREMENT PRIMARY KEY,
                week_end DATE NOT NULL,
                bu VARCHAR(64) NULL,
                target VARCHAR(255) NULL,
                pl_id VARCHAR(255) NULL,
                timelines VARCHAR(512) NULL,
                pdt_test_status VARCHAR(64) NULL,
                number_of_devices INT NULL,
                number_of_builds INT NULL,
                total_hours DECIMAL(12,2) NULL,
                total_crashes INT NULL,
                unique_crs INT NULL,
                mtbf DECIMAL(12,2) NULL,
                updated_by VARCHAR(128) NULL,
                created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_week_target_pl (week_end, target, pl_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        # Migration: add columns and fix unique keys for existing tables.
        # Each ALTER is wrapped individually so a pre-existing column/index
        # does not abort the remaining migrations.
        for alter_sql in (
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD COLUMN pl_id VARCHAR(255) NULL AFTER target",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD COLUMN build_type VARCHAR(16) NULL AFTER pl_id",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD COLUMN mtbf DECIMAL(12,2) NULL AFTER devices",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD COLUMN meta_build VARCHAR(255) NULL AFTER build_label",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD COLUMN bu VARCHAR(64) NULL AFTER crash_details",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD COLUMN es_date DATE NULL AFTER bu",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD COLUMN fc_date DATE NULL AFTER es_date",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD COLUMN cs_date DATE NULL AFTER fc_date",
            # FIX: drop the old narrow unique key (week_start, week_end, target, build_label)
            # that was created before pl_id/build_type columns existed. The new key
            # (week_start, week_end, target, pl_id, build_type, build_label) allows
            # multiple PL-IDs under the same target without a duplicate-key error.
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` DROP INDEX uq_week_target_build",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` ADD UNIQUE KEY uq_week_target_build (week_start, week_end, target, pl_id, build_type, build_label)",
            f"ALTER TABLE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` ADD COLUMN pl_id VARCHAR(255) NULL AFTER target",
            f"ALTER TABLE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` DROP INDEX uq_week_target",
            f"ALTER TABLE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` ADD UNIQUE KEY uq_week_target_pl (week_end, target, pl_id)",
            # Dedup: add stability_ticket + meta_build + unique key to weekly_qipl_data
            f"ALTER TABLE `{_QIPL_DB}`.`{_QIPL_TABLE}` ADD COLUMN stability_ticket VARCHAR(128) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_QIPL_TABLE}` ADD COLUMN meta_build VARCHAR(255) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_QIPL_TABLE}` ADD UNIQUE KEY uq_stability_ticket (stability_ticket)",
        ):
            try:
                cur.execute(alter_sql)
            except Exception:
                pass
        conn.commit()
    except Exception:
        pass
    finally:
        cur.close()
        conn.close()


# Ã¢Â”Â€Ã¢Â”Â€ parse Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€

def _clean_cell(val):
    """Strip Excel color-code prefixes like #0000FF, #FF0000 etc. from cell values."""
    if not val:
        return val
    s = str(val).strip()
    import re
    # remove leading hex color codes e.g. #0000FF, #FF0000, #RRGGBB
    s = re.sub(r'^#[0-9A-Fa-f]{6}', '', s).strip()
    return s or None


def _build_row(raw_headers: list, values: list, filepath: str, uploaded_by: str):
    if not any(str(v).strip() for v in values if v is not None):
        return None

    full = {}
    for h, v in zip(raw_headers, values):
        k = str(h).strip() if h else ''
        if k:
            raw = str(v).strip() if v is not None and str(v).strip() else None
            full[k] = _clean_cell(raw)

    if not full:
        return None

    def _get(norm_name):
        for k, v in full.items():
            if _norm(k) == norm_name:
                return v
        return None

    jira_dt = _safe_date(_get('jira_date'))
    ws = we = None
    if jira_dt:
        ws, we = _jira_week(jira_dt)

    return {
        'week_start':        ws.isoformat() if ws else None,
        'week_end':          we.isoformat() if we else None,
        'row_data':          json.dumps(full, ensure_ascii=False),
        'jira_date':         jira_dt,
        'cr_date':           _safe_date(_get('cr_date')),
        'jira_category':     _get('jira_category'),
        'cr_current_ticket': _get('cr_current_ticket'),
        'cr_si':             _get('cr_si'),
        'cr_title':          _get('cr_title'),
        'jira_title':        _get('jira_title'),
        'ticket_status':     _get('cr_current_ticket_status'),
        'resolution':        _get('resolution'),
        'jira_reporter':     _get('jira_reporter'),
        'fetched_date':      _safe_date(_get('fetched_date')),
        'target':            _get('target'),
        'jira_component':    _get('jira_component'),
        'pl_id':             _get('pl_id') or _get('pl-id'),
        'host_name':         _get('host_name'),
        'type_of_farm':      _get('type_of_farm'),
        'cr_status':         _get('cr_status'),
        'cr_area':           _get('cr_area'),
        'cr_age':            _safe_int(_get('cr_age')),
        'stability_ticket':  _get('stability_ticket') or _get('stabilityticket'),
        'meta_build':        _get('metabuild') or _get('meta_build'),
    }


def _parse_file(filepath: str, uploaded_by: str):
    ext = Path(filepath).suffix.lower()
    if ext == '.csv':
        return _parse_csv(filepath, uploaded_by)
    return _parse_excel(filepath, uploaded_by)


def _parse_csv(filepath: str, uploaded_by: str):
    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            rows_out = []
            with open(filepath, newline='', encoding=enc) as fh:
                reader = _csv_mod.reader(fh)
                headers = next(reader, [])
                if not headers:
                    return [], []
                for raw in reader:
                    r = _build_row(headers, raw, filepath, uploaded_by)
                    if r:
                        rows_out.append(r)
            return rows_out, headers
        except UnicodeDecodeError:
            continue
    return [], []


def _parse_excel(filepath: str, uploaded_by: str):
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    rows_out = []
    headers  = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row < 2:
            continue
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            row  = _build_row(headers, vals, filepath, uploaded_by)
            if row:
                rows_out.append(row)
    return rows_out, headers


# Ã¢Â”Â€Ã¢Â”Â€ upsert Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€

def _upsert_rows(rows: list):
    if not rows:
        return 0, 0, 'No rows'
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return 0, 0, 'DB connection failed'
    cur = conn.cursor()
    try:
        for stmt in (
            "SET SESSION wait_timeout=600",
            "SET SESSION interactive_timeout=600",
            "SET SESSION net_read_timeout=600",
            "SET SESSION net_write_timeout=600",
        ):
            try:
                cur.execute(stmt)
            except Exception:
                pass

        weeks = set(
            (r['week_start'], r['week_end'])
            for r in rows if r.get('week_start') and r.get('week_end')
        )

        # Step 1: find common rows (stability_tickets already in table for this week)
        # Step 2: delete those week rows from table (to get latest JIRA status)
        # Step 3: insert all rows fresh from CSV
        deleted = 0
        for ws_del, we_del in weeks:
            cur.execute(
                f"DELETE FROM `{_QIPL_DB}`.`{_QIPL_TABLE}`"
                " WHERE week_start=%s AND week_end=%s",
                (ws_del, we_del)
            )
            deleted += cur.rowcount
        conn.commit()

        sql = f"""
            INSERT INTO `{_QIPL_DB}`.`{_QIPL_TABLE}`
            (row_data, week_start, week_end, jira_date, cr_date,
             jira_category, cr_current_ticket, cr_si, cr_title, jira_title,
             ticket_status, resolution, jira_reporter, fetched_date,
             target, jira_component, pl_id, host_name, type_of_farm,
             cr_status, cr_area, cr_age, stability_ticket, meta_build)
            VALUES
            (%(row_data)s, %(week_start)s, %(week_end)s, %(jira_date)s, %(cr_date)s,
             %(jira_category)s, %(cr_current_ticket)s, %(cr_si)s, %(cr_title)s,
             %(jira_title)s, %(ticket_status)s, %(resolution)s, %(jira_reporter)s,
             %(fetched_date)s, %(target)s, %(jira_component)s, %(pl_id)s,
             %(host_name)s, %(type_of_farm)s, %(cr_status)s, %(cr_area)s, %(cr_age)s,
             %(stability_ticket)s, %(meta_build)s)
        """

        BATCH = 500
        inserted = 0
        for i in range(0, len(rows), BATCH):
            cur.executemany(sql, rows[i:i + BATCH])
            inserted += cur.rowcount
            conn.commit()

        return inserted, deleted, 'OK'
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        return 0, 0, str(exc)
    finally:
        cur.close()
        conn.close()


# Ã¢Â”Â€Ã¢Â”Â€ fetch Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€

def _fetch_rows(week_start: date, week_end: date) -> list:
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT row_data, week_start, week_end, jira_date, cr_date, jira_category,
                   cr_current_ticket, cr_si, cr_title, jira_title,
                   ticket_status, resolution, jira_reporter, fetched_date,
                   target, jira_component, pl_id, host_name, type_of_farm,
                   cr_status, cr_area, cr_age, stability_ticket, meta_build
            FROM `{_QIPL_DB}`.`{_QIPL_TABLE}`
            WHERE week_start=%s AND week_end=%s
            ORDER BY cr_age DESC
        """, (week_start.isoformat(), week_end.isoformat()))
        result = []
        for row in cur.fetchall() or []:
            try:
                data = json.loads(row['row_data'])
            except Exception:
                data = {}
            for k in ('week_start', 'week_end', 'jira_date', 'cr_date', 'jira_category', 'cr_current_ticket',
                      'cr_si', 'cr_title', 'jira_title', 'ticket_status',
                      'resolution', 'jira_reporter', 'fetched_date', 'target',
                      'jira_component', 'pl_id', 'host_name', 'type_of_farm',
                      'cr_status', 'cr_area', 'cr_age', 'stability_ticket', 'meta_build'):
                v = row.get(k)
                data[k] = str(v)[:10] if isinstance(v, (date, datetime)) else v
            result.append(data)
        return result
    except Exception:
        return []
    finally:
        cur.close()
        conn.close()


def _get_available_weeks() -> list:
    """Return Weekly QIPL DB weeks from May 1, 2026 onward.

    Do not hide future-dated weeks here. The upstream QIPL/Unique-CR reports can
    publish a week-ending workbook before that week has elapsed, and those rows
    should still be selectable if they already exist in the database.
    """
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = conn.cursor()
    try:
        cur.execute(f"""
            SELECT DISTINCT week_start, week_end
            FROM `{_QIPL_DB}`.`{_QIPL_TABLE}`
            WHERE week_start IS NOT NULL
              AND week_end IS NOT NULL
              AND week_end >= %s
            ORDER BY week_start DESC LIMIT 52
        """, (_QIPL_MIN_DATE.isoformat(),))
        return [(r[0], r[1]) for r in cur.fetchall()]
    except Exception:
        return []
    finally:
        cur.close()
        conn.close()


def _get_week_ranges(n=16):
    """Fallback completed Monday-Sunday week ranges from May 1, 2026 onward."""
    today = date.today()
    this_monday = today - timedelta(days=today.weekday())
    lm = this_monday - timedelta(days=7)
    ranges = [(lm - timedelta(weeks=i), lm - timedelta(weeks=i) + timedelta(days=6)) for i in range(n)]
    return [(s, e) for s, e in ranges if e >= _QIPL_MIN_DATE and e < today]


def _list_excel_files():
    """List upload/select files from the QIPL share with a short TTL cache."""
    import time
    now = time.time()
    if now - float(_QIPL_EXCEL_FILES_CACHE.get('ts') or 0) < _SHARE_LIST_TTL_SECONDS:
        return list(_QIPL_EXCEL_FILES_CACHE.get('value') or [])
    try:
        if not os.path.isdir(_QIPL_WEEKLY_EXCEL_DIR):
            return []
        value = sorted(
            [f for f in os.listdir(_QIPL_WEEKLY_EXCEL_DIR)
             if f.lower().endswith(('.xlsx', '.xls', '.csv')) and not f.startswith('~$')],
            reverse=True
        )
        _QIPL_EXCEL_FILES_CACHE.update({'ts': now, 'value': value})
        return list(value)
    except Exception:
        return list(_QIPL_EXCEL_FILES_CACHE.get('value') or [])


def _qipl_file_date(fname: str):
    """Extract generated date from QIPL source filenames like 2026y_05m_23d."""
    import re
    m = re.search(r'(\d{4})y_(\d{2})m_(\d{2})d', str(fname or ''), re.IGNORECASE)
    if not m:
        return None
    try:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except Exception:
        return None


def _qipl_exe_output_log_ready(csv_path: str) -> tuple[bool, str]:
    """Return true only after matching QIPL_CR_AGE_Exe_output_*.txt reports completion."""
    import re
    try:
        folder = os.path.dirname(csv_path)
        csv_name = os.path.basename(csv_path)
        m = re.search(r'(\d{4}y_\d{2}m_\d{2}d)(?:_(\d{2}h_\d{2}m_\d{2}s))?', csv_name, re.IGNORECASE)
        if not m or not folder or not os.path.isdir(folder):
            return False, 'completion_log_lookup_failed'

        date_token = m.group(1).lower()
        datetime_token = f"{date_token}_{m.group(2).lower()}" if m.group(2) else date_token
        exact_prefix = f'qipl_cr_age_exe_output_{datetime_token}'
        date_prefix = f'qipl_cr_age_exe_output_{date_token}'

        exact_matches = []
        date_matches = []
        for fname in os.listdir(folder):
            lower = fname.lower()
            if not lower.endswith('.txt'):
                continue
            path = os.path.join(folder, fname)
            if lower.startswith(exact_prefix):
                exact_matches.append(path)
            elif lower.startswith(date_prefix):
                date_matches.append(path)

        candidates = exact_matches or date_matches
        if not candidates:
            return False, 'completion_log_missing'
        candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        log_path = candidates[0]
        with open(log_path, 'r', encoding='utf-8', errors='ignore') as fh:
            if _QIPL_EXE_OUTPUT_DONE_MARKER.casefold() in fh.read().casefold():
                return True, 'completion_log_ready'
        return False, f'completion_marker_missing:{os.path.basename(log_path)}'
    except PermissionError:
        return False, 'completion_log_locked'
    except Exception as exc:
        return False, f'completion_log_error:{exc}'


def _list_qipl_source_files() -> list:

    """
    List auto-load candidates from the QIPL weekly share.
    Prefer consolidated CR_TAT_Jira CSV; ignore CrInfo/error/blacklist/log files.
    Short TTL cache avoids slow repeated UNC directory scans during page load.
    """
    import time
    now = time.time()
    if now - float(_QIPL_SOURCE_FILES_CACHE.get('ts') or 0) < _SHARE_LIST_TTL_SECONDS:
        return [dict(x) for x in (_QIPL_SOURCE_FILES_CACHE.get('value') or [])]
    try:
        if not os.path.isdir(_QIPL_WEEKLY_EXCEL_DIR):
            return []
        files = []
        for fname in os.listdir(_QIPL_WEEKLY_EXCEL_DIR):
            lower = fname.lower()
            if fname.startswith('~$'):
                continue
            if not lower.endswith(('.csv', '.xlsx', '.xls')):
                continue
            if 'cr_tat_jira' not in lower:
                continue
            fdate = _qipl_file_date(fname)
            if not fdate:
                continue
            path = os.path.join(_QIPL_WEEKLY_EXCEL_DIR, fname)
            try:
                mtime = os.path.getmtime(path)
            except Exception:
                mtime = 0
            files.append({'name': fname, 'path': path, 'file_date': fdate, 'mtime': mtime})
        files.sort(key=lambda x: (x['file_date'], x['mtime']), reverse=True)
        _QIPL_SOURCE_FILES_CACHE.update({'ts': now, 'value': files})
        return [dict(x) for x in files]
    except Exception:
        return [dict(x) for x in (_QIPL_SOURCE_FILES_CACHE.get('value') or [])]


def _get_qipl_file_weeks() -> list:
    """Return source-file weeks from May 1, 2026 onward.

    Future week-ending files are included when present so pre-published weekly
    data, e.g. a 06/07 week, can be selected and displayed.
    """
    seen = set()
    weeks = []
    for entry in _list_qipl_source_files():
        ws, we = _jira_week(entry['file_date'])
        if we < _QIPL_MIN_DATE:
            continue
        key = (ws, we)
        if key in seen:
            continue
        seen.add(key)
        weeks.append(key)
    weeks.sort(key=lambda x: x[0], reverse=True)
    return weeks


def _merge_week_ranges(*ranges, limit=52) -> list:
    """Merge DB weeks and file-derived weeks, newest first.

    Do not filter out future week_end values here. Fallback-generated ranges are
    already limited to completed weeks, while DB/file-derived ranges represent
    real data that should be visible as soon as it is available.
    """
    seen = set()
    merged = []
    for group in ranges:
        for ws, we in group or []:
            if not ws or not we:
                continue
            if we < _QIPL_MIN_DATE:
                continue
            key = (ws, we)
            if key in seen:
                continue
            seen.add(key)
            merged.append(key)
    merged.sort(key=lambda x: x[0], reverse=True)
    return merged[:limit]


def _qipl_file_fingerprint(path: str) -> dict:
    """Stable file identity based on path + current size + mtime."""
    import hashlib
    st = os.stat(path)
    mtime_dt = datetime.fromtimestamp(st.st_mtime)
    raw = f"{os.path.abspath(path).lower()}|{st.st_size}|{int(st.st_mtime)}"
    return {
        'key': hashlib.sha1(raw.encode('utf-8', errors='ignore')).hexdigest(),
        'size': int(st.st_size),
        'mtime': mtime_dt,
    }


def _is_qipl_file_ready(path: str, min_age_seconds: int = 180, settle_seconds: float = 2.0) -> tuple[bool, str]:
    """
    Avoid reading CSV while producer is still writing it.
    Ready means: exists, non-empty, older than min_age_seconds, size stable,
    and openable for shared read.
    """
    try:
        if not path or not os.path.isfile(path):
            return False, 'missing'
        st1 = os.stat(path)
        if st1.st_size <= 0:
            return False, 'empty'
        age = (datetime.now() - datetime.fromtimestamp(st1.st_mtime)).total_seconds()
        if age < min_age_seconds:
            return False, f'too_new_{int(age)}s'

        import time
        time.sleep(settle_seconds)
        st2 = os.stat(path)
        if st1.st_size != st2.st_size or int(st1.st_mtime) != int(st2.st_mtime):
            return False, 'still_changing'

        # Try opening. On Windows/share this can fail if another process holds
        # an exclusive write lock.
        with open(path, 'rb') as fh:
            fh.read(1024)


        log_ready, log_reason = _qipl_exe_output_log_ready(path)
        if not log_ready:
            return False, log_reason

        return True, 'ready'

    except PermissionError:
        return False, 'locked'
    except OSError as exc:
        return False, f'os_error:{exc}'
    except Exception as exc:
        return False, f'error:{exc}'


def _get_import_audit(file_key: str):
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return None
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT * FROM `{_QIPL_DB}`.`{_QIPL_IMPORT_AUDIT_TABLE}`
            WHERE file_key=%s
            LIMIT 1
        """, (file_key,))
        return cur.fetchone()
    except Exception:
        return None
    finally:
        cur.close()
        conn.close()


def _begin_import_audit(path: str, week_start: date, week_end: date, username: str) -> tuple[bool, str, dict]:
    """
    Atomically claim a file for import. Returns (claimed, reason, fingerprint).
    Existing done/in_progress rows prevent duplicate/concurrent imports.
    """
    fp = _qipl_file_fingerprint(path)
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return False, 'db_connection_failed', fp
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT status, row_count FROM `{_QIPL_DB}`.`{_QIPL_IMPORT_AUDIT_TABLE}`
            WHERE file_key=%s
            LIMIT 1
        """, (fp['key'],))
        existing = cur.fetchone()
        if existing:
            status = str(existing.get('status') or '')
            if status == 'done':
                return False, 'already_imported', fp
            if status == 'in_progress':
                return False, 'import_in_progress', fp

        cur.execute(f"""
            INSERT INTO `{_QIPL_DB}`.`{_QIPL_IMPORT_AUDIT_TABLE}`
            (file_key, file_name, file_path, file_size, file_mtime,
             week_start, week_end, status, row_count, message, started_by, started_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,'in_progress',0,NULL,%s,NOW())
            ON DUPLICATE KEY UPDATE
                status='in_progress', row_count=0, message=NULL,
                started_by=VALUES(started_by), started_at=NOW(), finished_at=NULL
        """, (
            fp['key'], os.path.basename(path), path, fp['size'], fp['mtime'],
            week_start.isoformat(), week_end.isoformat(), username or 'auto'
        ))
        conn.commit()
        return True, 'claimed', fp
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        return False, f'audit_begin_failed:{exc}', fp
    finally:
        cur.close()
        conn.close()


def _finish_import_audit(file_key: str, status: str, row_count: int = 0, message: str = ''):
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return
    cur = conn.cursor()
    try:
        cur.execute(f"""
            UPDATE `{_QIPL_DB}`.`{_QIPL_IMPORT_AUDIT_TABLE}`
            SET status=%s, row_count=%s, message=%s, finished_at=NOW()
            WHERE file_key=%s
        """, (status, int(row_count or 0), str(message or '')[:2000], file_key))
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        cur.close()
        conn.close()


def _find_qipl_source_file_for_week(week_start: date, week_end: date) -> str:
    """Find the latest ready CR_TAT_Jira source file generated within the selected week."""
    candidates = [
        e for e in _list_qipl_source_files()
        if week_start <= e['file_date'] <= week_end and os.path.isfile(e['path'])
    ]
    candidates.sort(key=lambda x: (x['file_date'], x['mtime']), reverse=True)
    for entry in candidates:
        ready, reason = _is_qipl_file_ready(entry['path'])
        if not ready:
            continue
        fp = _qipl_file_fingerprint(entry['path'])
        audit = _get_import_audit(fp['key'])
        if audit and str(audit.get('status') or '') in ('done', 'in_progress'):
            continue
        return entry['path']
    return ''


def _auto_load_qipl_week(week_start: date, week_end: date, username: str) -> dict:
    """
    Auto-import the matching weekly CR_TAT_Jira CSV only when safe.
    Duplicate/concurrent imports are prevented by the import audit table, and
    files still being generated are skipped until stable.
    """
    src_path = _find_qipl_source_file_for_week(week_start, week_end)
    if not src_path:
        return {'loaded': False, 'reason': 'no_ready_unimported_source_file', 'path': ''}

    ready, ready_reason = _is_qipl_file_ready(src_path)
    if not ready:
        return {'loaded': False, 'reason': ready_reason, 'path': src_path}

    claimed, claim_reason, fp = _begin_import_audit(src_path, week_start, week_end, username)
    if not claimed:
        return {'loaded': False, 'reason': claim_reason, 'path': src_path}

    try:
        rows, raw_headers = _parse_file(src_path, username or 'auto')
        ws = week_start.isoformat()
        we = week_end.isoformat()
        selected_rows = [
            r for r in rows
            if r.get('week_start') == ws and r.get('week_end') == we
        ]
        if not selected_rows:
            msg = f"No rows for selected week. Headers: {[str(h) for h in raw_headers[:10]]}"
            _finish_import_audit(fp['key'], 'failed', 0, msg)
            return {'loaded': False, 'reason': 'no_rows_for_selected_week', 'path': src_path}

        inserted, deleted, msg = _upsert_rows(selected_rows)
        if inserted:
            _finish_import_audit(fp['key'], 'done', inserted, msg)
        else:
            _finish_import_audit(fp['key'], 'failed', 0, msg)
        return {
            'loaded': bool(inserted),
            'inserted': inserted,
            'deleted': deleted,
            'message': msg,
            'path': src_path,
        }
    except Exception as exc:
        _finish_import_audit(fp['key'], 'failed', 0, str(exc))
        return {'loaded': False, 'reason': 'error', 'message': str(exc), 'path': src_path}


# Ã¢Â”Â€Ã¢Â”Â€ card data Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€

def _build_card_data(table_rows: list) -> dict:
    cr_mapped = [
        r for r in table_rows
        if str(r.get('jira_category') or '').strip().lower() == 'cr mapped'
    ]

    # deduplicate by cr_current_ticket Ã¢Â€Â” same as detail page
    seen_cr = set()
    unique_crs = []
    for r in cr_mapped:
        cr_key = str(r.get('cr_current_ticket') or '').strip()
        if cr_key and cr_key not in seen_cr:
            seen_cr.add(cr_key)
            unique_crs.append(r)

    # age buckets from unique CRs only
    age_buckets = {'0-20': 0, '21-40': 0, '41-60': 0, '61-90': 0, '90+': 0}
    for row in unique_crs:
        try:
            age = int(row.get('cr_age') or 0)
        except Exception:
            age = 0
        if   age <= 20: age_buckets['0-20']  += 1
        elif age <= 40: age_buckets['21-40'] += 1
        elif age <= 60: age_buckets['41-60'] += 1
        elif age <= 90: age_buckets['61-90'] += 1
        else:           age_buckets['90+']   += 1

    pie_status = Counter(str(r.get('cr_status') or 'Unknown').strip() for r in unique_crs)
    pie_area   = Counter(str(r.get('cr_area')   or 'Unknown').strip() for r in unique_crs)

    # New / Old CR counts on unique CRs
    def _week_num(val):
        try:
            d = val if hasattr(val, 'year') else datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
            return d.isocalendar()[1]
        except Exception:
            return None

    seen_new = set()
    seen_old = set()
    for r in unique_crs:
        cr_key  = str(r.get('cr_current_ticket') or '').strip()
        if not cr_key:
            continue
        jira_wk = _week_num(r.get('jira_date'))
        cr_wk   = _week_num(r.get('cr_date'))
        if jira_wk and cr_wk:
            if cr_wk >= jira_wk:
                seen_new.add(cr_key)
            else:
                seen_old.add(cr_key)
    total_new = len(seen_new)
    total_old = len(seen_old - seen_new)

    sp_rows = [r for r in table_rows if not str(r.get('cr_current_ticket') or '').strip()]

    return {
        'age_buckets':     age_buckets,
        'pie_status_data': [{'name': k, 'y': v} for k, v in sorted(pie_status.items(), key=lambda x: -x[1])],
        'pie_area_data':   [{'name': k, 'y': v} for k, v in sorted(pie_area.items(),   key=lambda x: -x[1])],
        'sp_rows':         sp_rows,
        'total_new_crs':   total_new,
        'total_old_crs':   total_old,
        'unique_summary':  {'total': len(unique_crs), 'total_all': len(table_rows),
                            'unmapped': len(sp_rows), 'built': 0, 'undisposed': 0},
        'cr_mapped_count': len(unique_crs),
    }


def _norm_target(t: str) -> str:
    """Normalize target name so case variants merge."""
    import re
    if not t:
        return 'Unknown'
    parts = str(t).strip().split('.')
    result = []
    for p in parts:
        if re.search(r'\d', p) or len(p) <= 3:
            result.append(p.upper())
        else:
            result.append(p.lower().capitalize())
    return '.'.join(result)


# Ã¢Â”Â€Ã¢Â”Â€ Overall MDM PL list Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€
# Any target whose normalised name matches one of these is bucketed as
# 'Overall_MDM' in the pivot table and pie charts.
_OVERALL_MDM_TARGETS = {
    _norm_target(t) for t in [
        'Olympic.LE', 'Olympic.MN',
        'MDM9205.TX',
        'SDX12.LE',
        'SA515M.LE', 'SA410M.LE', 'SA415M.LE',
        'SA2150P_SA415M.LE_LE', 'SA2150P_SA515M.LE_LE',
        'Waipio_Olympic.LA_MN',
        'MDM9607.LE',
    ]
}
_OVERALL_MDM_LABEL = 'Overall_MDM'


def _resolve_target(raw: str) -> str:
    """Return _OVERALL_MDM_LABEL if target is in the MDM list, else normalised name."""
    n = _norm_target(raw)
    return _OVERALL_MDM_LABEL if n in _OVERALL_MDM_TARGETS else n


def _build_cr_pie_card(table_rows: list) -> dict:
    """
    CR Pie card Ã¢Â€Â” ALL cr_mapped rows (no dedup).
    Pivot: CR Area x Target  (count of every instance/row).
    Per-target pie: slice = CR Area, value = instance count.
    """
    from collections import defaultdict
    cr_mapped = [
        r for r in table_rows
        if str(r.get('jira_category') or '').strip().lower() == 'cr mapped'
    ]

    areas   = sorted({str(r.get('cr_area')  or 'Unknown').strip() for r in cr_mapped})
    targets = sorted({_resolve_target(str(r.get('target') or 'Unknown')) for r in cr_mapped})

    # area x target matrix Ã¢Â€Â” count every row (not unique CRs)
    matrix       = defaultdict(lambda: defaultdict(int))
    area_total   = defaultdict(int)
    target_total = defaultdict(int)
    for r in cr_mapped:
        area   = str(r.get('cr_area')  or 'Unknown').strip()
        target = _resolve_target(str(r.get('target') or 'Unknown'))
        matrix[area][target] += 1
        area_total[area]     += 1
        target_total[target] += 1

    # sort areas by total desc
    areas = sorted(areas, key=lambda a: -area_total[a])

    # per-target pie  [{target, total, pie:[{name,y}]}]
    per_target_pies = []
    for tgt in sorted(targets, key=lambda t: -target_total[t]):
        pie = [
            {'name': area, 'y': matrix[area][tgt]}
            for area in areas if matrix[area][tgt] > 0
        ]
        pie.sort(key=lambda x: -x['y'])
        per_target_pies.append({
            'target': tgt,
            'total':  target_total[tgt],
            'pie':    pie,
        })

    per_target_pies = [
        p for p in per_target_pies if p.get('pie')
    ]

    overall_pie = [{'name': a, 'y': area_total[a]} for a in areas if area_total[a] > 0]

    return {
        'pie_areas':         areas,
        'pie_targets':       targets,
        'pie_matrix':        {a: dict(matrix[a]) for a in areas},
        'pie_area_totals':   dict(area_total),
        'pie_target_totals': dict(target_total),
        'per_target_pies':   per_target_pies,
        'overall_pie':       overall_pie,
        'pie_unique_count':  len(cr_mapped),
    }


# Ã¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•Â

_CR_STATUSES = [
    'Analysis', 'Ready', 'InProgress', 'Open', 'Fix',
    'Built', 'Withdrawn', 'Postponed', 'CannotDuplicate',
    'NotApplicable', 'Closed'
]


def _build_cr_age_card(table_rows: list, sel_start: date, sel_end: date) -> dict:
    """
    Matches VBA macro logic exactly:
    1. Filter CR Mapped rows
    2. Build CR_Count from full set (before dedup)
    3. Deduplicate by CR (cr_current_ticket) Ã¢Â†Â’ unique CR table
    4. Compute New/Old on unique table using year*54+weeknum
    5. Build pivot from unique table
    6. Pie charts from unique table where CR Age > 15
    """
    def _week_num(d):
        d = _safe_date(d)
        if not d:
            return None
        return d.year * 54 + int(d.strftime('%W'))

    # step 1: CR Mapped only
    cr_mapped = [
        r for r in table_rows
        if str(r.get('jira_category') or '').strip().lower() == 'cr mapped'
    ]

    # step 2: CR_Count = occurrences of each CR in full cr_mapped set
    cr_count_map = Counter(
        str(r.get('cr_current_ticket') or '').strip()
        for r in cr_mapped
        if r.get('cr_current_ticket')
    )

    # step 3: deduplicate by cr_current_ticket (keep first occurrence)
    seen = set()
    unique_rows = []
    for r in cr_mapped:
        cr_key = str(r.get('cr_current_ticket') or '').strip()
        if cr_key and cr_key not in seen:
            seen.add(cr_key)
            unique_rows.append(r)
        elif not cr_key:
            unique_rows.append(r)  # keep rows with no CR key

    # step 4: compute New/Old + CR_Count on unique rows
    for r in unique_rows:
        jira_wk = _week_num(r.get('jira_date'))
        cr_wk   = _week_num(r.get('cr_date'))
        if jira_wk and cr_wk:
            r['is_new'] = 1 if cr_wk >= jira_wk else 0
            r['is_old'] = 0 if cr_wk >= jira_wk else 1
        else:
            r['is_new'] = 0
            r['is_old'] = 0
        r['cr_count'] = cr_count_map.get(
            str(r.get('cr_current_ticket') or '').strip(), 1
        )

    # step 5: pivot by Target x CR Status
    from collections import defaultdict
    pivot = defaultdict(lambda: {
        'new_crs': 0, 'old_crs': 0,
        'status_counts': defaultdict(int), 'total': 0
    })
    for r in unique_rows:
        tgt    = _resolve_target(str(r.get('target') or 'Unknown'))
        status = str(r.get('cr_status') or 'Unknown').strip()
        pivot[tgt]['total']                 += 1
        pivot[tgt]['status_counts'][status] += 1
        pivot[tgt]['new_crs']               += r['is_new']
        pivot[tgt]['old_crs']               += r['is_old']

    pivot_rows = [
        {
            'target':        tgt,
            'new_crs':       v['new_crs'],
            'old_crs':       v['old_crs'],
            'status_counts': dict(v['status_counts']),
            'total':         v['total'],
        }
        for tgt, v in sorted(pivot.items(), key=lambda x: -x[1]['total'])
    ]

    # step 6: pie charts from unique rows where CR Age > 15
    age15 = [r for r in unique_rows
             if _safe_int(r.get('cr_age') or 0) and _safe_int(r.get('cr_age') or 0) > 15]
    pie_status_age15 = Counter(str(r.get('cr_status') or 'Unknown').strip() for r in age15)
    pie_area_age15   = Counter(str(r.get('cr_area')   or 'Unknown').strip() for r in age15)

    pie_target_age15 = Counter(_resolve_target(str(r.get('target') or 'Unknown')) for r in age15)

    return {
        'cr_statuses':           _CR_STATUSES,
        'pivot_rows':            pivot_rows,
        'cr_mapped_rows':        unique_rows,
        'cr_mapped_count':       len(unique_rows),
        'status_totals':         {st: sum(r['status_counts'].get(st, 0) for r in pivot_rows)
                                  for st in _CR_STATUSES},
        'pie_status_data_age14': [{'name': k, 'y': v} for k, v in
                                  sorted(pie_status_age15.items(), key=lambda x: -x[1])],
        'pie_area_data_age14':   [{'name': k, 'y': v} for k, v in
                                  sorted(pie_area_age15.items(),   key=lambda x: -x[1])],
        'pie_target_data_age14': [{'name': k, 'y': v} for k, v in
                                  sorted(pie_target_age15.items(), key=lambda x: -x[1])],
    }


# Ã¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•Â
# UNIQUE CR REPORT  helpers
# Ã¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•ÂÃ¢Â•Â

import re as _re

# Ã¢Â”Â€Ã¢Â”Â€ filename date parsing Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€
# Pattern: UNIQUECRSREPORT_WEEKENDING_2026Y_05M_17D_..._Unique_CRs-...
_UCR_FNAME_RE = _re.compile(
    r'UNIQUECRSREPORT_WEEKENDING_'
    r'(\d{4})Y_(\d{2})M_(\d{2})D',
    _re.IGNORECASE
)

def _ucr_file_week_end(fname: str):
    """Return (year, month, day) from a Unique-CR filename, or None."""
    m = _UCR_FNAME_RE.search(fname)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3))


def _ucr_week_end_date(fname: str):
    """Return a date object for the week-ending encoded in the filename."""
    t = _ucr_file_week_end(fname)
    if not t:
        return None
    try:
        return date(t[0], t[1], t[2])
    except Exception:
        return None


# Ã¢Â”Â€Ã¢Â”Â€ UCR_MIN_DATE: only show files from May 2026 onwards Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€
_UCR_MIN_DATE = date(2026, 5, 1)


def _list_ucr_files(force: bool = False) -> list:
    """List Unique CR RawData CSV/Excel files, newest week-ending first.

    force=True bypasses the short cache so a regenerated same-week source file
    is detected immediately and the newest modified file wins.
    """
    import time
    now = time.time()
    if (not force) and now - float(_UCR_RAW_FILES_CACHE.get('ts') or 0) < _SHARE_LIST_TTL_SECONDS:
        return [dict(x) for x in (_UCR_RAW_FILES_CACHE.get('value') or [])]

    result = []
    try:
        if not os.path.isdir(_UNIQUE_CR_RAW_DIR):
            return result
        if force:
            file_iter = ((root, fname) for root, _dirs, files in os.walk(_UNIQUE_CR_RAW_DIR) for fname in files)
        else:
            file_iter = ((_UNIQUE_CR_RAW_DIR, fname) for fname in os.listdir(_UNIQUE_CR_RAW_DIR))
        for root, fname in file_iter:
            if fname.startswith('~$'):
                continue
            lower = fname.lower()
            # Only pick the actual Unique CRs data file.
            # Exclude log/error/blacklist/stdoutput/crinfo files.
            if lower.endswith('.txt'):
                continue
            if any(x in lower for x in ('stdoutput', 'errorfile', 'black_list', 'crinfo', 'cr_tat_jira')):
                continue
            is_data_csv = 'uniquecrsreport_weekending_' in lower and ('unique_crs-' in lower or 'unique_crs_' in lower)
            is_source_xl = lower.endswith(('.xlsx', '.xlsm')) and 'uniquecrsreport_weekending_' in lower
            if not (is_data_csv or is_source_xl):
                continue
            we = _ucr_week_end_date(fname)
            if not we or we < _UCR_MIN_DATE:
                continue
            fpath = os.path.join(root, fname)
            if not os.path.isfile(fpath):
                continue
            try:
                mtime = os.path.getmtime(fpath)
            except Exception:
                mtime = 0
            ws_d = we - timedelta(days=6)
            iso_wk = we.isocalendar()[1]
            result.append({
                'path': fpath,
                'filename': fname,
                'week_end_date': we,
                'week': iso_wk,
                'mtime': mtime,
                'label': f"{ws_d.strftime('%b %d')} Ã¢Â€Â“ {we.strftime('%b %d, %Y')} (Wk {iso_wk}) {Path(fname).suffix.upper().lstrip('.') or 'FILE'}",
            })
    except Exception:
        return [dict(x) for x in (_UCR_RAW_FILES_CACHE.get('value') or [])]
    result.sort(key=lambda x: (x['week_end_date'], x.get('mtime') or 0), reverse=True)
    _UCR_RAW_FILES_CACHE.update({'ts': now, 'value': result})
    return [dict(x) for x in result]


def _find_ucr_file_by_week_end(week_end_str: str, force: bool = False) -> str:
    """Return newest RawData Unique_CRs data CSV/Excel for the requested week-ending date.
    Log/error/blacklist/crinfo files are excluded by _list_ucr_files().
    """
    try:
        target_we = date.fromisoformat(str(week_end_str)[:10])
    except Exception:
        return ''
    # Some generated source filenames encode the Monday after the displayed
    # weekly range (e.g. UI Jun 15-Jun 21, file WEEKENDING_2026Y_06M_22D).
    # Treat target week-end and target+1 day as the same report week, then pick
    # the newest regenerated file by modified time.
    valid_dates = {target_we, target_we + timedelta(days=1)}
    candidates = [e for e in _list_ucr_files(force=force) if e.get('week_end_date') in valid_dates]
    if not candidates:
        return ''
    candidates.sort(key=lambda x: x.get('mtime') or 0, reverse=True)
    return candidates[0]['path']


def _ensure_ucr_excel_for_week(week_end_date: date, farm_map: dict | None = None, force_refresh: bool = False) -> dict:
    """
    Ensure the generated Unique CR Excel exists for the selected week.

    If Excel already exists, do not read CSV unless force_refresh=True. When
    force_refresh=True, re-read the matching UNIQUECRSREPORT_WEEKENDING RawData
    CSV for that same week and overwrite/update the generated Excel results.
    """
    info = {
        'attempted': False,
        'success': False,
        'message': '',
        'path': _ucr_excel_path(week_end_date) if week_end_date else '',
    }
    if not week_end_date:
        info['message'] = 'No week selected'
        return info

    xl_path = _ucr_excel_path(week_end_date)
    if os.path.isfile(xl_path) and not force_refresh:
        info.update({'success': True, 'message': 'Excel exists', 'path': xl_path})
        return info

    src_path = _find_ucr_file_by_week_end(week_end_date.isoformat(), force=force_refresh)
    if not src_path or not os.path.isfile(src_path):
        info['message'] = 'No Excel workbook or matching RawData CSV/Excel found for selected week'
        return info

    info['attempted'] = True
    try:
        farm_map = farm_map if farm_map is not None else _load_farm_station_map()
        parsed = _parse_ucr_source_file(src_path)
        if not any(parsed.get('rows', {}).values()):
            info['message'] = 'RawData source parsed with no rows'
            return info
        out_path = _generate_ucr_excel(parsed, farm_map, week_end_date)
        info.update({
            'success': True,
            'message': f"Refreshed results from CSV/source file: {os.path.basename(src_path)}",
            'path': out_path,
        })
        return info
    except Exception as exc:
        info['message'] = str(exc)
        return info


# Ã¢Â”Â€Ã¢Â”Â€ Farm KPI station map Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€

def _load_farm_station_map() -> dict:
    """Read Farm_KPI station map with a short cache to avoid repeated UNC IO."""
    import time
    now = time.time()
    if now - float(_FARM_STATION_MAP_CACHE.get('ts') or 0) < _FARM_MAP_TTL_SECONDS:
        return dict(_FARM_STATION_MAP_CACHE.get('value') or {})

    result = {}
    try:
        if not os.path.isdir(_FARM_KPI_DIR):
            return result
        for fname in os.listdir(_FARM_KPI_DIR):
            if not fname.lower().endswith('.txt'):
                continue
            farm_name = os.path.splitext(fname)[0]   # e.g. 'CA', 'Mobility'
            fpath = os.path.join(_FARM_KPI_DIR, fname)
            try:
                with open(fpath, 'r', encoding='utf-8', errors='ignore') as fh:
                    for line in fh:
                        station = line.strip().upper()
                        if station:
                            result[station] = farm_name
            except Exception:
                pass
    except Exception:
        return dict(_FARM_STATION_MAP_CACHE.get('value') or {})
    _FARM_STATION_MAP_CACHE.update({'ts': now, 'value': result})
    return dict(result)


# Ã¢Â”Â€Ã¢Â”Â€ CSV parser Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€

def _ucr_text_key(value) -> str:
    """Case-insensitive sort/group key for Unique CR rows."""
    return str(value or '').strip().casefold()


def _is_ucr_separator_value(value) -> bool:
    """True for CSV/Excel visual separator cells like ########."""
    s = str(value or '').strip()
    if not s:
        return False
    compact = s.replace(' ', '')
    return len(compact) >= 3 and set(compact) <= {'#'}


def _is_ucr_non_data_marker(value) -> bool:
    s = str(value or '').strip().upper()
    if not s:
        return False
    if s in {'QIPL CRS', 'SD CRS', 'CH CRS', 'QIPL', 'SD', 'CH'} or _is_ucr_separator_value(s):
        return True
    # Source workbooks sometimes contain section header rows in the Target
    # column, for example "FF CRs".  They are not real Target values and should
    # never be displayed/saved as Target+PL-ID records.
    return bool(_re.fullmatch(r'[A-Z0-9 ._/-]{1,40}\s+CRS?', s))


def _is_ucr_data_row(row: dict) -> bool:
    """Reject section title/separator rows from Unique CR CSV/Excel."""
    if not row:
        return False
    vals = [str(v or '').strip() for v in row.values() if str(v or '').strip()]
    if not vals:
        return False
    # Rows such as ###########, QIPL CRs, SD CRs, CH CRs are visual separators.
    if all(_is_ucr_non_data_marker(v) for v in vals):
        return False
    for key in ('Target', 'PL-ID', 'CRID'):
        if _is_ucr_separator_value(row.get(key)):
            return False
    # Reject visual/section headers such as "FF CRs" even if other defaulted
    # columns like CR Count/Farm are populated by parsing fallbacks.
    if _is_ucr_non_data_marker(row.get('Target')):
        return False
    return True


def _sort_ucr_site_rows(rows: list) -> list:
    """
    Sort Unique CR rows so identical Target/PL-ID records are adjacent.

    Both the web page and Excel writer merge only consecutive equal Target/PL-ID
    cells, so grouping by Target first is required before rowspan/merge works.
    """
    return sorted(rows or [], key=lambda r: (
        _ucr_text_key(r.get('Target')),
        _ucr_text_key(r.get('PL-ID')),
        _ucr_text_key(r.get('CR Area')),
        _ucr_text_key(r.get('CR Subsystem')),
        _ucr_text_key(r.get('CR Functionality')),
        _ucr_text_key(r.get('CRID')),
        _ucr_text_key(r.get('Station')),
    ))


def _sort_ucr_rows_by_site(parsed: dict) -> dict:
    """Sort every site bucket in a parsed Unique CR payload in place."""
    if isinstance(parsed, dict):
        rows_by_site = parsed.get('rows') or {}
        for site, rows in list(rows_by_site.items()):
            rows_by_site[site] = _sort_ucr_site_rows(rows)
    return parsed


def _empty_ucr_parse_result(filepath: str, source: str) -> dict:
    result = {
        'headers': ['Target', 'PL-ID', 'CRID', 'CR Count', 'CR Date',
                    'CR Area', 'CR Subsystem', 'CR Functionality', 'CR SI',
                    'CR Status', 'Device ID', 'Station', 'Farm',
                    'Test/Setup Details', 'Labels'],
        'sites': [],
        'rows': {'QIPL': [], 'SD': [], 'CH': []},
        'file_name': os.path.basename(filepath or ''),
        'week_end_date': '',
        'source': source,
    }
    we = _ucr_week_end_date(result['file_name'])
    if we:
        result['week_end_date'] = we.isoformat()
    return result


def _parse_ucr_csv(filepath: str) -> dict:
    """Parse a Unique CR RawData CSV into QIPL/SD/CH site buckets.

    This is used only by Publish Report to generate the official Excel workbook.
    The UI cards continue to read the generated Excel workbook afterward.
    """
    result = _empty_ucr_parse_result(filepath, 'csv')

    if not filepath or not os.path.isfile(filepath):
        return result

    farm_map = _load_farm_station_map()

    def _norm_header(v):
        return _norm(v).replace('__', '_')

    def _pick(row, *names):
        wanted = {_norm_header(n) for n in names}
        for k, v in row.items():
            if _norm_header(k) in wanted:
                return str(v).strip() if v is not None else ''
        return ''

    def _site(row):
        raw = _pick(row, 'Site', 'Location', 'Lab', 'Source', 'Owner', 'Team')
        txt = str(raw or '').strip().upper()
        if 'QIPL' in txt or txt in ('QP', 'QCT'):
            return 'QIPL'
        if txt in ('SD', 'SAN DIEGO', 'SANDIEGO') or 'SAN DIEGO' in txt:
            return 'SD'
        if txt in ('CH', 'CHINA') or 'CHINA' in txt or 'SHANGHAI' in txt:
            return 'CH'
        labels = _pick(row, 'Labels', 'Label')
        up = str(labels or '').upper()
        if 'QIPL' in up:
            return 'QIPL'
        if 'SD' in up or 'SAN DIEGO' in up:
            return 'SD'
        if 'CH' in up or 'CHINA' in up:
            return 'CH'
        return 'QIPL'

    def _row_to_web(row):
        station = _pick(row, 'Station', 'HostPC', 'Host PC', 'Host', 'Machine', 'Station Name')
        farm = _pick(row, 'Farm', 'FarmType', 'Farm Type') or farm_map.get(str(station).strip().upper(), 'PDT')
        return {
            'Target': _pick(row, 'Target', 'Build Target', 'Chipset', 'Target Name'),
            'PL-ID': _pick(row, 'PL-ID', 'PL ID', 'PLID', 'Target PL', 'Target.PL', 'TargetPL'),
            'CRID': _pick(row, 'CRID', 'CR ID', 'CR', 'CR Number', 'CR#', 'CR Current Ticket'),
            'CR Count': _pick(row, 'CR Count', 'CR Instances', 'Occurrence', 'Occurrences', 'Count') or '1',
            'CR Date': _pick(row, 'CR Date', 'CR Created Date', 'Created Date', 'Creation Date'),
            'CR Area': _pick(row, 'CR Area', 'Area', 'Tech Area'),
            'CR Subsystem': _pick(row, 'CR Subsystem', 'CR SubSystem', 'Subsystem', 'SubSystem'),
            'CR Functionality': _pick(row, 'CR Functionality', 'Functionality', 'CR Function'),
            'CR SI': _pick(row, 'CR SI', 'SI'),
            'CR Status': _pick(row, 'CR Status', 'Status'),
            'Device ID': _pick(row, 'Device ID', 'Device', 'DeviceID'),
            'Station': station,
            'Farm': farm,
            'Test/Setup Details': _pick(row, 'Test/Setup Details', 'Setup', 'Test Setup', 'Setup Details', 'Test Details'),
            'Labels': _pick(row, 'Labels', 'Label'),
        }

    for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(filepath, newline='', encoding=enc, errors='strict') as fh:
                sample = fh.read(8192)
                fh.seek(0)
                try:
                    dialect = _csv_mod.Sniffer().sniff(sample, delimiters=',\t|;')
                except Exception:
                    dialect = _csv_mod.excel
                reader = _csv_mod.DictReader(fh, dialect=dialect)
                for raw in reader:
                    if not raw or not any(str(v or '').strip() for v in raw.values()):
                        continue
                    if all(_is_ucr_non_data_marker(v) for v in raw.values() if str(v or '').strip()):
                        continue
                    web_row = _row_to_web(raw)
                    if not _is_ucr_data_row(web_row):
                        continue
                    site = _site(raw)
                    if site not in result['rows']:
                        result['rows'][site] = []
                    result['rows'][site].append(web_row)
            break
        except UnicodeDecodeError:
            continue
        except Exception:
            break

    result['rows'] = {k: v for k, v in result['rows'].items() if v}
    _sort_ucr_rows_by_site(result)
    result['sites'] = [s for s in ('QIPL', 'SD', 'CH') if result['rows'].get(s)]
    return result


def _parse_ucr_excel_source(filepath: str) -> dict:
    """Parse UNIQUECRSREPORT_WEEKENDING_*.xlsx source workbooks.

    Raw Unique CR reports already contain Target and PL-ID columns.  This parser
    reads those columns directly from the source workbook so the generated saved
    report and the Sharepoint consolidated table can use the same Target+PL-ID
    data.
    """
    result = _empty_ucr_parse_result(filepath, 'excel_source')
    if not filepath or not os.path.isfile(filepath):
        return result

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        return result

    farm_map = _load_farm_station_map()

    def _norm_header(v):
        return _norm(v).replace('__', '_')

    def _pick(row, *names):
        wanted = {_norm_header(n) for n in names}
        for k, v in (row or {}).items():
            if _norm_header(k) in wanted:
                return str(v).strip() if v is not None else ''
        return ''

    def _site(row, sheet_name):
        raw = _pick(row, 'Site', 'Location', 'Lab', 'Source', 'Owner', 'Team') or sheet_name
        txt = str(raw or '').strip().upper()
        if 'QIPL' in txt or txt in ('QP', 'QCT'):
            return 'QIPL'
        if txt in ('SD', 'SAN DIEGO', 'SANDIEGO') or 'SAN DIEGO' in txt:
            return 'SD'
        if txt in ('CH', 'CHINA') or 'CHINA' in txt or 'SHANGHAI' in txt:
            return 'CH'
        labels = _pick(row, 'Labels', 'Label')
        up = str(labels or '').upper()
        if 'QIPL' in up:
            return 'QIPL'
        if 'SD' in up or 'SAN DIEGO' in up:
            return 'SD'
        if 'CH' in up or 'CHINA' in up:
            return 'CH'
        return 'QIPL'

    def _row_to_web(row):
        station = _pick(row, 'Station', 'HostPC', 'Host PC', 'Host', 'Machine', 'Station Name')
        farm = _pick(row, 'Farm', 'FarmType', 'Farm Type') or farm_map.get(str(station).strip().upper(), 'PDT')
        return {
            'Target': _pick(row, 'Target', 'Build Target', 'Chipset', 'Target Name'),
            'PL-ID': _pick(row, 'PL-ID', 'PL ID', 'PLID', 'Target PL', 'Target.PL', 'TargetPL'),
            'CRID': _pick(row, 'CRID', 'CR ID', 'CR', 'CR Number', 'CR#', 'CR Current Ticket'),
            'CR Count': _pick(row, 'CR Count', 'CR Instances', 'Occurrence', 'Occurrences', 'Count') or '1',
            'CR Date': _pick(row, 'CR Date', 'CR Created Date', 'Created Date', 'Creation Date'),
            'CR Area': _pick(row, 'CR Area', 'Area', 'Tech Area'),
            'CR Subsystem': _pick(row, 'CR Subsystem', 'CR SubSystem', 'Subsystem', 'SubSystem'),
            'CR Functionality': _pick(row, 'CR Functionality', 'Functionality', 'CR Function'),
            'CR SI': _pick(row, 'CR SI', 'SI'),
            'CR Status': _pick(row, 'CR Status', 'Status'),
            'Device ID': _pick(row, 'Device ID', 'Device', 'DeviceID'),
            'Station': station,
            'Farm': farm,
            'Test/Setup Details': _pick(row, 'Test/Setup Details', 'Setup', 'Test Setup', 'Setup Details', 'Test Details'),
            'Labels': _pick(row, 'Labels', 'Label'),
        }

    try:
        for sname in wb.sheetnames:
            ws = wb[sname]
            if ws.max_row < 2:
                continue
            header_row_num = None
            headers = []
            for ri, vals in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True), start=1):
                cand = [str(v or '').strip() for v in vals]
                normed = {_norm_header(h) for h in cand if h}
                if 'target' in normed and ('pl_id' in normed or 'plid' in normed or 'target_pl' in normed or 'target.pl' in normed):
                    header_row_num = ri
                    headers = cand
                    break
            if not header_row_num:
                continue
            for vals in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
                raw = {}
                for ci, h in enumerate(headers):
                    if not h:
                        continue
                    v = vals[ci] if ci < len(vals) else None
                    if isinstance(v, datetime):
                        v = v.date().isoformat()
                    elif v is not None:
                        v = str(v).strip()
                    raw[h] = v or ''
                if not any(str(v or '').strip() for v in raw.values()):
                    continue
                web_row = _row_to_web(raw)
                if not _is_ucr_data_row(web_row):
                    continue
                site = _site(raw, sname)
                if site not in result['rows']:
                    result['rows'][site] = []
                result['rows'][site].append(web_row)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    result['rows'] = {k: v for k, v in result['rows'].items() if v}
    _sort_ucr_rows_by_site(result)
    result['sites'] = [s for s in ('QIPL', 'SD', 'CH') if result['rows'].get(s)]
    return result


def _parse_ucr_source_file(filepath: str) -> dict:
    """Parse Unique CR source as CSV or Excel, based on extension."""
    ext = Path(filepath or '').suffix.lower()
    if ext in ('.xlsx', '.xlsm'):
        return _parse_ucr_excel_source(filepath)
    return _parse_ucr_csv(filepath)


# Ã¢Â”Â€Ã¢Â”Â€ Excel output helpers Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€

# Column palettes (openpyxl-compatible hex, no leading #)
_XL_HDR_FILL   = 'FF1F3864'   # dark navy
_XL_HDR_FONT   = 'FFFFFFFF'   # white
_XL_ALT_FILL   = 'FFD9E1F2'   # light blue
_XL_TOTAL_FILL = 'FF2E4057'   # dark slate
_XL_TOTAL_FONT = 'FFFFFFFF'
_XL_PDT_FILL   = 'FFFFF2CC'   # light yellow  (PDT default farm)
_XL_FARM_COLORS = {
    'CA':       'FFD6E4BC',
    'Mobility': 'FFFCE4D6',
    'Thermal':  'FFE2EFDA',
    'Vo-Wifi':  'FFDAE3F3',
    'WCNSS':    'FFEDE7F6',
    'PDT':      'FFFFF2CC',
}


def _xl_hdr(ws, row, col, value, bold=True, bg=_XL_HDR_FILL, fg=_XL_HDR_FONT,
            wrap=False, align='center'):
    """Write a styled header cell."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill    = PatternFill('solid', fgColor=bg)
    cell.font    = Font(bold=bold, color=fg, size=9)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                               wrap_text=wrap)
    thin = Side(style='thin', color='FF8EA9C1')
    cell.border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    return cell


def _xl_data(ws, row, col, value, bold=False, bg=None, fg='FF000000',
             align='left', wrap=False, num_fmt=None):
    """Write a styled data cell."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    cell.font  = Font(bold=bold, color=fg, size=9)
    cell.alignment = Alignment(horizontal=align, vertical='center',
                               wrap_text=wrap)
    thin = Side(style='thin', color='FFD0D7E0')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _xl_autofit(ws, min_w=8, max_w=40):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        length = max(
            (len(str(c.value)) if c.value is not None else 0)
            for c in col_cells
        )
        ws.column_dimensions[
            col_cells[0].column_letter
        ].width = max(min_w, min(length + 2, max_w))


def _ucr_excel_path(week_end_date: date) -> str:
    """
    Return the expected Excel path for a given week-end date.
    Pattern: WeeklyUniqueCRs/{YYYY}/Unique_CRs_{YYYY}_Week_{WW}.xlsx
    """
    year    = week_end_date.year
    iso_wk  = week_end_date.isocalendar()[1]
    folder  = os.path.join(_UNIQUE_CR_EXCEL_BASE, str(year))
    fname   = f'Unique_CRs_{year}_Week_{iso_wk}.xlsx'
    return os.path.join(folder, fname)


def _list_ucr_excel_files() -> list:
    """List generated Unique CR Excel files, cached to avoid repeated UNC scans."""
    import time
    now = time.time()
    if now - float(_UCR_EXCEL_FILES_CACHE.get('ts') or 0) < _SHARE_LIST_TTL_SECONDS:
        return [dict(x) for x in (_UCR_EXCEL_FILES_CACHE.get('value') or [])]

    result = []
    try:
        base = _UNIQUE_CR_EXCEL_BASE
        if not os.path.isdir(base):
            return result
        pat = _re.compile(r'Unique_CRs_(\d{4})_Week_(\d+)\.xlsx$', _re.IGNORECASE)
        for yr_dir in os.listdir(base):
            try:
                yr = int(yr_dir)
            except ValueError:
                continue
            if yr < 2026:
                continue
            yr_path = os.path.join(base, yr_dir)
            if not os.path.isdir(yr_path):
                continue
            for fname in os.listdir(yr_path):
                m = pat.match(fname)
                if not m:
                    continue
                year = int(m.group(1))
                week = int(m.group(2))
                try:
                    we = date.fromisocalendar(year, week, 7)
                except Exception:
                    continue
                ws_d = we - timedelta(days=6)
                result.append({
                    'path':          os.path.join(yr_path, fname),
                    'filename':      fname,
                    'year':          year,
                    'week':          week,
                    'week_end_date': we,
                    'label':         f"{ws_d.strftime('%b %d')} \u2013 {we.strftime('%b %d, %Y')} (Wk {week})",
                })
    except Exception:
        return [dict(x) for x in (_UCR_EXCEL_FILES_CACHE.get('value') or [])]
    result.sort(key=lambda x: x['week_end_date'], reverse=True)
    _UCR_EXCEL_FILES_CACHE.update({'ts': now, 'value': result})
    return [dict(x) for x in result]


def _find_ucr_excel_for_week(week_end_date: date):
    """
    Return the dict entry from _list_ucr_excel_files() whose week_end_date
    matches exactly, or None.
    """
    for entry in _list_ucr_excel_files():
        if entry['week_end_date'] == week_end_date:
            return entry
    return None


def _ucr_excel_site_counts(week_end_date: date) -> tuple[dict, bool, str]:
    """
    Return QIPL/SD/CH row counts from the selected week's Excel workbook.

    Fast landing-page path:
    - direct workbook path, no directory listing
    - in-process cache keyed by path + file size + mtime
    - uses sheet dimensions instead of iterating every populated cell

    If the workbook for the selected week is missing, values are blank.
    """
    counts = {'QIPL': '', 'SD': '', 'CH': ''}
    if not week_end_date:
        return counts, False, ''

    xl_path = _ucr_excel_path(week_end_date)
    try:
        if not os.path.isfile(xl_path):
            return counts, False, xl_path
        stat = os.stat(xl_path)
    except Exception:
        return counts, False, xl_path

    cache_key = (xl_path, int(stat.st_size), int(stat.st_mtime))
    cached = _UCR_LANDING_COUNTS_CACHE.get(cache_key)
    if cached is not None:
        return dict(cached), True, xl_path

    wb = None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)
        for site in ('QIPL', 'SD', 'CH'):
            if site not in wb.sheetnames:
                counts[site] = 0
                continue
            ws = wb[site]
            # The generated sheets have one header row and contiguous data rows.
            # max_row is much faster than scanning every cell over the network.
            counts[site] = max(int(ws.max_row or 0) - 1, 0)


        _UCR_LANDING_COUNTS_CACHE.clear()
        _UCR_LANDING_COUNTS_CACHE[cache_key] = dict(counts)
        return counts, True, xl_path
    except Exception:
        return {'QIPL': '', 'SD': '', 'CH': ''}, False, xl_path
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _read_ucr_pdtsite_chart_data(path: str) -> dict:
    """Read the small PDTSite summary sheet for the web bar chart.

    If an older workbook has no PDTSite sheet, fall back to scanning only the
    CR Area column from QIPL/CH/SD sheets, not the full row payload.
    """
    chart = {'categories': [], 'series': [], 'totals': {}}
    if not path or not os.path.isfile(path):
        return chart

    wb = None
    try:
        import openpyxl
        from collections import Counter as _Counter
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        site_order = ['QIPL', 'CH', 'SD']
        site_counts = {s: _Counter() for s in site_order}

        if 'PDTSite' in wb.sheetnames:
            ws = wb['PDTSite']
            headers = [str(v or '').strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
            idx = {h: i for i, h in enumerate(headers)}
            area_i = idx.get('CR Area')
            q_i = idx.get('QIPL CRs')
            ch_i = idx.get('CH CRs')
            sd_i = idx.get('SD CRs')
            for vals in ws.iter_rows(min_row=2, values_only=True):
                area = str(vals[area_i] if area_i is not None and area_i < len(vals) else '').strip()
                if not area or area.lower() == 'grand total':
                    continue
                site_counts['QIPL'][area] = int(vals[q_i] or 0) if q_i is not None and q_i < len(vals) else 0
                site_counts['CH'][area] = int(vals[ch_i] or 0) if ch_i is not None and ch_i < len(vals) else 0
                site_counts['SD'][area] = int(vals[sd_i] or 0) if sd_i is not None and sd_i < len(vals) else 0
        else:
            for site in site_order:
                if site not in wb.sheetnames:
                    continue
                ws = wb[site]
                headers = [str(v or '').strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
                area_i = None
                for i, h in enumerate(headers):
                    if h in ('CR Area', 'Area'):
                        area_i = i
                        break
                if area_i is None:
                    continue
                for vals in ws.iter_rows(min_row=2, values_only=True):
                    if area_i < len(vals):
                        area = str(vals[area_i] or '').strip() or 'Unknown'
                        site_counts[site][area] += 1

        area_totals = _Counter()
        for site in site_order:
            chart['totals'][site] = sum(site_counts[site].values())
            area_totals.update(site_counts[site])
        areas = [a for a, _ in sorted(area_totals.items(), key=lambda x: (-x[1], x[0]))]
        chart['categories'] = areas
        chart['series'] = [
            {'name': site, 'data': [site_counts[site].get(area, 0) for area in areas]}
            for site in site_order
        ]
        return chart
    except Exception:
        return chart
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _build_ucr_pdtsite_sheet(wb, parsed: dict):
    """Create/replace PDTSite sheet with the site-wise CR Area distribution chart."""
    from collections import Counter as _Counter
    from openpyxl.chart import BarChart, Reference

    if 'PDTSite' in wb.sheetnames:
        wb.remove(wb['PDTSite'])

    ws = wb.create_sheet('PDTSite', 0)
    ws.freeze_panes = 'A2'

    site_area_counts = {site: _Counter() for site in ('QIPL', 'CH', 'SD')}
    area_totals = _Counter()
    rows_by_site = (parsed or {}).get('rows') or {}
    for site in ('QIPL', 'CH', 'SD'):
        for row in rows_by_site.get(site, []) or []:
            area = str(row.get('CR Area') or 'Unknown').strip() or 'Unknown'
            site_area_counts[site][area] += 1
            area_totals[area] += 1

    areas = [area for area, _ in sorted(area_totals.items(), key=lambda x: (-x[1], x[0]))]

    headers = ['CR Area', 'QIPL CRs', 'CH CRs', 'SD CRs', 'Total']
    for ci, header in enumerate(headers, start=1):
        _xl_hdr(ws, 1, ci, header, bg='FF4472C4')

    for ri, area in enumerate(areas, start=2):
        qipl = site_area_counts['QIPL'].get(area, 0)
        ch = site_area_counts['CH'].get(area, 0)
        sd = site_area_counts['SD'].get(area, 0)
        bg = 'FFD9E1F2' if ri % 2 == 0 else 'FFFFFFFF'
        _xl_data(ws, ri, 1, area, bg=bg)
        _xl_data(ws, ri, 2, qipl, bg=bg, align='center')
        _xl_data(ws, ri, 3, ch, bg=bg, align='center')
        _xl_data(ws, ri, 4, sd, bg=bg, align='center')
        _xl_data(ws, ri, 5, qipl + ch + sd, bg=bg, align='center', bold=True)

    total_row = len(areas) + 2
    _xl_hdr(ws, total_row, 1, 'Grand Total', bg='FF2E4057')
    _xl_hdr(ws, total_row, 2, sum(site_area_counts['QIPL'].values()), bg='FF2E4057')
    _xl_hdr(ws, total_row, 3, sum(site_area_counts['CH'].values()), bg='FF2E4057')
    _xl_hdr(ws, total_row, 4, sum(site_area_counts['SD'].values()), bg='FF2E4057')
    _xl_hdr(ws, total_row, 5, sum(area_totals.values()), bg='FF2E4057')

    if areas:
        chart = BarChart()
        chart.type = 'col'
        chart.style = 10
        chart.title = 'PDT Site wise Unique CRs report'
        chart.y_axis.title = 'Count of CRID'
        chart.x_axis.title = 'CR Area'
        chart.width = 28
        chart.height = 12
        data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=len(areas) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(areas) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend.position = 'b'
        ws.add_chart(chart, 'G2')

    _xl_autofit(ws, min_w=10, max_w=34)


def _ensure_ucr_pdtsite_sheet(path: str, parsed: dict):
    """Persist PDTSite sheet into an existing workbook, best-effort."""
    if not path or not os.path.isfile(path):
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        _build_ucr_pdtsite_sheet(wb, parsed)
        wb.save(path)
        wb.close()
    except Exception:
        pass


def _read_ucr_excel(path: str, farm_map: dict) -> dict:
    """
    Read an existing Unique_CRs_*.xlsx workbook.
    Returns same structure as _parse_ucr_csv:
      {headers, sites, rows:{QIPL:[...], SD:[...], CH:[...]},
       file_name, week_end_date, source:'excel'}
    The Farm column is read from the workbook (user may have edited it).
    """
    import openpyxl
    result = {'headers': [], 'sites': [], 'rows': {}, 'file_name': '',
              'week_end_date': '', 'source': 'excel'}
    if not path or not os.path.isfile(path):
        return result

    result['file_name'] = os.path.basename(path)
    # derive week_end from filename
    pat = _re.compile(r'Unique_CRs_(\d{4})_Week_(\d+)\.xlsx$', _re.IGNORECASE)
    m   = pat.match(result['file_name'])
    if m:
        try:
            we = date.fromisocalendar(int(m.group(1)), int(m.group(2)), 7)
            result['week_end_date'] = we.isoformat()
        except Exception:
            pass

    wb = openpyxl.load_workbook(path, data_only=True)
    # Site sheets: QIPL, SD, CH
    _SITE_SHEETS = {'QIPL': 'QIPL', 'SD': 'SD', 'CH': 'CH'}
    common_headers = ['Target', 'PL-ID', 'CRID', 'CR Count', 'CR Date',
                      'CR Area', 'CR Subsystem', 'CR Functionality',
                      'CR SI', 'CR Status', 'Device ID', 'Station', 'Farm',
                      'Test/Setup Details']
    result['headers'] = common_headers

    for site, sname in _SITE_SHEETS.items():
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        if ws.max_row < 2:
            continue
        # read header row
        hdrs = [str(ws.cell(1, c).value or '').strip()
                for c in range(1, ws.max_column + 1)]
        rows_out = []
        # Excel merged cells store the value only in the first row of the merge.
        # Fill these values down while reading so the web table can compute
        # rowspan groups and filters correctly.
        last_merged_values = {'Target': '', 'PL-ID': ''}
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if not any(v for v in vals if v is not None and str(v).strip()):
                continue
            raw_row = {}
            for ci, h in enumerate(hdrs):
                v = vals[ci] if ci < len(vals) else None
                if isinstance(v, datetime):
                    v = v.date().isoformat()
                elif v is not None:
                    v = str(v).strip()
                raw_row[h] = v or ''

            # Normalize Excel sheet headers to the keys used by the web table.
            # This keeps the Unique CR card Excel-only while supporting the
            # workbook labels produced by the existing manual/generated format.
            def _v(*names):
                for name in names:
                    if raw_row.get(name):
                        return raw_row.get(name)
                return ''

            row_dict = {
                'Target':             _v('Target'),
                'PL-ID':              _v('PL-ID', 'PL ID', 'PLID', 'Target PL', 'Target.PL', 'TargetPL'),
                'CRID':               _v('CRID'),
                'CR Count':           _v('CR Count', 'CR Instances'),
                'CR Date':            _v('CR Date'),
                'CR Area':            _v('CR Area'),
                'CR Subsystem':       _v('CR Subsystem'),
                'CR Functionality':   _v('CR Functionality'),
                'CR SI':              _v('CR SI'),
                'CR Status':          _v('CR Status'),
                'Device ID':          _v('Device ID', 'Device'),
                'Station':            _v('Station'),
                'Farm':               _v('Farm'),
                'Test/Setup Details': _v('Test/Setup Details', 'Setup'),
                'Labels':             _v('Labels'),
            }

            for merge_key in ('Target', 'PL-ID'):
                if row_dict.get(merge_key):
                    last_merged_values[merge_key] = row_dict[merge_key]
                elif last_merged_values.get(merge_key):
                    row_dict[merge_key] = last_merged_values[merge_key]

            if not _is_ucr_data_row(row_dict):
                continue

            # ensure Farm column Ã¢Â€Â” prefer stored value, fallback to farm_map
            if not row_dict.get('Farm'):
                station = str(row_dict.get('Station') or '').strip().upper()
                row_dict['Farm'] = farm_map.get(station, 'PDT')
            rows_out.append(row_dict)
        result['rows'][site] = _sort_ucr_site_rows(rows_out)
        if site not in result['sites']:
            result['sites'].append(site)
    _sort_ucr_rows_by_site(result)
    try:
        wb.close()
    except Exception:
        pass
    return result


def _read_ucr_excel_sheet(path: str, site: str, farm_map: dict) -> dict:
    """Read only one site sheet from a generated Unique CR Excel workbook."""
    import openpyxl
    site = str(site or '').strip().upper()
    if site not in ('QIPL', 'SD', 'CH'):
        return {'headers': [], 'rows': [], 'site': site}
    if not path or not os.path.isfile(path):
        return {'headers': [], 'rows': [], 'site': site}

    wb = None
    rows_out = []
    common_headers = ['Target', 'PL-ID', 'CRID', 'CR Count', 'CR Date',
                      'CR Area', 'CR Subsystem', 'CR Functionality',
                      'CR SI', 'CR Status', 'Device ID', 'Station', 'Farm',
                      'Test/Setup Details']
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if site not in wb.sheetnames:
            return {'headers': common_headers, 'rows': [], 'site': site}
        ws = wb[site]
        if ws.max_row < 2:
            return {'headers': common_headers, 'rows': [], 'site': site}

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        hdrs = [str(v or '').strip() for v in header_row]
        last_merged_values = {'Target': '', 'PL-ID': ''}

        for vals in ws.iter_rows(min_row=2, values_only=True):
            if not any(v for v in vals if v is not None and str(v).strip()):
                continue
            raw_row = {}
            for ci, h in enumerate(hdrs):
                v = vals[ci] if ci < len(vals) else None
                if isinstance(v, datetime):
                    v = v.date().isoformat()
                elif v is not None:
                    v = str(v).strip()
                raw_row[h] = v or ''

            def _v(*names):
                for name in names:
                    if raw_row.get(name):
                        return raw_row.get(name)
                return ''

            row_dict = {
                'Target':             _v('Target'),
                'PL-ID':              _v('PL-ID', 'PL ID', 'PLID', 'Target PL', 'Target.PL', 'TargetPL'),
                'CRID':               _v('CRID'),
                'CR Count':           _v('CR Count', 'CR Instances'),
                'CR Date':            _v('CR Date'),
                'CR Area':            _v('CR Area'),
                'CR Subsystem':       _v('CR Subsystem'),
                'CR Functionality':   _v('CR Functionality'),
                'CR SI':              _v('CR SI'),
                'CR Status':          _v('CR Status'),
                'Device ID':          _v('Device ID', 'Device'),
                'Station':            _v('Station'),
                'Farm':               _v('Farm'),
                'Test/Setup Details': _v('Test/Setup Details', 'Setup'),
                'Labels':             _v('Labels'),
            }
            for merge_key in ('Target', 'PL-ID'):
                if row_dict.get(merge_key):
                    last_merged_values[merge_key] = row_dict[merge_key]
                elif last_merged_values.get(merge_key):
                    row_dict[merge_key] = last_merged_values[merge_key]
            if not _is_ucr_data_row(row_dict):
                continue
            if not row_dict.get('Farm'):
                station = str(row_dict.get('Station') or '').strip().upper()
                row_dict['Farm'] = farm_map.get(station, 'PDT')
            rows_out.append(row_dict)

        return {'headers': common_headers, 'rows': _sort_ucr_site_rows(rows_out), 'site': site}
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _generate_ucr_excel(parsed: dict, farm_map: dict, week_end_date: date) -> str:
    """
    Generate the Unique_CRs_YYYY_Week_WW.xlsx workbook from parsed CSV data.
    Sheets: QIPL_AreasWiseDistribution, QIPL, Farm_KPI, SD, CH
    Saves to WeeklyUniqueCRs/{YYYY}/
    Returns the saved filepath.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint
    from collections import Counter as _Counter

    year   = week_end_date.year
    iso_wk = week_end_date.isocalendar()[1]
    folder = os.path.join(_UNIQUE_CR_EXCEL_BASE, str(year))
    os.makedirs(folder, exist_ok=True)
    out_path = os.path.join(folder, f'Unique_CRs_{year}_Week_{iso_wk}.xlsx')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default Sheet

    # Ã¢Â”Â€Ã¢Â”Â€ helper: farm colour Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€
    def _farm_bg(farm):
        return _XL_FARM_COLORS.get(str(farm).strip(), _XL_PDT_FILL)

    # Ã¢Â”Â€Ã¢Â”Â€ QIPL_AreasWiseDistribution sheet Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€
    ws_pie = wb.create_sheet('QIPL_AreasWiseDistribution')
    qipl_rows = parsed['rows'].get('QIPL', [])
    area_ctr  = _Counter(
        str(r.get('CR Area') or 'Unknown').strip()
        for r in qipl_rows if str(r.get('CR Area') or '').strip()
    )
    total_qipl = sum(area_ctr.values())
    # write table starting at row 6 (matches existing format)
    _xl_hdr(ws_pie, 6, 1, 'CR Area',   bg='FF4472C4')
    _xl_hdr(ws_pie, 6, 2, 'CRs Count', bg='FF4472C4')
    pie_data_start = 7
    for i, (area, cnt) in enumerate(
            sorted(area_ctr.items(), key=lambda x: -x[1]), start=pie_data_start):
        bg = 'FFD9E1F2' if i % 2 == 0 else 'FFFFFFFF'
        _xl_data(ws_pie, i, 1, area, bg=bg)
        _xl_data(ws_pie, i, 2, cnt,  bg=bg, align='center')
    total_row = pie_data_start + len(area_ctr)
    _xl_hdr(ws_pie, total_row, 1, 'Grand Total', bg='FF2E4057')
    _xl_hdr(ws_pie, total_row, 2, total_qipl,    bg='FF2E4057')
    # metadata cells (row 2-3, matches existing)
    ws_pie.cell(2, 1, total_qipl)
    ws_pie.cell(3, 1, 'crTable')
    # Pie chart
    pie_chart = PieChart()
    pie_chart.title  = f'QIPL CR Area Distribution Ã¢Â€Â” Week {iso_wk}'
    pie_chart.style  = 10
    pie_chart.width  = 18
    pie_chart.height = 14
    labels = Reference(ws_pie, min_col=1, min_row=pie_data_start,
                       max_row=pie_data_start + len(area_ctr) - 1)
    data   = Reference(ws_pie, min_col=2, min_row=pie_data_start - 1,
                       max_row=pie_data_start + len(area_ctr) - 1)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.dataLabels = None
    ws_pie.add_chart(pie_chart, 'D2')
    _xl_autofit(ws_pie)

    # Ã¢Â”Â€Ã¢Â”Â€ site sheets: QIPL, SD, CH Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€
    _SITE_COLS = {
        'QIPL': [
            ('S.No',            'its '),
            ('Target',          'Target'),
            ('PL-ID',           'PL-ID'),
            ('CRID',            'CRID'),
            ('CR Instances',    'CR Count'),
            ('CR Date',         'CR Date'),
            ('CR Area',         'CR Area'),
            ('CR Subsystem',    'CR Subsystem'),
            ('CR Functionality','CR Functionality'),
            ('CR SI',           'CR SI'),
            ('CR Status',       'CR Status'),
            ('Station',         'Station'),
            ('Device',          'Device ID'),
            ('Farm',            'Farm'),
            ('Setup',           'Test/Setup Details'),
        ],
        'SD': [
            ('S.No.',           'its '),
            ('Target',          'Target'),
            ('Target PL',       'PL-ID'),
            ('CRID',            'CRID'),
            ('CR Count',        'CR Count'),
            ('CR Date',         'CR Date'),
            ('CR Area',         'CR Area'),
            ('CR Subsystem',    'CR Subsystem'),
            ('CR Functionality','CR Functionality'),
            ('CR SI',           'CR SI'),
            ('CR Status',       'CR Status'),
            ('Device ID',       'Device ID'),
            ('Station',         'Station'),
            ('Farm',            'Farm'),
        ],
        'CH': [
            ('S.No.',           'its '),
            ('Target',          'Target'),
            ('Target.PL',       'PL-ID'),
            ('CRID',            'CRID'),
            ('CR Count',        'CR Count'),
            ('CR Date',         'CR Date'),
            ('CR Area',         'CR Area'),
            ('CR Subsystem',    'CR Subsystem'),
            ('CR Functionality','CR Functionality'),
            ('CR SI',           'CR SI'),
            ('CR Status',       'CR Status'),
            ('Device ID',       'Device ID'),
            ('Station',         'Station'),
            ('Farm',            'Farm'),
            ('Labels',          'Labels'),
        ],
    }

    # Columns to merge per site: list of src_key values that should be
    # merged when consecutive rows share the same value.
    _MERGE_KEYS = {
        'QIPL': ['Target', 'PL-ID'],
        'SD':   ['Target', 'PL-ID'],
        'CH':   ['Target', 'PL-ID'],
    }

    _sort_ucr_rows_by_site(parsed)

    for site in ['QIPL', 'SD', 'CH']:
        site_rows = parsed['rows'].get(site, [])
        cols      = _SITE_COLS[site]
        ws_s      = wb.create_sheet(site)
        # header row
        for ci, (hdr_label, _) in enumerate(cols, start=1):
            _xl_hdr(ws_s, 1, ci, hdr_label)
        ws_s.freeze_panes = 'A2'

        # build col-index map: src_key -> 1-based column index
        col_idx = {src_key: ci for ci, (_, src_key) in enumerate(cols, start=1)}

        # data rows
        for ri, r in enumerate(site_rows, start=2):
            station  = str(r.get('Station') or '').strip().upper()
            farm_val = r.get('Farm') or farm_map.get(station, 'PDT')
            row_bg   = _farm_bg(farm_val) if farm_val else ('FFD9E1F2' if ri % 2 == 0 else 'FFFFFFFF')
            for ci, (_, src_key) in enumerate(cols, start=1):
                if src_key == 'its ':
                    val = ri - 1
                elif src_key == 'Farm':
                    val = farm_val
                else:
                    val = r.get(src_key) or ''
                    if src_key == 'CR Date' and val:
                        try:
                            val = datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
                        except Exception:
                            pass
                is_farm_col = (src_key == 'Farm')
                cell_bg   = _farm_bg(farm_val) if is_farm_col else row_bg
                cell_bold = is_farm_col and bool(farm_val)
                _xl_data(ws_s, ri, ci, val, bg=cell_bg, bold=cell_bold,
                         align='center' if src_key in ('its ', 'CR Count', 'CR Date') else 'left')

        # Ã¢Â”Â€Ã¢Â”Â€ merge consecutive identical cells in Target / PL-ID columns Ã¢Â”Â€Ã¢Â”Â€
        from openpyxl.styles import Alignment as _Align
        for merge_key in _MERGE_KEYS.get(site, []):
            ci = col_idx.get(merge_key)
            if ci is None:
                continue
            n_rows    = len(site_rows)
            data_start = 2                    # first data row (1-indexed)
            data_end   = data_start + n_rows  # exclusive
            run_start  = data_start
            run_val    = str(ws_s.cell(data_start, ci).value or '')
            for row_i in range(data_start + 1, data_end + 1):
                cur_val = str(ws_s.cell(row_i, ci).value or '') if row_i < data_end else None
                if cur_val == run_val and cur_val != '':
                    continue   # still in the same run
                # flush the run
                if row_i - run_start > 1:   # 2+ rows Ã¢Â†Â’ merge
                    ws_s.merge_cells(
                        start_row=run_start, start_column=ci,
                        end_row=row_i - 1,   end_column=ci
                    )
                    top_cell = ws_s.cell(run_start, ci)
                    top_cell.alignment = _Align(
                        horizontal='left', vertical='center',
                        wrap_text=False
                    )
                run_start = row_i
                run_val   = cur_val or ''

        _xl_autofit(ws_s)

    # Ã¢Â”Â€Ã¢Â”Â€ PDTSite sheet: same data/chart as the web PDT Site bar chart Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€
    _build_ucr_pdtsite_sheet(wb, parsed)

    # Ã¢Â”Â€Ã¢Â”Â€ Farm_KPI sheet Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€Ã¢Â”Â€
    ws_farm = wb.create_sheet('Farm_KPI')
    _xl_hdr(ws_farm, 1, 1, 'FarmType', bg='FF4472C4')
    _xl_hdr(ws_farm, 1, 2, 'HostPC',   bg='FF4472C4')
    ri = 2
    for station, farm in sorted(farm_map.items(), key=lambda x: (x[1], x[0])):
        bg = _farm_bg(farm)
        _xl_data(ws_farm, ri, 1, farm,    bg=bg, bold=True)
        _xl_data(ws_farm, ri, 2, station, bg=bg)
        ri += 1
    _xl_autofit(ws_farm)

    wb.save(out_path)
    return out_path

def _sp_pick(row: dict, *names) -> str:
    wanted = {_norm(n) for n in names}
    for k, v in (row or {}).items():
        if _norm(k) in wanted and v is not None and str(v).strip():
            return str(v).strip()
    return ''


def _clean_sp_ticket(value: str) -> str:
    """Remove CHIPMD ticket tokens and return a clean comma-separated ticket list."""
    tokens = _re.split(r'[,;\s]+', str(value or '').strip())
    cleaned = []
    for tok in tokens:
        tok = tok.strip().strip(',;')
        if not tok:
            continue
        if tok.upper().startswith('CHIPMD'):
            continue
        cleaned.append(tok.upper())
    return ', '.join(dict.fromkeys(cleaned))


def _sp_primary_ticket(row: dict) -> str:
    """Prefer CR ticket values; use Stability Ticket only when no CR exists."""
    cr = _clean_sp_ticket(_sp_pick(
        row,
        'CR/Current Ticket', 'CR Current Ticket', 'CR Current', 'Current Ticket',
        'cr_current_ticket', 'CR', 'CRID', 'CR ID', 'CR Number', 'CR#'
    ))
    if cr:
        return cr
    return _clean_sp_ticket(_sp_pick(row, 'Stability Ticket', 'StabilityTicket', 'stability_ticket'))




def _sp_pl_id(row: dict) -> str:
    return (_sp_pick(row, 'PL-ID', 'PL ID', 'pl_id', 'JIRA Component', 'jira_component') or 'Unknown PL-ID').strip()


def _sp_build_type(row: dict) -> str:
    raw = (_sp_pick(row, 'Build Type', 'build_type', 'CRM/ENG', 'CRM ENG', 'Release Type', 'release_type') or '').strip().upper()
    if raw in ('CRM', 'ENG'):
        return raw
    return 'CRM'


def _sp_metabuild(row: dict) -> str:
    return (_sp_pick(row, 'MetaBuild', 'Meta Build', 'Build', 'Build ID', 'BuildID', 'meta_build') or 'Unknown Build').strip()


def _sp_allow_qipl_selector_target(target: str) -> bool:
    """Allow SQL weekly_qipl_data to seed SharePoint Target/PL selectors.

    SWPDT JSON is still preferred, but historical/uploaded weekly SQL rows can
    contain valid targets that are absent from the JSON for that week. Skip
    obvious synthetic combined labels that caused noise in the main dropdown.
    """
    text = str(target or '').strip()
    if not text or _is_ucr_non_data_marker(text):
        return False
    # Example synthetic label from older data: Kailua_Sariska.LA_MN.
    if '_' in text and _re.search(r'_[A-Za-z0-9]+\.', text):
        return False
    return True


def _sp_row_text(row: dict) -> str:
    parts = []
    for v in (row or {}).values():
        if isinstance(v, (str, int, float)) and str(v).strip():
            parts.append(str(v))
    return ' | '.join(parts)


def _extract_sp_crash_items(text: str) -> list:
    """Extract CR/QSTABILITY/CNSSDEBUG tokens like CR123(x2) from text."""
    items = []
    seen = set()
    for m in _re.finditer(r'\b((?:CR\d+)|(?:QSTABILITY-\d+)|(?:CNSSDEBUG-\d+)|(?:[A-Z]+-\d+))\s*(?:\(\s*x\s*(\d+)\s*\))?', str(text or ''), _re.IGNORECASE):
        cr = m.group(1).upper()
        if cr.startswith('CHIPMD'):
            continue
        cnt = int(m.group(2) or 1)
        key = (cr, cnt)
        if key in seen:
            continue
        seen.add(key)
        items.append({'cr': cr, 'count': cnt, 'label': f'{cr}(x{cnt})'})
    return items


def _build_sp_ticket_pivot(ticket_counter, events: list | None = None) -> dict:
    """Build the CR/Current Ticket pivot stored with each Target+PL+Build.

    ticket_counter is a Counter of {ticket_id: occurrence_count}.
    Returns a dict with:
      items   Â– list of {ticket, count, label} sorted by ticket name
      total   Â– sum of all occurrence counts  (this is the crash count)
      details Â– human-readable string like 'CR123(x5), CR456(x3)'
    """
    if not isinstance(ticket_counter, Counter):
        ticket_counter = Counter(ticket_counter or {})
    items = [
        {'ticket': str(t), 'count': int(c or 0), 'label': f'{t}(x{int(c or 0)})'}
        for t, c in sorted(ticket_counter.items(), key=lambda x: str(x[0]))
        if str(t or '').strip() and int(c or 0) > 0
    ]
    clean_events = []
    reporters = []
    for ev in (events or []):
        if not isinstance(ev, dict):
            continue
        ticket = str(ev.get('ticket') or ev.get('cr') or '').strip().upper()
        if not ticket or ticket.startswith('CHIPMD'):
            continue
        reporter = str(ev.get('reporter') or '').strip()
        count = int(ev.get('count') or 1)
        clean_events.append({'ticket': ticket, 'reporter': reporter, 'count': count})
        if reporter and reporter not in reporters:
            reporters.append(reporter)
    total = sum(item['count'] for item in items)
    return {
        'items':   items,
        'total':   total,
        'details': ', '.join(item['label'] for item in items),
        'events': clean_events,
        'reporters': sorted(reporters, key=lambda x: x.lower()),
    }




def _sp_build_match_sql_expr() -> str:
    """SQL expression used to match build IDs like the Sharepoint pivot."""
    return """
        UPPER(TRIM(COALESCE(
            NULLIF(meta_build, ''),
            JSON_UNQUOTE(JSON_EXTRACT(row_data,'$.MetaBuild')),
            JSON_UNQUOTE(JSON_EXTRACT(row_data,'$.Metabuild')),
            JSON_UNQUOTE(JSON_EXTRACT(row_data,'$.meta_build')),
            JSON_UNQUOTE(JSON_EXTRACT(row_data,'$.\"Meta Build\"')),
            JSON_UNQUOTE(JSON_EXTRACT(row_data,'$.Build'))
        )))
    """



def _sp2_pl_group(value: str) -> str:
    """Normalize Axiom/QIPL PL-ID the same way Smart Build groups rows."""
    import re as _sp2_re
    return _sp2_re.sub(r'\.r\d+$', '', str(value or '').strip(), flags=_sp2_re.IGNORECASE)


def _sp2_weekly_crash_map(week_start, week_end) -> dict:
    """Return crash counts keyed by (meta_build_upper, pl_id_upper).

    Rules (confirmed from data analysis):
      - Each CSV row has a unique stability_ticket (never NULL).
      - stability_ticket LIKE 'CHIPMD%'  -> HWPDT crashes -> EXCLUDE.
      - stability_ticket LIKE 'QSTABILITY%' or 'DROIDBUG%' -> PDT crashes -> COUNT.
      - Crash count = COUNT(DISTINCT stability_ticket) per meta_build + pl_id.
      - Key: (meta_build.strip().upper(), pl_id.strip().upper())
    """
    ws = _safe_date(week_start)
    we = _safe_date(week_end)
    if not ws or not we:
        return {}
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {}
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT meta_build, pl_id,
                   COUNT(DISTINCT stability_ticket) AS crash_count
            FROM `{_QIPL_DB}`.`{_QIPL_TABLE}`
            WHERE week_start=%s AND week_end=%s
              AND stability_ticket IS NOT NULL
              AND stability_ticket != ''
              AND stability_ticket NOT LIKE 'CHIPMD%%'
            GROUP BY meta_build, pl_id
        """, (ws.isoformat(), we.isoformat()))
        result = {}
        for row in cur.fetchall() or []:
            mb  = str(row.get('meta_build') or '').strip().upper()
            pl  = str(row.get('pl_id')      or '').strip().upper()
            cnt = int(row.get('crash_count') or 0)
            if mb and pl and cnt > 0:
                # accumulate in case same meta_build appears under multiple pl_id variants
                result[(mb, pl)] = result.get((mb, pl), 0) + cnt
        return result
    except Exception:
        return {}
    finally:
        try:
            cur.close(); conn.close()
        except Exception:
            pass


def _sp2_crash_count_for_build(crash_map: dict, build_name: str = '', build_id: str = '', pl_id: str = '') -> int:
    """Look up crash count for one build row from the crash_map.

    Key is (meta_build.upper(), pl_id.upper()).
    Tries build_name first, then build_id, then last path component of each.
    pl_id is matched exactly (e.g. 'SA515M.LE.2.1.1') - no stripping.
    """
    pl_upper = str(pl_id or '').strip().upper()
    candidates = []
    for raw in (build_name, build_id):
        text = str(raw or '').strip().upper()
        if not text:
            continue
        candidates.append(text)
        # last path component (e.g. from a full path)
        last = text.replace('\\', '/').rstrip('/').split('/')[-1]
        if last and last != text:
            candidates.append(last)
    for key in dict.fromkeys(c for c in candidates if c):
        if (key, pl_upper) in crash_map:
            return int(crash_map[(key, pl_upper)] or 0)
    # fallback: match meta_build key ignoring pl_id (sum across all PLs for this build)
    for (mb, pl), cnt in crash_map.items():
        if mb in candidates:
            return int(cnt or 0)
    return 0



def _count_sharepoint_crashes_from_weekly_qipl(cur, target: str, pl_id: str, build_ids, week_start=None, week_end=None, jira_reporters=None) -> dict:
    """Count Target+PL-ID+build crashes from weekly_qipl_data using pivot rules."""
    normalized_builds = sorted({str(b or '').strip().upper() for b in (build_ids or []) if str(b or '').strip()})
    if not normalized_builds:
        pivot = _build_sp_ticket_pivot(Counter())
        pivot['row_count'] = 0
        pivot['stability_count'] = 0
        return pivot

    where = ["TRIM(target)=TRIM(%s)", "TRIM(pl_id)=TRIM(%s)"]
    params = [str(target or '').strip(), str(pl_id or '').strip()]
    ws = _safe_date(week_start)
    we = _safe_date(week_end)
    if ws:
        where.append("week_start=%s")
        params.append(ws.isoformat())
    if we:
        where.append("week_end=%s")
        params.append(we.isoformat())

    ph = ','.join(['%s'] * len(normalized_builds))
    where.append(f"{_sp_build_match_sql_expr()} IN ({ph})")
    params.extend(normalized_builds)

    cur.execute(f"""
        SELECT row_data, cr_current_ticket, stability_ticket, meta_build, target, pl_id, jira_reporter
        FROM `{_QIPL_DB}`.`{_QIPL_TABLE}`
        WHERE {' AND '.join(where)}
    """, tuple(params))
    live_rows = cur.fetchall() or []

    selected_reporters = {str(r or '').strip().upper() for r in (jira_reporters or []) if str(r or '').strip()}
    stab_seen = set()
    ticket_counter = Counter()
    ticket_events = []
    counted_rows = 0
    for lr in live_rows:
        try:
            d = json.loads(lr.get('row_data') or '{}') if isinstance(lr, dict) else json.loads(lr['row_data'] or '{}')
        except Exception:
            d = {}
        if isinstance(lr, dict):
            for k in ('cr_current_ticket', 'stability_ticket', 'meta_build', 'target', 'pl_id', 'jira_reporter'):
                if lr.get(k) not in (None, ''):
                    d[k] = lr.get(k)

        stab = str(_sp_pick(d, 'Stability Ticket', 'StabilityTicket', 'stability_ticket') or '').strip()
        if stab:
            stab_key = stab.upper()
            if stab_key in stab_seen:
                continue
            stab_seen.add(stab_key)
        reporter = str(_sp_pick(d, 'JIRA Reporter', 'Jira Reporter', 'jira_reporter', 'Reporter', 'Reported By') or '').strip()
        if selected_reporters and reporter.upper() not in selected_reporters:
            continue
        # Compute BU: exclude crashes where JIRA title contains "LKD"
        _jira_title = str(_sp_pick(d, "JIRA Title", "Title", "Summary", "jira_title", "title", "summary") or "").strip().upper()
        _tgt_upper  = str(target or "").strip().upper()
        if "COMPUTE" in _tgt_upper and "LKD" in _jira_title:
            continue
        counted_rows += 1

        ticket = _sp_primary_ticket(d)
        if not ticket:
            continue
        row_seen = set()
        for tok in [x.strip() for x in ticket.split(',') if x.strip()]:
            up = tok.upper()
            if up.startswith('CHIPMD') or up in row_seen:
                continue
            row_seen.add(up)
            ticket_counter[up] += 1
            ticket_events.append({'ticket': up, 'reporter': reporter, 'count': 1})

    pivot = _build_sp_ticket_pivot(ticket_counter, ticket_events)
    pivot['row_count'] = counted_rows
    pivot['stability_count'] = len(stab_seen)
    return pivot


def _build_sharepoint_context(sp_rows: list, week_start: date, week_end: date) -> dict:
    targets = []
    rows_by_target = {}
    build_rows_by_target = {}
    pl_ids_by_target = {}
    grouped = {}
    used_build_keys = set()
    saved = _fetch_sharepoint_summaries(week_start, week_end)
    #print(f'[CONSOLIDATE] week_end={week_end} | build_summary rows fetched: {len(saved)} | targets: {list(set(r.get("target") for r in saved))[:10]}')
    for rec in saved:
        saved_meta = str(rec.get('meta_build') or '').strip()
        if saved_meta:
            used_build_keys.add((
                str(rec.get('target') or '').strip(),
                str(rec.get('pl_id') or '').strip(),
                str(rec.get('build_type') or 'CRM').strip().upper(),
                saved_meta.upper(),
            ))
    for idx, row in enumerate(sp_rows or []):
        tgt = str(row.get('target') or row.get('Target') or '').strip()
        if not tgt:
            continue
        ticket = _sp_primary_ticket(row)
        if not ticket and not _is_snapdragon_auto_target(tgt):
            continue
        pl_id = _sp_pl_id(row)
        build_type = _sp_build_type(row)
        meta_build = _sp_metabuild(row)
        if tgt not in rows_by_target:
            rows_by_target[tgt] = []
            build_rows_by_target[tgt] = []
            pl_ids_by_target[tgt] = []
            targets.append(tgt)
        if pl_id not in pl_ids_by_target[tgt]:
            pl_ids_by_target[tgt].append(pl_id)
        rows_by_target[tgt].append(row)
        grp_key = (tgt, pl_id, build_type, meta_build)
        grouped.setdefault(grp_key, {
            'id': f'sp_{idx}',
            'target': tgt,
            'pl_id': pl_id,
            'build_type': build_type,
            'meta_build': meta_build,
            'tickets': [],
            'ticket_set': set(),
            'ticket_counter': Counter(),
            'ticket_events': [],
            'stab_seen': set(),   # dedup by Stability Ticket = unique crash event
            'row_count': 0,
        })
        # -- Dedup by Stability Ticket -----------------------------------------
        # Each Stability Ticket = one unique crash event.
        # The DB has duplicate rows (same crash ingested twice), so we skip any
        # row whose Stability Ticket we have already counted for this build.
        stab_ticket = str(
            _sp_pick(row, 'Stability Ticket', 'StabilityTicket', 'stability_ticket') or ''
        ).strip()
        if stab_ticket and stab_ticket in grouped[grp_key]['stab_seen']:
            continue   # exact duplicate row Â— skip
        if stab_ticket:
            grouped[grp_key]['stab_seen'].add(stab_ticket)
        grouped[grp_key]['row_count'] += 1
        # -- Count CR/Current Ticket (primary) per unique crash event ----------
        # ticket = CR/Current Ticket if present, else Stability Ticket (fallback)
        # Each unique crash event contributes exactly 1 count to its CR bucket.
        reporter = str(_sp_pick(row, 'JIRA Reporter', 'Jira Reporter', 'jira_reporter', 'Reporter', 'Reported By') or '').strip()
        row_tokens_seen = set()
        for tok in [x.strip() for x in str(ticket).split(',') if x.strip()]:
            up = tok.upper()
            if up.startswith('CHIPMD'):
                continue
            if up in row_tokens_seen:
                continue
            row_tokens_seen.add(up)
            # One unique crash event -> one count for this CR
            grouped[grp_key]['ticket_counter'][up] += 1
            grouped[grp_key]['ticket_events'].append({'ticket': up, 'reporter': reporter, 'count': 1})
            if up not in grouped[grp_key]['ticket_set']:
                grouped[grp_key]['ticket_set'].add(up)
                grouped[grp_key]['tickets'].append(up)
    for (_tgt, _pl, _type, _build), info in grouped.items():
        # For Snapdragon Auto targets the same build appears under multiple PL
        # aliases â€” do NOT exclude it just because it was saved under a different
        # PL alias.  Only exclude when target + pl_id + build all match exactly.
        if not _is_snapdragon_auto_target(_tgt):
            if (_tgt, _pl, _type, _build.upper()) in used_build_keys:
                continue
        # Snapdragon Auto: never exclude builds â€” each PL alias is a separate
        # save entry and the user must be able to select the same build again
        # under a different alias. The duplicate-key constraint on the DB
        # prevents actual double-saves.
        ticket_pivot = _build_sp_ticket_pivot(info['ticket_counter'], info.get('ticket_events') or [])
        build_rows_by_target.setdefault(_tgt, []).append({
            'id': info['id'],
            'target': _tgt,
            'pl_id': _pl,
            'build_type': _type,
            'meta_build': _build,
            'tickets': info['tickets'],
            'ticket_pivot': ticket_pivot,
            'ticket_count': ticket_pivot['total'],
            'crash_details': ticket_pivot['details'],
            'row_count': info['row_count'],
            'label': _build,
        })
    _pairs_needing_hist = set()
    for rec in saved:
        _tgt = str(rec.get('target') or '').strip()
        _pl  = str(rec.get('pl_id')  or '').strip()
        if not _tgt:
            continue
        _existing = [r for r in (build_rows_by_target.get(_tgt) or [])
                     if str(r.get('pl_id') or '') == _pl]
        if not _existing:
            _pairs_needing_hist.add((_tgt, _pl))
    if _pairs_needing_hist:
        _hist_targets = list({p[0] for p in _pairs_needing_hist})
        _hist_records = []
        try:
            _hist_conn = get_mysql_connection_db(bu_key=None)
            if _hist_conn:
                _hist_cur = _hist_conn.cursor(dictionary=True)
                try:
                    _ph = ','.join(['%s'] * len(_hist_targets))
                    _hist_cur.execute(
                        f"SELECT target, pl_id, build_type, meta_build, selected_items_json"
                        f" FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`"
                        f" WHERE target IN ({_ph})"
                        f" ORDER BY week_end DESC, id DESC LIMIT 500",
                        tuple(_hist_targets)
                    )
                    _hist_records = _hist_cur.fetchall() or []
                finally:
                    _hist_cur.close()
                    _hist_conn.close()
        except Exception:
            pass
        _hist_seen_upper = set()
        for (_et, _ep, _etype, _eb) in used_build_keys:
            _hist_seen_upper.add((_et, _ep, _eb.upper()))
        for _hrec in _hist_records:
            _tgt  = str(_hrec.get('target')     or '').strip()
            _pl   = str(_hrec.get('pl_id')      or '').strip()
            _type = str(_hrec.get('build_type') or 'CRM').strip().upper()
            if (_tgt, _pl) not in _pairs_needing_hist:
                continue
            _builds_in_rec = []
            _top = str(_hrec.get('meta_build') or '').strip()
            if _top:
                _builds_in_rec.append((_top, []))
            try:
                _sitems = json.loads(_hrec.get('selected_items_json') or '[]')
            except Exception:
                _sitems = []
            for _si in (_sitems or []):
                if not isinstance(_si, dict):
                    continue
                _sb = str(_si.get('meta_build') or '').strip()
                if _sb:
                    _builds_in_rec.append((_sb, _si.get('tickets') or []))
            for (_b, _tickets) in _builds_in_rec:
                _dedup_key = (_tgt, _pl, _b.upper())
                # For Snapdragon Auto: allow same build under different PL aliases
                if not _is_snapdragon_auto_target(_tgt):
                    if _dedup_key in _hist_seen_upper:
                        continue
                else:
                    if _dedup_key in _hist_seen_upper:
                        continue
                _hist_seen_upper.add(_dedup_key)
                _hist_counter = Counter()
                for _ticket in (_tickets or []):
                    for _tok in [x.strip().upper() for x in str(_ticket or '').split(',') if x.strip()]:
                        if _tok and not _tok.startswith('CHIPMD'):
                            _hist_counter[_tok] += 1
                _hist_pivot = _build_sp_ticket_pivot(_hist_counter)
                build_rows_by_target.setdefault(_tgt, []).append({
                    'id': f'hist_{_hrec.get("id", "")}',
                    'target': _tgt, 'pl_id': _pl, 'build_type': _type,
                    'meta_build': _b, 'tickets': _tickets,
                    'ticket_pivot': _hist_pivot,
                    'ticket_count': _hist_pivot['total'],
                    'crash_details': _hist_pivot['details'],
                    'row_count': 0, 'label': _b,
                })
            if _tgt not in targets:
                targets.append(_tgt)
            if _tgt not in pl_ids_by_target:
                pl_ids_by_target[_tgt] = []
            if _pl and _pl not in pl_ids_by_target[_tgt]:
                pl_ids_by_target[_tgt].append(_pl)
    saved_by_target = {}
    for rec in saved:
        saved_by_target.setdefault(rec.get('target') or '', []).append(rec)
    for _tgt in build_rows_by_target:
        build_rows_by_target[_tgt].sort(key=lambda x: (str(x.get('build_type') or ''), str(x.get('meta_build') or '').lower()))
        saved_pairs = {
        (str(rec.get('target') or '').strip(), str(rec.get('pl_id') or '').strip())
        for rec in saved
    }
    # Make the PL-ID dropdown complete.  Source Sharepoint rows only contain
    # builds that produced stability tickets, so no-crash / Unique-CR-only PLs
    # can be missing unless we also seed the selector from saved rows and the
    # selected week's Unique CR workbook.
    def _add_pl_option(raw_target, raw_pl_id, strict_target: bool = False):
        raw_target = str(raw_target or '').strip()
        raw_pl_id = str(raw_pl_id or '').strip()
        if not raw_target or not raw_pl_id:
            return
        target_key = raw_target
        if not strict_target:
            target_aliases = set(_ucr_target_aliases(raw_target))
            for existing_target in list(targets):
                existing_aliases = set(_ucr_target_aliases(existing_target))
                if _ucr_match_text(existing_target) == _ucr_match_text(raw_target) or (target_aliases and existing_aliases and target_aliases.intersection(existing_aliases)):
                    target_key = existing_target
                    break
        if target_key not in targets:
            targets.append(target_key)
        rows_by_target.setdefault(target_key, [])
        build_rows_by_target.setdefault(target_key, [])
        pl_ids_by_target.setdefault(target_key, [])
        if raw_pl_id not in pl_ids_by_target[target_key]:
            pl_ids_by_target[target_key].append(raw_pl_id)



        

    # Seed the Sharepoint Target / PL-ID selectors from QIPL SWPDT JSON only.
    # This avoids synthetic combined names from weekly_qipl_data such as
    # Kailua_Sariska.LA_MN appearing in the main dropdown unless they truly
    # exist in the QIPL JSON source.
    swpdt_pairs = []
    strict_targets = set()
    try:
        swpdt_pairs = (_swpdt_weekly_target_pl_options(week_start, week_end) or {}).get('rows') or []
    except Exception:
        swpdt_pairs = []

    if swpdt_pairs:
        targets = []
        rows_by_target = {k: v for k, v in rows_by_target.items() if k in build_rows_by_target}
        build_rows_by_target = {k: v for k, v in build_rows_by_target.items()}
        pl_ids_by_target = {}
        for pair in swpdt_pairs:
            strict_targets.add(str(pair.get('target') or '').strip())
            _add_pl_option(pair.get('target'), pair.get('pl_id'), strict_target=True)
    for rec in saved:
        _rt = str(rec.get('target') or '').strip()
        _add_pl_option(_rt, rec.get('pl_id'), strict_target=(_rt in strict_targets))

    # Also seed selectors from the selected week's SQL rows. This covers valid
    # PLs present in weekly_qipl_data/Excel, e.g. SA515M.LE for week ending 6/14,
    # even when the SWPDT JSON source does not emit that Target/PL for the week.
    for row in sp_rows or []:
        _rt = str(row.get('target') or row.get('Target') or '').strip()
        if not _sp_allow_qipl_selector_target(_rt):
            continue
        _pl = _sp_pl_id(row)
        if _pl and _pl != 'Unknown PL-ID':
            _add_pl_option(_rt, _pl, strict_target=True)



    missing_targets = []
    saved_builds_by_pair = {}
    for rec in saved:
        _pair = (str(rec.get('target') or '').strip(), str(rec.get('pl_id') or '').strip())
        _top_meta = str(rec.get('meta_build') or '').strip()
        if _top_meta:
            saved_builds_by_pair.setdefault(_pair, set()).add(_top_meta.upper())
        for _item in (rec.get('selected_items') or []):
            if not isinstance(_item, dict):
                continue
            _meta = str(_item.get('meta_build') or '').strip()
            if _meta:
                saved_builds_by_pair.setdefault(_pair, set()).add(_meta.upper())

        json_builds_by_pair = {}
    try:
        _payload, _ = _load_swpdt_json_payload()
        for _build in _flatten_swpdt_build_entries(_payload):
            _submitted_date = _axiom_date_from_value(_build.get('submitted'))
            if week_start and _submitted_date and _submitted_date < week_start:
                continue
            if week_end and _submitted_date and _submitted_date > week_end:
                continue
            _pl = str(_build.get('software_product') or '').strip()
            _tgt = _swpdt_target_from_product(_pl) or _pl
            _bid = str(_build.get('build_id') or '').strip()
            if not _tgt or not _pl or not _bid:
                continue
            _keys = sorted(_build_key_variants(_bid))
            _store_bid = _keys[0] if _keys else str(_bid).strip().upper()
            json_builds_by_pair.setdefault((_tgt, _pl), set()).add(_store_bid)
    except Exception:
        json_builds_by_pair = {}


    for (tgt, pl_id), json_builds in sorted(json_builds_by_pair.items(), key=lambda x: (str(x[0][0]).lower(), str(x[0][1]).lower())):
        saved_builds = saved_builds_by_pair.get((tgt, pl_id), set())
        missing = []
        for _bid in sorted(json_builds, key=lambda x: x.lower()):
            if _bid.upper() in saved_builds:
                continue
            missing.append(_bid)
        if not missing:
            continue
        missing_targets.append({'target': tgt, 'pl_id': pl_id, 'build_ids': ', '.join(missing)})


    return {
        'sp_targets': sorted(targets, key=lambda x: x.lower()),
        'sp_rows_by_target': rows_by_target,
        'sp_pl_ids_by_target': {k: sorted(v, key=lambda x: x.lower()) for k, v in pl_ids_by_target.items()},
        'sp_build_rows_by_target': build_rows_by_target,
        'sp_crash_items_by_target': {},
        'sp_saved_summaries': saved,
        'sp_saved_by_target': saved_by_target,
        'sp_missing_targets': missing_targets,
    }

def _fmt_mmddyyyy(val) -> str:
    d = _safe_date(val)
    return d.strftime('%m/%d/%Y') if d else ''


def _fmt_iso_date(val) -> str:
    d = _safe_date(val)
    return d.isoformat() if d else ''


def _sp_timeline(es, fc, cs) -> str:
    return f"( ES - {_fmt_mmddyyyy(es)}, FC - {_fmt_mmddyyyy(fc)}, CS - {_fmt_mmddyyyy(cs)} )"


def _sp_manual_meta(rec: dict) -> dict:
    for item in (rec.get('selected_items') or []):
        if isinstance(item, dict) and item.get('manual_no_crash'):
            return item
    return {}


def _sp_saved_build_count(rec: dict) -> int:
    """Count distinct selected builds represented by one saved Sharepoint row."""
    builds = set()
    selected_items = rec.get('selected_items') or []
    if not selected_items:
        try:
            selected_items = json.loads(rec.get('selected_items_json') or '[]')
        except Exception:
            selected_items = []
    for item in selected_items or []:
        if not isinstance(item, dict) or item.get('manual_bu'):
            continue
        meta = str(item.get('meta_build') or '').strip()
        if meta:
            builds.add(meta.upper())
    top_meta = str(rec.get('meta_build') or '').strip()
    if top_meta:
        builds.add(top_meta.upper())
    return max(len(builds), 1)


def _sp_pair_key(target: str, pl_id: str) -> str:

    return f"{str(target or '').strip().upper()}||{str(pl_id or '').strip().upper()}"


def _sharepoint_build_label_for_meta(week_start: date, week_end: date, target: str, pl_id: str, build_type: str, meta_build: str) -> str:
    """Return stable Build label for same week+target+PL+type+meta build.

    Rules:
    - same target + PL-ID + build_type + meta_build => reuse existing build_label
    - same target + PL-ID + build_type + different meta_build => next build_label number
    """
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return '1'
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT build_label, meta_build
            FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`
            WHERE week_start=%s AND week_end=%s AND target=%s
              AND COALESCE(pl_id,'')=%s AND COALESCE(build_type,'CRM')=%s
            ORDER BY id
        """, (week_start.isoformat(), week_end.isoformat(), target, pl_id, build_type or 'CRM'))
        rows = cur.fetchall() or []
        wanted_meta = str(meta_build or '').strip().upper()
        nums = []
        for r in rows:
            label = str(r.get('build_label') or '').strip()
            existing_meta = str(r.get('meta_build') or '').strip().upper()
            if wanted_meta and existing_meta == wanted_meta and label:
                return label
            m = _re.search(r'(\d+)', label)
            if m:
                nums.append(int(m.group(1)))
        return str(max(nums or [0]) + 1)
    except Exception:
        return '1'
    finally:
        cur.close(); conn.close()





def _sp_bu_options() -> list:
    """BU dropdown options for Sharepoint manual target entry."""
    keys = set((BU_DATABASE_MAPPING or {}).keys()) | set((STATIC_BUSINESS_UNITS or {}).keys())
    skip = {'WEEKLY_QIPL_REPORTS'}
    opts = []
    for key in sorted(k for k in keys if k and k not in skip):
        meta = (STATIC_BUSINESS_UNITS or {}).get(key) or {}
        label = str(meta.get('display_name') or key).strip()
        opts.append({'key': key, 'label': label})
    return opts


def _normalize_bu(bu: str) -> str:
    """Normalize BU aliases to canonical keys used in BU_DATABASE_MAPPING.
    e.g. IOT_WEARABLES -> IOT (same schema, same BU, different display name).
    """
    _aliases = {
        'IOT_WEARABLES': 'IOT',
        'IOT_WEARABLE':  'IOT',
        'WEARABLES':     'IOT',
        'AUTOMOTIVE':    'AUTO',
    }
    b = str(bu or '').strip().upper()
    return _aliases.get(b, b)


def _resolve_sharepoint_build_milestones(target: str, rec: dict | None = None, manual: dict | None = None) -> dict:
    """Resolve BU + milestone dates from explicit columns, manual JSON, and dashboard target lookup.

    Normal weekly report flows must not call OneView synchronously because that
    endpoint can timeout and make Sharepoint saves/refreshes feel stuck. Use the
    explicit milestone/backfill routes when OneView refresh is required.
    """
    rec = rec or {}
    manual = manual or {}
    target = str(target or rec.get('target') or '').strip()
    sp_name = str(rec.get('sp_name') or manual.get('sp_name') or target).strip()
    info = {
        'bu': _normalize_bu(str(rec.get('bu') or manual.get('bu') or '').strip()),
        'es': _fmt_iso_date(rec.get('es_date')) or _fmt_iso_date(manual.get('es')),
        'fc': _fmt_iso_date(rec.get('fc_date')) or _fmt_iso_date(manual.get('fc')),
        'cs': _fmt_iso_date(rec.get('cs_date')) or _fmt_iso_date(manual.get('cs')),
    }
    dash = _find_dashboard_target_info(target, rec.get('pl_id') or manual.get('pl_id')) if target or rec.get('pl_id') or manual.get('pl_id') else {}
    for key, src_key in (('bu', 'bu'), ('es', 'es'), ('fc', 'fc'), ('cs', 'cs')):
        if not info.get(key) and dash.get(src_key):
            info[key] = dash.get(src_key)
    return info


def _fetch_previous_sharepoint_pair_info(target: str, pl_id: str, before_week_start: date | None = None) -> dict:
    """Latest saved no-crash metadata for Target+PL/PL before selected week.

    Hours/devices are intentionally not returned for UI auto-fill; only static
    milestone/BU/SP details carry forward from a previous week such as 05/24 to
    the next selected week such as 05/31. If target is blank, PL-ID alone is
    enough to find the latest matching saved record.
    """
    target = str(target or '').strip()
    pl_id = str(pl_id or '').strip()
    if not target and not pl_id:
        return {}
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {}
    cur = conn.cursor(dictionary=True)
    try:
        clauses = []
        params = []
        if target:
            clauses.append('target=%s')
            params.append(target)
        if pl_id:
            clauses.append('COALESCE(pl_id,\'\')=%s')
            params.append(pl_id)
        if before_week_start:
            clauses.append('week_start < %s')
            params.append(before_week_start.isoformat())
        where_sql = ' AND '.join(clauses) if clauses else '1=0'
        cur.execute(f"""
            SELECT * FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`
            WHERE {where_sql}
            ORDER BY week_start DESC, week_end DESC, updated_at DESC, id DESC
            LIMIT 20
        """, tuple(params))
        for rec in cur.fetchall() or []:
            try:
                items = json.loads(rec.get('selected_items_json') or '[]')
            except Exception:
                items = []
            # Check for manual_no_crash item first (carries full metadata)
            for item in items or []:
                if isinstance(item, dict) and item.get('manual_no_crash'):
                    out = dict(item)
                    out.update({
                        'target': target or str(rec.get('target') or ''),
                        'pl_id': str(rec.get('pl_id') or pl_id or ''),
                        'build_type': str(rec.get('build_type') or item.get('build_type') or 'CRM').upper(),
                        'meta_build': str(rec.get('meta_build') or item.get('meta_build') or ''),
                        'source': 'previous_week',
                        'source_week_start': _fmt_iso_date(rec.get('week_start')),
                        'source_week_end': _fmt_iso_date(rec.get('week_end')),
                    })
                    return out
            # Fallback: crash-path rows have no manual_no_crash item but
            # still carry bu/target/pl_id directly on the row columns.
            row_bu = _normalize_bu(str(rec.get('bu') or '').strip())
            if row_bu:
                return {
                    'target': str(rec.get('target') or target or ''),
                    'pl_id': str(rec.get('pl_id') or pl_id or ''),
                    'bu': row_bu,
                    'build_type': str(rec.get('build_type') or 'CRM').upper(),
                    'meta_build': str(rec.get('meta_build') or ''),
                    'es': _fmt_iso_date(rec.get('es_date')),
                    'fc': _fmt_iso_date(rec.get('fc_date')),
                    'cs': _fmt_iso_date(rec.get('cs_date')),
                    'source': 'previous_week',
                    'source_week_start': _fmt_iso_date(rec.get('week_start')),
                    'source_week_end': _fmt_iso_date(rec.get('week_end')),
                }
        return {}
    except Exception:
        return {}
    finally:
        cur.close(); conn.close()


def _find_dashboard_target_info(target: str, pl_id: str | None = None) -> dict:
    """Resolve BU/milestone metadata from dashboard_status using Target and PL-ID aliases."""
    target = str(target or '').strip()
    pl_id = str(pl_id or '').strip()
    candidates = [v for v in (target, pl_id) if v]
    if not candidates:
        return {}

    def _norm_key(val: str) -> str:
        return _re.sub(r'\s+', '', str(val or '').strip().upper().replace('_', '.'))

    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {}
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT bu, target_name, sp_name, es_date, fc_date, cs_date
            FROM pdt_stats_dashboard.dashboard_status
            WHERE is_active=1
            ORDER BY id ASC
        """)
        rows = cur.fetchall() or []
        dash_map = {}
        for r in rows:
            tn = str(r.get('target_name') or '').strip()
            sp = str(r.get('sp_name') or '').strip()
            info = {
                'target': sp or tn or target,
                'target_name': tn,
                'sp_name': sp,
                'bu': _normalize_bu(str(r.get('bu') or '')),
                'es': _fmt_iso_date(r.get('es_date')),
                'fc': _fmt_iso_date(r.get('fc_date')),
                'cs': _fmt_iso_date(r.get('cs_date')),
            }
            aliases = [tn, sp, tn.replace('_', '.'), sp.replace('_', '.'), tn.replace('.', '_'), sp.replace('.', '_')]
            for raw in aliases:
                nk = _norm_key(raw)
                if nk and nk not in dash_map:
                    dash_map[nk] = info

        for cand in candidates:
            hit = dash_map.get(_norm_key(cand))
            if hit:
                out = dict(hit)
                out['matched_exact'] = True
                out['source'] = 'dashboard_status_exact'
                return out

        raw_match_map = {
            k: {'bu': v.get('bu'), 'ES': v.get('es'), 'FC': v.get('fc'), 'CS': v.get('cs'), 'sp_name': v.get('sp_name'), 'target_name': v.get('target_name')}
            for k, v in dash_map.items()
        }
        for cand in candidates:
            matched = _match_dashboard(cand, raw_match_map)
            if matched and matched.get('bu'):
                return {
                    'target': matched.get('sp_name') or matched.get('target_name') or target,
                    'target_name': matched.get('target_name') or '',
                    'sp_name': matched.get('sp_name') or '',
                    'bu': _normalize_bu(str(matched.get('bu') or '')),
                    'es': _fmt_iso_date(matched.get('ES')),
                    'fc': _fmt_iso_date(matched.get('FC')),
                    'cs': _fmt_iso_date(matched.get('CS')),
                    'matched_exact': False,
                    'source': 'dashboard_status_fuzzy',
                }
        return {}
    except Exception:
        return {}
    finally:
        cur.close(); conn.close()


def _extract_milestones_from_timelines(timelines: str) -> dict:
    text = str(timelines or '')
    out = {'ES': '', 'FC': '', 'CS': ''}
    for key in out:
        m = _re.search(rf'\b{key}\s*-\s*(\d{{1,2}}/\d{{1,2}}/\d{{2,4}})', text, _re.IGNORECASE)
        if m:
            out[key] = _fmt_iso_date(m.group(1))
    return out


def _compute_pdt_test_status(es, cs, fc) -> str:
    """Compute PDT milestone phase using normal ES -> FC -> CS chronology.

    Option A requested by user:
    - today < ES              => Pre-ES
    - ES <= today < FC        => Pre-FC
    - FC <= today < CS        => Post-FC
    - today >= CS             => Post-CS
    """
    today = date.today()
    es_d = _safe_date(es)
    cs_d = _safe_date(cs)
    fc_d = _safe_date(fc)

    if es_d and today < es_d:
        return 'Pre-ES'
    if es_d and fc_d and es_d <= today < fc_d:
        return 'Pre-FC'
    if fc_d and cs_d and fc_d <= today < cs_d:
        return 'Post-FC'
    if cs_d and today >= cs_d:
        return 'Post-CS'
    return ''



def _fetch_dashboard_status_map() -> dict:
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {}
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT bu, target_name, sp_name, es_date, fc_date, cs_date FROM pdt_stats_dashboard.dashboard_status WHERE is_active=1")
        rows = cur.fetchall() or []
        out = {}
        def _add(key, info):
            if key:
                out[key.upper()] = info
        for r in rows:
            info = {
                'bu': r.get('bu'),
                'ES': r.get('es_date'),
                'FC': r.get('fc_date'),
                'CS': r.get('cs_date'),
            }
            sp  = str(r.get('sp_name')    or '').strip()
            tn  = str(r.get('target_name') or '').strip()
            # index exact + dot/underscore swaps + version-stripped variants
            import re as _re_idx
            for raw in [sp, tn]:
                if not raw:
                    continue
                _add(raw, info)
                _add(raw.replace('_', '.'), info)
                _add(raw.replace('.', '_'), info)
                # also index first token (e.g. SKYROS from SKYROS.LA)
                first = raw.replace('_', '.').split('.')[0]
                if first and first.upper() not in out:
                    _add(first, info)
                # index version-stripped variants so Kobuk.LE.1.1 -> Kobuk.LE.1 -> Kobuk.LE
                _s = raw
                while True:
                    _s2 = _re_idx.sub(r'[._]\d+$', '', _s)
                    if _s2 == _s or not _s2:
                        break
                    _s = _s2
                    _add(_s, info)
                    _add(_s.replace('_', '.'), info)
                    _add(_s.replace('.', '_'), info)
        return out
    except Exception:
        return {}
    finally:
        cur.close()
        conn.close()


def _match_dashboard(tgt: str, dash_map: dict) -> dict:
    """Match SP target to dashboard row using exact then fuzzy prefix token matching.
    Prefers longer/more-specific key when scores are equal to avoid IOT matching IOT_WEARABLES.
    """
    if not tgt:
        return {}
    # normalise: replace underscores with dots, uppercase
    def norm(s):
        return s.replace('_', '.').upper().strip()

    tgt_norm = norm(tgt)
    tgt_parts = [p for p in tgt_norm.split('.') if p]

    # Level 1: exact normalised match
    for raw in [tgt, tgt.replace('_', '.'), tgt.replace('.', '_')]:
        hit = dash_map.get(raw.upper().strip())
        if hit:
            return hit

    # Level 2: best prefix token overlap
    # Tiebreak: prefer the key whose normalised form is LONGER (more specific)
    # e.g. IOT_WEARABLES beats IOT when both share score=1
    best = {}
    best_score = 0
    best_key_len = 0
    for key, info in dash_map.items():
        key_norm = norm(key)
        key_parts = [p for p in key_norm.split('.') if p]
        # count matching tokens from start
        score = 0
        for a, b in zip(tgt_parts, key_parts):
            if a == b:
                score += 1
            else:
                break
        # also try reverse: key tokens vs tgt tokens
        score2 = 0
        for a, b in zip(key_parts, tgt_parts):
            if a == b:
                score2 += 1
            else:
                break
        final_score = max(score, score2)
        key_len = len(key_norm)
        # Prefer higher score; on tie prefer longer (more specific) key
        if final_score > best_score or (final_score == best_score and final_score > 0 and key_len > best_key_len):
            best_score = final_score
            best_key_len = key_len
            best = info

    return best if best_score > 0 else {}


def _match_dashboard_with_fallback(tgt: str, dash_map: dict) -> dict:
    """Like _match_dashboard but also tries stripping trailing version tokens.
    Aurora.LA.3.1 -> try Aurora.LA.3 -> Aurora.LA -> Aurora if no exact match.
    This ensures versioned targets inherit BU from their base target row.
    """
    import re as _re_mdf
    hit = _match_dashboard(tgt, dash_map)
    if hit and hit.get('bu'):
        return hit
    # Strip trailing numeric version tokens one at a time
    s = str(tgt or '').strip()
    while True:
        # Remove trailing .N or _N segment
        s2 = _re_mdf.sub(r'[._]\d+$', '', s)
        if s2 == s or not s2:
            break
        s = s2
        hit2 = _match_dashboard(s, dash_map)
        if hit2 and hit2.get('bu'):
            return hit2
    return hit or {}


def _fetch_sharepoint_known_targets(week_end: date | None = None) -> dict:
    known = {}
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return known
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT bu, target_name, sp_name, es_date, fc_date, cs_date FROM pdt_stats_dashboard.dashboard_status WHERE is_active=1")
        for r in cur.fetchall() or []:
            tn = str(r.get('target_name') or '').strip()
            sp = str(r.get('sp_name') or '').strip()
            name = sp or tn
            info = {'bu': _normalize_bu(str(r.get('bu') or '')), 'target': name, 'target_name': tn, 'sp_name': sp, 'es': _fmt_iso_date(r.get('es_date')), 'fc': _fmt_iso_date(r.get('fc_date')), 'cs': _fmt_iso_date(r.get('cs_date'))}
            for alias in (tn, sp, tn.replace('_', '.'), sp.replace('_', '.'), tn.replace('.', '_'), sp.replace('.', '_')):
                if alias and alias not in known:
                    known[alias] = info
        params = []
        where = ''
        if week_end:
            where = 'WHERE week_end <= %s'
            params.append(week_end.isoformat())
        cur.execute(f"""
            SELECT target, bu, timelines FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`
            {where} ORDER BY week_end DESC, updated_at DESC LIMIT 500
        """, tuple(params))
        for r in cur.fetchall() or []:
            name = str(r.get('target') or '').strip()
            if name and name not in known:
                ms = _extract_milestones_from_timelines(r.get('timelines'))
                known[name] = {'bu': str(r.get('bu') or ''), 'target': name, 'es': ms.get('ES') or '', 'fc': ms.get('FC') or '', 'cs': ms.get('CS') or ''}
        return known
    except Exception:
        return known
    finally:
        cur.close(); conn.close()


def _fetch_consolidate_target_info_map(before_week_end: date | None = None) -> dict:
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {}
    cur = conn.cursor(dictionary=True)
    try:
        params = []
        where = ''
        if before_week_end:
            where = 'WHERE week_end < %s'
            params.append(before_week_end.isoformat())
        cur.execute(f"""
            SELECT target, pl_id, bu, timelines, pdt_test_status FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`
            {where} ORDER BY week_end DESC, updated_at DESC LIMIT 1000
        """, tuple(params))
        out = {}
        for r in cur.fetchall() or []:
            tgt = str(r.get('target') or '').strip()
            pl_id = str(r.get('pl_id') or '').strip()
            pair_key = _sp_pair_key(tgt, pl_id)
            if tgt and pl_id and pair_key not in out:
                out[pair_key] = r
            if tgt and tgt not in out:
                out[tgt] = r
        return out
    except Exception:
        return {}
    finally:
        cur.close(); conn.close()


def _fetch_sp2_previous_bu_map(before_week_start: date | None = None) -> dict:
    """Latest previous Smart Build BU by target and target+PL.

    Used to carry a target's BU forward to the selected/current week when the
    current Smart Build snapshot row has no BU yet. Keys are upper-cased
    `(target, pl_id)` tuples plus `(target, '')` target-level fallbacks.
    """
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {}
    cur = conn.cursor(dictionary=True)
    try:
        params = []
        where = "WHERE COALESCE(bu, '') <> ''"
        if before_week_start:
            where += ' AND week_start < %s'
            params.append(before_week_start.isoformat())
        cur.execute(f"""
            SELECT target, pl_id, bu
            FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
            {where}
            ORDER BY week_start DESC, week_end DESC, updated_at DESC
            LIMIT 2000
        """, tuple(params))
        out = {}
        for r in cur.fetchall() or []:
            tgt = str(r.get('target') or '').strip().upper()
            pl_id = str(r.get('pl_id') or '').strip().upper()
            bu = _normalize_bu(str(r.get('bu') or '').strip())
            if not tgt or not bu:
                continue
            if pl_id and (tgt, pl_id) not in out:
                out[(tgt, pl_id)] = bu
            if (tgt, '') not in out:
                out[(tgt, '')] = bu
        return out
    except Exception:
        return {}
    finally:
        cur.close(); conn.close()


def _ucr_match_text(value: str) -> str:
    """Normalize Unique CR/Sharepoint Target and PL-ID values for matching.

    Sharepoint and Unique CR workbooks may label/format the same PL identifier
    differently (for example ``PL-ID`` vs ``PL ID`` headers, mixed case, or
    embedded spaces). The comparison must therefore be done on a compact,
    case-insensitive value while preserving meaningful separators such as dots.
    """
    text = str(value or '').strip().upper().replace('_', '.')
    return _re.sub(r'\s+', '', text)


def _ucr_target_aliases(target: str) -> list:
    """Return exact and codename target aliases used to match Sharepoint rows."""
    exact = _ucr_match_text(target)
    if not exact:
        return []
    first = _re.split(r'[._]', exact)[0]
    aliases = [exact]
    if first and first != exact:
        aliases.append(first)
    return aliases


def _ucr_count_value(row: dict) -> int:
    """Return a positive row count fallback when CRID is blank."""
    val = _safe_int(row.get('CR Count'))
    return val if val and val > 0 else 1


def _refresh_ucr_excel_from_latest_csv_if_needed(week_end: date) -> dict:
    """Refresh generated Unique CR Excel from newest matching source CSV if needed.

    Used by landing, Smart Build, and Consolidate so Unique CR counts always use
    the latest regenerated UNIQUECRSREPORT_WEEKENDING source for the week.
    """
    info = {'source_path': '', 'excel_path': _ucr_excel_path(week_end) if week_end else '', 'refreshed': False}
    if not week_end:
        return info
    try:
        src_path = _find_ucr_file_by_week_end(week_end.isoformat(), force=True)
        info['source_path'] = src_path
        xl_path = _ucr_excel_path(week_end)
        src_mtime = os.path.getmtime(src_path) if src_path and os.path.isfile(src_path) else 0
        xl_mtime = os.path.getmtime(xl_path) if os.path.isfile(xl_path) else 0
        if src_path and (not os.path.isfile(xl_path) or src_mtime > xl_mtime):
            refresh_info = _ensure_ucr_excel_for_week(week_end, _load_farm_station_map(), force_refresh=True)
            info['refreshed'] = bool(refresh_info.get('success'))
            info['excel_path'] = refresh_info.get('path') or xl_path
    except Exception as exc:
        info['error'] = str(exc)
    return info



def _build_ucr_target_pl_summary_rows(week_end: date) -> list:
    """Read generated Unique CR Excel and return actual Target+PL-ID counts.

    Rows from UNIQUECRSREPORT_WEEKENDING_* contain Target and PL-ID.  After the
    saved Unique_CRs workbook is generated, this reads those values from the
    workbook and prepares rows that can be saved into the consolidate table even
    when a Target+PL-ID has no Sharepoint build summary yet.
    """
    path = _ucr_excel_path(week_end) if week_end else ''
    if not path or not os.path.isfile(path):
        return []

    display_values = {}
    cr_sets = {}
    count_fallbacks = Counter()
    farm_map = _load_farm_station_map()
    for site in ('QIPL',):  # Consolidated Report: QIPL only
        payload  = _read_ucr_excel_sheet(path, site, farm_map)

        for row in payload.get('rows') or []:
            target = str(row.get('Target') or '').strip()
            pl_id = str(row.get('PL-ID') or '').strip()
            tgt_key = _ucr_match_text(target)
            pl_key = _ucr_match_text(pl_id)
            if not tgt_key or not pl_key:
                continue
            key = (tgt_key, pl_key)
            display_values.setdefault(key, {'target': target, 'pl_id': pl_id})
            crid = _ucr_match_text(row.get('CRID'))
            if crid:
                cr_sets.setdefault(key, set()).add(crid)
            else:
                count_fallbacks[key] += _ucr_count_value(row)

    summaries = []
    for key, display in display_values.items():
        summaries.append({
            'target': display.get('target') or '',
            'pl_id': display.get('pl_id') or '',
            'unique_crs': len(cr_sets.get(key, set())) + int(count_fallbacks.get(key, 0) or 0),
        })
    return sorted(summaries, key=lambda r: (str(r.get('target') or '').lower(), str(r.get('pl_id') or '').lower()))





def _build_ucr_target_pl_count_map(week_end: date) -> dict:

    """
    Build Target+PL-ID -> Unique CR count from the selected Unique CR Excel.

    This is the source for the Sharepoint Consolidated Report "Unique CRs"
    column. Counts are grouped from /weekly-report/card/unique_report data by
    Target + PL-ID, and rows without PL-ID are intentionally ignored so a
    consolidated row displays '-' instead of an incorrect target-only count.
    """
    if not week_end:
        return {}
    path = _ucr_excel_path(week_end)
    if not path or not os.path.isfile(path):
        return {}

    wb = None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        cr_sets = {}
        count_fallbacks = Counter()

        for site in ('QIPL',):  # Consolidated Report: QIPL only
            if site not in wb.sheetnames:
                continue

            ws = wb[site]
            if ws.max_row < 2:
                continue
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            hdrs = [str(v or '').strip() for v in header_row]
            last_merged_values = {'Target': '', 'PL-ID': ''}

            for vals in ws.iter_rows(min_row=2, values_only=True):
                if not any(v for v in vals if v is not None and str(v).strip()):
                    continue
                raw_row = {}
                for ci, h in enumerate(hdrs):
                    v = vals[ci] if ci < len(vals) else None
                    if isinstance(v, datetime):
                        v = v.date().isoformat()
                    elif v is not None:
                        v = str(v).strip()
                    raw_row[h] = v or ''

                def _v(*names):
                    for name in names:
                        if raw_row.get(name):
                            return raw_row.get(name)
                    return ''

                row = {
                    'Target':   _v('Target'),
                    'PL-ID':    _v('PL-ID', 'PL ID', 'PLID', 'Target PL', 'Target.PL', 'TargetPL'),
                    'CRID':     _v('CRID'),
                    'CR Count': _v('CR Count', 'CR Instances'),
                }
                for merge_key in ('Target', 'PL-ID'):
                    if row.get(merge_key):
                        last_merged_values[merge_key] = row[merge_key]
                    elif last_merged_values.get(merge_key):
                        row[merge_key] = last_merged_values[merge_key]
                if not _is_ucr_data_row(row):
                    continue

                pl_key = _ucr_match_text(row.get('PL-ID'))
                if not pl_key:
                    continue
                crid = _ucr_match_text(row.get('CRID'))
                for tgt_key in _ucr_target_aliases(row.get('Target')):
                    key = (tgt_key, pl_key)
                    if crid:
                        cr_sets.setdefault(key, set()).add(crid)
                    else:
                        count_fallbacks[key] += _ucr_count_value(row)

        counts = {key: len(values) for key, values in cr_sets.items()}
        for key, value in count_fallbacks.items():
            counts[key] = counts.get(key, 0) + int(value or 0)
        return counts
    except Exception:
        return {}
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _ucr_count_for_sharepoint_pair(counts_by_pair: dict, target: str, pl_id: str):
    """Lookup QIPL Unique CR count for a Sharepoint/Smart Build PL row.

    Prefer Target+PL match. If target naming differs between Smart Build and the
    Unique CR CSV, fall back to PL-ID-only so the Consolidate Unique CR column is
    still populated w.r.t. PL.
    """
    pl_key = _ucr_match_text(pl_id)
    if not counts_by_pair or not pl_key:
        return None
    for tgt_key in _ucr_target_aliases(target):
        key = (tgt_key, pl_key)
        if key in counts_by_pair:
            return int(counts_by_pair[key] or 0)
    pl_matches = [int(v or 0) for (tgt_key, pair_pl), v in counts_by_pair.items() if pair_pl == pl_key]
    return max(pl_matches) if pl_matches else None


_FORCED_CONSOLIDATE_MILESTONES_BY_PL = {
    'SA8797P.ADAS.HGY.5.1.7.0': {'bu': 'AUTO', 'es': '2024-11-25', 'fc': '2026-03-31', 'cs': '2026-06-30'},
    'SA8797P.FLEX.HQX.5.7.7.0': {'bu': 'AUTO', 'es': '2024-11-25', 'fc': '2026-03-31', 'cs': '2026-06-30'},
    'SNAPDRAGON_AUTO.HQX.4.8.9.0.1.r1': {'bu': 'AUTO', 'es': '2023-03-31', 'fc': '2024-10-31', 'cs': '2024-09-05'},
    'SA525M.LE.3.0': {'bu': 'AUTO_TELEMATICS', 'es': '2022-11-13', 'fc': '2023-07-31', 'cs': '2023-10-30'},
    # IOT targets Ã¢Â€Â” milestones confirmed from OneView Milestones modal
    'QCM6490.LE.1.0.r1': {'bu': 'IOT', 'es': '2023-12-24', 'fc': '', 'cs': '2024-06-24'},
    'QCM6490.LE.1.0':    {'bu': 'IOT', 'es': '2023-12-24', 'fc': '', 'cs': '2024-06-24'},
    'QCS8300.LE.2.0':    {'bu': 'IOT', 'es': '2026-02-20', 'fc': '2026-04-24', 'cs': '2026-06-26'},
    'QCS9100.LE.2.0':    {'bu': 'IOT', 'es': '2026-02-20', 'fc': '2026-04-24', 'cs': '2026-06-26'},
    'SW5100.LW.5.1':     {'bu': 'IOT', 'es': '2025-09-05', 'fc': '2025-10-28', 'cs': '2025-12-27'},
}


def _backfill_sharepoint_crashes(week_start: date | None = None, week_end: date | None = None) -> dict:
    """Re-calculate crash_count + crash_details from live weekly_qipl_data for every
    saved sharepoint build row.  Hours and devices are NOT touched.
    MTBF is recalculated as hours / crash_count.
    Uses Stability Ticket dedup so each unique crash event is counted once.
    """
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {'success': False, 'updated': 0, 'checked': 0, 'message': 'DB connection failed'}
    cur = conn.cursor(dictionary=True)
    updated = 0
    checked = 0
    try:
        # -- 1. Load all saved build rows for the week -------------------------
        sql = f"SELECT * FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`"
        params = []
        if week_start and week_end:
            sql += " WHERE week_start=%s AND week_end=%s"
            params.extend([week_start.isoformat(), week_end.isoformat()])
        sql += " ORDER BY id"
        cur.execute(sql, tuple(params))
        saved_rows = cur.fetchall() or []

        # -- 2. For each saved row, find its builds and count crashes ----------
        for rec in saved_rows:
            checked += 1
            rec_target = str(rec.get('target') or '').strip()
            rec_pl     = str(rec.get('pl_id')  or '').strip()
            hours      = float(rec.get('hours') or 0)

            # Collect all build IDs referenced by this saved row
            build_ids = set()
            jira_reporters = set()
            top_build = str(rec.get('meta_build') or '').strip()
            if top_build:
                build_ids.add(top_build.upper())
            try:
                sitems = json.loads(rec.get('selected_items_json') or '[]')
            except Exception:
                sitems = []
            for si in (sitems or []):
                if not isinstance(si, dict):
                    continue
                sb = str(si.get('meta_build') or '').strip()
                if sb:
                    build_ids.add(sb.upper())
                for jr in (si.get('jira_reporters') or []):
                    if str(jr or '').strip():
                        jira_reporters.add(str(jr).strip())

            if not build_ids:
                continue

            # -- 3/4. Query weekly_qipl_data and count like the build-wise pivot
            pivot = _count_sharepoint_crashes_from_weekly_qipl(
                cur, rec_target, rec_pl, build_ids,
                week_start=rec.get('week_start'), week_end=rec.get('week_end'),
                jira_reporters=sorted(jira_reporters)
            )

            new_crashes  = pivot['total']
            new_details  = pivot['details']
            new_mtbf     = round(hours / new_crashes, 2) if new_crashes else 0.0

                        # -- 5. Update if count/details/MTBF changed -----------------------
            old_crashes = int(float(rec.get('crash_count') or 0) or 0)
            old_details = str(rec.get('crash_details') or '')
            old_mtbf = round(float(rec.get('mtbf') or 0), 2)
            if new_crashes == old_crashes and new_details == old_details and new_mtbf == old_mtbf:
                continue

            cur.execute(

                f"UPDATE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`"
                f" SET crash_count=%s, crash_details=%s, mtbf=%s"
                f" WHERE id=%s",
                (new_crashes, new_details, new_mtbf, rec.get('id'))
            )
            updated += 1

        conn.commit()
        return {
            'success': True,
            'updated': updated,
            'checked': checked,
            'message': f'Crashes backfilled: {updated} row(s) updated out of {checked} checked.',
        }
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        return {'success': False, 'updated': updated, 'checked': checked, 'message': str(exc)}
    finally:
        cur.close()
        conn.close()


def _backfill_sharepoint_build_milestones(week_start: date | None = None, week_end: date | None = None) -> dict:
    """Backfill weekly_sharepoint_build_summary BU + ES/FC/CS using target/SP lookup."""
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {'success': False, 'updated': 0, 'checked': 0, 'message': 'DB connection failed'}
    cur = conn.cursor(dictionary=True)
    updated = 0
    checked = 0
    try:
        sql = f"SELECT * FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` WHERE ((es_date IS NULL AND fc_date IS NULL AND cs_date IS NULL) OR bu IS NULL OR TRIM(bu)='')"
        params = []
        if week_start and week_end:
            sql += " AND week_start=%s AND week_end=%s"
            params.extend([week_start.isoformat(), week_end.isoformat()])
        sql += " ORDER BY id"
        cur.execute(sql, tuple(params))
        rows = cur.fetchall() or []
        for rec in rows:
            checked += 1
            target = str(rec.get('target') or '').strip()
            try:
                selected_items = json.loads(rec.get('selected_items_json') or '[]')
                if not isinstance(selected_items, list):
                    selected_items = []
            except Exception:
                selected_items = []
            manual = next((it for it in selected_items if isinstance(it, dict) and it.get('manual_no_crash')), {})
            info = _resolve_sharepoint_build_milestones(target, rec={
                'target': target,
                'bu': rec.get('bu'),
                'es_date': rec.get('es_date'),
                'fc_date': rec.get('fc_date'),
                'cs_date': rec.get('cs_date'),
                'sp_name': target,
            }, manual=manual)
            if info.get('bu') or info.get('es') or info.get('fc') or info.get('cs'):
                cur.execute(
                    f"UPDATE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` SET bu=%s, es_date=%s, fc_date=%s, cs_date=%s WHERE id=%s",
                    (info.get('bu') or rec.get('bu') or None, info.get('es') or None, info.get('fc') or None, info.get('cs') or None, rec.get('id'))
                )
                updated += 1
        conn.commit()
        return {'success': True, 'updated': updated, 'checked': checked, 'message': f'Updated {updated} row(s) out of {checked}'}
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        return {'success': False, 'updated': updated, 'checked': checked, 'message': str(exc)}
    finally:
        cur.close(); conn.close()


def _fetch_sharepoint_row_milestones_backfill(week_end: date | None = None) -> dict:
    """Fill consolidate timelines from dashboard_status using target, target_name or sp_name."""
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {'success': False, 'updated': 0, 'checked': 0, 'message': 'DB connection failed'}
    cur = conn.cursor(dictionary=True)
    updated = 0
    checked = 0
    try:
        sql = f"SELECT * FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`"
        params = []
        if week_end:
            sql += " WHERE week_end=%s"
            params.append(week_end.isoformat())
        sql += " ORDER BY id"
        cur.execute(sql, tuple(params))
        rows = cur.fetchall() or []

        cur.execute("""
            SELECT bu, target_name, sp_name, es_date, fc_date, cs_date
            FROM pdt_stats_dashboard.dashboard_status
            WHERE is_active=1
            ORDER BY id ASC
        """)
        dash_rows = cur.fetchall() or []

        def norm(val):
            return str(val or '').strip().upper().replace('_', '.').replace(' ', '')

        dash_exact = {}
        dash_rows_norm = []
        for dr in dash_rows:
            info = {
                'bu': str(dr.get('bu') or '').strip(),
                'es': _fmt_iso_date(dr.get('es_date')),
                'fc': _fmt_iso_date(dr.get('fc_date')),
                'cs': _fmt_iso_date(dr.get('cs_date')),
                'target_name': str(dr.get('target_name') or '').strip(),
                'sp_name': str(dr.get('sp_name') or '').strip(),
            }
            for key in (info['target_name'], info['sp_name']):
                nk = norm(key)
                if nk:
                    dash_exact[nk] = info
            dash_rows_norm.append((info, norm(info['target_name']), norm(info['sp_name'])))

        for rec in rows:
            checked += 1
            target = str(rec.get('target') or '').strip()
            pl_id = str(rec.get('pl_id') or '').strip()
            # Build a case-insensitive lookup: try pl_id first, then target
            _forced_upper = {k.upper(): v for k, v in _FORCED_CONSOLIDATE_MILESTONES_BY_PL.items()}
            forced = _forced_upper.get(pl_id.upper()) or _forced_upper.get(target.upper())
            if forced:
                es = forced.get('es') or ''
                fc = forced.get('fc') or ''
                cs = forced.get('cs') or ''
                bu = str(forced.get('bu') or '').strip()
                timelines = _sp_timeline(es, fc, cs)
                status = _compute_pdt_test_status(es, cs, fc)
                cur.execute(
                    f"UPDATE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` SET timelines=%s, pdt_test_status=%s, bu=CASE WHEN bu IS NULL OR TRIM(bu)='' THEN %s ELSE bu END WHERE id=%s",
                    (timelines, status, bu, rec.get('id'))
                )
                updated += 1
                continue
            nk = norm(target)
            info = dash_exact.get(nk)
            if not info and nk:
                target_token = nk.split('.')[0]
                for item, tn_norm, sp_norm in dash_rows_norm:
                    if nk == tn_norm or nk == sp_norm:
                        info = item; break
                    if target_token and (tn_norm.startswith(target_token) or sp_norm.startswith(target_token) or target_token.startswith(tn_norm.split('.')[0] if tn_norm else '') or target_token.startswith(sp_norm.split('.')[0] if sp_norm else '')):
                        info = item; break
            if not info:
                continue
            es = info.get('es') or ''
            fc = info.get('fc') or ''
            cs = info.get('cs') or ''
            bu = str(info.get('bu') or '').strip()
            timelines = _sp_timeline(es, fc, cs)
            status = _compute_pdt_test_status(es, cs, fc)
            cur.execute(
                f"UPDATE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` SET timelines=%s, pdt_test_status=%s, bu=CASE WHEN bu IS NULL OR TRIM(bu)='' THEN %s ELSE bu END WHERE id=%s",
                (timelines, status, bu, rec.get('id'))
            )
            updated += 1
        conn.commit()
        return {'success': True, 'updated': updated, 'checked': checked, 'message': f'Updated {updated} row(s) out of {checked}'}
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        return {'success': False, 'updated': updated, 'checked': checked, 'message': str(exc)}
    finally:
        cur.close(); conn.close()


def _build_and_save_consolidate_summary(week_end: date, username: str) -> list:
    week_start = week_end - timedelta(days=6)
    # Ensure the saved Unique_CRs workbook is refreshed from the newest matching
    # UNIQUECRSREPORT_WEEKENDING_* source first, preserving Target and PL-ID.
    _refresh_ucr_excel_from_latest_csv_if_needed(week_end)
    ucr_counts_by_pair = _build_ucr_target_pl_count_map(week_end)
    ucr_summary_rows = _build_ucr_target_pl_summary_rows(week_end)
    saved = _fetch_sharepoint_summaries(week_start, week_end)


    dash_map = _fetch_dashboard_status_map()
    previous_target_info = _fetch_consolidate_target_info_map(before_week_end=week_end)
    grouped = {}
    for rec in saved:
        tgt = str(rec.get('target') or '').strip()
        pl_id = str(rec.get('pl_id') or '').strip()
        manual = _sp_manual_meta(rec)
        prev = previous_target_info.get(_sp_pair_key(tgt, pl_id)) or previous_target_info.get(tgt) or {}
        milestone_info = _resolve_sharepoint_build_milestones(tgt, rec={**rec, 'sp_name': tgt}, manual=manual)
        bu = str(milestone_info.get('bu') or prev.get('bu') or '').strip()
        if milestone_info.get('es') or milestone_info.get('fc') or milestone_info.get('cs'):
            timelines = _sp_timeline(milestone_info.get('es'), milestone_info.get('fc'), milestone_info.get('cs'))
            status = _compute_pdt_test_status(milestone_info.get('es'), milestone_info.get('cs'), milestone_info.get('fc'))
        else:
            timelines = str(prev.get('timelines') or '')
            status = str(prev.get('pdt_test_status') or '')

        grp = grouped.setdefault((tgt, pl_id), {
            'week_end': week_end.isoformat(),
            'bu': bu,
            'target': tgt,
            'pl_id': pl_id,
            'timelines': timelines,
            'pdt_test_status': status,
            'number_of_devices': 0,
            'number_of_builds': 0,
            '_crm_max_label': 0,
            'total_hours': 0.0,
            'total_crashes': 0,
            'unique_crs': None,
            'mtbf': 0.0,
        })
        if not str(grp.get('pl_id') or '').strip() and pl_id:
            grp['pl_id'] = pl_id
        if not grp.get('bu') and bu:
            grp['bu'] = bu
        if not grp.get('timelines') and timelines:
            grp['timelines'] = timelines
        if not grp.get('pdt_test_status') and status:
            grp['pdt_test_status'] = status
        # number_of_builds = max build_label for CRM rows (avoids double-counting)
        _rec_build_type = str(rec.get('build_type') or 'CRM').strip().upper()
        if _rec_build_type == 'CRM':
            try:
                _lbl = int(str(rec.get('build_label') or '0').strip() or 0)
            except Exception:
                _lbl = 0
            if _lbl > grp.get('_crm_max_label', 0):
                grp['_crm_max_label'] = _lbl
        grp['number_of_devices'] += int(float(rec.get('devices') or 0) or 0)


        grp['total_hours'] += float(rec.get('hours') or 0)  # hours already week-bounded from sharepoint save
        grp['total_crashes'] += int(float(rec.get('crash_count') or 0) or 0)

    # Set number_of_builds = max CRM build_label (min 1 if any builds exist)
    for _grp in grouped.values():
        _max = _grp.pop('_crm_max_label', 0)
        if _max > 0:
            _grp['number_of_builds'] = _max
        # if no CRM rows found, keep existing count (ENG only targets)

    for ucr_row in ucr_summary_rows:
        tgt = str(ucr_row.get('target') or '').strip()
        pl_id = str(ucr_row.get('pl_id') or '').strip()
        if not tgt or not pl_id:
            continue
        grp = grouped.get((tgt, pl_id))
        if grp is None:
            prev = previous_target_info.get(_sp_pair_key(tgt, pl_id)) or previous_target_info.get(tgt) or {}
            milestone_info = _resolve_sharepoint_build_milestones(tgt, rec={'target': tgt, 'pl_id': pl_id, 'sp_name': tgt}, manual={})
            bu = str(milestone_info.get('bu') or prev.get('bu') or '').strip()
            if milestone_info.get('es') or milestone_info.get('fc') or milestone_info.get('cs'):
                timelines = _sp_timeline(milestone_info.get('es'), milestone_info.get('fc'), milestone_info.get('cs'))
                status = _compute_pdt_test_status(milestone_info.get('es'), milestone_info.get('cs'), milestone_info.get('fc'))
            else:
                timelines = str(prev.get('timelines') or '')
                status = str(prev.get('pdt_test_status') or '')

            grp = {

                'week_end': week_end.isoformat(),
                'bu': bu,
                'target': tgt,
                'pl_id': pl_id,
                'timelines': timelines,
                'pdt_test_status': status,
                'number_of_devices': 0,
                'number_of_builds': 0,
                'total_hours': 0.0,
                'total_crashes': 0,
                'unique_crs': None,
                'mtbf': 0.0,
            }
            grouped[(tgt, pl_id)] = grp
        grp['unique_crs'] = int(ucr_row.get('unique_crs') or 0)

    rows = sorted(
        [r for r in grouped.values()
         if float(r.get('total_hours') or 0) > 0
         or int(r.get('number_of_devices') or 0) > 0
         or int(r.get('number_of_builds') or 0) > 0],
        key=lambda x: (str(x.get('bu') or '').lower(), str(x.get('target') or '').lower(), str(x.get('pl_id') or '').lower())
    )


    for row in rows:
        row['mtbf'] = round((row['total_hours'] / row['total_crashes']), 2) if row['total_crashes'] else 0
        # Fetch Unique CR count from the selected Unique CR Report Excel by Target + PL-ID.
        # If a row came only from the Unique CR workbook, keep that workbook count.
        matched_ucr = _ucr_count_for_sharepoint_pair(ucr_counts_by_pair, row.get('target'), row.get('pl_id'))

        row['unique_crs'] = matched_ucr if matched_ucr is not None else row.get('unique_crs')
    conn = get_mysql_connection_db(bu_key=None)
    if conn:
        cur = conn.cursor()
        try:
            # Snapshot devices_count before DELETE so Refresh never wipes manual values
            _dc_snapshot = {}
            try:
                cur.execute(f"SELECT target, devices_count FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` WHERE week_end=%s AND devices_count IS NOT NULL", (week_end.isoformat(),))
                for _r in (cur.fetchall() or []):
                    _tgt = str(_r[0] or '').strip()
                    if _tgt: _dc_snapshot[_tgt] = _r[1]
            except Exception:
                _dc_snapshot = {}
            cur.execute(f"DELETE FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` WHERE week_end=%s", (week_end.isoformat(),))
            for row in rows:
                cur.execute(f"""
                    INSERT INTO `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`
                    (week_end, bu, target, pl_id, timelines, pdt_test_status, number_of_devices, number_of_builds, total_hours, total_crashes, unique_crs, mtbf, updated_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (week_end.isoformat(), row['bu'], row['target'], row.get('pl_id') or '', row['timelines'], row['pdt_test_status'], row['number_of_devices'], row['number_of_builds'], row['total_hours'], row['total_crashes'], row['unique_crs'], row['mtbf'], row.get('updated_by') or username))
            # Restore devices_count for all PLs under each snapshotted target
            if _dc_snapshot:
                for _tgt, _dc in _dc_snapshot.items():
                    cur.execute(f"UPDATE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` SET devices_count=%s WHERE week_end=%s AND target=%s", (_dc, week_end.isoformat(), _tgt))

            # Hours-proportional device split per PL.
            # For each target that has a devices_count set, distribute devices
            # across its PL rows proportionally to each PL's total_hours.
            # This ensures: SUM(number_of_devices per PL) == devices_count (target total)
            # which keeps Consolidate Report, Device Utilization, and Devices Tab in sync.
            try:
                cur.execute(
                    f"SELECT id, target, pl_id, total_hours, devices_count"
                    f" FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`"
                    f" WHERE week_end=%s AND devices_count IS NOT NULL AND devices_count > 0",
                    (week_end.isoformat(),)
                )
                _split_rows = cur.fetchall() or []
                # Group by target
                _tgt_groups = {}
                for _sr in _split_rows:
                    _t = str(_sr[1] or '').strip()
                    _tgt_groups.setdefault(_t, []).append(_sr)
                for _tgt, _pls in _tgt_groups.items():
                    _dc_total = int(_pls[0][4] or 0)  # devices_count same for all PLs
                    _hrs_total = sum(float(_p[3] or 0) for _p in _pls)
                    _assigned = 0
                    _sorted_pls = sorted(_pls, key=lambda x: float(x[3] or 0), reverse=True)
                    for _idx, _pl in enumerate(_sorted_pls):
                        _pl_hrs = float(_pl[3] or 0)
                        if _idx == len(_sorted_pls) - 1:
                            # Last PL gets the remainder to ensure exact total
                            _pl_dev = _dc_total - _assigned
                        elif _hrs_total > 0:
                            _pl_dev = round(_dc_total * _pl_hrs / _hrs_total)
                        else:
                            # No hours data: split evenly
                            _pl_dev = round(_dc_total / len(_sorted_pls))
                        _pl_dev = max(0, _pl_dev)
                        _assigned += _pl_dev
                        cur.execute(
                            f"UPDATE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`"
                            f" SET number_of_devices=%s WHERE id=%s",
                            (_pl_dev, _pl[0])
                        )
            except Exception:
                pass
            conn.commit()
        finally:
            cur.close()
            conn.close()
    # Save static JSON snapshot so Consolidate Report is frozen for this week
    _save_consolidate_json(week_end, rows)
    return rows


def _fetch_consolidate_summary(week_end: date) -> list:
    """Load consolidate rows: JSON snapshot first (static), fall back to DB."""
    # Try static JSON snapshot first â€” preserves data even if DB changes
    json_rows = _load_consolidate_json(week_end)
    if json_rows:
        return json_rows
    # Fall back to DB
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT * FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`
            WHERE week_end=%s
            ORDER BY COALESCE(bu,""), COALESCE(target,""), COALESCE(pl_id,"")
        """, (week_end.isoformat(),))
        rows = cur.fetchall() or []
        for r in rows:
            for k, v in list(r.items()):
                if isinstance(v, (date, datetime)):
                    r[k] = v.isoformat()[:10]
                elif hasattr(v, "__float__"):
                    try:
                        r[k] = float(v)
                    except Exception:
                        pass
        return rows
    except Exception:
        return []
    finally:
        cur.close(); conn.close()


def _consolidate_json_path(week_end) -> str:
    """Return path to the per-week consolidate JSON snapshot."""
    we = week_end.isoformat() if hasattr(week_end, "isoformat") else str(week_end)
    fname = f"consolidate_{we}.json"
    # Try network share first
    net = _CONSOLIDATE_JSON_NET
    if os.path.isdir(net):
        return os.path.join(net, fname)
    # Fall back to local
    local = _CONSOLIDATE_JSON_LOCAL
    os.makedirs(local, exist_ok=True)
    return os.path.join(local, fname)


def _save_consolidate_json(week_end, rows: list) -> None:
    """Save consolidate rows as a static JSON snapshot for this week."""
    path = _consolidate_json_path(week_end)
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump({"week_end": str(week_end), "rows": rows,
                       "saved_at": datetime.utcnow().isoformat() + "Z"}, fh, indent=2, default=str)
        os.replace(tmp, path)
    except Exception as _e:
        import logging as _log
        _log.getLogger("weekly_summary_routes").warning("[CONSOLIDATE JSON] save failed: %s", _e)


def _load_consolidate_json(week_end) -> list:
    """Load consolidate rows from static JSON snapshot. Returns [] if not found."""
    path = _consolidate_json_path(week_end)
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data.get("rows") or []
    except Exception:
        return []


def _current_user_identifier() -> str:
    for attr in ('username', 'email', 'id'):
        try:
            val = getattr(current_user, attr, None)
        except Exception:
            val = None
        if val is not None and str(val).strip():
            return str(val).strip()
    return ''


def _backfill_sharepoint_bu(week_start: date | None = None, week_end: date | None = None) -> dict:
    """Backfill NULL/blank BU in weekly_sharepoint_build_summary using available data."""
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {'success': False, 'updated': 0, 'checked': 0, 'message': 'DB connection failed'}
    cur = conn.cursor(dictionary=True)
    updated = 0
    checked = 0
    try:
        sql = f"SELECT * FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` WHERE (bu IS NULL OR TRIM(bu)='')"
        params = []
        if week_start and week_end:
            sql += " AND week_start=%s AND week_end=%s"
            params.extend([week_start.isoformat(), week_end.isoformat()])
        sql += " ORDER BY id"
        cur.execute(sql, tuple(params))
        rows = cur.fetchall() or []
        for rec in rows:
            checked += 1
            target = str(rec.get('target') or '').strip()
            fill_bu = ''
            if target:
                dash = _find_dashboard_target_info(target)
                fill_bu = str(dash.get('bu') or '').strip()
            if not fill_bu:
                try:
                    selected_items = json.loads(rec.get('selected_items_json') or '[]')
                except Exception:
                    selected_items = []
                for item in selected_items:
                    if isinstance(item, dict) and str(item.get('bu') or '').strip():
                        fill_bu = str(item.get('bu') or '').strip()
                        break
            if fill_bu:
                cur.execute(
                    f"UPDATE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` SET bu=%s WHERE id=%s",
                    (fill_bu, rec.get('id'))
                )
                updated += 1
        conn.commit()
        return {'success': True, 'updated': updated, 'checked': checked, 'message': f'Updated {updated} row(s)'}
    except Exception as exc:
        try:
            conn.rollback()
        except Exception:
            pass
        return {'success': False, 'updated': updated, 'checked': checked, 'message': str(exc)}
    finally:
        cur.close(); conn.close()


def _fetch_sharepoint_summaries(week_start: date, week_end: date) -> list:
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT * FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`
            WHERE week_end=%s
            ORDER BY COALESCE(bu,''), target, pl_id, build_type, id
        """, (week_end.isoformat(),))
        out = []
        for r in cur.fetchall() or []:
            try:
                r['selected_items'] = json.loads(r.get('selected_items_json') or '[]')
            except Exception:
                r['selected_items'] = []
            if not r.get('meta_build'):
                for it in r.get('selected_items') or []:
                    if isinstance(it, dict) and it.get('meta_build'):
                        r['meta_build'] = it.get('meta_build')
                        break
            if not str(r.get('bu') or '').strip():
                target = str(r.get('target') or '').strip()
                dash = _find_dashboard_target_info(target) if target else {}
                if dash and str(dash.get('bu') or '').strip():
                    r['bu'] = str(dash.get('bu') or '').strip()
                else:
                    manual_item = next((it for it in (r.get('selected_items') or []) if isinstance(it, dict) and str(it.get('bu') or '').strip()), {})
                    if manual_item:
                        r['bu'] = str(manual_item.get('bu') or '').strip()
            for k, v in list(r.items()):
                if isinstance(v, (date, datetime)):
                    r[k] = v.isoformat()[:10]
                elif hasattr(v, '__float__'):
                    try:
                        r[k] = float(v)
                    except Exception:
                        pass
            out.append(r)
        return out
    except Exception:
        return []
    finally:
        cur.close(); conn.close()


def _week_ranges_for_templates() -> list:
    ranges = _merge_week_ranges(_get_available_weeks(), _get_week_ranges(20), limit=52)
    return [(s.isoformat(), e.isoformat(), f"{s.strftime('%b %d')} - {e.strftime('%b %d, %Y')}") for s, e in ranges]


def _selected_week_from_request():
    ws = _safe_date(request.args.get('week_start'))
    # Keep Unique CR Report's dropdown isolated. ucr_week_end must not become the
    # global weekly-report week used by CR Age, CR Pie, Sharepoint, landing, etc.
    we = _safe_date(request.args.get('week_end'))
    if we and not ws:
        ws = we - timedelta(days=6)
    if not ws or not we:
        ranges = _merge_week_ranges(_get_available_weeks(), _get_week_ranges(20), limit=52)
        if ranges:
            ws, we = ranges[0]
        else:
            ws, we = current_monday_sunday()
    return ws, we


# ---------------------------------------------------------------------------
# HWPDT MSM Screening table for weekly report
# ---------------------------------------------------------------------------
def _get_hwpdt_msm_config(week_start, week_end, target_names):
    """
    Load MSM config for a week from DB.
    If no row exists for a target this week, auto-copy from the most recent
    previous week. Returns dict: target_name -> {idp_farm, idp_deployed, parts_planned, supply_note}
    """
    from src.utils import get_mysql_connection_db
    conn = get_mysql_connection_db(bu_key=None)
    cur  = conn.cursor(dictionary=True)
    try:
        # Load this week's config
        fmt = ','.join(['%s'] * len(target_names))
        cur.execute(
            f"SELECT * FROM pdt_stats_dashboard.hwpdt_msm_config "
            f"WHERE week_start=%s AND target_name IN ({fmt})",
            [week_start.isoformat()] + list(target_names)
        )
        this_week = {r['target_name']: r for r in (cur.fetchall() or [])}

        # For targets missing this week -> find most recent previous week
        missing = [t for t in target_names if t not in this_week]
        prev_config = {}
        if missing:
            fmt2 = ','.join(['%s'] * len(missing))
            cur.execute(
                f"""SELECT c.* FROM pdt_stats_dashboard.hwpdt_msm_config c
                    INNER JOIN (
                        SELECT target_name, MAX(week_start) AS max_ws
                        FROM pdt_stats_dashboard.hwpdt_msm_config
                        WHERE week_start < %s AND target_name IN ({fmt2})
                        GROUP BY target_name
                    ) latest ON c.target_name=latest.target_name AND c.week_start=latest.max_ws""",
                [week_start.isoformat()] + missing
            )
            prev_config = {r['target_name']: r for r in (cur.fetchall() or [])}

        result = {}
        for t in target_names:
            if t in this_week:
                result[t] = this_week[t]
            elif t in prev_config:
                # Copy from previous week (don't save yet â€” save on first edit)
                result[t] = dict(prev_config[t])
                result[t]['week_start'] = week_start.isoformat()
                result[t]['week_end']   = week_end.isoformat()
                result[t]['_from_prev'] = True
            else:
                result[t] = {'idp_farm': None, 'idp_deployed': None,
                             'parts_planned': None, 'supply_note': ''}
        return result
    finally:
        cur.close(); conn.close()


def _build_hwpdt_msm_table(sel_start, sel_end):
    """
    Build HWPDT MSM screening table rows for the weekly report.
    Matches DB sp_name to audit software_product using prefix matching.
    Loads editable fields (IDPs, planned, remarks) from hwpdt_msm_config table.
    """
    try:
        import sys as _sys
        _sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from dashboard_routes import _load_hwpdt_job_audit_data
        from src.utils import get_mysql_connection_db

        audit = _load_hwpdt_job_audit_data()
        jobs  = audit.get('jobs') or []

        ws_str = sel_start.isoformat()
        we_str = sel_end.isoformat()

        # Build sp_upper -> chips from audit
        week_chips  = {}
        total_chips = {}
        for job in jobs:
            sp    = (job.get('software_product') or '').strip().upper()
            if not sp:
                continue
            chips = [c for c in (job.get('chip_ids') or []) if c]
            st    = (job.get('start_time') or '')[:10]
            total_chips.setdefault(sp, set()).update(chips)
            if ws_str <= st <= we_str:
                week_chips.setdefault(sp, set()).update(chips)

        # Get HWPDT targets from DB
        conn = get_mysql_connection_db(bu_key=None)
        cur  = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT target_name, sp_name, chip_name, bu
            FROM pdt_stats_dashboard.dashboard_status
            WHERE is_hwpdt=1 AND is_active=1
            ORDER BY bu, target_name
        """)
        db_targets = cur.fetchall() or []
        cur.close(); conn.close()

        def _match_sp(db_sp):
            db_up = db_sp.strip().upper()
            w, a  = set(), set()
            for audit_sp, chips in total_chips.items():
                if (audit_sp == db_up or audit_sp.startswith(db_up + '.') or
                        audit_sp.startswith(db_up) or db_up.startswith(audit_sp)):
                    a.update(chips)
                    w.update(week_chips.get(audit_sp, set()))
            return w, a

        # Filter to targets with any activity
        active_targets = []
        chip_data      = {}
        for t in db_targets:
            sp = (t.get('sp_name') or '').strip()
            if not sp:
                continue
            w_chips, a_chips = _match_sp(sp)
            # Skip if no activity all-time OR no parts tested this week
            if not a_chips:  # only skip if no all-time activity
                continue
            active_targets.append(t)
            chip_data[t['target_name']] = (w_chips, a_chips)

        # Load editable config (IDPs, planned, remarks) â€” auto-copies from prev week
        target_names = [t['target_name'] for t in active_targets]
        cfg = _get_hwpdt_msm_config(sel_start, sel_end, target_names) if target_names else {}

        rows = []
        for sno, t in enumerate(active_targets, 1):
            tname  = t['target_name']
            chip   = t.get('chip_name') or ''
            sp     = (t.get('sp_name') or '').strip()
            farm   = (t.get('bu') or '').title()
            w_chips, a_chips = chip_data[tname]
            c = cfg.get(tname) or {}

            # Display name: "ChipName (SP)" e.g. "SM6850 (Skyros.LA.1.0)"
            sp_short = sp.split('.')[0] if sp else ''
            if chip and sp_short and sp_short.upper() not in chip.upper():
                display_target = chip + ' (' + sp_short + ')'
            else:
                display_target = chip or tname

            rows.append({
                's_no':               sno,
                'farm':               farm,
                'target':             display_target,
                'target_name':        tname,
                'team':               'HWPDT',
                'idp_farm':           c.get('idp_farm'),
                'idp_deployed':       c.get('idp_deployed'),
                'parts_tested_week':  len(w_chips),
                'parts_planned_week': c.get('parts_planned'),
                'parts_tested_sod':   len(a_chips),
                'supply_note':        c.get('supply_note') or '',
            })
        return rows
    except Exception as exc:
        import logging
        logging.getLogger(__name__).warning('[HWPDT MSM TABLE] %s', exc)
        return []


def _sp2_landing_summary(week_start, week_end):
    """Landing-card summary that matches api_sp2_builds exactly.

    Uses the same query (taxonomy_path = '/PDT/QIPL'), same grouping
    (build_name + pl_group), same crash counting (distinct tickets from
    weekly_qipl_data), and same unique-chip union so the numbers on the
    landing card are identical to what the Smart Build Report page shows.
    """
    import re as _re2
    import json as _json2

    # Preferred path: use the same frozen Smart Build rows returned by
    # /api/sp2/builds.  This keeps the landing card identical to the report
    # page after the weekly CSV-backed snapshot exists.
    try:
        _seed_sp2_build_type_overrides_from_axiom(week_start, week_end, _current_user_identifier())
        static_rows = _load_sp2_static_build_rows(week_start, week_end)
    except Exception:
        static_rows = []
    if static_rows:
        all_chips = set()
        total_hours = 0.0
        total_crashes = 0
        for r in static_rows:
            chips_raw = r.get('chip_ids') or '[]'
            try:
                chip_ids = _json2.loads(chips_raw) if isinstance(chips_raw, str) else list(chips_raw or [])
            except Exception:
                chip_ids = []
            all_chips.update(str(c).strip() for c in chip_ids if str(c).strip())
            total_hours += float(r.get('hours') or 0)
            total_crashes += int(r.get('total_crashes') or 0)
        return {
            'sp2_build_count': len(static_rows),
            'sp2_device_count': len(all_chips),
            'sp2_total_hours': round(total_hours, 1),
            'sp2_crash_count': total_crashes,
        }

    # 1. Axiom jobs for the week
    db_rows = []
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if conn:
            cur = conn.cursor(dictionary=True)
            try:
                _week_cap = week_end.isoformat() + " 23:59:59"
                _week_floor = week_start.isoformat() + " 00:00:00"
                live_h = (
                    "CASE"
                    " WHEN state IN ('Running','JobSetup') AND started_at IS NOT NULL"
                    " THEN ROUND(device_count *"
                    " TIMESTAMPDIFF(SECOND, started_at,"
                    " LEAST(NOW(), TIMESTAMP('" + _week_cap + "'))) / 3600.0 * 0.80, 3)"
                    " WHEN state IN ('Completed','Aborted') AND started_at IS NOT NULL AND ended_at IS NOT NULL"
                    " THEN ROUND(device_count *"
                    " TIMESTAMPDIFF(SECOND,"
                    " GREATEST(started_at, TIMESTAMP('" + _week_floor + "')),"
                    " LEAST(ended_at,      TIMESTAMP('" + _week_cap   + "'))) / 3600.0 * 0.80, 3)"
                    " ELSE 0 END"
                )
                cur.execute(f"""
                    SELECT job_id, build_id, build_name, software_product,
                           chip_ids, state, device_count, submitter,
                           ({live_h}) AS hours_live
                    FROM `pdt_stats_dashboard`.`axiom_job_summary`
                    WHERE taxonomy_path = '/PDT/QIPL'
                      AND started_at <= %s AND (ended_at >= %s OR state IN ('Running','JobSetup'))
                """, (week_end.isoformat(), week_start.isoformat()))
                db_rows = cur.fetchall() or []
            finally:
                cur.close(); conn.close()
    except Exception:
        pass

    if not db_rows:
        return {'sp2_build_count': 0, 'sp2_device_count': 0,
                'sp2_total_hours': 0.0, 'sp2_crash_count': 0}

        # 2. Week + PL bounded crash map from weekly_qipl_data.
    crash_map = _sp2_weekly_crash_map(week_start, week_end)


    # 3. Group by (build_name, pl_group) â€” same as api_sp2_builds
    def _pl_grp(sp):
        return _re2.sub(r'\.r\d+$', '', str(sp or ''), flags=_re2.IGNORECASE)

    grouped = {}
    for r in db_rows:
        chips_raw = r.get('chip_ids') or '[]'
        if isinstance(chips_raw, str):
            try:
                chip_ids = _json2.loads(chips_raw)
            except Exception:
                chip_ids = []
        else:
            chip_ids = list(chips_raw) if chips_raw else []

        _raw_dev = int(r.get('device_count') or 0)
        hours    = float(r.get('hours_live') or 0)
        if _raw_dev <= 0 and not chip_ids and hours <= 0.1:
            continue
        if str(r.get('submitter') or '').strip().upper() == 'AUTO':
            hours = round(hours * 0.80, 3)

        pl_grp     = _pl_grp(str(r.get('software_product') or '').strip())
        build_id   = str(r.get('build_id') or '').strip()
        build_name = str(r.get('build_name') or build_id).strip()

        crashes    = _sp2_crash_count_for_build(crash_map, build_name, build_id, pl_grp)


        grp_key = (build_name.upper(), pl_grp.upper())
        if grp_key not in grouped:
            grouped[grp_key] = {'hours': 0.0, 'chips': set(), 'crashes': 0}
        g = grouped[grp_key]
        g['hours']   += hours
        g['chips'].update(chip_ids)
        g['crashes'] += crashes

    # 4. Aggregate
    all_chips    = set()
    total_hours  = 0.0
    total_crashes = 0
    for g in grouped.values():
        all_chips.update(g['chips'])
        total_hours   += g['hours']
        total_crashes += g['crashes']

    return {
        'sp2_build_count':  len(grouped),
        'sp2_device_count': len(all_chips),
        'sp2_total_hours':  round(total_hours, 1),
        'sp2_crash_count':  total_crashes,
    }

@weekly_summary_bp.route('/weekly-report')
@login_required
def weekly_report_landing():
    # Do NOT call _ensure_weekly_qipl_table() or _auto_load_qipl_week() here.
    # Scheduler handles imports; running DDL/ALTER + UNC file scans on every
    # page request was the primary cause of slow landing page loads.
    sel_start, sel_end = _selected_week_from_request()
    rows = _fetch_rows(sel_start, sel_end)
    card_data = _build_card_data(rows)

    # Landing Unique Weekly Report should reflect the latest regenerated CSV for
    # the selected week. If the matching source CSV is newer than the generated
    # Excel, refresh Excel first, then read counts from it.
    ucr_latest_source_path = _find_ucr_file_by_week_end(sel_end.isoformat(), force=True)
    try:
        ucr_xl_path = _ucr_excel_path(sel_end)
        src_mtime = os.path.getmtime(ucr_latest_source_path) if ucr_latest_source_path and os.path.isfile(ucr_latest_source_path) else 0
        xl_mtime = os.path.getmtime(ucr_xl_path) if os.path.isfile(ucr_xl_path) else 0
        if ucr_latest_source_path and (not os.path.isfile(ucr_xl_path) or src_mtime > xl_mtime):
            _ensure_ucr_excel_for_week(sel_end, _load_farm_station_map(), force_refresh=True)
    except Exception:
        pass
    ucr_counts, ucr_exists, ucr_path = _ucr_excel_site_counts(sel_end)
    ucr_total = sum(int(v or 0) for v in ucr_counts.values() if str(v).strip() != '') if ucr_exists else ''
    hwpdt_msm_rows = _build_hwpdt_msm_table(sel_start, sel_end)
    sp2_summary = _sp2_landing_summary(sel_start, sel_end)
    return render_template(
        'weekly_reports_landing.html', cards=_CARDS, sel_start=sel_start, sel_end=sel_end,
        week_ranges=_week_ranges_for_templates(), table_rows=rows,
        excel_files=[], qipl_excel_dir=_QIPL_WEEKLY_EXCEL_DIR,
        qipl_auto_import_available=False,
        ucr_counts=ucr_counts, ucr_total=ucr_total,
        ucr_landing_counts=ucr_counts, ucr_landing_total=ucr_total,
        ucr_landing_has_excel=ucr_exists, ucr_landing_excel_path=ucr_path,
        ucr_landing_source_path=ucr_latest_source_path,
        hwpdt_msm_rows=hwpdt_msm_rows,
        **sp2_summary,
        **card_data
    )


@weekly_summary_bp.route('/weekly-report/card/<card_key>')
@login_required
def weekly_report_card(card_key):
    # Do NOT call _ensure_weekly_qipl_table() on every card request.
    # unique_report is Excel-backed; skip DB fetch entirely for that card.
    sel_start, sel_end = _selected_week_from_request()
    rows = [] if card_key == 'unique_report' else _fetch_rows(sel_start, sel_end)
    card_meta = next((c for c in _CARDS if c['key'] == card_key), {'key': card_key, 'title': card_key, 'icon': ''})
    ctx = {
        'card_key': card_key, 'card_meta': card_meta, 'sel_start': sel_start, 'sel_end': sel_end,
        'week_ranges': _week_ranges_for_templates(), 'table_rows': rows,
    }
    if card_key == 'cr_age':
        ctx.update(_build_cr_age_card(rows, sel_start, sel_end))
    elif card_key == 'cr_pie':
        ctx.update(_build_cr_pie_card(rows))
    elif card_key == 'smart_build':
        # Use the full weekly QIPL dataset for the Sharepoint build selector.
        # Older code filtered to rows where CR/Current Ticket was blank, so the
        # browser only saw fallback Stability Ticket IDs (QSTABILITY-*) and the
        # crash count could be lower than the backend/background count.
        ctx.update(_build_sharepoint_context(rows, sel_start, sel_end))
        ctx['sp_known_targets'] = _fetch_sharepoint_known_targets(sel_end)
        ctx['sp_bu_options'] = _sp_bu_options()

    elif card_key == 'unique_report':
        we = _safe_date(request.args.get('ucr_week_end')) or sel_end
        # Only check if the Excel file exists â€” do NOT parse RawData CSV,
        # do NOT call _ensure_ucr_excel_for_week (which reads/generates Excel),
        # do NOT scan _list_ucr_files() (RawData UNC dir scan),
        # do NOT call _read_ucr_pdtsite_chart_data (opens Excel over network),
        # do NOT call _load_farm_station_map (Farm_KPI UNC dir scan).
        # All of those were running synchronously on every page load.
        # Scheduler generates the Excel; page just checks if it exists.
        xl_path = _ucr_excel_path(we)
        exists = os.path.isfile(xl_path)
        # Page load must never regenerate. The saved Excel workbook is the
        # report source; if it is missing/stale, the user can explicitly click
        # "Refresh from CSV" to re-read the matching UNIQUECRSREPORT_WEEKENDING
        # source file and update the generated results.
        if os.path.isfile(xl_path):
            auto_info = {'attempted': False, 'success': True,
                         'message': 'Excel exists', 'path': xl_path}
        else:
            auto_info = {
                'attempted': False, 'success': False,
                'message': 'No generated results yet. Click Refresh from CSV to recheck the latest source file.' if _find_ucr_file_by_week_end(we.isoformat()) else 'No generated results yet. Click Refresh from CSV to scan for the latest UNIQUECRSREPORT file.',
                'path': xl_path,
            }

        counts, exists, _ = _ucr_excel_site_counts(we)
        ucr_sites = [s for s in ('QIPL', 'SD', 'CH') if counts.get(s) not in ('', None)]

        # PDT site-wise bar chart: read from PDTSite sheet (0.25s, fast).
        # Only runs when Excel exists; empty dict when not yet generated.
        site_area_chart = _read_ucr_pdtsite_chart_data(xl_path) if exists else {'categories': [], 'series': [], 'totals': {}}

        # File options from TTL-cached list (no extra UNC scan if cache is warm)
        file_options = [{'week_end': e['week_end_date'].isoformat(), 'label': e['label']}
                        for e in _list_ucr_excel_files()]
        if not file_options:
            lbl = f"{(we - timedelta(days=6)).strftime('%b %d')} - {we.strftime('%b %d, %Y')} (Wk {we.isocalendar()[1]})"
            file_options = [{'week_end': we.isoformat(), 'label': lbl}]

        # raw dates from TTL-cached list (no fresh UNC scan)
        raw_dates = {e['week_end_date'] for e in _list_ucr_files()}

        ctx.update({
            'ucr_week_end_date': we.isoformat(), 'ucr_sites': ucr_sites, 'ucr_rows': {},
            'ucr_total_counts': {k: int(v or 0) if str(v).strip() != '' else 0 for k, v in counts.items()},
            'ucr_file_name': os.path.basename(xl_path) if exists else '',
            'ucr_source': 'excel' if exists else '',
            'ucr_file_options': file_options,
            'ucr_auto_publish_info': auto_info,
            'ucr_site_area_chart': site_area_chart,
            'ucr_qipl_area_pie': [],
            'ucr_farm_names': sorted(_XL_FARM_COLORS.keys()),
                          'ucr_selected_has_raw': (we in raw_dates) or ((we + timedelta(days=1)) in raw_dates),
        })
    return render_template('weekly_card_detail.html', **ctx)


@weekly_summary_bp.route('/weekly-report/hwpdt-msm/save', methods=['POST'])
@login_required
def hwpdt_msm_save():
    """Save editable HWPDT MSM config for a week (IDPs, planned, remarks)."""
    try:
        from src.utils import get_mysql_connection_db
        payload     = request.get_json(force=True) or {}
        week_start  = payload.get('week_start', '').strip()
        week_end    = payload.get('week_end', '').strip()
        target_name = payload.get('target_name', '').strip()
        idp_farm    = payload.get('idp_farm')
        idp_deployed= payload.get('idp_deployed')
        parts_planned = payload.get('parts_planned')
        supply_note = payload.get('supply_note', '').strip()
        username    = getattr(current_user, 'username', 'unknown')

        if not week_start or not target_name:
            return jsonify({'success': False, 'message': 'week_start and target_name required'}), 400

        def _int_or_none(v):
            try: return int(v) if v not in (None, '', 'N/A') else None
            except: return None

        conn = get_mysql_connection_db(bu_key=None)
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO pdt_stats_dashboard.hwpdt_msm_config
                (week_start, week_end, target_name, idp_farm, idp_deployed,
                 parts_planned, supply_note, updated_by, updated_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,NOW())
            ON DUPLICATE KEY UPDATE
                week_end=%s, idp_farm=%s, idp_deployed=%s,
                parts_planned=%s, supply_note=%s,
                updated_by=%s, updated_at=NOW()
        """, (
            week_start, week_end, target_name,
            _int_or_none(idp_farm), _int_or_none(idp_deployed),
            _int_or_none(parts_planned), supply_note, username,
            # ON DUPLICATE KEY UPDATE values
            week_end, _int_or_none(idp_farm), _int_or_none(idp_deployed),
            _int_or_none(parts_planned), supply_note, username
        ))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({'success': True})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500

@weekly_summary_bp.route('/weekly-report/upload', methods=['POST'])
@login_required
def weekly_report_upload():
    _ensure_weekly_qipl_table()
    # â”€â”€ FULL FORM DUMP (temporary debug) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    import logging as _splog
    _splog.getLogger(__name__).warning(
        '[SP save DEBUG] form keys=%r  form_bu=%r  form_target=%r  form_pl_id=%r  selected_items_json=%r',
        list(request.form.keys()),
        request.form.get('bu'),
        request.form.get('target'),
        request.form.get('pl_id'),
        (request.form.get('selected_items_json') or '')[:400]
    )
    # â”€â”€ END FORM DUMP â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws = _safe_date(request.form.get('week_start'))
    we = _safe_date(request.form.get('week_end'))
    upload = request.files.get('excel_file')
    choice = str(request.form.get('file_choice') or '').strip()
    tmp_path = ''
    src_path = ''
    try:
        if upload and upload.filename:
            fname = secure_filename(upload.filename)
            tmp_path = os.path.join(tempfile.gettempdir(), fname)
            upload.save(tmp_path)
            src_path = tmp_path
        elif choice:
            src_path = os.path.join(_QIPL_WEEKLY_EXCEL_DIR, os.path.basename(choice))
        if not src_path or not os.path.isfile(src_path):
            flash('No valid file selected.', 'warning')
        else:
            rows, _ = _parse_file(src_path, getattr(current_user, 'username', ''))
            if ws and we:
                rows = [r for r in rows if r.get('week_start') == ws.isoformat() and r.get('week_end') == we.isoformat()]
            inserted, _deleted, msg = _upsert_rows(rows)
            flash(f'Imported {inserted} row(s). {msg}', 'success' if inserted else 'warning')
    except Exception as exc:
        flash(f'Upload failed: {exc}', 'danger')
    finally:
        if tmp_path:
            try: os.remove(tmp_path)
            except Exception: pass
    return redirect(url_for('weekly_summary_bp.weekly_report_landing', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/api/weekly_summary/<target_name>')
@login_required
def api_weekly_summary(target_name: str):
    """
    Return all weekly trend rows for a target from its managed_excel JSON.
    Called by the weekly_data.html chart via fetch('/api/weekly_summary/<target>').
    Reads the pre-generated JSON written by the weekly scheduler; never hits DB.
    """
    from weekly_summary_service import _target_weekly_path
    import json as _json

    target_key = str(target_name or '').strip().lower()
    if not target_key:
        return jsonify(success=False, message='No target', rows=[]), 400

    # Resolve BU from targets config
    try:
        import dashboard_common as _dc
        bu = (_dc.get_bu_for_target(target_key) or 'UNKNOWN').upper()
    except Exception:
        bu = 'UNKNOWN'

    path = _target_weekly_path(target_key, bu)
    if not path.exists():
        return jsonify(success=False, message=f'No weekly summary JSON found for {target_key}', rows=[]), 200

    try:
        payload = _json.loads(path.read_text(encoding='utf-8')) or {}
    except Exception as exc:
        return jsonify(success=False, message=str(exc), rows=[]), 200

    table_key = f'weekly_summary_{target_key}'
    rows = payload.get(table_key) or payload.get('weeks') or []

    # Apply the same completed-week filter used by the QIPL landing page
    rows = _weekly_trend_completed_rows(rows)

    # Ensure week_end_display label is present for x-axis
    for r in rows:
        if not r.get('week_end_display'):
            we = _safe_date(r.get('week_end'))
            r['week_end_display'] = we.strftime('%m/%d') if we else ''
        # Ensure numeric fields are ints not strings
        for fld in ('total_cr', 'built', 'undisposed', 'invalid', 'dup', 'new_crs', 'old_crs'):
            try:
                r[fld] = int(r[fld] or 0)
            except Exception:
                r[fld] = 0

    return jsonify(success=True, rows=rows, target=target_key, bu=bu,
                   updated_at=payload.get('updated_at', ''),
                   total_weeks=len(rows))


@weekly_summary_bp.route('/api/weekly_qipl_auto_import', methods=['POST'])
@login_required
def weekly_qipl_auto_import():
    data = request.get_json(silent=True) or {}
    ws = _safe_date(data.get('week_start'))
    we = _safe_date(data.get('week_end'))
    if not ws or not we:
        return jsonify(success=False, message='Invalid week'), 400
    info = _auto_load_qipl_week(ws, we, getattr(current_user, 'username', 'auto'))
    return jsonify(success=bool(info.get('loaded')), count=info.get('inserted') or 0, **info)

@weekly_summary_bp.route('/api/qipl_csv_import_now', methods=['POST'])
@login_required
def qipl_csv_import_now():
    """
    Manual on-demand import of the latest QIPL_CR_AGE__CR_TAT_Jira_*.csv.
    POST body (optional): {"week_start": "2026-05-18", "week_end": "2026-05-24"}
    If no week given, finds and imports the latest unimported file automatically.
    """
    data = request.get_json(silent=True) or {}
    ws   = _safe_date(data.get('week_start'))
    we   = _safe_date(data.get('week_end'))
    user = getattr(current_user, 'username', 'manual')

    # Specific week requested
    if ws and we:
        info = _auto_load_qipl_week(ws, we, user)
        return jsonify(success=bool(info.get('loaded')),
                       count=info.get('inserted') or 0,
                       week_start=ws.isoformat(), week_end=we.isoformat(), **info)

    # No week given - find latest unimported file
    files = _list_qipl_source_files()
    if not files:
        return jsonify(success=False, message='No CSV files found in share', path='')

    for entry in files:
        fdate = entry.get('file_date')
        if not fdate:
            continue
        file_ws, file_we = _jira_week(fdate)
        info = _auto_load_qipl_week(file_ws, file_we, user)
        if info.get('loaded'):
            return jsonify(success=True,
                           count=info.get('inserted') or 0,
                           week_start=file_ws.isoformat(),
                           week_end=file_we.isoformat(),
                           file=os.path.basename(info.get('path', '')),
                           **info)
        reason = info.get('reason') or ''
        if reason not in ('already_imported', 'no_ready_unimported_source_file'):
            return jsonify(success=False,
                           week_start=file_ws.isoformat(),
                           week_end=file_we.isoformat(),
                           file=os.path.basename(entry.get('path', '')),
                           **info)

    return jsonify(success=False, message='All available CSV files already imported', path='')


@weekly_summary_bp.route('/api/sharepoint/no_crash_prefill')
@login_required
def weekly_sharepoint_no_crash_prefill():
    target = str(request.args.get('target') or '').strip()
    pl_id = str(request.args.get('pl_id') or '').strip()
    ws = _safe_date(request.args.get('week_start'))
    if not target and not pl_id:
        return jsonify(success=False, message='Target or PL-ID is required'), 400

    prev = _fetch_previous_sharepoint_pair_info(target, pl_id, before_week_start=ws) or {}
    if not target and prev.get('target'):
        target = str(prev.get('target') or '').strip()

    try:
        dash = _find_dashboard_target_info(target, pl_id) or {}
    except Exception:
        dash = {}

    info = {}
    # Prefer exact Target+PL previous-week saved detail. Fill missing fields from dashboard.
    for src in (prev, dash):
        if not isinstance(src, dict):
            continue
        for key in ('target', 'pl_id', 'bu', 'sp_name', 'es', 'fc', 'cs', 'build_type', 'meta_build'):
            if src.get(key) and not info.get(key):
                info[key] = src.get(key)
    info['source'] = prev.get('source') or dash.get('source') or ''
    info['source_week_start'] = prev.get('source_week_start') or ''
    info['source_week_end'] = prev.get('source_week_end') or ''
    info['require_bu_selection'] = bool(target and not str(info.get('bu') or '').strip() and not bool(prev.get('bu')) and not bool(dash.get('matched_exact')))

    return jsonify(success=True, info=info)


@weekly_summary_bp.route('/api/sharepoint/no_crash_swpdt_options')
@login_required
def weekly_sharepoint_no_crash_swpdt_options():
    ws = _safe_date(request.args.get('week_start'))
    we = _safe_date(request.args.get('week_end'))
    if we and not ws:
        ws = we - timedelta(days=6)
    if ws and not we:
        we = ws + timedelta(days=6)
    if not ws or not we:
        ws, we = current_monday_sunday()
    try:
        result = _swpdt_weekly_target_pl_options(ws, we)
        return jsonify(success=True, week_start=ws.isoformat(), week_end=we.isoformat(), **result)
    except Exception as exc:
        return jsonify(success=False, message=str(exc), rows=[]), 500


@weekly_summary_bp.route('/api/sharepoint/no_crash_swpdt_builds')
@login_required
def weekly_sharepoint_no_crash_swpdt_builds():

    target = str(request.args.get('target') or '').strip()
    pl_id = str(request.args.get('pl_id') or '').strip()
    ws = _safe_date(request.args.get('week_start'))
    we = _safe_date(request.args.get('week_end'))
    if we and not ws:
        ws = we - timedelta(days=6)
    if ws and not we:
        we = ws + timedelta(days=6)
    if not ws or not we:
        ws, we = current_monday_sunday()
    try:
        result = _find_swpdt_no_crash_builds(ws, we, target=target, pl_id=pl_id)
        return jsonify(success=True, week_start=ws.isoformat(), week_end=we.isoformat(), **result)
    except Exception as exc:
        return jsonify(success=False, message=str(exc), rows=[]), 500


@weekly_summary_bp.route('/api/sharepoint/fetch_milestones', methods=['POST'])

@login_required
def weekly_sharepoint_fetch_milestones():
    data = request.get_json(silent=True) or {}
    sp_name = str(data.get('sp_name') or '').strip()
    if not sp_name:
        return jsonify(success=False, message='SP name is required'), 400
    if fetch_milestones_for_sp is None:
        return jsonify(success=False, message='OneView milestone helper is not available'), 500
    milestones, source = fetch_milestones_for_sp(sp_name)
    return jsonify(success=True, sp_name=sp_name, source=source, milestones=milestones or {})


@weekly_summary_bp.route('/weekly-report/card/sharepoint/backfill_build_milestones', methods=['POST'])
@login_required
def weekly_sharepoint_backfill_build_milestones():
    ws = _safe_date(request.form.get('week_start'))
    we = _safe_date(request.form.get('week_end'))
    result = _backfill_sharepoint_build_milestones(ws, we)
    flash(result.get('message') or ('Updated ' + str(result.get('updated') or 0) + ' row(s)'), 'success' if result.get('success') else 'error')
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/weekly-report/card/sharepoint/backfill_crashes', methods=['POST'])
@login_required
def weekly_sharepoint_backfill_crashes():
    ws = _safe_date(request.form.get('week_start'))
    we = _safe_date(request.form.get('week_end'))
    result = _backfill_sharepoint_crashes(ws, we)
    flash(result.get('message') or ('Updated ' + str(result.get('updated') or 0) + ' row(s)'), 'success' if result.get('success') else 'error')
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/api/sharepoint/refresh_crashes', methods=['GET'])
@login_required
def api_sharepoint_refresh_crashes():
    """Return live crash count for a single consolidated row (by id).
    Used by the Edit modal Refresh Crashes button.
    """
    row_id  = request.args.get('id', '').strip()
    target  = request.args.get('target', '').strip()
    pl_id   = request.args.get('pl_id', '').strip()
    if not row_id:
        return jsonify(success=False, error='Missing id')
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if not conn:
            return jsonify(success=False, error='DB connection failed')
        cur = conn.cursor(dictionary=True)
        try:
            # Load the saved row to get its build IDs
            cur.execute(
                f"SELECT * FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` WHERE id=%s",
                (row_id,)
            )
            rec = cur.fetchone()
            if not rec:
                return jsonify(success=False, error='Row not found')

            # Collect all build IDs
            build_ids = set()
            top_build = str(rec.get('meta_build') or '').strip()
            if top_build:
                build_ids.add(top_build.upper())
            try:
                sitems = json.loads(rec.get('selected_items_json') or '[]')
            except Exception:
                sitems = []
            for si in (sitems or []):
                if isinstance(si, dict):
                    sb = str(si.get('meta_build') or '').strip()
                    if sb:
                        build_ids.add(sb.upper())

            rec_target = str(rec.get('target') or target or '').strip()
            rec_pl     = str(rec.get('pl_id')  or pl_id  or '').strip()

            if not build_ids:
                return jsonify(success=True, crash_count=0, unique_tickets=0,
                               message='No builds found for this row')

                        # Query/count via the same helper used by backfill and the build-wise pivot
            pivot = _count_sharepoint_crashes_from_weekly_qipl(
                cur, rec_target, rec_pl, build_ids,
                week_start=rec.get('week_start'), week_end=rec.get('week_end')
            )
            return jsonify(
                success=True,
                crash_count=pivot['total'],
                crash_details=pivot.get('details') or '',
                unique_tickets=len(pivot.get('items') or []),
                row_count=pivot.get('row_count') or 0,
            )

        finally:
            cur.close()
            conn.close()
    except Exception as exc:
        return jsonify(success=False, error=str(exc))


@weekly_summary_bp.route('/weekly-report/card/sharepoint/backfill_bu', methods=['POST'])
@login_required
def weekly_sharepoint_backfill_bu():
    ws = _safe_date(request.form.get('week_start'))
    we = _safe_date(request.form.get('week_end'))
    result = _backfill_sharepoint_bu(ws, we)
    flash(result.get('message') or ('Updated ' + str(result.get('updated') or 0) + ' row(s)'), 'success' if result.get('success') else 'error')
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/weekly-report/card/sharepoint/save', methods=['POST'])
@login_required
def weekly_sharepoint_save():
    _ensure_weekly_qipl_table()
    ws = _safe_date(request.form.get('week_start'))
    we = _safe_date(request.form.get('week_end'))
    if not ws or not we:
        flash('Invalid week selected', 'error')
        return redirect(url_for('weekly_summary_bp.weekly_report_landing'))
    target = str(request.form.get('target') or '').strip()
    pl_id = str(request.form.get('pl_id') or '').strip()
    build_type = str(request.form.get('build_type') or 'CRM').strip().upper() or 'CRM'
    try:
        selected_items = json.loads(request.form.get('selected_items_json') or '[]')
        if not isinstance(selected_items, list):
            selected_items = []
    except Exception:
        selected_items = []
    manual_items = [it for it in selected_items if isinstance(it, dict) and it.get('manual_no_crash')]
    manual = manual_items[0] if manual_items else {}
    raw_form_pl_id = str(request.form.get('pl_id') or '').strip()



    # Extract BU from manual_bu item (crash save path) or form field
    manual_bu_item = next((it for it in selected_items if isinstance(it, dict) and it.get('manual_bu')), {})
    first_item_with_pl = next((it for it in selected_items if isinstance(it, dict) and str(it.get('pl_id') or '').strip()), {})

    bu = _normalize_bu(str(request.form.get('bu') or '').strip())
    # Pull BU from manual_bu sentinel (set by spApplyRequiredBu popup)
    if not bu and manual_bu_item.get('bu'):
        bu = _normalize_bu(str(manual_bu_item.get('bu') or '').strip())
    # When spTarget is empty (target not in dropdown), pull target/pl_id
    # from the manual_bu sentinel which carries them from the crash items
    if not target and manual_bu_item.get('target'):
        target = str(manual_bu_item.get('target') or '').strip()
    if not pl_id:
        # Try manual_bu sentinel first (most reliable for unknown targets)
        pl_id = str(manual_bu_item.get('pl_id') or '').strip()
    if not pl_id:
        pl_id = str(first_item_with_pl.get('pl_id') or '').strip()
    if manual:
        target = target or str(manual.get('target') or '').strip()
        pl_id = pl_id or str(manual.get('pl_id') or '').strip() or 'Manual'
        build_type = str(manual.get('build_type') or build_type or 'CRM').strip().upper()
        bu = _normalize_bu(bu or str(manual.get('bu') or '').strip())
        if not bu:
            flash('BU is mandatory for manually added Sharepoint Target / PL-ID.', 'error')
            return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat(), week_end=we.isoformat()))
    meta_build = ''
    for it in selected_items:
        if isinstance(it, dict) and str(it.get('meta_build') or '').strip():
            meta_build = str(it.get('meta_build')).strip()
            break
    if not meta_build:
        meta_build = str(request.form.get('meta_build') or '').strip() or 'Manual No-Crash Build'
    pl_id = _apply_snapdragon_pl_alias(target, pl_id, raw_form_pl_id)
    for it in selected_items:
        if isinstance(it, dict):
            _raw_item_pl = str(it.get('pl_id') or '').strip()
            it['pl_id'] = _apply_snapdragon_pl_alias(target or str(it.get('target') or '').strip(), _raw_item_pl, raw_form_pl_id)

    if not target and pl_id:
        target = pl_id
    # Always look up previous week â€” BU lookup is independent of pl_id.
    # Previously gated on `not pl_id or not bu` which skipped the lookup
    # when pl_id was set but bu was still empty (e.g. QCM6690 not in dashboard).
    _prev_lookup_needed = not pl_id or not bu
    if _prev_lookup_needed:
        prev = _fetch_previous_sharepoint_pair_info(target, pl_id, before_week_start=we + timedelta(days=1))
        if not pl_id:
            pl_id = str(prev.get('pl_id') or '').strip()
        if not bu:
            bu = _normalize_bu(str(prev.get('bu') or '').strip())
    # Even when pl_id was already set, still try previous week for BU alone
    if not bu:
        prev = _fetch_previous_sharepoint_pair_info(target, pl_id, before_week_start=we + timedelta(days=1))
        bu = _normalize_bu(str(prev.get('bu') or '').strip())
    if not bu:
        dash = _find_dashboard_target_info(target, pl_id)
        bu = _normalize_bu(str(dash.get('bu') or '').strip())
    # Last-resort: pull BU from any item in selected_items that carries it
    if not bu:
        for _it in selected_items:
            if isinstance(_it, dict) and str(_it.get('bu') or '').strip():
                bu = _normalize_bu(str(_it.get('bu') or '').strip())
                break
    if not bu:
        import logging as _lg2
        _lg2.getLogger(__name__).warning(
            '[SP save] BU unresolved: target=%r pl_id=%r selected_items=%r',
            target, pl_id, selected_items)
        flash('BU could not be resolved. Please select BU in the form and save again.', 'error')
        return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat(), week_end=we.isoformat()))
    try:
        hours_val = float(request.form.get('hours') or 0)
        devices_val = int(float(request.form.get('devices') or 0))  # 0 = not set, devices managed via Devices tab
    except Exception:
        hours_val = 0
        devices_val = 0
    if hours_val <= 0:
        flash('Hours is mandatory and must be greater than 0.', 'error')
        return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat(), week_end=we.isoformat()))
    if not pl_id:
        flash('PL-ID is missing. Please reselect Target / PL-ID before saving.', 'error')
        return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat(), week_end=we.isoformat()))
    conn = get_mysql_connection_db(bu_key=None)
    if conn:
        cur = conn.cursor(dictionary=True)
        try:
            rows_to_save = manual_items if manual_items else [{}]
            saved_count = 0
            label_state = {}

            def _label_for_row(row_target: str, row_pl_id: str, row_build_type: str, row_meta_build: str) -> str:
                key = (row_target, row_pl_id, row_build_type)
                if key not in label_state:
                    cur.execute(f"""
                        SELECT build_label, meta_build
                        FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`
                        WHERE week_start=%s AND week_end=%s AND target=%s
                          AND COALESCE(pl_id,'')=%s AND COALESCE(build_type,'CRM')=%s
                        ORDER BY id
                    """, (ws.isoformat(), we.isoformat(), row_target, row_pl_id, row_build_type or 'CRM'))
                    meta_to_label = {}
                    nums = []
                    for r in cur.fetchall() or []:
                        label = str(r.get('build_label') or '').strip()
                        existing_meta = str(r.get('meta_build') or '').strip().upper()
                        if existing_meta and label:
                            meta_to_label[existing_meta] = label
                        m = _re.search(r'(\d+)', label)
                        if m:
                            nums.append(int(m.group(1)))
                    label_state[key] = {'meta_to_label': meta_to_label, 'next': max(nums or [0]) + 1}
                state = label_state[key]
                wanted = str(row_meta_build or '').strip().upper()
                if wanted and wanted in state['meta_to_label']:
                    return state['meta_to_label'][wanted]
                label = str(state['next'])
                state['next'] += 1
                if wanted:
                    state['meta_to_label'][wanted] = label
                return label

            for save_item in rows_to_save:
                row_target = target or str(save_item.get('target') or '').strip()
                row_pl_id = _apply_snapdragon_pl_alias(row_target, pl_id or str(save_item.get('pl_id') or '').strip(), raw_form_pl_id)

                row_build_type = str(save_item.get('build_type') or build_type or 'CRM').strip().upper() or 'CRM'
                row_meta_build = str(save_item.get('meta_build') or meta_build or '').strip() or 'Manual No-Crash Build'
                row_bu = _normalize_bu(str(save_item.get('bu') or bu or '').strip())
                try:
                    row_devices = int(float(save_item.get('devices') or devices_val or 0))
                except Exception:
                    row_devices = devices_val
                if not row_target and row_pl_id:
                    row_target = row_pl_id
                if not row_pl_id or not row_bu:
                    import logging as _lg
                    _lg.getLogger(__name__).warning(
                        '[SP save] SKIPPED row: target=%r pl_id=%r bu=%r '
                        'save_item=%r', row_target, row_pl_id, row_bu, save_item)
                    continue

                build_label = _label_for_row(row_target, row_pl_id, row_build_type, row_meta_build)
                hours = hours_val
                devices = row_devices or devices_val
                try:
                    crash_raw = request.form.get('crash_count') or 0
                    if isinstance(save_item, dict) and save_item.get('crash_count') is not None:
                        crash_raw = save_item.get('crash_count')
                    crashes = int(float(crash_raw or 0))
                except Exception:
                    crashes = 0
                mtbf = float(request.form.get('mtbf') or (round(hours / crashes, 2) if crashes else 0))
                row_selected_items = [save_item] if save_item else selected_items
                # Keep save fast: do not resolve/fetch milestones while saving.
                # Manual no-crash rows may already carry ES/FC/CS from the form;
                # otherwise Consolidated Report Refresh can resolve timelines later.
                milestone_info = {
                    'bu': row_bu,
                    'es': (save_item.get('es') if isinstance(save_item, dict) else None),
                    'fc': (save_item.get('fc') if isinstance(save_item, dict) else None),
                    'cs': (save_item.get('cs') if isinstance(save_item, dict) else None),
                }
                crash_details_raw = request.form.get('crash_details')
                if isinstance(save_item, dict) and save_item.get('crash_details') is not None:
                    crash_details_raw = save_item.get('crash_details')
                crash_details = str(crash_details_raw or '')[:4000]
                cur.execute(f"""
                    INSERT INTO `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`
                    (week_start, week_end, target, pl_id, build_type, build_label, meta_build, selected_items_json, hours, devices, mtbf, crash_count, crash_details, bu, es_date, fc_date, cs_date, created_by)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE meta_build=VALUES(meta_build), selected_items_json=VALUES(selected_items_json), hours=VALUES(hours), devices=VALUES(devices), mtbf=VALUES(mtbf), crash_count=VALUES(crash_count), crash_details=VALUES(crash_details), bu=VALUES(bu), es_date=VALUES(es_date), fc_date=VALUES(fc_date), cs_date=VALUES(cs_date), created_by=VALUES(created_by)
                """, (ws.isoformat(), we.isoformat(), row_target, row_pl_id, row_build_type, build_label, row_meta_build, json.dumps(row_selected_items), hours, devices, mtbf, crashes, crash_details, milestone_info.get('bu') or row_bu, milestone_info.get('es') or None, milestone_info.get('fc') or None, milestone_info.get('cs') or None, _current_user_identifier()))
                saved_count += 1
            conn.commit()
            # Keep the save path fast for users. Consolidated Report is rebuilt
            # on demand by the Refresh Report button/API instead of blocking
            # every Save Build Summary / No-Crash save.
            if saved_count:
                flash(f'Sharepoint build summary saved ({saved_count} row(s)). Use Refresh Report in Consolidated Report if totals need updating.', 'success')
            else:
                flash('No Sharepoint rows were saved. Please verify Target, PL-ID and BU mapping.', 'error')



        except Exception as exc:
            try: conn.rollback()
            except Exception: pass
            flash(f'Save failed: {exc}', 'error')
        finally:
            cur.close(); conn.close()
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat(), week_end=we.isoformat()))


@weekly_summary_bp.route('/weekly-report/card/sharepoint/update', methods=['POST'])
@login_required
def weekly_sharepoint_update():
    ws = _safe_date(request.form.get('week_start')); we = _safe_date(request.form.get('week_end'))
    row_id = request.form.get('id')
    target = str(request.form.get('target') or '').strip()
    raw_form_pl_id = str(request.form.get('pl_id') or '').strip()
    pl_id = _apply_snapdragon_pl_alias(target, raw_form_pl_id, raw_form_pl_id)

    build_type = str(request.form.get('build_type') or 'CRM').strip().upper() or 'CRM'
    meta_build = str(request.form.get('meta_build') or '').strip()
    bu = str(request.form.get('bu') or '').strip()
    try:
        hours = float(request.form.get('hours') or 0)
        devices = int(float(request.form.get('devices') or 0))
        crashes = int(float(request.form.get('crash_count') or 0))
        mtbf = float(request.form.get('mtbf') or (round(hours / crashes, 2) if crashes else 0))
    except Exception:
        flash('Invalid numeric values for Sharepoint row update.', 'error')
        return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))
    if not target or not pl_id:
        flash('Target and PL-ID are required for row update.', 'error')
        return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))
    if hours <= 0 or devices <= 0:
        flash('Hours and Devices must be greater than 0.', 'error')
        return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))

    conn = get_mysql_connection_db(bu_key=None)
    if conn:
        cur = conn.cursor(dictionary=True)
        try:
            cur.execute(f"SELECT * FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` WHERE id=%s LIMIT 1", (row_id,))
            rec = cur.fetchone()
            if not rec:
                flash('Saved Sharepoint row not found.', 'error')
            else:
                try:
                    selected_items = json.loads(rec.get('selected_items_json') or '[]')
                    if not isinstance(selected_items, list):
                        selected_items = []
                except Exception:
                    selected_items = []
                updated_items = []
                for item in selected_items:
                    if not isinstance(item, dict):
                        updated_items.append(item)
                        continue
                    new_item = dict(item)
                    if target:
                        new_item['target'] = target
                    if pl_id:
                        new_item['pl_id'] = _apply_snapdragon_pl_alias(target, pl_id, raw_form_pl_id)

                    if build_type:
                        new_item['build_type'] = build_type
                    if bu:
                        new_item['bu'] = bu
                    if meta_build and str(new_item.get('meta_build') or '').strip() == str(rec.get('meta_build') or '').strip():
                        new_item['meta_build'] = meta_build
                    updated_items.append(new_item)
                milestone_info = _resolve_sharepoint_build_milestones(target, rec={'bu': bu, 'es_date': rec.get('es_date'), 'fc_date': rec.get('fc_date'), 'cs_date': rec.get('cs_date'), 'sp_name': target})
                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`
                    SET target=%s, pl_id=%s, build_type=%s, meta_build=%s,
                        selected_items_json=%s, hours=%s, devices=%s, mtbf=%s,
                        crash_count=%s, crash_details=%s, bu=%s, es_date=%s, fc_date=%s, cs_date=%s, created_by=%s
                    WHERE id=%s
                """, (target, pl_id, build_type, meta_build or (rec.get('meta_build') or ''), json.dumps(updated_items), hours, devices, mtbf, crashes, str(request.form.get('crash_details') or '')[:4000], milestone_info.get('bu') or bu, milestone_info.get('es') or None, milestone_info.get('fc') or None, milestone_info.get('cs') or None, _current_user_identifier(), row_id))
                conn.commit()
                flash('Sharepoint row updated successfully.', 'success')
        except Exception as exc:
            try: conn.rollback()
            except Exception: pass
            flash(f'Sharepoint row update failed: {exc}', 'error')
        finally:
            cur.close(); conn.close()
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/weekly-report/card/sharepoint/delete', methods=['POST'])
@login_required
def weekly_sharepoint_delete():
    ws = _safe_date(request.form.get('week_start')); we = _safe_date(request.form.get('week_end'))
    conn = get_mysql_connection_db(bu_key=None)
    if conn:
        cur = conn.cursor()
        try:
            cur.execute(f"DELETE FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}` WHERE id=%s", (request.form.get('id'),))
            conn.commit()
        finally:
            cur.close(); conn.close()
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/api/sharepoint/consolidate_data')
@login_required
def weekly_sharepoint_consolidate_data():
    we = _safe_date(request.args.get('week_end'))
    if not we:
        return jsonify(success=False, error='Invalid week_end')
    if request.args.get('refresh') == '1':
        _build_and_save_consolidate_summary(we, _current_user_identifier())
    rows = _fetch_consolidate_summary(we)
    return jsonify(success=True, rows=rows)


@weekly_summary_bp.route('/weekly-report/card/sharepoint/consolidate_backfill_milestones', methods=['POST'])
@login_required
def weekly_sharepoint_consolidate_backfill_milestones():
    we = _safe_date(request.form.get('week_end'))
    ws = _safe_date(request.form.get('week_start'))
    result = _fetch_sharepoint_row_milestones_backfill(we)
    flash(result.get('message') or ('Updated ' + str(result.get('updated') or 0) + ' row(s)'), 'success' if result.get('success') else 'error')
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/weekly-report/card/sharepoint/consolidate_update', methods=['POST'])
@login_required
def weekly_sharepoint_consolidate_update():
    ws = _safe_date(request.form.get('week_start')); we = _safe_date(request.form.get('week_end'))
    conn = get_mysql_connection_db(bu_key=None)
    if conn:
        cur = conn.cursor()
        try:
            cur.execute(f"""
                UPDATE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`
                SET bu=%s,target=%s,pl_id=%s,timelines=%s,pdt_test_status=%s,number_of_devices=%s,number_of_builds=%s,total_hours=%s,total_crashes=%s,unique_crs=%s,mtbf=%s,updated_by=%s
                WHERE id=%s
            """, (request.form.get('bu'), request.form.get('target'), request.form.get('pl_id'), request.form.get('timelines'), request.form.get('pdt_test_status'), request.form.get('number_of_devices') or 0, request.form.get('number_of_builds') or 0, request.form.get('total_hours') or 0, request.form.get('total_crashes') or 0, request.form.get('unique_crs') or None, request.form.get('mtbf') or 0, _current_user_identifier(), request.form.get('id')))
            conn.commit()
        finally:
            cur.close(); conn.close()
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/weekly-report/card/sharepoint/consolidate_delete', methods=['POST'])
@login_required
def weekly_sharepoint_consolidate_delete():
    ws = _safe_date(request.form.get('week_start')); we = _safe_date(request.form.get('week_end'))
    conn = get_mysql_connection_db(bu_key=None)
    if conn:
        cur = conn.cursor()
        try:
            cur.execute(f"DELETE FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` WHERE id=%s", (request.form.get('id'),))
            conn.commit()
        finally:
            cur.close(); conn.close()
    return redirect(url_for('weekly_summary_bp.weekly_report_card', card_key='sharepoint', week_start=ws.isoformat() if ws else '', week_end=we.isoformat() if we else ''))


@weekly_summary_bp.route('/api/sharepoint/device_utilization_data')
@login_required
def weekly_sharepoint_device_utilization_data():
    we = _safe_date(request.args.get('week_end')) or date.today()
    rows = _fetch_consolidate_summary(we)

        # Scan last 12 weeks, keep only those with real data, return up to 2 most-recent.
    # Each week also carries per-PL device counts so the table can show
    # every PL across all returned weeks with 0 for missing entries.
    candidate_weeks = []
    for i in range(11, -1, -1):
        w = we - timedelta(weeks=i)
        rr = _fetch_consolidate_summary(w)
        hrs     = sum(float(r.get('total_hours')       or 0) for r in rr)
        crashes = sum(float(r.get('total_crashes')     or 0) for r in rr)
        # number_of_devices is the hours-proportional per-PL split of devices_count.
        # SUM(number_of_devices) == devices_count target total.
        # Use number_of_devices so Device Utilization total matches Consolidate total.
        dev     = sum(float(r.get('number_of_devices') or 0) for r in rr)
        if not (hrs > 0 or crashes > 0 or dev > 0):
            continue  # skip weeks with no data at all
        # per-PL device map for this week
        pl_devices = {}
        for r in rr:
            key = (str(r.get('target') or ''), str(r.get('pl_id') or ''))
            pl_devices[key] = int(r.get('number_of_devices') or 0)
        d = w.strftime('%d-%b')
        candidate_weeks.append({
            'week_end':            w.isoformat(),
            'label':               d,
            'total_hours':         hrs,
            'total_crashes':       crashes,
            'total_devices':       dev,
            'device_usage_per_week': hrs / dev     if dev     else 0,
            'time_per_crash':        hrs / crashes if crashes else 0,
            'crash_per_mtp_week':    crashes / dev if dev     else 0,
            'pl_devices':          {f"{t}||{p}": d2 for (t, p), d2 in pl_devices.items()},
        })

        # Keep the most-recent 3 non-empty weeks (or fewer if not yet available)
    weeks = candidate_weeks[-3:]
    return jsonify(success=True, actual_week=we.isoformat(), rows=rows, weeks=weeks)


@weekly_summary_bp.route('/api/sharepoint/device_utilization_save', methods=['POST'])
@login_required
def weekly_sharepoint_device_utilization_save():
    data = request.get_json(silent=True) or {}
    conn = get_mysql_connection_db(bu_key=None)
    if conn:
        cur = conn.cursor()
        try:
            for u in data.get('updates') or []:
                # Write to devices_count - this is the field that Device Utilization
                # and Consolidate Report both read. number_of_devices is the auto-computed
                # sum from build-level entries and is separate.
                cur.execute(
                    f"UPDATE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`"
                    f" SET devices_count=%s WHERE week_end=%s AND target=%s",
                    (int(u.get('devices') or 0), u.get('week'), u.get('target'))
                )
            conn.commit()
        finally:
            cur.close(); conn.close()
    return jsonify(success=True)


# New: devices_count lookup by target
@weekly_summary_bp.route('/api/sharepoint/devices_lookup')
@login_required
def weekly_sharepoint_devices_lookup():
    we     = str(request.args.get('week_end') or '').strip()
    target = str(request.args.get('target')   or '').strip()
    if not we or not target:
        return jsonify(success=False, devices_count=None)
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return jsonify(success=False, devices_count=None)
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(
            f"SELECT devices_count FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` "
            f"WHERE week_end=%s AND target=%s AND devices_count IS NOT NULL LIMIT 1",
            (we, target)
        )
        row = cur.fetchone()
        dc = int(row['devices_count']) if row and row.get('devices_count') is not None else None
        return jsonify(success=True, devices_count=dc)
    except Exception as e:
        return jsonify(success=False, devices_count=None, error=str(e))
    finally:
        cur.close(); conn.close()


# New: manual devices_count save
@weekly_summary_bp.route('/api/sharepoint/devices_save', methods=['POST'])
@login_required
def weekly_sharepoint_devices_save():
    data = request.get_json(silent=True) or {}
    we   = str(data.get('week_end') or '').strip()
    rows = data.get('rows') or []
    if not we or not rows:
        return jsonify(success=False, error='Missing week_end or rows')
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return jsonify(success=False, error='DB connection failed')
    cur = conn.cursor()
    try:
        # Ensure devices_count column exists (safe for older MySQL)
        try:
            cur.execute(f"""
                SELECT COUNT(*) FROM information_schema.COLUMNS
                WHERE TABLE_SCHEMA=%s AND TABLE_NAME=%s AND COLUMN_NAME='devices_count'
            """, (_QIPL_DB, _CONSOLIDATE_SUMMARY_TABLE))
            row_chk = cur.fetchone()
            col_exists = (row_chk[0] if isinstance(row_chk, (list,tuple)) else list(row_chk.values())[0]) > 0
            if not col_exists:
                cur.execute(f"ALTER TABLE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` ADD COLUMN `devices_count` INT DEFAULT NULL")
                conn.commit()
        except Exception:
            pass
        for row in rows:
            tgt = str(row.get('target') or '').strip()
            cnt = row.get('devices_count')
            cnt = int(cnt) if cnt not in (None, '', 'null') else None
            # Save same devices_count for ALL PLs under this target
            cur.execute(
                f"UPDATE `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}` "
                f"SET devices_count=%s WHERE week_end=%s AND target=%s",
                (cnt, we, tgt)
            )
        conn.commit()
        return jsonify(success=True, updated=len(rows))
    except Exception as e:
        return jsonify(success=False, error=str(e))
    finally:
        cur.close(); conn.close()


@weekly_summary_bp.route('/api/sharepoint/stability_health_data')
@login_required
def weekly_sharepoint_stability_health_data():
    we = _safe_date(request.args.get('week_end')) or date.today()
    weeks = []
    for i in range(7, -1, -1):
        w = we - timedelta(weeks=i)
        rr = _fetch_consolidate_summary(w)
        hrs = sum(float(r.get('total_hours') or 0) for r in rr)
        crashes = sum(float(r.get('total_crashes') or 0) for r in rr)
        dev = sum(float(r.get('devices_count') or 0) for r in rr)
        weeks.append({'week_end': w.isoformat(), 'label': f"Wk {w.isocalendar()[1]} {w.strftime('%b %d, %Y')}", 'total_hours': hrs, 'total_crashes': crashes, 'total_devices': dev, 'device_usage_per_week': hrs / dev if dev else 0, 'time_per_crash': hrs / crashes if crashes else 0, 'crash_per_mtp_week': crashes / dev if dev else 0})
    return jsonify(success=True, weeks=weeks)


@weekly_summary_bp.route('/api/ucr/generate_excel', methods=['POST'])
@login_required
def weekly_ucr_generate_excel():
    data = request.get_json(silent=True) or {}
    we = _safe_date(data.get('week_end'))
    do_refresh = bool(data.get('force') or data.get('refresh'))
    if do_refresh:
        _UCR_RAW_FILES_CACHE.update({'ts': 0.0, 'value': []})
    info = _ensure_ucr_excel_for_week(we, _load_farm_station_map(), force_refresh=do_refresh)
    return jsonify(success=bool(info.get('success')), message=info.get('message'), path=info.get('path'))


@weekly_summary_bp.route('/api/ucr/sheet_data')
@login_required
def weekly_ucr_sheet_data():
    we = _safe_date(request.args.get('week_end'))
    site = str(request.args.get('site') or 'QIPL').upper()
    path = _ucr_excel_path(we) if we else ''
    payload = _read_ucr_excel_sheet(path, site, _load_farm_station_map())
    area_counts = Counter(str(r.get('CR Area') or 'Unknown') for r in payload.get('rows') or [])
    payload['success'] = True
    payload['area_pie'] = [{'name': k, 'y': v} for k, v in sorted(area_counts.items(), key=lambda x: -x[1])]
    return jsonify(payload)


@weekly_summary_bp.route('/api/ucr/save_farm', methods=['POST'])
@login_required
def weekly_ucr_save_farm():
    return jsonify(success=True)


# ---------------------------------------------------------------------------
# MONTHLY REPORT  (BU-level, date-range, SWPDT + optional HWPDT)
# ---------------------------------------------------------------------------


def _monthly_bu_list():
    """Return BU options with 'ALL' prepended."""
    opts = _sp_bu_options()
    return [{'key': 'ALL', 'label': 'All BUs'}] + opts


def _monthly_date_range_from_request():
    """Parse date_from / date_to from request args; default = last full calendar month."""
    today      = date.today()
    first_this = today.replace(day=1)
    last_prev  = first_this - timedelta(days=1)
    first_prev = last_prev.replace(day=1)
    date_from  = _safe_date(request.args.get('date_from')) or first_prev
    date_to    = _safe_date(request.args.get('date_to'))   or last_prev
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    return date_from, date_to


def _monthly_fetch_consolidate(bu: str, date_from, date_to) -> list:
    """
    Aggregate weekly_sharepoint_consolidate_summary rows across all week_ends
    that fall within [date_from, date_to], grouped by (bu, target, pl_id).
    """
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = conn.cursor(dictionary=True)
    try:
        params = [date_from.isoformat(), date_to.isoformat()]
        bu_clause = ''
        if bu and bu != 'ALL':
            bu_clause = "AND UPPER(COALESCE(bu,'')) = %s"
            params.append(bu.upper())
        cur.execute(f"""
            SELECT bu, target, pl_id,
                   SUM(COALESCE(total_hours,0))      AS total_hours,
                   SUM(COALESCE(total_crashes,0))    AS total_crashes,
                   SUM(COALESCE(number_of_builds,0)) AS number_of_builds,
                   MAX(COALESCE(number_of_devices,0)) AS number_of_devices,
                   MAX(COALESCE(unique_crs,0))        AS unique_crs,
                   MAX(COALESCE(mtbf,0))              AS mtbf,
                   MAX(pdt_test_status)               AS pdt_test_status,
                   MAX(timelines)                     AS timelines,
                   MIN(week_end)                      AS first_week,
                   MAX(week_end)                      AS last_week
            FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`
            WHERE week_end >= %s AND week_end <= %s
            {bu_clause}
            GROUP BY bu, target, pl_id
            ORDER BY COALESCE(bu,''), COALESCE(target,''), COALESCE(pl_id,'')
        """, params)
        rows = cur.fetchall() or []
        out  = []
        for r in rows:
            row = dict(r)
            for k, v in list(row.items()):
                if hasattr(v, 'isoformat'):
                    row[k] = v.isoformat()[:10]
                elif hasattr(v, '__float__'):
                    try:
                        row[k] = float(v)
                    except Exception:
                        pass
            out.append(row)
        return out
    except Exception:
        return []
    finally:
        cur.close(); conn.close()


def _monthly_fetch_build_rows(bu: str, date_from, date_to) -> list:
    """
    Fetch individual build rows from weekly_sharepoint_build_summary
    for the date range â€” used for the stability trend chart.
    """
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = conn.cursor(dictionary=True)
    try:
        params = [date_from.isoformat(), date_to.isoformat()]
        bu_clause = ''
        if bu and bu != 'ALL':
            bu_clause = "AND UPPER(COALESCE(bu,'')) = %s"
            params.append(bu.upper())
        cur.execute(f"""
            SELECT target, pl_id, bu, build_label, meta_build,
                   hours, crash_count, mtbf, week_end, week_start
            FROM `{_QIPL_DB}`.`{_SHAREPOINT_SUMMARY_TABLE}`
            WHERE week_end >= %s AND week_end <= %s
            {bu_clause}
            ORDER BY COALESCE(target,''), COALESCE(pl_id,''), week_end, build_label
        """, params)
        rows = cur.fetchall() or []
        out  = []
        for r in rows:
            row = dict(r)
            for k, v in list(row.items()):
                if hasattr(v, 'isoformat'):
                    row[k] = v.isoformat()[:10]
                elif hasattr(v, '__float__'):
                    try:
                        row[k] = float(v)
                    except Exception:
                        pass
            out.append(row)
        return out
    except Exception:
        return []
    finally:
        cur.close(); conn.close()


def _tbl_exists_cur(cur, fq_table: str) -> bool:
    """Check if a fully-qualified table exists using information_schema."""
    try:
        n = fq_table.replace('`', '')
        s, t = n.split('.', 1)
        cur.execute(
            "SELECT 1 FROM information_schema.tables "
            "WHERE table_schema=%s AND table_name=%s LIMIT 1",
            (s, t)
        )
        return cur.fetchone() is not None
    except Exception:
        return False


def _pl_id_to_db_prefix(pl_id: str) -> str:
    """
    Convert pl_id from consolidate table to db_prefix used in table names.
    e.g. 'Kobuk.LE.1.1' -> 'kobuk_le_1_1'
         'Pinnacles.LE.2.3' -> 'pinnacles_le_2_3'
    """
    return pl_id.replace('.', '_').replace('-', '_').lower()


def _target_to_base(target: str) -> str:
    """
    Convert target name to base prefix (first part before dot).
    e.g. 'Kobuk.LE' -> 'kobuk'
         'Pinnacles.LE' -> 'pinnacles'
    """
    return target.split('.')[0].lower()


def _get_schema_for_bu(bu: str) -> str:
    """Return DB schema for a BU key."""
    from config import BU_DATABASE_MAPPING
    bu_up = str(bu or '').strip().upper()
    return BU_DATABASE_MAPPING.get(bu_up, '')


def _monthly_fetch_target_cr_data(
    conn, bu: str, target: str, pl_id: str,
    date_from_s: str, date_to_s: str,
    include_hwpdt: bool = False,
) -> dict:
    """
    Fetch all CR/JIRA metrics for ONE pl_id within the date range.

    Table resolution:
      schema          = BU_DATABASE_MAPPING[bu]
      db_prefix       = pl_id.replace('.','_').lower()   e.g. kobuk_le_1_1
      unique_crs      = {schema}.{db_prefix}_unique_crs
      jiras           = {schema}.{db_prefix}_jiras
      openjiras       = {schema}.{db_prefix}_openjiras
      overall_crs     = try {schema}.{base}_overallcrs   e.g. kobuk_overallcrs
                        (shared across all pl_ids of same target)

    Returns:
      {
        'target': str, 'pl_id': str, 'bu': str,
        'pdt_crs':         list[dict],
        'overall_crs':     list[dict],
        'overall_enabled': bool,
        'total_jiras':     int,
        'open_jiras':      int,
        'unique_cr_count': int,   # COUNT DISTINCT mapped_cr from overall_crs
      }
    """
    schema = _get_schema_for_bu(bu)
    if not schema:
        return {}

    db_prefix    = _pl_id_to_db_prefix(pl_id)
    base_prefix  = _target_to_base(target)

    result = {
        'target':          target,
        'pl_id':           pl_id,
        'bu':              bu,
        'pdt_crs':         [],
        'overall_crs':     [],
        'overall_enabled': False,
        'total_jiras':     0,
        'open_jiras':      0,
        'unique_cr_count': 0,
        'error':           None,
    }

    try:
        cur = conn.cursor(dictionary=True)

        # â”€â”€ 1. unique_crs â†’ PDT CRs (jira_date filter) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        u_table = f'`{schema}`.`{db_prefix}_unique_crs`'
        if _tbl_exists_cur(cur, u_table):
            try:
                cur.execute(f'SHOW COLUMNS FROM {u_table}')
                u_cols = frozenset(r['Field'].lower() for r in (cur.fetchall() or []))

                def _sel(col, fallback="''"):
                    return f'`{col}`' if col.lower() in u_cols else fallback

                last_col = next(
                    (c for c in ('jira_date__last_instance',
                                 'qstability__last_instance',
                                 'jira_date_last_instance')
                     if c in u_cols), None)
                last_sel = (f'`{last_col}` AS jira_date_last'
                            if last_col else 'NULL AS jira_date_last')

                tt_clause = ''
                if 'test_team' in u_cols and not include_hwpdt:
                    tt_clause = " AND (test_team IS NULL OR UPPER(test_team) != 'PDT_QIPL_HWPDT')"

                date_clause = ''
                date_params = []
                if 'jira_date' in u_cols:
                    date_clause = ' AND jira_date >= %s AND jira_date <= %s'
                    date_params = [date_from_s, date_to_s]

                cur.execute(f"""
                    SELECT
                        mapped_cr,
                        {_sel('cr_status')},
                        {_sel('cr_area')},
                        {_sel('cr_subsystem')},
                        {_sel('cr_functionality')},
                        {_sel('cr_occurrence', '0')} AS cr_occurrence,
                        {_sel('cr_title')},
                        {_sel('cr_date', 'NULL')}    AS cr_date,
                        {_sel('jira_date', 'NULL')}  AS jira_date,
                        {last_sel},
                        {_sel('jira_count', '0')}    AS jira_count,
                        {_sel('image')},
                        {_sel('cr_age', '0')}        AS cr_age,
                        {_sel('test_team', "'SWPDT'")} AS test_team
                    FROM {u_table}
                    WHERE mapped_cr IS NOT NULL
                      AND TRIM(mapped_cr) <> ''
                      {tt_clause}
                      {date_clause}
                """, date_params)
                for r in (cur.fetchall() or []):
                    row = dict(r)
                    row['target_name'] = pl_id
                    for k, v in list(row.items()):
                        if hasattr(v, 'isoformat'):
                            row[k] = v.isoformat()[:10]
                    result['pdt_crs'].append(row)
            except Exception as e:
                result['error'] = f'unique_crs: {e}'

        # â”€â”€ 2. overall_crs â†’ PDT Reported CRs â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Try: {base}_overallcrs  (e.g. kobuk_overallcrs)
        # Also try: {db_prefix}_overall_crs as fallback
        o_table = None
        for candidate in [
            f'`{schema}`.`{base_prefix}_overallcrs`',
            f'`{schema}`.`{db_prefix}_overall_crs`',
            f'`{schema}`.`{db_prefix}_overallcrs`',
        ]:
            if _tbl_exists_cur(cur, candidate):
                o_table = candidate
                break

        if o_table:
            result['overall_enabled'] = True
            try:
                cur.execute(f'SHOW COLUMNS FROM {o_table}')
                o_cols = frozenset(r['Field'].lower() for r in (cur.fetchall() or []))

                # Column name mapping: overallcrs uses different names
                # crid -> mapped_cr, date -> jira_date, subs -> cr_subsystem,
                # func -> cr_functionality, reported_team -> test_team, count -> cr_occurrence
                def _ocol(preferred, *fallbacks, alias=None, default="''"):
                    for c in (preferred,) + fallbacks:
                        if c.lower() in o_cols:
                            a = alias or preferred
                            return f'`{c}` AS `{a}`' if c != a else f'`{c}`'
                    a = alias or preferred
                    return f'{default} AS `{a}`'

                # Determine date column for filtering
                date_col = next((c for c in ('jira_date', 'date', 'cr_date') if c in o_cols), None)
                o_date_clause = ''
                o_date_params = []
                if date_col:
                    o_date_clause = f' AND `{date_col}` >= %s AND `{date_col}` <= %s'
                    o_date_params = [date_from_s, date_to_s]

                # CR id column
                cr_col = 'crid' if 'crid' in o_cols else ('mapped_cr' if 'mapped_cr' in o_cols else None)
                if not cr_col:
                    result['overall_enabled'] = False
                else:
                    cur.execute(f"""
                        SELECT
                            `{cr_col}` AS mapped_cr,
                            {_ocol('cr_status', 'status',   alias='cr_status')},
                            {_ocol('cr_area',   'area',     alias='cr_area')},
                            {_ocol('cr_subsystem', 'subs',  alias='cr_subsystem')},
                            {_ocol('cr_functionality', 'func', alias='cr_functionality')},
                            {_ocol('cr_occurrence', 'count', alias='cr_occurrence', default='0')},
                            {_ocol('cr_title', 'label',     alias='cr_title')},
                            {_ocol('cr_date', 'date', 'jira_date', alias='cr_date', default='NULL')},
                            {_ocol('jira_date', 'date',     alias='jira_date', default='NULL')},
                            {_ocol('reported_team', 'test_team', alias='test_team')}
                        FROM {o_table}
                        WHERE `{cr_col}` IS NOT NULL
                          AND TRIM(`{cr_col}`) <> ''
                          {o_date_clause}
                    """, o_date_params)
                    for r in (cur.fetchall() or []):
                        row = dict(r)
                        row['target_name'] = pl_id
                        for k, v in list(row.items()):
                            if hasattr(v, 'isoformat'):
                                row[k] = v.isoformat()[:10]
                        result['overall_crs'].append(row)

                    seen = set()
                    for r in result['overall_crs']:
                        k = str(r.get('mapped_cr') or '').strip()
                        if k:
                            seen.add(k)
                    result['unique_cr_count'] = len(seen)
            except Exception as e:
                result['overall_enabled'] = False
                result['error'] = (result['error'] or '') + f' | overall_crs: {e}'

        # â”€â”€ 3. jiras â†’ total JIRAs â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        j_table = f'`{schema}`.`{db_prefix}_jiras`'
        if _tbl_exists_cur(cur, j_table):
            try:
                cur.execute(f'SHOW COLUMNS FROM {j_table}')
                j_cols = frozenset(r['Field'].lower() for r in (cur.fetchall() or []))
                j_date_clause = ''
                j_date_params = []
                if 'jira_date' in j_cols:
                    j_date_clause = 'WHERE jira_date >= %s AND jira_date <= %s'
                    j_date_params = [date_from_s, date_to_s]
                cur.execute(f"""
                    SELECT COUNT(DISTINCT stability_ticket) AS cnt
                    FROM {j_table}
                    {j_date_clause}
                """, j_date_params)
                row = cur.fetchone()
                result['total_jiras'] = int((row or {}).get('cnt') or 0)
            except Exception:
                pass

        # â”€â”€ 4. openjiras â†’ open JIRAs â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        oj_table = f'`{schema}`.`{db_prefix}_openjiras`'
        if _tbl_exists_cur(cur, oj_table):
            try:
                cur.execute(f'SHOW COLUMNS FROM {oj_table}')
                oj_cols = frozenset(r['Field'].lower() for r in (cur.fetchall() or []))
                oj_date_clause = ''
                oj_date_params = []
                if 'jira_date' in oj_cols:
                    oj_date_clause = 'WHERE jira_date >= %s AND jira_date <= %s'
                    oj_date_params = [date_from_s, date_to_s]
                cur.execute(f"""
                    SELECT COUNT(DISTINCT stability_ticket) AS cnt
                    FROM {oj_table}
                    {oj_date_clause}
                """, oj_date_params)
                row = cur.fetchone()
                result['open_jiras'] = int((row or {}).get('cnt') or 0)
            except Exception:
                pass

        cur.close()
    except Exception as e:
        result['error'] = str(e)

    return result


def _monthly_fetch_all_targets(
    bu: str, date_from, date_to, include_hwpdt: bool = False
) -> list:
    """
    Fetch CR/JIRA data for all pl_ids in the consolidate table for the selected BU.
    Uses pl_id -> db_prefix mapping (dots to underscores).
    """
    consolidate = _monthly_fetch_consolidate(bu, date_from, date_to)
    if not consolidate:
        return []

    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []

    date_from_s = date_from.isoformat()
    date_to_s   = date_to.isoformat()
    results     = []
    seen_pl_ids = set()

    for row in consolidate:
        pl_id  = str(row.get('pl_id')  or '').strip()
        target = str(row.get('target') or '').strip()
        bu_row = str(row.get('bu')     or bu).strip().upper()
        if not pl_id or pl_id in seen_pl_ids:
            continue
        seen_pl_ids.add(pl_id)
        try:
            data = _monthly_fetch_target_cr_data(
                conn, bu_row, target, pl_id,
                date_from_s, date_to_s, include_hwpdt
            )
            if data:
                results.append(data)
        except Exception:
            continue

    try:
        conn.close()
    except Exception:
        pass

    return results


def _monthly_fetch_all_targets(
    bu: str, date_from, date_to, include_hwpdt: bool = False
) -> list:
    """
    Fetch CR/JIRA data for all targets in the selected BU.
    Returns list of per-target dicts from _monthly_fetch_target_cr_data.
    """
    try:
        from dashboard_common import get_targets_config, get_business_units
    except Exception:
        return []

    try:
        bus     = get_business_units() or {}
        tgt_cfg = get_targets_config() or {}
        if bu and bu != 'ALL':
            bu_up   = bu.upper()
            bu_info = bus.get(bu_up) or {}
            targets = list(bu_info.get('targets') or [])
            if not targets:
                targets = [
                    k for k, v in tgt_cfg.items()
                    if str((v or {}).get('bu', '')).upper() == bu_up
                ]
        else:
            targets = list(tgt_cfg.keys())
    except Exception:
        return []

    if not targets:
        return []

    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []

    date_from_s = date_from.isoformat()
    date_to_s   = date_to.isoformat()
    results     = []

    for target_name in targets:
        try:
            data = _monthly_fetch_target_cr_data(
                conn, target_name, date_from_s, date_to_s, include_hwpdt
            )
            if data:
                results.append(data)
        except Exception:
            continue

    try:
        conn.close()
    except Exception:
        pass

    return results


def _monthly_build_status_table(consolidate_rows: list, target_data: list) -> list:
    """
    Build the 'PDT WBC Target-wise Test Status' table.
    Merges consolidate summary (hours/builds/devices/crashes/mtbf)
    with live CR/JIRA counts from target_data.

    Columns:
      S.No | PL ID | No. of Devices | No. of Builds | Total Hours
      | Total CRs by PDT | Unique CRs by PDT | Total Crashes
      | Total JIRAs | Open JIRAs | Overall PDT CRs (enabled/not)
    """
    # Build lookup from target_data keyed by target_name
    td_map = {d['pl_id']: d for d in (target_data or []) if d.get('pl_id')}

    # Build base rows from consolidate
    rows = []
    for r in consolidate_rows:
        tgt  = str(r.get('target') or '')
        pl_id_key = str(r.get('pl_id') or tgt)
        td   = td_map.get(pl_id_key) or td_map.get(tgt) or {}
        pdt_crs     = td.get('pdt_crs') or []
        overall_crs = td.get('overall_crs') or []
        overall_en  = td.get('overall_enabled', False)

        # unique PDT CRs = distinct mapped_cr from unique_crs
        seen_pdt = set()
        for cr in pdt_crs:
            k = str(cr.get('mapped_cr') or '').strip()
            if k:
                seen_pdt.add(k)

        rows.append({
            'pl_id':            str(r.get('pl_id') or r.get('target') or ''),
            'target':           tgt,
            'bu':               str(r.get('bu') or ''),
            'devices':          int(float(r.get('number_of_devices') or 0)),
            'builds':           int(float(r.get('number_of_builds') or 0)),
            'hours':            round(float(r.get('total_hours') or 0), 1),
            'crashes':          int(float(r.get('total_crashes') or 0)),
            'mtbf':             round(float(r.get('mtbf') or 0), 2),
            'pdt_test_status':  str(r.get('pdt_test_status') or ''),
            'timelines':        str(r.get('timelines') or ''),
            # CR/JIRA counts from live tables
            'total_pdt_crs':    len(pdt_crs),
            'unique_pdt_crs':   len(seen_pdt),
            'total_jiras':      td.get('total_jiras', 0),
            'open_jiras':       td.get('open_jiras', 0),
            # overall_crs (PDT Reported column)
            'overall_cr_count': td.get('unique_cr_count', 0),
            'overall_enabled':  overall_en,
        })
    return rows


def _monthly_build_cr_summary(target_data: list) -> dict:
    """
    Build aggregated CR summary across all targets:
      - pdt_crs:     all rows from unique_crs (flat list)
      - overall_crs: all rows from overall_crs (flat list)
      - by_target:   per-target summary dict
    """
    all_pdt     = []
    all_overall = []
    by_target   = []

    for td in (target_data or []):
        tgt         = td.get('pl_id') or td.get('target', '')
        pdt_crs     = td.get('pdt_crs') or []
        overall_crs = td.get('overall_crs') or []
        overall_en  = td.get('overall_enabled', False)

        seen_pdt = set()
        for cr in pdt_crs:
            k = str(cr.get('mapped_cr') or '').strip()
            if k:
                seen_pdt.add(k)

        all_pdt.extend(pdt_crs)
        all_overall.extend(overall_crs)

        by_target.append({
            'target':           tgt,
            'total_pdt_crs':    len(pdt_crs),
            'unique_pdt_crs':   len(seen_pdt),
            'overall_cr_count': td.get('unique_cr_count', 0),
            'overall_enabled':  overall_en,
            'total_jiras':      td.get('total_jiras', 0),
            'open_jiras':       td.get('open_jiras', 0),
        })

    # Deduplicate overall for unique count
    seen_overall = set()
    unique_overall = []
    for r in all_overall:
        k = str(r.get('mapped_cr') or '').strip()
        if k and k not in seen_overall:
            seen_overall.add(k)
            unique_overall.append(r)

    return {
        'pdt_crs':        all_pdt,
        'overall_crs':    all_overall,
        'unique_overall': unique_overall,
        'by_target':      sorted(by_target, key=lambda x: -x['total_pdt_crs']),
    }


def _monthly_build_area_chart(cr_rows: list) -> dict:
    """Build CR-area bar chart data: overall + per-target."""
    from collections import defaultdict, Counter
    area_total = Counter()
    by_target  = defaultdict(Counter)
    for r in cr_rows:
        area   = str(r.get('cr_area') or 'Unknown').strip() or 'Unknown'
        target = str(r.get('target_name') or 'Unknown').strip()
        area_total[area] += 1
        by_target[target][area] += 1
    overall = [{'area': a, 'count': c}
               for a, c in sorted(area_total.items(), key=lambda x: -x[1])]
    per_tgt = {
        t: [{'area': a, 'count': c}
            for a, c in sorted(cnts.items(), key=lambda x: -x[1])]
        for t, cnts in by_target.items()
    }
    return {'overall': overall, 'by_target': per_tgt}


def _monthly_build_stability_trend(build_rows: list) -> dict:
    """Build stability trend data per target: [{build_label, hours, mtbf, crashes}]."""
    from collections import defaultdict
    by_target = defaultdict(list)
    for r in build_rows:
        tgt     = str(r.get('target') or '').strip()
        if not tgt:
            continue
        hours   = float(r.get('hours') or 0)
        crashes = int(float(r.get('crash_count') or 0))
        mtbf    = float(r.get('mtbf') or 0)
        if not mtbf and hours > 0 and crashes > 0:
            mtbf = round(hours / crashes, 2)
        label   = str(r.get('meta_build') or r.get('build_label') or '').strip()
        by_target[tgt].append({
            'build_label': label,
            'hours':       round(hours, 1),
            'mtbf':        round(mtbf, 2) if crashes > 0 else None,
            'crashes':     crashes,
            'week_end':    str(r.get('week_end') or ''),
        })
    return dict(by_target)




@weekly_summary_bp.route('/monthly-report')
@login_required
def monthly_report_landing():
    """Monthly BU report landing page â€” rendered inside the BU shell sidebar."""
    import time as _time
    try:
        from dashboard_routes import _build_bu_shell_context
        shell_ctx = _build_bu_shell_context('MONTHLY_REPORT')
    except Exception:
        shell_ctx = {'active_bu_key': 'MONTHLY_REPORT', 'bu_list': [],
                     'BU_ICONS': {}, 'shell_title': 'Monthly BU Report'}
    shell_ctx['shell_title']  = 'Monthly BU Report'
    shell_ctx['cache_buster'] = int(_time.time())

    date_from, date_to = _monthly_date_range_from_request()
    sel_bu             = (request.args.get('bu') or 'ALL').upper()
    return render_template(
        'monthly_report.html',
        sel_bu=sel_bu,
        date_from=date_from.isoformat(),
        date_to=date_to.isoformat(),
        mr_bu_list=_monthly_bu_list(),
        **shell_ctx,
    )


@weekly_summary_bp.route('/api/monthly-report/data')
@login_required
def api_monthly_report_data():
    """
    Return all monthly report data as JSON.
    Params: bu, date_from, date_to, include_hwpdt (1/0)

    Response shape:
    {
      success, bu, date_from, date_to, include_hwpdt,
      status_table: [{
        pl_id, target, bu, devices, builds, hours, crashes, mtbf,
        total_pdt_crs, unique_pdt_crs,   <- from unique_crs table
        total_jiras, open_jiras,          <- from jiras / openjiras tables
        overall_cr_count, overall_enabled <- from overall_crs table
      }],
      trend: {target: [{build_label, hours, mtbf, crashes, week_end}]},
      pdt_crs: [...],          <- all rows from unique_crs (flat)
      overall_crs: [...],      <- all rows from overall_crs (flat)
      unique_overall: [...],   <- deduplicated overall_crs
      by_target: [{            <- per-target CR/JIRA summary
        target, total_pdt_crs, unique_pdt_crs,
        overall_cr_count, overall_enabled,
        total_jiras, open_jiras
      }],
      pdt_area_chart: {overall, by_target},
      overall_area_chart: {overall, by_target},
      totals: {
        total_pdt_crs, unique_pdt_crs,
        total_jiras, open_jiras,
        overall_cr_count, targets_with_overall
      }
    }
    """
    bu            = (request.args.get('bu') or 'ALL').upper()
    date_from, date_to = _monthly_date_range_from_request()
    include_hwpdt = request.args.get('include_hwpdt', '0') == '1'

    # 1. Consolidate summary (hours/builds/devices/crashes from sharepoint tables)
    consolidate  = _monthly_fetch_consolidate(bu, date_from, date_to)

    # 2. Build rows (stability trend)
    build_rows   = _monthly_fetch_build_rows(bu, date_from, date_to)
    trend        = _monthly_build_stability_trend(build_rows)

    # 3. Per-target CR/JIRA data from live tables
    target_data  = _monthly_fetch_all_targets(bu, date_from, date_to, include_hwpdt)

    # 4. Status table (merges consolidate + live CR/JIRA counts)
    status_table = _monthly_build_status_table(consolidate, target_data)

    # 5. CR summary
    cr_summary   = _monthly_build_cr_summary(target_data)

    # 6. Area charts
    pdt_area_chart     = _monthly_build_area_chart(cr_summary['pdt_crs'])
    overall_area_chart = _monthly_build_area_chart(cr_summary['overall_crs'])

    # 7. Totals
    totals = {
        'total_pdt_crs':         sum(r.get('total_pdt_crs', 0)   for r in status_table),
        'unique_pdt_crs':        sum(r.get('unique_pdt_crs', 0)  for r in status_table),
        'total_jiras':           sum(r.get('total_jiras', 0)     for r in status_table),
        'open_jiras':            sum(r.get('open_jiras', 0)      for r in status_table),
        'overall_cr_count':      sum(r.get('overall_cr_count', 0) for r in status_table),
        'targets_with_overall':  sum(1 for r in status_table if r.get('overall_enabled')),
        'total_targets':         len(status_table),
    }

    return jsonify({
        'success':            True,
        'bu':                 bu,
        'date_from':          date_from.isoformat(),
        'date_to':            date_to.isoformat(),
        'include_hwpdt':      include_hwpdt,
        'status_table':       status_table,
        'trend':              trend,
        'pdt_crs':            cr_summary['pdt_crs'],
        'overall_crs':        cr_summary['overall_crs'],
        'unique_overall':     cr_summary['unique_overall'],
        'by_target':          cr_summary['by_target'],
        'pdt_area_chart':     pdt_area_chart,
        'overall_area_chart': overall_area_chart,
        'totals':             totals,
    })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# SHAREPOINT 2 â€” Smart Build Report (auto-populated, no manual entry)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•


# ---------------------------------------------------------------------------
# Smart Build Report -- sp2_build_consolidate helpers
# ---------------------------------------------------------------------------


def _sp2_meta_build_key(bn: str) -> str:
    """Compute the meta-id for a build name.

    Rule: meta-key = Target.Version-[rN-]NNNNN[.NN]
    Everything after the 5-digit build number (flavour, _tags, suffixes) is stripped.
    Builds sharing the same target+version+number are ONE meta, regardless of flavour.

    Examples:
        Aldabra.LA.1.0-00291-STD.MAG.INT-1                           -> Aldabra.LA.1.0-00291
        Aldabra.LA.1.0-00291-PERF.INT-1                              -> Aldabra.LA.1.0-00291  (same meta)
        Aldabra.LA.1.0-00293-STD.INT-1                               -> Aldabra.LA.1.0-00293
        Aldabra.LA.1.0-00293-PERF.MAG.INT-1                          -> Aldabra.LA.1.0-00293  (same meta)
        Aldabra.LA.1.0-00293-PERF.INT-1                              -> Aldabra.LA.1.0-00293  (same meta)
        Aldabra.LA.1.0-00293.01-PERF.MAG.INT_QoSNThrottlesettings    -> Aldabra.LA.1.0-00293.01
        Aldabra.LA.1.0-00293.01-PERF.INT_QoSNThrottlesettings_0619   -> Aldabra.LA.1.0-00293.01  (same meta)
        Maili.LA.1.0-r1-00117-STD.INT-1_0619_M2_VP                   -> Maili.LA.1.0-r1-00117
        CQ2390.LA.1.0-00287-PERF.INT-1_power_val                     -> CQ2390.LA.1.0-00287
        Aldabra.LA.1.0-00268.10-STD.INT-1_Audio                      -> Aldabra.LA.1.0-00268.10
    """
    import re as _re_mbk
    s = str(bn or '').strip()
    # Capture: anything-[rN-]5digits[.NN] then drop everything after (flavour + _tags).
    m = _re_mbk.match(r'^(.+?-(?:r\d+-)?\d{5}(?:\.\d+)?)(?:[-_].+)?$', s)
    return m.group(1) if m else s
def _ensure_sp2_build_consolidate_table():
    """Create sp2_build_consolidate table if it does not exist, and add new columns."""
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return
    cur = conn.cursor()
    try:
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}` (
                id              BIGINT AUTO_INCREMENT PRIMARY KEY,
                week_start      DATE         NOT NULL,
                week_end        DATE         NOT NULL,
                target          VARCHAR(255) NOT NULL DEFAULT '',
                pl_id           VARCHAR(255) NOT NULL DEFAULT '',
                build_name      VARCHAR(512) NOT NULL DEFAULT '',
                build_type      VARCHAR(16)  NOT NULL DEFAULT 'CRM',
                total_hours     DECIMAL(12,3) NULL,
                total_crashes   INT          NULL,
                device_count    INT          NULL,
                chip_ids        JSON         NULL,
                bu              VARCHAR(128) NULL,
                timelines       TEXT         NULL,
                pdt_test_status VARCHAR(64)  NULL,
                unique_crs      INT          NULL,
                number_of_builds INT         NULL,
                updated_by      VARCHAR(128) NULL,
                updated_at      DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP
                                ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_sp2_build (week_start, week_end, build_name(255), pl_id(128)),
                KEY idx_sp2_week (week_start, week_end),
                KEY idx_sp2_target (target(64))
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
        # Add new columns to existing tables (safe ALTER â€” ignored if already exist)
        for _col_sql in [
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}` ADD COLUMN bu VARCHAR(128) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}` ADD COLUMN timelines TEXT NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}` ADD COLUMN pdt_test_status VARCHAR(64) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}` ADD COLUMN unique_crs INT NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}` ADD COLUMN number_of_builds INT NULL",
        ]:
            try:
                cur.execute(_col_sql)
                conn.commit()
            except Exception:
                pass  # column already exists
    except Exception:
        pass
    finally:
        cur.close(); conn.close()


def _ensure_sp2_build_type_overrides_table():
    """Separate table that persists per-build-name CRM/Eng overrides.
    Survives consolidate rebuilds (which DELETE+INSERT sp2_build_consolidate).
    """
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return
    cur = conn.cursor()
    try:
        cur.execute(f"""
            CREATE TABLE IF NOT EXISTS `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` (
                id          BIGINT AUTO_INCREMENT PRIMARY KEY,
                week_start  DATE         NOT NULL,
                week_end    DATE         NOT NULL,
                build_name  VARCHAR(512) NOT NULL DEFAULT '',
                pl_id       VARCHAR(255) NOT NULL DEFAULT '',
                build_type  VARCHAR(16)  NOT NULL DEFAULT 'CRM',
                updated_by  VARCHAR(128) NULL,
                updated_at  DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP
                            ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uq_sp2_override (week_start, week_end, build_name(255), pl_id(128))
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        conn.commit()
    except Exception:
        pass
    finally:
        cur.close(); conn.close()


def _upsert_sp2_build_type(ws, we, build_name: str, pl_id: str, build_type: str):
    """Persist build_type override into the dedicated overrides table.
    This table is never wiped by consolidate rebuilds.
    """
    _ensure_sp2_build_type_overrides_table()
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return
    cur = conn.cursor()
    try:
        cur.execute(f"""
            INSERT INTO `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                (week_start, week_end, build_name, pl_id, build_type, updated_by)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                build_type=VALUES(build_type),
                updated_by=VALUES(updated_by),
                updated_at=CURRENT_TIMESTAMP
        """, (ws.isoformat(), we.isoformat(), build_name, pl_id, build_type, _current_user_identifier()))
        conn.commit()
    finally:
        cur.close(); conn.close()


def _ensure_sp2_override_snapshot_columns():
    """Migrate build_type_overrides into the static Smart Build source table.

    The table started as a small CRM/Eng override table.  Smart Build now uses it
    as the one-time weekly snapshot built from CSV crash data + axiom_job_summary;
    user edits stay here and later consolidate refreshes read this table instead
    of refetching/recomputing from CSV.
    """
    _ensure_sp2_build_type_overrides_table()
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return
    cur = conn.cursor()
    try:
        for sql in (
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN target VARCHAR(255) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN bu VARCHAR(128) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN build_id VARCHAR(512) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN job_ids TEXT NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN submitted_at DATETIME NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN completed_at DATETIME NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN state VARCHAR(64) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN device_count INT NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN chip_ids LONGTEXT NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN hours DECIMAL(14,3) NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN total_crashes INT NULL",
            f"ALTER TABLE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}` ADD COLUMN source VARCHAR(64) NULL DEFAULT 'axiom_csv_snapshot'",
        ):
            try:
                cur.execute(sql)
                conn.commit()
            except Exception:
                pass
    finally:
        cur.close(); conn.close()


def _sp2_static_snapshot_count(ws, we) -> int:
    _ensure_sp2_override_snapshot_columns()
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return 0
    cur = conn.cursor()
    try:
        cur.execute(f"""
            SELECT COUNT(*)
            FROM `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
            WHERE week_start=%s AND week_end=%s AND hours IS NOT NULL
        """, (ws.isoformat(), we.isoformat()))
        row = cur.fetchone()
        return int((row[0] if isinstance(row, (tuple, list)) else list(row.values())[0]) or 0)
    except Exception:
        return 0
    finally:
        cur.close(); conn.close()


def _load_sp2_static_build_rows(ws, we) -> list:
    """Read the static weekly Smart Build snapshot from build_type_overrides."""
    _ensure_sp2_override_snapshot_columns()
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT target, pl_id, bu, build_name, build_id, job_ids, submitted_at,
                   completed_at, state, device_count, chip_ids, hours,
                   total_crashes, build_type
            FROM `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
            WHERE week_start=%s AND week_end=%s AND hours IS NOT NULL
            ORDER BY target, pl_id, submitted_at, build_name
        """, (ws.isoformat(), we.isoformat()))
        return cur.fetchall() or []
    except Exception:
        return []
    finally:
        cur.close(); conn.close()


def _seed_sp2_build_type_overrides_from_axiom(ws, we, username: str = '') -> int:
    """One-time weekly seed of Smart Build static rows.

    Reads axiom_job_summary plus weekly CSV crash data and inserts missing rows
    into sp2_build_type_overrides. Existing rows are never overwritten, so user
    CRM/Eng, hour, crash, or BU edits remain static until explicitly changed.
    """
    _ensure_sp2_override_snapshot_columns()

    # Insert missing snapshot rows every time. INSERT IGNORE below preserves any
    # existing/static user-edited rows, but this avoids a partial snapshot (for
    # example from an earlier failed seed) permanently limiting Consolidate.
    # Seed only after the weekly CRM CSV data is available. Until then the page
    # can still show live Axiom rows, but no static snapshot is frozen.
    conn_chk = get_mysql_connection_db(bu_key=None)
    if not conn_chk:
        return 0
    cur_chk = conn_chk.cursor()
    try:
        cur_chk.execute(f"SELECT COUNT(*) FROM `{_QIPL_DB}`.`{_QIPL_TABLE}` WHERE week_start=%s AND week_end=%s", (ws.isoformat(), we.isoformat()))
        row = cur_chk.fetchone()
        if int((row[0] if isinstance(row, (tuple, list)) else list(row.values())[0]) or 0) <= 0:
            return 0
    except Exception:
        return 0
    finally:
        cur_chk.close(); conn_chk.close()

        # Crash counts from the same CRM CSV import used by the CR cards.
    crash_map = _sp2_weekly_crash_map(ws, we)


    # Group duplicate Axiom job runs into one static build row.
    import re as _re_seed
    def _pl_group(sp):
        return _re_seed.sub(r'\.r\d+$', '', str(sp or ''), flags=_re_seed.IGNORECASE)

    grouped = {}
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return 0
    cur = conn.cursor(dictionary=True)
    cur2 = None
    try:
        _week_cap = we.isoformat() + " 23:59:59"
        _week_floor = ws.isoformat() + " 00:00:00"
        live_h = (
            "CASE"
            " WHEN state IN ('Running','JobSetup') AND started_at IS NOT NULL"
            " THEN ROUND(device_count *"
            " TIMESTAMPDIFF(SECOND, started_at,"
            " LEAST(NOW(), TIMESTAMP('" + _week_cap + "'))) / 3600.0 * 0.80, 3)"
            " WHEN state IN ('Completed','Aborted') AND started_at IS NOT NULL AND ended_at IS NOT NULL"
            " THEN ROUND(device_count *"
            " TIMESTAMPDIFF(SECOND,"
            " GREATEST(started_at, TIMESTAMP('" + _week_floor + "')),"
            " LEAST(ended_at,      TIMESTAMP('" + _week_cap   + "'))) / 3600.0 * 0.80, 3)"
            " ELSE 0 END"
        )
        cur.execute(f"""
            SELECT job_id, build_id, build_name, software_product,
                   state, device_count, chip_ids, submitted_at, ended_at,
                   submitter, ({live_h}) AS hours_live
            FROM `pdt_stats_dashboard`.`axiom_job_summary`
            WHERE taxonomy_path = '/PDT/QIPL'
              AND started_at <= %s AND (ended_at >= %s OR state IN ('Running','JobSetup'))
            ORDER BY submitted_at
        """, (we.isoformat(), ws.isoformat()))
        for r in cur.fetchall() or []:
            chips_raw = r.get('chip_ids') or '[]'
            if isinstance(chips_raw, str):
                try:
                    chips = json.loads(chips_raw)
                except Exception:
                    chips = []
            else:
                chips = list(chips_raw) if chips_raw else []
            _raw_dev = int(r.get('device_count') or 0)
            _raw_hrs = float(r.get('hours_live') or 0)
            if _raw_dev <= 0 and not chips and _raw_hrs <= 0.1:
                continue
            pl_exact = str(r.get('software_product') or '').strip()
            pl_id = _pl_group(pl_exact)
            target = _swpdt_target_from_product(pl_id) or pl_id
            build_id = str(r.get('build_id') or '').strip()
            build_name = str(r.get('build_name') or build_id).strip()
            if not build_name:
                continue
            key = (build_name.upper(), pl_id.upper())
            acc = grouped.setdefault(key, {
                'target': target, 'pl_id': pl_id, 'build_name': build_name,
                'build_id': build_id, 'job_ids': [], 'submitted_at': None,
                'completed_at': None, 'state': 'Completed', 'hours': 0.0,
                'chip_ids': set(), 'device_count': 0,
            })
            acc['job_ids'].append(str(r.get('job_id') or ''))
            acc['hours'] += _raw_hrs
            acc['chip_ids'].update(str(c).strip() for c in chips if str(c).strip())
            acc['device_count'] = max(acc['device_count'], _raw_dev)
            if str(r.get('state') or '').lower() in ('running', 'jobsetup'):
                acc['state'] = str(r.get('state') or 'Running')
            sub = r.get('submitted_at')
            end = r.get('ended_at')
            if sub and (not acc['submitted_at'] or sub < acc['submitted_at']):
                acc['submitted_at'] = sub
            if end and (not acc['completed_at'] or end > acc['completed_at']):
                acc['completed_at'] = end

        ins = 0
        cur2 = conn.cursor()
        _rows_to_insert = []
        for acc in grouped.values():
            crashes = _sp2_crash_count_for_build(
                crash_map,
                acc.get('build_name'),
                acc.get('build_id'),
                acc.get('pl_id'),
            )
            chips = sorted(acc['chip_ids'])
            _rows_to_insert.append((
                ws.isoformat(), we.isoformat(), acc['target'], acc['pl_id'],
                acc['build_name'], acc['build_id'], json.dumps(acc['job_ids']),
                acc['submitted_at'], acc['completed_at'], acc['state'],
                max(int(acc['device_count'] or 0), len(chips)), json.dumps(chips),
                round(float(acc['hours'] or 0), 3), int(crashes or 0),
                username or _current_user_identifier(),
            ))

        import time as _time
        _insert_sql = (
            f"INSERT INTO `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`"
            " (week_start, week_end, target, pl_id, build_name, build_id,"
            "  job_ids, submitted_at, completed_at, state, device_count,"
            "  chip_ids, hours, total_crashes, build_type, source, updated_by)"
            " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'CRM','axiom_csv_snapshot',%s)"
            " ON DUPLICATE KEY UPDATE"
            "  total_crashes=VALUES(total_crashes),"
            "  source=VALUES(source),"
            "  updated_by=VALUES(updated_by),"
            "  updated_at=CURRENT_TIMESTAMP"
        )
        for _attempt in range(3):
            try:
                for _row_params in _rows_to_insert:
                    cur2.execute(_insert_sql, _row_params)
                    ins += max(int(cur2.rowcount or 0), 0)
                conn.commit()
                break
            except Exception as _dl_exc:
                _errno = getattr(_dl_exc, "errno", None)
                try: conn.rollback()
                except Exception: pass
                if _errno == 1213 and _attempt < 2:
                    ins = 0
                    _time.sleep(0.5 * (2 ** _attempt))
                    continue
                raise
        return ins
    except Exception as exc:
        try: conn.rollback()
        except Exception: pass
        import logging as _log
        _log.getLogger('weekly_summary_routes').warning('[SP2 STATIC SEED] %s', exc, exc_info=True)
        return 0
    finally:
        try: cur.close()
        except Exception: pass
        try:
            if cur2: cur2.close()
        except Exception: pass
        conn.close()


def _build_and_save_sp2_consolidate_from_static(ws, we, username: str) -> bool:
    """Build consolidate sentinel rows from the static Smart Build snapshot."""
    static_rows = _load_sp2_static_build_rows(ws, we)
    if not static_rows:
        return False

        
    dash_map = _fetch_dashboard_status_map()
    previous_bu_map = _fetch_sp2_previous_bu_map(before_week_start=ws)

    try:
        _refresh_ucr_excel_from_latest_csv_if_needed(we)
        ucr_counts = _build_ucr_target_pl_count_map(we) or {}
    except Exception:
        ucr_counts = {}

    saved_bu = {}
    saved_tl = {}
    conn0 = get_mysql_connection_db(bu_key=None)
    if conn0:
        cur0 = conn0.cursor(dictionary=True)
        try:
            cur0.execute(f"""
                SELECT target, pl_id, bu, timelines, pdt_test_status
                FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                WHERE week_start=%s AND week_end=%s
            """, (ws.isoformat(), we.isoformat()))
            for r in cur0.fetchall() or []:
                key = (str(r.get('target') or '').strip().upper(), str(r.get('pl_id') or '').strip().upper())
                if str(r.get('bu') or '').strip():
                    saved_bu[key] = str(r.get('bu') or '').strip()
                    saved_bu[(key[0], '')] = str(r.get('bu') or '').strip()
                if str(r.get('timelines') or '').strip():
                    saved_tl[key] = r
        except Exception:
            pass
        finally:
            cur0.close(); conn0.close()

    grouped = {}
    order = []
    for r in static_rows:
        target = str(r.get('target') or '').strip() or (_swpdt_target_from_product(r.get('pl_id')) or '')
        pl_id = str(r.get('pl_id') or '').strip()
        key = (target.upper(), pl_id.upper())
        if key not in grouped:
            grouped[key] = {'target': target, 'pl_id': pl_id, 'bu': '', 'build_names': set(), 'hours': 0.0,
                            'crashes': 0, 'chip_ids_set': set(), 'device_count_sum': 0,
                            'bt_votes': {'CRM': 0, 'Eng': 0}}
            order.append(key)
        g = grouped[key]
        row_bu = str(r.get('bu') or '').strip()
        if row_bu and not g.get('bu'):
            g['bu'] = row_bu
        build_name = str(r.get('build_name') or r.get('build_id') or '').strip()
        bt = str(r.get('build_type') or 'CRM')
        if bt not in ('CRM', 'Eng'):
            bt = 'CRM'
        g['bt_votes'][bt] += 1
        chips_raw = r.get('chip_ids') or '[]'
        try:
            chips = json.loads(chips_raw) if isinstance(chips_raw, str) else list(chips_raw or [])
        except Exception:
            chips = []
        # Device pool is target-level and includes all builds. Later we split
        # each target's unique devices across its PLs by CRM hours so the
        # Consolidate total matches the Builds tab total device count.
        g['chip_ids_set'].update(str(c).strip() for c in chips if str(c).strip())
        g['device_count_sum'] = max(g['device_count_sum'], int(r.get('device_count') or 0), len(g['chip_ids_set']))
        # Consolidate metrics are CRM-only: Eng rows remain in the snapshot but
        # do not contribute builds/hours/crashes to the Consolidate report.
        if bt == 'CRM':
            if build_name:
                g['build_names'].add(_sp2_meta_build_key(build_name).upper())
            g['hours'] += float(r.get('hours') or 0)
            g['crashes'] += int(r.get('total_crashes') or 0)

    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return False
    cur = conn.cursor()
    try:
        cur.execute(f"""
            DELETE FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
            WHERE week_start=%s AND week_end=%s AND build_name LIKE '__consolidated__%'
        """, (ws.isoformat(), we.isoformat()))
        # Target-level BU: if a target has a BU, all of its PLs use the same BU.
        # Milestones remain PL-specific and are resolved from dashboard_status.sp_name
        # matching the PL name.
        target_bu_map = {}
        target_keys = {}
        for _key in order:
            _g = grouped[_key]
            _tu = str(_g['target'] or '').strip().upper()
            target_keys.setdefault(_tu, []).append(_key)
        for _tu, _keys in target_keys.items():
            _first = grouped[_keys[0]]
            _dash_tgt = _match_dashboard_with_fallback(_first['target'], dash_map) or {}
            _bu = (
                saved_bu.get((_tu, ''))
                or str(_first.get('bu') or '').strip()
                or previous_bu_map.get((_tu, ''))
                or str(_dash_tgt.get('bu') or '').strip()
            )
            if not _bu:
                for _key in _keys:
                    _pl_dash = _match_dashboard_with_fallback(grouped[_key]['pl_id'], dash_map) or {}
                    _bu = (
                        str(grouped[_key].get('bu') or '').strip()
                        or previous_bu_map.get(_key)
                        or previous_bu_map.get((_key[0], ''))
                        or str(_pl_dash.get('bu') or '').strip()
                    )
                    if _bu:
                        break
            target_bu_map[_tu] = _normalize_bu(_bu)

        # Device distribution: use each target's unique device pool and split it
        # across that target's PL rows by CRM hours. This preserves totals:
        # sum(consolidate.device_count) == Builds tab total unique devices.
        device_count_by_key = {}
        for _tu, _keys in target_keys.items():
            _chips = set()
            _total_hours = 0.0
            for _key in _keys:
                _chips.update(grouped[_key].get('chip_ids_set') or set())
                _total_hours += float(grouped[_key].get('hours') or 0)
            _total_devices = len(_chips)
            if _total_devices <= 0:
                for _key in _keys:
                    device_count_by_key[_key] = 0
                continue
            if len(_keys) == 1:
                device_count_by_key[_keys[0]] = _total_devices
                continue
            if _total_hours > 0:
                _raw = {_key: (_total_devices * float(grouped[_key].get('hours') or 0) / _total_hours) for _key in _keys}
            else:
                _raw = {_key: (_total_devices / len(_keys)) for _key in _keys}
            _base = {_key: int(_raw[_key]) for _key in _keys}
            _assigned = sum(_base.values())
            _remainders = sorted(_keys, key=lambda k: (_raw[k] - int(_raw[k]), float(grouped[k].get('hours') or 0)), reverse=True)
            for _key in _remainders:
                if _assigned >= _total_devices:
                    break
                _base[_key] += 1
                _assigned += 1
            device_count_by_key.update(_base)

        for key in order:
            g = grouped[key]
            votes = g['bt_votes']
            build_type = 'Eng' if votes.get('Eng', 0) > votes.get('CRM', 0) else 'CRM'
            dash_tgt = _match_dashboard_with_fallback(g['target'], dash_map) or {}
            dash_pl = _match_dashboard_with_fallback(g['pl_id'], dash_map) or {}
            bu = _normalize_bu(
                saved_bu.get(key)
                or target_bu_map.get(str(g['target'] or '').strip().upper())
                or previous_bu_map.get(key)
                or previous_bu_map.get((str(g['target'] or '').strip().upper(), ''))
                or str(dash_tgt.get('bu') or dash_pl.get('bu') or '').strip()
            )
            if bu:
                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                    SET bu=%s, updated_at=CURRENT_TIMESTAMP
                    WHERE week_start=%s AND week_end=%s
                      AND target=%s AND pl_id=%s
                """, (bu, ws.isoformat(), we.isoformat(), g['target'], g['pl_id']))
            # Milestones are PL-specific: PL name == dashboard_status.sp_name.
            milestone_dash = dash_pl or dash_tgt or {}
            es = _fmt_iso_date(milestone_dash.get('ES')) if milestone_dash.get('ES') else ''
            fc = _fmt_iso_date(milestone_dash.get('FC')) if milestone_dash.get('FC') else ''
            cs = _fmt_iso_date(milestone_dash.get('CS')) if milestone_dash.get('CS') else ''
            tl_saved = saved_tl.get(key) or {}
            if str(tl_saved.get('timelines') or '').strip():
                # Manual/saved timelines override dashboard/OneView so user edits
                # survive Refresh & Save rebuilds.
                timelines = str(tl_saved.get('timelines') or '')
                pdt_status = str(tl_saved.get('pdt_test_status') or '')
            elif es or fc or cs:
                timelines = _sp_timeline(es, fc, cs)
                pdt_status = _compute_pdt_test_status(es, cs, fc)
            else:
                timelines = ''
                pdt_status = ''
            chip_ids = sorted(g['chip_ids_set'])
            cur.execute(f"""
                INSERT INTO `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                    (week_start, week_end, target, pl_id, build_name, build_type,
                     total_hours, total_crashes, device_count, chip_ids,
                     bu, timelines, pdt_test_status, unique_crs,
                     number_of_builds, updated_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    total_hours=VALUES(total_hours), total_crashes=VALUES(total_crashes),
                    device_count=VALUES(device_count), chip_ids=VALUES(chip_ids),
                    build_type=VALUES(build_type), bu=VALUES(bu), timelines=VALUES(timelines),
                    pdt_test_status=VALUES(pdt_test_status), unique_crs=VALUES(unique_crs),
                    number_of_builds=VALUES(number_of_builds), updated_by=VALUES(updated_by),
                    updated_at=CURRENT_TIMESTAMP
            """, (ws.isoformat(), we.isoformat(), g['target'], g['pl_id'],
                  f"__consolidated__{g['target']}__{g['pl_id']}", build_type,
                  round(g['hours'], 3), int(g['crashes'] or 0),
                  int(device_count_by_key.get(key, 0)), json.dumps(chip_ids),
                  bu, timelines, pdt_status, _ucr_count_for_sharepoint_pair(ucr_counts, g['target'], g['pl_id']),
                  len(g['build_names']), username))
        conn.commit()
        return True
    except Exception as exc:
        try: conn.rollback()
        except Exception: pass
        import logging as _log
        _log.getLogger('weekly_summary_routes').error('[SP2 STATIC CONSOLIDATE] %s', exc, exc_info=True)
        return False
    finally:
        cur.close(); conn.close()


def _build_and_save_sp2_consolidate(ws, we, username: str):
    """Rebuild sp2_build_consolidate â€” ONE ROW per (target, pl_id).

    Build count  = distinct build_names (deduped set).
    BU/Timelines = from pdt_stats_dashboard.dashboard_status first,
                   then OneView milestone resolver, then NULL.
    Overrides    = read from sp2_build_type_overrides (never wiped).
    """
    import re as _re
    _ensure_sp2_build_consolidate_table()
    _ensure_sp2_override_snapshot_columns()
    _seed_sp2_build_type_overrides_from_axiom(ws, we, username)
    if _build_and_save_sp2_consolidate_from_static(ws, we, username):
        return

    # ------------------------------------------------------------------ #
    # 1. Axiom jobs for the week
    # ------------------------------------------------------------------ #
    db_rows = []
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if conn:
            cur = conn.cursor(dictionary=True)
            try:
                _week_cap = we.isoformat() + " 23:59:59"
                _week_floor = ws.isoformat() + " 00:00:00"
                live_h = (
                    "CASE"
                    " WHEN state IN ('Running','JobSetup') AND started_at IS NOT NULL"
                    " THEN ROUND(device_count *"
                    " TIMESTAMPDIFF(SECOND, started_at,"
                    " LEAST(NOW(), TIMESTAMP('" + _week_cap + "'))) / 3600.0 * 0.80, 3)"
                    " WHEN state IN ('Completed','Aborted') AND started_at IS NOT NULL AND ended_at IS NOT NULL"
                    " THEN ROUND(device_count *"
                    " TIMESTAMPDIFF(SECOND,"
                    " GREATEST(started_at, TIMESTAMP('" + _week_floor + "')),"
                    " LEAST(ended_at,      TIMESTAMP('" + _week_cap   + "'))) / 3600.0 * 0.80, 3)"
                    " ELSE 0 END"
                )
                cur.execute(f"""
                    SELECT job_id, build_id, build_name, software_product,
                           state, device_count, chip_ids,
                           ({live_h}) AS hours_live
                    FROM `pdt_stats_dashboard`.`axiom_job_summary`
                    WHERE taxonomy_path = '/PDT/QIPL'
                      AND started_at <= %s AND (ended_at >= %s OR state IN ('Running','JobSetup'))
                """, (we.isoformat(), ws.isoformat()))
                db_rows = cur.fetchall() or []
                import logging as _log_dbg
                _log_dbg.getLogger('weekly_summary_routes').info(
                    '[SP2 CONSOLIDATE] Axiom rows fetched: %d for %s to %s',
                    len(db_rows), ws.isoformat(), we.isoformat())
            finally:
                cur.close(); conn.close()
    except Exception as _axiom_exc:
        import logging as _log_ax
        _log_ax.getLogger('weekly_summary_routes').error('[SP2 CONSOLIDATE] Axiom query failed: %s', _axiom_exc)
        db_rows = []

        # ------------------------------------------------------------------ #
    # 2. Week + PL bounded crash map from weekly_qipl_data
    # ------------------------------------------------------------------ #
    crash_map = _sp2_weekly_crash_map(ws, we)


    # ------------------------------------------------------------------ #
    # 3. Build-type overrides from dedicated overrides table
    # ------------------------------------------------------------------ #
    build_type_map = {}   # (build_name_upper, pl_id_upper) -> 'CRM'|'Eng'
    try:
        conn3 = get_mysql_connection_db(bu_key=None)
        if conn3:
            cur3 = conn3.cursor(dictionary=True)
            try:
                cur3.execute(f"""
                    SELECT build_name, pl_id, build_type
                    FROM `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                    WHERE week_start=%s AND week_end=%s
                """, (ws.isoformat(), we.isoformat()))
                for row in cur3.fetchall() or []:
                    k = (str(row.get('build_name') or '').strip().upper(),
                         str(row.get('pl_id') or '').strip().upper())
                    build_type_map[k] = str(row.get('build_type') or 'CRM')
            except Exception:
                pass
            finally:
                try: cur3.close()
                except Exception: pass
                try: conn3.close()
                except Exception: pass
    except Exception:
        pass

    # ------------------------------------------------------------------ #
    # 4. BU / Timelines source 1: dashboard_status (fast, primary)
    # ------------------------------------------------------------------ #
    dash_map = _fetch_dashboard_status_map()   # {TARGET_UPPER: {bu, ES, FC, CS}}

    # ------------------------------------------------------------------ #
    # 5. Unique CRs from UCR Excel
    # ------------------------------------------------------------------ #
    ucr_counts = {}
    try:
        _refresh_ucr_excel_from_latest_csv_if_needed(we)
        ucr_counts = _build_ucr_target_pl_count_map(we) or {}
    except Exception:
        pass

    # ------------------------------------------------------------------ #
    # 5b. Pre-load weekly_sharepoint_consolidate_summary as BU/Timeline
    #     source 2 - keyed by (target_upper, pl_id_upper)
    # ------------------------------------------------------------------ #
    consolidate_meta_map = {}
    try:
        _cm_conn = get_mysql_connection_db(bu_key=None)
        if _cm_conn:
            _cm_cur = _cm_conn.cursor(dictionary=True)
            try:
                _cm_cur.execute(f"""
                    SELECT target, pl_id, bu, timelines, pdt_test_status
                    FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`
                    ORDER BY week_end DESC
                """)
                for _cm_row in (_cm_cur.fetchall() or []):
                    _cm_key = (
                        str(_cm_row.get('target') or '').strip().upper(),
                        str(_cm_row.get('pl_id')   or '').strip().upper(),
                    )
                    if _cm_key not in consolidate_meta_map:
                        consolidate_meta_map[_cm_key] = _cm_row
            finally:
                _cm_cur.close(); _cm_conn.close()
    except Exception:
        pass

    # ------------------------------------------------------------------ #
    # 6. Group Axiom jobs by (target, pl_id) â€” ONE group per target+PL
    # ------------------------------------------------------------------ #
    def _pl_group(sp):
        return _re.sub(r'\.r\d+$', '', str(sp or ''), flags=_re.IGNORECASE)
    # Use module-level _sp2_meta_build_key for consistent meta-id computation.
    # Meta-id = <Target.Version>-[rN-]<NNNNN[.NN]>-<flavour> with _suffix stripped.
    # Builds with same number + same flavour = same meta; different flavour = different build.
    def _meta_build_key(bn):
        return _sp2_meta_build_key(bn)


    grouped = {}
    group_order = []

    for r in db_rows:
        chips_raw = r.get('chip_ids') or '[]'
        if isinstance(chips_raw, str):
            try:
                chip_ids = json.loads(chips_raw)
            except Exception:
                chip_ids = []
        else:
            chip_ids = list(chips_raw) if chips_raw else []

        pl_id      = str(r.get('software_product') or '').strip()
        pl_grp     = _pl_group(pl_id)
        target     = _swpdt_target_from_product(pl_grp) or pl_grp
        build_id   = str(r.get('build_id') or '').strip()
        build_name = str(r.get('build_name') or build_id).strip()
        hours      = float(r.get('hours_live') or 0)

                
        crashes    = _sp2_crash_count_for_build(crash_map, build_name, build_id, pl_grp)

        # Per-build CRM/Eng override

        bn_key = (build_name.upper(), pl_grp.upper())
        bt = build_type_map.get(bn_key, 'CRM')

        grp_key = (target.upper(), pl_grp.upper())
        if grp_key not in grouped:
            grouped[grp_key] = {
                'target':       target,
                'pl_id':        pl_grp,
                'build_names':  set(),
                'hours':        0.0,
                'chip_ids_set': set(),
                'crashes':      0,
                'bt_votes':     {'CRM': 0, 'Eng': 0},
            }
            group_order.append(grp_key)

        g = grouped[grp_key]
        g['build_names'].add(_meta_build_key(build_name).upper())  # deduplicate by meta key
        g['bt_votes'][bt if bt in ('CRM', 'Eng') else 'CRM'] += 1
        # Devices (chip_ids) counted for ALL builds â€” they are physical hardware.
        # Hours and crashes are CRM-only (Eng builds excluded from those metrics).
        g['chip_ids_set'].update(chip_ids)
        if bt == 'Eng':
            continue
        g['hours']       += hours
        g['crashes']     += crashes
        import logging as _log_grp
    _log_grp.getLogger('weekly_summary_routes').info(
        '[SP2 CONSOLIDATE] db_rows=%d grouped=%d', len(db_rows), len(grouped))

    # ------------------------------------------------------------------ #
    # 7. Resolve BU / Timelines per group
    #    Priority:
    #      1. pdt_stats_dashboard.dashboard_status  (primary)
    #      2. weekly_sharepoint_consolidate_summary (existing saved data)
    #      3. OneView milestone resolver            (last resort)
    # ------------------------------------------------------------------ #
    import re as _re_tl

    # â”€â”€ Pre-load BU + milestone maps ONCE (no per-row DB calls in the loop) â”€â”€â”€â”€â”€â”€
    # Map 1: BU already saved in sp2_build_consolidate for this week (builds tab set it)
    _sp2_existing_bu_map = {}   # TARGET_UPPER -> bu
    try:
        _eb_conn = get_mysql_connection_db(bu_key=None)
        if _eb_conn:
            _eb_cur = _eb_conn.cursor(dictionary=True)
            try:
                _eb_cur.execute(f"""
                    SELECT target, pl_id, bu
                    FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                    WHERE week_start=%s AND week_end=%s
                      AND bu IS NOT NULL AND bu != ''
                """, (ws.isoformat(), we.isoformat()))
                for _r in (_eb_cur.fetchall() or []):
                    _bu_val = str(_r.get('bu') or '').strip()
                    if _bu_val:
                        _sp2_existing_bu_map[str(_r.get('target') or '').strip().upper()] = _bu_val
                        _sp2_existing_bu_map[str(_r.get('pl_id')  or '').strip().upper()] = _bu_val
            finally:
                _eb_cur.close(); _eb_conn.close()
    except Exception:
        pass

    # Map 2: milestone dates from dashboard_status indexed by sp_name variants
    # (dash_map already loaded above has ES/FC/CS â€” reuse it as _sp2_milestone_map)
    _sp2_milestone_map = dash_map   # same dict, already has ES/FC/CS per target key

    # Map 3: BU from _find_dashboard_target_info for all unique targets (one batch query)
    _sp2_bu_map = {}   # TARGET_UPPER -> bu
    try:
        _fd_conn = get_mysql_connection_db(bu_key=None)
        if _fd_conn:
            _fd_cur = _fd_conn.cursor(dictionary=True)
            try:
                _fd_cur.execute("""
                    SELECT bu, target_name, sp_name
                    FROM pdt_stats_dashboard.dashboard_status
                    WHERE is_active=1 AND bu IS NOT NULL AND bu != ''
                """)
                for _r in (_fd_cur.fetchall() or []):
                    _bu_val = str(_r.get('bu') or '').strip()
                    for _raw in [_r.get('target_name'), _r.get('sp_name')]:
                        if not _raw: continue
                        _raw = str(_raw).strip()
                        _sp2_bu_map[_raw.upper()] = _bu_val
                        _sp2_bu_map[_raw.replace('_','.').upper()] = _bu_val
                        # version-strip: Aurora.LA.3.1 -> Aurora.LA
                        import re as _re_bu
                        _s = _raw
                        while True:
                            _s2 = _re_bu.sub(r'[._]\d+$', '', _s)
                            if _s2 == _s or not _s2: break
                            _s = _s2
                            _sp2_bu_map[_s.upper()] = _bu_val
                            _sp2_bu_map[_s.replace('_','.').upper()] = _bu_val
            finally:
                _fd_cur.close(); _fd_conn.close()
    except Exception:
        pass


    # Map 4: existing saved timelines/bu from sp2_build_consolidate (preserve on re-save)
    _sp2_saved_timelines_map = {}   # (TARGET_UPPER, PL_UPPER) -> {es,fc,cs,bu}
    try:
        _st_conn = get_mysql_connection_db(bu_key=None)
        if _st_conn:
            _st_cur = _st_conn.cursor(dictionary=True)
            try:
                _st_cur.execute(f"""
                    SELECT target, pl_id, bu, timelines
                    FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                    WHERE week_start=%s AND week_end=%s
                      AND build_name LIKE '__consolidated__%'
                      AND timelines IS NOT NULL AND timelines != ''
                """, (ws.isoformat(), we.isoformat()))
                import re as _re_tl2
                for _r in (_st_cur.fetchall() or []):
                    _tl_str = str(_r.get('timelines') or '')
                    _es2 = (_re_tl2.search(r'ES[\:\s]+([\d]{4}-[\d]{2}-[\d]{2})', _tl_str) or None)
                    _fc2 = (_re_tl2.search(r'FC[\:\s]+([\d]{4}-[\d]{2}-[\d]{2})', _tl_str) or None)
                    _cs2 = (_re_tl2.search(r'CS[\:\s]+([\d]{4}-[\d]{2}-[\d]{2})', _tl_str) or None)
                    _sp2_saved_timelines_map[
                        (str(_r.get('target') or '').strip().upper(),
                         str(_r.get('pl_id')  or '').strip().upper())
                    ] = {
                        'es': _es2.group(1) if _es2 else None,
                        'fc': _fc2.group(1) if _fc2 else None,
                        'cs': _cs2.group(1) if _cs2 else None,
                        'bu': str(_r.get('bu') or '').strip(),
                    }
            finally:
                _st_cur.close(); _st_conn.close()
    except Exception:
        pass

    for g in grouped.values():
        votes = g.pop('bt_votes')
        g['build_type'] = 'Eng' if votes.get('Eng', 0) > votes.get('CRM', 0) else 'CRM'

        tgt   = g['target']
        pl_id = g['pl_id']

        # â”€â”€ BU + Milestones: all from pre-loaded maps, zero per-row DB calls â”€â”€â”€â”€â”€â”€
        # Source 1: dashboard_status map (pre-loaded once)
        dash = _match_dashboard_with_fallback(tgt, dash_map) or _match_dashboard_with_fallback(pl_id, dash_map) or {}
        bu = str(dash.get('bu') or '').strip()
        es = dash.get('ES')
        fc = dash.get('FC')
        cs = dash.get('CS')

        # Source 2: BU from existing sp2_build_consolidate rows (builds tab already resolved it)
        if not bu:
            bu = _sp2_existing_bu_map.get(tgt.upper()) or _sp2_existing_bu_map.get(pl_id.upper()) or ''

        # Source 3: old consolidate_summary table (has saved timelines strings)
        if not (es or fc or cs):
            _cm = (consolidate_meta_map.get((tgt.upper(), pl_id.upper()))
                   or consolidate_meta_map.get((tgt.upper(), ''))
                   or {})
            if not bu:
                bu = str(_cm.get('bu') or '').strip()
            if _cm.get('timelines'):
                _tl = str(_cm.get('timelines') or '')
                _es_m = _re_tl.search(r'ES[\:\s]+([\d]{4}-[\d]{2}-[\d]{2})', _tl)
                _fc_m = _re_tl.search(r'FC[\:\s]+([\d]{4}-[\d]{2}-[\d]{2})', _tl)
                _cs_m = _re_tl.search(r'CS[\:\s]+([\d]{4}-[\d]{2}-[\d]{2})', _tl)
                if _es_m: es = _es_m.group(1)
                if _fc_m: fc = _fc_m.group(1)
                if _cs_m: cs = _cs_m.group(1)

        # Source 4: milestone map (dashboard_status sp_name indexed, pre-loaded)
        # NOTE: milestones are NOT fetched from OneView here.
        # Use 'Fetch Missing Milestones' button to pull from OneView.
        if not (es or fc or cs):
            _mm = _sp2_milestone_map.get(tgt.upper()) or _sp2_milestone_map.get(pl_id.upper()) or {}
            if _mm.get('ES') or _mm.get('FC') or _mm.get('CS'):
                es = _mm.get('ES'); fc = _mm.get('FC'); cs = _mm.get('CS')
            if not bu:
                bu = str(_mm.get('bu') or '').strip()

        # Source 5: BU map (version-strip aware, pre-loaded)
        if not bu:
            bu = _sp2_bu_map.get(tgt.upper()) or _sp2_bu_map.get(pl_id.upper()) or ''

        # Source 6: preserve existing timelines from DB (don't overwrite with empty)
        # If we have no milestones now, keep whatever was saved before
        _existing = _sp2_saved_timelines_map.get((tgt.upper(), pl_id.upper()), {})
        if not (es or fc or cs):
            es = _existing.get('es'); fc = _existing.get('fc'); cs = _existing.get('cs')
        if not bu:
            bu = str(_existing.get('bu') or '').strip()

        # Build timeline string and PDT status from resolved dates
        if es or fc or cs:
            timelines  = _sp_timeline(es, fc, cs)
            pdt_status = _compute_pdt_test_status(es, cs, fc)
        else:
            timelines  = ''
            pdt_status = ''

        # Unique CRs
        unique_crs = _ucr_count_for_sharepoint_pair(ucr_counts, tgt, pl_id)

        g['bu']               = bu
        g['timelines']        = timelines
        g['pdt_status']       = pdt_status
        g['unique_crs']       = unique_crs
        g['number_of_builds'] = len(g['build_names'])
    # ------------------------------------------------------------------ #
    # 8. Save to DB â€” DELETE existing sentinel rows then INSERT fresh
    # ------------------------------------------------------------------ #
    if not grouped:
        return
    conn4 = get_mysql_connection_db(bu_key=None)
    if not conn4:
        return
    cur4 = conn4.cursor()
    try:
        # Only delete the consolidated sentinel rows, not per-build-name rows
        cur4.execute(f"""
            DELETE FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
            WHERE week_start=%s AND week_end=%s
              AND build_name LIKE '__consolidated__%'
        """, (ws.isoformat(), we.isoformat()))

        # â”€â”€ Per-target hours-proportional device split â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        # Same physical device may run builds under multiple PL-IDs for the same
        # target. We split the target's TOTAL unique devices proportionally by
        # each PL-ID's hours so that SUM(device_count per PL) == target total.
        # Rule: device_count[PL] = max(1, floor(total_devices * pl_hours / target_hours))
        # Remainder distributed to highest-hours PLs. Min 1 per active PL.
        _tgt_total_chips = {}   # target_upper -> total unique chip count
        _tgt_total_hours = {}   # target_upper -> sum of hours across all PLs
        _tgt_pl_keys     = {}   # target_upper -> [grp_key, ...] in group_order
        for _gk in group_order:
            _g  = grouped[_gk]
            _tu = _g['target'].upper()
            if _tu not in _tgt_total_chips:
                _tgt_total_chips[_tu] = set()
                _tgt_total_hours[_tu] = 0.0
                _tgt_pl_keys[_tu]     = []
            _tgt_total_chips[_tu].update(_g['chip_ids_set'])
            _tgt_total_hours[_tu] += _g['hours']
            _tgt_pl_keys[_tu].append(_gk)

        _group_dev_count = {}   # grp_key -> integer device count
        for _tu, _pl_keys in _tgt_pl_keys.items():
            _total_dev = len(_tgt_total_chips[_tu])
            _total_hrs = _tgt_total_hours[_tu]
            if len(_pl_keys) == 1 or _total_dev == 0:
                # Only one PL or no devices: assign all to that PL
                for _gk in _pl_keys:
                    _group_dev_count[_gk] = _total_dev
                continue
            # Proportional split
            if _total_hrs > 0:
                _raw = {_gk: _total_dev * grouped[_gk]['hours'] / _total_hrs
                        for _gk in _pl_keys}
            else:
                # No hours at all: split evenly
                _even = _total_dev / len(_pl_keys)
                _raw  = {_gk: _even for _gk in _pl_keys}
            # Floor each, guarantee minimum 1 per PL that has chips or hours
            _floored = {}
            for _gk in _pl_keys:
                _g = grouped[_gk]
                _active = _g['hours'] > 0 or len(_g['chip_ids_set']) > 0
                _floored[_gk] = max(1 if _active else 0, int(_raw[_gk]))
            # Distribute remainder to highest-hours PLs
            _assigned = sum(_floored.values())
            _remainder = _total_dev - _assigned
            if _remainder > 0:
                _sorted_by_hrs = sorted(_pl_keys,
                    key=lambda k: grouped[k]['hours'], reverse=True)
                for _gk in _sorted_by_hrs:
                    if _remainder <= 0:
                        break
                    _floored[_gk] += 1
                    _remainder   -= 1
            elif _remainder < 0:
                # Over-assigned (due to min-1 floor): reduce from lowest-hours PLs
                _sorted_asc = sorted(_pl_keys,
                    key=lambda k: grouped[k]['hours'])
                for _gk in _sorted_asc:
                    if _remainder >= 0:
                        break
                    if _floored[_gk] > 1:
                        _floored[_gk] -= 1
                        _remainder    += 1
            _group_dev_count.update(_floored)

        for g in grouped.values():
            _gk2     = (g['target'].upper(), g['pl_id'].upper())
            chip_ids = sorted(g['chip_ids_set'])  # kept for chip_ids JSON column
            _dev_cnt = _group_dev_count.get(_gk2, len(chip_ids))
            cur4.execute(f"""
                INSERT INTO `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                    (week_start, week_end, target, pl_id, build_name, build_type,
                     total_hours, total_crashes, device_count, chip_ids,
                     bu, timelines, pdt_test_status, unique_crs,
                     number_of_builds, updated_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE
                    target=VALUES(target),
                    total_hours=VALUES(total_hours),
                    total_crashes=VALUES(total_crashes),
                    device_count=VALUES(device_count),
                    chip_ids=VALUES(chip_ids),
                    bu=VALUES(bu),
                    timelines=VALUES(timelines),
                    pdt_test_status=VALUES(pdt_test_status),
                    unique_crs=VALUES(unique_crs),
                    number_of_builds=VALUES(number_of_builds),
                    updated_by=VALUES(updated_by),
                    updated_at=CURRENT_TIMESTAMP
            """, (
                ws.isoformat(), we.isoformat(),
                g['target'], g['pl_id'],
                f"__consolidated__{g['target']}__{g['pl_id']}",
                g['build_type'],
                round(g['hours'], 3),
                g['crashes'],
                _dev_cnt,
                json.dumps(chip_ids),
                g['bu'],
                g['timelines'],
                g['pdt_status'],
                g['unique_crs'],
                g['number_of_builds'],
                username,
            ))
        conn4.commit()
        import logging as _log_sv
        _log_sv.getLogger('weekly_summary_routes').info(
            '[SP2 CONSOLIDATE] Saved %d sentinel rows for %s to %s',
            len(grouped), ws.isoformat(), we.isoformat())
    except Exception as _exc:
        import logging as _log2
        _log2.getLogger('weekly_summary_routes').error('[SP2 CONSOLIDATE SAVE] %s', _exc, exc_info=True)
        try: conn4.rollback()
        except Exception: pass
    finally:
        cur4.close(); conn4.close()

def _fetch_sp2_consolidate(ws, we, crm_only: bool = True) -> list:
    """Fetch sp2 consolidate rows â€” one sentinel row per (target, pl_id).

    New schema: _build_and_save_sp2_consolidate writes exactly ONE row per
    (target, pl_id) with build_name='__consolidated__<target>__<pl_id>'.
    All enriched columns (bu, timelines, pdt_test_status, unique_crs,
    number_of_builds) are stored directly â€” no GROUP BY needed.

    Backwards compat: if no sentinel rows exist (old schema), falls back to
    reading all rows and deduplicating by (target, pl_id) in Python.
    """
    _ensure_sp2_build_consolidate_table()
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = conn.cursor(dictionary=True)
    try:
        # --- Try new schema first: sentinel rows only ---
        # NOTE: bt_filter removed â€” sentinel rows already have CRM-only metrics
        # baked in by _build_and_save_sp2_consolidate. All groups are shown.
        cur.execute(f"""
            SELECT target, pl_id, build_type,
                   number_of_builds, total_hours, total_crashes,
                   device_count, chip_ids,
                   bu, timelines, pdt_test_status, unique_crs
            FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
            WHERE week_start=%s AND week_end=%s
              AND build_name LIKE '__consolidated__%'
            ORDER BY target, pl_id
        """, (ws.isoformat(), we.isoformat()))
        sp2_rows = cur.fetchall() or []

        # --- Fallback: old schema (per-build-name rows) â€” deduplicate in Python ---
        if not sp2_rows:
            cur.execute(f"""
                SELECT target, pl_id, build_type,
                       total_hours, total_crashes, device_count,
                       chip_ids, bu, timelines, pdt_test_status, unique_crs,
                       number_of_builds
                FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                WHERE week_start=%s AND week_end=%s
                ORDER BY target, pl_id
            """, (ws.isoformat(), we.isoformat()))
            raw_rows = cur.fetchall() or []
            # Deduplicate: keep first row per (target, pl_id), accumulate hours/crashes/chips
            seen = {}
            for r in raw_rows:
                key = (str(r.get('target') or '').strip().upper(),
                       str(r.get('pl_id')   or '').strip().upper())
                if key not in seen:
                    seen[key] = dict(r)
                    seen[key]['_build_count'] = 1
                else:
                    seen[key]['total_hours']   = float(seen[key].get('total_hours') or 0) + float(r.get('total_hours') or 0)
                    seen[key]['total_crashes'] = int(seen[key].get('total_crashes') or 0) + int(r.get('total_crashes') or 0)
                    seen[key]['_build_count'] += 1
            sp2_rows = list(seen.values())
            for r in sp2_rows:
                if not r.get('number_of_builds'):
                    r['number_of_builds'] = r.get('_build_count', 1)

        # --- Fallback meta from old consolidate table (BU/Timelines if still empty) ---
        meta_map = {}
        try:
            cur.execute(f"""
                SELECT target, pl_id, bu, timelines, pdt_test_status, unique_crs
                FROM `{_QIPL_DB}`.`{_CONSOLIDATE_SUMMARY_TABLE}`
                WHERE week_end=%s
            """, (we.isoformat(),))
            for m in (cur.fetchall() or []):
                key = (str(m.get('target') or '').strip().upper(),
                       str(m.get('pl_id')   or '').strip().upper())
                meta_map[key] = m
        except Exception:
            pass

        rows = []
        for r in sp2_rows:
            total_hrs = float(r.get('total_hours') or 0)
            total_cr  = int(r.get('total_crashes') or 0)
            mtbf      = round(total_hrs / total_cr, 2) if total_cr else 0

            tgt   = str(r.get('target') or '').strip()
            pl_id = str(r.get('pl_id')  or '').strip()
            nb    = int(r.get('number_of_builds') or 0)

            bu         = str(r.get('bu') or '').strip()
            timelines  = str(r.get('timelines') or '').strip()
            pdt_status = str(r.get('pdt_test_status') or '').strip()
            unique_crs = r.get('unique_crs')

            # Fallback to old consolidate table if still empty
            if not bu or not timelines:
                meta = meta_map.get((tgt.upper(), pl_id.upper())) or {}
                if not bu:         bu         = str(meta.get('bu') or '').strip()
                if not timelines:  timelines  = str(meta.get('timelines') or '').strip()
                if not pdt_status: pdt_status = str(meta.get('pdt_test_status') or '').strip()
                if unique_crs is None: unique_crs = meta.get('unique_crs')

            # Final BU fallback: dashboard_status lookup (handles versioned PLs)
            if not bu:
                try:
                    _di = _find_dashboard_target_info(tgt, pl_id)
                    bu = str(_di.get('bu') or '').strip()
                except Exception:
                    pass

            rows.append({
                'target':           tgt,
                'pl_id':            pl_id,
                'build_type':       str(r.get('build_type') or 'CRM'),
                'number_of_builds': nb,
                'total_hours':      round(total_hrs, 2),
                'total_crashes':    total_cr,
                'device_count':     int(r.get('device_count') or 0),
                'mtbf':             mtbf,
                'bu':               bu,
                'timelines':        timelines,
                'pdt_test_status':  pdt_status,
                'unique_crs':       unique_crs,
            })
        return rows
    except Exception as _exc:
        import logging as _log3
        _log3.getLogger('weekly_summary_routes').warning('[SP2 FETCH CONSOLIDATE] %s', _exc)
        return []
    finally:
        cur.close(); conn.close()

@weekly_summary_bp.route('/weekly-report/smart-build-report')
@login_required
def sharepoint2_page():
    """Smart Build Report page - fully auto-populated from Axiom DB."""
    sel_start, sel_end = _selected_week_from_request()
    week_ranges = _week_ranges_for_templates()
    try:
        from dashboard_routes import _build_bu_shell_context
        shell_ctx = _build_bu_shell_context('WEEKLY_QIPL_REPORTS')
    except Exception:
        shell_ctx = {'active_bu_key': 'WEEKLY_QIPL_REPORTS', 'bu_list': [],
                     'BU_ICONS': {}, 'shell_title': 'Smart Build Report'}
    shell_ctx['shell_title'] = 'Smart Build Report'
    return render_template(
        'sharepoint2.html',
        sel_start=sel_start,
        sel_end=sel_end,
        week_ranges=week_ranges,
        **shell_ctx,
    )


@weekly_summary_bp.route('/api/sp2/builds')
@login_required
def api_sp2_builds():
    """Return /PDT/QIPL Axiom builds for the selected week.

    Grouping: same build_name + pl_id are merged into one row.
    Hours summed, chip_ids unioned (unique devices), crashes summed.
    Each row has build_type (CRM/Eng) that the user can toggle per row.
    Hours are week-bounded: running builds capped at week_end Sunday 23:59:59.
    Crashes from weekly_qipl_data (same CSV as CR Pie / CR Age).
    """
    import re as _re
    ws_arg = request.args.get('week_start', '').strip()
    we_arg = request.args.get('week_end', '').strip()
    ws = _safe_date(ws_arg)
    we = _safe_date(we_arg)
    if not ws or not we:
        ws, we = _selected_week_from_request()

    # Static mode: once the weekly CSV is present, seed/read frozen build rows
    # from sp2_build_type_overrides. User edits update this table, so page
    # refreshes do not recalculate/overwrite hours or crashes from CSV/Axiom.
    _seed_sp2_build_type_overrides_from_axiom(ws, we, _current_user_identifier())
    static_rows = _load_sp2_static_build_rows(ws, we)
    if static_rows:
        dash_map_static = _fetch_dashboard_status_map()
        # Build target->bu lookup from sp2_build_consolidate (user-saved target-level BU)
        _cons_bu_map = {}
        try:
            _cconn = get_mysql_connection_db(bu_key=None)
            if _cconn:
                _ccur = _cconn.cursor(dictionary=True)
                _ccur.execute(f"""
                    SELECT DISTINCT target, bu
                    FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                    WHERE week_start=%s AND week_end=%s
                      AND bu IS NOT NULL AND bu != ''
                """, (ws.isoformat(), we.isoformat()))
                for _cr in (_ccur.fetchall() or []):
                    if _cr.get('target') and _cr.get('bu'):
                        _cons_bu_map[str(_cr['target']).strip()] = str(_cr['bu']).strip()
                _ccur.close(); _cconn.close()
        except Exception:
            pass
        out = []
        all_chips = set()
        for r in static_rows:
            chips_raw = r.get('chip_ids') or '[]'
            try:
                chip_ids = json.loads(chips_raw) if isinstance(chips_raw, str) else list(chips_raw or [])
            except Exception:
                chip_ids = []
            chip_ids = sorted(str(c).strip() for c in chip_ids if str(c).strip())
            all_chips.update(chip_ids)
            target = str(r.get('target') or '').strip() or (_swpdt_target_from_product(r.get('pl_id')) or '')
            pl_id = str(r.get('pl_id') or '').strip()
            dash = _match_dashboard_with_fallback(target, dash_map_static) or _match_dashboard_with_fallback(pl_id, dash_map_static) or {}
            # BU priority: 1) consolidate target-level (user saved)  2) row-level  3) dashboard_status
            row_bu = _cons_bu_map.get(target) or str(r.get('bu') or dash.get('bu') or '').strip()
            job_ids_raw = r.get('job_ids') or '[]'
            try:
                job_ids = json.loads(job_ids_raw) if isinstance(job_ids_raw, str) else list(job_ids_raw or [])
            except Exception:
                job_ids = []
            state = str(r.get('state') or '').lower()
            out.append({
                'job_ids':      job_ids,
                'job_id':       job_ids[0] if job_ids else '',
                'target':       target,
                'pl_id':        pl_id,
                'pl_id_exact':  pl_id,
                'build_id':     str(r.get('build_id') or ''),
                'build_name':   str(r.get('build_name') or r.get('build_id') or ''),
                'submitted':    str(r.get('submitted_at') or '')[:10],
                'completed_at': str(r.get('completed_at') or '')[:10],
                'status':       'running' if state in ('running', 'jobsetup') else 'completed',
                'hours':        round(float(r.get('hours') or 0), 3),
                'device_count': len(chip_ids) or int(r.get('device_count') or 0),
                'chip_ids':     chip_ids,
                'crashes':      int(r.get('total_crashes') or 0),
                'build_type':   str(r.get('build_type') or 'CRM'),
                'bu':           row_bu,
                'meta_id':      _sp2_meta_build_key(str(r.get('build_name') or r.get('build_id') or '')),
                'run_count':    len(job_ids) or 1,
            })
        out.sort(key=lambda x: (x['target'].lower(), x['pl_id'].lower(), x['submitted']))
        bu_opts = {str(o.get('key') or '').strip() for o in _sp_bu_options() if str(o.get('key') or '').strip()}
        bu_opts.update({b['bu'] for b in out if b['bu']})
        return jsonify(success=True, builds=out, total_devices=len(all_chips),
                       bu_list=sorted(bu_opts),
                       week_start=ws.isoformat(), week_end=we.isoformat(), static=True)

    # 1. Query axiom_job_summary
    db_rows = []
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if conn:
            cur = conn.cursor(dictionary=True)
            try:
                _week_cap = we.isoformat() + " 23:59:59"
                _week_floor = ws.isoformat() + " 00:00:00"
                live_h = (
                    "CASE"
                    " WHEN state IN ('Running','JobSetup') AND started_at IS NOT NULL"
                    " THEN ROUND(device_count *"
                    " TIMESTAMPDIFF(SECOND, started_at,"
                    " LEAST(NOW(), TIMESTAMP('" + _week_cap + "'))) / 3600.0 * 0.80, 3)"
                    " WHEN state IN ('Completed','Aborted') AND started_at IS NOT NULL AND ended_at IS NOT NULL"
                    " THEN ROUND(device_count *"
                    " TIMESTAMPDIFF(SECOND,"
                    " GREATEST(started_at, TIMESTAMP('" + _week_floor + "')),"
                    " LEAST(ended_at,      TIMESTAMP('" + _week_cap   + "'))) / 3600.0 * 0.80, 3)"
                    " ELSE 0 END"
                )
                cur.execute(f"""
                    SELECT job_id, build_id, build_name, software_product,
                           taxonomy_path, team, state, device_count, chip_ids,
                           submitted_at, started_at, ended_at,
                           ({live_h}) AS hours_live,
                           product_flavor, submitter, site
                    FROM `pdt_stats_dashboard`.`axiom_job_summary`
                    WHERE taxonomy_path = '/PDT/QIPL'
                      AND started_at <= %s AND (ended_at >= %s OR state IN ('Running','JobSetup'))
                    ORDER BY submitted_at DESC
                """, (we.isoformat(), ws.isoformat()))
                db_rows = cur.fetchall() or []
            finally:
                cur.close(); conn.close()
    except Exception as _exc:
        import logging as _log
        _log.getLogger('weekly_summary_routes').warning('[SP2 BUILDS] DB read failed: %s', _exc)

    if not db_rows:
        return jsonify(success=True, builds=[], total_devices=0,
                       week_start=ws.isoformat(), week_end=we.isoformat())

        # 2. Week + PL bounded crash map from weekly_qipl_data.
    crash_map = _sp2_weekly_crash_map(ws, we)


    # 3. Load saved build_type overrides from sp2_build_type_overrides
    #    (sp2_build_consolidate only holds sentinel rows, not per-build overrides)
    build_type_map = {}
    try:
        _ensure_sp2_build_type_overrides_table()
        conn3 = get_mysql_connection_db(bu_key=None)
        if conn3:
            cur3 = conn3.cursor(dictionary=True)
            try:
                cur3.execute(f"""
                    SELECT build_name, pl_id, build_type
                    FROM `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                    WHERE week_start=%s AND week_end=%s
                """, (ws.isoformat(), we.isoformat()))
                for row in cur3.fetchall() or []:
                    k = (str(row.get('build_name') or '').strip().upper(),
                         str(row.get('pl_id') or '').strip().upper())
                    build_type_map[k] = str(row.get('build_type') or 'CRM')
            except Exception:
                pass
            finally:
                cur3.close(); conn3.close()
    except Exception:
        pass

        # 4. Normalise + GROUP by (meta_build_key, pl_group)
    # 4. Normalise each DB row into a flat build entry.
    #    Builds tab shows ALL individual builds with full names (no grouping).
    #    Consolidate tab uses _build_and_save_sp2_consolidate which groups by meta_key.
    import re as _re2
    def _pl_group(sp):
        return _re.sub(r'\.r\d+$', '', str(sp or ''), flags=_re.IGNORECASE)

    # Load dashboard_status for BU lookup per target
    _dash_map_builds = _fetch_dashboard_status_map()

    # Load saved BU overrides keyed by target_upper -> bu
    _bu_override_map = {}
    try:
        _ensure_sp2_build_consolidate_table()
        _bo_conn = get_mysql_connection_db(bu_key=None)
        if _bo_conn:
            _bo_cur = _bo_conn.cursor(dictionary=True)
            try:
                _bo_cur.execute(f"""
                    SELECT target, bu
                    FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                    WHERE week_start=%s AND week_end=%s
                      AND bu IS NOT NULL AND bu != ''
                    ORDER BY updated_at DESC
                """, (ws.isoformat(), we.isoformat()))
                for _bo_row in (_bo_cur.fetchall() or []):
                    _tgt_key = str(_bo_row.get('target') or '').strip().upper()
                    if _tgt_key and _tgt_key not in _bu_override_map:
                        _bu_override_map[_tgt_key] = str(_bo_row.get('bu') or '').strip()
            finally:
                _bo_cur.close(); _bo_conn.close()
    except Exception:
        pass

    # 5. Group by (build_name, pl_id) so duplicate Axiom job runs for the
    #    same build are merged into one row.
    #    Hours are SUMMED, chip_ids are UNIONED (unique devices),
    #    crashes come from crash_map (keyed by build_name, not job_id),
    #    status = running if ANY job for that build is still running.
    grouped = {}   # key: (build_name_upper, pl_grp_upper) -> accumulator dict
    all_chips = set()

    for r in db_rows:
        chips_raw = r.get('chip_ids') or '[]'
        if isinstance(chips_raw, str):
            try:
                chip_ids = json.loads(chips_raw)
            except Exception:
                chip_ids = []
        else:
            chip_ids = list(chips_raw) if chips_raw else []

        # Skip ghost/auto jobs: no devices assigned AND negligible hours.
        # These are Axiom bookkeeping rows (device_count=0, chip_ids=[],
        # hours<=0.1) that inflate the build count without real execution.
        _raw_dev = int(r.get('device_count') or 0)
        _raw_hrs = float(r.get('hours_live') or 0)
        if _raw_dev <= 0 and not chip_ids and _raw_hrs <= 0.1:
            continue

        pl_id      = str(r.get('software_product') or '').strip()
        pl_grp     = _pl_group(pl_id)
        target     = _swpdt_target_from_product(pl_grp) or pl_grp
        build_id   = str(r.get('build_id') or '').strip()
        build_name = str(r.get('build_name') or build_id).strip()
        hours      = float(r.get('hours_live') or 0)

        # AUTO submitter: reduce hours by 20% (Axiom auto-scheduled jobs
        # run at reduced farm priority; their wall-clock hours overcount
        # actual PDT execution time).
        _submitter = str(r.get('submitter') or '').strip().upper()
        if _submitter == 'AUTO':
            hours = round(hours * 0.80, 3)

        state      = str(r.get('state') or '').lower()
        is_running = state in ('running', 'jobsetup')
        job_id     = str(r.get('job_id') or '')
        submitted  = str(r.get('submitted_at') or '')[:10]
        completed  = str(r.get('ended_at') or '')[:10]

        grp_key = (build_name.upper(), pl_grp.upper())
        if grp_key not in grouped:
            # BU: saved override first, then dashboard_status
            _dash = _match_dashboard(target, _dash_map_builds) or {}
            _bu   = (_bu_override_map.get(target.upper())
                     or str(_dash.get('bu') or '').strip())
            bt    = build_type_map.get((build_name.upper(), pl_grp.upper()), 'CRM')
            grouped[grp_key] = {
                'job_ids':      [],
                'job_id':       job_id,   # first job_id (representative)
                'target':       target,
                'pl_id':        pl_grp,
                'pl_id_exact':  pl_id,
                'build_id':     build_id,
                'build_name':   build_name,
                'submitted':    submitted,
                'completed_at': completed,
                'is_running':   False,
                'hours':        0.0,
                'chip_ids_set': set(),
                'crashes':      0,        # filled after loop
                'build_type':   bt,
                'bu':           _bu,
                'meta_id':      _sp2_meta_build_key(build_name),
            }

        acc = grouped[grp_key]
        acc['job_ids'].append(job_id)
        acc['hours']        += hours
        acc['chip_ids_set'].update(str(c).strip() for c in chip_ids if str(c).strip())
        all_chips.update(chip_ids)
        if is_running:
            acc['is_running'] = True
        # keep earliest submitted, latest completed
        if submitted and (not acc['submitted'] or submitted < acc['submitted']):
            acc['submitted'] = submitted
        if completed and completed > acc.get('completed_at', ''):
            acc['completed_at'] = completed

    # Resolve crashes per unique build_name and finalise output list
    out = []
    for acc in grouped.values():
        build_name = acc['build_name']
        build_id   = acc['build_id']
        crashes    = _sp2_crash_count_for_build(crash_map, build_name, build_id, acc.get('pl_id'))


        chip_ids_sorted = sorted(acc.pop('chip_ids_set'))
        out.append({
            'job_ids':      acc['job_ids'],
            'job_id':       acc['job_id'],
            'target':       acc['target'],
            'pl_id':        acc['pl_id'],
            'pl_id_exact':  acc['pl_id_exact'],
            'build_id':     acc['build_id'],
            'build_name':   build_name,
            'submitted':    acc['submitted'],
            'completed_at': acc['completed_at'],
            'status':       'running' if acc['is_running'] else 'completed',
            'hours':        round(acc['hours'], 3),
            'device_count': len(chip_ids_sorted),
            'chip_ids':     chip_ids_sorted,
            'crashes':      crashes,
            'build_type':   acc['build_type'],
            'bu':           acc['bu'],
            'meta_id':      acc['meta_id'],
            'run_count':    len(acc['job_ids']),   # how many Axiom jobs ran for this build
        })

    # Collect BU list for dropdown
    bu_opts = {str(o.get('key') or '').strip() for o in _sp_bu_options() if str(o.get('key') or '').strip()}
    bu_opts.update({b['bu'] for b in out if b['bu']})
    bu_list = sorted(bu_opts)

    out.sort(key=lambda x: (x['target'].lower(), x['pl_id'].lower(), x['submitted']))
    return jsonify(success=True, builds=out, total_devices=len(all_chips),
                   bu_list=bu_list,
                   week_start=ws.isoformat(), week_end=we.isoformat())


@weekly_summary_bp.route('/api/sp2/debug_consolidate')
@login_required
def api_sp2_debug_consolidate():
    ws_arg = request.args.get('week_start','').strip()
    we_arg = request.args.get('week_end','').strip()
    ws = _safe_date(ws_arg)
    we = _safe_date(we_arg)
    if not we: _, we = _selected_week_from_request()
    if not ws: ws = we - timedelta(days=6)
    result = {'week_start': ws.isoformat(), 'week_end': we.isoformat()}
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if not conn:
            return jsonify({'error':'no db connection'})
        cur = conn.cursor(dictionary=True)
        try:
            # A: total QIPL rows (no date filter)
            cur.execute("SELECT COUNT(*) as c FROM `pdt_stats_dashboard`.`axiom_job_summary` WHERE taxonomy_path='/PDT/QIPL'")
            result['A_total_qipl'] = int((cur.fetchone() or {}).get('c') or 0)
            # B: date range in table
            cur.execute("SELECT MIN(DATE(submitted_at)) mn, MAX(DATE(submitted_at)) mx FROM `pdt_stats_dashboard`.`axiom_job_summary` WHERE taxonomy_path='/PDT/QIPL'")
            r = cur.fetchone() or {}
            result['B_date_range'] = {'min': str(r.get('mn') or ''), 'max': str(r.get('mx') or '')}
            # C: count for requested week
            cur.execute("SELECT COUNT(*) as c FROM `pdt_stats_dashboard`.`axiom_job_summary` WHERE taxonomy_path='/PDT/QIPL' AND started_at <= %s AND (ended_at >= %s OR state IN ('Running','JobSetup'))", (we.isoformat(), ws.isoformat()))
            result['C_count_for_week'] = int((cur.fetchone() or {}).get('c') or 0)
            # D: latest 5 rows
            cur.execute("SELECT software_product, build_name, DATE(submitted_at) sub FROM `pdt_stats_dashboard`.`axiom_job_summary` WHERE taxonomy_path='/PDT/QIPL' ORDER BY submitted_at DESC LIMIT 5")
            result['D_latest_5'] = [{k:str(v) for k,v in (row or {}).items()} for row in (cur.fetchall() or [])]
            # E: saved sentinel rows
            cur.execute(f"SELECT target,pl_id,device_count,total_hours,number_of_builds,bu FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}` WHERE week_start=%s AND week_end=%s AND build_name LIKE '__consolidated__%%' ORDER BY target,pl_id", (ws.isoformat(), we.isoformat()))
            result['E_saved_rows'] = [{k:str(v) for k,v in (row or {}).items()} for row in (cur.fetchall() or [])]
            result['E_saved_count'] = len(result['E_saved_rows'])
        finally:
            cur.close(); conn.close()
    except Exception as e:
        result['error'] = str(e)
    return jsonify(result)


@weekly_summary_bp.route('/api/sp2/consolidate')
@login_required
def api_sp2_consolidate():
    """Smart Build consolidate: CRM rows shown in UI, Eng rows saved but hidden."""
    ws_arg = request.args.get('week_start', '').strip()
    we_arg = request.args.get('week_end', '').strip()
    ws = _safe_date(ws_arg)
    we = _safe_date(we_arg)
    if not we:
        _, we = _selected_week_from_request()
    if not ws:
        ws = we - timedelta(days=6)
    # Always rebuild from Axiom on every load so consolidate matches builds tab
    _build_and_save_sp2_consolidate(ws, we, _current_user_identifier())
    rows = _fetch_sp2_consolidate(ws, we, crm_only=True)
    return jsonify(success=True, rows=rows, week_end=we.isoformat())


@weekly_summary_bp.route('/api/sp2/stability_health')
@login_required
def api_sp2_stability_health():
    """PDT Stability Health trend.

    Uses Smart Build consolidate for weeks that exist there. For older weeks,
    falls back to weekly_sharepoint_consolidate_summary totals so the chart can
    still show the previous 2-3 weeks.
    """
    we = _safe_date(request.args.get('week_end')) or date.today()
    try:
        count = int(request.args.get('count') or 20)
    except Exception:
        count = 20
    count = max(1, min(count, 60))
    min_week_end = date(2026, 6, 14)
    weeks = []
    for i in range(count - 1, -1, -1):
        week_end = we - timedelta(weeks=i)
        if week_end < min_week_end:
            continue
        week_start = week_end - timedelta(days=6)
        rows = []
        old_rows = None
        weekly_cr_mapped_distinct = 0
        source = 'sp2'
        conn = get_mysql_connection_db(bu_key=None)
        if conn:
            cur = conn.cursor(dictionary=True)
            try:
                cur.execute(
                    f"""SELECT total_hours, total_crashes, device_count
                      FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                      WHERE week_start=%s AND week_end=%s
                        AND build_name LIKE '__consolidated__%'""",
                    (week_start.isoformat(), week_end.isoformat()))
                rows = cur.fetchall() or []

                # Number of CRs for the JIRA/CR chart comes directly from
                # pdt_stats_dashboard.weekly_qipl_data for the same week:
                # distinct CR Current Ticket where jira_category = 'CR Mapped'.
                cur.execute(
                    f"""SELECT COUNT(DISTINCT NULLIF(TRIM(cr_current_ticket), '')) AS cr_count
                      FROM `{_QIPL_DB}`.`{_QIPL_TABLE}`
                      WHERE week_start=%s AND week_end=%s
                        AND LOWER(TRIM(COALESCE(jira_category,'')))='cr mapped'""",
                    (week_start.isoformat(), week_end.isoformat()))
                cr_row = cur.fetchone() or {}
                weekly_cr_mapped_distinct = int(cr_row.get('cr_count') or 0)
            finally:
                cur.close(); conn.close()

        hrs     = sum(float(r.get('total_hours')   or 0) for r in rows)
        crashes = sum(float(r.get('total_crashes') or 0) for r in rows)
        dev     = sum(float(r.get('device_count')  or 0) for r in rows)

        # Distinct CR count is sourced from weekly_qipl_data, not from the
        # SharePoint consolidate unique_crs column.
        old_rows = _fetch_consolidate_summary(week_end)
        unique_crs = weekly_cr_mapped_distinct

        if not (hrs > 0 or crashes > 0 or dev > 0):
            source = 'weekly_sharepoint_consolidate_summary'
            hrs     = sum(float(r.get('total_hours')       or 0) for r in old_rows)
            crashes = sum(float(r.get('total_crashes')     or 0) for r in old_rows)
            dev     = sum(float(r.get('number_of_devices') or 0) for r in old_rows)

        if not (hrs > 0 or crashes > 0 or dev > 0 or unique_crs > 0):
            continue
        weeks.append({
            'week_end':              week_end.isoformat(),
            'week_start':            week_start.isoformat(),
            'label':                 week_end.strftime('%d-%b'),
            'label_year':            week_end.strftime('%d-%b-%Y'),
            'source':                source,
            'total_hours':           round(hrs, 1),
            'total_crashes':         int(crashes),
            'total_jiras':           int(crashes),
            'total_unique_crs':      int(unique_crs),
            'total_devices':         int(dev),
            'device_usage_per_week': round(hrs / dev,     2) if dev     else 0,
            'time_per_crash':        round(hrs / crashes, 2) if crashes else 0,
            'crash_per_mtp_week':    round(crashes / dev, 2) if dev     else 0,
        })
    return jsonify(success=True, weeks=weeks[-count:])


@weekly_summary_bp.route('/api/sp2/fetch_all_missing_milestones', methods=['POST'])
@login_required
def api_sp2_fetch_all_missing_milestones():
    """Fetch ES/FC/CS from OneView for ALL consolidate rows that have no timelines.
    Processes sequentially with per-PL retry to handle OneView timeouts.
    POST body: {week_start, week_end}
    Returns: {results: [{pl_id, target, status, timelines, error}], total, updated, failed, skipped}
    """
    data = request.get_json(force=True, silent=True) or {}
    ws   = _safe_date(data.get('week_start'))
    we   = _safe_date(data.get('week_end'))
    if not ws or not we:
        return jsonify(success=False, error='week_start and week_end are required'), 400
    if not fetch_milestones_for_sp:
        return jsonify(success=False, error='fetch_milestones_for_sp not available'), 503

    # Load all sentinel rows that are missing timelines
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return jsonify(success=False, error='DB connection failed'), 500
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT target, pl_id, timelines
            FROM `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
            WHERE week_start=%s AND week_end=%s
              AND build_name LIKE '__consolidated__%'
            ORDER BY target, pl_id
        """, (ws.isoformat(), we.isoformat()))
        all_rows = cur.fetchall() or []
    finally:
        cur.close(); conn.close()

    # Separate missing vs already-filled
    missing = [r for r in all_rows if not str(r.get('timelines') or '').strip()]
    skipped = [r for r in all_rows if str(r.get('timelines') or '').strip()]

    results = []
    updated = 0
    failed  = 0

    for row in missing:
        target = str(row.get('target') or '').strip()
        pl_id  = str(row.get('pl_id')  or '').strip()
        if not pl_id:
            results.append({'target': target, 'pl_id': pl_id, 'status': 'skipped', 'error': 'empty pl_id'})
            failed += 1
            continue

        # DB-first: check dashboard_status before hitting OneView
        _ms = _resolve_pl_milestones(target, pl_id)
        es = _ms['ES']; fc = _ms['FC']; cs = _ms['CS']

        if not (es or fc or cs):
            results.append({'target': target, 'pl_id': pl_id, 'status': 'failed',
                            'error': 'No milestone dates found in DB or OneView'})
            failed += 1
            continue

        timelines  = _sp_timeline(es, fc, cs)
        pdt_status = _compute_pdt_test_status(es, cs, fc)
        sentinel   = f'__consolidated__{target}__{pl_id}'

        try:
            conn2 = get_mysql_connection_db(bu_key=None)
            if conn2:
                cur2 = conn2.cursor()
                try:
                    cur2.execute(f"""
                        UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                        SET timelines=%s, pdt_test_status=%s, updated_by=%s
                        WHERE week_start=%s AND week_end=%s AND build_name=%s
                    """, (timelines, pdt_status, _current_user_identifier(),
                            ws.isoformat(), we.isoformat(), sentinel))
                    conn2.commit()
                finally:
                    cur2.close(); conn2.close()
        except Exception as _db_e:
            import logging as _log
            _log.getLogger('weekly_summary_routes').warning('[SP2 BULK TL] DB update failed: %s', _db_e)

        results.append({'target': target, 'pl_id': pl_id, 'status': 'updated',
                        'timelines': timelines, 'pdt_status': pdt_status,
                        'es': es, 'fc': fc, 'cs': cs})
        updated += 1

    return jsonify(
        success=True,
        results=results,
        total=len(all_rows),
        updated=updated,
        failed=failed,
        skipped=len(skipped)
    )


def _resolve_pl_milestones(target: str, pl_id: str) -> dict:
    """Resolve ES/FC/CS for a PL-ID.
    Priority:
      1. dashboard_status DB with version-strip fallback:
         Kobuk.LE.1.1 -> Kobuk.LE.1 -> Kobuk.LE -> finds DB row instantly
      2. OneView API (fetch_milestones_for_sp) -- tries pl_id then target
    Returns dict with keys: ES, FC, CS, source ('db' or 'oneview' or 'none')
    """
    # Source 1: dashboard_status (version-strip fallback)
    try:
        _dm = _fetch_dashboard_status_map()
        for _cand in [pl_id, target]:
            if not _cand:
                continue
            _info = _match_dashboard_with_fallback(_cand, _dm)
            if _info and (_info.get('ES') or _info.get('FC') or _info.get('CS')):
                return {
                    'ES': _fmt_iso_date(_info.get('ES')),
                    'FC': _fmt_iso_date(_info.get('FC')),
                    'CS': _fmt_iso_date(_info.get('CS')),
                    'source': 'db',
                }
    except Exception:
        pass

    # Source 2: OneView API
    if fetch_milestones_for_sp:
        for _sp in [pl_id, target]:
            if not _sp:
                continue
            try:
                _kd, _src = fetch_milestones_for_sp(_sp)
                if _kd.get('ES') or _kd.get('FC') or _kd.get('CS'):
                    return {
                        'ES': _fmt_iso_date(_kd.get('ES')),
                        'FC': _fmt_iso_date(_kd.get('FC')),
                        'CS': _fmt_iso_date(_kd.get('CS')),
                        'source': 'oneview',
                    }
            except Exception:
                continue

    return {'ES': '', 'FC': '', 'CS': '', 'source': 'none'}

@weekly_summary_bp.route('/api/sp2/refetch_timelines', methods=['POST'])
@login_required
def api_sp2_refetch_timelines():
    """Refetch ES/FC/CS timelines from OneView for a specific PL-ID.
    Updates the sp2_build_consolidate DB row and returns fresh timelines.
    POST body: {week_start, week_end, target, pl_id}
    """
    data      = request.get_json(force=True, silent=True) or {}
    ws        = _safe_date(data.get('week_start'))
    we        = _safe_date(data.get('week_end'))
    target    = str(data.get('target') or '').strip()
    pl_id     = str(data.get('pl_id') or '').strip()
    if not ws or not we or not pl_id:
        return jsonify(success=False, error='week_start, week_end and pl_id are required'), 400

    # DB-first: check dashboard_status before hitting OneView
    _ms = _resolve_pl_milestones(target, pl_id)
    es = _ms['ES']; fc = _ms['FC']; cs = _ms['CS']
    source = _ms['source']

    if not (es or fc or cs):
        return jsonify(success=False,
                       error=f'No milestone dates found in DB or OneView for PL-ID: {pl_id}',
                       source=source), 404

    timelines  = _sp_timeline(es, fc, cs)
    pdt_status = _compute_pdt_test_status(es, cs, fc)

    # Update the sentinel row in sp2_build_consolidate
    sentinel_name = f'__consolidated__{target}__{pl_id}'
    try:
        conn = get_mysql_connection_db(bu_key=None)
        if conn:
            cur = conn.cursor()
            try:
                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                    SET timelines=%s, pdt_test_status=%s, updated_by=%s
                    WHERE week_start=%s AND week_end=%s AND build_name=%s
                """, (timelines, pdt_status, _current_user_identifier(),
                        ws.isoformat(), we.isoformat(), sentinel_name))
                conn.commit()
            finally:
                cur.close(); conn.close()
    except Exception as _db_exc:
        import logging as _log
        _log.getLogger('weekly_summary_routes').warning('[SP2 REFETCH TL] DB update failed: %s', _db_exc)

    return jsonify(success=True, timelines=timelines, pdt_status=pdt_status,
                   es=es, fc=fc, cs=cs, source=source)


@weekly_summary_bp.route('/api/sp2/save_build_type', methods=['POST'])
@login_required
def api_sp2_save_build_type():
    """Save build_type (CRM/Eng) for a build row; triggers background consolidate update."""
    data = request.get_json(force=True, silent=True) or {}
    ws   = _safe_date(data.get('week_start'))
    we   = _safe_date(data.get('week_end'))
    build_name = str(data.get('build_name') or '').strip()
    pl_id      = str(data.get('pl_id') or '').strip()
    build_type = str(data.get('build_type') or 'CRM').strip()
    if build_type not in ('CRM', 'Eng'):
        build_type = 'CRM'
    if not ws or not we or not build_name:
        return jsonify(success=False, error='Missing required fields'), 400
    try:
        _upsert_sp2_build_type(ws, we, build_name, pl_id, build_type)
        try:
            _build_and_save_sp2_consolidate(ws, we, _current_user_identifier())
        except Exception:
            pass
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@weekly_summary_bp.route('/api/sp2/save_pl_rows', methods=['POST'])
@login_required
def api_sp2_save_pl_rows():
    """Save edited hours/crashes/build_type for all builds under a target+PL.
    Called when user clicks the Save button on a PL section.
    """
    data = request.get_json(force=True, silent=True) or {}
    ws     = _safe_date(data.get('week_start'))
    we     = _safe_date(data.get('week_end'))
    target = str(data.get('target') or '').strip()
    pl_id  = str(data.get('pl_id')  or '').strip()
    rows   = data.get('rows') or []
    if not ws or not we or not target or not rows:
        return jsonify(success=False, error='Missing required fields'), 400
    try:
        _ensure_sp2_override_snapshot_columns()
        conn = get_mysql_connection_db(bu_key=None)
        if not conn:
            return jsonify(success=False, error='DB unavailable'), 500
        cur = conn.cursor()
        saved = 0
        try:
            for row in rows:
                build_name = str(row.get('build_name') or '').strip()
                build_type = str(row.get('build_type') or 'CRM').strip()
                hours      = float(row.get('hours') or 0)
                crashes    = int(row.get('crashes') or 0)
                if build_type not in ('CRM', 'Eng'):
                    build_type = 'CRM'
                if not build_name:
                    continue
                # Update the static source row. These values are what Builds and
                # Consolidate will show after refresh; they are no longer added
                # onto sentinel totals.
                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                    SET target=%s,
                        build_type=%s,
                        hours=%s,
                        total_crashes=%s,
                        updated_by=%s,
                        updated_at=CURRENT_TIMESTAMP
                    WHERE week_start=%s AND week_end=%s
                      AND build_name=%s AND pl_id=%s
                """, (target, build_type, hours, crashes, _current_user_identifier(),
                      ws.isoformat(), we.isoformat(), build_name, pl_id))
                if cur.rowcount == 0:
                    cur.execute(f"""
                        INSERT INTO `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                            (week_start, week_end, target, pl_id, build_name,
                             build_type, hours, total_crashes, updated_by)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON DUPLICATE KEY UPDATE
                            target=VALUES(target),
                            build_type=VALUES(build_type),
                            hours=VALUES(hours),
                            total_crashes=VALUES(total_crashes),
                            updated_by=VALUES(updated_by),
                            updated_at=CURRENT_TIMESTAMP
                    """, (ws.isoformat(), we.isoformat(), target, pl_id, build_name,
                          build_type, hours, crashes, _current_user_identifier()))
                saved += 1
            conn.commit()
        finally:
            cur.close(); conn.close()
        _build_and_save_sp2_consolidate(ws, we, _current_user_identifier())
        return jsonify(success=True, saved=saved)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@weekly_summary_bp.route('/api/sp2/save_bu', methods=['POST'])
@login_required
def api_sp2_save_bu():
    """Save BU for a target â€” persists to sp2_build_consolidate for all
    sentinel rows of that target so it survives page reloads.
    """
    data   = request.get_json(force=True, silent=True) or {}
    ws     = _safe_date(data.get('week_start'))
    we     = _safe_date(data.get('week_end'))
    target = str(data.get('target') or '').strip()
    bu     = str(data.get('bu')     or '').strip()
    if not ws or not we or not target or not bu:
        return jsonify(success=False, error='Missing required fields'), 400
    try:
        _ensure_sp2_build_consolidate_table()
        _ensure_sp2_override_snapshot_columns()
        conn = get_mysql_connection_db(bu_key=None)
        if not conn:
            return jsonify(success=False, error='DB unavailable'), 500
        cur = conn.cursor()
        try:
            pl_id = str(data.get('pl_id') or '').strip()

            # 1. Persist BU in the static Builds source table too. This is the
            #    main source for /api/sp2/builds and survives consolidate rebuilds.
            if pl_id:
                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                    SET bu=%s, target=%s, updated_at=CURRENT_TIMESTAMP
                    WHERE week_start=%s AND week_end=%s
                      AND pl_id=%s
                """, (bu, target, ws.isoformat(), we.isoformat(), pl_id))
            else:
                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                    SET bu=%s, target=%s, updated_at=CURRENT_TIMESTAMP
                    WHERE week_start=%s AND week_end=%s
                      AND (target=%s OR pl_id LIKE %s)
                """, (bu, target, ws.isoformat(), we.isoformat(), target, target + '%'))

            # 2. Update ALL existing consolidate sentinel rows for this target+week.
            cur.execute(f"""
                UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                SET bu=%s, updated_at=CURRENT_TIMESTAMP
                WHERE week_start=%s AND week_end=%s
                  AND target=%s
            """, (bu, ws.isoformat(), we.isoformat(), target))

            # 3. If no rows existed yet, insert a target-level placeholder
            #    so the BU is persisted even before consolidate runs
            if cur.rowcount == 0:
                cur.execute(f"""
                    INSERT INTO `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                        (week_start, week_end, target, pl_id, build_name,
                         build_type, bu, updated_by)
                    VALUES (%s,%s,%s,'','__bu_placeholder__{target}',
                            'CRM',%s,%s)
                    ON DUPLICATE KEY UPDATE
                        bu=VALUES(bu),
                        updated_at=CURRENT_TIMESTAMP
                """, (ws.isoformat(), we.isoformat(), target,
                       bu, _current_user_identifier()))
            conn.commit()
        finally:
            cur.close(); conn.close()
        return jsonify(success=True)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500



@weekly_summary_bp.route('/api/sp2/save_target', methods=['POST'])
@login_required
def api_sp2_save_target():
    """Save everything for a target in one shot:
       - BU (target-level)
       - hours / crashes / build_type for every build row under the target
    Replaces the separate save_bu + per-PL save_pl_rows flow.
    """
    data   = request.get_json(force=True, silent=True) or {}
    ws     = _safe_date(data.get('week_start'))
    we     = _safe_date(data.get('week_end'))
    target = str(data.get('target') or '').strip()
    bu     = str(data.get('bu')     or '').strip()
    rows   = data.get('rows') or []   # all build rows across all PLs for this target

    if not ws or not we or not target:
        return jsonify(success=False, error='Missing required fields'), 400

    try:
        _ensure_sp2_build_consolidate_table()
        _ensure_sp2_override_snapshot_columns()
        conn = get_mysql_connection_db(bu_key=None)
        if not conn:
            return jsonify(success=False, error='DB unavailable'), 500
        cur = conn.cursor()
        saved_rows = 0
        try:
            user = _current_user_identifier()

            # ── 1. Update BU on every override row for this target ──────────
            if bu:
                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                    SET bu=%s, target=%s, updated_at=CURRENT_TIMESTAMP
                    WHERE week_start=%s AND week_end=%s
                      AND (target=%s OR pl_id LIKE %s)
                """, (bu, target,
                      ws.isoformat(), we.isoformat(),
                      target, target + '%'))

                # Also stamp consolidate rows
                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                    SET bu=%s, updated_at=CURRENT_TIMESTAMP
                    WHERE week_start=%s AND week_end=%s AND target=%s
                """, (bu, ws.isoformat(), we.isoformat(), target))

                if cur.rowcount == 0:
                    # No consolidate rows yet — insert placeholder so BU persists
                    cur.execute(f"""
                        INSERT INTO `{_QIPL_DB}`.`{_SP2_BUILD_CONSOLIDATE_TABLE}`
                            (week_start, week_end, target, pl_id, build_name,
                             build_type, bu, updated_by)
                        VALUES (%s,%s,%s,'','__bu_placeholder__',
                                'CRM',%s,%s)
                        ON DUPLICATE KEY UPDATE
                            bu=VALUES(bu), updated_at=CURRENT_TIMESTAMP
                    """, (ws.isoformat(), we.isoformat(), target, bu, user))

            # ── 2. Update each build row (hours / crashes / build_type) ─────
            for row in rows:
                build_name = str(row.get('build_name') or '').strip()
                pl_id      = str(row.get('pl_id')      or '').strip()
                build_type = str(row.get('build_type') or 'CRM').strip()
                hours      = float(row.get('hours')    or 0)
                crashes    = int(row.get('crashes')    or 0)
                if build_type not in ('CRM', 'Eng'):
                    build_type = 'CRM'
                if not build_name:
                    continue

                cur.execute(f"""
                    UPDATE `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                    SET target=%s, bu=%s, build_type=%s,
                        hours=%s, total_crashes=%s,
                        updated_by=%s, updated_at=CURRENT_TIMESTAMP
                    WHERE week_start=%s AND week_end=%s
                      AND build_name=%s AND pl_id=%s
                """, (target, bu or None, build_type,
                      hours, crashes, user,
                      ws.isoformat(), we.isoformat(),
                      build_name, pl_id))

                if cur.rowcount == 0:
                    cur.execute(f"""
                        INSERT INTO `{_QIPL_DB}`.`{_SP2_BUILD_TYPE_OVERRIDES_TABLE}`
                            (week_start, week_end, target, pl_id, build_name,
                             build_type, hours, total_crashes, bu, updated_by)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                        ON DUPLICATE KEY UPDATE
                            target=VALUES(target),
                            bu=VALUES(bu),
                            build_type=VALUES(build_type),
                            hours=VALUES(hours),
                            total_crashes=VALUES(total_crashes),
                            updated_by=VALUES(updated_by),
                            updated_at=CURRENT_TIMESTAMP
                    """, (ws.isoformat(), we.isoformat(),
                          target, pl_id, build_name,
                          build_type, hours, crashes,
                          bu or None, user))
                saved_rows += 1

            conn.commit()
        finally:
            cur.close(); conn.close()

        # Rebuild consolidate snapshot after save
        try:
            _build_and_save_sp2_consolidate(ws, we, _current_user_identifier())
        except Exception:
            pass

        return jsonify(success=True, saved=saved_rows)
    except Exception as e:
        return jsonify(success=False, error=str(e)), 500


@weekly_summary_bp.route('/api/sp2/unique_devices')
@login_required
def api_sp2_unique_devices():
    """Return unique device (chip_id) list per target+PL from Axiom for the week.

    Groups chip_ids by target -> pl_id, computes per-device hours split
    proportionally across builds that used that chip.
    """
    ws_arg = request.args.get('week_start', '').strip()
    we_arg = request.args.get('week_end', '').strip()
    ws = _safe_date(ws_arg)
    we = _safe_date(we_arg)
    if not ws or not we:
        ws, we = _selected_week_from_request()

    payload, _src = _load_swpdt_json_payload()
    raw_builds = _flatten_swpdt_build_entries(payload)

    from collections import defaultdict
    target_pl_chip_hours = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for b in raw_builds:
        sub_dt = _axiom_date_from_value(b.get('submitted'))
        if not sub_dt or not (ws <= sub_dt <= we):
            continue
        pl_id = str(b.get('software_product') or '').strip()
        target = _swpdt_target_from_product(pl_id) or pl_id
        chip_ids = b.get('chip_ids') or []
        hours = float(b.get('hours') or 0)
        per_chip = (hours / len(chip_ids)) if chip_ids else 0
        for chip in chip_ids:
            chip = str(chip).strip()
            if chip:
                target_pl_chip_hours[target][pl_id][chip] += per_chip

    by_target = []
    total_devices = 0
    for target in sorted(target_pl_chip_hours.keys()):
        pl_map = target_pl_chip_hours[target]
        all_target_chips = defaultdict(float)
        for pl_id, chip_hours in pl_map.items():
            for chip, hrs in chip_hours.items():
                all_target_chips[chip] += hrs
        total_devices += len(all_target_chips)
        max_hours = max(all_target_chips.values()) if all_target_chips else 1

        pl_entries = []
        for pl_id in sorted(pl_map.keys()):
            chip_hours = pl_map[pl_id]
            devices = [
                {'chip_id': chip, 'hours': round(hrs, 2),
                 'pl_id': pl_id, 'software_product': pl_id}
                for chip, hrs in sorted(chip_hours.items(), key=lambda x: -x[1])
            ]
            pl_entries.append({'pl_id': pl_id, 'devices': devices,
                               'total_hours': round(sum(chip_hours.values()), 2)})

        flat_devices = [
            {'chip_id': chip, 'hours': round(hrs, 2),
             'pl_id': ', '.join(
                 p for p in sorted(pl_map.keys()) if chip in pl_map[p]
             )}
            for chip, hrs in sorted(all_target_chips.items(), key=lambda x: -x[1])
        ]

        by_target.append({
            'target':      target,
            'total_chips': len(all_target_chips),
            'total_hours': round(sum(all_target_chips.values()), 2),
            'max_hours':   round(max_hours, 2),
            'pl_entries':  pl_entries,
            'devices':     flat_devices,
        })

    return jsonify(
        success=True,
        by_target=by_target,
        total_devices=total_devices,
        week_start=ws.isoformat(),
        week_end=we.isoformat(),
    )

