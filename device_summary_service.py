import logging
import os
import json
import openpyxl
from datetime import datetime as _dt

logger = logging.getLogger(__name__)
from src.axiom_client import get_devices_by_chipset
from config import AXIOM_ENABLED_CHIPS

# ── paths ────────────────────────────────────────────────────────────────────
_ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
# Persistent user/data storage. Do NOT keep generated Excel/config under static/,
# because static/ can be replaced when Buddy is rebuilt/recompiled/redeployed.
_PDTBUDDY_DATA_ROOT = os.environ.get(
    'PDTBUDDY_DATA_ROOT',
    r'\\sphere\pdtqipl_internal\PDTBuddy'
)
_DS_STATIC_DIR      = os.path.join(_PDTBUDDY_DATA_ROOT, 'device_summary_data')
_DS_CONFIG_PATH     = os.path.join(_PDTBUDDY_DATA_ROOT, 'config', 'target_excel_page_config.json')
_OLD_STATIC_CONFIG_PATH = os.path.join(_ROOT_DIR, 'static', 'target_excel_page_config.json')
_MANAGED_EXCEL_ROOT = os.path.join(_PDTBUDDY_DATA_ROOT, 'managed_excel')


# ── generic JSON helpers ──────────────────────────────────────────────────────
def _load_json(path):
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            return json.load(fh)
    except Exception:
        return {}


def _save_json(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)
    return data


# ── Excel config (per-target, per-page) ──────────────────────────────────────
def _is_static_app_path(path_value):
    try:
        p = os.path.abspath(_normalize_path(path_value))
        static_root = os.path.abspath(os.path.join(_ROOT_DIR, 'static'))
        return p == static_root or p.startswith(static_root + os.sep)
    except Exception:
        return False


def get_ds_excel_config(target_name):
    """Return device_summary config dict for a target.

    Network config is authoritative. If it does not exist yet, migrate from the
    legacy static config once, but ignore stale static/managed paths because
    static/ is not persistent after rebuilds.
    """
    data = _load_json(_DS_CONFIG_PATH)
    cfg = (data.get(target_name) or {}).get('device_summary', {}) if isinstance(data, dict) else {}
    if cfg:
        return cfg

    old_data = _load_json(_OLD_STATIC_CONFIG_PATH)
    old_cfg = (old_data.get(target_name) or {}).get('device_summary', {}) if isinstance(old_data, dict) else {}
    if old_cfg and not _is_static_app_path(old_cfg.get('excel_path', '')):
        save_ds_excel_config(
            target_name,
            old_cfg.get('excel_path', ''),
            old_cfg.get('summary_sheet') or 'SW PDT Summary',
            old_cfg.get('devices_sheet') or 'Devices',
            old_cfg.get('data_mode') or 'excel',
        )
        return old_cfg
    return {}


def save_ds_excel_config(target_name, excel_path, summary_sheet, devices_sheet, data_mode='excel'):
    """Persist device_summary Excel config for a target."""
    all_cfg = _load_json(_DS_CONFIG_PATH)
    if not isinstance(all_cfg, dict):
        all_cfg = {}
    target_cfg = all_cfg.get(target_name) or {}
    target_cfg['device_summary'] = {
        'excel_path':    excel_path,
        'summary_sheet': summary_sheet,
        'devices_sheet': devices_sheet,
        'data_mode':     data_mode,
        'updated_at':    _dt.utcnow().isoformat() + 'Z',
    }
    all_cfg[target_name] = target_cfg
    _save_json(_DS_CONFIG_PATH, all_cfg)
    return target_cfg['device_summary']


# ── static device-summary data (JSON store) ───────────────────────────────────
def _static_path(target_name):
    return os.path.join(_DS_STATIC_DIR, f'{target_name.lower()}.json')


def load_static_data(target_name):
    return _load_json(_static_path(target_name))


def save_static_data(target_name, data):
    return _save_json(_static_path(target_name), data)


# ── Excel helpers ─────────────────────────────────────────────────────────────
def _normalize_path(path_value):
    if not path_value:
        return ''
    p = str(path_value).strip()
    return os.path.expanduser(p)


def get_sheet_names(excel_path):
    """Return list of sheet names from an Excel file."""
    path = _normalize_path(excel_path)
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f'Excel file not found: {path}')
    xl = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = list(xl.sheetnames)
    xl.close()
    return names


def _build_merge_map(ws):
    """Return dict (row,col)->value for all merged cells."""
    merge_map = {}
    for mr in list(ws.merged_cells.ranges):
        val = ws.cell(mr.min_row, mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = val
    return merge_map


def _cv(ws, merge_map, r, c):
    """Cell value respecting merge map; returns '' for None."""
    v = merge_map.get((r, c), ws.cell(r, c).value)
    if v is None:
        return ''
    if isinstance(v, _dt):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()


def _safe_int(v):
    try:
        s = str(v).strip()
        if s in ('', '-', '—', 'None', 'nan', 'none'):
            return 0
        return int(float(s))
    except Exception:
        return 0


def _has_number(v):
    try:
        s = str(v).strip()
        if s in ('', '-', '—', 'None', 'nan', 'none'):
            return False
        float(s)
        return True
    except Exception:
        return False


def _is_hidden_row(ws, row_number):
    try:
        dim = ws.row_dimensions[row_number]
        return bool(getattr(dim, 'hidden', False))
    except Exception:
        return False


def _is_del_header(value):
    text = str(value or '').strip().upper().replace('.', '')
    return text in ('DEL', 'DELIVERED') or text.startswith('DELIV')


def _is_dep_header(value):
    text = str(value or '').strip().upper().replace('.', '')
    return text in ('DEP', 'DEPLOYED') or text.startswith('DEPLO')


# ── Deployment table parser ───────────────────────────────────────────────────
def build_deployment_table(excel_path, sheet_name):
    """
    Parse SW PDT deployment table from Excel.
    Returns dict: {sites, rows, totals, grand_del, grand_dep, mcn_chart_data}
    grand_del/grand_dep = Excel Grand Total row values if found,
    else sum of all data row totals.
    """
    path = _normalize_path(excel_path)
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f'Excel file not found: {path}')
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found.')
    ws = wb[sheet_name]
    mm = _build_merge_map(ws)
    cv = lambda r, c: _cv(ws, mm, r, c)

    max_col = ws.max_column
    max_row = ws.max_row

    # ── Read header rows ──────────────────────────────────────────────────────
    # Find which row is the column-header row (row1) and sub-header row (row2).
    # Some sheets have a title row above the actual headers, so scan the first
    # 5 rows for the one that contains MCN / Form Factor.
    header_row_idx = 1
    for probe in range(1, 6):
        probe_vals = [cv(probe, c).upper() for c in range(1, max_col + 1)]
        if any('MCN' in v or 'FORM' in v for v in probe_vals):
            header_row_idx = probe
            break
    sub_header_row_idx = header_row_idx + 1
    data_start_row = sub_header_row_idx + 1

    row1 = [cv(header_row_idx, c) for c in range(1, max_col + 1)]
    row2 = [cv(sub_header_row_idx, c) for c in range(1, max_col + 1)]

    logger.debug(f'[DS PARSER] header_row={header_row_idx} row1={row1}')
    logger.debug(f'[DS PARSER] sub_header_row={sub_header_row_idx} row2={row2}')

    # ── Detect fixed columns ──────────────────────────────────────────────────
    SKIP = {'FORM FACTOR','FORM','FACTOR','MCN','STORAGE','STORAGE TYPE',
            'TOTAL','GRAND TOTAL','','DELIVERED','DEPLOYED',
            'DEL','DEL.','DEP','DEP.'}
    ff_col = mcn_col = sto_col = None
    for i, v in enumerate(row1):
        vu = v.upper().replace('.', '')
        if ('FORM' in vu or vu == 'FF') and ff_col is None:  ff_col  = i
        elif 'MCN' in vu and mcn_col is None:                 mcn_col = i
        elif 'STOR' in vu and sto_col is None:                sto_col = i

    # ── Detect site + total columns ───────────────────────────────────────────
    TOTAL_W = {'TOTAL', 'GRAND TOTAL'}
    site_cols  = {}  # site_name -> {del: 0-based-idx, dep: 0-based-idx}
    total_cols = {'del': None, 'dep': None}
    i = 0
    while i < len(row1):
        v, vu = row1[i], row1[i].upper().strip()
        if v and vu in TOTAL_W:
            for j in range(i, min(i + 4, len(row2))):
                if _is_del_header(row2[j]) and total_cols['del'] is None:
                    total_cols['del'] = j
                elif _is_dep_header(row2[j]) and total_cols['dep'] is None:
                    total_cols['dep'] = j
            i = max(x for x in [total_cols['del'], total_cols['dep'], i] if x is not None) + 1
            continue
        if v and vu not in SKIP and vu not in TOTAL_W:
            del_idx = dep_idx = None
            for j in range(i, min(i + 4, len(row2))):
                if _is_del_header(row2[j]) and del_idx is None:   del_idx = j
                elif _is_dep_header(row2[j]) and dep_idx is None: dep_idx = j
            if del_idx is not None or dep_idx is not None:
                site_cols[v] = {'del': del_idx, 'dep': dep_idx}
                i = max(x for x in [del_idx, dep_idx, i] if x is not None) + 1
                continue
        i += 1

    logger.debug(f'[DS PARSER] site_cols={site_cols} total_cols={total_cols}')

    if not site_cols:
        return None

    sites  = list(site_cols.keys())
    totals = {s: {'del': 0, 'dep': 0} for s in sites}
    rows_out = []
    grand_del = grand_dep = 0
    excel_grand_del = excel_grand_dep = None
    last_ff = ''

    for rn in range(data_start_row, max_row + 1):
        if _is_hidden_row(ws, rn):
            continue
        row = [cv(rn, c) for c in range(1, max_col + 1)]
        if not any(row):
            continue

        ff  = row[ff_col]  if ff_col  is not None and ff_col  < len(row) else ''
        mcn = row[mcn_col] if mcn_col is not None and mcn_col < len(row) else ''
        sto = row[sto_col] if sto_col is not None and sto_col < len(row) else ''

        # ── Detect Grand Total footer row ─────────────────────────────────────
        row_text = ' '.join(str(x or '').strip().upper() for x in row)
        ff_upper = str(ff).strip().upper()
        is_footer = (
            'GRAND TOTAL' in row_text
            or ff_upper in ('TOTAL', 'GRAND TOTAL')
            or (not mcn and any(_has_number(x) for x in row)  # numeric row with no MCN = likely footer
                and rn > data_start_row + 2)
        )
        if is_footer:
            # Try configured total columns first
            td = row[total_cols['del']] if total_cols['del'] is not None and total_cols['del'] < len(row) else ''
            tp = row[total_cols['dep']] if total_cols['dep'] is not None and total_cols['dep'] < len(row) else ''
            if _has_number(td): excel_grand_del = _safe_int(td)
            if _has_number(tp): excel_grand_dep = _safe_int(tp)
            # Fallback: last two numbers in the row
            nums = [_safe_int(x) for x in row if _has_number(x)]
            logger.debug(f'[DS PARSER] footer row {rn}: nums={nums} td={td!r} tp={tp!r}')
            if nums:
                if excel_grand_del is None: excel_grand_del = nums[-2] if len(nums) >= 2 else nums[-1]
                if excel_grand_dep is None: excel_grand_dep = nums[-1]
            continue

        if not mcn:
            continue
        if ff:  last_ff = ff
        else:   ff = last_ff

        site_data = {}
        site_del_sum = site_dep_sum = 0
        for site, cols in site_cols.items():
            d = _safe_int(row[cols['del']]) if cols['del'] is not None and cols['del'] < len(row) else 0
            p = _safe_int(row[cols['dep']]) if cols['dep'] is not None and cols['dep'] < len(row) else 0
            site_data[site] = {'del': d, 'dep': p}
            totals[site]['del'] += d
            totals[site]['dep'] += p
            site_del_sum += d
            site_dep_sum += p

        td_raw = row[total_cols['del']] if total_cols['del'] is not None and total_cols['del'] < len(row) else ''
        tp_raw = row[total_cols['dep']] if total_cols['dep'] is not None and total_cols['dep'] < len(row) else ''
        row_del = _safe_int(td_raw) if _has_number(td_raw) else site_del_sum
        row_dep = _safe_int(tp_raw) if _has_number(tp_raw) else site_dep_sum

        grand_del += row_del
        grand_dep += row_dep
        rows_out.append({'form_factor': ff, 'mcn': mcn, 'storage': sto,
                         'sites': site_data, 'total_del': row_del, 'total_dep': row_dep})

    logger.debug(f'[DS PARSER] computed grand_del={grand_del} excel_grand_del={excel_grand_del}')

    # Always prefer Excel Grand Total row over computed sum
    if excel_grand_del is not None:
        grand_del = excel_grand_del
    if excel_grand_dep is not None:
        grand_dep = excel_grand_dep

    mcn_agg = {}
    for r in rows_out:
        k = r['mcn'] or 'Unknown'
        mcn_agg[k] = mcn_agg.get(k, 0) + r['total_del']
    mcn_chart = [{'mcn': k, 'total': v}
                 for k, v in sorted(mcn_agg.items(), key=lambda x: -x[1])]

    return {'sites': sites, 'rows': rows_out, 'totals': totals,
            'grand_del': grand_del, 'grand_dep': grand_dep,
            'mcn_chart_data': mcn_chart,
            'computed_del': sum(r['total_del'] for r in rows_out),
            'excel_grand_del_found': excel_grand_del}


# ── Devices list reader ───────────────────────────────────────────────────────
# Mandatory display columns (always shown in popup + add-device form)
DEVICE_MANDATORY_COLS = [
    'Serial', 'DDR Type', 'Memory bundle', 'Storage', 'Platform', 'MCN',
    'Axiom Deploy', 'Station'
]

DEFAULT_DEVICE_HEADERS = [
    'S.No', 'Asset Tag', 'Serial', 'Adb id', 'DDR Type', 'Memory bundle',
    'Storage', 'UFS', 'RF card', 'Platform', 'QRDVersion', 'MCN',
    'Receive Date', 'Axiom Deploy', 'INT/EXT', 'Team', 'Station',
    'Rework Part Details', 'Rework date', 'Rework Type', 'DDR rework',
    'Location', 'WCN Card', 'Comments'
]


def _safe_segment(value):
    import re
    text = str(value or '').strip() or 'UNKNOWN'
    return re.sub(r'[^A-Za-z0-9_.-]+', '_', text).strip('_') or 'UNKNOWN'


def get_managed_target_dirs(target_name):
    """Return and create managed BU/Target folders for Device Summary and MTBF Excel files."""
    try:
        from dashboard_common import get_bu_for_target
        bu_key = get_bu_for_target(target_name) or 'UNKNOWN_BU'
    except Exception:
        bu_key = 'UNKNOWN_BU'
    base = os.path.join(_MANAGED_EXCEL_ROOT, _safe_segment(bu_key).upper(), _safe_segment(target_name).upper())
    devices_dir = os.path.join(base, 'Devices')
    mtbf_dir = os.path.join(base, 'MTBF')
    os.makedirs(devices_dir, exist_ok=True)
    os.makedirs(mtbf_dir, exist_ok=True)
    return {'base': base, 'devices': devices_dir, 'mtbf': mtbf_dir, 'bu_key': bu_key}


def ensure_device_summary_workbook(target_name, excel_path='', summary_sheet='SW PDT Summary', devices_sheet='Devices'):
    """Use configured workbook or create a managed BU/Target/Devices workbook if missing."""
    path = _normalize_path(excel_path)
    if not path:
        dirs = get_managed_target_dirs(target_name)
        path = os.path.join(dirs['devices'], f'{_safe_segment(target_name).upper()}_device_summary.xlsx')
    os.makedirs(os.path.dirname(path), exist_ok=True)

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = summary_sheet or 'SW PDT Summary'

    if summary_sheet and summary_sheet not in wb.sheetnames:
        ws_summary = wb.create_sheet(summary_sheet)
        ws_summary.append(['Form Factor', 'MCN', 'Storage', 'QIPL', '', 'CH', '', 'SD', '', 'Total', ''])
        ws_summary.append(['', '', '', 'Delivered', 'Deployed', 'Delivered', 'Deployed', 'Delivered', 'Deployed', 'Delivered', 'Deployed'])

    if devices_sheet not in wb.sheetnames:
        ws_devices = wb.create_sheet(devices_sheet)
        ws_devices.append(DEFAULT_DEVICE_HEADERS)
    else:
        ws_devices = wb[devices_sheet]
        if ws_devices.max_row < 1 or not any(ws_devices.cell(1, c).value for c in range(1, ws_devices.max_column + 1)):
            ws_devices.append(DEFAULT_DEVICE_HEADERS)

    wb.save(path)
    wb.close()
    return path, summary_sheet, devices_sheet


def _get_excel_device_context(target_name, create_if_missing=True):
    cfg = get_ds_excel_config(target_name) or {}
    excel_path = cfg.get('excel_path', '')
    summary_sheet = cfg.get('summary_sheet') or 'SW PDT Summary'
    devices_sheet = cfg.get('devices_sheet') or 'Devices'
    if create_if_missing and (not excel_path or not os.path.exists(_normalize_path(excel_path))):
        excel_path, summary_sheet, devices_sheet = ensure_device_summary_workbook(
            target_name, excel_path, summary_sheet, devices_sheet
        )
        save_ds_excel_config(target_name, excel_path, summary_sheet, devices_sheet, cfg.get('data_mode', 'excel'))
    return excel_path, summary_sheet, devices_sheet


def get_or_create_device_excel_config(target_name):
    excel_path, summary_sheet, devices_sheet = _get_excel_device_context(target_name, create_if_missing=True)
    return get_ds_excel_config(target_name) or save_ds_excel_config(
        target_name, excel_path, summary_sheet, devices_sheet, 'excel'
    )


def _normalize_device_row(row, width, row_number=None):
    values = [str(v) if v is not None else '' for v in (row or [])]
    values = (values + [''] * width)[:width]
    if width and row_number is not None:
        values[0] = str(row_number)
    return values


def read_all_devices_sheet(excel_path, sheet_name):
    headers, rows, total = read_devices_sheet(excel_path, sheet_name, max_rows=100000)
    return headers, rows, total


def replace_devices_sheet_rows(excel_path, sheet_name, headers, rows):
    """Rewrite the device sheet with headers + rows, preserving a simple Excel source of truth."""
    from copy import copy
    path = _normalize_path(excel_path)
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    headers = list(headers or DEFAULT_DEVICE_HEADERS)
    width = len(headers)

    hdr_styles = []
    if ws.max_row >= 1:
        for c in range(1, max(width, ws.max_column) + 1):
            cell = ws.cell(1, c)
            hdr_styles.append({
                'font': copy(cell.font), 'fill': copy(cell.fill),
                'alignment': copy(cell.alignment), 'border': copy(cell.border),
            })

    ws.delete_rows(1, ws.max_row or 1)
    for c, val in enumerate(headers, 1):
        cell = ws.cell(1, c, val)
        if c <= len(hdr_styles):
            st = hdr_styles[c - 1]
            cell.font = st['font']; cell.fill = st['fill']
            cell.alignment = st['alignment']; cell.border = st['border']

    normalized = []
    for idx, row in enumerate(rows or [], 1):
        vals = _normalize_device_row(row, width, idx)
        if not any(v for v in vals[1:] if width > 1):
            continue
        normalized.append(vals)
        for c, val in enumerate(vals, 1):
            ws.cell(idx + 1, c, val)

    wb.save(path)
    wb.close()
    return len(normalized)


def append_device_row(target_name, row):
    excel_path, _summary_sheet, devices_sheet = _get_excel_device_context(target_name, create_if_missing=True)
    headers, rows, _total = read_all_devices_sheet(excel_path, devices_sheet)
    if not headers:
        headers = DEFAULT_DEVICE_HEADERS
    rows.append(row)
    total = replace_devices_sheet_rows(excel_path, devices_sheet, headers, rows)
    return {'excel_path': excel_path, 'devices_sheet': devices_sheet, 'headers': headers, 'total': total}


def edit_device_row(target_name, index, row):
    excel_path, _summary_sheet, devices_sheet = _get_excel_device_context(target_name, create_if_missing=True)
    headers, rows, _total = read_all_devices_sheet(excel_path, devices_sheet)
    if index < 0 or index >= len(rows):
        raise ValueError(f'Invalid index {index}')
    rows[index] = row
    total = replace_devices_sheet_rows(excel_path, devices_sheet, headers or DEFAULT_DEVICE_HEADERS, rows)
    return {'excel_path': excel_path, 'devices_sheet': devices_sheet, 'headers': headers, 'total': total, 'row': row}


def delete_device_row(target_name, index):
    excel_path, _summary_sheet, devices_sheet = _get_excel_device_context(target_name, create_if_missing=True)
    headers, rows, _total = read_all_devices_sheet(excel_path, devices_sheet)
    if index < 0 or index >= len(rows):
        raise ValueError(f'Invalid index {index}')
    removed = rows.pop(index)
    total = replace_devices_sheet_rows(excel_path, devices_sheet, headers or DEFAULT_DEVICE_HEADERS, rows)
    return {'excel_path': excel_path, 'devices_sheet': devices_sheet, 'headers': headers, 'total': total, 'removed': removed}


def read_devices_sheet(excel_path, sheet_name, max_rows=1000):
    """
    Read devices list sheet.
    Returns (headers: list[str], rows: list[list[str]], total: int)
    """
    path = _normalize_path(excel_path)
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f'Excel file not found: {path}')
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Sheet "{sheet_name}" not found.')
    ws = wb[sheet_name]
    mm = _build_merge_map(ws)
    cv = lambda r, c: _cv(ws, mm, r, c)

    max_col = ws.max_column
    headers = [cv(1, c) for c in range(1, max_col + 1)]
    while headers and not headers[-1]:
        headers.pop()
    ncols = len(headers)

    rows = []
    for rn in range(2, ws.max_row + 1):
        row = [cv(rn, c) for c in range(1, ncols + 1)]
        if not any(row): continue
        rows.append(row)
    total = len(rows)
    return headers, rows[:max_rows], total


# ── Build deployment table from devices sheet ────────────────────────────────
def build_deployment_table_from_devices(excel_path, devices_sheet, prefiltered_headers=None, prefiltered_rows=None):
    """
    Build the SW PDT deployment table by aggregating the Devices sheet.
    Groups by MCN + Storage, counts per Location (QIPL / CH / SD / etc.).
    This is the fallback when no separate summary sheet exists.
    Pass prefiltered_headers/prefiltered_rows to use already-filtered data
    (e.g. after project filter) instead of re-reading from Excel.
    """
    if prefiltered_headers is not None and prefiltered_rows is not None:
        headers, rows = prefiltered_headers, prefiltered_rows
    else:
        headers, rows, _ = read_devices_sheet(excel_path, devices_sheet, max_rows=100000)
    if not headers or not rows:
        return None

    # Find column indices (case-insensitive)
    def _col_idx(names):
        for name in names:
            for i, h in enumerate(headers):
                if name.lower() in str(h or '').lower():
                    return i
        return None

    mcn_idx  = _col_idx(['mcn'])
    sto_idx  = _col_idx(['storage'])
    loc_idx  = _col_idx(['location'])
    ff_idx   = _col_idx(['form factor', 'platform', 'ff'])
    dep_idx  = _col_idx(['axiom deploy', 'deployed', 'deploy'])

    if mcn_idx is None:
        return None

    SITE_ORDER = ['QIPL', 'CH', 'SD']

    # Aggregate: key = (form_factor, mcn, storage) -> site -> count
    from collections import defaultdict
    agg   = defaultdict(lambda: defaultdict(int))
    sites_seen = set()

    for row in rows:
        def _get(idx):
            if idx is None or idx >= len(row): return ''
            return str(row[idx] or '').strip()

        mcn = _get(mcn_idx)
        if not mcn or mcn in ('-', '—'):
            continue

        sto = _get(sto_idx) or 'N/A'
        ff  = _get(ff_idx)  or 'Unknown'

        # Normalise location to site abbreviation
        raw_loc = _get(loc_idx).upper()
        if 'QIPL' in raw_loc or 'HYD' in raw_loc:
            site = 'QIPL'
        elif raw_loc in ('CH', 'CHINA', 'SH', 'SHANGHAI'):
            site = 'CH'
        elif raw_loc in ('SD', 'SAN DIEGO', 'SANDIEGO'):
            site = 'SD'
        elif raw_loc:
            site = raw_loc  # keep unknown sites as-is
        else:
            site = 'QIPL'  # default

        sites_seen.add(site)
        key = (ff, mcn, sto)
        agg[key][site] += 1

    if not agg:
        return None

    # Order sites: known order first, then any extras alphabetically
    sites = [s for s in SITE_ORDER if s in sites_seen] + \
            sorted(s for s in sites_seen if s not in SITE_ORDER)

    totals   = {s: {'del': 0, 'dep': 0} for s in sites}
    rows_out = []
    grand_del = 0
    last_ff   = ''

    for (ff, mcn, sto), site_counts in sorted(agg.items()):
        site_data = {}
        row_del   = 0
        for s in sites:
            cnt = site_counts.get(s, 0)
            site_data[s] = {'del': cnt, 'dep': cnt}
            totals[s]['del'] += cnt
            totals[s]['dep'] += cnt
            row_del += cnt
        grand_del += row_del
        rows_out.append({
            'form_factor': ff,
            'mcn':         mcn,
            'storage':     sto,
            'sites':       site_data,
            'total_del':   row_del,
            'total_dep':   row_del,
        })

    mcn_agg = {}
    for r in rows_out:
        k = r['mcn'] or 'Unknown'
        mcn_agg[k] = mcn_agg.get(k, 0) + r['total_del']
    mcn_chart = [{'mcn': k, 'total': v}
                 for k, v in sorted(mcn_agg.items(), key=lambda x: -x[1])]

    return {
        'sites':       sites,
        'rows':        rows_out,
        'totals':      totals,
        'grand_del':   grand_del,
        'grand_dep':   grand_del,
        'mcn_chart_data': mcn_chart,
        'source':      'devices_sheet',
    }


# ── Page data loader (called by route) ───────────────────────────────────────
def _find_latest_excel_in_dir(dir_path):
    try:
        if not dir_path or not os.path.isdir(dir_path):
            return ''
        candidates = []
        for name in os.listdir(dir_path):
            low = name.lower()
            if low.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')) and not name.startswith('~$'):
                full = os.path.join(dir_path, name)
                candidates.append((os.path.getmtime(full), full))
        candidates.sort(reverse=True)
        return candidates[0][1] if candidates else ''
    except Exception:
        return ''


def _get_compute_default_device_excel(target_name):
    try:
        from dashboard_common import get_bu_for_target
        bu = (get_bu_for_target(target_name) or '').upper()
        if bu != 'COMPUTE':
            return '', ''
    except Exception:
        return '', ''
    devices_dir = os.path.join(_MANAGED_EXCEL_ROOT, 'COMPUTE', _safe_segment(target_name).upper(), 'Devices')
    latest = _find_latest_excel_in_dir(devices_dir)
    return latest, 'LATEST DEVICE DATA_ALL'


def _filter_compute_project(devices_headers, devices_rows, project_filter):
    project_filter = str(project_filter or '').strip()
    if not devices_headers or not devices_rows:
        return devices_headers, devices_rows, []
    proj_idx = next((i for i, h in enumerate(devices_headers) if str(h or '').strip().upper() == 'PROJECT'), None)
    if proj_idx is None:
        return devices_headers, devices_rows, []

    def _canon_project(value):
        """Normalize project names so KALAMBO/Kalambo/kalambo are treated as one project."""
        return str(value or '').strip().upper()

    normalized_rows = []
    project_values = set()
    for row in devices_rows:
        new_row = list(row)
        project_value = _canon_project(new_row[proj_idx] if proj_idx < len(new_row) else '')
        if project_value:
            project_values.add(project_value)
            if proj_idx < len(new_row):
                new_row[proj_idx] = project_value
        normalized_rows.append(new_row)

    values = sorted(project_values)
    if not project_filter or project_filter.lower() == 'all':
        return devices_headers, normalized_rows, values

    wanted = _canon_project(project_filter)
    filtered = [r for r in normalized_rows if _canon_project(r[proj_idx] if proj_idx < len(r) else '') == wanted]
    return devices_headers, filtered, values



def load_page_data(target_name, project_filter='All'):
    """
    Load all data needed for device_summary_page.
    Returns dict with keys:
      excel_path, summary_sheet, devices_sheet, data_mode,
      sheet_names, sw_table, devices_headers, devices_rows,
      devices_total, page_error
    """
    cfg           = get_ds_excel_config(target_name)
    excel_path    = cfg.get('excel_path', '')
    summary_sheet = cfg.get('summary_sheet', '')
    devices_sheet = cfg.get('devices_sheet', '')
    data_mode     = cfg.get('data_mode', 'excel')

    if not excel_path:
        compute_excel, compute_sheet = _get_compute_default_device_excel(target_name)
        if compute_excel:
            excel_path = compute_excel
            devices_sheet = devices_sheet or compute_sheet
            data_mode = 'excel'

    sheet_names     = []
    sw_table        = None
    devices_headers = []
    devices_rows    = []
    devices_total   = 0
    error           = ''

    if data_mode == 'excel' and excel_path:
        try:
            sheet_names = get_sheet_names(excel_path)
            # Always read devices sheet first (no row limit)
            if devices_sheet:
                devices_headers, devices_rows, devices_total = read_devices_sheet(
                    excel_path, devices_sheet, max_rows=100000
                )
            devices_headers, devices_rows, project_options = _filter_compute_project(devices_headers, devices_rows, project_filter)
            devices_total = len(devices_rows)
            # Try summary sheet only if configured AND different from devices sheet
            if summary_sheet and summary_sheet != devices_sheet:
                try:
                    sw_table = build_deployment_table(excel_path, summary_sheet)
                except Exception as sum_exc:
                    logger.warning(f'[DS] Summary sheet parse failed ({sum_exc}), using devices aggregation')
                    sw_table = None
            # Build from devices sheet if no summary or it returned nothing
            if not sw_table and devices_sheet:
                try:
                    sw_table = build_deployment_table_from_devices(
                        excel_path, devices_sheet,
                        prefiltered_headers=devices_headers,
                        prefiltered_rows=devices_rows
                    )
                    logger.info(f'[DS] Built deployment table from devices sheet: {devices_total} rows, sites={(sw_table or {}).get("sites")}')
                except Exception as agg_exc:
                    logger.warning(f'[DS] Devices aggregation failed: {agg_exc}')
            elif sw_table and project_filter and str(project_filter).lower() != 'all':
                # Re-aggregate using only the filtered device rows
                try:
                    sw_table = build_deployment_table_from_devices(
                        excel_path, devices_sheet,
                        prefiltered_headers=devices_headers,
                        prefiltered_rows=devices_rows
                    ) if devices_sheet else sw_table
                except Exception as agg_exc:
                    logger.warning(f'[DS] Filtered aggregation failed: {agg_exc}')
        except Exception as exc:
            error = str(exc)
            logger.warning(f'[DS] Excel load error for {target_name}: {exc}')
    elif data_mode == 'static':
        static = load_static_data(target_name)
        sw_table        = static.get('sw_table')
        devices_headers = static.get('devices_headers', [])
        devices_rows    = static.get('devices_rows', [])
        devices_total   = len(devices_rows)

    return {
        'excel_path':      excel_path,
        'summary_sheet':   summary_sheet,
        'devices_sheet':   devices_sheet,
        'data_mode':       data_mode,
        'sheet_names':     sheet_names,
        'sw_table':        sw_table,
        'devices_headers': devices_headers,
        'devices_rows':    devices_rows,
        'devices_total':   devices_total,
        'page_error':      error,
        'mandatory_cols':  DEVICE_MANDATORY_COLS,
        'project_filter':  project_filter or 'All',
        'project_options': locals().get('project_options', []),
    }


# ── legacy Axiom helpers (kept for backward compat) ───────────────────────────
def get_device_summary(target_name):
    if AXIOM_ENABLED_CHIPS:
        try:
            devices = get_devices_by_chipset(target_name)
        except Exception as e:
            logger.info(f'Axiom call failed: {e}')
            devices = []
    else:
        devices = []
    return devices


def filter_devices_by_taxonomy(devices, taxonomy_path=None, exclude=False):
    if not taxonomy_path:
        return devices
    if exclude:
        return [d for d in devices if d.get('taxonomy_path') != taxonomy_path]
        return [d for d in devices if d.get('taxonomy_path') == taxonomy_path]


def update_hwpdt_dep_with_mcn(hw_devices):
    taxonomy_devices = filter_devices_by_taxonomy(hw_devices, '/PDT/QIPL/HW')
    for device in taxonomy_devices:
        if device.get('mcn'):
            device['dep'] = device['mcn']
    return hw_devices