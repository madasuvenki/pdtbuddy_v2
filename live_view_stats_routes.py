import json
import os
import re
from datetime import date, datetime
from glob import glob
from typing import Any, Dict, List, Optional

from flask import Blueprint, jsonify, redirect, render_template, request, url_for
from flask_login import login_required

from dashboard_common import get_bu_for_target, get_display_name_for_target, get_mysql_connection_db, get_schema_for_target

live_view_stats_bp = Blueprint("live_view_stats_bp", __name__)

_DATA_ROOT = os.environ.get("PDTBUDDY_DATA_ROOT", r"\\Sphere\pdtqipl_internal\PDTBuddy")
_LOCAL_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
_DEFAULT_EXCEL_ROOT = os.environ.get("LIVE_VIEW_STATS_EXCEL_ROOT", r"C:\Dropbox")
_STATS_HEADERS = ["S.No", "Date", "Meta-ID", "Hours", "System Crashes", "SSR Crashes", "Process Crashes", "Total Crashes", "MTBF"]


def _is_auto_target(target_name: str) -> bool:
    target = str(target_name or "").strip().upper()
    bu = str(get_bu_for_target(target_name) or "").strip().upper()
    return bu in {"AUTO", "AUTOMOTIVE"} or target.startswith("NORD") or "NORD_" in target or "NORD." in target


def _target_slug(target_name: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(target_name or "").strip()).strip("_") or "target"


def _sheet_slug(sheet_name: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(sheet_name or "").strip()).strip("_")
    return slug[:80] or "sheet"


def _stats_folder(target_name: str, *, create: bool = True) -> str:
    folder = os.path.join(_DATA_ROOT, "managed_excel", "AUTO", "LIVE_VIEW_STATS", _target_slug(target_name))
    if create:
        try:
            os.makedirs(folder, exist_ok=True)
            probe = os.path.join(folder, ".write_probe")
            with open(probe, "w", encoding="utf-8") as fh:
                fh.write("ok")
            os.remove(probe)
            return folder
        except Exception:
            pass
    if os.path.isdir(folder):
        return folder
    local = os.path.join(_LOCAL_ROOT, "live_view_stats", _target_slug(target_name))
    if create:
        os.makedirs(local, exist_ok=True)
    return local


def _config_path(target_name: str) -> str:
    return os.path.join(_stats_folder(target_name), "config.json")


def _index_path(target_name: str) -> str:
    return os.path.join(_stats_folder(target_name), "sheets_index.json")


def _sheet_json_path(target_name: str, sheet_name: str) -> str:
    return os.path.join(_stats_folder(target_name), f"sheet_{_sheet_slug(sheet_name)}.json")


def _json_default(value: Any) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def _atomic_write_json(path: str, payload: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False, default=_json_default)
    os.replace(tmp, path)


def _read_json(path: str, default: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, dict) else (default or {})
    except Exception:
        return default or {}


def _load_config(target_name: str) -> Dict[str, Any]:
    cfg = _read_json(_config_path(target_name), {})
    cfg.setdefault("target", target_name)
    cfg.setdefault("excel_path", "")
    cfg.setdefault("excel_root", _DEFAULT_EXCEL_ROOT)
    cfg.setdefault("sheet_tables", {})
    cfg.setdefault("sp_names", [])
    cfg.setdefault("updated_at", "")
    return cfg


def _save_config(target_name: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    cfg = _load_config(target_name)
    if "excel_path" in payload:
        cfg["excel_path"] = str(payload.get("excel_path") or "").strip()
    if "excel_root" in payload:
        cfg["excel_root"] = str(payload.get("excel_root") or "").strip() or _DEFAULT_EXCEL_ROOT
    if isinstance(payload.get("sheet_tables"), dict):
        cleaned = {}
        for sheet, row in payload.get("sheet_tables", {}).items():
            if not str(sheet or "").strip():
                continue
            row = row if isinstance(row, dict) else {}
            cleaned[str(sheet)] = {
                "target_table": str(row.get("target_table") or "").strip(),
                "jiras_table": str(row.get("jiras_table") or row.get("target_table") or "").strip(),
                "openjiras_table": str(row.get("openjiras_table") or "").strip(),
                "unique_crs_table": str(row.get("unique_crs_table") or "").strip(),
            }
        cfg["sheet_tables"] = cleaned
    # sp_names: explicit ordered list of SP names for Nord HQX/HGY-style pages
    if isinstance(payload.get("sp_names"), list):
        cfg["sp_names"] = [str(s).strip() for s in payload["sp_names"] if str(s).strip()]
    cfg["target"] = target_name
    cfg["updated_at"] = datetime.utcnow().isoformat() + "Z"
    _atomic_write_json(_config_path(target_name), cfg)
    return cfg


def _is_admin_user() -> bool:
    """Return True if the current user is an admin or TARGET_GROUP editor."""
    try:
        from flask_login import current_user as _cu
        from config import ADMIN_USERS, TARGET_GROUP
        uid = str(getattr(_cu, "id", "") or "").strip().lower()
        if uid in ADMIN_USERS:
            return True
        try:
            import app as _app
            return bool(_app.is_user_in_group(uid, TARGET_GROUP))
        except Exception:
            return False
    except Exception:
        return False


def _get_sp_names(target_name: str) -> List[str]:
    """Return the ordered SP name list for a target (Nord HQX/HGY-style pages)."""
    cfg = _load_config(target_name)
    return list(cfg.get("sp_names") or [])


def _add_sp_name(target_name: str, sp_name: str) -> List[str]:
    """Append a new SP name to the target's sp_names list. No-op if already present."""
    sp_name = str(sp_name or "").strip()
    if not sp_name:
        raise ValueError("SP name cannot be empty.")
    if len(sp_name) > 80:
        raise ValueError("SP name must be 80 characters or fewer.")
    cfg = _load_config(target_name)
    names: List[str] = list(cfg.get("sp_names") or [])
    if sp_name in names:
        return names  # already exists
    names.append(sp_name)
    cfg["sp_names"] = names
    cfg["updated_at"] = datetime.utcnow().isoformat() + "Z"
    _atomic_write_json(_config_path(target_name), cfg)
    return names


def _remove_sp_name(target_name: str, sp_name: str) -> List[str]:
    """Remove an SP name from the target's sp_names list."""
    sp_name = str(sp_name or "").strip()
    if not sp_name:
        raise ValueError("SP name cannot be empty.")
    cfg = _load_config(target_name)
    names: List[str] = list(cfg.get("sp_names") or [])
    if sp_name not in names:
        raise ValueError(f"SP '{sp_name}' not found.")
    names = [n for n in names if n != sp_name]
    cfg["sp_names"] = names
    cfg["updated_at"] = datetime.utcnow().isoformat() + "Z"
    _atomic_write_json(_config_path(target_name), cfg)
    return names


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("_", " ").replace("-", " ")
    text = re.sub(r"[^a-z0-9 ]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _find_col(headers: List[str], candidates: List[str]) -> int:
    normalized = [_norm_header(h) for h in headers]
    candidate_norm = [_norm_header(c) for c in candidates]
    for cand in candidate_norm:
        if cand in normalized:
            return normalized.index(cand)
    for i, header in enumerate(normalized):
        if any(cand and cand in header for cand in candidate_norm):
            return i
    return -1


def _num(value: Any, integer: bool = False) -> Any:
    if value in (None, ""):
        return ""
    try:
        n = float(str(value).replace(",", "").strip())
        return int(n) if integer else round(n, 2)
    except Exception:
        return ""


def _serial_cell(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return "" if value is None else str(value).strip()


def _sheet_to_payload(target_name: str, excel_path: str, sheet_name: str) -> Dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    actual_sheet = sheet_name if sheet_name in wb.sheetnames else next((s for s in wb.sheetnames if s.lower() == sheet_name.lower()), "")
    if not actual_sheet:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = wb[actual_sheet]

    merge_map: Dict[tuple, Any] = {}
    for mr in list(ws.merged_cells.ranges):
        val = ws.cell(mr.min_row, mr.min_col).value
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merge_map[(row, col)] = val

    def cv(row: int, col: int) -> str:
        return _serial_cell(merge_map.get((row, col), ws.cell(row, col).value))

    header_tokens = {
        "date", "week", "meta id", "meta", "build", "build id", "hours", "total hours",
        "system crashes", "ssr crashes", "process crashes", "total crashes", "crashes", "mtbf",
    }
    best_header_row, best_score = 1, -1
    for rr in range(1, min(ws.max_row or 1, 30) + 1):
        values = [cv(rr, c) for c in range(1, (ws.max_column or 1) + 1)]
        score = len({_norm_header(v) for v in values if v} & header_tokens)
        if score > best_score:
            best_header_row, best_score = rr, score

    headers = [cv(best_header_row, c) or f"Column {c}" for c in range(1, (ws.max_column or 1) + 1)]
    rows = []
    for rr in range(best_header_row + 1, (ws.max_row or 1) + 1):
        values = [cv(rr, c) for c in range(1, (ws.max_column or 1) + 1)]
        if not any(str(v).strip() for v in values):
            continue
        row_map = {headers[i]: values[i] for i in range(len(headers))}
        rows.append({"excel_row": rr, "values": values, "row": row_map})

    idx_date = _find_col(headers, ["date", "week", "build date"])
    idx_meta = _find_col(headers, ["meta id", "meta-id", "meta", "build", "build id", "builds full id"])
    idx_hours = _find_col(headers, ["hours", "total hours", "tested hours"])
    idx_system = _find_col(headers, ["system crashes", "system", "sys crashes"])
    idx_ssr = _find_col(headers, ["ssr crashes", "ssr"])
    idx_process = _find_col(headers, ["process crashes", "process"])
    idx_total = _find_col(headers, ["total crashes", "crashes", "crash count"])
    idx_mtbf = _find_col(headers, ["mtbf", "product mtbf", "qc mtbf"])

    chart_rows = []
    for i, row in enumerate(rows, start=1):
        values = row.get("values") or []
        def at(idx: int) -> str:
            return values[idx] if 0 <= idx < len(values) else ""
        meta_id = at(idx_meta).strip() if idx_meta >= 0 else ""
        if not meta_id:
            continue
        system_c = _num(at(idx_system), integer=True) if idx_system >= 0 else ""
        ssr_c = _num(at(idx_ssr), integer=True) if idx_ssr >= 0 else ""
        process_c = _num(at(idx_process), integer=True) if idx_process >= 0 else ""
        total_c = _num(at(idx_total), integer=True) if idx_total >= 0 else ""
        if total_c == "":
            total_c = sum(int(x or 0) for x in (system_c, ssr_c, process_c) if x != "")
        hours = _num(at(idx_hours)) if idx_hours >= 0 else ""
        mtbf = _num(at(idx_mtbf)) if idx_mtbf >= 0 else ""
        if mtbf == "" and hours not in ("", 0) and total_c not in ("", 0):
            mtbf = round(float(hours) / int(total_c), 2)
        chart_rows.append({
            "id": f"{_sheet_slug(actual_sheet)}_{i}",
            "s_no": i,
            "date": at(idx_date)[:10] if idx_date >= 0 else "",
            "meta_id": meta_id,
            "hours": hours,
            "system_crashes": system_c,
            "ssr_crashes": ssr_c,
            "process_crashes": process_c,
            "total_crashes": total_c,
            "mtbf": mtbf,
        })

    return {
        "target": target_name,
        "sheet_name": actual_sheet,
        "excel_path": excel_path,
        "headers": headers,
        "rows": rows,
        "chart_rows": chart_rows,
        "mtbf_headers": list(_STATS_HEADERS),
        "header_row": best_header_row,
        "detected_header_score": best_score,
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }


def _workbook_sheets(path: str) -> List[str]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    return list(wb.sheetnames)


def _db_table_options(target_name: str) -> List[Dict[str, str]]:
    schema = str(get_schema_for_target(target_name) or "").strip("`")
    if not schema:
        return []
    conn = get_mysql_connection_db(bu_key=schema)
    if not conn:
        conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    cur = None
    try:
        cur = conn.cursor(dictionary=True)
        target_parts = [p for p in re.split(r"[^a-z0-9]+", str(target_name or "").lower()) if p]
        like_terms = ["%_jiras", "%_openjiras", "%_unique_crs", "%overall%crs%"]
        params: List[str] = [schema]
        table_filters = ["TABLE_NAME LIKE %s" for _ in like_terms]
        params.extend(like_terms)
        if target_parts:
            table_filters.extend(["TABLE_NAME LIKE %s" for _ in target_parts[:4]])
            params.extend([f"%{part}%" for part in target_parts[:4]])
        cur.execute(
            f"""
            SELECT TABLE_NAME
            FROM information_schema.TABLES
            WHERE TABLE_SCHEMA = %s
              AND ({' OR '.join(table_filters)})
            ORDER BY
              CASE
                WHEN TABLE_NAME LIKE '%%_jiras' THEN 0
                WHEN TABLE_NAME LIKE '%%_openjiras' THEN 1
                WHEN TABLE_NAME LIKE '%%_unique_crs' THEN 2
                WHEN TABLE_NAME LIKE '%%overall%%crs%%' THEN 3
                ELSE 9
              END,
              TABLE_NAME
            LIMIT 500
            """,
            tuple(params),
        )
        rows = cur.fetchall() or []
        options = []
        for row in rows:
            name = str(row.get("TABLE_NAME") or row.get("table_name") or "").strip()
            if not name:
                continue
            fq = f"`{schema}`.`{name}`"
            lower = name.lower()
            if lower.endswith("_openjiras"):
                kind = "openjiras"
            elif lower.endswith("_unique_crs"):
                kind = "unique_crs"
            elif "overall" in lower and "cr" in lower:
                kind = "overallcrs"
            elif lower.endswith("_jiras"):
                kind = "jiras"
            else:
                kind = "other"
            options.append({"name": name, "fq": fq, "kind": kind, "label": f"{name} ({kind})"})
        return options
    except Exception:
        return []
    finally:
        if cur:
            cur.close()
        try:
            conn.close()
        except Exception:
            pass


def _route_live_view_stats(target_name: str):
    """Central dispatcher: redirect to the correct live-view-stats page for a target.

    Priority order:
      1. Auto Gen4.5 (canonical target = auto_gen4.5, or BU AUTO/AUTOMOTIVE with 4.8)
         → /automotive/live_view_stats/<target>
      2. WBC (BU = WBC)
         → /wbc/live_view_status
      3. AUTO / Automotive (Nord HQX, HGY, SECA, etc.)
         → render live_view_stats.html  (existing behaviour)
      4. Everything else (XR, Mobile, IoT, MBB, Compute, …)
         → /others/live_view_stats/<target>
    """
    # 1. Auto Gen4.5
    try:
        from automotive_live_view_stats_routes import _is_auto_gen45_target
        if _is_auto_gen45_target(target_name):
            return redirect(
                url_for(
                    "automotive_live_view_stats_bp.automotive_live_view_stats_page",
                    target_name=target_name,
                )
            )
    except Exception:
        pass

    # 2. WBC
    bu = str(get_bu_for_target(target_name) or "").strip().upper()
    if bu == "WBC":
        return redirect(url_for("wbc_live_view_stats_bp.wbc_live_view_status_page"))

    # 3. AUTO / Automotive (Nord HQX, HGY, SECA, …)
    is_auto = _is_auto_target(target_name)
    if is_auto:
        slug = str(target_name or "").strip().upper().replace(".", "_")
        is_nord_sp_managed = (
            slug in {"NORD_HQX", "NORD_HGY"}
            or "NORD_HQX" in slug
            or "NORD_HGY" in slug
        )
        return render_template(
            "live_view_stats.html",
            target_name=target_name,
            target_display=get_display_name_for_target(target_name) or target_name,
            default_excel_root=_DEFAULT_EXCEL_ROOT,
            is_admin=_is_admin_user(),
            is_nord_sp_managed=is_nord_sp_managed,
            is_auto=is_auto,
        )

    # 4. Others (XR, Mobile, IoT, MBB, Compute, …)
    return redirect(
        url_for(
            "others_live_view_stats_bp.others_live_view_stats_page",
            target_name=target_name,
        )
    )


@live_view_stats_bp.route("/live_view_stats/<string:target_name>", methods=["GET"])
@login_required
def live_view_stats_page(target_name: str):
    """Smart dispatcher: routes to the correct live-view-stats page based on BU."""
    return _route_live_view_stats(target_name)


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/config", methods=["GET", "POST"])
@login_required
def api_live_view_stats_config(target_name: str):
    if request.method == "POST":
        cfg = _save_config(target_name, request.get_json(force=True, silent=True) or {})
    else:
        cfg = _load_config(target_name)
    sheets = []
    if cfg.get("excel_path") and os.path.exists(cfg.get("excel_path")):
        try:
            sheets = _workbook_sheets(cfg["excel_path"])
        except Exception:
            sheets = []
    index = _read_json(_index_path(target_name), {"sheets": []})
    return jsonify({"ok": True, "config": cfg, "workbook_sheets": sheets, "index": index})


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/db_tables", methods=["GET"])
@login_required
def api_live_view_stats_db_tables(target_name: str):
    tables = _db_table_options(target_name)
    return jsonify({"ok": True, "target": target_name, "tables": tables})


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/workbooks", methods=["GET"])
@login_required
def api_live_view_stats_workbooks(target_name: str):
    root = str(request.args.get("root") or _load_config(target_name).get("excel_root") or _DEFAULT_EXCEL_ROOT).strip()
    patterns = [os.path.join(root, "*.xlsx"), os.path.join(root, "*.xlsm")]
    files: List[str] = []
    for pattern in patterns:
        files.extend(glob(pattern))
    if str(request.args.get("recursive") or "").lower() in {"1", "true", "yes"}:
        for pattern in (os.path.join(root, "**", "*.xlsx"), os.path.join(root, "**", "*.xlsm")):
            files.extend(glob(pattern, recursive=True))
    files = sorted(dict.fromkeys(files), key=lambda p: os.path.getmtime(p) if os.path.exists(p) else 0, reverse=True)[:200]
    return jsonify({"ok": True, "root": root, "files": files})


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/sync", methods=["POST"])
@login_required
def api_live_view_stats_sync(target_name: str):
    if not _is_auto_target(target_name):
        return jsonify({"ok": False, "error": "Live View Stats is AUTO-only."}), 404
    payload = request.get_json(force=True, silent=True) or {}
    cfg = _save_config(target_name, payload) if any(k in payload for k in ("excel_path", "excel_root", "sheet_tables")) else _load_config(target_name)
    excel_path = str(payload.get("excel_path") or cfg.get("excel_path") or "").strip()
    if not excel_path or not os.path.exists(excel_path):
        return jsonify({"ok": False, "error": f"Excel file not found: {excel_path or '(blank)'}"}), 400
    sheets = payload.get("sheets") or []
    if not sheets:
        sheets = _workbook_sheets(excel_path)
    synced = []
    errors = []
    for sheet in sheets:
        try:
            data = _sheet_to_payload(target_name, excel_path, str(sheet))
            data["db_config"] = (cfg.get("sheet_tables") or {}).get(data["sheet_name"], {})
            _atomic_write_json(_sheet_json_path(target_name, data["sheet_name"]), data)
            synced.append({
                "sheet_name": data["sheet_name"],
                "rows": len(data.get("rows") or []),
                "chart_rows": len(data.get("chart_rows") or []),
                "updated_at": data.get("updated_at"),
                "json_file": os.path.basename(_sheet_json_path(target_name, data["sheet_name"])),
            })
        except Exception as exc:
            errors.append({"sheet_name": str(sheet), "error": str(exc)})
    index = {"target": target_name, "excel_path": excel_path, "sheets": synced, "errors": errors, "updated_at": datetime.utcnow().isoformat() + "Z"}
    _atomic_write_json(_index_path(target_name), index)
    return jsonify({"ok": not errors or bool(synced), "synced": synced, "errors": errors, "index": index})


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/sheets", methods=["GET"])
@login_required
def api_live_view_stats_sheets(target_name: str):
    cfg = _load_config(target_name)
    index = _read_json(_index_path(target_name), {"target": target_name, "sheets": []})
    return jsonify({"ok": True, "config": cfg, "index": index})


# ---------------------------------------------------------------------------
# SP name management endpoints  (Nord HQX/HGY flexible SP list)
# ---------------------------------------------------------------------------

@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/sp_names", methods=["GET"])
@login_required
def api_live_view_stats_sp_names_get(target_name: str):
    """Return the current SP name list for a target."""
    try:
        names = _get_sp_names(target_name)
        return jsonify({"ok": True, "target": target_name, "sp_names": names})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/sp_names/add", methods=["POST"])
@login_required
def api_live_view_stats_sp_names_add(target_name: str):
    """Admin-only: add a new SP name to the target's list."""
    if not _is_admin_user():
        return jsonify({"ok": False, "error": "Admin access required."}), 403
    payload = request.get_json(force=True, silent=True) or {}
    sp_name = str(payload.get("sp_name") or "").strip()
    try:
        names = _add_sp_name(target_name, sp_name)
        return jsonify({"ok": True, "target": target_name, "sp_names": names,
                        "message": f"SP '{sp_name}' added."})
    except ValueError as ve:
        return jsonify({"ok": False, "error": str(ve)}), 400
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/sp_names/remove", methods=["POST"])
@login_required
def api_live_view_stats_sp_names_remove(target_name: str):
    """Admin-only: remove an SP name from the target's list."""
    if not _is_admin_user():
        return jsonify({"ok": False, "error": "Admin access required."}), 403
    payload = request.get_json(force=True, silent=True) or {}
    sp_name = str(payload.get("sp_name") or "").strip()
    try:
        names = _remove_sp_name(target_name, sp_name)
        return jsonify({"ok": True, "target": target_name, "sp_names": names,
                        "message": f"SP '{sp_name}' removed."})
    except ValueError as ve:
        return jsonify({"ok": False, "error": str(ve)}), 400
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/sp_names/reorder", methods=["POST"])
@login_required
def api_live_view_stats_sp_names_reorder(target_name: str):
    """Admin-only: replace the full ordered SP name list."""
    if not _is_admin_user():
        return jsonify({"ok": False, "error": "Admin access required."}), 403
    payload = request.get_json(force=True, silent=True) or {}
    sp_names = payload.get("sp_names")
    if not isinstance(sp_names, list):
        return jsonify({"ok": False, "error": "sp_names must be a list."}), 400
    try:
        cfg = _save_config(target_name, {"sp_names": sp_names})
        return jsonify({"ok": True, "target": target_name, "sp_names": cfg.get("sp_names") or [],
                        "message": "SP list reordered."})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/sheet/<path:sheet_name>", methods=["GET"])
@login_required
def api_live_view_stats_sheet(target_name: str, sheet_name: str):
    path = _sheet_json_path(target_name, sheet_name)
    if not os.path.exists(path):
        cfg = _load_config(target_name)
        excel_path = str(cfg.get("excel_path") or "").strip()
        if excel_path and os.path.exists(excel_path):
            try:
                data = _sheet_to_payload(target_name, excel_path, sheet_name)
                data["db_config"] = (cfg.get("sheet_tables") or {}).get(data["sheet_name"], {})
                _atomic_write_json(_sheet_json_path(target_name, data["sheet_name"]), data)
                return jsonify({"ok": True, "sheet": data})
            except Exception as exc:
                return jsonify({"ok": False, "error": str(exc)}), 500
        return jsonify({"ok": False, "error": "Sheet JSON not found. Configure Excel and click Sync Excel."}), 404
    data = _read_json(path, {})
    cfg = _load_config(target_name)
    data["db_config"] = (cfg.get("sheet_tables") or {}).get(data.get("sheet_name") or sheet_name, {})
    return jsonify({"ok": True, "sheet": data})


# ---------------------------------------------------------------------------
# Generic Saved JQL Tabs API — works for any BU/target
# The centralized scheduler in live_view_saved_jql_service is already running
# and will auto-refresh due jobs for every BU without any extra wiring.
# ---------------------------------------------------------------------------

def _sjql_domain(target_name: str) -> str:
    """Return the saved-JQL domain string for a target (BU-based)."""
    bu = str(get_bu_for_target(target_name) or "").strip().upper() or "GENERAL"
    return bu


def _sjql_resolve_filter(jql: str) -> str:
    """Resolve a JIRA filter ID to actual JQL if the value looks like a filter reference."""
    import re as _re
    text = str(jql or "").strip()
    if not text:
        return text
    # Pure numeric → treat as filter ID
    fid = ""
    if text.isdigit():
        fid = text
    else:
        m = _re.match(r"^\s*filter(?:Id)?\s*=\s*(\d+)\s*(?:ORDER\s+BY\s+.+)?$", text, flags=_re.I)
        if m:
            fid = m.group(1)
        else:
            m2 = _re.search(r"[?&]filter(?:Id)?=(\d+)", text, flags=_re.I)
            if m2:
                fid = m2.group(1)
    if not fid:
        return text
    try:
        from dashboard_routes import _resolve_jira_filter_jql
        resolved = _resolve_jira_filter_jql(fid)
        return str(resolved).strip() if resolved else text
    except Exception:
        return text


def _sjql_extract_filter_ids(jql_text: str) -> List[str]:
    """Extract all filter IDs from a JQL string (e.g. 'filter = 346152' → ['346152'])."""
    import re as _re
    return _re.findall(r'filter\s*=\s*(\d+)', str(jql_text or ""), flags=_re.I)


def _sjql_run_report(target_name: str, domain: str, tab_id: str, force: bool = False):
    """Execute a saved-JQL report for any BU target and cache the result."""
    import sys as _sys
    from datetime import datetime as _dt, timedelta as _td
    from live_view_saved_jql_service import (
        get_cached_report, get_tab, set_cached_report,
    )
    from config import JIRA_PDT_FILTER_ID

    tab = get_tab(target_name, domain, tab_id)
    if not tab:
        return {"ok": False, "error": "Saved JQL tab not found"}

    raw_jql = str(tab.get("jql") or "").strip()
    jql = _sjql_resolve_filter(raw_jql)
    if not jql:
        return {"ok": False, "error": "Saved JQL is empty"}

    # Extract filter IDs from raw JQL for meta display in the UI header
    filter_ids = _sjql_extract_filter_ids(raw_jql) or _sjql_extract_filter_ids(jql)
    meta_from_filter = ", ".join(filter_ids) if filter_ids else ""

    if not force:
        cached = get_cached_report(target_name, domain, tab_id)
        if cached:
            cached = dict(cached)
            cached.update({"ok": True, "from_cache": True, "tab": tab, "jql": jql,
                           "cache_status": "cached"})
            # Inject meta_from_filter if the cached payload doesn't have it yet
            if not cached.get("meta_from_filter") and meta_from_filter:
                cached["meta_from_filter"] = meta_from_filter
                cached["meta_ids"] = filter_ids
            return cached

    now = _dt.utcnow()
    try:
        scripts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scripts")
        if scripts_dir not in _sys.path:
            _sys.path.insert(0, scripts_dir)
        from fetch_consolidated_report import run_consolidated_report
        filter_id = str(tab.get("filter_id") or "").strip()
        raw_report = run_consolidated_report(
            build_ids=[],
            filter_id=filter_id or JIRA_PDT_FILTER_ID,
            traverse=True,
            enrich_orbit=True,
            target_name=str(target_name or "") or None,
            custom_jql=jql,
        )
        rows = []
        for cr_row in (raw_report.get("hierarchical_report") or []):
            cr = cr_row.get("cr") or "NO_CR"
            for jira in (cr_row.get("jiras") or []):
                trav = jira.get("traversal") or {}
                rows.append({
                    "CR": cr,
                    "CR Title": cr_row.get("cr_title") or "",
                    "CR Status": cr_row.get("cr_status") or "",
                    "CR Area": cr_row.get("cr_area") or "",
                    "JIRA": jira.get("key") or "",
                    "JIRA Title": jira.get("title") or jira.get("summary") or "",
                    "JIRA Status": jira.get("status") or "",
                    "Final Ticket": jira.get("final_key") or trav.get("final_key") or "",
                    "Final Status": jira.get("final_status") or trav.get("final_status") or "",
                    "Final Resolution": jira.get("final_resolution") or trav.get("final_resolution") or "",
                    "Created": jira.get("created") or "",
                })
        if not rows:
            rows = raw_report.get("rows") or raw_report.get("flat_rows") or []
        ttl = _td(minutes=30)
        report = {
            "ok": True,
            "tab": tab,
            "target_name": target_name,
            "domain": domain,
            "generated_at": now.isoformat() + "Z",
            "from_cache": False,
            "cache_status": "fresh",
            "source": "Saved JQL consolidated report",
            "jql": jql,
            "raw_jql": raw_jql,
            "filter_id": filter_id,
            "meta_from_filter": meta_from_filter,
            "meta_ids": filter_ids,
            "rows": rows,
            "flat_rows": rows,
            "row_count": len(rows),
            "cr_count": len({str(r.get("CR") or "").strip() for r in rows if str(r.get("CR") or "").strip() and str(r.get("CR") or "").strip() != "NO_CR"}),
            "jira_count": len({str(r.get("JIRA") or "").strip() for r in rows if str(r.get("JIRA") or "").strip()}),
            "next_run_at": (now + ttl).isoformat() + "Z",
            "next_auto_refresh_at": (now + ttl).isoformat() + "Z",
            "summary": raw_report.get("summary") or {},
        }
        stored = set_cached_report(target_name, domain, tab_id, report)
        # Use the registry-computed next_run_at (respects per-job refresh_minutes)
        report["generated_at"] = stored.get("generated_at") or report["generated_at"]
        report["next_run_at"] = stored.get("next_run_at") or report["next_run_at"]
        report["next_auto_refresh_at"] = stored.get("next_auto_refresh_at") or report["next_auto_refresh_at"]
        return report
    except Exception as exc:
        return {
            "ok": False,
            "tab": tab,
            "generated_at": now.isoformat() + "Z",
            "source": "Saved JQL consolidated report",
            "jql": jql,
            "raw_jql": raw_jql,
            "meta_from_filter": meta_from_filter,
            "meta_ids": filter_ids,
            "run_error": str(exc),
            "rows": [],
            "flat_rows": [],
        }


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/saved_jql_tabs", methods=["GET"])
@login_required
def api_lvs_saved_jql_tabs_list(target_name: str):
    from live_view_saved_jql_service import get_cached_report_raw, list_tabs
    domain = _sjql_domain(target_name)
    tabs = []
    for tab in list_tabs(target_name, domain):
        row = dict(tab)
        cached = get_cached_report_raw(target_name, domain, row.get("id")) or {}
        from datetime import datetime as _dt, timezone as _tz, timedelta as _td
        gen_at = cached.get("generated_at") or ""
        next_at = cached.get("next_run_at") or cached.get("next_auto_refresh_at") or ""
        rows = cached.get("rows") or cached.get("flat_rows") or []
        row.update({
            "has_cached_report": bool(cached),
            "cached_report_stale": bool(gen_at and next_at and _dt.fromisoformat(next_at.replace("Z", "+00:00")) <= _dt.now(_tz.utc)),
            "last_run_at": gen_at,
            "next_run_at": next_at,
            "cached_row_count": cached.get("row_count", len(rows)),
            "cached_cr_count": cached.get("cr_count", 0),
            "cached_jira_count": cached.get("jira_count", 0),
        })
        tabs.append(row)
    return jsonify({"ok": True, "target": target_name, "domain": domain, "tabs": tabs})


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/saved_jql_tabs", methods=["POST"])
@login_required
def api_lvs_saved_jql_tabs_save(target_name: str):
    if not _is_admin_user():
        return jsonify({"ok": False, "error": "Access denied"}), 403
    from flask_login import current_user as _cu
    from live_view_saved_jql_service import list_tabs, save_tab
    payload = request.get_json(force=True, silent=True) or {}
    domain = _sjql_domain(target_name)
    username = str(getattr(_cu, "id", "") or getattr(_cu, "username", "") or "unknown")
    try:
        tab = save_tab(
            target_name, domain,
            tab_id=str(payload.get("id") or "").strip() or None,
            name=str(payload.get("name") or "").strip(),
            jql=str(payload.get("jql") or "").strip(),
            username=username,
        )
        return jsonify({"ok": True, "tab": tab, "tabs": list_tabs(target_name, domain)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 400


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/saved_jql_tabs/<tab_id>", methods=["DELETE"])
@login_required
def api_lvs_saved_jql_tabs_delete(target_name: str, tab_id: str):
    if not _is_admin_user():
        return jsonify({"ok": False, "error": "Access denied"}), 403
    from live_view_saved_jql_service import delete_tab, list_tabs
    domain = _sjql_domain(target_name)
    deleted = delete_tab(target_name, domain, tab_id)
    return jsonify({"ok": True, "deleted": bool(deleted), "tabs": list_tabs(target_name, domain)})


@live_view_stats_bp.route("/api/live_view_stats/<string:target_name>/saved_jql_tabs/<tab_id>/report", methods=["GET", "POST"])
@login_required
def api_lvs_saved_jql_tab_report(target_name: str, tab_id: str):
    force = str(request.args.get("force") or "").lower() in ("1", "true", "yes")
    domain = _sjql_domain(target_name)
    result = _sjql_run_report(target_name, domain, tab_id, force=force)
    # Ensure cache_status is always present for UI display
    if "cache_status" not in result:
        result["cache_status"] = "cached" if result.get("from_cache") else "fresh"
    status = 200 if result.get("ok") or result.get("run_error") else 404
    return jsonify(result), status


# ---------------------------------------------------------------------------
# Non-AUTO BU live view  (Bonsai, MOBILE, COMPUTE, IOT, MBB, etc.)
# Uses the same common dashboard JSON API as /dashboard/<target>/mtbf-excel
# so that read, edit, and update all share one canonical data path.
# ---------------------------------------------------------------------------

@live_view_stats_bp.route("/live_view_stats/nonau/<string:target_name>", methods=["GET"])
@login_required
def nonau_live_view_stats_page(target_name: str) -> "flask.Response":
    """Live view stats page for non-automotive BUs.

    Reads MTBF data via ``GET /api/dashboard/<target>/excel/full_table`` —
    the same endpoint used by the dashboard MTBF edit page — so the live view
    and the dashboard always show the same JSON-backed data.  Editors can add
    builds (``POST /api/dashboard/<target>/excel/add_build``) and save the
    full table (``POST /api/dashboard/<target>/excel/save_table``) without
    any separate storage layer.
    """
    bu = (get_bu_for_target(target_name) or "").upper()
    return render_template(
        "nonau_live_view_stats.html",
        target_name=target_name,
        target_display=get_display_name_for_target(target_name) or target_name,
        bu=bu,
        is_admin=_is_admin_user(),
    )
