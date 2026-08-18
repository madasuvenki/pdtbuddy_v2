import logging
logger = logging.getLogger(__name__)
import os
import re
import sys
import json
import traceback
from datetime import datetime as _dt

from flask_login import login_required
from flask import abort, Blueprint, current_app, render_template, request, url_for, jsonify
from src.axiom_client import get_devices_by_chipset, get_devices_by_taxonomy_path
from dashboard_common import get_chip_name_for_target, is_axiom_enabled_for_target
from qdt_client import get_rework_info_from_qdt
import device_summary_service as ds_svc
from src.utils import get_mysql_connection_db

from pathlib import Path

device_summary_api_bp = Blueprint("device_summary_api_bp", __name__)

# ------ pending device rows waiting for Excel unlock ------
_pending_device_rows = {}  # target_name -> list of {row}
_inventory_summary_cache = {}  # (target, pdt, sp_names) -> {ts, chip_name, raw_devices, saved_at, active_map}
_INVENTORY_SUMMARY_CACHE_TTL_SECONDS = 120


def _get_excel_lock_info(path):
    """Return (is_locked, locked_by). Detects ~$filename.xlsx owner file."""
    owner = os.path.join(os.path.dirname(path), '~$' + os.path.basename(path))
    if os.path.exists(owner):
        try:
            with open(owner, 'rb') as f:
                data = f.read()
            raw  = data[8:]
            name = raw.split(b'\x00\x00')[0].decode('utf-16-le', errors='ignore').strip()
            return True, name or 'another user'
        except Exception:
            return True, 'another user'
    return False, None


# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# NEW: Excel-driven Device Summary APIs
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


@device_summary_api_bp.route('/api/ds/<string:target_name>/upload', methods=['POST'])
@login_required
def api_ds_upload_excel(target_name):
    """Upload a Device Summary Excel - stores managed copy, hides path from UI."""
    try:
        from werkzeug.utils import secure_filename
        upload = request.files.get('file')
        if not upload or not upload.filename:
            return jsonify({'success': False, 'message': 'Please choose an Excel file to upload.'}), 400
        original_name = secure_filename(upload.filename) or 'devices.xlsx'
        ext = os.path.splitext(original_name)[1].lower()
        if ext not in ('.xlsx', '.xlsm'):
            return jsonify({'success': False, 'message': 'Please upload an .xlsx or .xlsm file.'}), 400

        dirs = ds_svc.get_managed_target_dirs(target_name)
        upload_dir = dirs.get('devices') or dirs.get('base')
        os.makedirs(upload_dir, exist_ok=True)
        timestamp = _dt.utcnow().strftime('%Y%m%d_%H%M%S')
        stored_name = f'{target_name}_{timestamp}_{original_name}'
        stored_path = os.path.join(upload_dir, stored_name)
        upload.save(stored_path)

        sheet_names = ds_svc.get_sheet_names(stored_path)
        devices_sheet = next((s for s in sheet_names if 'device' in s.lower()), sheet_names[0] if sheet_names else '')

        ds_svc.save_ds_excel_config(
            target_name,
            excel_path=stored_path,
            summary_sheet=devices_sheet,
            devices_sheet=devices_sheet,
            data_mode='excel',
        )
        return jsonify({
            'success': True,
            'message': f'Uploaded {original_name}.',
            'sheet_names': sheet_names,
            'devices_sheet': devices_sheet,
            'original_filename': original_name,
        })
    except Exception as e:
        logger.exception('DS upload failed for %s', target_name)
        return jsonify({'success': False, 'message': str(e)}), 500

@device_summary_api_bp.route('/api/ds/<string:target_name>/sheets', methods=['POST'])
@login_required
def api_ds_get_sheets(target_name):
    """Return sheet names from an Excel file."""
    try:
        payload    = request.get_json(force=True) or {}
        excel_path = payload.get('excel_path', '')
        sheets     = ds_svc.get_sheet_names(excel_path)
        return jsonify({'success': True, 'sheets': sheets})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400


@device_summary_api_bp.route('/api/ds/<string:target_name>/config/save', methods=['POST'])
@login_required
def api_ds_save_config(target_name):
    """Save Excel config (path + sheets + mode) for a target."""
    try:
        p             = request.get_json(force=True) or {}
        excel_path    = str(p.get('excel_path')    or '').strip()
        summary_sheet = str(p.get('summary_sheet') or '').strip()
        devices_sheet = str(p.get('devices_sheet') or 'Devices').strip()
        data_mode     = str(p.get('data_mode')     or 'excel').strip()
        # Only create a managed workbook when no path is given (Managed Excel button)
        # When a real path is provided, just save the config --- never touch the existing file
        if not excel_path:
            excel_path, summary_sheet, devices_sheet = ds_svc.ensure_device_summary_workbook(
                target_name, '', summary_sheet, devices_sheet
            )
        saved = ds_svc.save_ds_excel_config(
            target_name, excel_path, summary_sheet, devices_sheet, data_mode
        )
        return jsonify({'success': True, 'config': saved})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400


@device_summary_api_bp.route('/api/ds/<string:target_name>/debug', methods=['GET'])
@login_required
def api_ds_debug(target_name):
    """Dump raw Excel parse info --- devices sheet headers + aggregation result."""
    try:
        import openpyxl
        cfg           = ds_svc.get_ds_excel_config(target_name) or {}
        excel_path    = cfg.get('excel_path', '')
        devices_sheet = cfg.get('devices_sheet', '')
        path          = ds_svc._normalize_path(excel_path)

        if not path or not os.path.exists(path):
            return jsonify({'error': f'Excel not found: {path}', 'cfg': cfg}), 404

        # Read raw headers + first 5 rows from devices sheet
        wb = openpyxl.load_workbook(path, data_only=True)
        dev_ws = wb[devices_sheet] if devices_sheet in wb.sheetnames else None
        raw_headers = []
        sample_rows = []
        if dev_ws:
            mm = ds_svc._build_merge_map(dev_ws)
            raw_headers = [ds_svc._cv(dev_ws, mm, 1, c) for c in range(1, dev_ws.max_column + 1)]
            for rn in range(2, min(dev_ws.max_row + 1, 7)):
                sample_rows.append([ds_svc._cv(dev_ws, mm, rn, c) for c in range(1, dev_ws.max_column + 1)])
        wb.close()

        # Try building the deployment table
        sw_table    = None
        parse_error = None
        try:
            sw_table = ds_svc.build_deployment_table_from_devices(excel_path, devices_sheet)
        except Exception as e:
            parse_error = str(e)

        return jsonify({
            'excel_path':    path,
            'devices_sheet': devices_sheet,
            'config':        cfg,
            'raw_headers':   raw_headers,
            'sample_rows':   sample_rows,
            'parse_error':   parse_error,
            'sw_table_sites':     (sw_table or {}).get('sites'),
            'sw_table_row_count': len((sw_table or {}).get('rows', [])),
            'grand_del':          (sw_table or {}).get('grand_del'),
        })
    except Exception as exc:
        return jsonify({'error': str(exc), 'trace': traceback.format_exc()}), 500


@device_summary_api_bp.route('/api/ds/<string:target_name>/refresh', methods=['POST', 'GET'])
@login_required
def api_ds_refresh_from_excel(target_name):
    """Force a fresh parse/recompile of Device Summary dashboard data from Excel."""
    try:
        ds_svc.get_or_create_device_excel_config(target_name)
        data = ds_svc.load_page_data(target_name)
        sw_table = data.get('sw_table') or {}
        return jsonify({
            'success': True,
            'excel_path': data.get('excel_path', ''),
            'summary_sheet': data.get('summary_sheet', ''),
            'devices_sheet': data.get('devices_sheet', ''),
            'deployment_total': sw_table.get('grand_del', 0),
            'deployment_deployed_total': sw_table.get('grand_dep', 0),
            'devices_total': data.get('devices_total', 0),
            'summary_rows': len(sw_table.get('rows') or []),
            'page_error': data.get('page_error', ''),
        })
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400


@device_summary_api_bp.route('/api/ds/<string:target_name>/devices/list', methods=['GET'])
@login_required
def api_ds_devices_list(target_name):
    """Return a fresh devices list from the configured/managed Excel sheet."""
    try:
        ds_svc.get_or_create_device_excel_config(target_name)
        data = ds_svc.load_page_data(target_name)
        return jsonify({
            'success':  True,
            'headers':  data['devices_headers'],
            'rows':     data['devices_rows'],
            'total':    data['devices_total'],
            'mode':     data['data_mode'],
            'excel_path': data['excel_path'],
            'devices_sheet': data['devices_sheet'],
        })
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400


@device_summary_api_bp.route('/api/ds/<string:target_name>/devices/add', methods=['POST'])
@login_required
def api_ds_add_device(target_name):
    """Add a device row and save it to the configured/managed Excel devices sheet."""
    try:
        p   = request.get_json(force=True) or {}
        row = [str(v) for v in (p.get('row') or [])]
        cfg = ds_svc.get_or_create_device_excel_config(target_name)
        path = ds_svc._normalize_path(cfg.get('excel_path', ''))
        locked, locked_by = _get_excel_lock_info(path)
        if locked:
            _pending_device_rows.setdefault(target_name, []).append({'op': 'add', 'row': row})
            return jsonify({'success': True, 'excel': 'locked', 'locked': True, 'locked_by': locked_by})

        saved = ds_svc.append_device_row(target_name, row)
        data = ds_svc.load_page_data(target_name)
        return jsonify({
            'success': True,
            'total': data['devices_total'],
            'row': row,
            'excel': 'saved',
            'locked': False,
            'locked_by': None,
            'headers': data['devices_headers'],
            'rows': data['devices_rows'],
            'excel_path': saved['excel_path'],
        })
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400


@device_summary_api_bp.route('/api/ds/<string:target_name>/devices/retry_sync', methods=['POST'])
@login_required
def api_ds_retry_sync(target_name):
    """Retry writing pending device rows to Excel after lock is released."""
    try:
        pending = _pending_device_rows.get(target_name, [])
        if not pending:
            return jsonify({'success': True, 'message': 'Nothing to sync.'})
        cfg = ds_svc.get_or_create_device_excel_config(target_name)
        path = ds_svc._normalize_path(cfg.get('excel_path', ''))
        locked, locked_by = _get_excel_lock_info(path)
        if locked:
            return jsonify({'success': False, 'locked': True, 'locked_by': locked_by,
                            'message': f'Still locked by {locked_by}.'})
        synced = 0
        for item in pending:
            if item.get('op', 'add') == 'add':
                ds_svc.append_device_row(target_name, item['row'])
                synced += 1
        _pending_device_rows[target_name] = []
        return jsonify({'success': True, 'message': f'Synced {synced} device(s) to Excel.'})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 500


@device_summary_api_bp.route('/api/ds/<string:target_name>/devices/delete', methods=['POST'])
@login_required
def api_ds_delete_device(target_name):
    """Delete a device row by 0-based index from the Excel devices sheet."""
    try:
        p = request.get_json(force=True) or {}
        idx = int(p.get('index', -1))
        cfg = ds_svc.get_or_create_device_excel_config(target_name)
        path = ds_svc._normalize_path(cfg.get('excel_path', ''))
        locked, locked_by = _get_excel_lock_info(path)
        if locked:
            return jsonify({'success': False, 'locked': True, 'locked_by': locked_by,
                            'message': f'Excel is locked by {locked_by}.'}), 423
        deleted = ds_svc.delete_device_row(target_name, idx)
        data = ds_svc.load_page_data(target_name)
        return jsonify({'success': True, 'total': data['devices_total'], 'removed': deleted['removed'],
                        'headers': data['devices_headers'], 'rows': data['devices_rows']})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400


@device_summary_api_bp.route('/api/ds/<string:target_name>/devices/edit', methods=['POST'])
@login_required
def api_ds_edit_device(target_name):
    """Edit a device row by 0-based index in the Excel devices sheet."""
    try:
        p = request.get_json(force=True) or {}
        idx = int(p.get('index', -1))
        row = [str(v) for v in (p.get('row') or [])]
        cfg = ds_svc.get_or_create_device_excel_config(target_name)
        path = ds_svc._normalize_path(cfg.get('excel_path', ''))
        locked, locked_by = _get_excel_lock_info(path)
        if locked:
            return jsonify({'success': False, 'locked': True, 'locked_by': locked_by,
                            'message': f'Excel is locked by {locked_by}.'}), 423
        ds_svc.edit_device_row(target_name, idx, row)
        data = ds_svc.load_page_data(target_name)
        return jsonify({'success': True, 'total': data['devices_total'], 'row': row,
                        'headers': data['devices_headers'], 'rows': data['devices_rows']})
    except Exception as exc:
        return jsonify({'success': False, 'message': str(exc)}), 400


# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# EXISTING Axiom-based APIs (unchanged below)
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def _app_root_dir():
    """Return the directory next to the exe (frozen) or project root (dev)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _axiom_cache_dir():
    root = os.path.join(_app_root_dir(), 'static', 'axiom_cache')
    os.makedirs(root, exist_ok=True)
    return root


def _cache_path_for(chipset, pdt_type):
    chip = (chipset or '').strip().upper() or 'UNKNOWN'
    pdt = (pdt_type or 'SWPDT').strip().upper()
    return os.path.join(_axiom_cache_dir(), "{}_{}.json".format(chip, pdt))


def _unified_cache_path(chipset, pdt_type):
    chip = (chipset or '').strip().upper() or 'UNKNOWN'
    pdt = (pdt_type or 'SWPDT').strip().upper()
    return os.path.join(_axiom_cache_dir(), "{}_{}_unified.json".format(chip, pdt))


def _overrides_dir():
    root = os.path.join(_axiom_cache_dir(), 'overrides')
    os.makedirs(root, exist_ok=True)
    return root


def _sw_del_override_path(chipset):
    return os.path.join(_overrides_dir(), "{}_SWPDT_DEL.json".format((chipset or '').strip().upper()))


def _hw_metrics_override_path(chipset):
    return os.path.join(_overrides_dir(), "{}_HWPDT_metrics.json".format((chipset or '').strip().upper()))


def _device_pool_override_path(chipset, pdt_type):
    chip = (chipset or '').strip().upper()
    pdt  = (pdt_type or 'SWPDT').strip().upper()
    return os.path.join(_overrides_dir(), "{}_{}_pool.json".format(chip, pdt))


def _load_device_pool_overrides(chipset, pdt_type):
    data = _load_json(_device_pool_override_path(chipset, pdt_type))
    if not data:
        return {'removed': [], 'edits': {}}
    return data


def _save_device_pool_overrides(chipset, pdt_type, payload):
    payload['chipset']  = (chipset or '').strip().upper()
    payload['pdt_type'] = (pdt_type or 'SWPDT').strip().upper()
    payload['saved_at'] = _dt.utcnow().isoformat() + 'Z'
    _save_json(_device_pool_override_path(chipset, pdt_type), payload)


def _load_json(path):
    if not os.path.exists(path):
        return None
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            return json.load(fh)
    except Exception:
        return None


def _save_json(path, payload):
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)


def _merge_axiom_qdt_devices(chip_name, pdt_type, devices):
    qdt_rows = []
    qdt_ok = True
    try:
        qdt_rows = get_rework_info_from_qdt(chip_name) or []
    except Exception as _qdt_err:
        qdt_ok = False
        logger.info(f'[DEVICE SUMMARY] QDT fetch failed (non-fatal): {_qdt_err}')
        qdt_rows = []

    qdt_map = {}
    for item in qdt_rows:
        # Match by serial_number OR u_serial_number_qc_mfg
        for key in ('SERIAL_NO', 'SERIAL_NO_QC'):
            serial = str(item.get(key) or '').strip().upper()
            if serial:
                qdt_map[serial] = item
                break

    merged_devices = []
    for dev in devices or []:
        merged = dict(dev)
        serial = str(
            dev.get('serialNumber')
            or dev.get('serial_number')
            or dev.get('serial_no')
            or dev.get('serial')
            or dev.get('SERIAL_NO')
            or ''
        ).strip().upper()
        q_item = qdt_map.get(serial)
        if q_item:
            merged['qdt_rework_info'] = str(q_item.get('REWORK_INFO') or '').strip()
            merged['qdt_location']    = str(q_item.get('LOCATION') or '').strip()
            merged['qdt_asset_tag']   = str(q_item.get('ASSET_TAG') or '').strip()
            merged['qdt_assigned_to'] = str(q_item.get('ASSIGNED_TO') or '').strip()
            merged['qdt_condition']   = str(q_item.get('DEVICE_CONDITION') or '').strip()
            merged['qdt_mes_build']   = str(q_item.get('MES_BUILD') or '').strip()
            merged['qdt_model_desc']  = str(q_item.get('MODEL_DESC') or '').strip()
            # Parse MCN and storage from dv_model string e.g.
            # "DTP, MTP, INT-SECURE SM4850, 6GB LP4X + 128GB UFS2.2, ..."
            model_desc = merged['qdt_model_desc']
            import re as _re
            mcn_m = _re.search(r'(\d{2}-\d{5}-\d{3,4})', model_desc)
            merged['qdt_mcn'] = mcn_m.group(1) if mcn_m else ''
            ufs_m = _re.search(r'(\d+GB\s+UFS[\d\.]*)', model_desc, _re.I)
            lp_m  = _re.search(r'(\d+GB\s+LP[\dA-Z]+)', model_desc, _re.I)
            merged['qdt_storage'] = (ufs_m or lp_m).group(1).strip() if (ufs_m or lp_m) else ''
        merged_devices.append(merged)

    return {
        'chipset': chip_name.upper(),
        'pdt_type': (pdt_type or 'SWPDT').strip().upper(),
        'saved_at': _dt.utcnow().isoformat() + 'Z',
        'count': len(merged_devices),
        'devices': merged_devices,
        'qdt_ok': qdt_ok,
    }


@device_summary_api_bp.route('/api/device_summary_data/<string:target_name>')
@login_required
def api_device_summary_data(target_name):
    pdt_type = (request.args.get('pdt') or 'SWPDT').strip().upper()
    refresh = (request.args.get('refresh') or '0').strip() in ('1', 'true', 'True')

    chip_name = get_chip_name_for_target(target_name) or ''
    if not chip_name:
        return jsonify({'success': False, 'message': "No chip name configured for {}".format(target_name)}), 400

    unified_path = _unified_cache_path(chip_name, pdt_type)
    legacy_path = _cache_path_for(chip_name, pdt_type)
    data = None if refresh else _load_json(unified_path) or _load_json(legacy_path)
    source = 'cache'

    if refresh:
        data = None

    if data is None:
        if not is_axiom_enabled_for_target(target_name):
            return jsonify({'success': False, 'message': "Axiom not enabled for {}".format(chip_name)}), 400
        try:
            devices = get_devices_by_chipset(chip_name, pdt_type=pdt_type, include_site_details=True) or []
            if pdt_type == 'SWPDT':
                devices = [
                    d for d in devices
                    if not str(
                        (d.get('_raw') or {}).get('taxonomyPath')
                        or d.get('taxonomy_path')
                        or ''
                    ).strip().upper().startswith('/PDT/QIPL/HW')
                ]
            data = _merge_axiom_qdt_devices(chip_name, pdt_type, devices)
            _save_json(unified_path, data)
            _save_json(legacy_path, data)
            source = 'axiom' if not data.get('qdt_ok', True) else 'axiom+qdt'
        except OSError as e:
            # Missing credentials --- tell the UI clearly instead of 500
            return jsonify({
                'success': False,
                'message': str(e),
                'hint': 'Add AXIOM_CLIENT_ID and AXIOM_CLIENT_SECRET to the .env file next to BuddyApp.exe and restart.'
            }), 503
        except Exception as e:
            logger.debug(traceback.format_exc())
            return jsonify({'success': False, 'message': str(e)}), 500

    payload = {'success': True, 'source': source, 'cache_file': unified_path, 'legacy_cache_file': legacy_path}
    payload.update(data)
    return jsonify(payload)


@device_summary_api_bp.route('/api/device_summary_data/<string:target_name>/sync_status')
@login_required
def api_sync_status(target_name):
    """Return current cache status for both PDT types --- used by the sync banner to show progress."""
    chip_name = get_chip_name_for_target(target_name) or ''
    result = {}
    for pdt in ('SWPDT', 'HWPDT'):
        path = _unified_cache_path(chip_name, pdt)
        legacy = _cache_path_for(chip_name, pdt)
        data = _load_json(path) or _load_json(legacy)
        if data:
            result[pdt] = {
                'ready': True,
                'count': data.get('count', len(data.get('devices', []))),
                'saved_at': data.get('saved_at', ''),
                'qdt_ok': data.get('qdt_ok', True),
                'source': data.get('source', 'cache'),
            }
        else:
            result[pdt] = {'ready': False, 'count': 0}
    return jsonify({'success': True, 'chip_name': chip_name, 'status': result})



@login_required
def api_device_taxonomy_data(target_name):
    path = request.args.get('path', '').strip()
    if not path:
        return jsonify({'success': False, 'message': "Taxonomy path is required"}), 400
        
    chip_name = get_chip_name_for_target(target_name) or ''
    if not chip_name:
        return jsonify({'success': False, 'message': "No chip name configured for {}".format(target_name)}), 400

    hw_data = _load_json(_unified_cache_path(chip_name, 'HWPDT')) or _load_json(_cache_path_for(chip_name, 'HWPDT')) or {'devices': []}
    sw_data = _load_json(_unified_cache_path(chip_name, 'SWPDT')) or _load_json(_cache_path_for(chip_name, 'SWPDT')) or {'devices': []}

    filtered_devices = []
    path_upper = path.strip().upper()
    for device in hw_data.get('devices', []) + sw_data.get('devices', []):
        if str(device.get('taxonomy_path') or '').strip().upper() == path_upper:
            filtered_devices.append(device)
    
    return jsonify({
        'success': True,
        'path': path,
        'chipset': chip_name,
        'count': len(filtered_devices),
        'devices': filtered_devices,
    })


@device_summary_api_bp.route('/api/device_summary_data/save_sw_del/<string:target_name>', methods=['POST'])
@login_required
def api_save_sw_del_data(target_name):
    chip_name = get_chip_name_for_target(target_name) or ''
    if not chip_name:
        return jsonify({'success': False, 'message': "No chip name configured for {}".format(target_name)}), 400
    payload = request.get_json(force=True) or {}
    data = {
        'chipset': chip_name.upper(),
        'saved_at': _dt.utcnow().isoformat() + 'Z',
        'rows': payload.get('rows') or [],
    }
    _save_json(_sw_del_override_path(chip_name), data)
    return jsonify({'success': True, 'saved_at': data['saved_at']})


@device_summary_api_bp.route('/api/device_summary_data/save_hw_metrics/<string:target_name>', methods=['POST'])
@login_required
def api_save_hw_metrics_data(target_name):
    chip_name = get_chip_name_for_target(target_name) or ''
    if not chip_name:
        return jsonify({'success': False, 'message': "No chip name configured for {}".format(target_name)}), 400
    payload = request.get_json(force=True) or {}
    data = {
        'chipset': chip_name.upper(),
        'saved_at': _dt.utcnow().isoformat() + 'Z',
        'columns': payload.get('columns') or ['REV0', 'REV1', 'Part Type', 'Total'],
        'rows': payload.get('rows') or [],
    }
    _save_json(_hw_metrics_override_path(chip_name), data)
    return jsonify({'success': True, 'saved_at': data['saved_at']})


# ---------------------------------------------------------------------------
# Device Pool Override APIs  (remove / edit / restore per device)
# ---------------------------------------------------------------------------

@device_summary_api_bp.route('/api/device_pool/<string:target_name>/remove', methods=['POST'])
@login_required
def api_device_pool_remove(target_name):
    chip_name = get_chip_name_for_target(target_name) or ''
    if not chip_name:
        return jsonify({'success': False, 'message': 'No chip name for target'}), 400
    body      = request.get_json(force=True) or {}
    device_id = str(body.get('device_id') or '').strip()
    pdt_type  = str(body.get('pdt_type') or 'SWPDT').strip().upper()
    if not device_id:
        return jsonify({'success': False, 'message': 'device_id required'}), 400
    ov = _load_device_pool_overrides(chip_name, pdt_type)
    removed = list(ov.get('removed') or [])
    if device_id not in removed:
        removed.append(device_id)
    ov['removed'] = removed
    _save_device_pool_overrides(chip_name, pdt_type, ov)
    return jsonify({'success': True, 'removed_count': len(removed)})


@device_summary_api_bp.route('/api/device_pool/<string:target_name>/restore', methods=['POST'])
@login_required
def api_device_pool_restore(target_name):
    chip_name = get_chip_name_for_target(target_name) or ''
    if not chip_name:
        return jsonify({'success': False, 'message': 'No chip name for target'}), 400
    body      = request.get_json(force=True) or {}
    device_id = str(body.get('device_id') or '').strip()
    pdt_type  = str(body.get('pdt_type') or 'SWPDT').strip().upper()
    ov = _load_device_pool_overrides(chip_name, pdt_type)
    ov['removed'] = [d for d in (ov.get('removed') or []) if d != device_id]
    _save_device_pool_overrides(chip_name, pdt_type, ov)
    return jsonify({'success': True, 'removed_count': len(ov['removed'])})


@device_summary_api_bp.route('/api/device_pool/<string:target_name>/edit', methods=['POST'])
@login_required
def api_device_pool_edit(target_name):
    chip_name = get_chip_name_for_target(target_name) or ''
    if not chip_name:
        return jsonify({'success': False, 'message': 'No chip name for target'}), 400
    body     = request.get_json(force=True) or {}
    pdt_type = str(body.get('pdt_type') or 'SWPDT').strip().upper()
    edits    = body.get('edits') or {}
    ov = _load_device_pool_overrides(chip_name, pdt_type)
    existing = dict(ov.get('edits') or {})
    for dev_id, fields in edits.items():
        if not dev_id:
            continue
        existing[dev_id] = {
            'mcn_display':         str(fields.get('mcn_display') or '').strip(),
            'storage_display':     str(fields.get('storage_display') or '').strip(),
            'location_display':    str(fields.get('location_display') or '').strip(),
            'rework_info_display': str(fields.get('rework_info_display') or '').strip(),
        }
    ov['edits'] = existing
    _save_device_pool_overrides(chip_name, pdt_type, ov)
    return jsonify({'success': True, 'edited_count': len(existing)})


@device_summary_api_bp.route('/api/device_pool/<string:target_name>/overrides')
@login_required
def api_device_pool_overrides(target_name):
    chip_name = get_chip_name_for_target(target_name) or ''
    pdt_type  = (request.args.get('pdt') or 'SWPDT').strip().upper()
    ov = _load_device_pool_overrides(chip_name, pdt_type)
    return jsonify({'success': True, 'chip_name': chip_name, 'pdt_type': pdt_type,
                    'removed': ov.get('removed', []), 'edits': ov.get('edits', {})})



def _ds_first_non_empty(*values):
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ''


_DSMCN_RE = re.compile(r'\b(\d{2}-\d{5}-\d{3,4})(?!\d)')


def _ds_device_identity(device: dict) -> str:
    raw = device.get('_raw') or {}
    return _ds_first_non_empty(
        device.get('chip_id'), device.get('chipId'), device.get('device_id'),
        device.get('serialNumber'), device.get('serial_number'), device.get('serial_no'),
        device.get('serial'), raw.get('chipId'), raw.get('chip_id'),
        raw.get('serialNumber'), raw.get('serial_number'), raw.get('serial_no'),
        raw.get('serial'), device.get('hostname'), raw.get('hostname')
    )


def _ds_device_host(device: dict) -> str:
    raw = device.get('_raw') or {}
    return _ds_first_non_empty(
        device.get('host_pc'), device.get('hostname'), device.get('hostName'),
        device.get('host'), raw.get('hostname'), raw.get('hostName'), raw.get('host')
    )


def _ds_device_mcn(device: dict) -> str:
    """Extract MCN from device dict.

    Priority (per Axiom API docs):
    1. mcn_display (user override)
    2. qdt_mcn (from QDT)
    3. mcn (already normalised from axiom_client — checks deviceMcn + RfCard.mcn + mcnRev)
    4. properties.deviceMcn (TestResourcePropertiesDto.deviceMcn)
    5. properties.RfCard.mcn (RF card MCN — shown in Axiom API docs)
    6. dependencies.mcnRev (MCN revision — shown in Axiom API docs as 'dependencies.mcnRev')
    7. raw.description / qdt_model_desc (regex pattern match)
    """
    raw = device.get('_raw') or {}
    props = raw.get('properties') or {}
    deps = raw.get('dependencies') or {}
    rf_card = props.get('RfCard') or {}

    # Direct MCN fields from Axiom API
    device_mcn_direct = props.get('deviceMcn') or ''
    rf_card_mcn = rf_card.get('mcn') or ''
    mcn_rev = deps.get('mcnRev') or device.get('mcn_rev') or ''

    for text in (
        device.get('mcn_display'), device.get('qdt_mcn'), device.get('mcn'),
        device_mcn_direct,
        rf_card_mcn,
        mcn_rev,
        raw.get('mcn'), raw.get('description'), device.get('description'),
        device.get('qdt_model_desc')
    ):
        if not text:
            continue
        text_str = str(text).strip()
        if text_str and text_str not in ('', 'None', 'null'):
            match = _DSMCN_RE.search(text_str)
            if match:
                return match.group(1)
            # Return direct MCN fields even if they don't match the regex pattern
            # (e.g. deviceMcn might be a non-standard format)
            if text in (device_mcn_direct, rf_card_mcn) and text_str:
                return text_str
    return _ds_first_non_empty(device.get('mcn_display'), device.get('qdt_mcn'), device.get('mcn'), device_mcn_direct, rf_card_mcn)


def _ds_device_storage(device: dict) -> str:
    raw = device.get('_raw') or {}
    deps = raw.get('dependencies') or {}
    return _ds_first_non_empty(
        device.get('storage_display'), device.get('storage'), device.get('Storage'),
        device.get('qdt_storage'), deps.get('storage Type'), deps.get('storage'),
        raw.get('storage')
    )


def _ds_device_location(device: dict) -> str:
    raw = device.get('_raw') or {}
    tax = str(raw.get('taxonomyPath') or device.get('taxonomy_path') or '').upper()
    loc = str(device.get('location') or device.get('site') or raw.get('location') or raw.get('site') or '').upper()
    if '/PDT/QIPL' in tax or 'QIPL' in loc or 'HYDERABAD' in loc:
        return 'QIPL'
    if '/PDT/CHINA' in tax or 'CHINA' in loc or 'SHANGHAI' in loc:
        return 'CH'
    if '/PDT/SD' in tax or 'SAN DIEGO' in tax or 'SANDIEGO' in tax or '/SD' in loc:
        return 'SD'
    return _ds_first_non_empty(device.get('location_display'), device.get('site'), raw.get('site'))


def _ds_device_quarantine(device: dict) -> bool:
    """Return True only when explicit inventory state/status fields say quarantined.

    Do not scan the full JSON blob/description because many Axiom/QDT records can
    contain words like "not quarantined" or historical text, which incorrectly
    marks every device as quarantine.
    """
    raw = device.get('_raw') or {}
    fields = (
        device.get('quarantine'),
        device.get('is_quarantined'),
        device.get('quarantine_status'),
        device.get('device_state'),
        device.get('device_status'),
        device.get('status'),
        device.get('state'),
        raw.get('quarantine'),
        raw.get('isQuarantined'),
        raw.get('quarantineStatus'),
        raw.get('deviceState'),
        raw.get('deviceStatus'),
        raw.get('status'),
        raw.get('state'),
    )
    for value in fields:
        if isinstance(value, bool):
            if value:
                return True
            continue
        text = str(value or '').strip().lower()
        if not text or text in ('false', '0', 'no', 'none', 'null', 'available', 'active', 'ready', 'idle'):
            continue
        if 'not quarant' in text or 'non quarant' in text:
            continue
        if 'quarantine' in text or 'quarantined' in text:
            return True
    return False


def _ds_load_cached_devices_for_target(target_name: str, pdt_type: str) -> tuple[str, list, str]:
    chip_name = (get_chip_name_for_target(target_name) or str(target_name or '')).strip().upper()
    data = (
        _load_json(_unified_cache_path(chip_name, pdt_type))
        or _load_json(_cache_path_for(chip_name, pdt_type))
        or {}
    )
    return chip_name, list(data.get('devices') or []), str(data.get('saved_at') or '')


def _build_global_chip_id_map() -> dict:
    """Build a map of chip_id → device data from ALL available caches.

    Indexes by BOTH serial number AND ADB IDs (properties.adbId), because
    axiom_job_summary.chip_ids stores ADB IDs for some chipsets (e.g. ALANA.LA.1.0
    uses ADB IDs like '046AA46518C2') while others use serial numbers.
    """
    chip_map = {}
    try:
        cache_dir = _axiom_cache_dir()
        for fname in os.listdir(cache_dir):
            if not fname.endswith('_unified.json'):
                continue
            fpath = os.path.join(cache_dir, fname)
            try:
                data = _load_json(fpath) or {}
                for dev in data.get('devices') or []:
                    # Index by serial number (primary identity)
                    dev_id = _ds_device_identity(dev)
                    if dev_id:
                        chip_map[dev_id.upper()] = dev
                    # Also index by ADB IDs (used in axiom_job_summary chip_ids)
                    raw = dev.get('_raw') or {}
                    props = raw.get('properties') or {}
                    adb_ids = props.get('adbId') or []
                    if isinstance(adb_ids, list):
                        for adb_id in adb_ids:
                            if adb_id:
                                chip_map[str(adb_id).strip().upper()] = dev
                    elif adb_ids:
                        chip_map[str(adb_ids).strip().upper()] = dev
            except Exception:
                pass
    except Exception:
        pass
    return chip_map


def _get_taxonomy_paths_for_sp(sp_names: list) -> list:
    """Return distinct Axiom taxonomy paths from job_summary for the given SP names.

    Used to know which taxonomy paths to query when fetching device details
    for SP-only mode (when the target has no chipset cache).
    """
    if not sp_names:
        return []
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    try:
        cur = conn.cursor(dictionary=True)
        where = ' AND software_product IN (' + ','.join(['%s'] * len(sp_names)) + ')'
        cur.execute("""
            SELECT DISTINCT taxonomy_path
            FROM pdt_stats_dashboard.axiom_job_summary
            WHERE state IN ('Running','JobSetup')
              AND taxonomy_path IS NOT NULL
              AND taxonomy_path != ''
        """ + where + " LIMIT 20", tuple(sp_names))
        paths = [str(r['taxonomy_path']).strip() for r in (cur.fetchall() or []) if r.get('taxonomy_path')]
        cur.close()
        conn.close()
        return list(dict.fromkeys(paths))  # deduplicate preserving order
    except Exception:
        return []


# Module-level progress tracker for SP-only Axiom fetch (per target)
_sp_fetch_progress: dict = {}  # target_name → {status, found, total, current_path}


def _normalise_axiom_device(raw_dev: dict, taxonomy_path: str = '', host_pc: str = '') -> dict:
    """Return the same normalized shape used by Axiom device cache rows."""
    raw_dev = raw_dev or {}
    props = raw_dev.get('properties') or {}
    deps = raw_dev.get('dependencies') or {}
    rf_card = props.get('RfCard') or {}
    mcn_val = (
        props.get('deviceMcn')
        or rf_card.get('mcn')
        or deps.get('mcnRev')
        or deps.get('mcn')
        or ''
    )
    storage_val = props.get('storageType') or deps.get('storageType') or ''
    return {
        'id':            raw_dev.get('id'),
        'serial_number': props.get('serialNumber') or '',
        'hostname':      raw_dev.get('hostname') or props.get('hostname') or '',
        'host_pc':       host_pc or raw_dev.get('hostName') or raw_dev.get('host') or '',
        'location':      raw_dev.get('location') or '',
        'asset_tag_id':  props.get('assetTagId') or '',
        'chipset':       deps.get('chipset') or '',
        'chipset_rev':   deps.get('chipsetRev') or '',
        'mcn_rev':       deps.get('mcnRev') or '',
        'form_factor':   deps.get('formFactor') or '',
        'device_type':   props.get('deviceType') or '',
        'mcn':           str(mcn_val or ''),
        'storage':       str(storage_val or ''),
        'heartbeat':     raw_dev.get('heartbeat') or '',
        'taxonomy_path': raw_dev.get('taxonomyPath') or taxonomy_path,
        '_raw':          raw_dev,
    }


def _index_axiom_device_aliases(out: dict, entry: dict, wanted: set, extra_aliases=None) -> None:
    """Index a normalized Axiom device by all chip-id-like aliases present."""
    raw = entry.get('_raw') or {}
    props = raw.get('properties') or {}
    aliases = set(str(x or '').strip().upper() for x in (extra_aliases or []) if str(x or '').strip())
    for value in (
        entry.get('serial_number'), entry.get('device_id'), entry.get('chip_id'),
        raw.get('serialNumber'), raw.get('chipId'), raw.get('id'), props.get('serialNumber'),
        props.get('hwId'), props.get('macAddress'),
    ):
        text = str(value or '').strip().upper()
        if text:
            aliases.add(text)
            aliases.add(text.replace(':', '').replace('-', '').replace('.', ''))
    for field in ('adbId', 'edlId', 'deviceSerialNumbers', 'chipIdSerialNumbers', 'serialNumbers'):
        value = props.get(field) if field in props else raw.get(field)
        if isinstance(value, list):
            aliases.update(str(x or '').strip().upper() for x in value if str(x or '').strip())
        elif value:
            aliases.add(str(value).strip().upper())
    for alias in aliases:
        if alias and (not wanted or alias in wanted):
            out[alias] = entry


def _fetch_resource_for_sp_serial(serial: str, taxonomy_path: str = '', resource_id: str = '', host_pc: str = '') -> dict:
    """Fetch one Axiom Device resource using serial/resource identifiers from job playlists."""
    from src.axiom_client import _paginate, axiom_get

    serial = str(serial or '').strip()
    resource_id = str(resource_id or '').strip()
    taxonomy_path = str(taxonomy_path or '').strip()
    candidates = []

    if resource_id:
        candidates.append(('direct', f"/axiom/v1/public/resources/{resource_id}"))

    if serial:
        query_serial = serial.replace(' ', '%20')
        if taxonomy_path:
            candidates.append(('query', f"/axiom/v1/public/resources?taxonomyPath={taxonomy_path}&type=Device&serialNumber={query_serial}"))
        candidates.append(('query', f"/axiom/v1/public/resources?taxonomyPath=/PDT&type=Device&serialNumber={query_serial}"))

    for mode, path in candidates:
        try:
            if mode == 'direct':
                raw = axiom_get(path)
                if isinstance(raw, dict) and raw:
                    return _normalise_axiom_device(raw, taxonomy_path=taxonomy_path, host_pc=host_pc)
            else:
                for raw in _paginate(path, page_size=20, max_pages=2):
                    return _normalise_axiom_device(raw, taxonomy_path=taxonomy_path, host_pc=host_pc)
        except Exception as exc:
            logger.debug('[SP FETCH JOBS] resource lookup failed path=%s: %s', path, exc)

    return {}


def _scan_sp_devices_by_taxonomy(sp_names: list, wanted: set, taxonomy_paths: list | None = None, target_name: str = '') -> dict:
    """Fallback SP device lookup by scanning Axiom /resources under job taxonomy paths.

    Job playlist track resource names do not always match `axiom_job_summary.chip_ids`
    directly. This fallback indexes each Device resource by serial/ADB/MAC/EDL aliases
    and fills only the requested chip IDs.
    """
    from src.axiom_client import _paginate

    wanted = {str(c or '').strip().upper() for c in (wanted or set()) if str(c or '').strip()}
    if not wanted:
        return {}

    paths = [str(p or '').strip() for p in (taxonomy_paths or []) if str(p or '').strip()]
    if not paths:
        paths = _get_taxonomy_paths_for_sp(sp_names)
    paths = list(dict.fromkeys(paths))
    out = {}

    for path_idx, tax_path in enumerate(paths):
        if len(out) >= len(wanted):
            break
        if target_name:
            _sp_fetch_progress.setdefault(target_name, {})
            _sp_fetch_progress[target_name]['current_path'] = f'resources {tax_path}'
            _sp_fetch_progress[target_name]['paths_done'] = path_idx
            _sp_fetch_progress[target_name]['paths_total'] = max(_sp_fetch_progress[target_name].get('paths_total', 0), len(paths))
        try:
            base = f"/axiom/v1/public/resources?taxonomyPath={tax_path}&type=Device"
            for raw_dev in _paginate(base, page_size=200):
                entry = _normalise_axiom_device(raw_dev, taxonomy_path=tax_path)
                before = len(out)
                _index_axiom_device_aliases(out, entry, wanted)
                if target_name and len(out) != before:
                    _sp_fetch_progress[target_name]['found'] = len(out)
                if len(out) >= len(wanted):
                    break
        except Exception as exc:
            logger.debug('[SP FETCH JOBS] taxonomy fallback failed taxonomy=%s: %s', tax_path, exc)

    return out


def _fetch_sp_devices_via_jobs(sp_names: list, chip_ids: set, target_name: str = '') -> dict:
    """Fetch SP device details via job playlists, then enrich from Axiom /resources.

    Flow:
      1. Read active job IDs from local axiom_job_summary for the selected SP names.
      2. Call /jobs/{id}/data/playlists for each job to map chip/resource serials to host PCs.
      3. Use those serial/resource IDs to fetch /resources device details (MCN, storage, taxonomy).
      4. Return dict keyed by requested chip_id.upper().
    """
    from src.axiom_client import axiom_get

    sp_names = [str(x).strip() for x in (sp_names or []) if str(x).strip()]
    wanted = {str(c).strip().upper() for c in (chip_ids or set()) if str(c).strip()}
    if not sp_names or not wanted:
        return {}

    if target_name:
        _sp_fetch_progress[target_name] = {
            'status': 'fetching',
            'found': 0,
            'total': len(wanted),
            'paths': [],
            'current_path': 'job playlists',
            'paths_done': 0,
            'paths_total': 0,
        }

    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {}

    jobs = []
    cur = conn.cursor(dictionary=True)
    try:
        where = ' AND software_product IN (' + ','.join(['%s'] * len(sp_names)) + ')'
        cur.execute("""
            SELECT job_id, chip_ids, taxonomy_path, site, city_team
            FROM `pdt_stats_dashboard`.`axiom_job_summary`
            WHERE state IN ('Running','JobSetup')
              AND job_id IS NOT NULL
              AND chip_ids IS NOT NULL
        """ + where + """
            ORDER BY started_at DESC
            LIMIT 300
        """, tuple(sp_names))
        for row in cur.fetchall() or []:
            try:
                chips = json.loads(row.get('chip_ids') or '[]') if isinstance(row.get('chip_ids'), str) else list(row.get('chip_ids') or [])
            except Exception:
                chips = []
            row_chips = {str(c or '').strip().upper() for c in chips if str(c or '').strip()}
            if row_chips & wanted:
                row['_chip_set'] = row_chips
                jobs.append(row)
    except Exception as exc:
        logger.warning('[SP FETCH JOBS] failed to load job_summary rows: %s', exc)
        return {}
    finally:
        cur.close(); conn.close()

    # Match the proven ALANA debug script flow as well:
    #   /axiom/v1/public/jobs?taxonomyPath=/PDT&softwareProduct=...&submittedFrom=...&expand=chipIdSerialNumbers&state=Running
    # The local DB can lag or store chip IDs differently, while the live /jobs
    # response exposes `jobId` + `chipIdSerialNumbers` directly.
    try:
        from datetime import timedelta
        from src.axiom_client import _paginate
        from_dt = (_dt.utcnow() - timedelta(days=2)).strftime('%Y-%m-%dT%H:%M:%SZ')
        seen_job_ids = {str(j.get('job_id') or '').strip() for j in jobs if str(j.get('job_id') or '').strip()}
        for sp in sp_names:
            for state in ('Running', 'JobSetup'):
                base = (
                    f"/axiom/v1/public/jobs?taxonomyPath=/PDT"
                    f"&softwareProduct={sp}"
                    f"&submittedFrom={from_dt}"
                    f"&expand=chipIdSerialNumbers"
                    f"&state={state}"
                )
                for live_job in _paginate(base, page_size=50, max_pages=6):
                    job_id = str(live_job.get('jobId') or live_job.get('job_id') or live_job.get('id') or '').strip()
                    if not job_id or job_id in seen_job_ids:
                        continue
                    chips = live_job.get('chipIdSerialNumbers') or live_job.get('chip_ids') or []
                    row_chips = {str(c or '').strip().upper() for c in chips if str(c or '').strip()}
                    if not (row_chips & wanted):
                        continue
                    seen_job_ids.add(job_id)
                    jobs.append({
                        'job_id': job_id,
                        'chip_ids': list(row_chips),
                        'taxonomy_path': str(live_job.get('taxonomyPath') or '/PDT'),
                        'site': str(live_job.get('site') or live_job.get('cityTeam') or ''),
                        'city_team': str(live_job.get('cityTeam') or live_job.get('city_team') or ''),
                        '_chip_set': row_chips,
                        '_source': 'axiom_live_jobs',
                    })
    except Exception as exc:
        logger.debug('[SP FETCH JOBS] live /jobs lookup failed, continuing with DB rows: %s', exc)

    if target_name:
        _sp_fetch_progress[target_name]['paths_total'] = len(jobs)

    out = {}
    resource_candidates = {}
    taxonomy_paths_seen = []

    def _playlist_candidates_for_job(idx_job):
        idx, job = idx_job
        job_id = str(job.get('job_id') or '').strip()
        job_taxonomy_path = str(job.get('taxonomy_path') or '').strip()
        if not job_id:
            return job_id, job_taxonomy_path, {}
        try:
            payload = axiom_get(f"/axiom/v1/public/jobs/{job_id}/data/playlists?pageNumber=0&pageSize=100")
        except Exception as exc:
            logger.debug('[SP FETCH JOBS] playlist fetch failed job=%s: %s', job_id, exc)
            return job_id, job_taxonomy_path, {}

        candidates = {}
        items = payload.get('data') or payload.get('content') or payload.get('resources') or []
        for playlist in items:
            if not isinstance(playlist, dict):
                continue
            tracks = playlist.get('playlistStatusOfEachTrack') or []
            for track in tracks if isinstance(tracks, list) else []:
                if not isinstance(track, dict):
                    continue
                resource = track.get('testResource') or {}
                if not isinstance(resource, dict):
                    resource = {}
                aliases = {
                    str(resource.get('name') or '').strip().upper(),
                    str(resource.get('serialNumber') or '').strip().upper(),
                    str(resource.get('resourceId') or '').strip().upper(),
                    str(resource.get('id') or '').strip().upper(),
                    str(resource.get('adbId') or '').strip().upper(),
                    str(resource.get('macAddress') or '').strip().upper().replace(':', '').replace('-', '').replace('.', ''),
                    str(resource.get('edlId') or '').strip().upper(),
                }
                matched = wanted & {a for a in aliases if a}
                if not matched:
                    continue
                host_pc = str(track.get('hostName') or track.get('host') or job.get('site') or job.get('city_team') or '').strip()
                serial = str(resource.get('name') or resource.get('serialNumber') or '').strip()
                resource_id = str(resource.get('resourceId') or resource.get('id') or '').strip()
                for chip in matched:
                    candidates.setdefault(chip, {
                        'serial': serial or chip,
                        'resource_id': resource_id,
                        'host_pc': host_pc,
                        'taxonomy_path': job_taxonomy_path,
                    })
        return job_id, job_taxonomy_path, candidates

    try:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        max_workers = min(12, max(1, len(jobs)))
        with ThreadPoolExecutor(max_workers=max_workers) as pool:
            futures = [pool.submit(_playlist_candidates_for_job, item) for item in enumerate(jobs)]
            for done_count, fut in enumerate(as_completed(futures), start=1):
                job_id, tax_path, candidates = fut.result()
                if tax_path:
                    taxonomy_paths_seen.append(tax_path)
                for chip, cand in (candidates or {}).items():
                    resource_candidates.setdefault(chip, cand)
                if target_name:
                    _sp_fetch_progress[target_name]['current_path'] = f'playlists {done_count}/{len(jobs)}'
                    _sp_fetch_progress[target_name]['paths_done'] = done_count
                if len(resource_candidates) >= len(wanted):
                    break
    except Exception as exc:
        logger.debug('[SP FETCH JOBS] parallel playlist fetch failed: %s', exc)

    def _resource_entry_for_candidate(item):
        chip, cand = item
        entry = _fetch_resource_for_sp_serial(
            cand.get('serial') or chip,
            taxonomy_path=cand.get('taxonomy_path') or '',
            resource_id=cand.get('resource_id') or '',
            host_pc=cand.get('host_pc') or '',
        )
        return chip, cand, entry

    if resource_candidates:
        try:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            max_workers = min(16, max(1, len(resource_candidates)))
            with ThreadPoolExecutor(max_workers=max_workers) as pool:
                futures = [pool.submit(_resource_entry_for_candidate, item) for item in resource_candidates.items()]
                for fut in as_completed(futures):
                    chip, cand, entry = fut.result()
                    if not entry:
                        continue
                    if cand.get('host_pc') and not entry.get('host_pc'):
                        entry['host_pc'] = cand.get('host_pc')
                    _index_axiom_device_aliases(out, entry, wanted, extra_aliases=[chip, cand.get('serial'), cand.get('resource_id')])
                    if target_name:
                        _sp_fetch_progress[target_name]['found'] = len(out)
                    if len(out) >= len(wanted):
                        break
        except Exception as exc:
            logger.debug('[SP FETCH JOBS] parallel resource fetch failed: %s', exc)

    missing = wanted - set(out.keys())
    if missing:
        fallback = _scan_sp_devices_by_taxonomy(
            sp_names,
            missing,
            taxonomy_paths=list(dict.fromkeys(taxonomy_paths_seen)),
            target_name=target_name,
        )
        out.update(fallback)

    if target_name:
        _sp_fetch_progress[target_name]['status'] = 'done'
        _sp_fetch_progress[target_name]['found'] = len(out)

    logger.info('[SP FETCH JOBS] Done: found %d/%d devices via %d jobs for SPs=%s', len(out), len(wanted), len(jobs), sp_names)
    return out


def _fetch_sp_devices_from_axiom(sp_names: list, chip_ids: set, target_name: str = '') -> dict:
    """Backward-compatible wrapper for SP-only live fetch.

    Prefer the job-playlist path because it directly exposes resource serials and
    host PCs for the active jobs, avoiding broad taxonomy scans.
    """
    return _fetch_sp_devices_via_jobs(sp_names, chip_ids, target_name=target_name)


def _enrich_sp_rows_with_live_axiom_status(raw_devices: list, max_workers: int = 5) -> list:
    """Refresh SP-mode device details from live Axiom /resources in parallel.

    DB `axiom_job_summary` is used as the device list/source of active jobs.
    This helper uses Axiom public `/resources` as the realtime source for host,
    MCN, taxonomy/location, heartbeat, status, and quarantine-like fields.

    If Axiom cannot match a device, the DB-backed row is returned unchanged.
    """
    rows = list(raw_devices or [])
    if not rows:
        return rows

    def _enrich_one(device: dict) -> dict:
        dev_id = _ds_device_identity(device)
        host_pc = _ds_device_host(device)
        taxonomy_path = str(device.get('taxonomy_path') or '').strip()
        if not dev_id:
            return device

        live = _fetch_resource_for_sp_serial(
            dev_id,
            taxonomy_path=taxonomy_path,
            host_pc=host_pc,
        )
        if not live:
            return device

        merged = dict(device)
        # Keep DB/result-level device id stable so active_map/running jobs attach.
        merged.update({k: v for k, v in live.items() if v not in (None, '')})
        merged['chip_id'] = str(device.get('chip_id') or dev_id).strip()
        merged['device_id'] = str(device.get('device_id') or dev_id).strip()
        if host_pc and not merged.get('host_pc'):
            merged['host_pc'] = host_pc
        merged['_from_axiom_job_summary'] = device.get('_from_axiom_job_summary')
        merged['_live_axiom_refreshed'] = True
        return merged

    try:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        out = [None] * len(rows)
        with ThreadPoolExecutor(max_workers=max(1, min(int(max_workers or 5), 5, len(rows)))) as pool:
            future_map = {pool.submit(_enrich_one, row): idx for idx, row in enumerate(rows)}
            for fut in as_completed(future_map):
                idx = future_map[fut]
                try:
                    out[idx] = fut.result()
                except Exception as exc:
                    logger.debug('[DEVICE SUMMARY] live Axiom status refresh failed for row %d: %s', idx, exc)
                    out[idx] = rows[idx]
        return [row if row is not None else rows[idx] for idx, row in enumerate(out)]
    except Exception as exc:
        logger.debug('[DEVICE SUMMARY] parallel live Axiom status refresh failed: %s', exc)
        return rows


def _ds_active_axiom_device_map(device_ids: set[str], sp_names=None) -> dict:
    """Return active Axiom job info keyed by chip/device ID.

    Prefer the generated pdt_stats_dashboard.axiom_all_devices table populated
    by scripts/update_axiom_job_summary.py.  That table already expands active
    jobs to one row per device with host/SP/build/date, so Device Summary can
    avoid live Axiom job/playlist checks for pages such as Hamoa_LA_1_0.
    Falls back to axiom_job_summary for older deployments before the generated
    table exists.
    """
    sp_names = [str(x).strip() for x in (sp_names or []) if str(x).strip()]
    if not device_ids and not sp_names:
        return {}
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return {}
    cur = conn.cursor(dictionary=True)
    active = {}
    try:
        wanted = {str(x).strip().upper() for x in device_ids if str(x).strip()}
        include_all_chips = not wanted and bool(sp_names)

        where_extra = ''
        params = []
        if sp_names:
            where_extra = ' AND software_product IN (' + ','.join(['%s'] * len(sp_names)) + ')'
            params.extend(sp_names)

        # Fast path: generated all-devices table from update_axiom_job_summary.py.
        try:
            cur.execute("""
                SELECT device_id, host_name, software_product, build_running,
                       build_name, job_id, state, taxonomy_path, site, city_team,
                       started_at, job_date, heartbeat
                FROM `pdt_stats_dashboard`.`axiom_all_devices`
                WHERE state IN ('Running','JobSetup')
            """ + where_extra + """
                ORDER BY started_at DESC, generated_at DESC
                LIMIT 20000
            """, tuple(params))
            for row in cur.fetchall() or []:
                key = str(row.get('device_id') or '').strip().upper()
                if not key or (not include_all_chips and key not in wanted):
                    continue
                active.setdefault(key, []).append({
                    'job_id': str(row.get('job_id') or ''),
                    'build_name': str(row.get('build_running') or row.get('build_name') or ''),
                    'pl_id': str(row.get('software_product') or ''),
                    'state': str(row.get('state') or ''),
                    'site': str(row.get('host_name') or row.get('site') or row.get('city_team') or ''),
                    'host_pc': str(row.get('host_name') or ''),
                    'taxonomy_path': str(row.get('taxonomy_path') or ''),
                    'submitter': '',
                    'started_at': str(row.get('started_at') or row.get('job_date') or ''),
                    'heartbeat': str(row.get('heartbeat') or ''),
                    '_source': 'axiom_all_devices',
                })
            if active:
                return active
        except Exception as exc:
            logger.debug('[DEVICE SUMMARY] axiom_all_devices unavailable, falling back to job_summary: %s', exc)

        cur.execute("""
            SELECT job_id, build_name, software_product, state, chip_ids, site,
                   city_team, taxonomy_path, submitter, started_at,
                   device_host_map, device_hostnames
            FROM `pdt_stats_dashboard`.`axiom_job_summary`
            WHERE state IN ('Running','JobSetup')
              AND chip_ids IS NOT NULL
        """ + where_extra + """
            ORDER BY started_at DESC
            LIMIT 5000
        """, tuple(params))
        include_all_chips = not wanted and bool(sp_names)
        for row in cur.fetchall() or []:
            try:
                chips = json.loads(row.get('chip_ids') or '[]') if isinstance(row.get('chip_ids'), str) else list(row.get('chip_ids') or [])
            except Exception:
                chips = []
            base_job = {
                'job_id': str(row.get('job_id') or ''),
                'build_name': str(row.get('build_name') or ''),
                'pl_id': str(row.get('software_product') or ''),
                'state': str(row.get('state') or ''),
                'site': str(row.get('site') or row.get('city_team') or ''),
                'taxonomy_path': str(row.get('taxonomy_path') or ''),
                'submitter': str(row.get('submitter') or ''),
                'started_at': str(row.get('started_at') or ''),
            }
            for chip in chips:
                key = str(chip or '').strip().upper()
                if key and (include_all_chips or key in wanted):
                    active.setdefault(key, []).append(dict(base_job))

            # New DB-backed device-host mapping populated by the Axiom updater:
            #   device_host_map = {testCaseTestResourceName: testCaseHostName}
            # Example: {"TDC00002MCCH": "Lab7181"}. These are the actual
            # device IDs/host PCs shown by /jobs/{id}/results for ALANA.
            try:
                host_map = row.get('device_host_map') or {}
                if isinstance(host_map, str):
                    host_map = json.loads(host_map or '{}')
            except Exception:
                host_map = {}
            if isinstance(host_map, dict):
                for dev_id, host_name in host_map.items():
                    key = str(dev_id or '').strip().upper()
                    if not key:
                        continue
                    if include_all_chips or key in wanted:
                        job = dict(base_job)
                        job['site'] = str(host_name or '').strip() or job.get('site', '')
                        job['host_pc'] = str(host_name or '').strip()
                        active.setdefault(key, []).append(job)
        return active
    except Exception:
        return {}
    finally:
        cur.close(); conn.close()


@device_summary_api_bp.route('/api/device_summary_data/<string:target_name>/debug_axiom_sp')
@login_required
def api_debug_axiom_sp(target_name):
    """Debug: show what Axiom API returns for SP-only mode devices.

    Usage: GET /api/device_summary_data/{target}/debug_axiom_sp?sp=ALANA.LA.1.0
    Returns sample device records from Axiom to understand the identifier format.
    """
    sp_names = [x.strip() for x in re.split(r'[,;\n]+', str(request.args.get('sp') or '')) if x.strip()]
    if not sp_names:
        return jsonify({'error': 'sp parameter required. Usage: ?sp=ALANA.LA.1.0'}), 400

    taxonomy_paths = _get_taxonomy_paths_for_sp(sp_names)
    if not taxonomy_paths:
        return jsonify({'error': 'No taxonomy paths found for SP names', 'sp_names': sp_names}), 404

    from src.axiom_client import _paginate, AXIOM_FETCH_DISABLED
    if AXIOM_FETCH_DISABLED:
        return jsonify({'error': 'Axiom fetch disabled (AXIOM_FETCH_DISABLED=True)'}), 400

    results = {}
    for tax_path in taxonomy_paths[:3]:  # limit to first 3 paths
        base = f"/axiom/v1/public/resources?taxonomyPath={tax_path}&type=Device"
        try:
            sample_devices = []
            for raw_dev in _paginate(base, page_size=10):
                props = raw_dev.get('properties') or {}
                deps  = raw_dev.get('dependencies') or {}
                rf_card = props.get('RfCard') or {}
                sample_devices.append({
                    'id':           raw_dev.get('id'),
                    'hostname':     raw_dev.get('hostname'),
                    'taxonomyPath': raw_dev.get('taxonomyPath'),
                    'location':     raw_dev.get('location'),
                    # Identity fields — which one matches axiom_job_summary.chip_ids?
                    'serialNumber': props.get('serialNumber'),
                    'adbId':        props.get('adbId'),
                    'macAddress':   props.get('macAddress'),
                    'edlId':        props.get('edlId'),
                    'hwId':         props.get('hwId'),
                    # MCN fields
                    'deviceMcn':    props.get('deviceMcn'),
                    'rfCard_mcn':   rf_card.get('mcn'),
                    'mcnRev':       deps.get('mcnRev'),
                    # Chipset
                    'chipset':      deps.get('chipset'),
                    'chipsetRev':   deps.get('chipsetRev'),
                })
                if len(sample_devices) >= 5:
                    break
            results[tax_path] = {'sample_count': len(sample_devices), 'devices': sample_devices}
        except Exception as exc:
            results[tax_path] = {'error': str(exc)}

    # Also show what chip_ids look like in job_summary for these SPs
    conn = get_mysql_connection_db(bu_key=None)
    sample_chip_ids = []
    if conn:
        try:
            cur = conn.cursor(dictionary=True)
            where = ' AND software_product IN (' + ','.join(['%s'] * len(sp_names)) + ')'
            cur.execute("""
                SELECT chip_ids, taxonomy_path, site
                FROM pdt_stats_dashboard.axiom_job_summary
                WHERE state IN ('Running','JobSetup') AND chip_ids IS NOT NULL
            """ + where + " LIMIT 3", tuple(sp_names))
            for row in cur.fetchall() or []:
                try:
                    chips = json.loads(row.get('chip_ids') or '[]') if isinstance(row.get('chip_ids'), str) else list(row.get('chip_ids') or [])
                    sample_chip_ids.append({'chip_ids': chips[:5], 'taxonomy_path': row.get('taxonomy_path'), 'site': row.get('site')})
                except Exception:
                    pass
            cur.close()
            conn.close()
        except Exception:
            pass

    return jsonify({
        'success': True,
        'sp_names': sp_names,
        'taxonomy_paths': taxonomy_paths,
        'job_summary_chip_ids_sample': sample_chip_ids,
        'axiom_device_samples': results,
        'note': 'Compare chip_ids from job_summary with adbId/macAddress/serialNumber/edlId from Axiom to find the matching field',
    })


@device_summary_api_bp.route('/api/device_summary_data/<string:target_name>/sp_fetch_progress')
@login_required
def api_sp_fetch_progress(target_name):
    """Return current SP-only Axiom fetch progress for the UI progress indicator."""
    prog = _sp_fetch_progress.get(target_name) or {}
    return jsonify({
        'success': True,
        'status': prog.get('status', 'idle'),
        'found': prog.get('found', 0),
        'total': prog.get('total', 0),
        'current_path': prog.get('current_path', ''),
        'paths_done': prog.get('paths_done', 0),
        'paths_total': prog.get('paths_total', 0),
        'paths': prog.get('paths', []),
    })


def _ds_default_sp_names_for_target(target_name: str) -> list[str]:
    """Resolve default Axiom software product/SP names for a dashboard target.

    Primary source is dashboard_status.sp_name. If it is missing, infer recent
    software_product values from axiom_job_summary using target/chip terms.
    """
    target = str(target_name or '').strip()
    if not target:
        return []
    conn = get_mysql_connection_db(bu_key=None)
    if not conn:
        return []
    names = []
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT target_name, target_display, sp_name, chip_name
            FROM pdt_stats_dashboard.dashboard_status
            WHERE is_active = 1
              AND (
                    LOWER(target_name) = LOWER(%s)
                 OR LOWER(REPLACE(target_name, '_', '.')) = LOWER(%s)
                 OR LOWER(sp_name) = LOWER(%s)
                 OR LOWER(REPLACE(sp_name, '.', '_')) = LOWER(%s)
              )
            ORDER BY id DESC
            LIMIT 1
        """, (target, target, target, target))
        row = cur.fetchone() or {}
        sp = str(row.get('sp_name') or '').strip()
        if sp and sp.upper() not in ('N/A', 'NA', 'NONE', 'NULL'):
            names.append(sp)

        # Fallback: find recent software_product rows matching target/chip/display terms.
        if not names:
            terms = []
            for value in (target, row.get('chip_name'), row.get('target_display')):
                value = str(value or '').strip()
                if not value:
                    continue
                terms.extend([value, value.replace('_', '.'), value.replace('.', '_'), value.split('_')[0].split('.')[0]])
            terms = [t for t in dict.fromkeys(terms) if len(t) >= 3]
            if terms:
                where = ' OR '.join(['software_product LIKE %s OR build_name LIKE %s OR build_id LIKE %s'] * len(terms))
                params = []
                for term in terms:
                    like = f'%{term}%'
                    params.extend([like, like, like])
                cur.execute(f"""
                    SELECT software_product, MAX(started_at) AS last_seen, COUNT(*) AS cnt
                    FROM pdt_stats_dashboard.axiom_job_summary
                    WHERE software_product IS NOT NULL
                      AND software_product != ''
                      AND ({where})
                    GROUP BY software_product
                    ORDER BY last_seen DESC, cnt DESC
                    LIMIT 3
                """, tuple(params))
                for r in cur.fetchall() or []:
                    sp = str(r.get('software_product') or '').strip()
                    if sp and sp not in names:
                        names.append(sp)
        cur.close()
    except Exception as exc:
        logger.debug('[DEVICE SUMMARY] default SP lookup failed target=%s: %s', target_name, exc)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return names[:3]


@device_summary_api_bp.route('/api/device_summary_data/<string:target_name>/default_sp')
@login_required
def api_device_summary_default_sp(target_name):
    names = _ds_default_sp_names_for_target(target_name)
    return jsonify({
        'success': True,
        'target': target_name,
        'sp_names': names,
        'sp_names_text': ', '.join(names),
    })


@device_summary_api_bp.route('/api/device_summary_data/<string:target_name>/inventory_summary')
@login_required
def api_device_inventory_summary(target_name):
    """Device Summary inventory grouped by MCN/host with running/quarantine status.

    Query filters:
      pdt=SWPDT|HWPDT, mcn=<mcn>, host=<host>, status=all|running|idle|quarantine, q=<text>
    """
    pdt_type = (request.args.get('pdt') or 'SWPDT').strip().upper()
    mcn_filter = str(request.args.get('mcn') or '').strip().lower()
    host_filter = str(request.args.get('host') or '').strip().lower()
    site_filter = str(request.args.get('site') or '').strip().lower()
    status_filter = str(request.args.get('status') or 'all').strip().lower()
    q = str(request.args.get('q') or '').strip().lower()
    sp_names = [
        x.strip() for x in re.split(r'[,;\n]+', str(request.args.get('sp_names') or request.args.get('sp') or ''))
        if x.strip()
    ]
    if not sp_names and str(request.args.get('auto_sp') or '0').strip().lower() in ('1', 'true', 'yes', 'on'):
        sp_names = _ds_default_sp_names_for_target(target_name)
    refresh = str(request.args.get('refresh') or '0').strip().lower() in ('1', 'true', 'yes', 'on')

    cache_key = (
        str(target_name or '').strip().lower(),
        pdt_type,
        tuple(sorted(str(x).strip().upper() for x in sp_names if str(x).strip())),
    )
    cache_hit = None
    if not refresh:
        try:
            import time as _time
            cached = _inventory_summary_cache.get(cache_key) or {}
            if cached and (_time.time() - float(cached.get('ts') or 0)) <= _INVENTORY_SUMMARY_CACHE_TTL_SECONDS:
                cache_hit = cached
        except Exception:
            cache_hit = None

    if cache_hit:
        chip_name = cache_hit.get('chip_name') or ''
        raw_devices = list(cache_hit.get('raw_devices') or [])
        saved_at = str(cache_hit.get('saved_at') or '')
        active_map = cache_hit.get('active_map') or {}
    else:
        chip_name, raw_devices, saved_at = _ds_load_cached_devices_for_target(target_name, pdt_type)

    if not cache_hit and sp_names:
        # SP-backed pages can now use the generated all-devices table directly.
        # Do this before any live Axiom device fetch so /device-summary/<SP>
        # can render from local DB data produced by update_axiom_job_summary.py.
        active_map = _ds_active_axiom_device_map(set(), sp_names=sp_names)

    # If the target has not been synced yet, or user clicks Refresh Inventory,
    # pull Axiom/QDT live and save the same unified cache used by the existing
    # Device Summary API. Without this, a never-synced target shows 0 rows.
    if not cache_hit and not sp_names and (refresh or not raw_devices):
        if is_axiom_enabled_for_target(target_name):
            try:
                live_devices = get_devices_by_chipset(chip_name, pdt_type=pdt_type, include_site_details=True) or []
                if pdt_type == 'SWPDT':
                    live_devices = [
                        d for d in live_devices
                        if not str(
                            (d.get('_raw') or {}).get('taxonomyPath')
                            or d.get('taxonomy_path')
                            or ''
                        ).strip().upper().startswith('/PDT/QIPL/HW')
                    ]
                merged = _merge_axiom_qdt_devices(chip_name, pdt_type, live_devices)
                _save_json(_unified_cache_path(chip_name, pdt_type), merged)
                _save_json(_cache_path_for(chip_name, pdt_type), merged)
                raw_devices = list(merged.get('devices') or [])
                saved_at = str(merged.get('saved_at') or '')
            except Exception as exc:
                logger.warning('[DEVICE SUMMARY INVENTORY] live refresh failed target=%s pdt=%s: %s', target_name, pdt_type, exc)

    if not cache_hit and 'active_map' not in locals():
        device_ids = {_ds_device_identity(d) for d in raw_devices}
        active_map = _ds_active_axiom_device_map(device_ids, sp_names=sp_names)

    # If user supplied SP/software-product names, build the rows from the active
    # job chip IDs. Do this not only when inventory is empty, but also on refresh
    # or when the existing cache only contains placeholder rows with blank MCN.
    # Otherwise a previous "Unknown MCN/site" SP-only cache would keep masking the
    # newly fetched Axiom /resources details.
    needs_sp_rebuild = bool(sp_names and active_map) and (
        refresh
        or not raw_devices
        or not any(_ds_device_mcn(d) for d in raw_devices)
    )
    if needs_sp_rebuild:
        chip_ids_set = set(active_map.keys())

        # Step 1: Try global chip map (serial numbers + ADB IDs from all caches)
        global_chip_map = _build_global_chip_id_map()

        # Step 2: If refresh requested OR chip IDs not in global map, fetch from Axiom.
        # NOTE: is_axiom_enabled_for_target() checks chip_name in dashboard_status.
        # For SP-only mode the target name IS the SP name (e.g. 'ALANA.LA.1.0'), not a
        # chipset, so that check always returns False.  Instead, check Axiom credentials.
        axiom_adb_map = {}
        if refresh or not any(global_chip_map.get(c.upper()) for c in chip_ids_set):
            from src.axiom_client import AXIOM_FETCH_DISABLED
            axiom_creds_ok = (not AXIOM_FETCH_DISABLED) and bool(os.getenv('AXIOM_CLIENT_ID', ''))
            if axiom_creds_ok:
                axiom_adb_map = _fetch_sp_devices_via_jobs(sp_names, chip_ids_set, target_name=target_name)
                # Cache the fetched devices into the unified cache for future use
                if axiom_adb_map:
                    fetched_devs = list({id(v): v for v in axiom_adb_map.values()}.values())
                    merged = _merge_axiom_qdt_devices(chip_name, pdt_type, fetched_devs)
                    _save_json(_unified_cache_path(chip_name, pdt_type), merged)
                    _save_json(_cache_path_for(chip_name, pdt_type), merged)
                    # Rebuild global map with newly cached devices
                    global_chip_map = _build_global_chip_id_map()

        previous_raw_devices = list(raw_devices or [])
        rebuilt_devices = []
        for chip_id in sorted(active_map.keys()):
            key = chip_id.upper()
            cached_dev = axiom_adb_map.get(key) or global_chip_map.get(key)
            if cached_dev:
                # Use cached/enriched device data — has MCN (deviceMcn), hostname, storage, etc.
                # Preserve the active-map chip key as the row identity so running jobs
                # still attach when Axiom /resources returns a different serial field.
                enriched_dev = dict(cached_dev)
                enriched_dev['chip_id'] = key
                enriched_dev['device_id'] = key
                rebuilt_devices.append(enriched_dev)

        if rebuilt_devices:
            raw_devices = rebuilt_devices
        elif previous_raw_devices:
            # Refresh found active SP jobs but no matching Axiom/cache device data.
            # Keep the previous inventory state instead of replacing it with
            # Unknown placeholder rows.
            raw_devices = previous_raw_devices
        else:
            # Use DB-backed axiom_job_summary rows when updater has device_host_map.
            # This gives Device Summary the actual device id + host immediately,
            # even before /resources/MCN enrichment is available.
            raw_devices = []
            for chip_id in sorted(active_map.keys()):
                jobs_for_chip = active_map.get(chip_id, [])
                first_job = jobs_for_chip[0] if jobs_for_chip else {}
                host_pc = str(first_job.get('host_pc') or first_job.get('site') or '').strip()
                if not host_pc:
                    continue
                raw_devices.append({
                    'chip_id': chip_id,
                    'device_id': chip_id,
                    'mcn': '',
                    'host': host_pc,
                    'host_pc': host_pc,
                    'storage': '',
                    'location': '',
                    'taxonomy_path': str(first_job.get('taxonomy_path') or '').strip(),
                    '_sp_only': True,
                    '_from_axiom_job_summary': True,
                })

    if sp_names and raw_devices and refresh:
        raw_devices = _enrich_sp_rows_with_live_axiom_status(raw_devices, max_workers=5)

    if not cache_hit:
        try:
            import time as _time
            _inventory_summary_cache[cache_key] = {
                'ts': _time.time(),
                'chip_name': chip_name,
                'raw_devices': list(raw_devices or []),
                'saved_at': saved_at,
                'active_map': active_map,
            }
            if len(_inventory_summary_cache) > 64:
                oldest_key = min(_inventory_summary_cache, key=lambda k: _inventory_summary_cache[k].get('ts') or 0)
                _inventory_summary_cache.pop(oldest_key, None)
        except Exception:
            pass

    devices = []
    for d in raw_devices:
        dev_id = _ds_device_identity(d)
        dev_key = dev_id.upper()
        host = _ds_device_host(d)
        mcn = _ds_device_mcn(d)
        quarantine = _ds_device_quarantine(d)
        jobs = active_map.get(dev_key, [])
        running = bool(jobs)
        status = 'quarantine' if quarantine else ('running' if running else 'idle')
        job_sites = sorted({j.get('site') for j in jobs if j.get('site')})
        # Site = CH/QIPL/SD derived from the Axiom taxonomy path of the device.
        # This comes ONLY from the Axiom /resources API (taxonomyPath field).
        # Do NOT use axiom_job_summary.site (anusat/wig/snowcone) — that is the
        # job execution site, not the device inventory site.
        site = _ds_device_location(d)
        # In SP-only mode (no Axiom device data yet), host falls back to job site name
        if d.get('_sp_only') and not host:
            host = (job_sites[0] if job_sites else '') or 'SP-only'
        # Extract rework and other details from QDT/Axiom data
        raw_d = d.get('_raw') or {}
        props_d = raw_d.get('properties') or {}
        row = {
            'device_id': dev_id,
            'host': host,
            'mcn': mcn,
            'storage': _ds_device_storage(d),
            'location': site,
            'site': site,
            'status': status,
            'running': running,
            'quarantine': quarantine,
            'running_jobs': jobs,
            'running_job_count': len(jobs),
            'pl_ids': sorted({j.get('pl_id') for j in jobs if j.get('pl_id')}),
            # Additional details from QDT/Axiom
            'rework_info': str(d.get('qdt_rework_info') or '').strip(),
            'asset_tag': str(d.get('qdt_asset_tag') or props_d.get('assetTagId') or d.get('asset_tag_id') or '').strip(),
            'assigned_to': str(d.get('qdt_assigned_to') or '').strip(),
            'condition': str(d.get('qdt_condition') or '').strip(),
            'mes_build': str(d.get('qdt_mes_build') or '').strip(),
            'serial_number': str(d.get('serial_number') or props_d.get('serialNumber') or '').strip(),
            'form_factor': str(d.get('form_factor') or '').strip(),
            'device_type': str(d.get('device_type') or props_d.get('deviceType') or '').strip(),
            'taxonomy_path': str(d.get('taxonomy_path') or raw_d.get('taxonomyPath') or '').strip(),
            'heartbeat': str(d.get('heartbeat') or raw_d.get('heartbeat') or '').strip(),
            'is_quarantined': bool(raw_d.get('isQuarantined') or False),
            'quarantine_reason': str(raw_d.get('quarantineReason') or '').strip(),
        }
        hay = ' '.join(str(v) for v in (row.get('device_id'), row.get('host'), row.get('mcn'), row.get('storage'), row.get('location'), row.get('site'), row.get('status'))).lower()
        if mcn_filter and mcn_filter not in str(row['mcn']).lower():
            continue
        if host_filter and host_filter not in str(row['host']).lower():
            continue
        if site_filter and site_filter not in str(row['site']).lower():
            continue
        if status_filter in ('running', 'idle', 'quarantine') and row['status'] != status_filter:
            continue
        if q and q not in hay:
            continue
        devices.append(row)

    def _bucket(key_name):
        out = {}
        for row in devices:
            key = str(row.get(key_name) or 'Unknown').strip() or 'Unknown'
            rec = out.setdefault(key, {'key': key, 'total': 0, 'running': 0, 'idle': 0, 'quarantine': 0, 'hosts': set(), 'mcns': set()})
            rec['total'] += 1
            rec[row['status']] += 1
            if row.get('host'):
                rec['hosts'].add(row['host'])
            if row.get('mcn'):
                rec['mcns'].add(row['mcn'])
        rows = []
        for rec in out.values():
            rec['hosts'] = sorted(rec['hosts'])
            rec['mcns'] = sorted(rec['mcns'])
            rows.append(rec)
        return sorted(rows, key=lambda r: (-r['total'], r['key']))

    totals = {
        'total': len(devices),
        'running': sum(1 for d in devices if d['running']),
        'quarantine': sum(1 for d in devices if d['quarantine']),
    }
    totals['idle'] = max(totals['total'] - totals['running'] - totals['quarantine'], 0)

    return jsonify({
        'success': True,
        'target': target_name,
        'chipset': chip_name,
        'pdt_type': pdt_type,
        'saved_at': saved_at,
        'filters': {'mcn': mcn_filter, 'host': host_filter, 'site': site_filter, 'status': status_filter, 'q': q, 'sp_names': sp_names},
        'totals': totals,
        'devices': devices,
        'by_mcn': _bucket('mcn'),
        'by_host': _bucket('host'),
        'by_site': _bucket('site'),
        'running_hosts': [r for r in _bucket('host') if r.get('running')],
        'mcn_options': sorted({d.get('mcn') for d in devices if d.get('mcn')}),
        'host_options': sorted({d.get('host') for d in devices if d.get('host')}),
        'site_options': sorted({d.get('site') for d in devices if d.get('site')}),
    })


@device_summary_api_bp.route('/device-summary-devices/<string:target_name>')
@login_required


def device_summary_devices_page(target_name):
    pdt_type = (request.args.get("pdt") or "SWPDT").strip().upper()

    if not target_name:
        abort(400, description="Invalid target")

    # Resolve chip_name from target config (e.g. ALDABRA -> SM4850)
    chip_name = get_chip_name_for_target(target_name) or ''
    if not chip_name:
        # fallback: use target_name itself as chip_name
        chip_name = str(target_name).strip().upper()
    chip_name = chip_name.strip().upper()

    # ------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------
    def _first_non_empty(*values):
        for v in values:
            if v is None:
                continue
            s = str(v).strip()
            if s:
                return s
        return ""

    def _normalize_storage(value):
        s = str(value or "").strip()
        if not s:
            return ""
        s = re.sub(r"\s+", " ", s)
        s = s.replace("UFS ", "UFS")
        s = s.replace("LPDDR ", "LPDDR")
        return s.strip()

    def _parse_from_description(desc):
        text = str(desc or "").strip()
        if not text:
            return {"mcn": "", "ddr": "", "storage": ""}

        t = text.upper()
        t = t.replace("+", " + ")
        t = t.replace("/", " / ")
        t = t.replace(",", " , ")
        t = re.sub(r"[\(\)\[\]\|]", " ", t)
        t = re.sub(r"\s+", " ", t).strip()

        mcn = ""
        ddr = ""
        storage = ""

        # MCN
        m = re.search(r"\bMCN[-_\s:]*([A-Z0-9\-]+)\b", t)
        if m:
            val = m.group(1).strip()
            mcn = val if val.startswith("MCN") else f"MCN{val}"

        # DDR
        ddr_patterns = [
            r"\b\d+\s*GB\s*(?:LPDDR[0-9A-Z\.]+|DDR[0-9A-Z\.]+)\b",
            r"\b(?:LPDDR[0-9A-Z\.]+|DDR[0-9A-Z\.]+)\s*\d+\s*GB\b",
            r"\b(?:LPDDR[0-9A-Z\.]+|DDR[0-9A-Z\.]+)\b",
        ]
        for pat in ddr_patterns:
            m = re.search(pat, t)
            if m:
                ddr = re.sub(r"\s+", " ", m.group(0)).strip()
                break

        # Storage
        storage_patterns = [
            r"\b\d+\s*GB\s*UFS\s*[0-9\.]*\b",
            r"\bUFS\s*[0-9\.]+\b",
            r"\bUFS\b",
            r"\b\d+\s*GB\s*EMMC\s*[0-9\.]*\b",
            r"\bEMMC\s*[0-9\.]+\b",
            r"\bEMMC\b",
        ]
        for pat in storage_patterns:
            m = re.search(pat, t)
            if m:
                storage = re.sub(r"\s+", " ", m.group(0)).strip()
                break

        return {
            "mcn": mcn,
            "ddr": ddr,
            "storage": _normalize_storage(storage),
        }

    def _normalize_taxonomy(taxonomy_path=""):
        return str(taxonomy_path or "").strip()

    def _include_for_pdt(pdt_type_value, taxonomy_path):
        tx = _normalize_taxonomy(taxonomy_path).upper()
        pdt = str(pdt_type_value or "").strip().upper()

        # HWPDT => only /PDT/QIPL/HW
        if pdt == "HWPDT":
            return tx.startswith("/PDT/QIPL/HW")

        # SWPDT => exclude /PDT/QIPL/HW
        return not tx.startswith("/PDT/QIPL/HW")

    def _location_from_values(taxonomy_path="", raw_location=""):
        tx = str(taxonomy_path or "").strip().upper()
        loc = str(raw_location or "").strip().upper()

        # taxonomy mapping --- check most specific first
        # /PDT/QIPL/HW is excluded for SWPDT, but map it to QIPL if it ever appears
        if "/PDT/QIPL" in tx:
            return "QIPL"
        if "/PDT/CHINA" in tx:
            return "CH"
        if "/PDT/SD" in tx or "/PDT/SAN DIEGO" in tx or "/PDT/SANDIEGO" in tx:
            return "SD"

        # fallback: derive from physical location path
        if "/AP/SHANGHAI" in loc or "/AP/CHINA" in loc:
            return "CH"
        if "/AP/SAN DIEGO" in loc or "/AP/SANDIEGO" in loc or "/AP/SD" in loc:
            return "SD"
        if "/AP/HYDERABAD" in loc or "QIPL" in loc:
            return "QIPL"

        return ""

    # MCN pattern: 10-XXXXX-XXX or 10-XXXXX-XXXX (followed by non-digit or end)
    _MCN_RE = re.compile(r'\b(\d{2}-\d{5}-\d{3,4})(?!\d)')

    def _extract_mcn(device_obj):
        """Extract MCN (10-XXXXX-XXXX) from _raw.description or qdt_model_desc.
        Never fall back to chipsetRev (V1.0) --- that is a hardware revision, not MCN."""
        raw  = device_obj.get("_raw") or {}
        desc = str(raw.get("description") or device_obj.get("description") or "")
        m = _MCN_RE.search(desc)
        if m:
            return m.group(1)
        # QDT enrichment: parse from dv_model string
        qdt_desc = str(device_obj.get("qdt_model_desc") or "")
        m2 = _MCN_RE.search(qdt_desc)
        if m2:
            return m2.group(1)
        return ""  # blank is better than showing V1.0

    def _extract_storage(device_obj):
        raw = device_obj.get("_raw") or {}
        deps = raw.get("dependencies") or {}

        storage = _first_non_empty(
            device_obj.get("storage_display"),
            device_obj.get("storage"),
            device_obj.get("Storage"),
            device_obj.get("qdt_storage"),
            deps.get("storage Type"),
            deps.get("storage"),
            raw.get("storage"),
        )

        if storage:
            return _normalize_storage(storage)

        desc = _first_non_empty(
            device_obj.get("description"),
            device_obj.get("Description"),
            raw.get("description"),
        )
        parsed = _parse_from_description(desc)
        return _normalize_storage(parsed.get("storage"))

    def _extract_rework_info(device_obj):
        raw = device_obj.get("_raw") or {}
        return _first_non_empty(
            device_obj.get("rework_info_display"),
            device_obj.get("rework_info"),
            device_obj.get("Rework Info"),
            device_obj.get("rework"),
            raw.get("rework_info"),
            raw.get("rework"),
        )

    def _extract_host_pc(device_obj):
        raw = device_obj.get("_raw") or {}
        return _first_non_empty(
            device_obj.get("host_pc"),
            device_obj.get("hostname"),
            device_obj.get("hostName"),
            raw.get("hostname"),
            raw.get("hostName"),
        )

    def _extract_device_id(device_obj):
        raw = device_obj.get("_raw") or {}
        return _first_non_empty(
            device_obj.get("device_id"),
            device_obj.get("device_key"),
            device_obj.get("serialNumber"),
            device_obj.get("serial_number"),
            device_obj.get("serial_no"),
            device_obj.get("serial"),
            raw.get("serialNumber"),
            raw.get("serial_number"),
            raw.get("serial_no"),
            raw.get("serial"),
            device_obj.get("hostname"),
            raw.get("hostname"),
        )

    # ------------------------------------------------------------
    # Load JSON
    # ------------------------------------------------------------
    root = Path(current_app.root_path)
    axiom_cache_dir = root / "static" / "axiom_cache"

    candidate_files = [
        axiom_cache_dir / f"{chip_name}_{pdt_type}_unified.json",
        axiom_cache_dir / f"{chip_name}_{pdt_type}.json",
        root / "cache" / f"{chip_name}_{pdt_type.lower()}_devices.json",
        root / "cache" / f"{chip_name}_{pdt_type}_devices.json",
        root / "cache" / f"{chip_name}_devices.json",
        root / "cache" / f"{chip_name}.json",
        root / "unified_cache" / f"{chip_name}_{pdt_type.lower()}_devices.json",
        root / "unified_cache" / f"{chip_name}.json",
    ]

    json_path = None
    for fp in candidate_files:
        if fp.exists():
            json_path = fp
            break

    if not json_path:
        current_app.logger.warning(
            "No device JSON found for target=%s pdt=%s. Checked: %s",
            chip_name, pdt_type, [str(x) for x in candidate_files]
        )
        # ------ No cache at all: render the "syncing" waiting page ------
        return render_template(
            "device_summary_devices.html",
            target_name=target_name,
            chip_name=chip_name,
            pdt_type=pdt_type,
            devices=[],
            total_devices=0,
            refresh_qdt_url=url_for("dashboard_bp.device_summary_page", target_name=target_name, pdt=pdt_type),
            cache_missing=True,
            sync_api_url=url_for("device_summary_api_bp.api_device_summary_data", target_name=target_name, pdt=pdt_type, refresh=1),
        )
    else:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)

    raw_devices = data.get("devices") or []

    # ------------------------------------------------------------
    # Build output rows
    # ------------------------------------------------------------
    devices_out = []

    for d in raw_devices:
        raw = d.get("_raw") or {}

        # Always prefer _raw.taxonomyPath --- the top-level taxonomy_path is often just "/PDT"
        taxonomy_path = _first_non_empty(
            raw.get("taxonomyPath"),
            d.get("taxonomy_path"),
            d.get("taxonomy"),
            d.get("taxonomy_display"),
            raw.get("taxonomy_path"),
            raw.get("taxonomy"),
        )

        # HWPDT / SWPDT split --- exclude /PDT/QIPL/HW from SWPDT
        if not _include_for_pdt(pdt_type, taxonomy_path):
            continue

        raw_location = _first_non_empty(
            d.get("location"),
            d.get("site"),
            d.get("lab"),
            raw.get("location"),
            raw.get("site"),
            raw.get("lab"),
        )

        location_display = _location_from_values(taxonomy_path, raw_location)

        device_row = {
            "device_id": _extract_device_id(d),
            "host_pc": _extract_host_pc(d),
            "mcn_display": _extract_mcn(d),
            "storage_display": _extract_storage(d),
            "location_display": location_display,
            "rework_info_display": _extract_rework_info(d),
            "taxonomy_display": taxonomy_path,
        }

        devices_out.append(device_row)

    # sort rows
    devices_out.sort(
        key=lambda x: (
            str(x.get("location_display") or ""),
            str(x.get("host_pc") or ""),
            str(x.get("device_id") or "")
        )
    )

    refresh_url = url_for("dashboard_bp.device_summary_page", target_name=target_name, pdt=pdt_type)

    return render_template(
        "device_summary_devices.html",
        target_name=target_name,
        chip_name=chip_name,
        pdt_type=pdt_type,
        devices=devices_out,
        total_devices=len(devices_out),
        refresh_qdt_url=refresh_url,
    )
